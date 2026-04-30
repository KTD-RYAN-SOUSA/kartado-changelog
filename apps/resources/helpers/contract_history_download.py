import datetime
import os
import random
import shutil
import string
import uuid
from collections import OrderedDict
from datetime import timedelta

import boto3
import pytz
from django.conf import settings
from django.contrib.admin.utils import flatten
from django.db.models import Case, CharField, Q, Value, When
from django.utils.timezone import now
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from zappa.asynchronous import task

from apps.companies.models import Firm, SubCompany
from apps.resources.models import Contract, ContractService
from apps.service_orders.models import ServiceOrderActionStatus
from apps.users.models import User
from helpers.apps.contract_utils import get_provisioned_price
from helpers.dates import utc_to_local
from RoadLabsAPI.settings import credentials


class ContractHistoryDownload:
    def __init__(
        self,
        company_id="",
        pk=None,
        filename=None,
        object_name=None,
    ):

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )
        if filename:
            self.filename = filename
        else:
            self.filename = self.get_filename()

        if object_name:
            self.object_name = object_name
        else:
            self.object_name = self.get_object_name()

        if pk:
            self.contract = Contract.objects.get(uuid=pk)
            self.contract_services = ContractService.objects.filter(
                Q(unit_price_service_contracts=self.contract)
                | Q(administration_service_contracts=self.contract)
                | Q(performance_service_contracts=self.contract)
            ).distinct()
            self.company_id = company_id
            self.users = {
                user.pk: user.get_full_name()
                for user in User.objects.filter(
                    companies__uuid=self.company_id
                ).distinct()
            }
            self.firms = {
                firm.pk: firm.name
                for firm in Firm.objects.filter(company_id=self.company_id)
            }
            self.cell_alignment = Alignment(
                vertical="center", wrap_text=True, wrapText=True
            )
            self.cell_font = Font(name="Calibri", size=11, color="000000")
            self.cell_font_bold = Font(
                name="Calibri", size=11, color="000000", bold=True
            )
            self.CONTRACT_TRANSLATION = {
                "r_c_number": "Número do objeto",
                "firm": "Fornecedor",
                "subcompany": "Fornecedor",
                "name": "Descrição",
                "contract_start": "Início",
                "contract_end": "Fim",
                "accounting_classification": "Classificação contábil",
                "provisioned_price": "Provisionado",
                "total_price": "Contratado",
                "spent_price": "Realizado",
                "responsibles_hirer": "Prepostos da contratante",
                "responsibles_hired": "Prepostos da contratada",
                "status": "Status do Objeto",
                "performance_months": "Duração em meses",
            }
            self.CONTRACT_SERVICE_TRANSLATION = {
                "description": "Descrição",
                "firms": "Equipes",
                "price": "Custo Provisionado",
                "weight": "Peso referente à seção",
                "unit_price": "Preço Unitário",
                "administration": "Administração",
                "performance": "Performance",
            }
            self.contract_service_types = {}
            self.EMPTY_VALUES = ["", None, []]

    def get_random_string(self):
        return "".join(
            random.SystemRandom().choice(string.ascii_lowercase + string.digits)
            for _ in range(10)
        )

    def get_filename(self):
        filename = "historico_objeto"

        return f"{filename}.xlsx"

    def get_object_name(self):
        random_string = self.get_random_string()
        object_name = "media/private/{}_{}.xlsx".format(
            self.filename.split(".xlsx")[0], random_string
        )

        return object_name

    def get_s3_url(self):
        empty = {"url": "", "name": ""}

        url = self.s3.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                "Key": self.object_name,
            },
        )

        if not url:
            return empty

        return url

    def upload_file(self, path):
        expires = now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)

        try:
            self.s3.upload_file(
                path,
                settings.AWS_STORAGE_BUCKET_NAME,
                self.object_name,
                ExtraArgs={"Expires": expires},
            )
        except Exception:
            return

    def copy_and_rename(self, old_folder, new_folder, temp_file):
        os.makedirs(new_folder, exist_ok=True)
        shutil.copy(old_folder + temp_file, new_folder + temp_file)
        os.rename(
            new_folder + temp_file,
            new_folder + self.filename + ".xlsx",
        )
        return

    def return_file_path(self, new_folder):
        return new_folder + self.filename + ".xlsx"

    def load_file(self, new_file_path):
        wb = load_workbook(filename=new_file_path)
        return wb

    def get_provisioned_price(self, obj):
        provisioned_price = get_provisioned_price(obj)
        return provisioned_price

    def return_user_list(self, user_set):
        if not isinstance(user_set, set):
            user_set = set(user_set)
        if not user_set:
            return ""

        user_name_list = [self.users.get(user) for user in user_set]
        return ", ".join(user_name_list)

    def return_firm_list(self, firm_set):
        if not isinstance(firm_set, set):
            firm_set = set(firm_set)
        if not firm_set:
            return ""

        firm_name_list = [self.firms.get(firm) for firm in firm_set]
        return ", ".join(firm_name_list)

    def translate_value(self, value, field=None):
        if isinstance(value, datetime.date):
            return value.strftime("%d/%m/%Y")
        elif isinstance(value, bool):
            return "Sim" if value else "Não"
        elif isinstance(value, uuid.UUID):
            if field == "status":
                return ServiceOrderActionStatus.objects.get(uuid=value).name
            elif field == "firm":
                return Firm.objects.get(uuid=value).name
            elif field == "subcompany":
                return SubCompany.objects.get(uuid=value).name
        return value if value is not None else ""

    def apply_header(
        self,
        ws,
        row,
        text,
        header_style=None,
        subtexts=[],
        subtext_styles={},
    ):

        cell = ws.cell(row=row, column=1)
        cell.value = text
        if header_style:
            cell._style = header_style

        if subtexts:
            ws.merge_cells(
                "{0}{1}:{2}{1}".format(
                    get_column_letter(1), row, get_column_letter(len(subtexts))
                )
            )
            for column_index, subtext in enumerate(subtexts):
                subtext_cell = ws.cell(row=row + 1, column=column_index + 1)
                subtext_cell.value = subtext
                if subtext_styles:
                    for attr, value in subtext_styles.items():
                        setattr(subtext_cell, attr, value)

    def format_contract_history(self, history, is_created=False):

        line_data = []

        if is_created:
            for item in history:
                if item["field"] in ["spend_schedule", "total_price", "spent_price"]:
                    formatted_new_value = f'{item["newValue"]:.2f}'.replace(".", ",")

                    line_data.append(
                        f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi criado com "R$ {formatted_new_value}"'
                    )
                elif item["field"] in ["responsibles_hirer", "responsibles_hired"]:
                    line_data.append(
                        f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi criado com [{item["newValue"]}]'
                    )
                else:
                    line_data.append(
                        f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi criado com "{item["newValue"]}"'
                    )
        else:
            for item in history:
                if item["field"] in ["spend_schedule", "total_price", "spent_price"]:
                    formatted_old_value = f'{item["oldValue"]:.2f}'.replace(".", ",")
                    formatted_new_value = f'{item["newValue"]:.2f}'.replace(".", ",")

                    line_data.append(
                        f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi alterado de "R$ {formatted_old_value}" para R$ "{formatted_new_value}"'
                    )
                elif item["field"] in ["responsibles_hirer", "responsibles_hired"]:
                    line_data.append(
                        f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi alterado de [{item["oldValue"]}] para [{item["newValue"]}]'
                    )
                else:
                    line_data.append(
                        f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi removido'
                        if item["action"] == "removed"
                        else f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi criado com "{item["newValue"]}"'
                        if item["oldValue"] in self.EMPTY_VALUES
                        and item["newValue"] not in self.EMPTY_VALUES
                        else f'"{self.CONTRACT_TRANSLATION.get(item["field"])}" foi alterado de "{item["oldValue"]}" para "{item["newValue"]}"'
                    )

        return "\n".join(line_data)

    def format_contract_service_history(self, history, is_created=False):

        line_data = []

        line_values = {}

        if is_created:
            for items in history:
                for item in items:
                    if item["field"] in [
                        "description",
                        "firms",
                        "weight",
                        "price",
                        "contract_service_type",
                    ]:
                        line_values.update({item["field"]: item["newValue"]})
                section_text = f'Seção ({self.CONTRACT_SERVICE_TRANSLATION.get(line_values["contract_service_type"])}) foi criado com [Descrição "{line_values["description"]}"]'
                if line_values.get("firms"):
                    section_text += f' e [Equipes: {line_values["firms"]}]'
                if line_values["contract_service_type"] == "performance":
                    formatted_price = f'{line_values["price"]:.2f}'.replace(".", ",")
                    formatted_weight = f'{line_values["weight"]:.2f}'.replace(".", ",")
                    section_text += f", [Custo Provisionado: R$ {formatted_price}], [Peso referente à seção: {formatted_weight}%]"
                line_data.append(section_text)
        else:
            for items in history:
                for item in items:
                    if item["field"] in [
                        "description",
                        "firms",
                        "weight",
                        "price",
                    ]:
                        line_values.update(
                            {item["field"]: (item["oldValue"], item["newValue"])}
                        )
                    elif item["field"] == "contract_service_type":
                        line_values.update({item["field"]: item["newValue"]})
                item_type = line_values.pop("contract_service_type")
                section_text = f"Em seção ({self.CONTRACT_SERVICE_TRANSLATION.get(item_type)}) foi alterado "
                for field, (old, new) in line_values.items():
                    if field in ["firms"]:
                        section_text += f'"{self.CONTRACT_SERVICE_TRANSLATION.get(field)}" de [{old}] para [{new}], '
                    elif field in ["price"]:
                        formatted_old_value = f"{old:.2f}".replace(".", ",")
                        formatted_new_value = f"{new:.2f}".replace(".", ",")
                        section_text += f'"{self.CONTRACT_SERVICE_TRANSLATION.get(field)}" de "R$ {formatted_old_value}" para "R$ {formatted_new_value}", '
                    elif field in ["weight"]:
                        formatted_old_value = f"{old:.2f}".replace(".", ",")
                        formatted_new_value = f"{new:.2f}".replace(".", ",")
                        section_text += f'"{self.CONTRACT_SERVICE_TRANSLATION.get(field)}" de "{formatted_old_value}%" para "{formatted_new_value}%", '
                    else:
                        section_text += f'"{self.CONTRACT_SERVICE_TRANSLATION.get(field)}" de "{old}" para "{new}", '
                line_data.append(section_text.strip()[:-1])

        return "\n".join(line_data)

    def get_contract_history(self):

        data = []

        first_history = self.contract.history.filter(history_type="+").first()

        first_values = [
            {
                "field": "r_c_number",
                "newValue": first_history.extra_info.get("r_c_number", ""),
            },
            {
                "field": "firm",
                "newValue": first_history.firm.name
                if first_history.firm
                else first_history.subcompany.name,
            },
            {"field": "name", "newValue": first_history.name},
            {
                "field": "contract_start",
                "newValue": first_history.contract_start.strftime("%d/%m/%Y"),
            },
            {
                "field": "contract_end",
                "newValue": first_history.contract_end.strftime("%d/%m/%Y"),
            },
            {
                "field": "accounting_classification",
                "newValue": first_history.extra_info.get(
                    "accounting_classification", ""
                ),
            },
            {
                "field": "provisioned_price",
                "newValue": self.get_provisioned_price(first_history),
            },
            {
                "field": "total_price",
                "newValue": getattr(first_history, "total_price", 0) or 0,
            },
            {
                "field": "spent_price",
                "newValue": getattr(first_history, "spent_price", 0) or 0,
            },
            {
                "field": "responsibles_hirer",
                "newValue": self.return_user_list(
                    first_history.responsibles_hirer.values_list("pk", flat=True)
                )
                if getattr(first_history, "responsibles_hirer")
                else "",
            },
            {
                "field": "responsibles_hired",
                "newValue": self.return_user_list(
                    first_history.responsibles_hired.values_list("pk", flat=True)
                )
                if getattr(first_history, "responsibles_hired")
                else "",
            },
            {
                "field": "status",
                "newValue": first_history.status.name if first_history.status else "",
            },
            {
                "field": "performance_months",
                "newValue": first_history.performance_months
                if first_history.performance_months is not None
                else "",
            },
        ]

        first_values = [
            {
                "field": item["field"],
                "action": "created",
                "oldValue": "",
                "newValue": item["newValue"],
            }
            for item in first_values
            if item["newValue"] not in self.EMPTY_VALUES
        ]

        data.append(
            {
                "historyDate": utc_to_local(first_history.history_date).replace(
                    second=0, microsecond=0
                ),
                "historyUser": self.users.get(first_history.history_user.uuid)
                if first_history.history_user
                else "",
                "historyChanges": first_values,
                "isCreated": True,
            }
        )

        remaining_histories = self.contract.history.exclude(history_type="+").order_by(
            "history_date"
        )

        for history in remaining_histories:

            previous_history = history.prev_record
            delta = history.diff_against(
                previous_history,
                included_fields=[
                    "extra_info",
                    "firm",
                    "subcompany",
                    "name",
                    "contract_start",
                    "contract_end",
                    "spend_schedule",
                    "total_price",
                    "spent_price",
                    "responsibles_hirer",
                    "responsibles_hired",
                    "status",
                    "performance_months",
                ],
            )
            new_history = {
                "historyDate": utc_to_local(history.history_date).replace(
                    second=0, microsecond=0
                ),
                "historyUser": self.users.get(history.history_user.uuid)
                if history.history_user
                else "",
                "historyChanges": [],
                "isCreated": False,
            }
            for change in delta.changes:
                if change.field in ["status", "firm", "subcompany"]:
                    new_history["historyChanges"].append(
                        {
                            "field": change.field,
                            "action": "updated"
                            if change.new not in self.EMPTY_VALUES
                            else "removed",
                            "oldValue": self.translate_value(change.old, change.field),
                            "newValue": self.translate_value(change.new, change.field),
                        }
                    )
                elif change.field == "spend_schedule":
                    old_value = self.get_provisioned_price(previous_history)
                    new_value = self.get_provisioned_price(history)
                    if old_value != new_value:
                        new_history["historyChanges"].append(
                            {
                                "field": "provisioned_price",
                                "action": "updated"
                                if new_value not in self.EMPTY_VALUES
                                else "removed",
                                "oldValue": old_value,
                                "newValue": new_value,
                            }
                        )
                elif change.field in ["contract_start", "contract_end"]:
                    new_history["historyChanges"].append(
                        {
                            "field": change.field,
                            "action": "updated"
                            if change.new not in self.EMPTY_VALUES
                            else "removed",
                            "oldValue": self.translate_value(change.old),
                            "newValue": self.translate_value(change.new),
                        }
                    )
                elif change.field in [
                    "name",
                    "total_price",
                    "spent_price",
                    "performance_months",
                ]:
                    new_history["historyChanges"].append(
                        {
                            "field": change.field,
                            "action": "updated"
                            if change.new not in self.EMPTY_VALUES
                            else "removed",
                            "oldValue": self.translate_value(change.old),
                            "newValue": self.translate_value(change.new),
                        }
                    )
                elif change.field == "extra_info":
                    old_r_c_number = change.old.get("r_c_number", "")
                    new_r_c_number = change.new.get("r_c_number", "")
                    if old_r_c_number != new_r_c_number:
                        new_history["historyChanges"].append(
                            {
                                "field": "r_c_number",
                                "action": "updated"
                                if new_r_c_number not in self.EMPTY_VALUES
                                else "removed",
                                "oldValue": old_r_c_number,
                                "newValue": new_r_c_number,
                            }
                        )
                    old_accounting_classification = change.old.get(
                        "accounting_classification", ""
                    )
                    new_accounting_classification = change.new.get(
                        "accounting_classification", ""
                    )
                    if old_accounting_classification != new_accounting_classification:
                        new_history["historyChanges"].append(
                            {
                                "field": "accounting_classification",
                                "action": "updated"
                                if new_accounting_classification
                                not in self.EMPTY_VALUES
                                else "removed",
                                "oldValue": old_accounting_classification,
                                "newValue": new_accounting_classification,
                            }
                        )

                elif change.field in ["responsibles_hirer", "responsibles_hired"]:
                    current_users = set(
                        getattr(history, change.field).values_list("user_id", flat=True)
                    )
                    previous_users = set(
                        getattr(previous_history, change.field).values_list(
                            "user_id", flat=True
                        )
                    )

                    new_history["historyChanges"].append(
                        {
                            "field": change.field,
                            "action": "updated",
                            "oldValue": self.return_user_list(previous_users),
                            "newValue": self.return_user_list(current_users),
                        }
                    )

                else:
                    pass

            if new_history.get("historyChanges", []) != []:
                data.append(new_history)

        data.sort(key=lambda x: x["historyChanges"][0].get("field", ""), reverse=True)
        data.sort(key=lambda x: x["historyDate"])
        data.sort(key=lambda x: x["isCreated"], reverse=True)

        temp_dict = OrderedDict()
        for item in data:
            temp_dict.setdefault(
                (item["historyDate"], item["historyUser"], item["isCreated"]), []
            ).append(item["historyChanges"])

        new_history = [
            {
                "historyDate": k[0],
                "historyUser": k[1],
                "historyChanges": flatten(v.pop() if len(v) == 1 else v),
                "isCreated": k[2],
            }
            for k, v in temp_dict.items()
        ]

        return new_history

    def get_contract_service_history(self):

        data = []

        type_checking = (
            self.contract_services.annotate(
                type_name=Case(
                    When(
                        contract_item_unit_prices__isnull=False,
                        then=Value("unit_price"),
                    ),
                    When(
                        contract_item_administration__isnull=False,
                        then=Value("administration"),
                    ),
                    When(
                        contract_item_performance__isnull=False,
                        then=Value("performance"),
                    ),
                    default=Value(""),
                    output_field=CharField(),
                ),
            )
            .exclude(type_name="")
            .in_bulk(field_name="uuid")
        )

        for contract_service in self.contract_services:
            first_history = contract_service.history.filter(history_type="+").first()

            contract_service_item = type_checking.get(first_history.uuid)

            if contract_service_item:

                contract_service_type = getattr(contract_service_item, "type_name")

                self.contract_service_types.update(
                    {first_history.uuid: contract_service_type}
                )

                first_values = [
                    {
                        "field": "description",
                        "newValue": first_history.description,
                    },
                    {
                        "field": "firms",
                        "newValue": self.return_firm_list(
                            first_history.firms.values_list("pk", flat=True)
                        )
                        if getattr(first_history, "firms")
                        else "",
                    },
                ]
                if contract_service_type == "performance":
                    first_values.extend(
                        [
                            {
                                "field": "price",
                                "newValue": first_history.price,
                            },
                            {
                                "field": "weight",
                                "newValue": first_history.weight,
                            },
                        ]
                    )
                first_values.append(
                    {
                        "field": "contract_service_type",
                        "newValue": contract_service_type,
                    }
                )
                first_values = [
                    {
                        "field": item["field"],
                        "action": "created",
                        "oldValue": "",
                        "newValue": item["newValue"],
                    }
                    for item in first_values
                    if item["newValue"] not in self.EMPTY_VALUES
                ]

                data.append(
                    {
                        "historyDate": utc_to_local(first_history.history_date).replace(
                            second=0, microsecond=0
                        ),
                        "historyUser": self.users.get(first_history.history_user.uuid)
                        if first_history.history_user
                        else "",
                        "historyChanges": first_values,
                        "isCreated": True,
                    }
                )

        remaining_histories = (
            ContractService.history.model.objects.filter(
                uuid__in=list(type_checking.keys())
            )
            .exclude(history_type="+")
            .order_by("history_date")
        )

        for history in remaining_histories:
            previous_history = history.prev_record
            delta = history.diff_against(
                previous_history,
                included_fields=["description", "firms", "price", "weight"],
            )
            new_history = {
                "historyDate": utc_to_local(history.history_date).replace(
                    second=0, microsecond=0
                ),
                "historyUser": self.users.get(history.history_user.uuid)
                if history.history_user
                else "",
                "historyChanges": [],
                "isCreated": False,
            }
            for change in delta.changes:
                if change.field in ["description", "price", "weight"]:
                    new_history["historyChanges"].append(
                        {
                            "field": change.field,
                            "action": "updated"
                            if change.new not in self.EMPTY_VALUES
                            else "removed",
                            "oldValue": self.translate_value(change.old),
                            "newValue": self.translate_value(change.new),
                        }
                    )
                elif change.field in ["firms"]:
                    current_firms = set(
                        getattr(history, change.field).values_list("firm_id", flat=True)
                    )
                    previous_firms = set(
                        getattr(previous_history, change.field).values_list(
                            "firm_id", flat=True
                        )
                    )

                    new_history["historyChanges"].append(
                        {
                            "field": change.field,
                            "action": "updated",
                            "oldValue": self.return_firm_list(previous_firms),
                            "newValue": self.return_firm_list(current_firms),
                        }
                    )
                else:
                    pass

            if new_history.get("historyChanges", []) != []:
                contract_service_type = getattr(
                    type_checking.get(history.uuid), "type_name"
                )
                new_history["historyChanges"].append(
                    {
                        "field": "contract_service_type",
                        "action": "fixed",
                        "newValue": contract_service_type,
                    }
                )
                data.append(new_history)

        data.sort(key=lambda x: x["historyChanges"][0].get("field", ""), reverse=True)
        data.sort(key=lambda x: x["historyDate"])

        temp_dict = OrderedDict()
        for item in data:
            temp_dict.setdefault(
                (item["historyDate"], item["historyUser"], item["isCreated"]), []
            ).append(item["historyChanges"])

        new_history = [
            {
                "historyDate": k[0],
                "historyUser": k[1],
                "historyChanges": v,
                "isCreated": k[2],
            }
            for k, v in temp_dict.items()
        ]

        return new_history

    def fill_contract_history(self):
        data_contract = self.get_contract_history()

        dados = self.wb["Dados"]

        start_line = 3

        for line_index, item in enumerate(data_contract):
            date_cell = dados.cell(row=line_index + start_line, column=1)
            date_cell.value = item.get("historyDate").strftime("%d/%m/%Y, %H:%M")
            user_cell = dados.cell(row=line_index + start_line, column=2)
            user_cell.value = item.get("historyUser", "")
            changes_cell = dados.cell(row=line_index + start_line, column=3)
            changes_cell.value = self.format_contract_history(
                item.get("historyChanges"),
                is_created=True if item.get("isCreated", False) is True else False,
            )

            user_cell.alignment = self.cell_alignment
            date_cell.alignment = self.cell_alignment
            changes_cell.alignment = self.cell_alignment

            user_cell.font = self.cell_font
            date_cell.font = self.cell_font
            changes_cell.font = self.cell_font

    def fill_contract_service_history(self):

        dados = self.wb["Dados"]
        start_line = dados.max_row + 2

        header_style = dados["A1"]._style

        self.apply_header(
            dados,
            start_line,
            "INFORMAÇÕES DO CONJUNTO DO OBJETO",
            header_style,
            ["Data", "Usuário", "Descrição"],
            {"font": self.cell_font_bold},
        )

        start_line = dados.max_row + 1

        data_contract_service = self.get_contract_service_history()
        for line_index, item in enumerate(data_contract_service):
            date_cell = dados.cell(row=line_index + start_line, column=1)
            date_cell.value = item.get("historyDate").strftime("%d/%m/%Y, %H:%M")
            user_cell = dados.cell(row=line_index + start_line, column=2)
            user_cell.value = item.get("historyUser", "")
            changes_cell = dados.cell(row=line_index + start_line, column=3)
            changes_cell.value = self.format_contract_service_history(
                item.get("historyChanges"),
                is_created=True if item.get("isCreated", False) is True else False,
            )

            user_cell.alignment = self.cell_alignment
            date_cell.alignment = self.cell_alignment
            changes_cell.alignment = self.cell_alignment

            user_cell.font = self.cell_font
            date_cell.font = self.cell_font
            changes_cell.font = self.cell_font

    def generate_file(self):
        old_folder = "apps/resources/templates/"
        new_folder = "/tmp/contract_history_download/"
        temp_file = "historico_contrato.xlsx"

        self.copy_and_rename(old_folder, new_folder, temp_file)
        new_file_path = self.return_file_path(new_folder)
        self.wb = self.load_file(new_file_path)
        self.fill_contract_history()
        self.fill_contract_service_history()

        self.wb.save(new_file_path)
        self.upload_file(new_file_path)
        os.remove(new_file_path)
        os.rmdir(new_folder)


@task
def contract_history_download_async(company_id, pk, filename, object_name):
    contract_history_object = ContractHistoryDownload(
        company_id=company_id, pk=pk, filename=filename, object_name=object_name
    )
    contract_history_object.generate_file()
