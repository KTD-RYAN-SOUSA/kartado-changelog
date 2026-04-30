from io import BytesIO
from tempfile import TemporaryFile
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl.compat import deprecated
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.reader.excel import ExcelReader, _find_workbook_part
from openpyxl.reader.workbook import WorkbookParser
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.xml.constants import XLTM, XLTX

from helpers.kartado_excel.worksheet import KartadoWorksheetWriter


class KartadoExcelWriter(ExcelWriter):
    def write_worksheet(self, ws):
        ws._drawing = SpreadsheetDrawing()
        ws._drawing.charts = ws._charts
        ws._drawing.images = ws._images
        if self.workbook.write_only:
            if not ws.closed:
                ws.close()
            writer = ws._writer
        else:
            writer = KartadoWorksheetWriter(ws)
            writer.write()

        ws._rels = writer._rels
        self._archive.write(writer.out, ws.path[1:])
        self.manifest.append(ws)
        writer.cleanup()


class KartadoExcelReader(ExcelReader):
    def read_workbook(self):
        wb_part = _find_workbook_part(self.package)
        self.parser = KartadoWorkbookParser(
            self.archive, wb_part.PartName[1:], keep_links=self.keep_links
        )
        self.parser.parse()
        wb = self.parser.wb
        wb._sheets = []
        wb._data_only = self.data_only
        wb._read_only = self.read_only
        wb.template = wb_part.ContentType in (XLTX, XLTM)

        # If are going to preserve the vba then attach a copy of the archive to the
        # workbook so that is available for the save.
        if self.keep_vba:
            wb.vba_archive = ZipFile(BytesIO(), "a", ZIP_DEFLATED)
            for name in self.valid_files:
                wb.vba_archive.writestr(name, self.archive.read(name))

        if self.read_only:
            wb._archive = self.archive

        self.wb = wb


def save_workbook(workbook, filename):
    """Save the given workbook on the filesystem under the name filename.

    :param workbook: the workbook to save
    :type workbook: :class:`openpyxl.workbook.Workbook`

    :param filename: the path to which save the workbook
    :type filename: string

    :rtype: bool

    """
    archive = ZipFile(filename, "w", ZIP_DEFLATED, allowZip64=True)
    writer = KartadoExcelWriter(workbook, archive)
    writer.save()
    return True


@deprecated("Use a NamedTemporaryFile")
def save_virtual_workbook(workbook):
    """Return an in-memory workbook, suitable for a Django response."""
    tmp = TemporaryFile()
    archive = ZipFile(tmp, "w", ZIP_DEFLATED, allowZip64=True)

    writer = KartadoExcelWriter(workbook, archive)
    writer.save()

    tmp.seek(0)
    virtual_workbook = tmp.read()
    tmp.close()

    return virtual_workbook


def load_workbook(
    filename, read_only=False, keep_vba=False, data_only=False, keep_links=True
):
    """Open the given filename and return the workbook

    :param filename: the path to open or a file-like object
    :type filename: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`

    :param read_only: optimised for reading, content cannot be edited
    :type read_only: bool

    :param keep_vba: preseve vba content (this does NOT mean you can use it)
    :type keep_vba: bool

    :param data_only: controls whether cells with formulae have either the formula (default) or the value stored the last time Excel read the sheet
    :type data_only: bool

    :param keep_links: whether links to external workbooks should be preserved. The default is True
    :type keep_links: bool

    :rtype: :class:`openpyxl.workbook.Workbook`

    .. note::

        When using lazy load, all worksheets will be :class:`openpyxl.worksheet.iter_worksheet.IterableWorksheet`
        and the returned workbook will be read-only.

    """
    reader = KartadoExcelReader(filename, read_only, keep_vba, data_only, keep_links)
    reader.read()
    return reader.wb


class KartadoWorkbook(Workbook):
    def save(self, filename):
        """Save the current workbook under the given `filename`.
        Use this function instead of using an `ExcelWriter`.

        .. warning::
            When creating your workbook using `write_only` set to True,
            you will only be able to call this function once. Subsequents attempts to
            modify or save the file will raise an :class:`openpyxl.shared.exc.WorkbookAlreadySaved` exception.
        """
        if self.read_only:
            raise TypeError("""Workbook is read-only""")
        if self.write_only and not self.worksheets:
            self.create_sheet()
        save_workbook(self, filename)


class KartadoWorkbookParser(WorkbookParser):
    def __init__(self, archive, workbook_part_name, keep_links=True):
        self.archive = archive
        self.workbook_part_name = workbook_part_name
        self.wb = KartadoWorkbook()
        self.keep_links = keep_links
        self.sheets = []
