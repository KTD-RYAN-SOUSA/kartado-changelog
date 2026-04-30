import logging
import os
import shutil
from collections import defaultdict
from math import ceil, floor

import sentry_sdk
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from zappa.asynchronous import task

from apps.resources.models import ContractService, FieldSurveyExport, FieldSurveyRoad
from apps.roads.models import Road


@task
def generate_survey(field_export_id):

    try:

        field_export = FieldSurveyExport.objects.get(pk=field_export_id)

    except Exception as e:

        logging.error("FieldSurveyExport not found")
        sentry_sdk.capture_exception(e)

    else:
        field_survey = field_export.field_survey
        number = field_survey.number

        error = True
        old_folder = "apps/resources/templates/"
        new_folder = "/tmp/field_survey/"
        temp_file = "temp_survey.xlsx"

        CONTRACT_ROW = 2
        SURVEY_ROW = 6

        cabecalho_style = NamedStyle(name="cabecalho_style")
        cabecalho_style.font = Font(name="Calibri", size=10, color="000000")
        cabecalho_style.alignment = Alignment(horizontal="left")
        cabecalho_style.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        cabecalho_style.protection = Protection(hidden=True)

        detalhe_title_style = NamedStyle(name="detalhe_title_style")
        detalhe_title_style.font = Font(name="Calibri", size=10, color="000000")
        detalhe_title_style.alignment = Alignment(horizontal="center")
        detalhe_title_style.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        detalhe_title_style.protection = Protection(hidden=True)

        detalhe_subtitle_style = NamedStyle(name="detalhe_subtitle_style")
        detalhe_subtitle_style.font = Font(name="Calibri", size=10, color="000000")
        detalhe_subtitle_style.alignment = Alignment(horizontal="center")
        detalhe_subtitle_style.fill = PatternFill("solid", fgColor="9BC2E6")
        detalhe_subtitle_style.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        detalhe_subtitle_style.protection = Protection(hidden=True)

        simple_style = NamedStyle(name="simple_style")
        simple_style.font = Font(name="Calibri", size=10, color="000000")
        simple_style.alignment = Alignment(horizontal="left")
        simple_style.protection = Protection(hidden=True)

        def move_and_rename(survey_number):
            os.makedirs(new_folder, exist_ok=True)
            shutil.copy(old_folder + temp_file, new_folder + temp_file)
            os.rename(
                new_folder + temp_file,
                new_folder + "Avaliação de Campo " + survey_number + ".xlsx",
            )
            return

        def return_file_path(survey_number):

            return new_folder + "Avaliação de Campo " + survey_number + ".xlsx"

        def load_file(survey_number):

            wb = load_workbook(filename=return_file_path(survey_number))
            return wb

        def get_contract_data(contract):

            try:
                object_number = contract.extra_info["r_c_number"]
            except Exception:
                raise Exception('Campo  "Nº do Objeto" não está preenchido')

            try:
                subcompany_name = contract.subcompany.name
            except Exception:
                try:
                    subcompany_name = contract.firm.name
                except Exception:
                    raise Exception('Campo "Fornecedor" não está preenchido')

            try:
                contract_status = contract.status.name
            except Exception:
                contract_status = None

            try:
                hirer_list = ", ".join(
                    sorted(
                        [
                            hirer.get_full_name()
                            for hirer in contract.responsibles_hirer.all()
                        ]
                    )
                )
            except Exception:
                raise Exception(
                    'Erro no preenchimento do campo "Prepostos da Contratante"'
                )

            try:
                hired_list = ", ".join(
                    sorted(
                        [
                            hired.get_full_name()
                            for hired in contract.responsibles_hired.all()
                        ]
                    )
                )
            except Exception:
                raise Exception(
                    'Erro no preenchimento do campo "Prepostos do Contratado"'
                )

            performance_provisioned = round(
                sum(
                    list(contract.performance_services.values_list("price", flat=True))
                ),
                2,
            )

            return [
                object_number,
                subcompany_name,
                contract_status,
                hirer_list,
                hired_list,
                performance_provisioned,
            ]

        def get_survey_data(field_survey):
            def get_signature_status(obj, hire_type):
                if hire_type == "hirer":
                    signed_list = obj.signatures.filter(hirer__isnull=False)
                elif hire_type == "hired":
                    signed_list = obj.signatures.filter(hired__isnull=False)
                signed_status = [
                    True if item.signed_at else False for item in signed_list
                ]

                return "Assinado" if all(signed_status) else "Aguardando assinatura"

            return [
                field_survey.number,
                field_survey.name,
                ", ".join(
                    sorted(
                        [
                            hirer.get_full_name()
                            for hirer in field_survey.responsibles_hirer.all()
                        ]
                    )
                ),
                get_signature_status(field_survey, "hirer"),
                ", ".join(
                    sorted(
                        [
                            hired.get_full_name()
                            for hired in field_survey.responsibles_hired.all()
                        ]
                    )
                ),
                get_signature_status(field_survey, "hired"),
                field_survey.created_at.strftime("%d/%m/%Y"),
                field_survey.contract.firm.company.entity_set.first().name
                if field_survey.contract.firm
                else field_survey.contract.subcompany.company.entity_set.first().name,
                "{} ({})".format(
                    field_survey.measurement_bulletin.description
                    if field_survey.measurement_bulletin.description
                    else None,
                    field_survey.measurement_bulletin.number
                    if field_survey.measurement_bulletin.number
                    else None,
                )
                if field_survey.measurement_bulletin
                else None,
            ]

        def fill_contract_info(field_survey, workbook):

            data_contract = get_contract_data(field_survey.contract)

            try:
                cabecalho = workbook["Cabeçalho"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            for index, item in enumerate(data_contract):

                cell = cabecalho.cell(row=CONTRACT_ROW, column=index + 1)

                cell.value = item
                cell.style = cabecalho_style
                if get_column_letter(index + 1) in ["F", "G", "H"]:
                    cell.number_format = "R$ #,##0.00"

        def fill_survey_info(field_survey, workbook):

            data_survey = get_survey_data(field_survey)

            try:
                cabecalho = workbook["Cabeçalho"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            for index, item in enumerate(data_survey):

                cell = cabecalho.cell(row=SURVEY_ROW, column=index + 1)

                cell.value = item
                cell.style = cabecalho_style
                if get_column_letter(index + 1) in ["G"]:
                    cell.number_format = "dd/mm/yyyy"

        def get_grades_length(worksheet, column=2):
            return len(
                [
                    cell
                    for cell in list(worksheet[get_column_letter(column)])
                    if cell.value is not None
                ]
            )

        def get_and_fill_detail_info(field_survey, workbook):
            try:
                detalhe = workbook["Detalhes da Avaliação"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            detail_column = 3
            title_column = 3
            style_flag = True
            style_title_flag = True

            services = ContractService.objects.filter(
                performance_service_contracts=field_survey.contract
            ).order_by("uuid")

            validated_data = defaultdict(list)
            used_uuids = defaultdict(set)
            average_denominator = defaultdict(set)

            for service in services:
                for perf_item in service.contract_item_performance.all().order_by(
                    "order", "uuid"
                ):
                    validated_data[service.description].append(
                        perf_item.resource.resource.name
                        if perf_item.resource and perf_item.resource.resource
                        else ""
                    )
                    detail_row = 3
                    for road, perf_uuid in field_survey.grades.items():
                        road_flag = 0
                        road_name = Road.objects.get(id=road).name
                        item_data = perf_uuid.get(str(perf_item.uuid), None)
                        survey_road = FieldSurveyRoad.objects.get(
                            contract=field_survey.contract, road_id=road
                        )
                        start_km = survey_road.start_km
                        end_km = survey_road.end_km
                        average_denominator[road] = ceil(end_km) - floor(start_km)
                        if item_data or item_data == {}:
                            used_uuids[str(service.uuid)].add(str(perf_item.uuid))
                            for i in range(floor(start_km), ceil(end_km)):
                                if not road_flag:
                                    road_cell = detalhe.cell(row=detail_row, column=1)
                                    road_cell.value = road_name
                                    road_cell.style = simple_style

                                    km_cell = detalhe.cell(row=detail_row, column=2)
                                    km_cell.value = f"Km {str(i)} – Km {str(i + 1)}"
                                    km_cell.style = simple_style
                                grade_value = item_data.get(str(i), None)
                                grade_cell = detalhe.cell(
                                    row=detail_row, column=detail_column
                                )
                                grade_cell.value = (
                                    grade_value if grade_value is not None else None
                                )
                                grade_cell.style = simple_style
                                grade_cell.alignment = Alignment(horizontal="right")
                                grade_cell.border = Border(
                                    left=Side(border_style="thin", color="000000"),
                                    right=Side(border_style="thin", color="000000"),
                                )
                                if style_flag:
                                    grade_cell.fill = PatternFill(
                                        "solid", fgColor="DEE6EF"
                                    )
                                detail_row += 1
                            road_flag += 1
                    detail_column += 1
                style_flag = not style_flag

            for service, perf_item in validated_data.items():
                cell_title = detalhe.cell(row=1, column=title_column)
                cell_title.value = service
                cell_title.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                )
                for item in perf_item:
                    cell_perf = detalhe.cell(row=2, column=title_column)
                    cell_perf.value = item
                    cell_perf.style = detalhe_subtitle_style
                    detalhe.column_dimensions[
                        get_column_letter(title_column)
                    ].width = max(10, len(item))
                    title_column += 1
                merge_len = len(perf_item)
                detalhe.merge_cells(
                    "{0}{2}:{1}{2}".format(
                        get_column_letter(title_column - merge_len),
                        get_column_letter(title_column - 1),
                        1,
                    )
                )

                cell_merged = detalhe.cell(row=1, column=title_column - merge_len)
                cell_merged.style = detalhe_title_style
                if style_title_flag:
                    cell_merged.fill = PatternFill("solid", fgColor="DEE6EF")

                style_title_flag = not style_title_flag

            services_length = len([b for a in list(validated_data.values()) for b in a])

            total_km = 0
            for kms in average_denominator.values():
                total_km += kms

            grades_length = (
                get_grades_length(detalhe) + 3
            )  # Adding the title columns offset
            for column in range(3, services_length + 3):
                average = 0
                for line in range(3, grades_length + 3):
                    grade = detalhe.cell(row=line, column=column).value
                    if grade is not None:
                        average += grade
                average = average / total_km
                cell_average = detalhe.cell(row=grades_length, column=column)
                cell_average.value = average
                cell_average.style = detalhe_subtitle_style
                cell_average.alignment = Alignment(horizontal="right")

            media_cell = detalhe.cell(row=grades_length, column=1)
            media_cell.value = "Média"
            media_cell.style = detalhe_subtitle_style
            media_cell.alignment = Alignment(horizontal="left")
            detalhe.merge_cells(
                "{0}{2}:{1}{2}".format(
                    get_column_letter(1), get_column_letter(2), grades_length
                )
            )

            road_cell = detalhe.cell(row=3, column=1)
            road_cell.border = Border(top=Side(border_style="thin", color="000000"))
            km_cell = detalhe.cell(row=3, column=2)
            km_cell.border = Border(top=Side(border_style="thin", color="000000"))

            detalhe.column_dimensions["B"].width = 36

            return used_uuids

        def get_and_fill_summary_info(field_survey, workbook, used_uuids):

            try:
                resumo = workbook["Resumo da Avaliação"]
                detalhe = workbook["Detalhes da Avaliação"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            detalhe_row = 2
            service_row = 2

            average_row = (
                get_grades_length(detalhe) + 3
            )  # Adding the title columns offset

            services = ContractService.objects.filter(
                performance_service_contracts=field_survey.contract
            ).order_by("uuid")
            title_len = []
            subtitle_len = []
            service_grades = []
            for service in services:
                if str(service.uuid) in used_uuids.keys():
                    title_column = 1
                    service_average = []
                    title_len.append(len(service.description))
                    title_cell = resumo.cell(row=detalhe_row, column=title_column)
                    title_cell.value = service.description
                    title_cell.style = simple_style
                    title_column += 1
                    weight_cell = resumo.cell(row=detalhe_row, column=title_column)
                    weight_cell.value = service.weight / 100
                    weight_cell.style = simple_style
                    weight_cell.alignment = Alignment(horizontal="right")
                    weight_cell.number_format = FORMAT_PERCENTAGE_00
                    for perf_item in service.contract_item_performance.all().order_by(
                        "order", "uuid"
                    ):
                        if str(perf_item.uuid) in set().union(*used_uuids.values()):
                            subitem_column = 3
                            sort_cell = resumo.cell(
                                row=detalhe_row, column=subitem_column
                            )
                            sort_cell.value = perf_item.sort_string
                            sort_cell.style = simple_style
                            sort_cell.alignment = Alignment(horizontal="right")
                            subitem_column += 1
                            perf_name = (
                                perf_item.resource.resource.name
                                if perf_item.resource and perf_item.resource.resource
                                else ""
                            )
                            subtitle_len.append(len(perf_name))
                            perf_cell = resumo.cell(
                                row=detalhe_row, column=subitem_column
                            )
                            perf_cell.value = perf_name
                            perf_cell.style = simple_style
                            subitem_column += 1
                            subweight_cell = resumo.cell(
                                row=detalhe_row, column=subitem_column
                            )
                            subweight_cell.value = perf_item.weight / 100
                            subweight_cell.style = simple_style
                            subweight_cell.alignment = Alignment(horizontal="right")
                            subweight_cell.number_format = FORMAT_PERCENTAGE_00
                            subitem_column += 1
                            average_cell_value = detalhe.cell(
                                row=average_row, column=detalhe_row + 1
                            ).value
                            average_cell = resumo.cell(
                                row=detalhe_row, column=subitem_column
                            )
                            average_cell_value = (
                                average_cell_value
                                if not field_survey.manual
                                else field_survey.final_grade / 10
                            )
                            average_cell.value = average_cell_value
                            average_cell.style = simple_style
                            average_cell.alignment = Alignment(horizontal="right")
                            subitem_column += 1
                            subweight_average_cell_value = (
                                average_cell_value * perf_item.weight
                            ) / 10
                            service_average.append(subweight_average_cell_value)
                            subweight_average_cell = resumo.cell(
                                row=detalhe_row, column=subitem_column
                            )
                            subweight_average_cell.value = subweight_average_cell_value
                            subweight_average_cell.style = simple_style
                            subweight_average_cell.alignment = Alignment(
                                horizontal="right"
                            )

                            detalhe_row += 1
                    service_average_value = sum(service_average)
                    service_average_cell = resumo.cell(
                        row=service_row, column=subitem_column + 1
                    )
                    service_average_cell.value = service_average_value
                    service_average_cell.style = simple_style
                    service_average_cell.alignment = Alignment(horizontal="right")
                    service_grade_value = (service_average_value * service.weight) / 100
                    service_grades.append(service_grade_value)
                    service_average_weight_cell = resumo.cell(
                        row=service_row, column=subitem_column + 2
                    )
                    service_average_weight_cell.value = service_grade_value
                    service_average_weight_cell.style = simple_style
                    service_average_weight_cell.alignment = Alignment(
                        horizontal="right"
                    )
                    service_row = detalhe_row

            final_grade_cell = resumo.cell(row=2, column=10)
            final_grade_cell.value = sum(service_grades) / 100
            final_grade_cell.style = simple_style
            final_grade_cell.alignment = Alignment(horizontal="right")
            final_grade_cell.number_format = FORMAT_PERCENTAGE_00

            survey_title = f"Nota Final – {field_survey.name}"
            survey_title_cell = resumo.cell(row=1, column=10)
            survey_title_cell.value = survey_title
            survey_title_cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )
            survey_title_cell.alignment = Alignment(horizontal="center")

            resumo.column_dimensions["A"].width = max(title_len, default=30)
            resumo.column_dimensions["D"].width = max(subtitle_len, default=30)
            resumo.column_dimensions["J"].width = len(survey_title)

        try:
            move_and_rename(number)
            workbook = load_file(number)
            fill_contract_info(field_survey, workbook)
            fill_survey_info(field_survey, workbook)
            used_uuids = get_and_fill_detail_info(field_survey, workbook)
            get_and_fill_summary_info(field_survey, workbook, used_uuids)
            if field_survey.manual is True:
                del workbook["Detalhes da Avaliação"]
            workbook.save(return_file_path(number))

            filename = os.listdir(new_folder)[0]

            with open(new_folder + filename, "rb") as excel_file:
                field_export.exported_file.save(
                    filename, ContentFile(excel_file.read())
                )
            os.remove(return_file_path(number))
            os.rmdir(new_folder)

            error = False
        except Exception as e:
            error = True
            logging.error(
                "Untreated exception found while exporting file. Check Sentry."
            )
            sentry_sdk.capture_exception(e)
            pass

        finally:
            field_export.error = error
            field_export.done = True
            field_export.save()
