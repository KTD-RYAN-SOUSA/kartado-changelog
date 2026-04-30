import shutil
import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Firm, Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import get_form_data, new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_reporting_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_custom_option,
    get_end_km,
    get_km,
    get_previous_campaign_report,
)
from helpers.strings import clean_latin_string


class XlsxHandler:
    def __init__(
        self,
        s3,
        list_reporting: List[Reporting],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.__sheet_target = sheet_target
        self.__list_reporting = list_reporting
        self.__xlsx_file = (
            "./fixtures/reports/ccr_report_road_markings_horizontal_signage.xlsx"
        )
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet: Worksheet = self._workbook["Marcas Viarias"]

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="P1:Q5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B5",
            resize_method=ResizeMethod.ProportionalLeft,
        )

    @classmethod
    def __as_float(cls, num):
        try:
            return float(str(num))
        except ValueError:
            return 0.0

    def fill_sheet(self, data_list: list):
        data_dict = {}
        order_dict = {}
        for data in data_list:
            if data["road_name"] not in data_dict.keys():
                data_dict[data["road_name"]] = [data]
                order_dict[data["road_name"]] = [data["direction"]]
            else:
                data_dict[data["road_name"]].append(data)
                order_dict[data["road_name"]].append(data["direction"])
        for k, v in order_dict.items():
            __direction = list(set(v))
            if len(__direction) == 1:
                if str(__direction[0]).lower() == "sul":
                    new_list = sorted(
                        data_dict[k], key=lambda x: x["km"], reverse=False
                    )
                    data_dict[k] = new_list
                elif str(__direction[0]).lower() == "norte":
                    new_list = sorted(data_dict[k], key=lambda x: x["km"], reverse=True)
                    data_dict[k] = new_list
                else:
                    new_list = sorted(
                        data_dict[k], key=lambda x: (x["direction"], x["km"])
                    )
                    data_dict[k] = new_list
            else:
                new_list = sorted(data_dict[k], key=lambda x: (x["direction"], x["km"]))
                data_dict[k] = new_list

        bold = Font(bold=True)

        files = []
        for road, datalist in data_dict.items():
            init_row = 8
            lanes = []
            direction = []
            team_list = []
            subcompany_list = []
            company_list = []
            dates = []
            for data in datalist:
                team_list.append(data["team"])
                subcompany_list.append(data["subcompany"])
                company_list.append(data["company"])
                dates.append(data["found_at"])
                if data["lane"] not in lanes:
                    lanes.append(data["lane"])
                if data["direction"] not in direction:
                    direction.append(data["direction"])

                len_retro = max(len(data["surveys"]), 1)
                end_row = len_retro + init_row
                for col in ["A", "B", "C", "D", "L", "M", "N", "O", "P", "Q"]:
                    self._worksheet.merge_cells(f"{col}{init_row}:{col}{end_row-1}")
                    self._worksheet[f"{col}{init_row}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                self._worksheet[f"A{init_row}"].font = bold
                self._worksheet[f"A{init_row}"].value = data["km"]
                self._worksheet[f"B{init_row}"].value = data["lane"]
                self._worksheet[f"C{init_row}"].value = data["lat"]
                self._worksheet[f"D{init_row}"].value = data["long"]
                average_pitogram = XlsxHandler.__as_float(num=data["average_pitogram"])
                residual_value = XlsxHandler.__as_float(num=data["residual_value"])
                average_pitogram_last_year = XlsxHandler.__as_float(
                    num=data["average_pitogram_last_year"]
                )
                minimal_value_general = XlsxHandler.__as_float(
                    num=data["minimal_value_general"]
                )
                if average_pitogram:
                    self._worksheet[f"L{init_row}"].value = f"{average_pitogram:.2f}"
                if residual_value:
                    self._worksheet[f"M{init_row}"].value = data["residual_value"]
                if average_pitogram_last_year:
                    self._worksheet[
                        f"N{init_row}"
                    ].value = f"{average_pitogram_last_year:.2f}"
                self._worksheet[f"O{init_row}"].value = data["minimal_value_general"]
                self._worksheet[f"P{init_row}"].value = data["found_at"]
                self._worksheet[f"Q{init_row}"].value = data["notes"]
                all_cols = [chr(i) for i in range(ord("A"), ord("R"))]
                for row in range(init_row, end_row):
                    self._worksheet.row_dimensions[row].height = 69.75
                    for col in all_cols:
                        self._worksheet[f"{col}{row}"].border = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000"),
                        )
                if average_pitogram and minimal_value_general:
                    if average_pitogram < minimal_value_general:
                        self._worksheet[f"L{init_row}"].font = bold
                        self._worksheet[f"L{init_row}"].font = Font(
                            color="FF0000", bold=True
                        )
                if residual_value:
                    if residual_value < 0:
                        self._worksheet[f"M{init_row}"].font = bold
                        self._worksheet[f"M{init_row}"].font = Font(
                            color="FF0000", bold=True
                        )

                for retro in data["surveys"]:
                    self._worksheet[f"E{init_row}"].value = retro["code_type"]
                    self._worksheet[f"F{init_row}"].value = retro["color"]
                    if retro["image"]:
                        range_str = f"G{init_row}"
                        insert_picture_2(
                            self._worksheet,
                            range_str,
                            Image(retro["image"]),
                            self.__sheet_target,
                            (1, 1, 1, 1),
                            ResizeMethod.ProportionalCentered,
                        )
                    self._worksheet[f"H{init_row}"].value = retro["first"]
                    self._worksheet[f"I{init_row}"].value = retro["second"]
                    self._worksheet[f"J{init_row}"].value = retro["third"]
                    self._worksheet[f"K{init_row}"].value = retro["average"]
                    average = XlsxHandler.__as_float(num=retro["average"])
                    lifespan = XlsxHandler.__as_float(num=retro["lifespan"])
                    if average and lifespan:
                        if average < lifespan:
                            self._worksheet[f"K{init_row}"].font = Font(
                                color="FF0000", bold=True
                            )
                    for col in ["E", "F", "H", "I", "J", "K"]:
                        self._worksheet[f"{col}{init_row}"].alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    init_row += 1
                init_row = end_row

            team_list = list(set(team_list))
            query_set_teams = Firm.objects.filter(uuid__in=team_list).all()
            names = []
            for team in query_set_teams:
                users_query_set = team.users.all()
                intern_list = []
                for user in users_query_set:
                    intern_list.append(user.get_full_name())
                names.append(intern_list)
            names = [", ".join(_) for _ in names]
            team_text = f"Operador: {' - '.join(names)}"
            subcompany_text = f"Empresa: {' - '.join(list(set(subcompany_list)))}"
            company_text = f"Concessionária: {' - '.join(list(set(company_list)))}"
            road_text = f"Rodovia: {data['road_name']}"
            reporting_text = f"Monitoração: Marcas Viárias -  {' - '.join(list(set(lanes)))} {' - '.join(list(set(direction)))}"
            if len(dates) > 1:
                dates_new = sorted(
                    dates,
                    key=lambda x: (x.split("/")[2], x.split("/")[1], x.split("/")[0]),
                )
                date_text = f"Data: {dates_new[0]} à {dates_new[-1]}"
            elif dates:
                date_text = f"Data: {dates[0]}"
            else:
                date_text = "Data: "
            self._worksheet["E2"] = subcompany_text
            self._worksheet["E3"] = team_text
            self._worksheet["E4"] = date_text
            self._worksheet["K2"] = road_text
            self._worksheet["K3"] = reporting_text
            self._worksheet["K4"] = company_text
            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                target=self.__sheet_target,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
            )
            result = f"Fichas de Marcas Viárias {road} {'-'.join(sorted(list(set((lanes)))))} {'-'.join(sorted(list(set((direction)))))}"
            result = clean_latin_string(result.replace(".", "").replace("/", ""))
            result = f"/tmp/{result}.xlsx"
            files.append(result)
            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook["Marcas Viarias"]
        return {"files": files}

    def create_dict(self, reporting: Reporting, s3) -> dict:
        km = get_km(reporting)
        km_end = get_end_km(reporting)
        direction = get_custom_option(reporting, "direction", "Pista")
        lat = get_form_data(reporting, "latitudedecimal", "latitudedecimal")
        long = get_form_data(reporting, "longitudedecimal", "longitudedecimal")
        average_pitogram = get_form_data(
            reporting, "average_pitogram", "averagePitogram"
        )
        average_pitogram_last_year = 0
        previous_reporting = get_previous_campaign_report(
            reporting.occurrence_type, reporting, "form_data"
        )
        if previous_reporting is not None:
            average_pitogram_last_year = get_form_data(
                previous_reporting, "average_pitogram", "averagePitogram"
            )

        minimal_value_general = get_form_data(
            reporting, "minimal_value_general", "minimalValueGeneral"
        )
        found_at = reporting.found_at.strftime("%d/%m/%Y")
        notes = get_form_data(reporting, "notes", "notes")
        reporting_files = download_reporting_pictures(
            s3=s3,
            path=self.temp_file,
            reporting=reporting,
            width=337,
            height=242,
            enable_include_dnit=False,
            enable_is_shared_antt=True,
        )
        residual_value = ""
        if average_pitogram and minimal_value_general:
            residual_value = int(
                ((average_pitogram - minimal_value_general) / minimal_value_general)
                * 100.0
            )
        surveys = []
        retros = new_get_form_data(reporting, "retro")
        if retros:
            for retro in retros:
                code_type = [
                    x for x in [retro.get("type"), retro.get("description")] if x
                ]
                color = get_form_data(
                    reporting,
                    data_name="color",
                    subgroup="retro",
                    field_name="color",
                    value=retro.get("color"),
                )
                img = ""
                try:
                    img = reporting_files["images"][0]["path"]
                except Exception:
                    pass

                average = ""
                first = retro.get("first", None)
                second = retro.get("second", None)
                third = retro.get("third", None)
                try:
                    average = (float(first) + float(second) + float(third)) / 3
                except Exception:
                    pass

                surveys.append(
                    {
                        "code_type": " - ".join(code_type),
                        "color": color,
                        "lifespan": retro.get("lifespan"),
                        "first": first,
                        "second": second,
                        "third": third,
                        "average": average,
                        "image": img,
                    }
                )

        data = {
            "road_name": reporting.road_name,
            "reporting": reporting,
            "lane": get_custom_option(reporting, "lane", "Faixa"),
            "subcompany": reporting.firm.subcompany.__dict__.get("name"),
            "company": reporting.company.__dict__.get("name"),
            "team": str(reporting.firm.__dict__.get("uuid")),
            "km": km,
            "km_end": km_end,
            "direction": direction,
            "lat": lat,
            "long": long,
            "average_pitogram": average_pitogram,
            "average_pitogram_last_year": average_pitogram_last_year,
            "minimal_value_general": minimal_value_general,
            "found_at": found_at,
            "notes": notes,
            "residual_value": residual_value,
            "surveys": surveys,
        }
        for k, v in data.items():
            if v is None and k not in ["station_one", "station_two", "station_three"]:
                data[k] = ""
        return data

    def execute(self):
        data = []
        for reporting in self.__list_reporting:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRRoadMarkingsHorizontalSignage(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""
        road_names = []
        directions = []
        lanes = []
        reportings = Reporting.objects.filter(uuid__in=self.uuids).prefetch_related(
            "occurrence_type", "firm", "firm__subcompany", "company"
        )
        for reporting in reportings:
            lane = ""
            lane = get_custom_option(reporting, "lane", "Faixa")
            direction = get_custom_option(reporting, "direction", "Pista")

            if reporting.road_name not in road_names:
                road_names.append(reporting.road_name)
            if lane not in lanes:
                lanes.append(lane)
            if direction not in directions:
                directions.append(direction)

        file_name = f"Fichas de Marcas Viárias {'-'.join(sorted(list(set(road_names))))} {'-'.join(sorted(list(set((lanes)))))} {'-'.join(sorted(list(set((directions)))))}"
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        if len(road_names) > 1:
            file_name = f"{file_name}.zip"
        else:
            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{file_name}.{extension}"

        return file_name

    def __get_repotings_obj(self):
        form = OccurrenceType.objects.get(name="Retrorrefletância de Marcas Viárias")
        query_set = Reporting.objects.filter(
            occurrence_type=form, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def export(self):
        list_reporting = self.__get_repotings_obj()
        s3 = get_s3()
        obj = XlsxHandler(s3, list_reporting, self.sheet_target()).execute()
        files = obj["files"]
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_road_markings_horizontal_signage_async_handler(reporter_dict: dict):
    reporter = CCRRoadMarkingsHorizontalSignage.from_dict(reporter_dict)
    reporter.export()
