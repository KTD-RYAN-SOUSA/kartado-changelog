import datetime
import io
import logging
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from tempfile import NamedTemporaryFile

import requests
from django.contrib.gis.db.models import IntegerField
from django.core.files.base import ContentFile
from django.db.models import Count, F, Func, OuterRef, Prefetch, Q, Subquery
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils.cell import get_column_letter
from PIL import Image as PILImage
from PIL import UnidentifiedImageError
from sentry_sdk import capture_exception
from storages.utils import clean_name
from zappa.asynchronous import task

from apps.companies.models import Company, Firm
from apps.occurrence_records.models import OccurrenceType, RecordPanel
from apps.reportings.models import HistoricalReporting, Reporting, ReportingFile
from apps.reportings.serializers import LightReportingSerializer
from apps.reportings.views import ReportingFilter
from apps.templates.models import ReportingExport
from apps.users.models import User
from apps.work_plans.models import Job
from helpers.apps.daily_reports import (
    get_exporter_extra_columns,
    get_exporter_extra_columns_parsed_infos,
    get_fields_to_hide_reporting_location,
    get_updated_by,
    get_values_from_reporting_extra_columns,
    remove_fields_to_hide_reporting_location,
    translate_reporting_value,
)
from helpers.apps.record_panel import handle_field_name
from helpers.dates import utc_to_local
from helpers.permissions import PermissionManager, join_queryset
from helpers.strings import (
    clean_invalid_characters,
    get_obj_from_path,
    to_camel_case,
    to_snake_case,
)

MAX_REPORTING_FILES = 100000
THREADING_LIMIT = 30


class ReportingFileCountExceededException(Exception):
    pass


def get_reporting_queryset(user_company, user, permissions):
    queryset = None

    allowed_queryset = permissions.get_allowed_queryset()

    if "none" in allowed_queryset:
        queryset = join_queryset(queryset, Reporting.objects.none())
    if (
        "self" in allowed_queryset
        or "firm" in allowed_queryset
        or "self_and_created_by_firm" in allowed_queryset
    ):
        user_firms = list(
            (user.user_firms.filter(company_id=user_company)).union(
                user.user_firms_manager.filter(company_id=user_company)
            )
        )
        jobs = (
            Job.objects.filter(
                Q(company_id=user_company)
                & (
                    Q(worker=user)
                    | Q(created_by=user)
                    | Q(watcher_users=user)
                    | Q(firm__in=user_firms)
                    | Q(watcher_firms__in=user_firms)
                    | Q(watcher_subcompanies__subcompany_firms__in=user_firms)
                )
            )
            .distinct()
            .values_list("uuid", flat=True)
        )
        if "self" in allowed_queryset:
            queryset = join_queryset(
                queryset,
                Reporting.objects.filter(Q(created_by=user) | Q(job__in=jobs)),
            )
        if "firm" in allowed_queryset or "self_and_created_by_firm" in allowed_queryset:
            # Get users related to the request user's firms
            related_users = User.objects.filter(user_firms__in=user_firms).distinct()

            if "firm" in allowed_queryset:
                created_by_sub = (
                    User.objects.filter(reportings=OuterRef("uuid"))
                    .order_by()
                    .annotate(
                        user_firm_count=Func(F("user_userinfirm"), function="Count")
                    )
                    .values("user_firm_count")
                )
                queryset = join_queryset(
                    queryset,
                    (
                        Reporting.objects.annotate(
                            user_firms_count=Subquery(
                                created_by_sub, output_field=IntegerField()
                            )
                        )
                        .filter(
                            Q(company_id=user_company)
                            & (
                                Q(firm__in=user_firms)
                                | (
                                    Q(created_by__in=related_users)
                                    # & Q(user_firms_count__lte=1)
                                )
                                | Q(job__in=jobs)
                            )
                        )
                        .exclude(
                            ~Q(firm__in=user_firms)
                            & ~Q(job__in=jobs)
                            & Q(created_by__in=related_users)
                            & Q(user_firms_count__gt=1)
                        )
                    ),
                )

            if "self_and_created_by_firm" in allowed_queryset:
                # Expand user_firms to include firms created by related users
                related_firms = Firm.objects.filter(
                    created_by__in=related_users
                ).distinct()
                user_firms.extend(related_firms)
                user_firms = list(set(user_firms))

                # Get users of new related firms and add them to related_users
                related_firm_users = User.objects.filter(
                    user_firms__in=related_firms
                ).distinct()
                related_users = (related_users | related_firm_users).distinct()

                queryset = join_queryset(
                    queryset,
                    (
                        Reporting.objects.filter(
                            Q(company_id=user_company)
                            & (
                                Q(firm__in=user_firms)
                                | Q(created_by__in=related_users)
                                | Q(job__in=jobs)
                            )
                        )
                    ),
                )
    if "artesp" in allowed_queryset:
        queryset = join_queryset(
            queryset,
            Reporting.objects.filter(
                company_id=user_company,
                form_data__artesp_code__isnull=False,
            ).exclude(form_data__artesp_code__exact=""),
        )
    if "artesp_entrevias" in allowed_queryset:
        try:
            company = Company.objects.get(pk=user_company)
        except Exception:
            queryset = join_queryset(queryset, Reporting.objects.none())
        else:
            queryset_company = Reporting.objects.filter(
                company_id=user_company
            ).distinct()

            possible_path_kinds = "artesp_exclude__occurrence_kind"
            kinds = get_obj_from_path(company.metadata, possible_path_kinds)

            possible_path_firms = "artesp_exclude__historical_firm"
            firms = get_obj_from_path(company.metadata, possible_path_firms)

            if kinds and isinstance(kinds, list) and firms and isinstance(firms, list):
                histories = HistoricalReporting.objects.filter(
                    history_type="+", firm__in=firms
                )
                queryset = join_queryset(
                    queryset,
                    queryset_company.filter(found_at__gte="2020-01-01").exclude(
                        (
                            Q(occurrence_type__occurrence_kind__in=kinds)
                            | Q(historicalreporting__in=histories)
                        )
                        & (
                            Q(form_data__artesp_code__isnull=True)
                            | Q(form_data__artesp_code__exact="")
                        )
                    ),
                )
            else:
                queryset = join_queryset(queryset, Reporting.objects.none())

    if "antt_supervisor_agency" in allowed_queryset:
        queryset = join_queryset(
            queryset,
            Reporting.objects.filter(company=user_company, shared_with_agency=True),
        )
    if "supervisor_agency" in allowed_queryset:
        queryset = join_queryset(
            queryset,
            Reporting.objects.filter(
                Q(company_id=user_company)
                & (
                    Q(reporting_construction_progresses__construction__origin="AGENCY")
                    | (
                        Q(form_data__artesp_code__isnull=False)
                        & ~Q(form_data__artesp_code__exact="")
                    )
                )
            ),
        )
    if "all" in allowed_queryset:
        queryset = join_queryset(
            queryset, Reporting.objects.filter(company_id=user_company)
        )

    # If queryset isn't set by any means above
    if queryset is None:
        user_companies = user.companies.all()
        queryset = Reporting.objects.filter(company__in=user_companies)

    return queryset.exclude(occurrence_type__occurrence_kind="2")


def get_inventory_queryset(user_company, user, permissions):
    queryset = None

    allowed_queryset = permissions.get_allowed_queryset()

    if "none" in allowed_queryset:
        queryset = join_queryset(queryset, Reporting.objects.none())
    if "self" in allowed_queryset:
        queryset = join_queryset(
            queryset,
            Reporting.objects.filter(created_by=user),
        )
    if "all" in allowed_queryset:
        queryset = join_queryset(
            queryset, Reporting.objects.filter(company_id=user_company)
        )

    # If queryset isn't set by any means above
    if queryset is None:
        user_companies = user.companies.all()
        queryset = Reporting.objects.filter(company__in=user_companies)

    return queryset.filter(occurrence_type__occurrence_kind="2").distinct()


def as_text(value):
    return str(value) if value is not None else ""


@task
def generate_reporting_export(reporting_export_id):
    logging.info(
        f"[ReportingExport] generate_reporting_export iniciado | id={reporting_export_id} | fargate={bool(os.environ.get('ECS_CONTAINER_METADATA_URI_V4'))}"
    )

    try:
        reporting_export = ReportingExport.objects.get(pk=reporting_export_id)
    except ReportingExport.DoesNotExist as e:
        logging.error("ReportingExport not found")
        capture_exception(e)
    else:
        error = True  # Error until proven otherwise
        start_time = time.time()
        _fargate_counts = {"item_count": None, "image_count": None}

        extra_info = reporting_export.extra_info
        export_type = reporting_export.export_type
        is_inventory = reporting_export.is_inventory
        filters = reporting_export.filters
        created_by = reporting_export.created_by
        company = reporting_export.company

        export_resources = extra_info.get("export_resources", False)
        export_photos = extra_info.get("export_photos", False)
        export_kind = extra_info.get("export_kind", False)
        export_date = extra_info.get("export_date", False)
        export_description = extra_info.get("export_description", False)
        photo_order = extra_info.get("photo_order", "uploaded_at")
        include_city = extra_info.get("include_city", False)

        basic_columns = {
            "number": "Serial",
            "parent_number": "Serial Inventário Vinculado",
            "road": "Rodovia",
            "cityCalc": "Município",
            "km": "km inicial",
            "end_km": "km final",
            "lot": "Lote",
            "latitude": "Latitude",
            "longitude": "Longitude",
            "occurrence_kind": "Natureza",
            "occurrence_type": "Classe",
            "length": "Comprimento",
            "width": "Largura",
            "height": "Espessura",
            "lane": "Faixa",
            "track": "Pista",
            "km_reference": "km de referência",
            "branch": "Ramo",
            "direction": "Sentido",
            "status": "Status",
            "created_by": "Criado por",
            "updated_by": "Atualizado por",
            "subcompany": "Empresa",
            "firm": "Equipe",
            "job": "Programação",
            "job_start_date": "Início da programação",
            "job_end_date": "Fim da programação",
            "created_at": "Criado em",
            "found_at": "Encontrado em",
            "updated_at": "Atualizado em",
            "executed_at": "Executado em",
            "notes": "Observações",
            "due_at": "Prazo",
        }

        HEADER_ROW = 1
        COLUMN_START_POSITION = 2
        NUMBER_FORMAT = "0.0000"
        DATE_FORMAT = "dd/mm/yyyy"
        PRICE_FORMAT = "R$ #,##0.00"
        RESOURCE_PATTERN = re.compile(r"^Recurso_\d+$")
        PHOTO_PATTERN = re.compile(r"^Foto \d+$")
        PHOTO_DATA_PATTERN = re.compile(r"Foto \d+ - \w+")
        ARRAY_PHOTO_PATTERN = re.compile(r"^[\w]+#[\d]+#[\w]+#[\w]+$")
        ARRAY_PHOTO_CONTENT_PATTERN = re.compile(r"^[\w ]+[\d]+: [\w ]+[\d]+")
        ARRAY_PHOTO_DATA_PATTERN = re.compile(r"^[\w ]+[\d]+: [\w ]+[\d]+ - [\w]+$")
        PHOTO_COLUMN_WIDTH = 30

        simple_style = NamedStyle(name="simple_style")
        simple_style.border = Border(
            right=Side(border_style="dashed", color="000000"),
            left=Side(border_style="dashed", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        simple_style.alignment = Alignment(
            vertical="center", wrap_text=True, wrapText=True
        )
        normal_style = NamedStyle(name="normal_style")
        normal_style.border = Border(
            right=Side(border_style="dashed", color="000000"),
            left=Side(border_style="dashed", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        normal_style.alignment = Alignment(vertical="center")
        last_row_style = NamedStyle(name="last_row_style")
        last_row_style.border = Border(
            right=Side(border_style="dashed", color="000000"),
            left=Side(border_style="dashed", color="000000"),
            bottom=Side(border_style="medium", color="000000"),
        )
        last_row_style.alignment = Alignment(
            vertical="center", wrapText=True, wrap_text=True
        )

        add_total_style = NamedStyle(name="add_total_style")
        add_total_style.border = Border(
            top=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000"),
        )

        rf_thread_pool = ThreadPoolExecutor(max_workers=THREADING_LIMIT)
        try:

            def config_static_columns(basic_columns):
                if not include_city:
                    basic_columns.pop("cityCalc", None)
                if not has_lot:
                    basic_columns.pop("lot", None)
                if not has_track:
                    basic_columns.pop("track", None)
                    basic_columns.pop("branch", None)
                    basic_columns.pop("km_reference", None)
                if not can_view_subcompany:
                    basic_columns.pop("subcompany", None)
                if not can_view_deadline:
                    basic_columns.pop("due_at", None)
                if not can_view_inventory:
                    basic_columns.pop("parent_number", None)
                if is_inventory:
                    basic_columns.pop("job", None)
                    basic_columns.pop("job_start_date", None)
                    basic_columns.pop("job_end_date", None)
                    basic_columns.pop("updated_by", None)
                    basic_columns.pop("executed_at", None)
                    basic_columns.pop("due_at", None)
                    basic_columns.pop("firm", None)
                    basic_columns.pop("subcompany", None)
                    basic_columns.pop("status", None)
                    basic_columns.pop("parent_number", None)

                if hide_reporting_location:
                    fields_to_hide = get_fields_to_hide_reporting_location()
                    basic_columns = remove_fields_to_hide_reporting_location(
                        fields_to_hide, basic_columns
                    )

                return basic_columns

            def fill_header_cell(pos, value):
                cell_name = cell_by_pos(pos, HEADER_ROW)
                templ_ws[cell_name] = clean_invalid_characters(value)
                templ_ws[cell_name]._style = copy(header_style)

            def cell_by_pos(column_pos, row_pos):
                column_letter = get_column_letter(column_pos)
                return cell(column_letter, row_pos)

            def cell(column_letter, row_pos):
                return "{}{}".format(column_letter, row_pos)

            def get_reporting_data(reporting, exporter_extra_columns=[]):
                reporting_data = {
                    "uuid": str(reporting.pk),
                    "number": reporting.number,
                    "road": reporting.road_name,
                    "cityCalc": reporting.city or "",
                    "km": reporting.km,
                    "end_km": (
                        reporting.end_km if reporting.end_km is not None else None
                    ),
                    "lot": (
                        translate_reporting_value(company, "lot", reporting.lot)
                        if reporting.lot
                        else None
                    ),
                    "longitude": reporting.point.coords[0] if reporting.point else None,
                    "latitude": reporting.point.coords[1] if reporting.point else None,
                    "occurrence_kind": next(
                        (
                            a.get("name", "")
                            for a in occurrence_kind_values
                            if reporting.occurrence_type
                            and a.get("value", "")
                            == reporting.occurrence_type.occurrence_kind
                        ),
                        "",
                    ),
                    "occurrence_type": (
                        reporting.occurrence_type.name
                        if reporting.occurrence_type
                        else None
                    ),
                    "length": reporting.form_data.get("length", 0),
                    "width": reporting.form_data.get("width", 0),
                    "height": reporting.form_data.get("height", 0),
                    "lane": translate_reporting_value(company, "lane", reporting.lane),
                    "track": translate_reporting_value(
                        company, "track", reporting.track
                    ),
                    "km_reference": reporting.km_reference or 0,
                    "branch": translate_reporting_value(
                        company, "branch", reporting.branch
                    ),
                    "direction": translate_reporting_value(
                        company, "direction", reporting.direction
                    ),
                    "status": reporting.status.name if reporting.status else None,
                    "created_by": (
                        reporting.created_by.get_full_name()
                        if reporting.created_by
                        else None
                    ),
                    "updated_by": get_updated_by(reporting),
                    "subcompany": (
                        reporting.firm.subcompany.name
                        if reporting.firm and reporting.firm.subcompany
                        else None
                    ),
                    "firm": reporting.firm.name if reporting.firm else None,
                    "job": reporting.job.title if reporting.job else None,
                    "job_start_date": (
                        utc_to_local(reporting.job.start_date)
                        if reporting.job
                        else None
                    ),
                    "job_end_date": (
                        utc_to_local(reporting.job.end_date)
                        if reporting.job and reporting.job.end_date
                        else None
                    ),
                    "created_at": utc_to_local(reporting.created_at),
                    "found_at": utc_to_local(reporting.found_at),
                    "updated_at": utc_to_local(reporting.updated_at),
                    "executed_at": (
                        utc_to_local(reporting.executed_at)
                        if reporting.executed_at
                        else None
                    ),
                    "notes": reporting.form_data.get("notes", ""),
                    "due_at": (
                        utc_to_local(reporting.due_at) if reporting.due_at else None
                    ),
                    "parent_number": (
                        reporting.parent.number if reporting.parent else None
                    ),
                    "extra_columns": get_values_from_reporting_extra_columns(
                        reporting, exporter_extra_columns, reference_values
                    ),
                }
                return reporting_data

            def fill_cell_reporting_section(
                pos, row_pos, value, number_format=None, style=None
            ):
                cell_name = cell_by_pos(pos, row_pos)
                templ_ws[cell_name] = (
                    str(clean_invalid_characters(value))
                    if isinstance(value, list)
                    else clean_invalid_characters(value)
                )
                if number_format:
                    templ_ws[cell_name].number_format = number_format
                if style:
                    templ_ws[cell_name].style = style

            def get_resource_data(reporting):
                resources = {}
                if export_resources and (can_view_resources or can_view_money):
                    for (
                        procedure_resource
                    ) in reporting.reporting_resources.all().prefetch_related(
                        "resource"
                    ):
                        resource_name = procedure_resource.resource.name
                        total_price = procedure_resource.total_price

                        if can_view_money:
                            if "total" in resources:
                                resources["total"] += total_price
                            else:
                                resources["total"] = total_price

                        if "proc_resources" not in resources:
                            resources["proc_resources"] = []
                        if can_view_resources:
                            resources["proc_resources"].append(
                                {
                                    "resource_name": resource_name,
                                    "resource_unit": procedure_resource.resource.unit,
                                    "amount": procedure_resource.amount,
                                    "unit_price": procedure_resource.unit_price,
                                    "total_price": total_price,
                                }
                            )
                return resources

            def add_resource_headers():
                try:
                    max_num_resources = max(
                        [
                            len(reporting_data["resources"]["proc_resources"])
                            for reporting_data in reportings_data
                            if reporting_data["resources"]
                        ]
                    )
                except Exception:
                    max_num_resources = 0

                new_column_pos = COLUMN_START_POSITION + len(columns_order)
                if can_view_money:
                    fill_header_cell(new_column_pos, "Valor total")
                    new_column_pos += 1

                if can_view_resources:
                    resource_fields = [
                        "Recurso",
                        "Quantidade",
                        "Valor",
                        "Valor Unitário",
                    ]
                    for i in range(1, max_num_resources + 1):
                        for field in resource_fields:
                            fill_header_cell(new_column_pos, "{}_{}".format(field, i))
                            new_column_pos += 1

            def translate_photo_value(company, field, value):
                options = get_obj_from_path(
                    company.custom_options,
                    "reporting_file__fields__{}__selectoptions__options".format(field),
                )
                try:
                    return next(a["name"] for a in options if a["value"] == value)
                except Exception:
                    return ""

            def download_picture_to_temp_file(upload_url):
                try:
                    resp = requests.get(upload_url, stream=True)
                    content = resp.content if resp.status_code == 200 else None
                    if not content:
                        return None

                    try:
                        PILImage.open(io.BytesIO(content)).verify()
                    except (UnidentifiedImageError, Exception):
                        logging.warning(
                            f"[ReportingExport] Imagem inválida ou corrompida ignorada: {upload_url}"
                        )
                        return None

                    temp_file = NamedTemporaryFile(delete=False)
                    temp_file.write(content)
                    temp_file.close()
                    return temp_file
                except Exception as e:
                    logging.error(f"Failed to download picture to temp file: {e}")
                    capture_exception(e)
                    return None

            def get_photos_data(reporting, inner_image_array_info=[]):
                if reporting_file_queryset == "none":
                    return []
                inner_uuids = []
                exclude_uuids = []
                if inner_image_array_info:
                    inner_uuids, exclude_uuids = get_inner_images_uuids(
                        reporting, inner_image_array_info
                    )

                reporting_files_data = []
                inner_files_data = []
                reporting_file_objects = reporting.reporting_files.all()

                exclude_uuids = [str(uuid) for uuid in exclude_uuids]
                for reporting_file in reporting_file_objects:
                    if (
                        reporting_file.upload
                        and str(reporting_file.uuid) not in exclude_uuids
                    ):
                        image_type = str(reporting_file.upload).split(".")
                        if image_type[-1].lower() not in ["jpeg", "jpg", "png", "gif"]:
                            continue
                        params = {}
                        params["Bucket"] = "{}-{}px".format(
                            reporting_file.upload.storage.bucket.name, 400
                        )
                        params["Key"] = reporting_file.upload.storage._normalize_name(
                            clean_name(reporting_file.upload.name)
                        )
                        upload_url = reporting_file.upload.storage.bucket.meta.client.generate_presigned_url(
                            "get_object", Params=params, ExpiresIn=3600
                        )

                        # Fetch image data
                        resp_future = rf_thread_pool.submit(
                            download_picture_to_temp_file, upload_url
                        )
                        try:
                            photo_data = {"content": resp_future}
                            if export_kind:
                                photo_data.update(
                                    {
                                        "kind": translate_photo_value(
                                            company, "kind", reporting_file.kind
                                        )
                                    }
                                )
                            if export_date:
                                photo_data.update(
                                    {
                                        "datetime": utc_to_local(
                                            reporting_file.datetime
                                        ).date()
                                    }
                                )
                            if export_description:
                                photo_data.update(
                                    {
                                        "description": getattr(
                                            reporting_file, "description", ""
                                        )
                                    }
                                )
                            reporting_files_data.append(photo_data)
                        except Exception as e:
                            logging.error(
                                "Error fetching photo data for reporting file"
                            )
                            capture_exception(e)

                if inner_image_array_info and inner_uuids:
                    for item in inner_uuids:
                        for k, v in item.items():
                            v = [str(uuid) for uuid in v]
                            image_index = -1
                            for reporting_file in reporting_file_objects:
                                if str(reporting_file.uuid) not in v:
                                    continue
                                image_index += 1
                                if reporting_file.upload:
                                    image_type = str(reporting_file.upload).split(".")
                                    if image_type[-1].lower() not in [
                                        "jpeg",
                                        "jpg",
                                        "png",
                                        "gif",
                                    ]:
                                        continue
                                    params = {}
                                    params["Bucket"] = "{}-{}px".format(
                                        reporting_file.upload.storage.bucket.name, 400
                                    )
                                    params[
                                        "Key"
                                    ] = reporting_file.upload.storage._normalize_name(
                                        clean_name(reporting_file.upload.name)
                                    )
                                    upload_url = reporting_file.upload.storage.bucket.meta.client.generate_presigned_url(
                                        "get_object", Params=params, ExpiresIn=3600
                                    )

                                    # Fetch image data
                                    resp_future = rf_thread_pool.submit(
                                        download_picture_to_temp_file, upload_url
                                    )

                                    try:
                                        key, position, inner_key = k.split("#")
                                        inner_files_data.append(
                                            {
                                                f"{to_camel_case(key)}#{str(position)}#{to_camel_case(inner_key)}{str(image_index)}#content": resp_future
                                            }
                                        )
                                        if export_kind:
                                            inner_files_data.append(
                                                {
                                                    f"{to_camel_case(key)}#{str(position)}#{to_camel_case(inner_key)}{str(image_index)}#kind": translate_photo_value(
                                                        company,
                                                        "kind",
                                                        reporting_file.kind,
                                                    )
                                                }
                                            )

                                        if export_date:
                                            inner_files_data.append(
                                                {
                                                    f"{to_camel_case(key)}#{str(position)}#{to_camel_case(inner_key)}{str(image_index)}#datetime": utc_to_local(
                                                        reporting_file.datetime
                                                    ).date()
                                                }
                                            )
                                        if export_description:
                                            inner_files_data.append(
                                                {
                                                    f"{to_camel_case(key)}#{str(position)}#{to_camel_case(inner_key)}{str(image_index)}#description": getattr(
                                                        reporting_file,
                                                        "description",
                                                        "",
                                                    )
                                                }
                                            )
                                    except Exception as e:
                                        logging.error(
                                            "Error fetching inner photo data for reporting file"
                                        )
                                        capture_exception(e)

                return reporting_files_data, inner_files_data

            def add_photo_headers():
                try:
                    max_num_photos = max(
                        [
                            len(reporting_data["reporting_files"])
                            for reporting_data in reportings_data
                            if "reporting_files" in reporting_data
                        ]
                    )
                except Exception:
                    max_num_photos = 0

                new_column_pos = templ_ws.max_column + 1

                resource_fields = ["Foto"]
                if export_kind:
                    resource_fields.append("tipo")
                if export_date:
                    resource_fields.append("data")
                if export_description:
                    resource_fields.append("descrição")
                for i in range(1, max_num_photos + 1):
                    for index, field in enumerate(resource_fields):
                        if index == 0:
                            fill_header_cell(new_column_pos, "{} {}".format(field, i))
                        else:
                            fill_header_cell(
                                new_column_pos, "Foto {} - {}".format(i, field)
                            )
                        new_column_pos += 1

            def remove_columns_normal_export(columns):
                columns.pop("latitude", None)
                columns.pop("longitude", None)
                columns.pop("occurrence_type", None)
                columns.pop("length", None)
                columns.pop("width", None)
                columns.pop("height", None)
                columns.pop("track", None)
                columns.pop("km_reference", None)
                columns.pop("branch", None)
                columns.pop("updated_by", None)
                columns.pop("job", None)
                columns.pop("job_start_date", None)
                columns.pop("job_end_date", None)
                columns.pop("notes", None)
                columns.pop("due_at", None)

                if not can_view_inventory:
                    basic_columns.pop("parent_number", None)

                if is_inventory:
                    basic_columns.pop("executed_at", None)
                    basic_columns.pop("firm", None)
                    basic_columns.pop("subcompany", None)
                    basic_columns.pop("status", None)
                    basic_columns.pop("parent_number", None)

                return columns

            def get_add_total_columns(measurement_columns):
                add_total_columns = []
                for extra_column in measurement_columns:
                    is_array = extra_column.get("isArray", False)
                    if not is_array:
                        add_total = extra_column.get("addTotal", False)
                        if add_total:
                            add_total_columns.append(extra_column.get("header", ""))
                    else:
                        add_total_columns_array = get_add_total_columns_array(
                            extra_column
                        )
                        if add_total_columns_array:
                            add_total_columns.extend(add_total_columns_array)

                return add_total_columns

            def get_add_total_columns_array(extra_column):
                add_total_columns_array = []
                header = extra_column.get("header", "")
                max_repetitions = extra_column.get("maxRepetitions", 5)
                fields = extra_column.get("fields", [])

                for i in range(0, max_repetitions):
                    for item in fields:
                        add_total = item.get("addTotal", False)
                        if add_total:
                            inner_header = item.get("header", "")
                            add_total_columns_array.append(
                                f"{header} {str(i+1)}: {inner_header}"
                            )
                return add_total_columns_array

            def get_inner_image_array_data(measurement_columns):
                inner_image_info = []
                for item in measurement_columns:
                    header = item.get("header", "")
                    key = item.get("key", "")
                    is_array = item.get("isArray", False)
                    max_repetitions = item.get("maxRepetitions", 5)
                    if is_array:
                        fields = item.get("fields", [])
                        for array_item in fields:
                            is_image = array_item.get("isImage", False)
                            if is_image:
                                inner_key = array_item.get("field")
                                inner_header = array_item.get("header", "")
                                repeat_in_the_end = array_item.get(
                                    "repeatInTheEnd", False
                                )
                                inner_image_info.append(
                                    {
                                        f"{to_snake_case(key)}.{to_snake_case(inner_key)}": {
                                            "header": f"{header}: {inner_header}",
                                            "repeat_in_the_end": repeat_in_the_end,
                                            "max_repetitions": max_repetitions,
                                        }
                                    }
                                )
                return inner_image_info

            def get_inner_images_uuids(reporting, inner_image_array_info):
                inner_uuids = []
                exclude_uuids = []
                form_data = reporting.form_data
                for item in inner_image_array_info:
                    for k, v in item.items():
                        key, inner_key = k.split(".")
                        repeat_in_the_end = v.get("repeat_in_the_end")
                        if key in form_data and form_data[key]:
                            inner_data = form_data[key]
                            for index, data in enumerate(inner_data):
                                if inner_key in data and data[inner_key]:
                                    inner_uuids.append(
                                        {
                                            f"{key}#{str(index)}#{inner_key}": data[
                                                inner_key
                                            ]
                                        }
                                    )
                                    if not repeat_in_the_end:
                                        exclude_uuids.extend(data[inner_key])
                                else:
                                    continue
                        else:
                            continue
                return inner_uuids, exclude_uuids

            def get_prefetched_reporting_files_queryset(qs):
                if export_photos:
                    rf_queryset = ReportingFile.objects.all()
                    if reporting_file_queryset == "self":
                        rf_queryset = ReportingFile.objects.filter(
                            Q(created_by=created_by)
                            | Q(reporting__created_by=created_by)
                        )
                    elif reporting_file_queryset == "antt_supervisor_agency":
                        rf_queryset = ReportingFile.objects.filter(
                            # Only consider shared approval steps
                            Q(reporting__approval_step__in=shared_approval_steps)
                            # Check if reporting is being shared
                            & Q(reporting__shared_with_agency=True)
                            # Check if reporting file is being shared
                            & Q(is_shared=True)
                        )
                    elif reporting_file_queryset == "none":
                        rf_queryset = None

                    if rf_queryset is not None:
                        rf_queryset = rf_queryset.order_by(photo_order, "uploaded_at")
                        qs = qs.prefetch_related(
                            Prefetch(lookup="reporting_files", queryset=rf_queryset)
                        )

                        qs = qs.annotate(reporting_files_count=Count("reporting_files"))
                        total_reporting_files = sum(r.reporting_files_count for r in qs)

                        logging.info(
                            f"reporting_export::generate_reporting_export::get_prefetched_reporting_files_queryset::total_reporting_files: {total_reporting_files}"
                        )
                        _fargate_counts["item_count"] = len(qs)
                        _fargate_counts["image_count"] = total_reporting_files
                        if total_reporting_files > MAX_REPORTING_FILES:
                            raise ReportingFileCountExceededException()
                return qs

            # occurrence_kind list

            occurrence_kind_values = get_obj_from_path(
                company.custom_options,
                "reporting__fields__occurrence_kind__selectoptions__options",
            )

            # Pre-cache occs for export
            reference_values = {
                str(a.uuid): a.name
                for a in OccurrenceType.objects.filter(company=company)
            }

            # Get some configs from Company
            hide_reporting_location = (
                get_obj_from_path(company.metadata, "hide_reporting_location") or False
            )
            has_lot = get_obj_from_path(
                company.custom_options, "reporting__fields__lot"
            )
            has_track = get_obj_from_path(company.metadata, "show_track")

            shared_approval_steps = get_obj_from_path(
                company.metadata, "shared_approval_steps"
            )

            # Get permission list for the user
            if is_inventory:
                permissions = PermissionManager(
                    str(company.pk), created_by, "Inventory"
                )
            else:
                permissions = PermissionManager(
                    str(company.pk), created_by, "Reporting"
                )

            # Get permissions for some columns

            can_view_subcompany = permissions.get_specific_model_permision(
                "SubCompany", "can_view"
            )
            can_view_deadline = permissions.get_specific_model_permision(
                "ReportingDeadline", "can_view"
            )
            can_view_money = permissions.get_specific_model_permision(
                "ProcedureResource", "can_view_money"
            )
            can_view_resources = permissions.get_specific_model_permision(
                "ProcedureResource", "can_view"
            ) and permissions.get_specific_model_permision("Resource", "can_view")

            can_view_inventory = permissions.get_specific_model_permision(
                "Inventory", "can_view"
            )

            reporting_file_queryset = permissions.get_specific_model_permision(
                "ReportingFile", "queryset"
            )

            # Build queryset
            if is_inventory:
                reporting_queryset = get_inventory_queryset(
                    company.pk, created_by, permissions
                )
            else:
                reporting_queryset = get_reporting_queryset(
                    company.pk, created_by, permissions
                )

            # Get sort param
            sort = filters.pop("sort", "number")
            if sort == "record_panel":
                if "record_panel" in filters:
                    record_panel = RecordPanel.objects.get(uuid=filters["record_panel"])
                    sort = [
                        (
                            "{}{}".format(
                                "-" if a["order"] == "DESC" else "",
                                handle_field_name(a["field"]),
                            )
                            if "order" in a
                            else handle_field_name(a["field"])
                        )
                        for a in record_panel.list_order_by
                    ]
                else:
                    sort = "number"

            # Filter queryset
            if is_inventory:
                filters.update({"occurrence_kind": "2"})

            filtered_queryset = ReportingFilter(filters, queryset=reporting_queryset).qs

            # Prefetch queryset and sort
            prefetch_related_fields = (
                LightReportingSerializer._PREFETCH_RELATED_FIELDS.copy()
            )
            prefetch_related_fields.remove("reporting_files")
            if type(sort) is list:
                queryset = filtered_queryset.order_by(*sort, "uuid").prefetch_related(
                    *prefetch_related_fields
                )

            else:
                queryset = filtered_queryset.order_by(sort, "uuid").prefetch_related(
                    *prefetch_related_fields
                )

            queryset = get_prefetched_reporting_files_queryset(queryset)

            all_columns = config_static_columns(basic_columns)

            reportings_data = []

            files_to_close = []

            # Fill simplified Excel
            if export_type == "SIMPLE":
                templ_wb = load_workbook(
                    filename="apps/templates/templates/reporting_export_simple_template.xlsx",
                    read_only=False,
                    keep_vba=False,
                )
                exporter_extra_columns = get_exporter_extra_columns(
                    company, is_inventory
                )
                exporter_simple_excel_columns_order = get_obj_from_path(
                    company.custom_options,
                    "{}__exporter__simple_excel_columns_order".format(
                        "inventory" if is_inventory else "reporting"
                    ),
                )
                extra_columns = get_exporter_extra_columns_parsed_infos(
                    exporter_extra_columns
                )
                all_columns = {**all_columns, **extra_columns}
                if exporter_simple_excel_columns_order:
                    columns_order = [
                        column
                        for column in exporter_simple_excel_columns_order
                        if column in all_columns
                    ]
                else:
                    columns_order = [*all_columns.keys()]

                templ_ws = templ_wb["Apontamentos"]

                header_style = templ_ws.cell(column=2, row=HEADER_ROW)._style

                data_row = 2

                # Fill basic headers
                for pos, column_key in enumerate(
                    columns_order, start=COLUMN_START_POSITION
                ):
                    fill_header_cell(pos, all_columns.get(column_key, ""))

                # Collect data
                for reporting in queryset:
                    reporting_data = get_reporting_data(
                        reporting, exporter_extra_columns
                    )

                    # Collect resources data
                    if not is_inventory:
                        resources = get_resource_data(reporting)
                        reporting_data["resources"] = resources

                    # Collect photo data
                    if export_photos:
                        try:
                            photos, _ = get_photos_data(reporting)
                            reporting_data["reporting_files"] = photos
                        except Exception:
                            reporting_data["reporting_files"] = []

                    reportings_data.append(reporting_data)

                # Fill resource and photo headers
                if (
                    not is_inventory
                    and export_resources
                    and (can_view_resources or can_view_money)
                ):
                    add_resource_headers()
                if export_photos:
                    photo_column_start = templ_ws.max_column + 1
                    add_photo_headers()

                row_count = len(reportings_data)

                # Fill data
                for reporting_data in reportings_data:
                    for pos, column_key in enumerate(
                        columns_order, start=COLUMN_START_POSITION
                    ):
                        is_extra_column = column_key in extra_columns
                        resource_value = (
                            reporting_data["extra_columns"].get(column_key)
                            if is_extra_column
                            else reporting_data.get(column_key)
                        )

                        row_cell = templ_ws[cell_by_pos(pos, data_row)]

                        if (
                            column_key
                            in [
                                "km",
                                "end_km",
                                "km_reference",
                                "project_km",
                                "project_end_km",
                                "latitude",
                                "longitude",
                                "length",
                                "width",
                                "height",
                            ]
                            and resource_value is not None
                        ):
                            fill_cell_reporting_section(
                                pos, data_row, resource_value, NUMBER_FORMAT
                            )

                        elif type(resource_value) is datetime.datetime:
                            date_value = (
                                utc_to_local(resource_value)
                                if is_extra_column and resource_value
                                else resource_value
                            )

                            fill_cell_reporting_section(
                                pos,
                                data_row,
                                date_value.replace(tzinfo=None),
                                DATE_FORMAT,
                            )
                        else:
                            fill_cell_reporting_section(pos, data_row, resource_value)

                    # Fill resources
                    resources = reporting_data.get("resources", [])
                    new_column_pos = COLUMN_START_POSITION + len(columns_order)
                    if not is_inventory and resources:
                        if can_view_money:
                            fill_cell_reporting_section(
                                new_column_pos,
                                data_row,
                                resources.get("total", 0),
                                PRICE_FORMAT,
                            )
                            new_column_pos += 1
                        for proc_resource in resources["proc_resources"]:
                            resource_values = [
                                {
                                    "value": "{} ({})".format(
                                        proc_resource["resource_name"],
                                        proc_resource["resource_unit"],
                                    ),
                                    "number_format": None,
                                },
                                {
                                    "value": proc_resource["amount"],
                                    "number_format": None,
                                },
                                {
                                    "value": proc_resource["total_price"],
                                    "number_format": PRICE_FORMAT,
                                },
                                {
                                    "value": proc_resource["unit_price"],
                                    "number_format": PRICE_FORMAT,
                                },
                            ]
                            for resource_value in resource_values:
                                fill_cell_reporting_section(
                                    new_column_pos,
                                    data_row,
                                    resource_value.get("value"),
                                    resource_value.get("number_format"),
                                )
                                new_column_pos += 1

                    # Fill photos
                    photos = reporting_data.get("reporting_files", [])
                    if photos:
                        new_column_pos = photo_column_start
                        for photo in photos:
                            try:
                                resp_future = photo["content"]
                                temp_file = resp_future.result()
                                if temp_file is None:
                                    continue
                                image = Image(temp_file.name)
                                temp_file.close()
                            except Exception:
                                continue
                            else:
                                from_anchor_marker = AnchorMarker(
                                    col=new_column_pos - 1, row=data_row - 1
                                )
                                to_anchor_marker = AnchorMarker(
                                    col=new_column_pos, row=data_row
                                )
                                anchor = TwoCellAnchor(
                                    _from=from_anchor_marker, to=to_anchor_marker
                                )
                                templ_ws.add_image(image, anchor)
                                files_to_close.append(temp_file)
                                templ_ws.column_dimensions[
                                    get_column_letter(new_column_pos)
                                ].width = PHOTO_COLUMN_WIDTH
                                new_column_pos += 1
                                if export_kind:
                                    fill_cell_reporting_section(
                                        new_column_pos, data_row, photo.get("kind")
                                    )
                                    new_column_pos += 1
                                if export_date:
                                    fill_cell_reporting_section(
                                        new_column_pos,
                                        data_row,
                                        photo.get("datetime"),
                                        DATE_FORMAT,
                                    )
                                    new_column_pos += 1
                                if export_description:
                                    fill_cell_reporting_section(
                                        new_column_pos,
                                        data_row,
                                        photo.get("description"),
                                    )
                                    new_column_pos += 1
                    data_row += 1

                last_column = templ_ws.max_column

                # Set styles
                for i, column in enumerate(templ_ws.columns, start=1):
                    if i == 1:
                        continue
                    column_value = column[0].value
                    if column_value is not None:
                        length = max(len(as_text(cell.value)) for cell in column)
                        if column_value == "Observações" or re.match(
                            PHOTO_DATA_PATTERN, column_value
                        ):
                            templ_ws.column_dimensions[
                                column[0].column_letter
                            ].width = 20
                        elif re.match(PHOTO_PATTERN, column_value):
                            pass
                        elif re.match(RESOURCE_PATTERN, column_value):
                            templ_ws.column_dimensions[
                                column[0].column_letter
                            ].width = 36
                        else:
                            templ_ws.column_dimensions[
                                column[0].column_letter
                            ].width = (max(length, 20) * 0.8)

                    for row_cell in column:
                        if row_cell.row == 1:
                            pass
                        elif row_cell.row != row_count + 1:
                            number_format = copy(row_cell.number_format)
                            row_cell.style = simple_style
                            row_cell.number_format = number_format
                        else:
                            number_format = copy(row_cell.number_format)
                            new_border = copy(simple_style.border)
                            new_border.bottom = Side(
                                border_style="medium", color="000000"
                            )
                            row_cell.style = simple_style
                            row_cell.border = new_border
                            row_cell.number_format = number_format

                        if i == COLUMN_START_POSITION:
                            new_border = copy(row_cell.border)
                            new_border.left = Side(
                                border_style="medium", color="000000"
                            )
                            row_cell.border = new_border
                        if i == last_column:
                            new_border = copy(row_cell.border)
                            new_border.right = Side(
                                border_style="medium", color="000000"
                            )
                            row_cell.border = new_border

                # Increase line height if there are exported photos
                if export_photos:
                    for row in range(2, templ_ws.max_row + 1):
                        templ_ws.row_dimensions[row].height = 80

                if is_inventory:
                    templ_ws.title = "Itens de Inventário"

                for named_style in templ_wb.style_names:
                    if named_style == "Normal":
                        continue
                    try:
                        del templ_wb._named_styles[
                            templ_wb.style_names.index(named_style)
                        ]
                    except Exception:
                        pass

                # Saving file
                with NamedTemporaryFile() as temp_file:
                    templ_wb.save(temp_file.name)
                    reporting_export.exported_file.save(
                        "{} (simplificado).xlsx".format(
                            "Inventário" if is_inventory else "Apontamentos"
                        ),
                        ContentFile(temp_file.read()),
                    )
                    error = False

            elif export_type == "NORMAL":
                templ_wb = load_workbook(
                    filename="apps/templates/templates/reporting_export_simple_template.xlsx",
                    read_only=False,
                    keep_vba=False,
                )

                all_columns_basic = remove_columns_normal_export(all_columns)

                templ_base = templ_wb["Apontamentos"]

                header_style = copy(templ_base.cell(column=2, row=HEADER_ROW)._style)

                data_row = 2

                grouped_header = {}

                grouped_data = {}

                grouped_add_total = {}

                # Collect data and group_it
                for reporting in queryset:
                    occ = reporting.occurrence_type
                    if occ:
                        measurement_columns = get_obj_from_path(
                            occ.form_fields, "measurement_columns"
                        )
                        occ_uuid = str(reporting.occurrence_type.uuid)
                        reporting_data = get_reporting_data(
                            reporting, measurement_columns
                        )
                        inner_image_array_info = get_inner_image_array_data(
                            measurement_columns
                        )
                        if export_photos:
                            try:
                                photos, inner_photos = get_photos_data(
                                    reporting, inner_image_array_info
                                )
                                reporting_data["reporting_files"] = photos
                                reporting_data["inner_photos"] = inner_photos
                            except Exception:
                                reporting_data["reporting_files"] = []
                                reporting_data["inner_photos"] = []

                        if occ_uuid not in grouped_data:
                            grouped_data[occ_uuid] = [reporting_data]
                            add_total_columns = get_add_total_columns(
                                measurement_columns
                            )
                            grouped_header[occ_uuid] = measurement_columns
                            grouped_add_total[occ_uuid] = add_total_columns
                        else:
                            grouped_data[occ_uuid].append(reporting_data)

                for occ_uuid, measurement_columns in copy(grouped_header).items():
                    extra_columns = get_exporter_extra_columns_parsed_infos(
                        measurement_columns,
                        reporting_export=True,
                        extra_info=extra_info,
                    )
                    grouped_header[occ_uuid] = extra_columns

                for occ_uuid, reportings_data in grouped_data.items():
                    # Get sheetname
                    occ_name = reportings_data[0].get("occurrence_type")

                    # First row with data
                    data_row = 2

                    # Manipulate sheetname and create new sheet
                    if occ_name:
                        new_sheetname = re.sub(r"[,\]\[*?:/\\]", "", occ_name)[:28]
                        is_sheetname_present = new_sheetname in templ_wb.sheetnames
                        if is_sheetname_present:
                            occurrences_counter = len(
                                list(
                                    filter(
                                        lambda x: x == new_sheetname,
                                        templ_wb.sheetnames,
                                    )
                                )
                            )
                            new_sheetname = (
                                f"{new_sheetname}_{str(occurrences_counter)}"
                            )
                        templ_ws = templ_wb.copy_worksheet(templ_base)
                        templ_ws.title = new_sheetname
                        extra_columns = grouped_header.get(occ_uuid, {})
                        add_total_columns = grouped_add_total.get(occ_uuid, [])
                        all_columns = {**all_columns_basic, **extra_columns}
                        columns_order = [*all_columns.keys()]

                        # Fill headers
                        for pos, column_key in enumerate(
                            columns_order, start=COLUMN_START_POSITION
                        ):
                            fill_header_cell(pos, all_columns.get(column_key, ""))

                        #  Fill photo headers
                        if export_photos:
                            photo_column_start = templ_ws.max_column + 1
                            add_photo_headers()

                        row_count = len(reportings_data)

                        # Clean the reporting_data
                        reportings_data = clean_invalid_characters(reportings_data)

                        # Fill data
                        for reporting_data in reportings_data:
                            for pos, column_key in enumerate(
                                columns_order, start=COLUMN_START_POSITION
                            ):
                                if re.match(ARRAY_PHOTO_PATTERN, column_key):
                                    resource_dict = next(
                                        (
                                            item
                                            for item in reporting_data.get(
                                                "inner_photos", []
                                            )
                                            if column_key in item
                                        ),
                                        {},
                                    )
                                    resource_value = resource_dict.get(column_key)
                                    if resource_value:
                                        (
                                            key,
                                            _,
                                            inner_key,
                                            control_string,
                                        ) = column_key.split("#")
                                        if control_string == "content":
                                            try:
                                                resp_future = resource_value
                                                temp_file = resp_future.result()
                                                if temp_file is None:
                                                    continue
                                                image = Image(temp_file.name)
                                                temp_file.close()
                                            except Exception:
                                                continue
                                            else:
                                                from_anchor_marker = AnchorMarker(
                                                    col=pos - 1, row=data_row - 1
                                                )
                                                to_anchor_marker = AnchorMarker(
                                                    col=pos, row=data_row
                                                )
                                                anchor = TwoCellAnchor(
                                                    _from=from_anchor_marker,
                                                    to=to_anchor_marker,
                                                )
                                                templ_ws.add_image(image, anchor)
                                                files_to_close.append(temp_file)
                                                templ_ws.column_dimensions[
                                                    get_column_letter(pos)
                                                ].width = PHOTO_COLUMN_WIDTH
                                        elif control_string == "kind":
                                            fill_cell_reporting_section(
                                                pos, data_row, resource_value
                                            )
                                        elif control_string == "datetime":
                                            fill_cell_reporting_section(
                                                pos,
                                                data_row,
                                                resource_value,
                                                DATE_FORMAT,
                                            )
                                        elif control_string == "description":
                                            fill_cell_reporting_section(
                                                pos,
                                                data_row,
                                                resource_value,
                                            )

                                else:
                                    is_extra_column = column_key in extra_columns
                                    resource_value = (
                                        reporting_data["extra_columns"].get(column_key)
                                        if is_extra_column
                                        else reporting_data.get(column_key)
                                    )

                                    row_cell = templ_ws[cell_by_pos(pos, data_row)]
                                    if (
                                        column_key
                                        in [
                                            "km",
                                            "end_km",
                                            "km_reference",
                                            "project_km",
                                            "project_end_km",
                                            "latitude",
                                            "longitude",
                                            "length",
                                            "width",
                                            "height",
                                        ]
                                        and resource_value is not None
                                    ):
                                        fill_cell_reporting_section(
                                            pos, data_row, resource_value, NUMBER_FORMAT
                                        )

                                    elif type(resource_value) is datetime.datetime:
                                        date_value = (
                                            utc_to_local(resource_value)
                                            if is_extra_column and resource_value
                                            else resource_value
                                        )

                                        fill_cell_reporting_section(
                                            pos,
                                            data_row,
                                            date_value.replace(tzinfo=None),
                                            DATE_FORMAT,
                                        )
                                    else:
                                        fill_cell_reporting_section(
                                            pos, data_row, resource_value
                                        )

                            # Fill photos
                            photos = reporting_data.get("reporting_files", [])
                            if photos:
                                new_column_pos = photo_column_start
                                for photo in photos:
                                    try:
                                        resp_future = photo["content"]
                                        temp_file = resp_future.result()
                                        if temp_file is None:
                                            continue
                                        image = Image(temp_file.name)
                                        temp_file.close()
                                    except Exception:
                                        continue
                                    else:
                                        from_anchor_marker = AnchorMarker(
                                            col=new_column_pos - 1, row=data_row - 1
                                        )
                                        to_anchor_marker = AnchorMarker(
                                            col=new_column_pos, row=data_row
                                        )
                                        anchor = TwoCellAnchor(
                                            _from=from_anchor_marker,
                                            to=to_anchor_marker,
                                        )
                                        templ_ws.add_image(image, anchor)
                                        files_to_close.append(temp_file)
                                        templ_ws.column_dimensions[
                                            get_column_letter(new_column_pos)
                                        ].width = PHOTO_COLUMN_WIDTH
                                        new_column_pos += 1
                                        if export_kind:
                                            fill_cell_reporting_section(
                                                new_column_pos,
                                                data_row,
                                                photo.get("kind"),
                                            )
                                            new_column_pos += 1
                                        if export_date:
                                            fill_cell_reporting_section(
                                                new_column_pos,
                                                data_row,
                                                photo.get("datetime"),
                                                DATE_FORMAT,
                                            )
                                            new_column_pos += 1
                                        if export_description:
                                            fill_cell_reporting_section(
                                                new_column_pos,
                                                data_row,
                                                photo.get("description"),
                                            )
                                            new_column_pos += 1
                            data_row += 1

                        last_column = templ_ws.max_column

                        # Set styles
                        for i, column in enumerate(templ_ws.columns, start=1):
                            if i == 1:
                                continue
                            column_value = column[0].value
                            if column_value is not None:
                                length = max(
                                    len(as_text(cell.value)) for cell in column
                                )
                                if (
                                    column_value == "Observações"
                                    or re.match(PHOTO_DATA_PATTERN, column_value)
                                    or re.match(ARRAY_PHOTO_DATA_PATTERN, column_value)
                                ):
                                    templ_ws.column_dimensions[
                                        column[0].column_letter
                                    ].width = 20
                                elif re.match(PHOTO_PATTERN, column_value) or re.match(
                                    ARRAY_PHOTO_CONTENT_PATTERN, column_value
                                ):
                                    pass
                                elif re.match(RESOURCE_PATTERN, column_value):
                                    templ_ws.column_dimensions[
                                        column[0].column_letter
                                    ].width = 36
                                else:
                                    templ_ws.column_dimensions[
                                        column[0].column_letter
                                    ].width = (max(length, 20) * 0.8)

                            for row_cell in column:
                                if row_cell.row == 1:
                                    pass
                                elif row_cell.row != row_count + 1:
                                    number_format = copy(row_cell.number_format)
                                    row_cell.style = normal_style
                                    row_cell.number_format = number_format
                                else:
                                    number_format = copy(row_cell.number_format)
                                    new_border = copy(normal_style.border)
                                    new_border.bottom = Side(
                                        border_style="medium", color="000000"
                                    )
                                    row_cell.style = normal_style
                                    row_cell.border = new_border
                                    row_cell.number_format = number_format

                                if i == COLUMN_START_POSITION:
                                    new_border = copy(row_cell.border)
                                    new_border.left = Side(
                                        border_style="medium", color="000000"
                                    )
                                    row_cell.border = new_border
                                if i == last_column:
                                    new_border = copy(row_cell.border)
                                    new_border.right = Side(
                                        border_style="medium", color="000000"
                                    )
                                    row_cell.border = new_border

                        # Apply line height

                        for row in range(2, templ_ws.max_row + 1):
                            templ_ws.row_dimensions[row].height = (
                                80 if export_photos else 15
                            )

                        # Add "TOTAIS" line

                        add_total_row = templ_ws.max_row + 2
                        add_total_title_cell = templ_ws[
                            "B{}".format(str(add_total_row))
                        ]
                        add_total_title_cell.value = "TOTAIS"
                        new_border = copy(add_total_style.border)
                        new_border.left = Side(border_style="medium", color="000000")
                        add_total_title_cell.border = new_border
                        add_total_title_cell.font = Font(color="000000", bold=True)

                        add_total_line = templ_ws[add_total_row]

                        for i, column_cell in enumerate(add_total_line, start=1):
                            if i in [1, 2]:
                                continue
                            column_cell.style = add_total_style
                            if i == 3:
                                new_border = copy(column_cell.border)
                                new_border.left = Side(
                                    border_style="thin", color="000000"
                                )
                                column_cell.border = new_border
                            elif i == last_column:
                                new_border = copy(column_cell.border)
                                new_border.right = Side(
                                    border_style="medium", color="000000"
                                )
                                column_cell.border = new_border
                            column_letter = column_cell.column_letter
                            if (
                                templ_ws["{}1".format(column_letter)].value
                                in add_total_columns
                            ):
                                column_cell.value = "=SUM({0}{1}:{0}{2})".format(
                                    column_letter, "2", str(add_total_row - 2)
                                )
                                column_cell.number_format = NUMBER_FORMAT
                                column_cell.font = Font(color="000000", bold=True)
                                column_cell.fill = PatternFill(
                                    "solid", fgColor="FFCCCCCC"
                                )

                templ_wb.remove(templ_base)

                for named_style in templ_wb.style_names:
                    if named_style == "Normal":
                        continue
                    try:
                        del templ_wb._named_styles[
                            templ_wb.style_names.index(named_style)
                        ]
                    except Exception:
                        pass

                # Saving file
                with NamedTemporaryFile() as temp_file:
                    templ_wb.save(temp_file.name)
                    reporting_export.exported_file.save(
                        "{}.xlsx".format(
                            "Inventário" if is_inventory else "Apontamentos"
                        ),
                        ContentFile(temp_file.read()),
                    )
                    error = False

            # Close image files
            for open_file in files_to_close:
                if not open_file.closed:
                    open_file.close()
                if os.path.exists(open_file.name):
                    os.remove(open_file.name)

        except ReportingFileCountExceededException:
            reporting_export.extra_info["error_cause"] = {"limit": MAX_REPORTING_FILES}
            error = True

        except Exception as e:
            logging.error(
                "Untreated exception found while exporting Reportings. Check Sentry."
            )
            capture_exception(e)
            error = True

        reporting_export.error = error
        reporting_export.done = True
        reporting_export.save()
        rf_thread_pool.shutdown(wait=False)

        logging.info(
            f"[ReportingExport] Finalizado | id={reporting_export_id} | error={error} | fargate={bool(os.environ.get('ECS_CONTAINER_METADATA_URI_V4'))}"
        )

        if os.environ.get("ECS_CONTAINER_METADATA_URI_V4"):
            if not error:
                try:
                    from apps.templates.notifications import send_email_reporting_export

                    send_email_reporting_export(reporting_export)
                except Exception as e:
                    logging.error(e)
                    capture_exception(e)

            try:
                from helpers.ecs_task_service import publish_fargate_metrics

                publish_fargate_metrics(
                    "ReportingExport",
                    time.time() - start_time,
                    not error,
                    _fargate_counts["item_count"],
                    _fargate_counts["image_count"],
                )
            except Exception as e:
                capture_exception(e)
            finally:
                from helpers.ecs_task_service import stop_current_fargate_task

                stop_current_fargate_task(reason="ReportingExport concluído")
