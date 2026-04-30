import tempfile
from typing import List

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from apps.companies.models import Firm
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.export_utils import insert_centered_value
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
)
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km


class XlsxHandlerResumeReportAntiglareScreens(object):
    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.wb = load_workbook("./fixtures/reports/ccr_resume_antiglare_screens.xlsx")
        self.s3 = s3
        self.__sheet_target = sheet_target
        self._worksheet: Worksheet = self.wb.active
        self._worksheet.title = "Cadastro Resumo"
        self.list_uuids = list_uuids
        self.reportings = Reporting.objects.filter(uuid=list_uuids[0]).prefetch_related(
            "company"
        )
        self.temp_file = tempfile.mkdtemp()

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="L1:M5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:A5",
            resize_method=ResizeMethod.ProportionalLeft,
        )

        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.resume_report_dict = []
        self.static_fields = {
            "index": "A",
            "road": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "side": "F",
            "length": "G",
            "height": "H",
            "corrosion": "I",
            "alignment": "J",
            "fixation_posts": "K",
            "screw": "L",
            "general_state": "M",
        }

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def __insert_new_rows(self, row: int):
        chars = [chr(ord("A") + i) for i in range(13)]

        for char in chars:
            self._worksheet[f"{char}{row}"].border = self.border
            self._worksheet.row_dimensions[row].height = 20

    def create_dict(self, reporting):
        result_index = new_get_form_data(reporting, "index", "idCcrAntt", "")
        direction = get_custom_option(reporting, "direction")

        side = new_get_form_data(reporting, "side", default="")

        height_value = new_get_form_data(reporting, "height", default="")

        corrosion = "X" if reporting.form_data.get("corrosion") else "-"
        alignment = "X" if reporting.form_data.get("alignment") else "-"
        screw = "X" if reporting.form_data.get("screw") else "-"

        general_state = new_get_form_data(reporting, "generalState", default="")

        result_general_state = general_state if general_state else "-"
        fixation_posts = reporting.form_data.get("fixation_posts")

        result_fixation_posts = "X" if fixation_posts else "-"

        length_value = reporting.form_data.get("length")
        data = {
            "index": result_index,
            "road": reporting.__dict__.get("road_name"),
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "end_km", 3),
            "raw_initial_km": reporting.km,
            "direction": direction.upper(),
            "side": side if side else "-",
            "height": height_value if height_value else "-",
            "length": length_value if length_value else "-",
            "corrosion": corrosion,
            "alignment": alignment,
            "fixation_posts": result_fixation_posts,
            "general_state": result_general_state,
            "screw": screw,
            "executed_at": reporting.executed_at,
            "team": str(reporting.firm.__dict__.get("uuid")) if reporting.firm else "",
            "road_name": reporting.__dict__.get("road_name"),
            "subcompany": (
                reporting.firm.subcompany.__dict__.get("name") if reporting.firm else ""
            ),
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def fill_sheet(self, *, data_list: list):
        row = 8
        team_list = []
        date_list = []
        road_name_list = []
        subcompany_list = []

        list_mapping = {
            "executed_at": date_list,
            "road_name": road_name_list,
            "subcompany": subcompany_list,
            "team": team_list,
        }

        for values in data_list:
            self.__insert_new_rows(row=row)
            for key, value in values.items():
                if key in list_mapping:
                    list_mapping[key].append(value)
                else:
                    if key == "raw_initial_km":
                        continue
                    key_value = f"{self.static_fields[key]}{row}"
                    self._worksheet[key_value] = value
                    XlsxHandlerResumeReportAntiglareScreens.__format_fonts(
                        cell=self._worksheet[key_value], size=11
                    )

            row += 1

        filtered_team_list = set([uuid for uuid in list_mapping["team"] if uuid != ""])
        query_set_teams = Firm.objects.filter(uuid__in=filtered_team_list).all()
        names = []
        for team in query_set_teams:
            users_query_set = team.users.all()
            intern_list = []
            for user in users_query_set:
                intern_list.append(user.full_name)
            names.append(intern_list)
        names = [", ".join(_) for _ in names]
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(names),
            cell="D3",
            horizontal="left",
        )

        filtered_date_list = [
            date for date in list_mapping["executed_at"] if date != ""
        ]
        filtered_date_list.sort()
        if len(filtered_date_list) == 0:
            date_text = ""
        elif len(filtered_date_list) == 1:
            date_text = filtered_date_list[0].strftime("%d/%m/%Y")
        else:
            date_text = f"{filtered_date_list[0].strftime('%d/%m/%Y')} até {filtered_date_list[-1].strftime('%d/%m/%Y')}"

        insert_centered_value(
            worksheet=self._worksheet, value=date_text, cell="D4", horizontal="left"
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value=self.company.name,
            cell="G4",
            horizontal="left",
        )

        subcompany = set(subcompany_list)
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(subcompany),
            cell="D2",
            horizontal="left",
        )

        road_name = set(list_mapping["road_name"])
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(road_name),
            cell="G2",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value="Elementos de Proteção e Segurança - Cadastro Resumo de Telas Antiofuscante",
            cell="G3",
            horizontal="left",
        )

        total_length = sum(
            [t.value for t in self._worksheet["G"] if isinstance(t.value, (int, float))]
        )
        reference_total_length = self._worksheet[f"{'G'}{self._worksheet.max_row + 1}"]
        reference_total_length.value = total_length
        reference_total_length.border = self.border
        XlsxHandlerResumeReportAntiglareScreens.__format_fonts(
            cell=reference_total_length, size=11, bold=True
        )

        reference_total_string = reference_total_length.offset(0, -1)
        reference_total_string.value = "TOTAL(m)"
        reference_total_string.border = self.border
        XlsxHandlerResumeReportAntiglareScreens.__format_fonts(
            cell=reference_total_string, size=11, bold=True
        )

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            target=self.__sheet_target,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

    def execute(self):
        query_set = (
            Reporting.objects.filter(occurrence_type=self.form, pk__in=self.list_uuids)
            .prefetch_related("occurrence_type", "firm", "firm__subcompany")
            .distinct()
        )

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        sorted_data = sorted(
            data,
            key=lambda x: (
                x["raw_initial_km"] if x["direction"] == "NORTE" else 0,
                -x["raw_initial_km"] if x["direction"] == "SUL" else 0,
                x["direction"] == "CANTEIRO CENTRAL",
                x["direction"] == "AMBOS",
                x["direction"] == "CRESCENTE",
                x["direction"] == "DECRESCENTE",
                x["direction"] == "LESTE",
                x["direction"] == "LESTE/OESTE",
                x["direction"] == "NORTE/SUL",
                x["direction"] == "TRANSVERSAL",
                x["direction"] == "OESTE",
                x["direction"],
            ),
            reverse=True,
        )

        self.fill_sheet(data_list=sorted_data)

        road_name = self.reportings.first().road_name

        file_name = "Cadastro Resumo - {} - Telas Antiofuscantes".format(road_name)

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        result = f"/tmp/{file_name}.xlsx"
        self.wb.save(result)
        return result
