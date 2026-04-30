import re
import shutil
import tempfile
from typing import List
from uuid import uuid4
from zipfile import ZipFile

from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Font, Side
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import form_data_images_grouped
from helpers.strings import (
    COMMON_IMAGE_TYPE,
    deep_keys_to_snake_case,
    format_km,
    int_set_zero_prefix,
)


def _get_instances_reporting_file_from_urls(list_urls: list) -> List:
    _list_images = [
        instance
        for instance in [_loc_file_name_in_reporting_file(_url) for _url in list_urls]
        if instance
    ]
    return _list_images


def _get_file_name_to_zip(all_roads: list) -> str:
    return f"Fichas de Drenagem Superficial ARTESP - {'_'.join([re.sub(r'[- ]', '', road) for road in all_roads])}.zip"


def _loc_file_name_in_reporting_file(
    url,
    query_set: ReportingFile = ReportingFile.objects.filter(upload__isnull=False),
    pk_excludes: list = [],
):
    if "http" in url:
        split_url = url.split("?X")

        tiny_url_split = split_url[0].split("/") if split_url else None
        if not tiny_url_split or not isinstance(tiny_url_split, list):
            return

        _file_name = tiny_url_split[-1]
        file_instance = (
            query_set.exclude(pk__in=pk_excludes).filter(upload=_file_name).first()
        )
    else:
        file_instance = query_set.exclude(pk__in=pk_excludes).filter(pk=url).first()
    return file_instance


def _km_format_to_prefix(km) -> str:
    try:
        if km:
            data_list = _ = str(km).split(".")
            _km = str(data_list[0].zfill(3))
            _mt = str(data_list[1].ljust(3, "0"))
            return f"{_km[:3]}{_mt[:3]}"
        return ""
    except Exception:
        return ""


border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
bold = Font(bold=True)


class XlsxHandler:
    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.__list_reporting = list_reporting
        self.__sheet_target = sheet_target
        self.temp_file = tempfile.mkdtemp()
        self.__xlsx_file = (
            "./fixtures/reports/ccr_report_artesp_surface_drainage_sheets.xlsx"
        )
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook.active
        # Custom config
        self.default_name_excel_sheet = "DS 000+000 - Ficha"
        self.default_name_excel_photo = "DS 000+000 - Fotos 1"
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="Y1:AC4",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:E4",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        self.croqui_range_string = "A25:AC53"
        self.max_sheets_croqui = 4
        self.max_photo_to_sheets = 6
        self.salt_row_default = 20

        self.static_fields_sheet = {
            "element": "E11",
            "id_ccr_antt": "E13",
            "road_name": "E15",
            "track": "E17",
            "initial_km": "E19",
            "end_km": "E21",
            "width": "P11",
            "height": "P13",
            "extension": "P15",
            "geometric_session": "P17",
            "material": "P19",
            "environment": "P21",
            "element_detail": "W11",
            "initial_connection": "W13",
            "end_connection": "W15",
            "zone": "Y17",
            "e1": "X19",
            "n1": "AA19",
            "e2": "X21",
            "n2": "AA21",
            "executed_at": "F58",
            "sub_company_name": "F60",
            "responsible": "R58",
            "vistor_names": "R60",
            "observer_geral_conservation": "E67",
            "diagnostic_ok": "B74",
            "to_repair": "B76",
            "local_extension_to_repair": "K76",
            "cleaner": "B78",
            "local_extension_cleaner": "K78",
            "deploy": "B80",
            "local_extension_deploy": "K80",
            "observation_diagnostic": "E82",
        }
        self.static_fields_photo = {
            "element": "E11",
            "id_ccr_antt": "E13",
            "road_name": "N11",
            "track": "N13",
            "initial_km": "V11",
            "end_km": "V13",
            "start_photo": "B19",
            "end_photo": "N34",
        }

    def __set_photos(self, data: dict):
        photos = data.get("photos")
        self.max_sheets = 4
        original_sheet = self._workbook.worksheets[-1]
        self._worksheet = original_sheet
        new_base_name_excel_photo = (self.default_name_excel_photo).replace(
            "000+000", data.get("initial_km")
        )
        self._worksheet.title = new_base_name_excel_photo

        for field, column in self.static_fields_photo.items():
            _data = data.get(field, None)

            if _data is None:
                continue

            self._worksheet[column] = _data

        if photos:
            count_sheet = 1
            count_photo = 1
            col_left_init = "B"
            col_left_end = "N"
            col_left_description = "F"
            col_right_init = "P"
            col_right_end = "AB"
            col_right_description = "T"

            while photos:
                row_photo_init = 19
                row_photo_end = 34
                row_data = 35

                # Pela regra de negocio só pode ter no máximo abas do excel se haver mais abas ele para a criação
                if count_sheet > self.max_sheets:
                    break

                photos_currency = []

                if count_sheet > 1:
                    title = new_base_name_excel_photo[
                        : (len(self.default_name_excel_photo) - 1)
                    ]
                    new_sheet = self._workbook.copy_worksheet(original_sheet)
                    new_sheet.title = f"{title}{count_sheet}"
                    new_sheet.print_options.horizontalCentered = True
                    new_sheet.print_area = "A1:AC78"
                    self._worksheet = new_sheet
                else:
                    title = self.default_name_excel_photo[
                        : (len(self.default_name_excel_photo) - 1)
                    ]
                    self._worksheet = self._workbook.worksheets[-1]
                    self._worksheet.title = f"{new_base_name_excel_photo[:(len(new_base_name_excel_photo) - 1)]}{count_sheet}"

                # Pega as fotos por pagina
                if len(photos) > 0:
                    for _ in range(self.max_photo_to_sheets):
                        if photos:
                            photo_remove = photos.pop(0)
                            photos_currency.append(photo_remove)

                ws = self._worksheet

                # Clean Campos
                salt_row = 0
                for _ in range(int(self.max_photo_to_sheets / 2)):
                    row_init_cleaner = row_photo_init + salt_row
                    row_data_cleaner = row_data + salt_row

                    coordination = f"{col_left_init}{row_init_cleaner}"
                    ws[coordination] = ""

                    coordination = f"{col_right_init}{row_init_cleaner}"
                    ws[coordination] = ""

                    coordination = f"{col_left_init}{row_data_cleaner}"
                    ws[coordination] = ""

                    coordination = f"{col_right_init}{row_data_cleaner}"
                    ws[coordination] = ""

                    coordination = f"{col_left_description}{row_data_cleaner}"
                    ws[coordination] = ""

                    coordination = f"{col_right_description}{row_data_cleaner}"
                    ws[coordination] = ""

                    salt_row += self.salt_row_default

                if photos_currency:
                    prefix_photo_name = data.get("prefix_photo_name")
                    salt_row = 0

                    for _obj in photos_currency:
                        path_image = _obj.get("path_image")
                        photo_number = int_set_zero_prefix(count_photo)
                        photo_name = f"{prefix_photo_name}F{photo_number}"

                        left = count_photo % 2 > 0
                        if left:
                            try:
                                insert_picture_2(
                                    self._worksheet,
                                    range_string=f"{col_left_init}{row_photo_init}:{col_left_end}{row_photo_end}",
                                    picture=Image(path_image),
                                    target=self.__sheet_target,
                                    resize_method=ResizeMethod.ProportionalCentered,
                                )
                            except Exception:
                                continue

                            coordination = f"{col_left_init}{row_data}"
                            ws[coordination] = photo_name

                            coordination = f"{col_left_description}{row_data}"
                            ws[coordination] = _obj.get("description")
                        else:
                            try:
                                insert_picture_2(
                                    self._worksheet,
                                    range_string=f"{col_right_init}{row_photo_init}:{col_right_end}{row_photo_end}",
                                    picture=Image(path_image),
                                    target=self.__sheet_target,
                                    resize_method=ResizeMethod.ProportionalCentered,
                                )
                            except Exception:
                                continue

                            coordination = f"{col_right_init}{row_data}"
                            ws[coordination] = photo_name

                            coordination = f"{col_right_description}{row_data}"
                            ws[coordination] = _obj.get("description")

                        count_photo += 1
                        if not left:
                            row_photo_init += self.salt_row_default
                            row_photo_end += self.salt_row_default
                            row_data += self.salt_row_default

                insert_logo_and_provider_logo(
                    worksheet=self._worksheet,
                    logo_company=self.data_logo_company,
                    provider_logo=self.data_provider_logo,
                    target=self.__sheet_target,
                )
                count_sheet += 1
        else:
            self._worksheet = self._workbook.worksheets[-1]

            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
                target=self.__sheet_target,
            )

    def fill_sheet(self, data_list: list):
        data_work = data_list
        files = []
        filenames = []
        all_roads = []

        for data in data_work:
            road_name = data.get("road_name")
            all_roads.append(road_name)

            self._worksheet = self._workbook[self.default_name_excel_sheet]
            self._worksheet.title = (self.default_name_excel_sheet).replace(
                "000+000", data.get("initial_km")
            )

            for field, column in self.static_fields_sheet.items():
                _data = data.get(field, None)

                if _data is None:
                    continue

                try:
                    self._worksheet[column] = _data
                except Exception as err:
                    print(err)

            path_croqui_image = data.get("croqui_image")
            if path_croqui_image:
                insert_picture_2(
                    self._worksheet,
                    range_string=self.croqui_range_string,
                    picture=Image(path_croqui_image),
                    target=self.__sheet_target,
                    resize_method=ResizeMethod.ProportionalCentered,
                )

            no_of_geral_conservation = data.get("no_of_geral_conservation")
            no_of_geral_conservation_column = "H65"
            if no_of_geral_conservation == 2:
                no_of_geral_conservation_column = "N65"
            elif no_of_geral_conservation == 1:
                no_of_geral_conservation_column = "S65"
            self._worksheet[no_of_geral_conservation_column] = "X"

            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
                target=self.__sheet_target,
            )

            self.__set_photos(data)

            file_name = data.get("file_name")
            if file_name in filenames:
                n_file = filenames.count(file_name)
                file_name += f"_{n_file}"

            filenames.append(file_name)
            result = f"/tmp/{file_name}.xlsx"
            self._workbook.save(result)
            files.append(result)

            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook.active

        all_roads = list(set(all_roads))
        all_roads.sort()

        return {"files": files, "names": filenames, "all_roads": all_roads}

    def _get_download_pictures(self, url):
        path_image = ""
        if url:
            try:
                file_path = url.split("?")[0].split(".com/")[1]
                bucket_name = url.split(".s3")[0].split("/")[-1]
                image_format = file_path.split(".")[-1]
                if image_format not in COMMON_IMAGE_TYPE:
                    return ""

                path_image = f"{self.temp_file}{uuid4()}.{image_format}"
                self.s3.download_file(bucket_name, file_path, path_image)
            except Exception:
                path_image = ""
        return path_image

    def __get_data_photo(self, reporting_file: ReportingFile) -> dict:
        url_image = reporting_file.upload.url if reporting_file.upload else ""
        path_imagem = self._get_download_pictures(url_image)

        return {
            "pk": str(reporting_file.pk),
            "path_image": path_imagem,
            "description": reporting_file.description,
            "datetime": reporting_file.datetime,
        }

    def create_dict(self, reporting: Reporting, s3) -> dict:
        form_data_display = reporting.get_form_data_display()
        form_data: dict = reporting.form_data
        fields: list = deep_keys_to_snake_case(
            reporting.occurrence_type.form_fields.get("fields", [])
        )

        file_name: str = form_data_display.get("id_ccr_antt") or reporting.number or "-"
        prefix_sheet = f"DS {format_km(reporting, 'km', 3)}"
        ID_CCR_ANTT = form_data_display.get("id_ccr_antt") or "-"
        ELEMENT = form_data_display.get("elemento") or "-"
        ROAD_NAME = reporting.road_name or "-"
        TRACK = form_data_display.get("pista")
        INITIAL_KM = format_km(reporting, "km", 3) or "-"
        END_KM = format_km(reporting, "end_km", 3) or "-"
        KM_FLOAT = reporting.km
        WIDTH = (
            form_data_display.get("largura") or form_data_display.get("width") or "-"
        )
        HEIGHT = (
            form_data_display.get("altura") or form_data_display.get("height") or "-"
        )
        EXTENSION = form_data_display.get("length") or "-"
        GEOMETRIC_SESSION = form_data_display.get("geometric_session") or "-"
        MATERIAL = form_data_display.get("material") or "-"
        ENVIRONMENT = form_data_display.get("ambiente") or "-"
        ELEMENT_DETAIL = form_data_display.get("detalheelemento") or "-"
        INITIAL_CONNECTION = form_data_display.get("conexao_inicio") or "-"
        END_CONNECTION = form_data_display.get("conexao_fim")
        ZONE = form_data_display.get("zona") or "-"
        E1 = form_data_display.get("xini") or "-"
        N1 = form_data_display.get("yini") or "-"
        E2 = form_data_display.get("xfim") or "-"
        N2 = form_data_display.get("yfim") or "-"

        inspection_campaign_year = form_data_display.get("inspection_campaign_year", "")
        type_element = form_data_display.get("tipoelemento", "-")
        no_of_road_name = re.sub(r"\D", "", ROAD_NAME)
        uf = form_data_display.get("uf", "")
        prefix_photo_name = f"{type_element}{inspection_campaign_year}{no_of_road_name}k{_km_format_to_prefix(KM_FLOAT)}{uf}"

        croqui_image = ""
        photos = []

        photos_report: list = form_data_display.get("fotos_relatorio", [])
        query = Q()
        for extension in COMMON_IMAGE_TYPE:
            query |= Q(upload__icontains=extension)

        pks_in_croqui = []
        instance_use_croqui = []
        if photos_report:
            for obj in photos_report:
                for _key, _array in obj.items():
                    if _array and _key == "croqui_image":
                        if not croqui_image:
                            instance_use_croqui = [
                                instance.pk
                                for instance in [
                                    _loc_file_name_in_reporting_file(_url)
                                    for _url in _array
                                ]
                                if instance
                            ]
                            pks_in_croqui.extend(instance_use_croqui)
                            url_croqui_image = _array[0]
                            croqui_image = self._get_download_pictures(url_croqui_image)

        last_monitoring_uuids = reporting.form_data.get("last_monitoring", [])
        flattened_last_monitoring_files = []
        for monitoring_dict in last_monitoring_uuids:
            if (
                isinstance(monitoring_dict, dict)
                and "last_monitoring_files" in monitoring_dict
            ):
                flattened_last_monitoring_files.extend(
                    monitoring_dict["last_monitoring_files"]
                )

        excluded_rf_uuids = pks_in_croqui + flattened_last_monitoring_files
        qs_reporting_files = (
            ReportingFile.objects.exclude(pk__in=excluded_rf_uuids)
            .filter(reporting=reporting)
            .distinct()
            .order_by("datetime", "uploaded_at")
        )

        image_arrays = form_data_images_grouped(reporting)
        for instance_file in qs_reporting_files:
            extension = str(instance_file.upload).split(".")
            if extension and (extension[-1]).lower() in COMMON_IMAGE_TYPE:
                result = self.__get_data_photo(instance_file)
                section = "Arquivos e Imagens"
                for _, image_array in image_arrays.items():
                    if str(instance_file.uuid) in image_array.uuids:
                        if image_array.sectionSubtitle is not None:
                            section = image_array.sectionSubtitle
                        elif image_array.display_name is not None:
                            section = image_array.display_name
                        else:
                            section = None
                        break
                result["description"] = f"{section} - {result['description']}"

                photos.append(result)

        EXECUTED_AT = (
            reporting.executed_at.strftime("%d/%m/%Y") if reporting.executed_at else "-"
        )

        sub_company_name = "-"
        responsible = "-"

        team = None
        vistor_names = "-"

        if getattr(reporting, "firm"):
            firm = reporting.firm
            responsible = firm.name
            vistor_names = ("/").join(
                [inspector.full_name for inspector in firm.inspectors.all()]
            ) or "-"

            if getattr(reporting.firm, "subcompany"):
                team = firm.subcompany
                sub_company_name = team.name

        NO_OF_GERAL_CONSERVATION = int(form_data.get("general_conservation_state", "3"))
        OBSERVER_GERAL_CONSERVATION = form_data_display.get(
            "observacoesconservacao", "-"
        )

        TO_REPAIR = "X" if form_data.get("reparar") else ""
        LOCAL_TO_REPAIR = form_data_display.get("local_reparo", "")
        EXTENSION_TO_REPAIR = form_data_display.get("extensaoreparo", "")

        CLEANER = "X" if form_data.get("limpeza") else ""
        LOCAL_CLEANER = form_data_display.get("local_limpeza", "")
        EXTENSION_CLEANER = form_data_display.get("extensaolimpeza", "")

        DEPLOY = "X" if form_data.get("implantar") else ""
        LOCAL_DEPLOY = form_data_display.get("local_implantacao", "")
        EXTENSION_DEPLOY = form_data_display.get("extensaoimplantacao", "")

        diagnostic_ok = ""
        local_extension_to_repair = "-"
        local_extension_cleaner = "-"
        local_extension_deploy = "-"

        if any([TO_REPAIR, CLEANER, DEPLOY]):
            unit = " m"
            if TO_REPAIR:
                local_extension_to_repair = LOCAL_TO_REPAIR
                if local_extension_to_repair:
                    local_extension_to_repair += " / "
                if EXTENSION_TO_REPAIR:
                    if not local_extension_to_repair:
                        local_extension_to_repair = "- / "
                    for field in fields:
                        if field.get("api_name") == "extensaoreparo":
                            unit = f" {field.get('unit')}"
                            break
                    local_extension_to_repair += f"{EXTENSION_TO_REPAIR}{unit}"

                if not local_extension_to_repair:
                    local_extension_to_repair = "-"

            if CLEANER:
                local_extension_cleaner = LOCAL_CLEANER
                if local_extension_cleaner:
                    local_extension_cleaner += " / "
                if EXTENSION_CLEANER:
                    if not local_extension_cleaner:
                        local_extension_cleaner = "- / "
                    for field in fields:
                        if field.get("api_name") == "extensaolimpeza":
                            unit = f" {field.get('unit')}"
                            break
                    local_extension_cleaner += f"{EXTENSION_CLEANER}{unit}"

                if not local_extension_cleaner:
                    local_extension_cleaner = "-"

            if DEPLOY:
                local_extension_deploy = LOCAL_DEPLOY
                if local_extension_deploy:
                    local_extension_deploy += " / "
                if EXTENSION_DEPLOY:
                    if not local_extension_deploy:
                        local_extension_deploy = "- / "
                    for field in fields:
                        if field.get("api_name") == "extensaoimplantar":
                            unit = f" {field.get('unit')}"
                            break
                    local_extension_deploy += f"{EXTENSION_DEPLOY}{unit}"

                if not local_extension_deploy:
                    local_extension_deploy = "-"
        else:
            diagnostic_ok = "X"

        OBSERVATION_DIAGNOSTIC = form_data_display.get("observacoes_diagnostico", "-")

        data = {
            "element": ELEMENT,
            "id_ccr_antt": ID_CCR_ANTT,
            "road_name": ROAD_NAME,
            "track": TRACK,
            "initial_km": INITIAL_KM,
            "end_km": END_KM,
            "width": WIDTH,
            "height": HEIGHT,
            "extension": EXTENSION,
            "geometric_session": GEOMETRIC_SESSION,
            "material": MATERIAL,
            "environment": ENVIRONMENT,
            "element_detail": ELEMENT_DETAIL,
            "type_element": type_element,
            "initial_connection": INITIAL_CONNECTION,
            "end_connection": END_CONNECTION,
            "zone": ZONE,
            "e1": E1,
            "n1": N1,
            "e2": E2,
            "n2": N2,
            "croqui_image": croqui_image,
            "executed_at": EXECUTED_AT,
            "sub_company_name": sub_company_name,
            "responsible": vistor_names,  # Inversão solicitada no KTD-6946
            "vistor_names": responsible,  # Inversão solicitada no KTD-6946
            "no_of_geral_conservation": NO_OF_GERAL_CONSERVATION,
            "observer_geral_conservation": OBSERVER_GERAL_CONSERVATION,
            "diagnostic_ok": diagnostic_ok,
            "to_repair": TO_REPAIR,
            "cleaner": CLEANER,
            "deploy": DEPLOY,
            "local_extension_to_repair": local_extension_to_repair,
            "local_extension_cleaner": local_extension_cleaner,
            "local_extension_deploy": local_extension_deploy,
            "observation_diagnostic": OBSERVATION_DIAGNOSTIC,
            "photos": photos,
            "km_float": KM_FLOAT,
            "file_name": file_name,
            "prefix_photo_name": prefix_photo_name,
            "prefix_sheet": prefix_sheet,
        }

        return data

    def execute(self):
        data = []
        for reporting in self.__list_reporting:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class XlsxHandlerReportSurfaceDrainageSheets(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        self.class_name = "Monitoração Drenagem Superficial"
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""
        reportings = Reporting.objects.filter(
            occurrence_type__name=self.class_name, uuid__in=self.uuids
        ).prefetch_related("road")

        if reportings.count() == 1:
            obj = reportings.first()
            file = obj.form_data.get("id_ccr_antt") or obj.number
            extension = "xlsx"
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            file_name = f"{file}.{extension}"
        else:
            all_roads = list(set([reporting.road_name for reporting in reportings]))
            all_roads.sort()
            file_name = _get_file_name_to_zip(all_roads)

        return file_name

    def __get_reportings_obj(self):
        query_set = Reporting.objects.filter(
            occurrence_type__name=self.class_name, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return list(query_set)

    def export(self):
        list_reporting = self.__get_reportings_obj()
        s3 = get_s3()
        obj = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()
        files = obj["files"]
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_surface_drainage_async_handler(reporter_dict: dict):
    reporter = XlsxHandlerReportSurfaceDrainageSheets.from_dict(reporter_dict)
    reporter.export()
