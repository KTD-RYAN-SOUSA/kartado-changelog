import re
import shutil
import tempfile
from datetime import datetime, timedelta
from typing import Dict, List
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType, RecordPanel
from apps.reportings.models import Reporting, ReportingFile, ReportingInReporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_image,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_brasilia_date,
    get_direction,
    get_end_km,
    get_identification,
    get_km,
    get_parent_serial,
    get_reporting_files,
    get_road_name,
    get_serial,
)
from helpers.apps.ccr_report_utils.workbook_utils import save_workbook, set_block_style
from helpers.apps.occurrence_records import convert_conditions_to_query_params
from helpers.strings import clean_latin_string


def get_conditions_date(panel_uuid: str, filter: str, data_pattern: str):
    conditions = (
        RecordPanel.objects.filter(uuid=panel_uuid)
        .only("conditions")[:1]
        .get()
        .conditions
    )
    conditions_query_params = str(
        convert_conditions_to_query_params(conditions["logic"])
    )
    pattern = r"\(\'" + filter + r"\'\, \'(" + data_pattern + r")\'\)"
    match = re.search(pattern, conditions_query_params).group(1)
    return match


def get_found_at_after_condition(panel_uuid: str):
    return get_conditions_date(panel_uuid, "found_at__date__gt", r"\d+\-\d+\-\d+")


def get_found_at_before_condition(panel_uuid: str):
    return get_conditions_date(panel_uuid, "found_at__date__lt", r"\d+\-\d+\-\d+")


class EmbankmentsAnnexFivePostProtocolXlsxHandler(object):

    __OCCURRENCE_TYPE_UUIDS: List[UUID] = []

    _RECUPERATION_NAMES: List[str] = [
        "Desassoreamento",
        "Esgotamento e destinação",
        "Hidrojateamento em drenagem",
        "Implantação Drenagem",
        "Limpeza/desobstrução manual",
        "Limpeza/desobstrução mecânica",
        "Pintura em elemento de drenagem",
        "Reconstrução Drenagem",
        "Recuperação de drenagem",
        "Reparo em drenagem",
    ]

    __DEFAULT_ROW_HEIGHT = 99.75

    __DEFAULT_BORDER_COLOR = "A5A5A5"

    __DEFAULT_FONT = Font(name="Calibri", size=11)
    __DEFAULT_SIDE = Side(border_style="thin", color=__DEFAULT_BORDER_COLOR)

    __DEFAULT_BORDER = Border(
        left=__DEFAULT_SIDE,
        right=__DEFAULT_SIDE,
        top=__DEFAULT_SIDE,
        bottom=__DEFAULT_SIDE,
    )

    __DEFAULT_ALIGNMENT = Alignment(
        vertical="center", horizontal="center", wrap_text=True
    )

    @classmethod
    def __get_linked_reportings(
        cls, reporting: Reporting, after: datetime
    ) -> List[Reporting]:
        try:
            query_set = (
                ReportingInReporting.objects.filter(
                    parent=reporting.uuid,
                    child__occurrence_type__name__in=EmbankmentsAnnexFivePostProtocolXlsxHandler._RECUPERATION_NAMES,
                    child__status__name="Executado",
                    reporting_relation__name="Recuperação",
                    child__executed_at__gte=after,
                )
                .only(
                    "child__uuid", "child__status__name", "child__occurrence_type__name"
                )
                .prefetch_related(
                    "child",
                    "child__occurrence_type",
                    "child__status",
                    "reporting_relation",
                )
                .order_by("uuid")
            )
            return [link.child for link in query_set]
        except Exception:
            return []

    @classmethod
    def __get_recuperation_images(
        cls,
        s3,
        dir: str,
        recuperations: List[Reporting],
        images_per_recuperation: int = 1,
        limit: int = 1,
    ) -> List[Image]:
        images: List[Image] = []

        for recuperation in recuperations:
            reporting_files = get_reporting_files(
                recuperation, True, "upload", "uuid", "datetime", "uploaded_at"
            )
            reporting_files = sorted(
                reporting_files,
                key=lambda reporting_file: (
                    reporting_file.datetime,
                    reporting_file.uploaded_at,
                ),
            )
            per_recuperation_limit = images_per_recuperation

            for reporting_file in reporting_files:
                image = None
                try:
                    image = get_image(
                        s3, dir, str(reporting_file.uuid), reporting_file, 150, 150
                    )
                except Exception as e:
                    print(e)

                if image is not None:
                    images.append(image)
                    per_recuperation_limit -= 1
                    limit -= 1
                if per_recuperation_limit == 0 or limit == 0:
                    break

            if limit == 0:
                break

        return images

    @classmethod
    def __get_condition_dre_sup(cls, reporting: Reporting) -> str:
        try:
            return new_get_form_data(reporting, "conditionDreSup")
        except Exception:
            return ""

    @classmethod
    def __get_condition_dre_prof(cls, reporting: Reporting) -> str:
        try:
            return new_get_form_data(reporting, "conditionDreProf")
        except Exception:
            return ""

    @classmethod
    def __get_drainage_pictures(
        cls, reporting: Reporting, limit: int
    ) -> List[ReportingFile]:
        try:
            drainage_pictures = new_get_form_data(reporting, "drainagePictures")
            drainage_picture_uuids = [
                UUID(uuid)
                for drainage_picture in drainage_pictures
                for uuid in drainage_picture["drainage_picture"]
            ][:limit]
            return list(
                ReportingFile.objects.filter(
                    uuid__in=drainage_picture_uuids, is_shared=True
                ).only("upload")
            )
        except Exception:
            return []

    def __init__(
        self,
        found_at,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.__sheet_target = sheet_target
        self.__sheet_reporting: Dict[str, List[Reporting]] = {}
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.occurrence_type = Reporting.objects.get(uuid=list_uuids[0]).occurrence_type

        self.reference_date: datetime = None
        try:
            self.reference_date = datetime.strptime(found_at["before"], "%Y-%m-%d")
        except Exception:
            self.reference_date = datetime.strptime(found_at["after"], "%Y-%m-%d")

        self.deadline = (self.reference_date + timedelta(days=30)).strftime("%d/%m/%Y")

        for e in OccurrenceType.objects.values("uuid", "name"):
            if (
                e["name"]
                in EmbankmentsAnnexFivePostProtocolXlsxHandler._RECUPERATION_NAMES
            ):
                EmbankmentsAnnexFivePostProtocolXlsxHandler.__OCCURRENCE_TYPE_UUIDS.append(
                    e["uuid"]
                )

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __append_entry(
        self,
        worksheet: Worksheet,
        reporting: Reporting,
        recuperation_reportings: List[Reporting],
    ) -> None:
        identification = get_identification(reporting, default="-")
        road_name = get_road_name(reporting)
        km = get_km(reporting)
        end_km = get_end_km(reporting)
        direction = get_direction(reporting)
        serial = get_serial(reporting)
        inventory_serial = get_parent_serial(reporting)
        recuperations = " - ".join(
            sorted(
                list(
                    {
                        recuperation_reporting.occurrence_type.name
                        for recuperation_reporting in recuperation_reportings
                    }
                )
            )
        )
        statuses = " - ".join(
            sorted(
                list(
                    {
                        recuperation_reporting.status.name
                        for recuperation_reporting in recuperation_reportings
                    }
                )
            )
        )

        text_entry: List[str | Cell] = [
            "",
            identification,
            road_name,
            km,
            end_km,
            direction,
            recuperations,
            self.deadline,
            statuses,
        ]
        text_entry.extend([""] * 7)
        text_entry.extend(
            [
                serial,
                inventory_serial,
            ]
        )
        worksheet.append(text_entry)

        last_row_index = worksheet.max_row
        condition_dre_sup = (
            EmbankmentsAnnexFivePostProtocolXlsxHandler.__get_condition_dre_sup(
                reporting
            )
        )
        condition_dre_prof = (
            EmbankmentsAnnexFivePostProtocolXlsxHandler.__get_condition_dre_prof(
                reporting
            )
        )

        set_block_style(
            worksheet,
            row_begin=last_row_index,
            row_end=last_row_index,
            col_begin="B",
            col_end="R",
            height=EmbankmentsAnnexFivePostProtocolXlsxHandler.__DEFAULT_ROW_HEIGHT,
            font=EmbankmentsAnnexFivePostProtocolXlsxHandler.__DEFAULT_FONT,
            border=EmbankmentsAnnexFivePostProtocolXlsxHandler.__DEFAULT_BORDER,
            alignment=EmbankmentsAnnexFivePostProtocolXlsxHandler.__DEFAULT_ALIGNMENT,
        )

        if condition_dre_sup != "Satisfatória" or condition_dre_prof != "Satisfatória":
            pictures_reporting_files = (
                EmbankmentsAnnexFivePostProtocolXlsxHandler.__get_drainage_pictures(
                    reporting, 2
                )
            )
            images = [
                get_image(
                    self.s3,
                    self.temp_dir,
                    "temp" + str(pictures_reporting_file.uuid) + "-file",
                    pictures_reporting_file,
                    150,
                    150,
                )
                for pictures_reporting_file in pictures_reporting_files
            ]

            try:
                insert_picture(
                    worksheet,
                    range_string=f"J{last_row_index}",
                    picture=images[0],
                    target=self.__sheet_target,
                    border_width=1,
                )
                insert_picture(
                    worksheet,
                    range_string=f"K{last_row_index}",
                    picture=images[1],
                    target=self.__sheet_target,
                    border_width=1,
                )
            except Exception:
                pass

        images = EmbankmentsAnnexFivePostProtocolXlsxHandler.__get_recuperation_images(
            self.s3, self.temp_dir, recuperation_reportings, 1, 5
        )

        curr_column = 12
        for image in images:
            try:
                insert_picture(
                    worksheet,
                    range_string=f"{get_column_letter(curr_column)}{last_row_index}",
                    picture=image,
                    target=self.__sheet_target,
                    border_width=1,
                )
                curr_column += 1
            except Exception:
                pass

    def __add_to_sheet_reporting(self, reporting: Reporting) -> None:

        if not (reporting.road_name) in self.__sheet_reporting:
            self.__sheet_reporting[reporting.road_name] = []

        self.__sheet_reporting[reporting.road_name].append(reporting)

    def __create_workbooks_files(self) -> List[str]:
        files: List[str] = []

        sorted_sheets = sorted(self.__sheet_reporting.items())
        for road_name, reportings in sorted_sheets:
            reportings.sort(
                key=lambda reporting: (get_identification(reporting), reporting.km)
            )

            workbook = load_workbook(
                "./fixtures/reports/ccr_embankments_annex_five_post_protocol.xlsx"
            )
            worksheet = workbook[workbook.sheetnames[0]]
            worksheet[
                "B1"
            ] = f"Pós-protocolo - Cronograma de Drenagem - Terraplenos e Estruturas de Contenção\n{reportings[0].road_name}"

            for reporting in reportings:

                linked_reportings = (
                    EmbankmentsAnnexFivePostProtocolXlsxHandler.__get_linked_reportings(
                        reporting, self.reference_date
                    )
                )
                if len(linked_reportings) == 0:
                    continue

                self.__append_entry(worksheet, reporting, linked_reportings)

            workbook_name = (
                "Anexo V Pos Protocolo - Cronograma de Drenagem - {}".format(road_name)
            )
            logo_path = get_logo_file(
                s3=self.s3,
                temp_prefix=self.temp_dir,
                reporting=reportings[0],
            )
            provider_logo_path = get_provider_logo_file(
                s3=self.s3,
                temp_prefix=self.temp_dir,
                reporting=reportings[0],
            )
            logo_config: dict = dict(
                path_image=logo_path,
                range_string="O1:P4",
                resize_method=ResizeMethod.ProportionalRight,
            )
            provider_logo_config: dict = dict(
                path_image=provider_logo_path,
                range_string="B1:B4",
                resize_method=ResizeMethod.ProportionalLeft,
            )
            insert_logo_and_provider_logo(
                worksheet=worksheet,
                logo_company=logo_config,
                provider_logo=provider_logo_config,
                target=self.__sheet_target,
            )

            worksheet.column_dimensions["Q"].hidden = True
            worksheet.column_dimensions["R"].width = worksheet.column_dimensions[
                "Q"
            ].width
            worksheet.column_dimensions["R"].hidden = True

            file = save_workbook(workbook_name, workbook)
            files.append(file)

        return files

    def execute(self) -> List[str]:

        query_set = (
            Reporting.objects.filter(
                occurrence_type=self.occurrence_type, uuid__in=self.list_uuids
            )
            .only(
                "uuid",
                "number",
                "road_name",
                "found_at",
                "km",
                "end_km",
                "direction",
                "form_data",
                "parent__number",
                "occurrence_type",
            )
            .prefetch_related("parent", "occurrence_type")
        )

        for report in query_set:
            self.__add_to_sheet_reporting(reporting=report)

        return self.__create_workbooks_files()


class CCREmbankmentsAnnexFivePostProtocol(CCRReport):
    def __init__(
        self,
        found_at=None,
        panel_uuid: str = None,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.found_at_filter: Dict = {}
        if found_at is None and not (panel_uuid is None):
            try:
                self.found_at_filter["before"] = get_found_at_before_condition(
                    panel_uuid
                )
            except Exception:
                pass
            try:
                self.found_at_filter["after"] = get_found_at_after_condition(panel_uuid)
            except Exception:
                pass
        else:
            self.found_at_filter = found_at
        super().__init__(uuids, report_format)

    def get_file_name(self):
        reference_date: datetime = None
        try:
            reference_date = datetime.strptime(
                self.found_at_filter["before"], "%Y-%m-%d"
            )
        except Exception:
            reference_date = datetime.strptime(
                self.found_at_filter["after"], "%Y-%m-%d"
            )
        reference_date = get_brasilia_date(reference_date)
        road_names = list(
            ReportingInReporting.objects.filter(
                parent__in=self.uuids,
                child__occurrence_type__name__in=EmbankmentsAnnexFivePostProtocolXlsxHandler._RECUPERATION_NAMES,
                child__status__name="Executado",
                reporting_relation__name="Recuperação",
                child__executed_at__gte=reference_date,
            )
            .only("parent__road_name")
            .distinct("parent__road_name")
            .values_list("parent__road_name", flat=True)
        )

        road_names.sort()
        file_name = "Anexo V Pos Protocolo - Cronograma de Drenagem - {}".format(
            "_".join(road_names)
        )
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        ext = "zip"
        if len(road_names) < 2:
            if self.report_format() == ReportFormat.PDF:
                ext = "pdf"
            else:
                ext = "xlsx"
        file_name = f"{file_name}.{ext}"
        return file_name

    def export(self):
        s3 = get_s3()
        files = EmbankmentsAnnexFivePostProtocolXlsxHandler(
            found_at=self.found_at_filter,
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_embankments_annex_five_post_protocol_async_handler(
    reporter_dict: dict,
):
    reporter = CCREmbankmentsAnnexFivePostProtocol.from_dict(reporter_dict)
    reporter.export()
