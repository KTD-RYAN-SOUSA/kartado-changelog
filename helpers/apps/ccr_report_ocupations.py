import tempfile
from datetime import datetime
from os.path import isfile
from typing import List
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_direction_letter,
    get_km_plus_meter,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import (
    get_form_array_iterator,
    new_get_form_data,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_image,
    get_logo_file,
    get_provider_logo_file,
    insert_picture,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import ThreadExecutor, synchronized_request_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, deep_keys_to_snake_case


class XlsxHandlerOccupations:
    _LOGO_CELL = "S1:U4"
    _PROVIDER_LOGO_CELL = "A1:G4"

    _LOGO_CELL_2 = "S51:U54"
    _PROVIDER_LOGO_CELL_2 = "A51:G54"

    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company"
        )
        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.__init_wb()

        self.static_fields = {
            "name": "F7",
            "address": "F8",
            "city_of_prop": "F9",
            "state": "R9",
            "owner_fone": "F10",
            "sheet": "F11",
            "cartory": "F12",
            "telephone": "V10",
            "postal_code": "V9",
            "celular": "V8",
            "matricule": "V11",
            "road_name": "E15",
            "km": "N15",
            "direction": "V15",
            "city": "E16",
            "uf": "V16",
            "latitude": "I17",
            "longitude": "U17",
            "lane_occupied": {
                "Faixa de Domínio": "F20",
                "Faixa não Edificante": "K20",
                "Faixa de Domínio e Faixa não Edificante": "S20",
            },
            "occupation": {
                "Comercial": "F21",
                "Industrial": "K21",
                "Residencial": "O21",
                "Passivo Ambiental": "F22",
                "Lavoura": "S21",
                "Público": "W21",
                "Outros": "K22",
            },
            "zone": {"Rural": "F23", "Urbana": "K23"},
            "situation": {"Regular": "F24", "Irregular": "K24"},
            "occupation_kind": {
                "Telecomunicação": "F25",
                "Ponto de Ônibus": "K25",
                "Elétrica": "O25",
                "Painel": "S25",
                "Edificação": "W25",
                "Outros": "F26",
            },
            "characteristics": {
                "Transversal": "F27",
                "Longitudinal": "K27",
                "Pontual": "O27",
            },
            "lane_extension": "H30",
            "lane_extension_not_edified": "H31",
            "lane_extension_edified_area": "U30",
            "lane_extension_not_edified_area": "U31",
            "observations": "A34",
            "photo_06": "A39",
            "photo_07": "N39",
            "image_croqui": "A67",
            "construction_kind_occupation": {
                "Construção em Madeira": "F57",
                "Construção Metálica": "F58",
                "Construção em Alvenaria": "O57",
                "Construção Mista": "O58",
            },
            "utilization_condition": {
                "Exclusivamente Residencial": "F61",
                "Exclusivamente Industrial": "F62",
                "Residencial e Industrial": "F63",
                "Exclusivamente Comercial": "O61",
                "Agropecuário": "O62",
                "Residencial e Comercial": "O63",
                "Pública": "O64",
                "Outros": "F64",
            },
            "id_antt": {"sheet_1": "W4", "sheet_2": "Y54"},
            "others_occupation": "N22",
            "others_ocupation_kind": "J26",
            "others_utilization_condition": "H64",
            "others_construction_kind": "A59",
        }

        self.action_static_fields = {
            "id_antt": "W4",
            "name": "F7",
            "road_name": "E10",
            "city": "E11",
            "km": "N10",
            "direction": "V10",
            "uf": "V11",
            "latitude": "I12",
            "longitude": "U12",
        }

    @classmethod
    def _insert_logos(
        cls,
        s3,
        temp_dir: str,
        sheet_target: SheetTarget,
        workbook: Workbook,
        reporting: Reporting,
    ) -> None:
        logo = get_logo_file(s3, temp_dir, reporting)
        provider_logo = get_provider_logo_file(s3, temp_dir, reporting)

        for worksheet in workbook.worksheets:
            try:
                insert_picture_2(
                    worksheet,
                    cls._LOGO_CELL,
                    Image(logo),
                    sheet_target,
                    border_width=(2, 2, 2, 2),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass
            try:
                insert_picture_2(
                    worksheet,
                    cls._PROVIDER_LOGO_CELL,
                    Image(provider_logo),
                    sheet_target,
                    border_width=(2, 2, 2, 2),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass

        worksheet = workbook.worksheets[0]
        try:
            insert_picture_2(
                worksheet,
                cls._LOGO_CELL_2,
                Image(logo),
                sheet_target,
                border_width=(2, 2, 2, 2),
                resize_method=ResizeMethod.ProportionalCentered,
            )
        except Exception:
            pass
        try:
            insert_picture_2(
                worksheet,
                cls._PROVIDER_LOGO_CELL_2,
                Image(provider_logo),
                sheet_target,
                border_width=(2, 2, 2, 2),
                resize_method=ResizeMethod.ProportionalCentered,
            )
        except Exception:
            pass

    def create_dict(self, reporting: Reporting):
        self.photo_06 = None
        self.photo_07 = None
        self.imagem_croqui = None

        result_occupation = new_get_form_data(reporting, "ocupation")

        occupation = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["occupation"], result_occupation
        )

        ocupation_others = (
            new_get_form_data(reporting, "ocupationOthers")
            if result_occupation == "Outros"
            else ""
        )

        state = new_get_form_data(reporting, "state", default="")

        city = new_get_form_data(reporting, "cityOfProp", default="")

        name = new_get_form_data(reporting, "name", default="")

        id_antt = new_get_form_data(reporting, "idCcrAntt", default="")

        county = new_get_form_data(reporting, "city", default="")

        uf = new_get_form_data(reporting, "uf", default="")

        latitude = new_get_form_data(reporting, "lat", default="")

        longitude = new_get_form_data(reporting, "long", default="")

        matricule = new_get_form_data(reporting, "matricule", default="")

        cartory = new_get_form_data(reporting, "cartory", default="")

        owner_phone = new_get_form_data(reporting, "ownerFone", default="")

        sheet = new_get_form_data(reporting, "sheet", default="")

        address = new_get_form_data(reporting, "address", default="")

        telephone_access = new_get_form_data(reporting, "celular", default="")

        postal_code = new_get_form_data(reporting, "postalCode", default="")

        telephone_delivery = new_get_form_data(reporting, "telephone", default="")

        observations = new_get_form_data(reporting, "observations", default="")

        result_lane_occupied = new_get_form_data(reporting, "laneOccupied", default="")
        lane_occupied = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["lane_occupied"], result_lane_occupied
        )

        result_zone = new_get_form_data(reporting, "zone")
        zone = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["zone"], result_zone
        )

        result_situation = new_get_form_data(reporting, "situation")
        situation = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["situation"], result_situation
        )

        result_occupation_kind = new_get_form_data(reporting, "ocupationKind")

        occupation_kind = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["occupation_kind"], result_occupation_kind
        )

        outhers_ocupation_kind = (
            new_get_form_data(reporting, "ocupationKindOthers")
            if result_occupation_kind == "Outros"
            else ""
        )

        result_characteristics = new_get_form_data(reporting, "characteristics")
        characteristics = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["characteristics"], result_characteristics
        )

        lane_extension = new_get_form_data(reporting, "laneExtension", default="-")

        lane_extension_not_edified = new_get_form_data(
            reporting, "laneExtensionNotEdified", default="-"
        )

        lane_extension_edified_area = new_get_form_data(
            reporting, "laneExtensionEdifiedArea", default="-"
        )

        lane_extension_not_edified_area = new_get_form_data(
            reporting, "laneExtensionNotEdifiedArea", default="-"
        )

        form_data = deep_keys_to_snake_case(reporting.form_data)

        reporting_files = form_data.get("imagens_ocupacao_faixa_dominio", [])

        picture_uuids: List[List[UUID]] = []

        it = get_form_array_iterator(reporting, "imagensOcupacaoFaixaDominio")
        try:
            while True:
                imagensOcupacao = it.get("imagensOcupacao")
                uuids = []
                if isinstance(imagensOcupacao, list):
                    for imagemOcupacao in imagensOcupacao:
                        try:
                            uuids.append(UUID(imagemOcupacao))
                        except Exception as e:
                            print(e)
                picture_uuids.append(uuids)
                it.inc()
        except Exception as e:
            print(e)

        reporting_files: List[ReportingFile] = [
            r for r in reporting.reporting_files.filter(is_shared=True)
        ]
        reporting_file_lists: List[List[ReportingFile]] = []
        for uuids in picture_uuids:
            curr_rfs = [r for r in reporting_files if r.uuid in uuids]
            curr_rfs = sorted(curr_rfs, key=lambda r: (r.datetime, r.uploaded_at))
            reporting_file_lists.append(curr_rfs)

        self.photo_06 = None
        self.photo_07 = None
        for reporting_file_list in reporting_file_lists:
            for rf in reporting_file_list:
                try:
                    image = get_image(
                        s3=self.s3,
                        dir=self.temp_file,
                        image_name=rf.upload.name,
                        reporting_file=rf,
                    )
                    if self.photo_06 is None:
                        self.photo_06 = image
                    elif self.photo_07 is None:
                        self.photo_07 = image
                    else:
                        break
                except Exception:
                    pass

                if self.photo_06 is not None and self.photo_07 is not None:
                    break
            if self.photo_06 is not None and self.photo_07 is not None:
                break

        croqui_files = form_data.get("imagens_croqui_faixa_dominio", "")
        if croqui_files:
            try:
                imagem_croqui = (
                    ReportingFile.objects.filter(
                        pk__in=croqui_files[0].get("imagens_croqui", []),
                    )
                    .order_by("datetime")
                    .first()
                )
                if imagem_croqui is not None:
                    self.imagem_croqui = get_image(
                        s3=self.s3,
                        dir=self.temp_file,
                        image_name=imagem_croqui.upload.name,
                        reporting_file=imagem_croqui,
                        width=337,
                        height=242,
                    )
            except Exception:
                self.imagem_croqui = ""

        result_construction_kind_occupation = new_get_form_data(
            reporting, "constructionKindOccupation"
        )

        outher_construction_occupation = (
            new_get_form_data(reporting, "qualFxDom")
            if result_construction_kind_occupation == "Outro"
            else ""
        )

        construction_kind_occupation = (
            XlsxHandlerOccupations.__result_ocupation_condition(
                self.static_fields["construction_kind_occupation"],
                result_construction_kind_occupation,
            )
        )

        result_utilization_condition = new_get_form_data(
            reporting, "utilizationCondition"
        )
        utilization_condition = XlsxHandlerOccupations.__result_ocupation_condition(
            self.static_fields["utilization_condition"], result_utilization_condition
        )

        utilization_condition_other = new_get_form_data(
            reporting, "utilizationConditionOther"
        )
        result_utilization_condition_other = "Outros"
        if result_utilization_condition == "Outros":
            result_utilization_condition_other = (
                f"Outros({utilization_condition_other})"
            )

        result_road_name = reporting.road_name
        road_name = result_road_name if result_road_name else ""

        direction = get_custom_option(reporting, "direction")

        result_km = reporting.km
        km = result_km if result_km else ""

        result_therapy = reporting.form_data.get("therapy")
        therapy = result_therapy if result_therapy else ""

        data = {
            "id_antt": id_antt,
            "name": name,
            "address": address,
            "city_of_prop": city,
            "state": state,
            "owner_fone": owner_phone,
            "sheet": sheet,
            "cartory": cartory,
            "telephone": telephone_access,
            "postal_code": postal_code,
            "celular": telephone_delivery,
            "matricule": matricule,
            "road_name": road_name.split(" ")[0] if road_name else "",
            "km": get_km_plus_meter(km),
            "direction": direction,
            "city": county,
            "uf": uf,
            "latitude": latitude,
            "longitude": longitude,
            "lane_occupied": lane_occupied,
            "occupation": occupation,
            "zone": zone,
            "situation": situation,
            "occupation_kind": occupation_kind,
            "characteristics": characteristics,
            "lane_extension": lane_extension,
            "lane_extension_not_edified": lane_extension_not_edified,
            "lane_extension_edified_area": lane_extension_edified_area,
            "lane_extension_not_edified_area": lane_extension_not_edified_area,
            "observations": observations,
            "photo_06": self.photo_06,
            "photo_07": self.photo_07,
            "construction_kind_occupation": construction_kind_occupation,
            "utilization_condition": utilization_condition,
            "therapy": therapy,
            "image_croqui": self.imagem_croqui,
            "others_occupation": ocupation_others,
            "others_ocupation_kind": outhers_ocupation_kind,
            "others_utilization_condition": result_utilization_condition_other,
            "others_construction_kind": outher_construction_occupation,
            "reporting": reporting,
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""

        return data

    def fill_actions_date_and_description(self, data_actions: list):
        row = 16
        limit_rows = 0
        count = 0
        step_rows = 7
        for action in data_actions:
            sum_coor = row + count
            description_coordinate = (
                f"D{sum_coor}" if limit_rows <= 6 else f"Q{sum_coor}"
            )
            date_coordinate = f"A{sum_coor}" if limit_rows <= 6 else f"N{sum_coor}"
            column_letter_description = self.action_worksheet[
                description_coordinate
            ].column_letter
            column_letter_date = self.action_worksheet[date_coordinate].column_letter

            cell_merge_description = (
                f"{description_coordinate}:M{sum_coor}"
                if column_letter_description == "D"
                else f"{description_coordinate}:Z{sum_coor}"
            )
            self.action_worksheet.merge_cells(cell_merge_description)

            cell_merge_date = (
                f"{date_coordinate}:C{sum_coor}"
                if column_letter_date == "A"
                else f"{date_coordinate}:P{sum_coor}"
            )

            self.action_worksheet.merge_cells(cell_merge_date)

            occurrence = action.get("occurrence_type", "")
            if occurrence:
                occurrence = OccurrenceType.objects.get(uuid=occurrence)

            self.action_worksheet[description_coordinate] = (
                occurrence.name if occurrence else ""
            )

            XlsxHandlerOccupations.__format_fonts(
                cell=self.action_worksheet[description_coordinate],
                size=11,
                color="0070C0",
                bold=True,
            )

            date = action.get("action_date", "")
            if date:
                date = datetime.strptime(date, "%Y-%m-%dT%H:%M:%S.%fZ").strftime(
                    "%d/%m/%Y"
                )
            self.action_worksheet[date_coordinate] = date
            XlsxHandlerOccupations.__format_fonts(
                cell=self.action_worksheet[date_coordinate],
                size=11,
                color="0070C0",
                bold=True,
            )

            if limit_rows == 12:
                limit_rows = 0
                row += step_rows
            limit_rows += 1
            count = 0 if count > 6 else +1

    def fill_action_worksheet(self, key, value):
        if key in [_ for _ in self.action_static_fields.keys()]:
            key_value = self.action_static_fields[key]
            self.action_worksheet[key_value] = value

            if key not in ("name", "latitude", "longitude"):
                return XlsxHandlerOccupations.__format_fonts(
                    cell=self.action_worksheet[key_value],
                    size=11,
                    color="0070C0",
                    bold=True,
                )
            XlsxHandlerOccupations.__format_fonts(
                cell=self.action_worksheet[key_value],
                horizontal="left",
                size=11,
                color="0070C0",
                bold=True,
            )

    def __init_wb(self) -> None:
        self.wb = load_workbook("./fixtures/reports/ccr_occupation.xlsx")
        self._worksheet = self.wb["Ocupação"]
        self.action_worksheet = self.wb["ACOES"]
        self.photo_06 = None
        self.photo_07 = None
        self.imagem_croqui = None

    def fill_sheet(self, *, data_list: list):
        convert_executor: ThreadExecutor = None
        if self.__report_format == ReportFormat.PDF:
            convert_executor = ThreadExecutor(50)
        count = 0
        list_files = list()
        for values in data_list:
            reporting = values.pop("reporting")
            for key, value in values.items():
                if key in [
                    "lane_occupied",
                    "occupation",
                    "zone",
                    "situation",
                    "occupation_kind",
                    "characteristics",
                    "construction_kind_occupation",
                    "utilization_condition",
                ]:
                    for _key, _value in value.items():
                        key_value = self.static_fields[key][_key]
                        self.__insert_status_values(
                            cell=key_value, value=_value, key=key
                        )

                elif key in ["photo_06"] and value:
                    insert_picture(
                        worksheet=self._worksheet,
                        range_string="A39:M49",
                        picture=value,
                        target=self.__sheet_target,
                    )

                elif key in ["photo_07"] and value:
                    insert_picture(
                        worksheet=self._worksheet,
                        range_string="N39:Z49",
                        picture=value,
                        target=self.__sheet_target,
                    )
                elif key in ["image_croqui"] and value:
                    insert_picture(
                        worksheet=self._worksheet,
                        range_string="A67:Z93",
                        picture=value,
                        target=self.__sheet_target,
                    )
                elif key == "therapy":
                    self.fill_actions_date_and_description(value)
                elif key == "id_antt":
                    self.fill_ids_in_sheets(self.static_fields["id_antt"], value)
                else:
                    self.fill_action_worksheet(key, value)
                    key_value = self.static_fields[key]
                    self._worksheet[key_value] = value
                    self.custom_formats(key, self._worksheet[key_value])

            count += 1
            road_name = values["road_name"]
            uf_access = values["uf"]
            km = values["km"]
            direction = (
                get_direction_letter(values["direction"])
                if values["direction"]
                else "NA"
            )
            file_name = f"{uf_access} {road_name} {km} {direction}"
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
            file_path = f"/tmp/{file_name}.xlsx"

            i = 1
            while isfile(file_path):
                file_path = f"/tmp/{file_name} ({i}).xlsx"
                i += 1

            temp_dir = tempfile.mkdtemp()
            self._insert_logos(
                self.s3, temp_dir, self.__sheet_target, self.wb, reporting
            )
            self.wb.save(file_path)
            self.__init_wb()
            if self.__report_format == ReportFormat.PDF:
                convert_executor.submit(synchronized_request_pdf, file_path)
            list_files.append(file_path)
            list_files = list(set(list_files))
            list_files.sort()
        if self.__report_format == ReportFormat.PDF:
            list_files = list(set(convert_executor.get()))
            list_files.sort()
        return list_files

    def fill_ids_in_sheets(self, static_field, value):
        sheet_1 = static_field["sheet_1"]
        sheet_2 = static_field["sheet_2"]
        self.action_worksheet["W4"] = value
        self._worksheet[sheet_1] = value
        self._worksheet[sheet_2] = value

    def custom_formats(self, key, cell):
        if key in (
            "name",
            "address",
            "city_of_prop",
            "titular_foil",
            "telephone_delivery",
            "owner_fone",
            "sheet",
            "cartory",
            "telephone",
            "postal_code",
            "celular",
            "matricule",
            "state",
            "latitude",
            "longitude",
            "others_occupation",
            "others_ocupation_kind",
            "others_construction_kind",
        ):
            XlsxHandlerOccupations.__format_fonts(
                cell=cell, size=11, horizontal="left", color="0070C0", bold=True
            )

        elif key == "others_utilization_condition":
            XlsxHandlerOccupations.__format_fonts(cell=cell, size=11, horizontal="left")
        elif key == "observations":
            XlsxHandlerOccupations.__format_fonts(
                cell=cell,
                size=11,
                horizontal="left",
                vertical="top",
                color="0070C0",
                bold=True,
            )
        else:
            XlsxHandlerOccupations.__format_fonts(
                cell=cell, size=11, color="0070C0", bold=True
            )

    def __insert_status_values(self, cell, value, key):
        if value:
            self._worksheet[cell] = "X"
            XlsxHandlerOccupations.__format_fonts(
                cell=self._worksheet[cell],
                size=11,
                horizontal="center",
                color="0070C0",
            )
        else:
            self._worksheet[cell] = ""

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Calibri",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        color="FF000000",
    ) -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold, color=color)

    @classmethod
    def __result_ocupation_condition(
        cls, ocupation_conditions: dict, result_get_form_data
    ):
        conditions_dict = dict()

        for occupation_condition in ocupation_conditions.keys():
            if result_get_form_data == occupation_condition:
                conditions_dict[occupation_condition] = "X"
            else:
                conditions_dict[occupation_condition] = ""
        return conditions_dict

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]
        data = [self.create_dict(reporting=reporting) for reporting in list_reporting]

        result_file = self.fill_sheet(data_list=data)

        return result_file


class CCROcupation(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = "Relatórios ANTT de Monitoração de Ocupações"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerOccupations(
            s3, self.uuids, self.sheet_target(), self.report_format()
        ).execute()

        result_file = "/tmp/" + self.file_name
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_occupation_async_handler(reporter_dict: dict):
    reporter = CCROcupation.from_dict(reporter_dict)
    reporter.export()
