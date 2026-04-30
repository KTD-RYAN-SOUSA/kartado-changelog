import json
import os
import random
import shutil
import string
from datetime import timedelta

import boto3
import pytz
from django.conf import settings
from django.utils.timezone import now
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from zappa.asynchronous import task

from apps.files.models import GenericFile
from apps.service_orders.models import ProcedureResource
from helpers.dates import utc_to_local
from RoadLabsAPI.settings import credentials


class ProcedureResourceExport:
    def __init__(
        self,
        file_uuid="",
        filename=None,
        object_name=None,
        queryset=[],
        company_name=None,
    ):
        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )
        if file_uuid:
            generic_file = GenericFile.objects.get(uuid=file_uuid)
            uuid_list = json.loads(generic_file.file.read())
            prefetch_fields_list = [
                "service_order_resource",
                "service_order_resource__contract",
                "service_order_resource__contract__firm",
                "service_order_resource__contract__subcompany",
                "service_order_resource__entity",
                "resource",
                "created_by",
                "measurement_bulletin",
                "approved_by",
                "reporting",
            ]
            self.queryset = ProcedureResource.objects.filter(
                pk__in=set(uuid_list)
            ).prefetch_related(*prefetch_fields_list)
        elif queryset:
            self.queryset = queryset
        self.company_name = company_name
        if filename:
            self.filename = filename
        else:
            self.filename = self.get_filename()

        if object_name:
            self.object_name = object_name
        else:
            self.object_name = self.get_object_name()
        self.approval_status_translation = {
            "APPROVED_APPROVAL": "Aprovado",
            "DENIED_APPROVAL": "Rejeitado",
            "WAITING_APPROVAL": "Pendente",
        }

    def get_random_string(self):
        return "".join(
            random.SystemRandom().choice(string.ascii_lowercase + string.digits)
            for _ in range(10)
        )

    def get_company_name(self):
        company_name = self.queryset.resource.company.name

        return company_name

    def get_filename(self):
        filename = f'[Kartado] - Relatório de uso - [{self.company_name.replace(".", "").replace("/", "")}]'

        return f"{filename}.xlsx"

    def get_object_name(self):
        random_string = self.get_random_string()
        object_name = "media/private/{}_{}.xlsx".format(
            self.filename.split(".xlsx")[0], random_string
        )

        return object_name

    def get_s3_url(self):
        empty = {"url": "", "name": ""}

        url = self.s3.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                "Key": self.object_name,
            },
        )

        if not url:
            return empty

        return url

    def upload_file(self, path):
        expires = now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)

        try:
            self.s3.upload_file(
                path,
                settings.AWS_STORAGE_BUCKET_NAME,
                self.object_name,
                ExtraArgs={"Expires": expires},
            )
        except Exception:
            return

    def copy_and_rename(self, old_folder, new_folder, temp_file):
        os.makedirs(new_folder, exist_ok=True)
        shutil.copy(old_folder + temp_file, new_folder + temp_file)
        os.rename(
            new_folder + temp_file,
            new_folder + self.filename + ".xlsx",
        )
        return

    def return_file_path(self, new_folder):
        return new_folder + self.filename + ".xlsx"

    def load_file(self, new_file_path):
        wb = load_workbook(filename=new_file_path)
        return wb

    def get_sheetname(self, resource):
        if resource.service_order_resource.contract.subcompany:
            first_part = resource.service_order_resource.contract.subcompany.name
        elif resource.service_order_resource.contract.firm:
            first_part = resource.service_order_resource.contract.firm.name
        else:
            first_part = ""

        return f"{first_part} - {self.company_name}"

    def get_item_data(self, procedure_resource):
        resource = procedure_resource.resource
        service_order_resource = procedure_resource.service_order_resource

        name = resource.name if resource else ""
        unit = resource.unit if resource else ""
        entity_name = (
            service_order_resource.entity.name
            if service_order_resource and service_order_resource.entity
            else ""
        )
        creation_date = utc_to_local(procedure_resource.creation_date).strftime(
            "%d/%m/%Y"
        )
        created_by = (
            procedure_resource.created_by.get_full_name()
            if procedure_resource.created_by
            else ""
        )
        reporting_number = (
            procedure_resource.reporting.number if procedure_resource.reporting else ""
        )

        approval_status = self.approval_status_translation.get(
            procedure_resource.approval_status, ""
        )
        approval_date = (
            utc_to_local(procedure_resource.approval_date).strftime("%d/%m/%Y %H:%M")
            if procedure_resource.approval_date
            else ""
        )
        bulletin_number = (
            procedure_resource.measurement_bulletin.number
            if procedure_resource.measurement_bulletin
            else ""
        )
        return [
            name,
            procedure_resource.amount,
            unit,
            entity_name,
            creation_date,
            created_by,
            procedure_resource.unit_price,
            procedure_resource.total_price,
            reporting_number,
            approval_status,
            approval_date,
            bulletin_number,
        ]

    def fill_workbook(self):
        relatorio = self.wb["relatorio"]

        start_line = 2

        for line_index, item in enumerate(self.queryset):
            procedure_resource_data = self.get_item_data(item)
            for column_index, data in enumerate(procedure_resource_data):
                data_cell = relatorio.cell(
                    row=line_index + start_line, column=column_index + 1
                )
                data_cell.value = data
                if get_column_letter(column_index + 1) in ["E", "K"]:
                    data_cell.number_format = "dd/mm/yyyy"
                if get_column_letter(column_index + 1) in ["G", "H"]:
                    data_cell.number_format = "R$ #,##0.00"

        sheetname = self.get_sheetname(self.queryset[0])
        relatorio.title = sheetname

        return

    def generate_file(self):
        old_folder = "apps/service_orders/templates/"
        new_folder = "/tmp/procedure_resource/"
        temp_file = "relatorio_de_exportacao.xlsx"

        self.copy_and_rename(old_folder, new_folder, temp_file)
        new_file_path = self.return_file_path(new_folder)
        self.wb = self.load_file(new_file_path)
        self.fill_workbook()

        self.wb.save(new_file_path)
        self.upload_file(new_file_path)
        os.remove(new_file_path)
        os.rmdir(new_folder)


@task
def procedure_resource_export_async(file_uuid, filename, object_name, company_name):
    procedure_resource_object = ProcedureResourceExport(
        file_uuid, filename, object_name, company_name=company_name
    )
    procedure_resource_object.generate_file()
