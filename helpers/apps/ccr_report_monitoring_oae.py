import io
import re
import shutil
import tempfile
from typing import List
from uuid import uuid4
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import pixels_to_EMU
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_picture,
    get_logo_file,
    insert_picture,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.apps.ccr_report_utils.workbook_utils import set_zoom
from helpers.kartado_excel.worksheet import copy_sheet_with_settings
from helpers.strings import (
    deep_keys_to_snake_case,
    format_km,
    int_set_zero_prefix,
    keys_to_snake_case,
    remove_ext_in_filename,
    remove_random_string_file_name_in_upload,
)

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
bold = Font(bold=True)
italic = Font(italic=True)
italic_bold = Font(italic=True, bold=True)


def get_file_name(reporting):
    inspection_year_campaign = keys_to_snake_case(reporting.form_data).get(
        "inspection_year_campaign", ""
    )
    road_number = re.sub(r"\D", "", reporting.road_name)
    n_oae = new_get_form_data(reporting, "oaeNumeroCodigoObra", default="")
    try:
        n_oae = f"{int(n_oae):03}"
    except Exception:
        pass
    uf = reporting.road_name[-2:]
    file_name = f"OAE{inspection_year_campaign}{road_number}{n_oae}{uf}"
    return file_name


class XlsxHandler:
    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.list_reporting = list_reporting
        self._sheet_target = sheet_target
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self._xlsx_file = "./fixtures/reports/ccr_report_monitoring_oae.xlsx"
        self._workbook = load_workbook(self._xlsx_file)
        self._sheetnames = self._workbook.sheetnames
        self._worksheet = self._workbook["FICHA 1"]
        self._worksheets = self._workbook.worksheets
        self._copy_antt_logo = True

    def __insert_img_size(
        self,
        image: str,
        row_init: int,
        row_end: int,
        col_init: int,
        col_end: int,
        width: int,
        height: int,
    ):
        img = Image(image)
        img.width = int(width)
        img.height = int(height)

        left = round((col_init + col_end) / 2)
        top = round((row_init + row_end) / 2)

        self._worksheet.add_image(
            img, f"{self._worksheet.cell(row=top, column=left).coordinate}"
        )

    def __insert_img_with_absolute_anchor(
        self, image: str, width: int, height: int, pos_x: int, pos_y: int
    ):
        img = Image(image)
        img.width = int(width)
        img.height = int(height)

        p2e = pixels_to_EMU
        h, w = img.height, img.width

        # position = XDRPoint2D(p2e(660), p2e(995))
        position = XDRPoint2D(p2e(pos_x), p2e(pos_y))
        size = XDRPositiveSize2D(p2e(w), p2e(h))

        img.anchor = AbsoluteAnchor(pos=position, ext=size)

        self._worksheet.add_image(img)

    def __insert_bold(self, row: int, col: str, value: str, horizontal="left"):
        insert_centered_value(
            worksheet=self._worksheet,
            value=value,
            cell=f"{col}{row}",
            horizontal=horizontal,
            wrapText=False,
        )
        self._worksheet[f"{col}{row}"].font = bold

    def __mark_x(self, row: int, col: str, horizontal="center"):
        self.__insert_bold(
            row=row,
            col=col,
            value="X",
            horizontal=horizontal,
        )

    def __insert_value(
        self,
        row: int,
        col: str,
        value: str,
        horizontal="left",
        font: Font = None,
        alignment: Alignment = None,
        wrapText: bool = False,
    ):
        try:
            if value is not None:
                str_val = str(value).strip()
                if len(str_val) == 0:
                    value = "-"
        except Exception:
            value = "-"

        insert_centered_value(
            worksheet=self._worksheet,
            value=value,
            cell=f"{col}{row}",
            horizontal=horizontal,
            wrapText=wrapText,
        )
        if font:
            self._worksheet[f"{col}{row}"].font = font
        if alignment:
            self._worksheet[f"{col}{row}"].alignment = alignment

    def __set_ficha_1(self, data: dict):
        self._worksheet = self._workbook["FICHA 1"]
        self.__insert_value(
            row=2, col="AF", value=data.get("codigo_obra"), horizontal="center"
        )
        self.__insert_value(row=5, col="B", value=data.get("codigo_agencia"))
        self.__insert_value(row=5, col="G", value=data.get("denominacao"))
        self.__insert_value(row=5, col="W", value=data.get("road_name_short"))
        self.__insert_value(row=5, col="Z", value=data.get("km"))
        self.__insert_value(row=5, col="AE", value=data.get("direction"))
        self.__insert_value(row=7, col="B", value=data.get("executed_at"))
        self.__insert_value(row=7, col="K", value=data.get("zona"))
        self.__insert_value(row=7, col="O", value=data.get("utm_e1"))
        self.__insert_value(row=7, col="W", value=data.get("utm_e2"))
        self.__insert_value(row=8, col="O", value=data.get("utm_e3"))
        self.__insert_value(row=8, col="W", value=data.get("utm_e4"))
        self.__insert_value(row=7, col="S", value=data.get("utm_n1"))
        self.__insert_value(row=7, col="Z", value=data.get("utm_n2"))
        self.__insert_value(row=8, col="S", value=data.get("utm_n3"))
        self.__insert_value(row=8, col="Z", value=data.get("utm_n4"))

        condicoes_estabilidade = data.get("condicoes_estabilidade")
        condicoes_estabilidade_target = None

        if str(condicoes_estabilidade) == "1":
            condicoes_estabilidade_target = "D"
        elif str(condicoes_estabilidade) == "2":
            condicoes_estabilidade_target = "F"
        elif str(condicoes_estabilidade) == "3":
            condicoes_estabilidade_target = "H"

        if isinstance(condicoes_estabilidade_target, str):
            self.__mark_x(row=14, col=condicoes_estabilidade_target)

        condicoes_conservacao = data.get("condicoes_conservacao")
        condicoes_conservacao_target = None

        if str(condicoes_conservacao) == "1":
            condicoes_conservacao_target = "R"
        elif str(condicoes_conservacao) == "2":
            condicoes_conservacao_target = "T"
        elif str(condicoes_conservacao) == "3":
            condicoes_conservacao_target = "V"
        elif str(condicoes_conservacao) == "4":
            condicoes_conservacao_target = "X"

        if isinstance(condicoes_conservacao_target, str):
            self.__mark_x(row=14, col=condicoes_conservacao_target)

        nivel_vibracao_tabuleiro = data.get("nivel_vibracao_tabuleiro")
        nivel_vibracao_tabuleiro_target = None

        if str(nivel_vibracao_tabuleiro) == "1":
            nivel_vibracao_tabuleiro_target = "D"
        elif str(nivel_vibracao_tabuleiro) == "2":
            nivel_vibracao_tabuleiro_target = "F"
        elif str(nivel_vibracao_tabuleiro) == "3":
            nivel_vibracao_tabuleiro_target = "H"

        if isinstance(nivel_vibracao_tabuleiro_target, str):
            self.__mark_x(
                row=16,
                col=nivel_vibracao_tabuleiro_target,
            )

        inspecao_especializada = data.get("inspecao_especializada")
        inspecao_especializada_target = None

        if inspecao_especializada is True:
            inspecao_especializada_target = "L"
        elif inspecao_especializada is False:
            inspecao_especializada_target = "N"

        if isinstance(inspecao_especializada_target, str):
            self.__mark_x(row=18, col=inspecao_especializada_target)

        urgente_antt = data.get("urgente_antt")
        urgente_antt_target = None

        if str(urgente_antt) == "1":
            urgente_antt_target = "T"
        elif str(urgente_antt) == "2":
            urgente_antt_target = "V"

        if isinstance(urgente_antt_target, str):
            self.__mark_x(row=18, col=urgente_antt_target)

        houve_anteriormente = data.get("houve_anteriormente")
        houve_anteriormente_target = None

        if houve_anteriormente is True:
            houve_anteriormente_target = "D"
        elif houve_anteriormente is False:
            houve_anteriormente_target = "F"

        if isinstance(houve_anteriormente_target, str):
            self.__mark_x(row=20, col=houve_anteriormente_target)

        self.__insert_value(
            row=22,
            col="F",
            value=data.get("historico_intervencoes_realizadas"),
            alignment=Alignment(vertical="top", horizontal="left", wrapText=True),
        )
        self.__insert_value(row=25, col="D", value=data.get("observacoes_adicionais"))

        self.__insert_value(
            row=17,
            col="AF",
            value=data.get("nota_tecnica_comentarios_gerais"),
            horizontal="center",
        )

        self.__insert_value(
            row=30,
            col="O",
            value=data.get("nota_tecnica_laje"),
            horizontal="center",
        )

        buraco_abertura1_laje = data.get("buraco_abertura1_laje") or []
        if "1" in buraco_abertura1_laje:
            self.__mark_x(row=32, col="D")
        if "2" in buraco_abertura1_laje:
            self.__mark_x(row=32, col="I")

        self.__insert_value(
            row=32, col="Q", value=data.get("buraco_abertura_local1_laje")
        )
        self.__insert_value(
            row=32,
            col="AA",
            value=data.get("buraco_abertura_observacoes_quantidade1_laje"),
        )

        armadura_exposta1_laje = data.get("armadura_exposta1_laje") or []
        if "1" in armadura_exposta1_laje:
            self.__mark_x(row=34, col="D")

        if "2" in armadura_exposta1_laje:
            self.__mark_x(row=34, col="I")

        self.__insert_value(
            row=34, col="Q", value=data.get("armadura_exposta_local1_laje")
        )
        self.__insert_value(
            row=34,
            col="AA",
            value=data.get("armadura_exposta_observacoes_quantidade1_laje"),
        )

        concreto_desagregado1_laje = data.get("concreto_desagregado1_laje") or []
        if "1" in concreto_desagregado1_laje:
            self.__mark_x(row=36, col="D")

        if "2" in concreto_desagregado1_laje:
            self.__mark_x(row=36, col="I")

        self.__insert_value(
            row=36, col="Q", value=data.get("concreto_desagregado_local1_laje")
        )
        self.__insert_value(
            row=36,
            col="AA",
            value=data.get("concreto_desagregado_observacoes_quantidade1_laje"),
        )

        fissuras1_laje = data.get("fissuras1_laje") or []
        if "1" in fissuras1_laje:
            self.__mark_x(row=38, col="D")

        if "2" in fissuras1_laje:
            self.__mark_x(row=38, col="I")

        self.__insert_value(row=38, col="Q", value=data.get("fissuras_local1_laje"))
        self.__insert_value(
            row=38,
            col="AA",
            value=data.get("fissuras_observacoes_quantidade1_laje"),
        )

        marcas_infiltracao1_laje = data.get("marcas_infiltracao1_laje") or []
        if "1" in marcas_infiltracao1_laje:
            self.__mark_x(row=40, col="D")
        if "2" in marcas_infiltracao1_laje:
            self.__mark_x(row=40, col="I")

        self.__insert_value(
            row=40, col="Q", value=data.get("marcas_infiltracao_local1_laje")
        )
        self.__insert_value(
            row=40,
            col="AA",
            value=data.get("marcas_infiltracao_observacoes_quantidade1_laje"),
        )

        aspecto_concreto1_laje = data.get("aspecto_concreto1_laje") or []
        if "1" in aspecto_concreto1_laje:
            self.__mark_x(row=42, col="D")

        self.__insert_value(
            row=42, col="Q", value=data.get("aspecto_concreto_local1_laje")
        )
        self.__insert_value(
            row=42,
            col="AA",
            value=data.get("aspecto_concreto_observacoes_quantidade1_laje"),
        )

        cobrimento1_laje = data.get("cobrimento1_laje") or []
        if "1" in cobrimento1_laje:
            self.__mark_x(row=44, col="D")

        self.__insert_value(row=44, col="Q", value=data.get("cobrimento_local1_laje"))
        self.__insert_value(
            row=44,
            col="AA",
            value=data.get("cobrimento_observacoes_quantidade1_laje"),
        )

        self.__insert_value(
            row=48,
            col="O",
            value=data.get("nota_tecnica_vigamento_principal"),
            horizontal="center",
        )

        fissuras_pequena_abertura2_vigamento_principal = (
            data.get("fissuras_pequena_abertura2_vigamento_principal") or []
        )
        if "1" in fissuras_pequena_abertura2_vigamento_principal:
            self.__mark_x(row=50, col="D")

        if "2" in fissuras_pequena_abertura2_vigamento_principal:
            self.__mark_x(row=50, col="I")

        self.__insert_value(
            row=50,
            col="Q",
            value=data.get("fissuras_pequena_abertura_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=50,
            col="AA",
            value=data.get(
                "fissuras_pequena_abertura_observacoes_quantidade2_vigamento_principal"
            ),
        )

        trincas2_vigamento_principal = data.get("trincas2_vigamento_principal") or []
        if "1" in trincas2_vigamento_principal:
            self.__mark_x(row=52, col="D")
        if "2" in trincas2_vigamento_principal:
            self.__mark_x(row=52, col="I")

        self.__insert_value(
            row=52,
            col="Q",
            value=data.get("trincas_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=52,
            col="AA",
            value=data.get("trincas_observacoes_quantidade2_vigamento_principal"),
        )

        armadura_principal2_vigamento_principal = (
            data.get("armadura_principal2_vigamento_principal") or []
        )
        if "1" in armadura_principal2_vigamento_principal:
            self.__mark_x(row=54, col="D")
        if "2" in armadura_principal2_vigamento_principal:
            self.__mark_x(row=54, col="I")

        self.__insert_value(
            row=54,
            col="Q",
            value=data.get("armadura_principal_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=54,
            col="AA",
            value=data.get(
                "armadura_principal_observacoes_quantidade2_vigamento_principal"
            ),
        )

        desagregamento_concreto2_vigamento_principal = (
            data.get("desagregamento_concreto2_vigamento_principal") or []
        )
        if "1" in desagregamento_concreto2_vigamento_principal:
            self.__mark_x(row=56, col="D")
        if "2" in desagregamento_concreto2_vigamento_principal:
            self.__mark_x(row=56, col="I")

        self.__insert_value(
            row=56,
            col="Q",
            value=data.get("desagregamento_concreto_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=56,
            col="AA",
            value=data.get(
                "desagregamento_concreto_observacoes_quantidade2_vigamento_principal"
            ),
        )

        dente_geber2_vigamento_principal = (
            data.get("dente_geber2_vigamento_principal") or []
        )
        if "1" in dente_geber2_vigamento_principal:
            self.__mark_x(row=58, col="D")
        if "2" in dente_geber2_vigamento_principal:
            self.__mark_x(row=58, col="I")

        self.__insert_value(
            row=58,
            col="Q",
            value=data.get("dente_geber_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=58,
            col="AA",
            value=data.get("dente_geber_observacoes_quantidade2_vigamento_principal"),
        )

        deformacao_flecha2_vigamento_principal = (
            data.get("deformacao_flecha2_vigamento_principal") or []
        )
        if "1" in deformacao_flecha2_vigamento_principal:
            self.__mark_x(row=60, col="D")

        self.__insert_value(
            row=60,
            col="Q",
            value=data.get("deformacao_flecha_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=60,
            col="AA",
            value=data.get(
                "deformacao_flecha_observacoes_quantidade2_vigamento_principal"
            ),
        )

        aspecto_concreto2_vigamento_principal = (
            data.get("aspecto_concreto2_vigamento_principal") or []
        )
        if "1" in aspecto_concreto2_vigamento_principal:
            self.__mark_x(row=62, col="D")

        self.__insert_value(
            row=62,
            col="Q",
            value=data.get("aspecto_concreto_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=62,
            col="AA",
            value=data.get(
                "aspecto_concreto_observacoes_quantidade2_vigamento_principal"
            ),
        )

        cobrimento2_vigamento_principal = (
            data.get("cobrimento2_vigamento_principal") or []
        )
        if "1" in cobrimento2_vigamento_principal:
            self.__mark_x(row=64, col="D")

        self.__insert_value(
            row=64,
            col="Q",
            value=data.get("cobrimento_local2_vigamento_principal"),
        )
        self.__insert_value(
            row=64,
            col="AA",
            value=data.get("cobrimento_observacoes_quantidade2_vigamento_principal"),
        )

    def __set_ficha_2(self, data: dict):
        self._worksheet = self._workbook["FICHA 2"]
        self.__insert_value(
            row=2, col="AE", value=data.get("codigo_obra"), horizontal="center"
        )
        self.__insert_value(row=5, col="B", value=data.get("codigo_agencia"))
        self.__insert_value(row=5, col="G", value=data.get("denominacao"))
        self.__insert_value(row=5, col="V", value=data.get("road_name_short"))
        self.__insert_value(row=5, col="X", value=data.get("km"))
        self.__insert_value(row=5, col="AD", value=data.get("direction"))
        self.__insert_value(row=7, col="B", value=data.get("executed_at"))
        self.__insert_value(row=7, col="L", value=data.get("zona"))
        self.__insert_value(row=7, col="P", value=data.get("utm_e1"))
        self.__insert_value(row=7, col="V", value=data.get("utm_e2"))
        self.__insert_value(row=8, col="P", value=data.get("utm_e3"))
        self.__insert_value(row=8, col="V", value=data.get("utm_e4"))
        self.__insert_value(row=7, col="T", value=data.get("utm_n1"))
        self.__insert_value(row=7, col="X", value=data.get("utm_n2"))
        self.__insert_value(row=8, col="T", value=data.get("utm_n3"))
        self.__insert_value(row=8, col="X", value=data.get("utm_n4"))

        self.__insert_value(
            row=12,
            col="P",
            value=data.get("nota_tecnica_mesoestrutura"),
            horizontal="center",
        )

        armadura_exposta3_mesoestrutura = (
            data.get("armadura_exposta3_mesoestrutura") or []
        )
        if "1" in armadura_exposta3_mesoestrutura:
            self.__mark_x(row=14, col="D")
        if "2" in armadura_exposta3_mesoestrutura:
            self.__mark_x(row=14, col="I")

        self.__insert_value(
            row=14,
            col="S",
            value=data.get("armadura_exposta_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=14,
            col="Y",
            value=data.get("armadura_exposta_observacoes_quantidade3_mesoestrutura"),
        )

        concreto_desagregado3_mesoestrutura = (
            data.get("concreto_desagregado3_mesoestrutura") or []
        )
        if "1" in concreto_desagregado3_mesoestrutura:
            self.__mark_x(row=16, col="D")
        if "2" in concreto_desagregado3_mesoestrutura:
            self.__mark_x(row=16, col="I")

        self.__insert_value(
            row=16,
            col="S",
            value=data.get("concreto_desagregado_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=16,
            col="Y",
            value=data.get(
                "concreto_desagregado_observacoes_quantidade3_mesoestrutura"
            ),
        )

        fissuras3_mesoestrutura = data.get("fissuras3_mesoestrutura") or []
        if "1" in fissuras3_mesoestrutura:
            self.__mark_x(row=18, col="D")
        if "2" in fissuras3_mesoestrutura:
            self.__mark_x(row=18, col="I")

        self.__insert_value(
            row=18,
            col="S",
            value=data.get("fissuras_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=18,
            col="Y",
            value=data.get("fissuras_observacoes_quantidade3_mesoestrutura"),
        )

        aparelho_apoio3_mesoestrutura = data.get("aparelho_apoio3_mesoestrutura") or []
        if "1" in aparelho_apoio3_mesoestrutura:
            self.__mark_x(row=20, col="D")
        if "2" in aparelho_apoio3_mesoestrutura:
            self.__mark_x(row=20, col="I")

        self.__insert_value(
            row=20,
            col="S",
            value=data.get("aparelho_apoio_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=20,
            col="Y",
            value=data.get("aparelho_apoio_observacoes_quantidade3_mesoestrutura"),
        )

        aspecto_concreto3_mesoestrutura = (
            data.get("aspecto_concreto3_mesoestrutura") or []
        )
        if "1" in aspecto_concreto3_mesoestrutura:
            self.__mark_x(row=22, col="D")

        self.__insert_value(
            row=22,
            col="S",
            value=data.get("aspecto_concreto_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=22,
            col="Y",
            value=data.get("aspecto_concreto_observacoes_quantidade3_mesoestrutura"),
        )

        cobrimento3_mesoestrutura = data.get("cobrimento3_mesoestrutura") or []
        if "1" in cobrimento3_mesoestrutura:
            self.__mark_x(row=24, col="D")

        self.__insert_value(
            row=24,
            col="S",
            value=data.get("cobrimento_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=24,
            col="Y",
            value=data.get("cobrimento_observacoes_quantidade3_mesoestrutura"),
        )

        desaprumo3_mesoestrutura = data.get("desaprumo3_mesoestrutura") or []
        if "1" in desaprumo3_mesoestrutura:
            self.__mark_x(row=26, col="D")

        self.__insert_value(
            row=26,
            col="S",
            value=data.get("desaprumo_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=26,
            col="Y",
            value=data.get("desaprumo_observacoes_quantidade3_mesoestrutura"),
        )

        deslocabilidade_pilares3_mesoestrutura = (
            data.get("deslocabilidade_pilares3_mesoestrutura") or []
        )
        if "1" in deslocabilidade_pilares3_mesoestrutura:
            self.__mark_x(row=28, col="D")

        self.__insert_value(
            row=28,
            col="S",
            value=data.get("deslocabilidade_pilares_local3_mesoestrutura"),
        )
        self.__insert_value(
            row=28,
            col="Y",
            value=data.get(
                "deslocabilidade_pilares_observacoes_quantidade3_mesoestrutura"
            ),
        )

        self.__insert_value(
            row=32,
            col="P",
            value=data.get("nota_tecnica_infraestrutura"),
            horizontal="center",
        )

        recalque_fundacao4_infraestrutura = data.get(
            "recalque_fundacao4_infraestrutura"
        )
        recalque_fundacao4_infraestrutura_target = None

        if str(recalque_fundacao4_infraestrutura) == "1":
            recalque_fundacao4_infraestrutura_target = "D"

        if isinstance(recalque_fundacao4_infraestrutura_target, str):
            self.__mark_x(
                row=34,
                col=recalque_fundacao4_infraestrutura_target,
            )

        self.__insert_value(
            row=34,
            col="S",
            value=data.get("recalque_fundacao_local4_infraestrutura"),
        )
        self.__insert_value(
            row=34,
            col="Y",
            value=data.get("recalque_fundacao_observacoes_quantidade4_infraestrutura"),
        )

        deslocamento_fundacao4_infraestrutura = data.get(
            "deslocamento_fundacao4_infraestrutura"
        )
        deslocamento_fundacao4_infraestrutura_target = None

        if str(deslocamento_fundacao4_infraestrutura) == "1":
            deslocamento_fundacao4_infraestrutura_target = "D"

        if isinstance(deslocamento_fundacao4_infraestrutura_target, str):
            self.__mark_x(
                row=36,
                col=deslocamento_fundacao4_infraestrutura_target,
            )

        self.__insert_value(
            row=36,
            col="S",
            value=data.get("deslocamento_fundacao_local4_infraestrutura"),
        )
        self.__insert_value(
            row=36,
            col="Y",
            value=data.get(
                "deslocamento_fundacao_observacoes_quantidade4_infraestrutura"
            ),
        )

        erosao_terreno_fundacao4_infraestrutura = data.get(
            "erosao_terreno_fundacao4_infraestrutura"
        )
        erosao_terreno_fundacao4_infraestrutura_target = None

        if str(erosao_terreno_fundacao4_infraestrutura) == "1":
            erosao_terreno_fundacao4_infraestrutura_target = "D"

        if isinstance(erosao_terreno_fundacao4_infraestrutura_target, str):
            self.__mark_x(row=38, col=erosao_terreno_fundacao4_infraestrutura_target)

        self.__insert_value(
            row=38,
            col="S",
            value=data.get("erosao_terreno_fundacao_local4_infraestrutura"),
        )
        self.__insert_value(
            row=38,
            col="Y",
            value=data.get(
                "erosao_terreno_fundacao_observacoes_quantidade4_infraestrutura"
            ),
        )

        estacas_dessenterradas4_infraestrutura = data.get(
            "estacas_dessenterradas4_infraestrutura"
        )
        estacas_dessenterradas4_infraestrutura_target = None

        if str(estacas_dessenterradas4_infraestrutura) == "1":
            estacas_dessenterradas4_infraestrutura_target = "D"

        if isinstance(estacas_dessenterradas4_infraestrutura_target, str):
            self.__mark_x(row=40, col=estacas_dessenterradas4_infraestrutura_target)

        self.__insert_value(
            row=40,
            col="S",
            value=data.get("estacas_dessenterradas_local4_infraestrutura"),
        )
        self.__insert_value(
            row=40,
            col="Y",
            value=data.get(
                "estacas_dessenterradas_observacoes_quantidade4_infraestrutura"
            ),
        )

        self.__insert_value(
            row=44,
            col="P",
            value=data.get("nota_tecnica_pista_acesso"),
            horizontal="center",
        )

        irregularidades_pavimento5_pista_acesso = (
            data.get("irregularidades_pavimento5_pista_acesso") or []
        )
        if "1" in irregularidades_pavimento5_pista_acesso:
            self.__mark_x(row=46, col="D")
        if "2" in irregularidades_pavimento5_pista_acesso:
            self.__mark_x(row=46, col="I")

        self.__insert_value(
            row=46,
            col="S",
            value=data.get("irregularidades_pavimento_local5_pista_acesso"),
        )
        self.__insert_value(
            row=46,
            col="Y",
            value=data.get(
                "irregularidades_pavimento_observacoes_quantidade5_pista_acesso"
            ),
        )

        junta_dilatacao5_pista_acesso = data.get("junta_dilatacao5_pista_acesso") or []
        if "1" in junta_dilatacao5_pista_acesso:
            self.__mark_x(row=48, col="D")
        if "2" in junta_dilatacao5_pista_acesso:
            self.__mark_x(row=48, col="I")

        self.__insert_value(
            row=48,
            col="S",
            value=data.get("junta_dilatacao_local5_pista_acesso"),
        )
        self.__insert_value(
            row=48,
            col="Y",
            value=data.get("junta_dilatacao_observacoes_quantidade5_pista_acesso"),
        )

        acesso_oae5_pista_acesso = data.get("acesso_oae5_pista_acesso") or []
        if "1" in acesso_oae5_pista_acesso:
            self.__mark_x(row=50, col="D")
        if "2" in acesso_oae5_pista_acesso:
            self.__mark_x(row=50, col="I")

        self.__insert_value(
            row=50,
            col="S",
            value=data.get("acesso_oae_local5_pista_acesso"),
        )
        self.__insert_value(
            row=50,
            col="Y",
            value=data.get("acesso_oae_observacoes_quantidade5_pista_acesso"),
        )

        acidente_veiculos5_pista_acesso = (
            data.get("acidente_veiculos5_pista_acesso") or []
        )
        if "1" in acidente_veiculos5_pista_acesso:
            self.__mark_x(row=52, col="D")
        if "2" in acidente_veiculos5_pista_acesso:
            self.__mark_x(row=52, col="I")

        self.__insert_value(
            row=52,
            col="S",
            value=data.get("acidente_veiculos_local5_pista_acesso"),
        )
        self.__insert_value(
            row=52,
            col="Y",
            value=data.get("acidente_veiculos_observacoes_quantidade5_pista_acesso"),
        )

        guarda_corpo = data.get("guarda_corpo")
        guarda_corpo_target = None

        if str(guarda_corpo) == "1":
            guarda_corpo_target = "I"
        elif str(guarda_corpo) == "2":
            guarda_corpo_target = "M"

        if isinstance(guarda_corpo_target, str):
            self.__mark_x(row=58, col=guarda_corpo_target)

        self.__insert_value(
            row=58,
            col="U",
            value=data.get("guarda_corpo_observacoes_quantidade"),
        )

        drenagem_limpeza = data.get("drenagem_limpeza")
        drenagem_limpeza_target = None

        if str(drenagem_limpeza) == "1":
            drenagem_limpeza_target = "I"
        elif str(drenagem_limpeza) == "2":
            drenagem_limpeza_target = "M"

        if isinstance(drenagem_limpeza_target, str):
            self.__mark_x(row=60, col=drenagem_limpeza_target)

        self.__insert_value(
            row=60,
            col="U",
            value=data.get("drenagem_limpeza_observacoes_quantidade"),
        )

        placa_gabarito_vertical = data.get("placa_gabarito_vertical")
        placa_gabarito_vertical_target = None

        if str(placa_gabarito_vertical) == "1":
            placa_gabarito_vertical_target = "I"
        elif str(placa_gabarito_vertical) == "2":
            placa_gabarito_vertical_target = "M"
        elif str(placa_gabarito_vertical) == "3":
            placa_gabarito_vertical_target = "Q"

        if isinstance(placa_gabarito_vertical_target, str):
            self.__mark_x(row=62, col=placa_gabarito_vertical_target)

        self.__insert_value(
            row=62,
            col="U",
            value=data.get("placa_gabarito_vertical_observacoes_quantidade"),
        )

        juntas_dilatacao_vida_util_remanescente = data.get(
            "juntas_dilatacao_vida_util_remanescente"
        )
        juntas_dilatacao_vida_util_remanescente_target = None

        if str(juntas_dilatacao_vida_util_remanescente) == "1":
            juntas_dilatacao_vida_util_remanescente_target = "I"
        elif str(juntas_dilatacao_vida_util_remanescente) == "2":
            juntas_dilatacao_vida_util_remanescente_target = "M"

        if isinstance(juntas_dilatacao_vida_util_remanescente_target, str):
            self.__mark_x(row=64, col=juntas_dilatacao_vida_util_remanescente_target)

        self.__insert_value(
            row=64,
            col="U",
            value=data.get(
                "juntas_dilatacao_vida_util_remanescente_observacoes_quantidade"
            ),
        )

        aparelho_apoio_vida_util_remanescente = data.get(
            "aparelho_apoio_vida_util_remanescente"
        )
        aparelho_apoio_vida_util_remanescente_target = None

        if str(aparelho_apoio_vida_util_remanescente) == "1":
            aparelho_apoio_vida_util_remanescente_target = "I"
        elif str(aparelho_apoio_vida_util_remanescente) == "2":
            aparelho_apoio_vida_util_remanescente_target = "M"

        if isinstance(aparelho_apoio_vida_util_remanescente_target, str):
            self.__mark_x(row=66, col=aparelho_apoio_vida_util_remanescente_target)

        self.__insert_value(
            row=66,
            col="U",
            value=data.get("aparelho_apoio_vida_util_remanescente_local"),
        )

    def __set_croqui(self, data):
        self._worksheet = self._workbook["CROQUI"]

        self.__insert_value(
            row=2, col="X", value=data.get("codigo_obra"), horizontal="center"
        )
        self.__insert_value(row=5, col="B", value=data.get("codigo_agencia"))
        self.__insert_value(row=5, col="E", value=data.get("denominacao"))
        self.__insert_value(row=5, col="O", value=data.get("road_name_short"))
        self.__insert_value(row=5, col="Q", value=data.get("km"))
        self.__insert_value(row=5, col="W", value=data.get("direction"))
        self.__insert_value(row=7, col="B", value=data.get("executed_at"))
        self.__insert_value(row=7, col="H", value=data.get("zona"))
        self.__insert_value(row=7, col="J", value=data.get("utm_e1"))
        self.__insert_value(row=7, col="O", value=data.get("utm_e2"))
        self.__insert_value(row=8, col="J", value=data.get("utm_e3"))
        self.__insert_value(row=8, col="O", value=data.get("utm_e4"))
        self.__insert_value(row=7, col="M", value=data.get("utm_n1"))
        self.__insert_value(row=7, col="Q", value=data.get("utm_n2"))
        self.__insert_value(row=8, col="M", value=data.get("utm_n3"))
        self.__insert_value(row=8, col="Q", value=data.get("utm_n4"))

        images_croqui = data.get("images_croqui")
        if images_croqui:
            origin_title_sheet = "CROQUI 01"
            self._worksheet.title = origin_title_sheet
            original_sheet = self._workbook[origin_title_sheet]
            original_images = original_sheet._images

            self._worksheet = original_sheet

            att_img = None
            if self._copy_antt_logo:
                att_img = original_images[0]

            for i, image in enumerate(images_croqui):
                crouqui_current = str(i + 1).zfill(2)
                new_sheet_name = f"CROQUI {crouqui_current}"

                croqui_sheets = [
                    sheet
                    for sheet in self._workbook.sheetnames
                    if sheet.startswith("CROQUI")
                ]
                last_croqui_index = max(
                    self._workbook.sheetnames.index(sheet) for sheet in croqui_sheets
                )
                if i > 0:
                    self._worksheet = copy_sheet_with_settings(
                        self._workbook,
                        original_sheet,
                        new_sheet_name,
                        last_croqui_index + 1,
                    )

                range_str = "A14:Z67"
                insert_picture_2(
                    self._worksheet,
                    range_str,
                    Image(image),
                    self._sheet_target,
                    (2, 2, 2, 2),
                    ResizeMethod.ProportionalCentered,
                )

                if self._copy_antt_logo:
                    self._worksheet.add_image(
                        Image(io.BytesIO(att_img._data())), att_img.anchor
                    )

    def __set_photos(self, data, logo=None, logo_team=""):
        photos = data.get("photos")
        original_title_sheet = "FOTOS 1"
        original_sheet = self._workbook[original_title_sheet]
        original_images = original_sheet._images

        att_img = None
        if self._copy_antt_logo:
            att_img = original_images[0]
        self._worksheet = original_sheet

        self.__insert_value(
            row=2, col="Z", value=data.get("codigo_obra"), horizontal="center"
        )
        self.__insert_value(row=5, col="B", value=data.get("codigo_agencia"))
        self.__insert_value(row=5, col="E", value=data.get("denominacao"))
        self.__insert_value(row=5, col="P", value=data.get("road_name_short"))
        self.__insert_value(row=5, col="R", value=data.get("km"))
        self.__insert_value(row=5, col="X", value=data.get("direction"))
        self.__insert_value(row=7, col="B", value=data.get("executed_at"))
        self.__insert_value(row=7, col="H", value=data.get("zona"))
        self.__insert_value(row=7, col="J", value=data.get("utm_e1"))
        self.__insert_value(row=7, col="P", value=data.get("utm_e2"))
        self.__insert_value(row=8, col="J", value=data.get("utm_e3"))
        self.__insert_value(row=8, col="P", value=data.get("utm_e4"))
        self.__insert_value(row=7, col="N", value=data.get("utm_n1"))
        self.__insert_value(row=7, col="R", value=data.get("utm_n2"))
        self.__insert_value(row=8, col="N", value=data.get("utm_n3"))
        self.__insert_value(row=8, col="R", value=data.get("utm_n4"))

        if photos:
            count_sheet = 1
            count_photo = 1

            while photos:
                # Pela regra de negocio só pode ter no maximo 13 abas do excel se haver mais abas ele para a criação
                if count_sheet > 13:
                    break

                photos_currency = []

                if count_sheet > 1:
                    new_sheet_name = f"FOTOS {count_sheet}"
                    photo_sheets = [
                        sheet
                        for sheet in self._workbook.sheetnames
                        if sheet.startswith("FOTO")
                    ]
                    last_photo_index = max(
                        self._workbook.sheetnames.index(sheet) for sheet in photo_sheets
                    )
                    self._worksheet = copy_sheet_with_settings(
                        self._workbook,
                        original_sheet,
                        new_sheet_name,
                        last_photo_index + 1,
                    )

                if len(photos) > 0:
                    for _ in range(4):
                        if photos:
                            photo_remove = photos.pop(0)
                            photos_currency.append(photo_remove)

                # Clean Campos
                self.__insert_value(row=37, col="C", value="")
                self.__insert_value(row=37, col="E", value="")
                self.__insert_value(row=37, col="N", value="")
                self.__insert_value(row=37, col="P", value="")
                self.__insert_value(row=65, col="C", value="")
                self.__insert_value(row=65, col="E", value="")
                self.__insert_value(row=65, col="N", value="")
                self.__insert_value(row=65, col="P", value="")

                if photos_currency:
                    # WIDTH_IMAGE = 465
                    # HEIGHT_IMAGE = 350
                    prefix_photo_name = data.get("prefix_photo_name")
                    for i, _obj in enumerate(photos_currency):
                        photo = _obj.get("photo")
                        photo_number = int_set_zero_prefix(count_photo)
                        photo_name = f"{prefix_photo_name}F{photo_number}:"

                        if photo:
                            if i == 0:
                                range_str = "C13:J36"
                                insert_picture(
                                    worksheet=self._worksheet,
                                    range_string=range_str,
                                    picture=Image(photo),
                                    target=self._sheet_target,
                                )
                                self.__insert_value(
                                    row=37,
                                    col="C",
                                    value=photo_name,
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                                self.__insert_value(
                                    row=37,
                                    col="E",
                                    value=_obj.get("description"),
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                            elif i == 1:
                                range_str = "N13:X36"
                                insert_picture(
                                    worksheet=self._worksheet,
                                    range_string=range_str,
                                    picture=Image(photo),
                                    target=self._sheet_target,
                                )
                                self.__insert_value(
                                    row=37,
                                    col="N",
                                    value=photo_name,
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                                self.__insert_value(
                                    row=37,
                                    col="P",
                                    value=_obj.get("description"),
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                            elif i == 2:
                                range_str = "C40:J64"
                                insert_picture(
                                    worksheet=self._worksheet,
                                    range_string=range_str,
                                    picture=Image(photo),
                                    target=self._sheet_target,
                                )
                                self.__insert_value(
                                    row=65,
                                    col="C",
                                    value=photo_name,
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                                self.__insert_value(
                                    row=65,
                                    col="E",
                                    value=_obj.get("description"),
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                            elif i == 3:
                                range_str = "N40:X64"
                                insert_picture(
                                    worksheet=self._worksheet,
                                    range_string=range_str,
                                    picture=Image(photo),
                                    target=self._sheet_target,
                                )
                                self.__insert_value(
                                    row=65,
                                    col="N",
                                    value=photo_name,
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )
                                self.__insert_value(
                                    row=65,
                                    col="P",
                                    value=_obj.get("description"),
                                    alignment=Alignment(
                                        vertical="top", horizontal="left", wrapText=True
                                    ),
                                )

                            count_photo += 1

                if logo:
                    self.__insert_img_size(
                        image=logo,
                        row_init=68,
                        row_end=70,
                        col_init=0,
                        col_end=2,
                        width=90,
                        height=30,
                    )
                if logo_team:
                    self.__insert_img_with_absolute_anchor(
                        image=logo_team, width=90, height=30, pos_x=660, pos_y=995
                    )
                if self._copy_antt_logo:
                    self._worksheet.add_image(
                        Image(io.BytesIO(att_img._data())), att_img.anchor
                    )

                count_sheet += 1

    def fill_sheet(self, data_list: list):
        files = []
        filenames = []
        all_roads = []

        for i, data in enumerate(data_list, 1):
            logo = data.get("logo", "")
            logo_team = data.get("logo_team")

            logo_data = {
                "image": logo,
                "width": 90,
                "height": 30,
            }
            logo_team_data = {
                "image": logo_team,
                "width": 90,
                "height": 30,
                "pos_x": 660,
                "pos_y": 995,
            }

            args_logo = dict(
                row_init=66,
                row_end=68,
                col_init=0,
                col_end=2,
                **logo_data,
            )

            self.__set_ficha_1(data)
            if logo:
                self.__insert_img_size(
                    **args_logo,
                )

            if logo_team:
                self.__insert_img_with_absolute_anchor(
                    **logo_team_data,
                )

            self.__set_ficha_2(data)
            if logo:
                args_logo.update(
                    dict(
                        row_init=68,
                        row_end=70,
                    )
                )
                self.__insert_img_size(**args_logo)

            if logo_team:
                self.__insert_img_with_absolute_anchor(
                    **logo_team_data,
                )

            self.__set_croqui(data)

            if logo:
                self.__insert_img_size(**args_logo)

            if logo_team:
                self.__insert_img_with_absolute_anchor(
                    **logo_team_data,
                )

            self.__set_photos(data, logo, logo_team)

            if logo:
                self.__insert_img_size(**args_logo)

            if logo_team:
                self.__insert_img_with_absolute_anchor(
                    **logo_team_data,
                )

            file_name = data.get("file_name")
            filenames.append(file_name)
            result = f"/tmp/{file_name}.xlsx"
            if result in files:
                result = f"/tmp/{file_name}({i}).xlsx"

            set_zoom(self._workbook, 50, "pageBreakPreview")

            self._workbook.save(result)
            self._workbook = load_workbook(self._xlsx_file)

            files.append(result)
        all_roads.sort()
        return {"files": files, "names": filenames, "all_roads": all_roads}

    def _imagem(self, url):
        path_image = ""
        if url:
            try:
                file_path = url.split("?")[0].split(".com/")[1]
                bucket_name = url.split(".s3")[0].split("/")[-1]
                image_format = file_path.split(".")[-1]
                path_image = f"{self.temp_file}{uuid4()}.{image_format}"
                self.s3.download_file(bucket_name, file_path, path_image)
            except Exception:
                path_image = ""
        return path_image

    def valid_rules_order_photo_oae(self, file_name: str) -> bool:
        """
        Utilizando o nome do arquivo como referência, aparecem primeiro
        as fotos com menor valor dos últimos 3 dígitos do nome do arquivo
        (os quais devem obrigatoriamente ser 3 números). O formato esperado para
        os nomes de arquivo segue o padrão do exemplo: OAE2023101095RJF001, em que:

        OAE = OAE
        2023 = ano
        101 = rodovia
        095 = Nº OAE
        SP ou RJ
        F001 = nº foto
        """
        try:
            file_name = file_name.upper()
            return (
                len(file_name) == 19
                and file_name[15] == "F"
                and file_name.startswith("OAE")  # OAE
                and file_name[3:7].isdigit()  # YEAR
                and file_name[7:10].isdigit()  # ROAD
                and file_name[10:13].isdigit()  # Nº OAE
                and file_name[13:15].isalpha()  # UF
                and file_name[16:].isdigit()  # Nº PHOTO
            )
        except Exception as e:
            print(f"Error: {e}")
            return False

    def create_dict(self, reporting: Reporting, s3) -> dict:
        form_data = deep_keys_to_snake_case(reporting.form_data)
        _form_data_display = reporting.get_form_data_display()

        INSPECTION_YEAR_CAMPAIGN = form_data.get("inspection_year_campaign") or ""
        KM = format_km(reporting, "km", 3) or "-"
        n_oae = new_get_form_data(reporting, "oaeNumeroCodigoObra", default="")
        try:
            n_oae = f"{int(n_oae):03}"
        except Exception:
            pass
        _road_name = (reporting.road_name.replace(" ", n_oae).split("-"))[1]
        PREFIX_PHOTO_NAME = f"OAE{INSPECTION_YEAR_CAMPAIGN}{_road_name}"
        DIRECTION = get_custom_option(reporting, "direction") or "-"
        DENOMINACAO = _form_data_display.get("denominacao", "-")
        CODIGO_AGENCIA = _form_data_display.get("codigo_agencia", "-")
        ZONA = _form_data_display.get("zona", "-")
        UTM_E1 = _form_data_display.get("utm_e1", "-")
        UTM_E2 = _form_data_display.get("utm_e2", "-")
        UTM_E3 = _form_data_display.get("utm_e3", "-")
        UTM_E4 = _form_data_display.get("utm_e4", "-")
        UTM_N1 = _form_data_display.get("utm_n1", "-")
        UTM_N2 = _form_data_display.get("utm_n2", "-")
        UTM_N3 = _form_data_display.get("utm_n3", "-")
        UTM_N4 = _form_data_display.get("utm_n4", "-")

        NOTA_TECNICA_COMENTARIOS_GERAIS = _form_data_display.get(
            "nota_tecnica_comentarios_gerais", "-"
        )
        NOTA_TECNICA_LAJE = _form_data_display.get("nota_tecnica_laje", "-")
        NOTA_TECNICA_VIGAMENTO_PRINCIPAL = _form_data_display.get(
            "nota_tecnica_vigamento_principal", "-"
        )

        oae_numero_codigo_obra = form_data.get("oae_numero_codigo_obra", "1")

        OAE_NUMERO_CODIGO_OBRA = int_set_zero_prefix(oae_numero_codigo_obra)

        CONDICOES_ESTABILIDADE = form_data.get("condicoes_estabilidade")
        CONDICOES_CONSERVACAO = form_data.get("condicoes_conservacao")
        NIVEL_VIBRACAO_TABULEIRO = form_data.get("nivel_vibracao_tabuleiro")
        INSPECAO_ESPECIALIZADA = form_data.get("inspecao_especializada")
        URGENTE_ANTT = form_data.get("urgente_antt")
        HOUVE_ANTERIORMENTE = form_data.get("houve_anteriormente")

        HISTORICO_INTERVENCOES_REALIZADAS = _form_data_display.get(
            "historico_intervencoes_realizadas", "-"
        )
        observacoes_adicionais = _form_data_display.get("observacoes_adicionais", "-")

        buraco_abertura1_laje = form_data.get("buraco_abertura1_laje")
        buraco_abertura_local1_laje = _form_data_display.get(
            "buraco_abertura_local1_laje", "-"
        )
        buraco_abertura_observacoes_quantidade1_laje = _form_data_display.get(
            "buraco_abertura_observacoes_quantidade1_laje", "-"
        )

        armadura_exposta1_laje = form_data.get("armadura_exposta1_laje")
        armadura_exposta_local1_laje = _form_data_display.get(
            "armadura_exposta_local1_laje", "-"
        )
        armadura_exposta_observacoes_quantidade1_laje = _form_data_display.get(
            "armadura_exposta_observacoes_quantidade1_laje", "-"
        )

        concreto_desagregado1_laje = form_data.get("concreto_desagregado1_laje")
        concreto_desagregado_local1_laje = _form_data_display.get(
            "concreto_desagregado_local1_laje", "-"
        )
        concreto_desagregado_observacoes_quantidade1_laje = _form_data_display.get(
            "concreto_desagregado_observacoes_quantidade1_laje", "-"
        )

        fissuras1_laje = form_data.get("fissuras1_laje")
        fissuras_local1_laje = _form_data_display.get("fissuras_local1_laje", "-")
        fissuras_observacoes_quantidade1_laje = _form_data_display.get(
            "fissuras_observacoes_quantidade1_laje", "-"
        )

        marcas_infiltracao1_laje = form_data.get("marcas_infiltracao1_laje")
        marcas_infiltracao_local1_laje = _form_data_display.get(
            "marcas_infiltracao_local1_laje", "-"
        )
        marcas_infiltracao_observacoes_quantidade1_laje = _form_data_display.get(
            "marcas_infiltracao_observacoes_quantidade1_laje", "-"
        )

        aspecto_concreto1_laje = form_data.get("aspecto_concreto1_laje")
        aspecto_concreto_local1_laje = _form_data_display.get(
            "aspecto_concreto_local1_laje", "-"
        )
        aspecto_concreto_observacoes_quantidade1_laje = _form_data_display.get(
            "aspecto_concreto_observacoes_quantidade1_laje", "-"
        )

        cobrimento1_laje = form_data.get("cobrimento1_laje")
        cobrimento_local1_laje = _form_data_display.get("cobrimento_local1_laje", "-")
        cobrimento_observacoes_quantidade1_laje = _form_data_display.get(
            "cobrimento_observacoes_quantidade1_laje", "-"
        )

        fissuras_pequena_abertura2_vigamento_principal = form_data.get(
            "fissuras_pequena_abertura2_vigamento_principal"
        )
        fissuras_pequena_abertura_local2_vigamento_principal = _form_data_display.get(
            "fissuras_pequena_abertura_local2_vigamento_principal", "-"
        )
        fissuras_pequena_abertura_observacoes_quantidade2_vigamento_principal = (
            _form_data_display.get(
                "fissuras_pequena_abertura_observacoes_quantidade2_vigamento_principal",
                "-",
            )
        )

        trincas2_vigamento_principal = form_data.get("trincas2_vigamento_principal")
        trincas_local2_vigamento_principal = _form_data_display.get(
            "trincas_local2_vigamento_principal", "-"
        )
        trincas_observacoes_quantidade2_vigamento_principal = _form_data_display.get(
            "trincas_observacoes_quantidade2_vigamento_principal",
            "-",
        )

        armadura_principal2_vigamento_principal = form_data.get(
            "armadura_principal2_vigamento_principal"
        )
        armadura_principal_local2_vigamento_principal = _form_data_display.get(
            "armadura_principal_local2_vigamento_principal", "-"
        )
        armadura_principal_observacoes_quantidade2_vigamento_principal = (
            _form_data_display.get(
                "armadura_principal_observacoes_quantidade2_vigamento_principal",
                "-",
            )
        )

        desagregamento_concreto2_vigamento_principal = form_data.get(
            "desagregamento_concreto2_vigamento_principal"
        )
        desagregamento_concreto_local2_vigamento_principal = _form_data_display.get(
            "desagregamento_concreto_local2_vigamento_principal", "-"
        )
        desagregamento_concreto_observacoes_quantidade2_vigamento_principal = (
            _form_data_display.get(
                "desagregamento_concreto_observacoes_quantidade2_vigamento_principal",
                "-",
            )
        )

        dente_geber2_vigamento_principal = form_data.get(
            "dente_geber2_vigamento_principal"
        )
        dente_geber_local2_vigamento_principal = _form_data_display.get(
            "dente_geber_local2_vigamento_principal", "-"
        )
        dente_geber_observacoes_quantidade2_vigamento_principal = (
            _form_data_display.get(
                "dente_geber_observacoes_quantidade2_vigamento_principal",
                "-",
            )
        )

        deformacao_flecha2_vigamento_principal = form_data.get(
            "deformacao_flecha2_vigamento_principal"
        )
        deformacao_flecha_local2_vigamento_principal = _form_data_display.get(
            "deformacao_flecha_local2_vigamento_principal", "-"
        )
        deformacao_flecha_observacoes_quantidade2_vigamento_principal = (
            _form_data_display.get(
                "deformacao_flecha_observacoes_quantidade2_vigamento_principal",
                "-",
            )
        )

        aspecto_concreto2_vigamento_principal = form_data.get(
            "aspecto_concreto2_vigamento_principal"
        )
        aspecto_concreto_local2_vigamento_principal = _form_data_display.get(
            "aspecto_concreto_local2_vigamento_principal", "-"
        )
        aspecto_concreto_observacoes_quantidade2_vigamento_principal = (
            _form_data_display.get(
                "aspecto_concreto_observacoes_quantidade2_vigamento_principal",
                "-",
            )
        )

        cobrimento2_vigamento_principal = form_data.get(
            "cobrimento2_vigamento_principal"
        )
        cobrimento_local2_vigamento_principal = _form_data_display.get(
            "cobrimento_local2_vigamento_principal", "-"
        )
        cobrimento_observacoes_quantidade2_vigamento_principal = _form_data_display.get(
            "cobrimento_observacoes_quantidade2_vigamento_principal",
            "-",
        )
        NOTA_TECNICA_MESOESTRUTURA = _form_data_display.get(
            "nota_tecnica_mesoestrutura", "-"
        )

        armadura_exposta3_mesoestrutura = form_data.get(
            "armadura_exposta3_mesoestrutura"
        )
        armadura_exposta_local3_mesoestrutura = _form_data_display.get(
            "armadura_exposta_local3_mesoestrutura", "-"
        )
        armadura_exposta_observacoes_quantidade3_mesoestrutura = _form_data_display.get(
            "armadura_exposta_observacoes_quantidade3_mesoestrutura",
            "-",
        )

        concreto_desagregado3_mesoestrutura = form_data.get(
            "concreto_desagregado3_mesoestrutura"
        )
        concreto_desagregado_local3_mesoestrutura = _form_data_display.get(
            "concreto_desagregado_local3_mesoestrutura", "-"
        )
        concreto_desagregado_observacoes_quantidade3_mesoestrutura = (
            _form_data_display.get(
                "concreto_desagregado_observacoes_quantidade3_mesoestrutura",
                "-",
            )
        )

        fissuras3_mesoestrutura = form_data.get("fissuras3_mesoestrutura")
        fissuras_local3_mesoestrutura = _form_data_display.get(
            "fissuras_local3_mesoestrutura", "-"
        )
        fissuras_observacoes_quantidade3_mesoestrutura = _form_data_display.get(
            "fissuras_observacoes_quantidade3_mesoestrutura",
            "-",
        )

        aparelho_apoio3_mesoestrutura = form_data.get("aparelho_apoio3_mesoestrutura")
        aparelho_apoio_local3_mesoestrutura = _form_data_display.get(
            "aparelho_apoio_local3_mesoestrutura", "-"
        )
        aparelho_apoio_observacoes_quantidade3_mesoestrutura = _form_data_display.get(
            "aparelho_apoio_observacoes_quantidade3_mesoestrutura",
            "-",
        )

        aspecto_concreto3_mesoestrutura = form_data.get(
            "aspecto_concreto3_mesoestrutura"
        )
        aspecto_concreto_local3_mesoestrutura = _form_data_display.get(
            "aspecto_concreto_local3_mesoestrutura", "-"
        )
        aspecto_concreto_observacoes_quantidade3_mesoestrutura = _form_data_display.get(
            "aspecto_concreto_observacoes_quantidade3_mesoestrutura",
            "-",
        )

        cobrimento3_mesoestrutura = form_data.get("cobrimento3_mesoestrutura")
        cobrimento_local3_mesoestrutura = _form_data_display.get(
            "cobrimento_local3_mesoestrutura", "-"
        )
        cobrimento_observacoes_quantidade3_mesoestrutura = _form_data_display.get(
            "cobrimento_observacoes_quantidade3_mesoestrutura",
            "-",
        )

        desaprumo3_mesoestrutura = form_data.get("desaprumo3_mesoestrutura")
        desaprumo_local3_mesoestrutura = _form_data_display.get(
            "desaprumo_local3_mesoestrutura", "-"
        )
        desaprumo_observacoes_quantidade3_mesoestrutura = _form_data_display.get(
            "desaprumo_observacoes_quantidade3_mesoestrutura",
            "-",
        )

        deslocabilidade_pilares3_mesoestrutura = form_data.get(
            "deslocabilidade_pilares3_mesoestrutura"
        )
        deslocabilidade_pilares_local3_mesoestrutura = _form_data_display.get(
            "deslocabilidade_pilares_local3_mesoestrutura", "-"
        )
        deslocabilidade_pilares_observacoes_quantidade3_mesoestrutura = (
            _form_data_display.get(
                "deslocabilidade_pilares_observacoes_quantidade3_mesoestrutura",
                "-",
            )
        )

        NOTA_TECNICA_INFRAESTRUTURA = _form_data_display.get(
            "nota_tecnica_infraestrutura", "-"
        )

        recalque_fundacao4_infraestrutura = form_data.get(
            "recalque_fundacao4_infraestrutura"
        )
        recalque_fundacao_local4_infraestrutura = _form_data_display.get(
            "recalque_fundacao_local4_infraestrutura", "-"
        )
        recalque_fundacao_observacoes_quantidade4_infraestrutura = (
            _form_data_display.get(
                "recalque_fundacao_observacoes_quantidade4_infraestrutura",
                "-",
            )
        )

        deslocamento_fundacao4_infraestrutura = form_data.get(
            "deslocamento_fundacao4_infraestrutura"
        )
        deslocamento_fundacao_local4_infraestrutura = _form_data_display.get(
            "deslocamento_fundacao_local4_infraestrutura", "-"
        )
        deslocamento_fundacao_observacoes_quantidade4_infraestrutura = (
            _form_data_display.get(
                "deslocamento_fundacao_observacoes_quantidade4_infraestrutura",
                "-",
            )
        )

        erosao_terreno_fundacao4_infraestrutura = form_data.get(
            "erosao_terreno_fundacao4_infraestrutura"
        )
        erosao_terreno_fundacao_local4_infraestrutura = _form_data_display.get(
            "erosao_terreno_fundacao_local4_infraestrutura", "-"
        )
        erosao_terreno_fundacao_observacoes_quantidade4_infraestrutura = (
            _form_data_display.get(
                "erosao_terreno_fundacao_observacoes_quantidade4_infraestrutura",
                "-",
            )
        )

        estacas_dessenterradas4_infraestrutura = form_data.get(
            "estacas_dessenterradas4_infraestrutura"
        )
        estacas_dessenterradas_local4_infraestrutura = _form_data_display.get(
            "estacas_dessenterradas_local4_infraestrutura", "-"
        )
        estacas_dessenterradas_observacoes_quantidade4_infraestrutura = (
            _form_data_display.get(
                "estacas_dessenterradas_observacoes_quantidade4_infraestrutura",
                "-",
            )
        )

        NOTA_TECNICA_PISTA_ACESSO = _form_data_display.get(
            "nota_tecnica_pista_acesso", "-"
        )

        irregularidades_pavimento5_pista_acesso = form_data.get(
            "irregularidades_pavimento5_pista_acesso"
        )
        irregularidades_pavimento_local5_pista_acesso = _form_data_display.get(
            "irregularidades_pavimento_local5_pista_acesso", "-"
        )
        irregularidades_pavimento_observacoes_quantidade5_pista_acesso = (
            _form_data_display.get(
                "irregularidades_pavimento_observacoes_quantidade5_pista_acesso",
                "-",
            )
        )

        junta_dilatacao5_pista_acesso = form_data.get("junta_dilatacao5_pista_acesso")
        junta_dilatacao_local5_pista_acesso = _form_data_display.get(
            "junta_dilatacao_local5_pista_acesso", "-"
        )
        junta_dilatacao_observacoes_quantidade5_pista_acesso = _form_data_display.get(
            "junta_dilatacao_observacoes_quantidade5_pista_acesso",
            "-",
        )

        acesso_oae5_pista_acesso = form_data.get("acesso_oae_cinco_pista_acesso")
        acesso_oae_local5_pista_acesso = _form_data_display.get(
            "acesso_oae_local_cinco_pista_acesso", "-"
        )
        acesso_oae_observacoes_quantidade5_pista_acesso = _form_data_display.get(
            "acesso_oae_observacoes_quantidade_cinco_pista_acesso",
            "-",
        )

        acidente_veiculos5_pista_acesso = form_data.get(
            "acidente_veiculos_cinco_pista_acesso"
        )
        acidente_veiculos_local5_pista_acesso = _form_data_display.get(
            "acidente_veiculos_local5_pista_acesso", "-"
        )
        acidente_veiculos_observacoes_quantidade5_pista_acesso = _form_data_display.get(
            "acidente_veiculos_observacoes_quantidade5_pista_acesso",
            "-",
        )

        guarda_corpo = form_data.get("guarda_corpo")
        guarda_corpo_observacoes_quantidade = _form_data_display.get(
            "guarda_corpo_observacoes_quantidade",
            "-",
        )

        drenagem_limpeza = form_data.get("drenagem_limpeza")
        drenagem_limpeza_observacoes_quantidade = _form_data_display.get(
            "drenagem_limpeza_observacoes_quantidade",
            "-",
        )

        placa_gabarito_vertical = form_data.get("placa_gabarito_vertical")
        placa_gabarito_vertical_observacoes_quantidade = _form_data_display.get(
            "placa_gabarito_vertical_observacoes_quantidade",
            "-",
        )

        juntas_dilatacao_vida_util_remanescente = form_data.get(
            "juntas_dilatacao_vida_util_remanescente"
        )
        juntas_dilatacao_vida_util_remanescente_observacoes_quantidade = (
            _form_data_display.get(
                "juntas_dilatacao_vida_util_remanescente_observacoes_quantidade",
                "-",
            )
        )
        aparelho_apoio_vida_util_remanescente = form_data.get(
            "aparelho_apoio_vida_util_remanescente"
        )
        aparelho_apoio_vida_util_remanescente_local = _form_data_display.get(
            "aparelho_apoio_vida_util_remanescente_local",
            "-",
        )

        EXECUTED_AT = (
            reporting.executed_at.strftime("%d/%m/%Y") if reporting.executed_at else "-"
        )
        ROAD_NAME = reporting.road_name
        ROAD_NAME_SHORT = (reporting.road_name[3:]).replace(" ", "/")
        LOGO = get_logo_file(self.s3, self.temp_file, reporting)

        team = None
        if getattr(reporting, "firm") and getattr(reporting.firm, "subcompany"):
            team = reporting.firm.subcompany

        LOGO_TEAM = self._imagem(team.logo.url) if getattr(team, "logo") else ""

        list_images_croqui = []

        # limit 10 images croqui
        limit_croqui = 10
        data_croqui = reporting.form_data.get("croqui")

        if data_croqui and isinstance(data_croqui, list):
            for vector_croqui in data_croqui:
                croquis = vector_croqui.get("croqui_image")
                if croquis and isinstance(croquis, list):
                    for file_pk in croquis:
                        if len(list_images_croqui) == limit_croqui:
                            break

                        file = (
                            ReportingFile.objects.filter(uuid=file_pk)
                            .only("upload", "is_shared")
                            .first()
                        )

                        if file and file.is_shared:
                            result = download_picture(
                                s3, self.temp_file, file.uuid, reporting_file=file
                            )
                            if result is not None:
                                list_images_croqui.append(result)

                if len(list_images_croqui) == limit_croqui:
                    break

        reportings_imagens = form_data.get("relatorio", [])
        picture_uuid_list = []

        for _vector in reportings_imagens:
            vector_photos = _vector.get("fotos_relatorio")
            if vector_photos and isinstance(vector_photos, list):
                picture_uuid_list.extend(vector_photos)

        qs_files = ReportingFile.objects.filter(
            uuid__in=picture_uuid_list, is_shared=True
        ).order_by("datetime")

        photos = []
        rest_photos = []

        for file in qs_files:
            file_name = remove_ext_in_filename(file.upload.name)
            file_name = remove_random_string_file_name_in_upload(file_name)

            _result = {
                "file_name": file_name,
                "photo": download_picture(
                    s3, self.temp_file, file.uuid, reporting_file=file
                ),
                "description": file.description,
            }

            if self.valid_rules_order_photo_oae(file_name):
                photos.append(_result)
            else:
                rest_photos.append(_result)

        photos.sort(key=lambda x: x.get("file_name"))
        photos.extend(rest_photos)

        FILE_NAME = get_file_name(reporting)

        data = {
            "inspection_year_campaign": INSPECTION_YEAR_CAMPAIGN,
            "denominacao": DENOMINACAO,
            "codigo_agencia": CODIGO_AGENCIA,
            "zona": ZONA,
            "utm_e1": UTM_E1,
            "utm_e2": UTM_E2,
            "utm_e3": UTM_E3,
            "utm_e4": UTM_E4,
            "utm_n1": UTM_N1,
            "utm_n2": UTM_N2,
            "utm_n3": UTM_N3,
            "utm_n4": UTM_N4,
            "codigo_obra": OAE_NUMERO_CODIGO_OBRA,
            "executed_at": EXECUTED_AT,
            "road_name": ROAD_NAME,
            "road_name_short": ROAD_NAME_SHORT,
            "km": KM,
            "prefix_photo_name": PREFIX_PHOTO_NAME,
            "direction": DIRECTION,
            "condicoes_estabilidade": CONDICOES_ESTABILIDADE,
            "condicoes_conservacao": CONDICOES_CONSERVACAO,
            "nivel_vibracao_tabuleiro": NIVEL_VIBRACAO_TABULEIRO,
            "inspecao_especializada": INSPECAO_ESPECIALIZADA,
            "urgente_antt": URGENTE_ANTT,
            "houve_anteriormente": HOUVE_ANTERIORMENTE,
            "historico_intervencoes_realizadas": HISTORICO_INTERVENCOES_REALIZADAS,
            "observacoes_adicionais": observacoes_adicionais,
            "nota_tecnica_comentarios_gerais": NOTA_TECNICA_COMENTARIOS_GERAIS,
            "nota_tecnica_laje": NOTA_TECNICA_LAJE,
            "buraco_abertura1_laje": buraco_abertura1_laje,
            "buraco_abertura_local1_laje": buraco_abertura_local1_laje,
            "buraco_abertura_observacoes_quantidade1_laje": buraco_abertura_observacoes_quantidade1_laje,
            "armadura_exposta1_laje": armadura_exposta1_laje,
            "armadura_exposta_local1_laje": armadura_exposta_local1_laje,
            "armadura_exposta_observacoes_quantidade1_laje": armadura_exposta_observacoes_quantidade1_laje,
            "concreto_desagregado1_laje": concreto_desagregado1_laje,
            "concreto_desagregado_local1_laje": concreto_desagregado_local1_laje,
            "concreto_desagregado_observacoes_quantidade1_laje": concreto_desagregado_observacoes_quantidade1_laje,
            "fissuras1_laje": fissuras1_laje,
            "fissuras_local1_laje": fissuras_local1_laje,
            "fissuras_observacoes_quantidade1_laje": fissuras_observacoes_quantidade1_laje,
            "marcas_infiltracao1_laje": marcas_infiltracao1_laje,
            "marcas_infiltracao_local1_laje": marcas_infiltracao_local1_laje,
            "marcas_infiltracao_observacoes_quantidade1_laje": marcas_infiltracao_observacoes_quantidade1_laje,
            "aspecto_concreto1_laje": aspecto_concreto1_laje,
            "aspecto_concreto_local1_laje": aspecto_concreto_local1_laje,
            "aspecto_concreto_observacoes_quantidade1_laje": aspecto_concreto_observacoes_quantidade1_laje,
            "cobrimento1_laje": cobrimento1_laje,
            "cobrimento_local1_laje": cobrimento_local1_laje,
            "cobrimento_observacoes_quantidade1_laje": cobrimento_observacoes_quantidade1_laje,
            "nota_tecnica_vigamento_principal": NOTA_TECNICA_VIGAMENTO_PRINCIPAL,
            "fissuras_pequena_abertura2_vigamento_principal": fissuras_pequena_abertura2_vigamento_principal,
            "fissuras_pequena_abertura_local2_vigamento_principal": fissuras_pequena_abertura_local2_vigamento_principal,
            "fissuras_pequena_abertura_observacoes_quantidade2_vigamento_principal": fissuras_pequena_abertura_observacoes_quantidade2_vigamento_principal,
            "trincas2_vigamento_principal": trincas2_vigamento_principal,
            "trincas_local2_vigamento_principal": trincas_local2_vigamento_principal,
            "trincas_observacoes_quantidade2_vigamento_principal": trincas_observacoes_quantidade2_vigamento_principal,
            "armadura_principal2_vigamento_principal": armadura_principal2_vigamento_principal,
            "armadura_principal_local2_vigamento_principal": armadura_principal_local2_vigamento_principal,
            "armadura_principal_observacoes_quantidade2_vigamento_principal": armadura_principal_observacoes_quantidade2_vigamento_principal,
            "desagregamento_concreto2_vigamento_principal": desagregamento_concreto2_vigamento_principal,
            "desagregamento_concreto_local2_vigamento_principal": desagregamento_concreto_local2_vigamento_principal,
            "desagregamento_concreto_observacoes_quantidade2_vigamento_principal": desagregamento_concreto_observacoes_quantidade2_vigamento_principal,
            "dente_geber2_vigamento_principal": dente_geber2_vigamento_principal,
            "dente_geber_local2_vigamento_principal": dente_geber_local2_vigamento_principal,
            "dente_geber_observacoes_quantidade2_vigamento_principal": dente_geber_observacoes_quantidade2_vigamento_principal,
            "deformacao_flecha2_vigamento_principal": deformacao_flecha2_vigamento_principal,
            "deformacao_flecha_local2_vigamento_principal": deformacao_flecha_local2_vigamento_principal,
            "deformacao_flecha_observacoes_quantidade2_vigamento_principal": deformacao_flecha_observacoes_quantidade2_vigamento_principal,
            "aspecto_concreto2_vigamento_principal": aspecto_concreto2_vigamento_principal,
            "aspecto_concreto_local2_vigamento_principal": aspecto_concreto_local2_vigamento_principal,
            "aspecto_concreto_observacoes_quantidade2_vigamento_principal": aspecto_concreto_observacoes_quantidade2_vigamento_principal,
            "cobrimento2_vigamento_principal": cobrimento2_vigamento_principal,
            "cobrimento_local2_vigamento_principal": cobrimento_local2_vigamento_principal,
            "cobrimento_observacoes_quantidade2_vigamento_principal": cobrimento_observacoes_quantidade2_vigamento_principal,
            "nota_tecnica_mesoestrutura": NOTA_TECNICA_MESOESTRUTURA,
            "armadura_exposta3_mesoestrutura": armadura_exposta3_mesoestrutura,
            "armadura_exposta_local3_mesoestrutura": armadura_exposta_local3_mesoestrutura,
            "armadura_exposta_observacoes_quantidade3_mesoestrutura": armadura_exposta_observacoes_quantidade3_mesoestrutura,
            "concreto_desagregado3_mesoestrutura": concreto_desagregado3_mesoestrutura,
            "concreto_desagregado_local3_mesoestrutura": concreto_desagregado_local3_mesoestrutura,
            "concreto_desagregado_observacoes_quantidade3_mesoestrutura": concreto_desagregado_observacoes_quantidade3_mesoestrutura,
            "fissuras3_mesoestrutura": fissuras3_mesoestrutura,
            "fissuras_local3_mesoestrutura": fissuras_local3_mesoestrutura,
            "fissuras_observacoes_quantidade3_mesoestrutura": fissuras_observacoes_quantidade3_mesoestrutura,
            "aparelho_apoio3_mesoestrutura": aparelho_apoio3_mesoestrutura,
            "aparelho_apoio_local3_mesoestrutura": aparelho_apoio_local3_mesoestrutura,
            "aparelho_apoio_observacoes_quantidade3_mesoestrutura": aparelho_apoio_observacoes_quantidade3_mesoestrutura,
            "aspecto_concreto3_mesoestrutura": aspecto_concreto3_mesoestrutura,
            "aspecto_concreto_local3_mesoestrutura": aspecto_concreto_local3_mesoestrutura,
            "aspecto_concreto_observacoes_quantidade3_mesoestrutura": aspecto_concreto_observacoes_quantidade3_mesoestrutura,
            "cobrimento3_mesoestrutura": cobrimento3_mesoestrutura,
            "cobrimento_local3_mesoestrutura": cobrimento_local3_mesoestrutura,
            "cobrimento_observacoes_quantidade3_mesoestrutura": cobrimento_observacoes_quantidade3_mesoestrutura,
            "desaprumo3_mesoestrutura": desaprumo3_mesoestrutura,
            "desaprumo_local3_mesoestrutura": desaprumo_local3_mesoestrutura,
            "desaprumo_observacoes_quantidade3_mesoestrutura": desaprumo_observacoes_quantidade3_mesoestrutura,
            "deslocabilidade_pilares3_mesoestrutura": deslocabilidade_pilares3_mesoestrutura,
            "deslocabilidade_pilares_local3_mesoestrutura": deslocabilidade_pilares_local3_mesoestrutura,
            "deslocabilidade_pilares_observacoes_quantidade3_mesoestrutura": deslocabilidade_pilares_observacoes_quantidade3_mesoestrutura,
            "nota_tecnica_infraestrutura": NOTA_TECNICA_INFRAESTRUTURA,
            "recalque_fundacao4_infraestrutura": recalque_fundacao4_infraestrutura,
            "recalque_fundacao_local4_infraestrutura": recalque_fundacao_local4_infraestrutura,
            "recalque_fundacao_observacoes_quantidade4_infraestrutura": recalque_fundacao_observacoes_quantidade4_infraestrutura,
            "deslocamento_fundacao4_infraestrutura": deslocamento_fundacao4_infraestrutura,
            "deslocamento_fundacao_local4_infraestrutura": deslocamento_fundacao_local4_infraestrutura,
            "deslocamento_fundacao_observacoes_quantidade4_infraestrutura": deslocamento_fundacao_observacoes_quantidade4_infraestrutura,
            "erosao_terreno_fundacao4_infraestrutura": erosao_terreno_fundacao4_infraestrutura,
            "erosao_terreno_fundacao_local4_infraestrutura": erosao_terreno_fundacao_local4_infraestrutura,
            "erosao_terreno_fundacao_observacoes_quantidade4_infraestrutura": erosao_terreno_fundacao_observacoes_quantidade4_infraestrutura,
            "estacas_dessenterradas4_infraestrutura": estacas_dessenterradas4_infraestrutura,
            "estacas_dessenterradas_local4_infraestrutura": estacas_dessenterradas_local4_infraestrutura,
            "estacas_dessenterradas_observacoes_quantidade4_infraestrutura": estacas_dessenterradas_observacoes_quantidade4_infraestrutura,
            "nota_tecnica_pista_acesso": NOTA_TECNICA_PISTA_ACESSO,
            "irregularidades_pavimento5_pista_acesso": irregularidades_pavimento5_pista_acesso,
            "irregularidades_pavimento_local5_pista_acesso": irregularidades_pavimento_local5_pista_acesso,
            "irregularidades_pavimento_observacoes_quantidade5_pista_acesso": irregularidades_pavimento_observacoes_quantidade5_pista_acesso,
            "junta_dilatacao5_pista_acesso": junta_dilatacao5_pista_acesso,
            "junta_dilatacao_local5_pista_acesso": junta_dilatacao_local5_pista_acesso,
            "junta_dilatacao_observacoes_quantidade5_pista_acesso": junta_dilatacao_observacoes_quantidade5_pista_acesso,
            "acesso_oae5_pista_acesso": acesso_oae5_pista_acesso,
            "acesso_oae_local5_pista_acesso": acesso_oae_local5_pista_acesso,
            "acesso_oae_observacoes_quantidade5_pista_acesso": acesso_oae_observacoes_quantidade5_pista_acesso,
            "acidente_veiculos5_pista_acesso": acidente_veiculos5_pista_acesso,
            "acidente_veiculos_local5_pista_acesso": acidente_veiculos_local5_pista_acesso,
            "acidente_veiculos_observacoes_quantidade5_pista_acesso": acidente_veiculos_observacoes_quantidade5_pista_acesso,
            "guarda_corpo": guarda_corpo,
            "guarda_corpo_observacoes_quantidade": guarda_corpo_observacoes_quantidade,
            "drenagem_limpeza": drenagem_limpeza,
            "drenagem_limpeza_observacoes_quantidade": drenagem_limpeza_observacoes_quantidade,
            "placa_gabarito_vertical": placa_gabarito_vertical,
            "placa_gabarito_vertical_observacoes_quantidade": placa_gabarito_vertical_observacoes_quantidade,
            "juntas_dilatacao_vida_util_remanescente": juntas_dilatacao_vida_util_remanescente,
            "juntas_dilatacao_vida_util_remanescente_observacoes_quantidade": juntas_dilatacao_vida_util_remanescente_observacoes_quantidade,
            "aparelho_apoio_vida_util_remanescente": aparelho_apoio_vida_util_remanescente,
            "aparelho_apoio_vida_util_remanescente_local": aparelho_apoio_vida_util_remanescente_local,
            "images_croqui": list_images_croqui,
            "photos": photos,
            "logo": LOGO,
            "file_name": FILE_NAME,
            "reporting": reporting,
            "logo_team": LOGO_TEAM,
        }
        return data

    def execute(self):
        data = [
            self.create_dict(reporting=reporting, s3=self.s3)
            for reporting in self.list_reporting
        ]
        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRReportMonitoringOAE(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""
        occurrence_type_uuid = (
            Reporting.objects.filter(uuid=self.uuids[0])
            .only("occurrence_type__uuid")[0]
            .occurrence_type.uuid
        )
        reportings = Reporting.objects.filter(
            occurrence_type__uuid=occurrence_type_uuid, pk__in=self.uuids
        ).prefetch_related("road")

        reporting = reportings.first()

        road_name = reporting.road_name

        if reportings.count() == 1:
            reporting = reportings.first()
            extension = ""
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{get_file_name(reporting)}.{extension}"
        else:
            file_name = f"{road_name} - Fichas Anexo II.zip"

        return file_name

    def _get_repotings_obj(self):
        occurrence_type_uuid = (
            Reporting.objects.filter(uuid=self.uuids[0])
            .only("occurrence_type__uuid")[0]
            .occurrence_type.uuid
        )
        query_set = Reporting.objects.filter(
            occurrence_type__uuid=occurrence_type_uuid, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def export(self):
        list_reporting = self._get_repotings_obj()
        s3 = get_s3()
        obj = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()
        files = obj["files"]
        road_names = ("-").join(list(set([x.road_name for x in list_reporting])))
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            self.file_name = f"{road_names} - Fichas Anexo II.zip"
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_monitoring_oae_async_handler(reporter_dict: dict):
    reporter = CCRReportMonitoringOAE.from_dict(reporter_dict)
    reporter.export()
