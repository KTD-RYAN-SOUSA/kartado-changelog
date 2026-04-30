import re
from datetime import datetime, timedelta
from typing import List, Union
from zipfile import ZipFile

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.base_ccr_xlsx_handler import BaseXlsxHandler
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import ReportFormat
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import deep_keys_to_snake_case, format_km, to_camel_case


def _get_name_to_file() -> str:
    return "Anexo VI - Diagnóstico e Ações - OAC"


def _get_file_name_to_zip(all_roads: list) -> str:
    tags = re.sub(r"[-\ ]", "", "_".join(all_roads))
    return f"{_get_name_to_file()} {tags}.zip"


def _get_file_name(road_name, form_data__inspection_campaign_year) -> str:
    return (
        f"{_get_name_to_file()} - {road_name} - {form_data__inspection_campaign_year}"
    )


class XlsxHandler(BaseXlsxHandler):
    def __init__(self, **kwargs) -> None:
        self.title = (
            "Anexo VI - Diagnóstico e Ações - Obras de Arte Corrente - {road_name}"
        )
        self.bond_class_pk = None
        self.height_row = 69.75
        self._api_name_status_conservation = [
            "brokenmon",
            "erosionmon",
            "rocada_mon",
            "cleaningmon",
            "box_damagemon",
            "tube_damagemon",
            "cover_damagemon",
            "desobstrucaomon",
            "desassoreamentomon",
            "forehead_damagemon",
        ]
        self.static_fields = {
            "data_sheet": "A",
            "road_name": "B",
            "initial_km": "C",
            "end_km": "D",
            "element": "E",
            "montante_direction": "F",
            "montante_estado_conservacao_caixa": "G",
            "montante_action": "H",
            "montante_material_rev_mont": "I",
            "jusante_direction": "J",
            "jusante_estado_conservacao_caixa": "K",
            "jusante_action": "L",
            "jusante_material_rev_mont": "M",
            "montante_ambient_mon": "N",
            "executed_at": "O",
            "montante_executed_at": "P",
            "jusante_executed_at": "Q",
            "hidden_serial_monitoring": "R",
            "hidden_serial_inventory": "S",
        }
        super().__init__(
            path_file_xlsx="./fixtures/reports/ccr_report_action_diagnosis.xlsx",
            split_xlsx_in=["road_name", "inspection_campaign_year"],
            logo_company_range_string="Q1:Q1",
            provider_logo_range_string="A1:A1",
            **kwargs,
        )

    @classmethod
    def format_fonts(
        cls,
        *,
        cell,
        name="Calibri",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        wrap_text: bool = True,
    ) -> None:
        cell.alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrap_text=wrap_text
        )
        cell.font = Font(name=name, sz=size, bold=bold)

    def fill_sheet(self, data_work: list) -> list:
        self.worksheet.freeze_panes = "A4"

        files = []
        filenames = []

        for key, data in data_work.items():
            road_name, year = key.split("|")
            self.worksheet.title = road_name
            initial_row = 5
            cell_title = "A1"

            self.format_fonts(
                cell=self.worksheet[cell_title],
                size=12,
                bold=True,
            )

            title = self.title.format(road_name=road_name)
            self.worksheet[cell_title] = title

            for values in data:
                for key, value in values.items():
                    self.insert_new_row(
                        row=initial_row,
                        number_col=len(self.static_fields),
                        height_row=self.height_row,
                    )
                    if key not in self.static_fields:
                        continue

                    col = self.static_fields[key]
                    cell = f"{col}{initial_row}"

                    data_format = dict(cell=self.worksheet[cell], size=11)
                    self.worksheet[cell] = value
                    self.format_fonts(**data_format)

                initial_row += 1

            last_col = self.worksheet.max_column - 2
            self.insert_logos(
                logo_company_config=self.data_logo_company,
                provider_logo_config=self.data_provider_logo,
            )

            self.worksheet.print_area = (
                f"A1:{get_column_letter(last_col)}{self.worksheet.max_row}"
            )

            file_name = _get_file_name(road_name, year)
            filenames.append(file_name)
            result = f"/tmp/{file_name}.xlsx"
            self.workbook.save(result)
            self.reload_workbook()
            files.append(result)

        return files

    def __add_month(self, date, delta):
        meses_ingles_para_portugues = {
            "Jan": "Jan",
            "Feb": "Fev",
            "Mar": "Mar",
            "Apr": "Abr",
            "May": "Mai",
            "Jun": "Jun",
            "Jul": "Jul",
            "Aug": "Ago",
            "Sep": "Set",
            "Oct": "Out",
            "Nov": "Nov",
            "Dec": "Dez",
        }

        initial_date = date.strftime("%b/%Y")
        format_initial_month = initial_date.split("/")
        new_string_initial_month = meses_ingles_para_portugues[format_initial_month[0]]

        new_date = date + timedelta(days=delta)
        new_date = new_date.strftime("%b/%Y")
        format_month = new_date.split("/")
        new_string_month = meses_ingles_para_portugues[format_month[0]]
        new_date = f"{new_string_initial_month}/{format_initial_month[1]} a {new_string_month}/{format_month[1]}"

        return new_date

    def create_dict(self, reporting: Reporting) -> dict:
        DATA_SHEET = reporting.get_single_form_data_display("id_ccr_antt", "-")
        ROAD_NAME = reporting.road_name
        INSPECTION_CAMPAIGN_YEAR = str(
            reporting.get_single_form_data_display("inspection_campaign_year", "")
        )
        KM_INITIAL = format_km(reporting, "km", 3)
        KM = reporting.km
        KM_END = format_km(reporting, "end_km", 3)
        ELEMENT = new_get_form_data(reporting, "holeKindCelular", default="-")

        obj_montante = None
        obj_jusante = None
        if not self.bond_class_pk:
            self.bond_class_pk = reporting.company.metadata.get(
                "report_action_diagnosis_element_form_id"
            )

        if not self.bond_class_pk:
            self.bond_class_pk = reporting.company.metadata.get(
                "report_action_diagnosis_element_form_id"
            )

        for _ in reporting.reporting_relation_parent.all():
            if str(_.child.occurrence_type_id) == self.bond_class_pk:
                if (_.reporting_relation.name).lower() == "montante":
                    obj_montante = _.child
                elif (_.reporting_relation.name).lower() == "jusante":
                    obj_jusante = _.child

        date_executed_at = reporting.executed_at
        executed_at = "-"

        if isinstance(date_executed_at, datetime):
            executed_at = date_executed_at.strftime("%d/%m/%Y")

        HIDDEN_SERIAL_MONITORING = reporting.number
        HIDDEN_SERIAL_INVENTORY = (
            reporting.parent.number if getattr(reporting, "parent") else "-"
        )

        MONTANTE_DIRECTION = "-"
        MONTANTE_ESTADO_CONSERVACAO_CAIXA = "-"
        montante_pathology = "-"
        MONTANTE_MATERIAL_REV_MONT = "-"
        JUSANTE_DIRECTION = "-"
        JUSANTE_ESTADO_CONSERVACAO_CAIXA = "-"
        jusante_pathology = "-"
        JUSANTE_MATERIAL_REV_MONT = "-"
        MONTANTE_AMBIENT_MON = "-"
        montante_executed_at = "-"
        jusante_executed_at = "-"

        if obj_montante:

            montante_form_fields: dict = obj_montante.occurrence_type.form_fields
            montante_form_data: dict = obj_montante.form_data
            MONTANTE_DIRECTION = get_custom_option(obj_montante, "direction", "-")

            MONTANTE_ESTADO_CONSERVACAO_CAIXA = (
                obj_montante.get_single_form_data_display("hole_classification", "-")
            )

            default_pathology = "Monitorar"
            good_montante_conservation = ""

            good_montante_conservation = ""
            if MONTANTE_ESTADO_CONSERVACAO_CAIXA:
                good_montante_conservation = (
                    montante_form_data.get("hole_classification") == "1"
                )

                montante_pathology = (
                    default_pathology if good_montante_conservation else ""
                )

                prev_day_montante_conservation = 90
                if (
                    not good_montante_conservation
                    and montante_form_data.get("hole_classification") == "3"
                ):
                    prev_day_montante_conservation = 30

            if not montante_pathology:
                fields = deep_keys_to_snake_case(montante_form_fields.get("fields", {}))
                list_pathology = []
                for status in self._api_name_status_conservation:
                    if montante_form_data.get(status):
                        status = to_camel_case(status)
                        for field in fields:
                            if field.get("api_name") == status:
                                list_pathology.append(field.get("display_name"))
                                break

                montante_pathology = (", ").join(list_pathology)

            MONTANTE_MATERIAL_REV_MONT = obj_montante.get_single_form_data_display(
                "material_rev_mont", "-"
            )

            MONTANTE_AMBIENT_MON = obj_montante.get_single_form_data_display(
                "ambient_mon", "-"
            )

            if isinstance(date_executed_at, datetime):
                montante_executed_at = (
                    "-"
                    if good_montante_conservation
                    or not MONTANTE_ESTADO_CONSERVACAO_CAIXA
                    else f"{self.__add_month(date_executed_at, prev_day_montante_conservation)}"
                )

        if obj_jusante:
            jusante_form_fields: dict = obj_jusante.occurrence_type.form_fields
            jusante_form_data: dict = obj_jusante.form_data
            JUSANTE_DIRECTION = get_custom_option(obj_jusante, "direction", "-")

            JUSANTE_ESTADO_CONSERVACAO_CAIXA = obj_jusante.get_single_form_data_display(
                "hole_classification", "-"
            )

            good_jusante_conservation = ""
            if JUSANTE_ESTADO_CONSERVACAO_CAIXA:
                good_jusante_conservation = (
                    jusante_form_data.get("hole_classification") == "1"
                )
                jusante_pathology = (
                    default_pathology if good_jusante_conservation else ""
                )

                prev_day_jusante_conservation = 90
                if (
                    not good_jusante_conservation
                    and jusante_form_data.get("hole_classification") == "3"
                ):
                    prev_day_jusante_conservation = 30

            if not jusante_pathology:
                fields = deep_keys_to_snake_case(jusante_form_fields.get("fields", {}))
                list_pathology = []
                for status in self._api_name_status_conservation:
                    if jusante_form_data.get(status):
                        status = to_camel_case(status)
                        for field in fields:
                            if field.get("api_name") == status:
                                display = str(field.get("display_name", "")).strip()
                                if display:
                                    list_pathology.append(display)
                                break

                jusante_pathology = (", ").join(list_pathology)

            JUSANTE_MATERIAL_REV_MONT = obj_jusante.get_single_form_data_display(
                "material_rev_mont", "-"
            )

            if isinstance(date_executed_at, datetime):
                jusante_executed_at = (
                    "-"
                    if good_jusante_conservation or not JUSANTE_ESTADO_CONSERVACAO_CAIXA
                    else f"{self.__add_month(date_executed_at, prev_day_jusante_conservation)}"
                )

        data = {
            "inspection_campaign_year": INSPECTION_CAMPAIGN_YEAR,
            "km": KM,
            "data_sheet": DATA_SHEET,
            "initial_km": KM_INITIAL,
            "end_km": KM_END,
            "element": ELEMENT,
            "montante_direction": MONTANTE_DIRECTION,
            "montante_estado_conservacao_caixa": MONTANTE_ESTADO_CONSERVACAO_CAIXA,
            "montante_pathology": montante_pathology,
            "montante_material_rev_mont": MONTANTE_MATERIAL_REV_MONT,
            "jusante_direction": JUSANTE_DIRECTION,
            "jusante_estado_conservacao_caixa": JUSANTE_ESTADO_CONSERVACAO_CAIXA,
            "jusante_pathology": jusante_pathology,
            "jusante_material_rev_mont": JUSANTE_MATERIAL_REV_MONT,
            "montante_ambient_mon": MONTANTE_AMBIENT_MON,
            "executed_at": executed_at,
            "montante_executed_at": montante_executed_at,
            "jusante_executed_at": jusante_executed_at,
            "road_name": ROAD_NAME,
            "hidden_serial_monitoring": HIDDEN_SERIAL_MONITORING,
            "hidden_serial_inventory": HIDDEN_SERIAL_INVENTORY,
        }

        return data


class XlsxHandlerReportActionDiagnosisAnnex6(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        self.uuids = uuids
        self.tags: Union[List[str], List[dict]] = []
        self.queryset_reportings_serializer: str = ""
        if uuids:
            self.__set_class(uuids)
        super().__init__(uuids, report_format)

    def get_file_name(self) -> str:
        if len(self.tags) == 1:
            file = _get_file_name(**(self.tags[0]))
            extension = "xlsx"
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"

            file_name = f"{file}.{extension}"

        else:
            self.tags.sort()
            file_name = _get_file_name_to_zip(self.tags)

        return file_name

    def __set_class(self, uuids):
        query_set = Reporting.objects.filter(uuid__in=uuids, company__isnull=False)
        company = query_set.first().company
        metadata: dict = deep_keys_to_snake_case(company.metadata)

        class_pk = metadata["report_action_diagnosis_monitoring_form_id"]

        query_set = (
            query_set.filter(occurrence_type_id=class_pk)
            .prefetch_related(
                "occurrence_type",
                "company",
                "parent",
                "reporting_relation_parent",
                "reporting_relation_parent__child",
                "reporting_relation_parent__reporting_relation",
                "reporting_relation_parent__child__company",
                "reporting_relation_parent__child__occurrence_type",
            )
            .order_by("road_name")
        ).distinct()
        self.uuids = [str(x) for x in query_set.values_list("pk", flat=True)]

        tags = list(
            query_set.values("road_name", "form_data__inspection_campaign_year")
        )
        if len(set([x["road_name"] for x in tags])) == 1:
            self.tags = tags
        else:
            self.tags = list(query_set.values_list("road_name", flat=True))

        self.queryset_reportings_serializer = self.serializer_queryset(query_set)

    def export(self):
        s3 = get_s3()
        files = XlsxHandler(
            query_set_serializer=self.queryset_reportings_serializer,
            sheet_target=self.sheet_target(),
        ).execute()
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_action_diagnosis_annex_6_async_handler(reporter_dict: dict):
    reporter = XlsxHandlerReportActionDiagnosisAnnex6.from_dict(reporter_dict)
    reporter.export()
