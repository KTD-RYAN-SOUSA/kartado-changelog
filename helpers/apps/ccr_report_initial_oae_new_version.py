import re
import shutil
import tempfile
from typing import List
from zipfile import ZipFile

import pyproj
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_picture,
    get_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.apps.ccr_report_utils.workbook_utils import set_zoom
from helpers.strings import (
    ZONE_MAP,
    format_km,
    int_set_zero_prefix,
    keys_to_snake_case,
    remove_ext_in_filename,
    remove_random_string_file_name_in_upload,
)


def get_file_name(reporting: Reporting):
    inspection_year_campaign = keys_to_snake_case(reporting.form_data).get(
        "inspection_year_campaign", ""
    )
    road_number = re.sub(r"\D", "", reporting.road_name)
    n_oae = new_get_form_data(reporting, "oaeNumeroCodigoObra", default="")
    try:
        n_oae = f"{int(n_oae):03}"
    except Exception:
        pass
    uf = reporting.road_name[-2:]
    file_name = f"OAE{inspection_year_campaign}{road_number}{n_oae}{uf}"
    return file_name


class XlsxHandler:
    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.__list_reporting = list_reporting
        self.__sheet_target = sheet_target
        self.temp_file = tempfile.mkdtemp()
        self.__xlsx_file = "./fixtures/reports/ccr_report_rotineira_OAE_novo.xlsx"
        self._workbook = load_workbook(self.__xlsx_file)
        # Custom config
        self.default_name_excel_sheet = "Rotineira OAE"
        self.default_name_excel_photo = "Rotineira OAE - Fotos 1"
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="F40:H40",
            resize_method=ResizeMethod.ProportionalCentered,
        )
        self.sheet_counter = 1

        self.form_fields = self.__list_reporting[0].occurrence_type.form_fields[
            "fields"
        ]

        # Extrai as opções de região e element_family do form_fields
        self.region_options = []
        self.element_family_options = []
        self.damage_options = []
        self.condition_state_options = []
        self.estrural_ineficiencies_options = []
        if self.form_fields:
            for field in self.form_fields:
                if field.get("apiName") == "therapy" and "innerFields" in field:
                    for inner_field in field["innerFields"]:
                        if (
                            inner_field.get("apiName") == "region"
                            and "selectOptions" in inner_field
                        ):
                            self.region_options = inner_field["selectOptions"][
                                "options"
                            ]
                        elif (
                            inner_field.get("apiName") == "elementFamily"
                            and "selectOptions" in inner_field
                        ):
                            self.element_family_options = inner_field["selectOptions"][
                                "options"
                            ]
                        elif (
                            inner_field.get("apiName") == "damage"
                            and "selectOptions" in inner_field
                        ):
                            self.damage_options = inner_field["selectOptions"][
                                "options"
                            ]
                        elif (
                            inner_field.get("apiName") == "conditionState"
                            and "selectOptions" in inner_field
                        ):
                            self.condition_state_options = inner_field["selectOptions"][
                                "options"
                            ]
                elif (
                    field.get("apiName") == "insuficienciaEstruturais"
                    and "innerFields" in field
                ):
                    for inner_field in field["innerFields"]:
                        if (
                            inner_field.get("apiName") == "insuficienciaEstrutural"
                            and "selectOptions" in inner_field
                        ):
                            self.estrural_ineficiencies_options = inner_field[
                                "selectOptions"
                            ]["options"]

        self.croqui_range_string = "A25:AC53"
        self.max_sheets_croqui = 4
        self.max_photo_to_sheets = 6
        self.salt_row_default = 20

        self.static_fields_sheet = {}
        self.static_fields_photo = {}

        self.static_basic_fields = {
            "provider_name": "A2",
            "oae_number": "O2",
            "executed_at": "B7",
            "road_name": "B9",
            "initial_km": "B11",
            "oae": "N7",
            "denomination": "N9",
            "direction": "N11",
            "general_tech_grade": "N14",
            "notes_1": "A18",  # Identificação do inspetor
            "notes_2": "C18",  # Coordenadas geográficas
            "notes_3": "M18",  # Histórico de intervenções realizadas
        }

        self.static_performance_parameters = {
            "guarda_corpo_param": "D30/H30",
            "guarda_corpo_observ": "M30",
            "drenagem_limpagem_param": "D32/H32",
            "drenagem_limpagem_observ": "M32",
            "placa_gabarito_vertical_param": "D34/H34/K34",
            "placa_gabarito_vertical_observ": "M34",
            "juntas_dilatacao_param": "D36/H36",
            "juntas_dilatacao_observ": "M36",
            "aparelho_apoio_param": "D38/H38",
            "aparelho_apoio_observ": "M38",
        }

        self.provider_name_translation = {
            "CCR - Rio Sp": "RIOSP - Concessionária do Sistema Rodoviário Rio - São Paulo S.A.",
        }

    def _generate_notes(self, form_data_display: dict) -> dict:
        """Gera as notas separadas com base nos dados do formulário"""
        notes = {
            "notes_1": "-",  # Identificação do inspetor - A18
            "notes_2": "-",  # Coordenadas geográficas - C18
            "notes_3": "-",  # Histórico de intervenções realizadas - J18
        }

        # Identificação do inspetor
        empresa_antt = form_data_display.get("empresa_antt", "-")
        if empresa_antt:
            notes["notes_1"] = f"Identificação do inspetor\n{empresa_antt}"

        # Coordenadas geográficas
        zona = form_data_display.get("zona", "-")

        coord_parts = []
        if zona and zona != "-":
            coord_parts.append(f"Coordenadas geográficas\nZona: {zona}")

            # Processa todas as UTMs (1 a 4)
            for i in range(1, 5):
                utm_e = form_data_display.get(f"utm_e{i}", "")
                utm_n = form_data_display.get(f"utm_n{i}", "")

                if utm_e and utm_n:
                    try:
                        lat, lon = self._convert_utm_to_lat_lon(
                            float(utm_e), float(utm_n), zona
                        )
                        if lat is not None and lon is not None:
                            coord_parts.append(
                                f"{i} - Latitude: {lat} Longitude: {lon}"
                            )
                    except (ValueError, TypeError):
                        pass

        if coord_parts:
            notes["notes_2"] = "\n".join(coord_parts)

        # Histórico de intervenções realizadas
        historico = form_data_display.get("historico_intervencoes_realizadas", "-")
        if historico:
            notes["notes_3"] = f"Histórico de intervenções realizadas\n{historico}"

        return notes

    def _convert_utm_to_lat_lon(self, utm_e, utm_n, zona):
        """Converte coordenadas UTM para latitude/longitude"""
        try:
            # Verifica se a zona existe no mapeamento
            if zona not in ZONE_MAP:
                return None, None

            # Define o sistema de coordenadas UTM
            utm_proj = pyproj.Proj(ZONE_MAP[zona])

            # Define o sistema de coordenadas WGS84 (lat/lon)
            lat_lon_proj = pyproj.Proj(proj="latlong", datum="WGS84")

            # Converte UTM para lat/lon
            lon, lat = pyproj.transform(utm_proj, lat_lon_proj, utm_e, utm_n)

            return round(lat, 6), round(lon, 6)
        except Exception:
            return None, None

    def _get_option_name(self, value, options):
        """Converte um valor para o nome legível baseado nas opções fornecidas"""
        if not options:
            return value

        # Usa next() para encontrar a primeira opção que corresponde ao valor
        option = next((opt for opt in options if opt.get("value") == str(value)), None)
        return option.get("name", value) if option else value

    def _remove_sheet(self, sheet_name):
        """Remove uma aba específica do workbook"""
        try:
            if sheet_name in self._workbook.sheetnames:
                self._workbook.remove(self._workbook[sheet_name])
            else:
                print(f"⚠️ Aba '{sheet_name}' não encontrada")
        except Exception as e:
            print(f"❌ Erro ao remover aba '{sheet_name}': {e}")

    def _reorder_sheets(self, sheet, position):
        """Reordena as abas do workbook na ordem desejada"""
        self._workbook._sheets.remove(sheet)
        self._workbook._sheets.insert(position, sheet)

    def _group_therapy_by_section(self, therapy_array):
        """Agrupa o array THERAPY pelo campo section de forma crescente"""
        if not therapy_array:
            return []

        # Agrupa por section
        grouped = {}
        for item in therapy_array:
            section = item.get("section", 0)
            if section not in grouped:
                grouped[section] = []
            grouped[section].append(item)

        # Ordena as seções e retorna como array de arrays
        sorted_sections = sorted(grouped.keys())
        return [grouped[section] for section in sorted_sections]

    def valid_rules_order_photo_oae(self, file_name: str) -> bool:
        """
        Utilizando o nome do arquivo como referência, aparecem primeiro
        as fotos com menor valor dos últimos 3 dígitos do nome do arquivo
        (os quais devem obrigatoriamente ser 3 números). O formato esperado para
        os nomes de arquivo segue o padrão do exemplo: OAE2023101095RJF001, em que:

        OAE = OAE
        2023 = ano
        101 = rodovia
        095 = Nº OAE
        SP ou RJ
        F001 = nº foto
        """
        try:
            file_name = file_name.upper()
            return (
                len(file_name) == 19
                and file_name[15] == "F"
                and file_name.startswith("OAE")  # OAE
                and file_name[3:7].isdigit()  # YEAR
                and file_name[7:10].isdigit()  # ROAD
                and file_name[10:13].isdigit()  # Nº OAE
                and file_name[13:15].isalpha()  # UF
                and file_name[16:].isdigit()  # Nº PHOTO
            )
        except Exception as e:
            print(f"Error: {e}")
            return False

    def _fill_basic_fields_sheet(self, data):
        self._worksheet = self._workbook["Dados Básicos"]

        for field, column in self.static_basic_fields.items():
            _data = data.get(field, None)

            if _data is None:
                continue

            try:
                self._worksheet[column] = _data
            except Exception as err:
                print(err)

        for field, column in self.static_performance_parameters.items():
            _data = data.get(field, None)

            if _data is None:
                continue

            try:
                if "/" in column:
                    parts = column.split("/")
                    value_positive = ["Bom", "Presente", "Atende"]
                    value_negative = [
                        "Ruim/Ausente",
                        "Ruim",
                        "Ausente",
                        "Não atende",
                    ]
                    value_na = ["Não se aplica"]

                    target_idx = None
                    if _data in value_positive:
                        target_idx = 0
                    elif _data in value_negative:
                        target_idx = 1 if len(parts) >= 2 else None
                    elif _data in value_na:
                        target_idx = 2 if len(parts) >= 3 else None

                    if target_idx is not None and target_idx < len(parts):
                        self._worksheet[parts[target_idx]] = "X"
                else:
                    self._worksheet[column] = _data
            except Exception as err:
                print(err)

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            logo_company=self.data_logo_company,
            target=self.__sheet_target,
        )

    def _fill_technical_grade_fields_sheet(self, data):
        therapy_grouped = data.get("therapy_grouped", [])

        for index, therapy_section in enumerate(therapy_grouped):
            line = 10
            section_number = therapy_section[0].get("section", "-")
            # Cria uma cópia da aba template
            template_sheet = self._workbook["Nota técnica Template"]
            new_sheet = self._workbook.copy_worksheet(template_sheet)
            new_sheet.title = f"Nota técnica - Tramo {section_number}"

            # Reordena as abas na ordem desejada
            self.sheet_counter += 1
            self._reorder_sheets(new_sheet, self.sheet_counter)

            new_sheet["B7"] = section_number

            for item in therapy_section:
                region_name = self._get_option_name(
                    item.get("region", "-"), self.region_options
                )
                element_family_name = self._get_option_name(
                    item.get("element_family", "-"), self.element_family_options
                )

                new_sheet[f"A{line}"] = region_name
                new_sheet[f"B{line}"] = element_family_name
                new_sheet[f"C{line}"] = item.get("individual_element", "-")
                new_sheet[f"D{line}"] = item.get("technical_note", "-")
                line += 1

            insert_picture_2(
                picture=Image(self.data_logo_company.get("path_image")),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="B40:C40",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalCentered
                ),
                border_width=(2, 2, 2, 2),
            )

            insert_picture_2(
                picture=Image("assets/static/antt_logo.png"),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="D40",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalRight
                ),
                border_width=(2, 2, 2, 2),
            )

        # Remove a aba template após criar todas as abas necessárias
        self._remove_sheet("Nota técnica Template")

    def _fill_elements_damage_fields_sheet(self, data):
        therapy_grouped = data.get("therapy_grouped", [])

        for index, therapy_section in enumerate(therapy_grouped):
            line = 10
            # Cria uma cópia da aba template
            template_sheet = self._workbook["Dano ao Elemento Template"]
            section_number = therapy_section[0].get("section", "-")
            new_sheet = self._workbook.copy_worksheet(template_sheet)
            new_sheet.title = f"Dano ao Elemento - Tramo {section_number}"

            # Reordena as abas na ordem desejada
            self.sheet_counter += 1
            self._reorder_sheets(new_sheet, self.sheet_counter)

            new_sheet["B7"] = section_number

            for item in therapy_section:
                region_name = self._get_option_name(
                    item.get("region", "-"), self.region_options
                )
                element_family_name = self._get_option_name(
                    item.get("element_family", "-"), self.element_family_options
                )
                damage_name = self._get_option_name(
                    item.get("damage", "-"), self.damage_options
                )
                condition_state_name = self._get_option_name(
                    item.get("condition_state", "-"), self.condition_state_options
                )

                new_sheet[f"A{line}"] = region_name
                new_sheet[f"B{line}"] = element_family_name
                new_sheet[f"C{line}"] = item.get("individual_element", "-")
                new_sheet[f"D{line}"] = damage_name
                new_sheet[f"E{line}"] = item.get("amount", "-")
                new_sheet[f"F{line}"] = item.get("location", "-")
                new_sheet[f"G{line}"] = item.get("coord_x", "-")
                new_sheet[f"H{line}"] = item.get("coord_y", "-")
                new_sheet[f"I{line}"] = item.get("coord_z", "-")
                new_sheet[f"J{line}"] = item.get("relative_extension", "-")
                new_sheet[f"K{line}"] = condition_state_name
                line += 1

            insert_picture_2(
                picture=Image(self.data_logo_company.get("path_image")),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="D40:F40",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalCentered
                ),
                border_width=(2, 2, 2, 2),
            )

            insert_picture_2(
                picture=Image("assets/static/antt_logo.png"),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="K40",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalRight
                ),
                border_width=(2, 2, 2, 2),
            )

        # Remove a aba template após criar todas as abas necessárias
        self._remove_sheet("Dano ao Elemento Template")

    def _fill_inspection_report_fields_sheet(self, data):
        inspection_report_grouped = data.get("inspection_report_grouped", [])

        for index, inspection_report_section in enumerate(inspection_report_grouped):
            line = 10
            # Cria uma cópia da aba template
            template_sheet = self._workbook["Insuficiências Template"]
            section_number = inspection_report_section[0].get("section", "-")
            new_sheet = self._workbook.copy_worksheet(template_sheet)
            new_sheet.title = f"Insuficiências - Tramo {section_number}"

            # Reordena as abas na ordem desejada
            self.sheet_counter += 1
            self._reorder_sheets(new_sheet, self.sheet_counter)

            new_sheet["B7"] = section_number

            for item in inspection_report_section:
                region_name = self._get_option_name(
                    item.get("region_ie", "-"), self.region_options
                )
                element_family_name = self._get_option_name(
                    item.get("element_family_ie", "-"), self.element_family_options
                )
                estrural_ineficiencies_name = self._get_option_name(
                    item.get("insuficiencia_estrutural", "-"),
                    self.estrural_ineficiencies_options,
                )

                new_sheet[f"A{line}"] = region_name
                new_sheet[f"B{line}"] = element_family_name
                new_sheet[f"C{line}"] = item.get("individual_element_ie", "-")
                new_sheet[f"D{line}"] = estrural_ineficiencies_name
                new_sheet[f"E{line}"] = item.get("causa_provavel", "-")
                line += 1

            insert_picture_2(
                picture=Image(self.data_logo_company.get("path_image")),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="C41",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalCentered
                ),
                border_width=(2, 2, 2, 2),
            )

            insert_picture_2(
                picture=Image("assets/static/antt_logo.png"),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="E41",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalRight
                ),
                border_width=(2, 2, 2, 2),
            )

        # Remove a aba template após criar todas as abas necessárias
        self._remove_sheet("Insuficiências Template")

    def _fill_appraisal_fields_sheet(self, data):
        appraisal_date = data.get("appraisal_date", [])
        appraisal_text = data.get("appraisal_text", [])

        self._worksheet = self._workbook["Laudo Especializado"]

        self._worksheet["B7"] = appraisal_date
        self._worksheet["B9"] = appraisal_text

        insert_picture_2(
            picture=Image(self.data_logo_company.get("path_image")),
            worksheet=self._worksheet,
            target=self.__sheet_target,
            range_string="B41:C41",
            resize_method=self.data_logo_company.get(
                "resize_method", ResizeMethod.ProportionalCentered
            ),
            border_width=(2, 2, 2, 2),
        )

    def _fill_croqui_photos_sheet(self, data):
        images_croqui = data.get("images_croqui")
        template_sheet = self._workbook["Croqui Template"]

        for croqui_index, croqui_photo in enumerate(images_croqui):
            # Cria uma nova aba para este grupo
            new_sheet = self._workbook.copy_worksheet(template_sheet)
            new_sheet.title = f"Croqui {croqui_index + 1}"

            # Insere a imagem na seção especificada
            insert_picture_2(
                picture=Image(croqui_photo),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="A7:Z55",
                resize_method=ResizeMethod.Stretch,
                border_width=(2, 2, 2, 2),
            )

            insert_picture_2(
                picture=Image(self.data_logo_company.get("path_image")),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="K56:M56",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalCentered
                ),
                border_width=(2, 2, 2, 2),
            )

            insert_picture_2(
                picture=Image("assets/static/antt_logo.png"),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="X56:Y56",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalRight
                ),
                border_width=(2, 2, 2, 2),
            )

            # Reordena as abas na ordem desejada
            self.sheet_counter += 1
            self._reorder_sheets(new_sheet, self.sheet_counter)

        # Remove a aba template após criar todas as abas necessárias
        self._remove_sheet("Croqui Template")

    def _fill_photos_sheet(self, data):
        photos = data.get("photos")
        prefix_photo_name = data.get("prefix_photo_name")

        template_sheet = self._workbook["Fotos Template"]
        # Define as seções onde as imagens serão inseridas
        image_sections = ["C6:J29", "N6:X28", "C33:J57", "N33:X57"]
        name_sections = ["C30", "N30", "C58", "N58"]
        description_sections = ["E30", "P30", "E58", "P58"]

        # Processa as fotos em grupos de 4
        for group_index in range(0, len(photos), 4):
            # Sai do loop se o número do grupo for maior que 13
            if group_index // 4 + 1 > 13:
                break

            # Pega até 4 fotos do grupo atual
            group_photos = photos[group_index : group_index + 4]

            # Cria uma nova aba para este grupo
            new_sheet = self._workbook.copy_worksheet(template_sheet)
            new_sheet.title = f"Fotos {group_index // 4 + 1}"

            # Insere as imagens nas seções correspondentes
            for photo_index, _obj in enumerate(group_photos):
                if photo_index < len(image_sections):
                    section = image_sections[photo_index]
                    photo = _obj.get("photo")
                    description = _obj.get("description")
                    photo_number = int_set_zero_prefix(
                        group_index // 4 + photo_index + 1
                    )
                    photo_name = f"{prefix_photo_name}F{photo_number}:"

                    # Insere o nome da foto na seção correspondente
                    new_sheet[name_sections[photo_index]] = photo_name

                    # Insere a descrição da foto na seção correspondente
                    new_sheet[description_sections[photo_index]] = description

                    # Insere a imagem na seção especificada
                    insert_picture_2(
                        picture=Image(photo),
                        worksheet=new_sheet,
                        target=self.__sheet_target,
                        range_string=section,
                        resize_method=ResizeMethod.Stretch,
                        border_width=(2, 2, 2, 2),
                    )

            insert_picture_2(
                picture=Image(self.data_logo_company.get("path_image")),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="L61",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalCentered
                ),
                border_width=(2, 2, 2, 2),
            )
            insert_picture_2(
                picture=Image("assets/static/antt_logo.png"),
                worksheet=new_sheet,
                target=self.__sheet_target,
                range_string="Z61:AA61",
                resize_method=self.data_logo_company.get(
                    "resize_method", ResizeMethod.ProportionalRight
                ),
                border_width=(2, 2, 2, 2),
            )

            # Reordena as abas na ordem desejada
            self.sheet_counter += 1
            self._reorder_sheets(new_sheet, self.sheet_counter)

        # Remove a aba template após criar todas as abas necessárias
        self._remove_sheet("Fotos Template")

    def fill_sheet(self, data_list: list):
        data_work = data_list
        files = []
        filenames = []

        for i, data in enumerate(data_work, 1):
            self._fill_basic_fields_sheet(data)
            self._fill_technical_grade_fields_sheet(data)
            self._fill_elements_damage_fields_sheet(data)
            self._fill_inspection_report_fields_sheet(data)
            self._fill_appraisal_fields_sheet(data)
            self._fill_croqui_photos_sheet(data)
            self._fill_photos_sheet(data)

            file_name = data.get("file_name")
            filenames.append(file_name)
            result = f"/tmp/{file_name}.xlsx"
            if result in files:
                result = f"/tmp/{file_name}({i}).xlsx"

            set_zoom(self._workbook, 50, "pageBreakPreview")

            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)

            files.append(result)

        return {"files": files}

    def create_dict(self, reporting: Reporting, s3) -> dict:
        form_data_display = reporting.get_form_data_display()

        FILE_NAME = get_file_name(reporting)

        # Basic Fields
        provider_name_key = getattr(reporting.company, "name", "-")
        PROVIDER_NAME = self.provider_name_translation.get(
            provider_name_key, provider_name_key
        )
        OAE_NUMBER = form_data_display.get("oae_numero_codigo_obra") or "-"
        EXECUTED_AT = (
            reporting.executed_at.strftime("%d/%m/%Y") if reporting.executed_at else "-"
        )
        ROAD_NAME = getattr(reporting, "road_name", "-")
        INITIAL_KM = format_km(reporting, "km", 3) or "-"
        OAE = form_data_display.get("codigo_agencia") or "-"
        DENOMINATION = form_data_display.get("denominacao") or "-"
        DIRECTION = get_custom_option(reporting, "direction") or "-"
        notes_dict = self._generate_notes(form_data_display)

        # Performance Parameters
        GUARDA_CORPO_PARAM = form_data_display.get("guarda_corpo") or 0
        GUARDA_CORPO_OBSERV = (
            form_data_display.get("guarda_corpo_observacoes_quantidade") or "-"
        )
        DRENAGEM_LIMPAGEM_PARAM = form_data_display.get("drenagem_limpeza") or 0
        DRENAGEM_LIMPAGEM_OBSERV = (
            form_data_display.get("drenagem_limpeza_observacoes_quantidade") or "-"
        )
        PLACA_GABARITO_VERTICAL_PARAM = (
            form_data_display.get("placa_gabarito_vertical") or 0
        )
        PLACA_GABARITO_VERTICAL_OBSERV = (
            form_data_display.get("placa_gabarito_vertical_observacoes_quantidade")
            or "-"
        )
        JUNTAS_DILATACAO_PARAM = (
            form_data_display.get("juntas_dilatacao_vida_util_remanescente") or 0
        )
        JUNTAS_DILATACAO_OBSERV = (
            form_data_display.get(
                "juntas_dilatacao_vida_util_remanescente_observacoes_quantidade"
            )
            or "-"
        )
        APARELHO_APOIO_PARAM = (
            form_data_display.get("aparelho_apoio_vida_util_remanescente") or 0
        )
        APARELHO_APOIO_OBSERV = (
            form_data_display.get("aparelho_apoio_vida_util_remanescente_local") or "-"
        )

        # Array "Danos aos Elementos"
        THERAPY = form_data_display.get("therapy") or []
        THERAPY_GROUPED = self._group_therapy_by_section(THERAPY)
        GENERAL_TECH_GRADE = min(
            [
                int(item.get("technical_note", 0))
                for item in THERAPY
                if item.get("technical_note")
            ],
            default=0,
        )

        # Array "Laudo"
        INSPECTION_REPORT = form_data_display.get("insuficiencia_estruturais") or []
        INSPECTION_REPORT_GROUPED = self._group_therapy_by_section(INSPECTION_REPORT)

        # Fields "Laudo"
        APPRAISAL_DATE = form_data_display.get("data_do_laudo")
        # Converte string ISO 8601 para datetime no fuso horário local e formata
        if APPRAISAL_DATE:
            try:
                value = parser.parse(APPRAISAL_DATE)
                APPRAISAL_DATE_FORMATED = value.strftime("%d/%m/%Y")
            except (ValueError, AttributeError, ImportError):
                APPRAISAL_DATE_FORMATED = str(APPRAISAL_DATE)
        else:
            APPRAISAL_DATE_FORMATED = "-"
        APPRAISAL_TEXT = form_data_display.get("laudo_especializado") or "-"

        # Croqui Photos
        list_images_croqui = []

        # limit 10 images croqui
        limit_croqui = 10
        data_croqui = (
            reporting.form_data.get("croqui_image_array", [])
            if reporting.form_data
            else []
        )

        if data_croqui and isinstance(data_croqui, list):
            for vector_croqui in data_croqui:
                croquis = vector_croqui.get("croqui_image")
                if croquis and isinstance(croquis, list):
                    for file_pk in croquis:
                        if len(list_images_croqui) == limit_croqui:
                            break

                        file = (
                            ReportingFile.objects.filter(uuid=file_pk)
                            .only("upload", "is_shared")
                            .first()
                        )

                        if file and file.is_shared:
                            result = download_picture(
                                s3, self.temp_file, file.uuid, reporting_file=file
                            )
                            if result is not None:
                                list_images_croqui.append(result)

                if len(list_images_croqui) == limit_croqui:
                    break

        # Report Photos
        INSPECTION_YEAR_CAMPAIGN = (
            form_data_display.get("inspection_year_campaign") or ""
        )
        _road_name = (reporting.road_name.replace(" ", str(OAE_NUMBER)).split("-"))[1]
        PREFIX_PHOTO_NAME = f"OAE{INSPECTION_YEAR_CAMPAIGN}{_road_name}"
        reportings_imagens = (
            reporting.form_data.get("fotos_relatorio_array", [])
            if reporting.form_data
            else []
        )
        picture_uuid_list = []

        for _vector in reportings_imagens:
            vector_photos = _vector.get("fotos_relatorio")
            if vector_photos and isinstance(vector_photos, list):
                picture_uuid_list.extend(vector_photos)

        qs_files = ReportingFile.objects.filter(
            uuid__in=picture_uuid_list, is_shared=True
        ).order_by("datetime")

        photos = []
        rest_photos = []

        for file in qs_files:
            file_name = remove_ext_in_filename(file.upload.name)
            file_name = remove_random_string_file_name_in_upload(file_name)

            _result = {
                "file_name": file_name,
                "photo": download_picture(
                    s3, self.temp_file, file.uuid, reporting_file=file
                ),
                "description": file.description,
            }

            if self.valid_rules_order_photo_oae(file_name):
                photos.append(_result)
            else:
                rest_photos.append(_result)

        photos.sort(key=lambda x: x.get("file_name"))
        photos.extend(rest_photos)

        data = {
            "file_name": FILE_NAME,
            "provider_name": PROVIDER_NAME,
            "oae_number": OAE_NUMBER,
            "executed_at": EXECUTED_AT,
            "road_name": ROAD_NAME,
            "initial_km": INITIAL_KM,
            "oae": OAE,
            "denomination": DENOMINATION,
            "direction": DIRECTION,
            "notes_1": notes_dict["notes_1"],  # A18 - Identificação do inspetor
            "notes_2": notes_dict["notes_2"],  # C18 - Coordenadas geográficas
            "notes_3": notes_dict[
                "notes_3"
            ],  # J18 - Histórico de intervenções realizadas
            "general_tech_grade": GENERAL_TECH_GRADE,
            "guarda_corpo_param": GUARDA_CORPO_PARAM,
            "guarda_corpo_observ": GUARDA_CORPO_OBSERV,
            "drenagem_limpagem_param": DRENAGEM_LIMPAGEM_PARAM,
            "drenagem_limpagem_observ": DRENAGEM_LIMPAGEM_OBSERV,
            "placa_gabarito_vertical_param": PLACA_GABARITO_VERTICAL_PARAM,
            "placa_gabarito_vertical_observ": PLACA_GABARITO_VERTICAL_OBSERV,
            "juntas_dilatacao_param": JUNTAS_DILATACAO_PARAM,
            "juntas_dilatacao_observ": JUNTAS_DILATACAO_OBSERV,
            "aparelho_apoio_param": APARELHO_APOIO_PARAM,
            "aparelho_apoio_observ": APARELHO_APOIO_OBSERV,
            "therapy_grouped": THERAPY_GROUPED,
            "inspection_report_grouped": INSPECTION_REPORT_GROUPED,
            "appraisal_date": APPRAISAL_DATE_FORMATED,
            "appraisal_text": APPRAISAL_TEXT,
            "images_croqui": list_images_croqui,
            "photos": photos,
            "prefix_photo_name": PREFIX_PHOTO_NAME,
        }

        return data

    def execute(self):
        data = []
        for reporting in self.__list_reporting:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )

                if path_logo_company:
                    self.data_logo_company["path_image"] = path_logo_company

        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class XlsxHandlerMonitoringOAENewVersion(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        self.class_name = "Monitoração OAE (Em Validação)"
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""

        if len(self.uuids) == 1:
            occurrence_type_uuid = (
                Reporting.objects.filter(uuid=self.uuids[0])
                .only("occurrence_type__uuid")[0]
                .occurrence_type.uuid
            )
            reportings = Reporting.objects.filter(
                occurrence_type__uuid=occurrence_type_uuid, pk__in=self.uuids
            ).prefetch_related("road")

            reporting = reportings.first()
            extension = ""
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{get_file_name(reporting)}.{extension}"
        else:
            file_name = "Relatório ANTT - Anexo II Ficha Inspeções de OAE.zip"

        return file_name

    def __get_reportings_obj(self):
        query_set = Reporting.objects.filter(
            occurrence_type__name=self.class_name, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return list(query_set)

    def export(self):
        list_reporting = self.__get_reportings_obj()
        s3 = get_s3()
        obj = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()
        files = obj["files"]

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}.zip"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(
            s3,
            result_file,
            self.object_name,
        )

        return True


@task
def ccr_report_monitoring_oae_new_version_async_handler(reporter_dict: dict):
    reporter = XlsxHandlerMonitoringOAENewVersion.from_dict(reporter_dict)
    reporter.export()
