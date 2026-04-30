import shutil
import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Font, Side
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Firm, Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data_selected_option
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_reporting_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_custom_option,
    get_km,
    get_previous_campaign_report,
)

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
bold = Font(bold=True)


def calculate_average_zebra(value_a, value_b):
    average = ""
    try:
        average = ((float(value_a) - float(value_b)) / float(value_b)) * 100
        average = round(average, 2)
    except Exception:
        pass
    return average


def data_work(data_list: list) -> dict:
    south = [obj for obj in data_list if obj["direction"].lower() == "sul"]
    south = (
        sorted(south, key=lambda x: x.get("km_float"), reverse=False)
        if south
        else south
    )
    north = [obj for obj in data_list if obj["direction"].lower() == "norte"]
    north = (
        sorted(north, key=lambda x: x.get("km_float"), reverse=True) if north else north
    )
    others = [
        obj for obj in data_list if obj["direction"].lower() not in ["sul", "norte"]
    ]
    data_work = {}

    for obj in south + north + others:
        key = f"{obj.get('road_name')} - {obj.get('direction')} - {obj.get('monitoring_kind')}"
        if key not in data_work.keys():
            data_work[key] = [obj]
        else:
            data_work[key].append(obj)
    return data_work


class XlsxHandler:
    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.s3 = s3
        self.__list_reporting = list_reporting
        self.__sheet_target = sheet_target
        self.temp_file = tempfile.mkdtemp()
        self.__xlsx_file = "./fixtures/reports/ccr_report_zebra_horizontal_signage.xlsx"
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Zebrados"]

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="W1:X3",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B3",
            resize_method=ResizeMethod.ProportionalLeft,
        )

    def __merge_insert_border(self, row: int, col: str, value: str):
        self._worksheet.merge_cells(f"{col}{row}:{col}{row+1}")
        insert_centered_value(
            worksheet=self._worksheet, value=value, cell=f"{col}{row}", wrapText=True
        )
        self._worksheet[f"{col}{row}"].border = border
        self._worksheet[f"{col}{row+1}"].border = border

    def __insert_border(self, row: int, col: str, value: str):
        insert_centered_value(
            worksheet=self._worksheet, value=value, cell=f"{col}{row}", wrapText=True
        )
        self._worksheet[f"{col}{row}"].border = border

    def __insert_bold(self, row: int, col: str, value: str):
        insert_centered_value(
            worksheet=self._worksheet,
            value=value,
            cell=f"{col}{row}",
            horizontal="left",
            wrapText=False,
        )
        self._worksheet[f"{col}{row}"].font = bold

    def fill_sheet(self, data_list: list):
        datadict = data_work(data_list=data_list)
        files = []
        filenames = []
        all_roads = []
        for sheetname, datalist in datadict.items():
            row = 8
            lanes = []
            subcompanies = []
            team_list = []
            roads = []
            dates = []
            road_name = ""
            direction = ""

            company = ""
            for data in datalist:
                average_lca = data.get("average_lca")
                average_zpa = data.get("average_zpa")
                lifespan_lca = data.get("lifespan_lca")
                lifespan_zpa = data.get("lifespan_zpa")
                self.__merge_insert_border(row=row, col="A", value=data.get("km"))
                self.__merge_insert_border(
                    row=row, col="B", value=data.get("direction")
                )
                self.__merge_insert_border(row=row, col="C", value=data.get("found_at"))
                self.__merge_insert_border(row=row, col="D", value=data.get("lat"))
                self.__merge_insert_border(row=row, col="E", value=data.get("long"))
                self.__merge_insert_border(row=row, col="F", value="ZEBRADOS")
                self.__insert_border(row=row, col="G", value="LINHA DE CANALIZAÇÃO")
                self.__insert_border(row=row + 1, col="G", value="PREENCHIMENTO")
                self.__insert_border(row=row, col="H", value=data.get("color"))
                self.__insert_border(row=row + 1, col="H", value=data.get("color"))
                self.__insert_border(row=row, col="I", value="")
                self.__insert_border(row=row + 1, col="I", value="")
                self._worksheet.row_dimensions[row].height = 69.75
                self._worksheet.row_dimensions[row + 1].height = 69.75
                if data.get("img"):
                    range_str = f"I{row}"
                    insert_picture(
                        self._worksheet,
                        range_str,
                        Image(data.get("img")),
                        self.__sheet_target,
                    )
                    range_str = f"I{row+1}"
                    insert_picture(
                        self._worksheet,
                        range_str,
                        Image(data.get("img")),
                        self.__sheet_target,
                    )
                self.__insert_border(row=row, col="J", value=data.get("first_lca"))
                self.__insert_border(row=row + 1, col="J", value=data.get("first_zpa"))
                self.__insert_border(row=row, col="K", value=data.get("second_lca"))
                self.__insert_border(row=row + 1, col="K", value=data.get("second_zpa"))
                self.__insert_border(row=row, col="L", value=data.get("third_lca"))
                self.__insert_border(row=row + 1, col="L", value=data.get("third_zpa"))
                self.__insert_border(row=row, col="M", value=data.get("fourth_lca"))
                self.__insert_border(row=row + 1, col="M", value=data.get("fourth_zpa"))
                self.__insert_border(row=row, col="N", value=data.get("fifth_lca"))
                self.__insert_border(row=row + 1, col="N", value=data.get("fifth_zpa"))
                self.__insert_border(row=row, col="O", value=data.get("sixth_lca"))
                self.__insert_border(row=row + 1, col="O", value=data.get("sixth_zpa"))
                self.__insert_border(row=row, col="P", value=data.get("seventh_lca"))
                self.__insert_border(
                    row=row + 1, col="P", value=data.get("seventh_zpa")
                )
                self.__insert_border(row=row, col="Q", value=data.get("eighth_lca"))
                self.__insert_border(row=row + 1, col="Q", value=data.get("eighth_zpa"))
                self.__insert_border(row=row, col="R", value=data.get("ninth_lca"))
                self.__insert_border(row=row + 1, col="R", value=data.get("ninth_zpa"))
                self.__insert_border(row=row, col="S", value=data.get("tenth_lca"))
                self.__insert_border(row=row + 1, col="S", value=data.get("tenth_zpa"))
                self.__insert_border(row=row, col="T", value=average_lca)
                self.__insert_border(row=row + 1, col="T", value=average_zpa)
                self.__insert_border(
                    row=row, col="U", value=data.get("residual_value_lca")
                )
                self.__insert_border(
                    row=row + 1, col="U", value=data.get("residual_value_zpa")
                )
                self.__insert_border(row=row, col="V", value=lifespan_lca)
                self.__insert_border(row=row + 1, col="V", value=lifespan_zpa)
                self.__insert_border(
                    row=row, col="W", value=data.get("average_lca_last_year")
                )
                self.__insert_border(
                    row=row + 1, col="W", value=data.get("average_zpa_last_year")
                )
                self.__merge_insert_border(row=row, col="X", value=data.get("notes"))
                if (average_lca or average_lca == 0) and (
                    lifespan_lca or lifespan_lca == 0
                ):
                    color = "000000"
                    try:
                        if float(average_lca) < float(lifespan_lca):
                            color = "FF0000"
                    except Exception:
                        pass
                    self._worksheet[f"T{row}"].font = Font(color=color)
                    self._worksheet[f"U{row}"].font = Font(color=color)
                if (average_zpa or average_zpa == 0) and (
                    lifespan_zpa or lifespan_zpa == 0
                ):
                    color = "000000"
                    try:
                        if float(average_zpa) < float(lifespan_zpa):
                            color = "FF0000"
                    except Exception:
                        pass
                    self._worksheet[f"T{row+1}"].font = Font(color=color)
                    self._worksheet[f"U{row+1}"].font = Font(color=color)
                if data.get("lane") not in lanes:
                    lanes.append(data.get("lane"))
                if data.get("subcompany") not in subcompanies:
                    subcompanies.append(data.get("subcompany"))
                if data.get("team") not in team_list:
                    team_list.append(data["team"])
                if data.get("road") not in roads:
                    roads.append(data.get("road"))
                if data.get("road") not in all_roads:
                    all_roads.append(data.get("road"))
                if not road_name:
                    road_name = data.get("road_name")
                if not direction:
                    direction = data.get("direction")
                if not company:
                    company = data.get("company")
                if data["executed_at"] not in dates:
                    dates.append(data["executed_at"])
                if not data["executed_at"] and data["found_at"] not in dates:
                    dates.append(data["found_at"])
                row += 2

            dates[:] = [item for item in dates if item != ""]
            if len(dates) > 1:
                dates_new = sorted(
                    dates,
                    key=lambda x: (x.split("/")[2], x.split("/")[1], x.split("/")[0]),
                )
                date_text = f"Data: {dates_new[0]} à {dates_new[-1]}"
            elif dates:
                date_text = f"Data: {dates[0]}"
            else:
                date_text = "Data:"
            team_list = list(set(team_list))
            query_set_teams = Firm.objects.filter(uuid__in=team_list).all()
            names = []
            for team in query_set_teams:
                users_query_set = team.users.all()
                intern_list = []
                for user in users_query_set:
                    intern_list.append(user.get_full_name())
                names.append(intern_list)
            names = [", ".join(_) for _ in names]
            team_text = f"{' - '.join(names)}"
            self.__insert_bold(
                row=1, col="E", value=f"Empresa: {'/'.join(subcompanies)}"
            )
            self.__insert_bold(row=2, col="E", value=f"Operador: {team_text}")
            self.__insert_bold(row=3, col="E", value=date_text)
            self.__insert_bold(row=1, col="K", value=f"Rodovia: {'-'.join(roads)}")
            self.__insert_bold(
                row=2,
                col="K",
                value=f"Monitoração: Zebrados - {direction} - {' - '.join(lanes)}",
            )
            self.__insert_bold(row=3, col="K", value=f"Concessionária: {company}")
            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                target=self.__sheet_target,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
            )

            file_name = (
                f"Fichas de Zebrados {road_name} {' - '.join(lanes)} {direction}"
            )
            file_name = file_name.replace("/", "")
            filenames.append(file_name)
            result = f"/tmp/{file_name}.xlsx"
            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook["Zebrados"]
            files.append(result)
        all_roads.sort()
        return {"files": files, "names": filenames, "all_roads": all_roads}

    def create_dict(self, reporting: Reporting, s3) -> dict:
        km = get_km(reporting)
        direction = get_custom_option(reporting, "direction")
        form_data: dict = reporting.form_data

        average_lca = form_data.get("average_lca")
        lifespan_lca = form_data.get("lifespan_lca")
        average_zpa = form_data.get("average_zpa")
        lifespan_zpa = form_data.get("lifespan_zpa")
        residual_value_lca = (
            calculate_average_zebra(value_a=average_lca, value_b=lifespan_lca)
            if average_lca and lifespan_lca
            else ""
        )
        residual_value_zpa = (
            calculate_average_zebra(value_a=average_zpa, value_b=lifespan_zpa)
            if average_zpa and lifespan_zpa
            else ""
        )

        average_lca_last_year = ""
        average_zpa_last_year = ""
        if reporting.parent:
            obj = get_previous_campaign_report(
                reporting.occurrence_type, reporting, "form_data"
            )
            if obj is not None and obj.form_data:
                form_data_ly: dict = obj.form_data
                if form_data_ly:
                    average_lca_last_year = form_data_ly.get("average_lca")
                    average_zpa_last_year = form_data_ly.get("average_zpa")

        reporting_files: dict = download_reporting_pictures(
            s3=s3,
            path=self.temp_file,
            reporting=reporting,
            width=337,
            height=242,
            enable_include_dnit=False,
            enable_is_shared_antt=True,
        )
        img = ""
        if reporting_files.get("images"):
            images = reporting_files.get("images")
            img = images[0]["path"]

        color = new_get_form_data_selected_option(
            reporting,
            "color",
            form_data.get("color"),
        )

        data = {
            "km": km,
            "km_float": reporting.km,
            "road_name": reporting.road_name,
            "road": reporting.road.name,
            "monitoring_kind": form_data.get("monitoringKind")
            or form_data.get("monitoring_kind")
            or form_data.get("monitoringkind"),
            "direction": direction,
            "found_at": (
                reporting.found_at.strftime("%d/%m/%Y") if reporting.found_at else ""
            ),
            "executed_at": (
                reporting.executed_at.strftime("%d/%m/%Y")
                if reporting.executed_at
                else ""
            ),
            "lat": form_data.get("latitudedecimal") or form_data.get("lat"),
            "long": form_data.get("longitudedecimal") or form_data.get("long"),
            "average_lca_last_year": average_lca_last_year,
            "average_zpa_last_year": average_zpa_last_year,
            "average_lca": average_lca,
            "average_zpa": average_zpa,
            "lifespan_lca": lifespan_lca,
            "lifespan_zpa": lifespan_zpa,
            "residual_value_lca": residual_value_lca,
            "residual_value_zpa": residual_value_zpa,
            "color": color,
            "first_lca": form_data.get("first_lca"),
            "first_zpa": form_data.get("first_zpa"),
            "second_lca": form_data.get("second_lca"),
            "second_zpa": form_data.get("second_zpa"),
            "third_lca": form_data.get("third_lca"),
            "third_zpa": form_data.get("third_zpa"),
            "fourth_lca": form_data.get("fourth_lca"),
            "fourth_zpa": form_data.get("fourth_zpa"),
            "fifth_lca": form_data.get("fifth_lca"),
            "fifth_zpa": form_data.get("fifth_zpa"),
            "sixth_lca": form_data.get("sixth_lca"),
            "sixth_zpa": form_data.get("sixth_zpa"),
            "seventh_lca": form_data.get("seventh_lca"),
            "seventh_zpa": form_data.get("seventh_zpa"),
            "eighth_lca": form_data.get("eighth_lca"),
            "eighth_zpa": form_data.get("eighth_zpa"),
            "ninth_lca": form_data.get("ninth_lca"),
            "ninth_zpa": form_data.get("ninth_zpa"),
            "tenth_lca": form_data.get("tenth_lca"),
            "tenth_zpa": form_data.get("tenth_zpa"),
            "notes": form_data.get("notes"),
            "lane": get_custom_option(reporting, "lane"),
            "img": img,
            "subcompany": reporting.firm.subcompany.__dict__.get("name"),
            "team": str(reporting.firm.__dict__.get("uuid")),
            "company": reporting.company.__dict__.get("name"),
            "reporting": reporting,
        }
        for k, v in data.items():
            if v is None and k not in ["station_one", "station_two", "station_three"]:
                data[k] = ""
        return data

    def execute(self):
        data = []
        for reporting in self.__list_reporting:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRReportZebraHorizontalSignage(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""

        reporting = Reporting.objects.get(uuid=self.uuids[0])

        reportings = Reporting.objects.filter(
            occurrence_type=reporting.occurrence_type, uuid__in=self.uuids
        ).prefetch_related("road")
        all_roads = []
        datawork = {}
        for obj in reportings:
            if obj.road.name not in all_roads:
                all_roads.append(obj.road.name)
            key = f"{obj.road_name} - {obj.direction} - {obj.form_data.get('monitoring_kind')}"
            if key not in datawork.keys():
                datawork[key] = [obj]
            else:
                datawork[key].append(obj)
        if len(datawork.keys()) == 1:
            direction = get_custom_option(reporting, "direction")
            lane = get_custom_option(reporting, "lane")
            road_name = reporting.road_name
            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"
            file_name = f"Fichas de Zebrados {road_name} {lane if lane else ''} {direction}.{extension}"
        else:

            all_roads.sort()
            file_name = f"Monitoração de Zebrados - Rodovia {'-'.join(all_roads)}.zip"
            self.file_name = file_name

        return file_name

    def __get_reportings_obj(self):
        form = OccurrenceType.objects.get(
            name="Retrorrefletância Horizontal de Zebrado"
        )
        query_set = Reporting.objects.filter(
            occurrence_type=form, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def export(self):
        list_reporting = self.__get_reportings_obj()
        s3 = get_s3()
        obj = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()
        files = obj["files"]
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_zebra_horizontal_signage_async_handler(reporter_dict: dict):
    reporter = CCRReportZebraHorizontalSignage.from_dict(reporter_dict)
    reporter.export()
