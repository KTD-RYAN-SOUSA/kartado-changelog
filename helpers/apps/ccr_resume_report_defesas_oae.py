import tempfile
from typing import List

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from apps.companies.models import Firm
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.export_utils import insert_centered_value
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
)
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km


class XlsxHandlerResumeReportDefensesOAE(object):
    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.wb = load_workbook(
            "./fixtures/reports/ccr_resume_report_defenses_oae.xlsx"
        )
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self._worksheet = self.wb.active
        self._worksheet.title = "Cad. Resumo de Defensas Em OAE"
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company"
        )
        self.temp_file = tempfile.mkdtemp()
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="U1:W5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:A5",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.resume_report_dict = []
        self.static_fields = {
            "index": "A",
            "road": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "side": "F",
            "length": "G",
            "height": "H",
            "metal_post": "I",
            "cantilevered": "J",
            "aerial_terminal": "K",
            "transition_triple_wave": "L",
            "fuel_terminal": "M",
            "areial_boolean_terminal": "N",
            "anchor_talude": "O",
            "defense_deflected": "P",
            "energy_terminal": "Q",
            "reflective_presence": "R",
            "delimitors_presence": "S",
            "corrosion": "T",
            "alignment": "U",
            "post": "V",
            "screw_nods": "W",
        }

    def __insert_new_rows(self, row: int):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        chars = [chr(ord("A") + i) for i in range(23)]

        for char in chars:
            self._worksheet[f"{char}{row}"].border = border
            self._worksheet.row_dimensions[row].height = 25

    def create_dict(self, reporting):
        result_index = new_get_form_data(reporting, "idCcrAntt")

        direction = get_custom_option(reporting, "direction")

        side = new_get_form_data(reporting, "side")

        height_value = new_get_form_data(reporting, "height")

        corrosion = "X" if reporting.form_data.get("corrosion") else "-"
        alignment = "X" if reporting.form_data.get("alignment") else "-"
        post = "X" if reporting.form_data.get("post") else "-"

        result_metal_post = new_get_form_data(reporting, "metalPost")
        metal_post = "Sim" if result_metal_post else "Não"

        result_cantilevered = new_get_form_data(reporting, "cantilevered")

        cantilevered = "X" if result_cantilevered else "-"

        result_aerial_terminal = new_get_form_data(reporting, "aerialTerminal")
        aerial_terminal = "X" if result_aerial_terminal else "-"

        result_transition_triple_wave = new_get_form_data(
            reporting, "transitionTripleWave"
        )

        transition_triple_wave = "X" if result_transition_triple_wave else "-"

        result_fuel_terminal = new_get_form_data(reporting, "fuelTerminal")

        fuel_terminal = "X" if result_fuel_terminal else "-"

        result_areial_boolean_terminal = new_get_form_data(
            reporting,
            "areialBooleanTerminal",
        )

        areial_boolean_terminal = "X" if result_areial_boolean_terminal else "-"

        result_anchor_talude = new_get_form_data(reporting, "anchorTalude")

        anchor_talude = "X" if result_anchor_talude else "-"

        result_defense_deflected = new_get_form_data(reporting, "defenseDeflected")

        defense_deflected = "X" if result_defense_deflected else "-"

        result_energy_terminal = new_get_form_data(reporting, "energyTerminal")

        energy_terminal = "X" if result_energy_terminal else "-"

        result_reflective_presence = new_get_form_data(reporting, "reflectivePresence")

        reflective_presence = "Sim" if result_reflective_presence else "Não"
        delimitors_presence = "Sim" if result_reflective_presence else "Não"

        result_screw_nods = new_get_form_data(reporting, "screwNods")

        screw_nods = "X" if result_screw_nods else "-"

        legth_value = reporting.form_data.get("length")

        data = {
            "index": result_index,
            "road": reporting.__dict__.get("road_name"),
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "end_km", 3),
            "raw_initial_km": reporting.km,
            "direction": direction.upper(),
            "side": side if side else "-",
            "height": height_value if height_value else "-",
            "length": legth_value if legth_value else "-",
            "metal_post": metal_post,
            "cantilevered": cantilevered,
            "aerial_terminal": aerial_terminal,
            "transition_triple_wave": transition_triple_wave,
            "fuel_terminal": fuel_terminal,
            "areial_boolean_terminal": areial_boolean_terminal,
            "anchor_talude": anchor_talude,
            "defense_deflected": defense_deflected,
            "energy_terminal": energy_terminal,
            "reflective_presence": reflective_presence,
            "delimitors_presence": delimitors_presence,
            "corrosion": corrosion,
            "alignment": alignment,
            "post": post,
            "screw_nods": screw_nods,
            "executed_at": reporting.executed_at,
            "team": str(reporting.firm.__dict__.get("uuid")) if reporting.firm else "",
            "road_name": reporting.__dict__.get("road_name"),
            "subcompany": (
                reporting.firm.subcompany.__dict__.get("name") if reporting.firm else ""
            ),
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""

        return data

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def fill_sheet(self, *, data_list: list):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        row = 8
        team_list = []
        date_list = []
        road_name_list = []
        subcompany_list = []

        list_mapping = {
            "executed_at": date_list,
            "road_name": road_name_list,
            "subcompany": subcompany_list,
            "team": team_list,
        }

        for values in data_list:
            for key, value in values.items():
                self.__insert_new_rows(row=row)
                if key in list_mapping:
                    list_mapping[key].append(value)
                else:
                    if key == "raw_initial_km":
                        continue
                    key_value = f"{self.static_fields[key]}{row}"
                    self._worksheet[key_value] = value
                    XlsxHandlerResumeReportDefensesOAE.__format_fonts(
                        cell=self._worksheet[key_value], size=11
                    )

            row += 1

        filtered_team_list = set([uuid for uuid in list_mapping["team"] if uuid != ""])
        query_set_teams = Firm.objects.filter(uuid__in=filtered_team_list).all()
        names = []
        for team in query_set_teams:
            users_query_set = team.users.all()
            intern_list = []
            for user in users_query_set:
                intern_list.append(user.full_name)
            names.append(intern_list)
        names = [", ".join(_) for _ in names]
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(names),
            cell="F3",
            horizontal="left",
        )

        filtered_date_list = [
            date for date in list_mapping["executed_at"] if date != ""
        ]
        filtered_date_list.sort()
        if len(filtered_date_list) == 0:
            date_text = ""
        elif len(filtered_date_list) == 1:
            date_text = filtered_date_list[0].strftime("%d/%m/%Y")
        else:
            date_text = (
                f"{filtered_date_list[0].strftime('%d/%m/%Y')} "
                f"até {filtered_date_list[-1].strftime('%d/%m/%Y')}"
            )

        insert_centered_value(
            worksheet=self._worksheet, value=date_text, cell="F4", horizontal="left"
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value=self.company.name,
            cell="F2",
            horizontal="left",
        )

        subcompany = set(subcompany_list)
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(subcompany),
            cell="O4",
            horizontal="left",
        )

        road_name = set(list_mapping["road_name"])
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(road_name),
            cell="O2",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value="Elementos de Proteção e Segurança - Cadastro Resumo de Defesas em OAE",
            cell="O3",
            horizontal="left",
        )

        total_length = sum(
            [t.value for t in self._worksheet["G"] if isinstance(t.value, (int, float))]
        )
        reference_total_legth = self._worksheet[f"{'G'}{self._worksheet.max_row + 1}"]
        reference_total_legth.value = total_length
        reference_total_legth.border = border
        XlsxHandlerResumeReportDefensesOAE.__format_fonts(
            cell=reference_total_legth, size=11, bold=True
        )

        reference_total_string = reference_total_legth.offset(0, -1)
        reference_total_string.value = "TOTAL(m)"
        reference_total_string.border = border
        XlsxHandlerResumeReportDefensesOAE.__format_fonts(
            cell=reference_total_string, size=11, bold=True
        )

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            target=self.__sheet_target,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        sorted_data = sorted(
            data,
            key=lambda x: (
                x["raw_initial_km"] if x["direction"] == "NORTE" else False,
                -x["raw_initial_km"] if x["direction"] == "SUL" else False,
                x["direction"] == "CANTEIRO CENTRAL",
                x["direction"] == "AMBOS",
                x["direction"] == "CRESCENTE",
                x["direction"] == "DECRESCENTE",
                x["direction"] == "LESTE",
                x["direction"] == "LESTE/OESTE",
                x["direction"] == "NORTE/SUL",
                x["direction"] == "TRANSVERSAL",
                x["direction"] == "OESTE",
                x["direction"],
            ),
            reverse=True,
        )

        self.fill_sheet(data_list=sorted_data)

        road_name = self.reportings.first().road_name
        file_name = "Cadastro Resumo - {} - Defensas em OAE".format(road_name)

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        result = f"/tmp/{file_name}.xlsx"
        self.wb.save(result)
        return result
