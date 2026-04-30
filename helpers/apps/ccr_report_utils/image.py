import imghdr
import os
import urllib.parse
from enum import IntEnum
from os.path import isfile
from typing import Iterable, List, Tuple, Union
from urllib.parse import unquote
from uuid import UUID, uuid4

from django.db.models import Q
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
    TwoCellAnchor,
)
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.units import (
    DEFAULT_COLUMN_WIDTH,
    DEFAULT_ROW_HEIGHT,
    pixels_to_EMU,
    points_to_pixels,
)
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as ImagePil
from PIL import ImageOps

from apps.companies.models import Firm
from apps.reportings.models import Reporting, ReportingFile
from helpers.strings import COMMON_IMAGE_TYPE, check_image_file


class SheetTarget(IntEnum):
    DesktopExcel = 0
    LibreOffice = 1
    GotenbergPrinter = 2


class ReportFormat(IntEnum):
    XLSX = 0
    PDF = 1


def _get_query_upload_valid_img():
    q = Q()
    for extension_file in COMMON_IMAGE_TYPE:
        q |= Q(upload__icontains=extension_file)
    return q


def download_picture(
    s3,
    dir: str,
    image_name: str,
    picture_uuid: str = None,
    reporting_file: ReportingFile = None,
    enable_is_shared: bool = False,
    quality: str = None,
) -> str:
    """
    Downloads a picture and returns the image local absolute path

    :param s3: boto s3 client instance
    :param dir: directory to download picture
    :param image_name: picture local name
    :param picture_uuid: picture ReportingFile UUID
    :param quality: possible values: None (original size), 400px, 1000px
    :returns: image local absolute path
    """
    try:
        upload: str = ""
        if picture_uuid is not None:
            if enable_is_shared:
                rf_queryset = ReportingFile.objects.filter(
                    uuid=picture_uuid, is_shared=True
                )
            else:
                rf_queryset = ReportingFile.objects.filter(uuid=picture_uuid)
            upload = rf_queryset.filter(_get_query_upload_valid_img())[0].upload

        else:
            if enable_is_shared:
                if reporting_file.is_shared:
                    upload = reporting_file.upload
            else:
                upload = reporting_file.upload
        file_path = upload.url.split("?")[0].split(".com/")[1]
        bucket_name = upload.url.split(".s3")[0].split("/")[-1]
        image_format = file_path.split(".")[-1]
    except Exception:
        return None

    if not check_image_file(file_path):
        return None

    image_path = "{}/{}.{}".format(dir, image_name, image_format)
    i = 1
    while isfile(image_path):
        image_path = "{}/{}({}).{}".format(dir, image_name, i, image_format)
        i += 1
    temp_file = f"{image_path}"
    downloaded = False
    if quality is not None:
        try:
            s3.download_file(bucket_name + "-" + quality, unquote(file_path), temp_file)
            downloaded = True
        except Exception:
            pass

    if not downloaded:
        try:
            s3.download_file(bucket_name, unquote(file_path), temp_file)
            downloaded = True
        except Exception:
            pass

    converted_image_path = None
    if downloaded:
        try:
            converted_image_path = save_image_as_png(temp_file)
            if converted_image_path != temp_file and isfile(temp_file):
                os.remove(temp_file)
        except Exception:
            pass
    return converted_image_path


class ResizeMethod(IntEnum):
    NoResize = 0
    Stretch = 1
    ProportionalLeft = 2
    ProportionalCentered = 3
    ProportionalRight = 4
    ProportionalLeftMiddle = 5


def get_range_dimensions(
    worksheet: Worksheet, range_string: str, target: SheetTarget, only_sum: bool = True
) -> Union[Tuple[float, float], Tuple[List[float], List[float]]]:
    """
    Returns width and height of a cell range given a worksheet and a cell range

    :param worksheet: Worksheet to insert the picture
    :param range_string: image coverage cell range. (As in "A1:B2" or just "A1")
    :returns: (width, height) tuple in pixels
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    if only_sum:
        range_width = 0
        range_height = 0
    else:
        widths: List[float] = []
        heights: List[float] = []

    curr_col = min_col + 1
    curr_max = None
    while (
        curr_max is None and curr_col > 0
    ):  # Probe back to find first column width and first min max range
        curr_col -= 1
        curr_max = worksheet.column_dimensions[get_column_letter(curr_col)].max
        curr_width = worksheet.column_dimensions[get_column_letter(curr_col)].width

    if curr_width is None:
        """
        Would happen in a corrupted workbook, autofit cells, or due to programmer intervention
        """
        curr_col = 1
        curr_width = DEFAULT_COLUMN_WIDTH

    for col in range(curr_col, min_col):
        """
        [1] This is necessary due to what i think is poor read-write segregation in openpyxl
            Reading a column dimension with null width might cause the column width
            to be overwritten by DEFAULT_COLUMN_WIDTH
            It makes necessary to set the width value to the expected column width
        """
        worksheet.column_dimensions[get_column_letter(col)].width = curr_width

    curr_col = min_col
    while curr_col <= max_col:  # Probe ahead collecting the widths until end of range
        curr_max = None
        while (
            curr_col <= max_col and curr_max is None
        ):  # Probe ahead in the current min max range
            worksheet.column_dimensions[
                get_column_letter(curr_col)
            ].width = curr_width  # [1]
            if only_sum:
                range_width += curr_width
            else:
                wt = curr_width * 7.0
                if target == SheetTarget.LibreOffice:
                    wt = wt * 96.0 / 72.0  # Fudge factor
                elif target == SheetTarget.GotenbergPrinter:
                    wt = wt * 96.0 / 91.0  # Fudge factor
                widths.append(wt)
            curr_col += 1
            curr_max = worksheet.column_dimensions[get_column_letter(curr_col)].max
        if curr_max is None:
            worksheet.column_dimensions[get_column_letter(curr_col)].width = curr_width
        curr_width = worksheet.column_dimensions[
            get_column_letter(curr_col)
        ].width  # Next min max range width

    for row in range(min_row, max_row + 1):
        ht = worksheet.row_dimensions[row].ht
        if ht is None:
            ht = DEFAULT_ROW_HEIGHT
        if worksheet.row_dimensions[row].customHeight:
            ht -= 0.7109375
        if only_sum:
            range_height += ht
        else:
            heights.append(points_to_pixels(ht))

    if only_sum:
        range_width = range_width * 7.0
        range_height = points_to_pixels(range_height)
        if target == SheetTarget.LibreOffice:
            range_width = range_width * 96.0 / 72.0  # Fudge factor
        elif target == SheetTarget.GotenbergPrinter:
            range_width = range_width * 96.0 / 91.0  # Fudge factor
        return (range_width, range_height)
    else:
        return (widths, heights)


def offset_correction(
    sizes: Iterable[float], offset: float, increment: int
) -> Tuple[int, float]:
    """
    Gets offset anc col/row increments to help in proportional image placement.
    Used to set openpyxl Anchors without relying on big offset values (that are not
    supported by Excel).

    :param sizes: sizes of rows/cols to iterate and inc/decrement
    :param offset: the total offset in pixels
    :param increment: 1 for left and top and -1 for right and bottom
    :returns: tuple of (cells to skip, remaining offset)
    """
    cell_inc = 0
    if increment < 0:
        sizes = reversed(sizes)
    for size in sizes:
        sub_offset = offset - size
        if sub_offset > 0:
            offset = sub_offset
            cell_inc += increment
        else:
            break

    return (cell_inc, offset)


def insert_picture_2(
    worksheet: Worksheet,
    range_string: str,
    picture: Image,
    target: SheetTarget = SheetTarget.DesktopExcel,
    border_width: Tuple[int, int, int, int] = (0, 0, 0, 0),
    resize_method: ResizeMethod = ResizeMethod.Stretch,
) -> None:
    """
    Inserts image in a cell range keeping proportions.
    This has a faster ProportionalCentered/Left/Right resizing
    (Does not rely on creating a new image)

    :param worksheet: Worksheet to insert the picture
    :param range_string: image coverage cell range. (As in "A1:B2" or just "A1")
    :param picture: picture to insert
    :param target: worksheet tool used to open worksheet
    :param border_width: tuple (top, left, bottom, right) in pixels
    to consider as border/margin of the range
    :param resize_method: specifies how image is resized
    :returns: None
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    top_off, left_off, bottom_off, right_off = border_width

    if resize_method != ResizeMethod.Stretch:
        range_widths, range_heights = get_range_dimensions(
            worksheet, range_string, target, only_sum=False
        )
        range_height = sum(range_heights) - (top_off + bottom_off)
        range_width = sum(range_widths) - (left_off + right_off)

        width_ratio = range_width / picture.width
        height_ratio = range_height / picture.height
        ratio = min(width_ratio, height_ratio)

        picture_width = picture.width * ratio
        picture_height = picture.height * ratio

        if width_ratio < height_ratio:
            offset = (range_height - picture_height) / 2
            min_row_inc, top_inc = offset_correction(range_heights, offset, 1)
            max_row_inc, bottom_inc = offset_correction(range_heights, offset, -1)
            min_row += min_row_inc
            max_row += max_row_inc
            top_off += top_inc
            bottom_off += bottom_inc
        else:
            if resize_method in [
                ResizeMethod.ProportionalLeft,
                ResizeMethod.ProportionalLeftMiddle,
            ]:
                offset = range_width - picture_width
                max_col_inc, right_inc = offset_correction(range_widths, offset, -1)
                max_col += max_col_inc
                right_off += right_inc
            elif resize_method == ResizeMethod.ProportionalCentered:
                offset = (range_width - picture_width) / 2
                min_col_inc, left_inc = offset_correction(range_widths, offset, 1)
                max_col_inc, right_inc = offset_correction(range_widths, offset, -1)
                min_col += min_col_inc
                max_col += max_col_inc
                left_off += left_inc
                right_off += right_inc
            elif resize_method == ResizeMethod.ProportionalRight:
                offset = range_width - picture_width
                min_col_inc, left_inc = offset_correction(range_widths, offset, 1)
                min_col += min_col_inc
                left_off += left_inc

    _from = AnchorMarker(
        col=min_col - 1,
        row=min_row - 1,
        colOff=pixels_to_EMU(left_off),
        rowOff=pixels_to_EMU(top_off),
    )
    _to = AnchorMarker(
        col=max_col,
        row=max_row,
        colOff=-pixels_to_EMU(right_off),
        rowOff=-pixels_to_EMU(bottom_off),
    )

    picture.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=_to)
    worksheet.add_image(picture)


def insert_picture(
    worksheet: Worksheet,
    range_string: str,
    picture: Image,
    target: SheetTarget = SheetTarget.DesktopExcel,
    width: int = 0,
    height: int = 0,
    border_width: int = 0,
    resize_method: ResizeMethod = ResizeMethod.Stretch,
) -> None:
    """
    Inserts image in a cell range
    Faster ProportionalCentered/Left/Right resizing are welcome

    :param worksheet: Worksheet to insert the picture
    :param range_string: image coverage cell range. (As in "A1:B2" or just "A1")
    :param picture: picture to insert
    :param target: worksheet tool used to open worksheet
    :param width: image height (when not 0, won't auto size to range_string area)
    :param height: worksheet tool used to open worksheet (when not 0, won't auto size to range_string area)
    :param border_width: decrement in pixels of bottom and right sides
    :param resize_method: specifies how image is resized
    :returns: None
    """

    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    if resize_method == ResizeMethod.Stretch:
        _from = AnchorMarker(
            col=min_col - 1,
            row=min_row - 1,
            colOff=pixels_to_EMU(border_width),
            rowOff=pixels_to_EMU(border_width),
        )
        _to = AnchorMarker(
            col=max_col,
            row=max_row,
            colOff=-pixels_to_EMU(border_width),
            rowOff=-pixels_to_EMU(border_width),
        )
        picture.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=_to)
        worksheet.add_image(picture)
        return
    if resize_method != ResizeMethod.NoResize:
        range_width, range_height = get_range_dimensions(
            worksheet, range_string, target
        )

        # ProportionalLeft/Centered/Right
        ratio = min(
            range_width / picture.width,
            range_height / picture.height,
        )
        picture_width = picture.width * ratio
        picture_height = picture.height * ratio
        picture.width = picture_width - (border_width * ratio)
        picture.height = picture_height - (border_width * ratio)

        if resize_method != ResizeMethod.ProportionalLeft:  # Super slower
            frame_width = range_width - border_width * ratio
            frame_height = points_to_pixels(range_height) - border_width * ratio

            if resize_method != ResizeMethod.ProportionalLeftMiddle:
                horizontal_padding = int(
                    abs(picture.width - range_width)
                )  # padding to align right
            else:
                horizontal_padding = 0
            vertical_padding = int(abs(picture.height - frame_height) / 2)
            if (
                resize_method == ResizeMethod.ProportionalCentered
            ):  # to align center, halves the left padding
                horizontal_padding = int(horizontal_padding / 2)

            pil_image = ImagePil.open(picture.ref).convert("RGBA")
            pil_image = pil_image.resize((int(picture.width), int(picture.height)))
            new_picture = ImagePil.new(
                "RGBA",
                (int(frame_width), int(frame_height)),
                (255, 255, 255, 0),
            )
            new_picture.paste(pil_image, (horizontal_padding, vertical_padding))
            new_picture.save(f"{picture.ref}.png")
            picture = Image(f"{picture.ref}.png")
        # Create the anchor marker
    anchor_maker = AnchorMarker(
        col=min_col - 1,
        row=min_row - 1,
        colOff=pixels_to_EMU(border_width),
        rowOff=pixels_to_EMU(border_width),
    )
    size = XDRPositiveSize2D(
        pixels_to_EMU(picture.width), pixels_to_EMU(picture.height)
    )
    anchor = OneCellAnchor(_from=anchor_maker, ext=size)

    worksheet.add_image(img=picture, anchor=anchor)


def download_files_pictures(
    s3,
    path: str,
    files: Iterable[ReportingFile],
    width: int,
    height: int,
    low_quality=False,
    limit=None,
):
    images = []
    count = 0
    for index, file_obj in enumerate(files):
        if file_obj.upload:
            try:
                file_path = file_obj.upload.url.split("?")[0].split(".com/")[1]
                bucket_name = file_obj.upload.url.split(".s3")[0].split("/")[-1]
                image_format = file_path.split(".")[-1]
                image_name = str(index)
            except Exception:
                continue

            image_path = "{}{}.{}".format(
                path, image_name + str(file_obj.uuid), image_format
            )

            if low_quality:
                try:
                    uncoded_file_name = urllib.parse.unquote(file_path)
                    s3.download_file(
                        bucket_name + "-400px", uncoded_file_name, image_path
                    )
                except Exception:
                    try:
                        uncoded_file_name = urllib.parse.unquote(file_path)
                        s3.download_file(bucket_name, uncoded_file_name, image_path)
                    except Exception:
                        continue
            else:
                try:
                    uncoded_file_name = urllib.parse.unquote(file_path)
                    s3.download_file(bucket_name, uncoded_file_name, image_path)
                except Exception:
                    continue
            try:
                image_path = save_image_as_png(image_path)
                images.append(
                    {
                        "path": image_path,
                        "description": file_obj.description,
                        "uuid": str(file_obj.uuid),
                    }
                )
                count += 1
                if limit and count >= limit:
                    break
            except Exception:
                continue
            else:
                resize_image(image_path, width, height)
    return {"status": True, "images": images}


def download_reporting_pictures(
    s3,
    path: str,
    reporting: Reporting,
    width: int,
    height: int,
    enable_is_shared_antt=False,
    enable_include_dnit=True,
    low_quality=False,
):
    files = reporting.reporting_files
    if enable_include_dnit:
        files = files.filter(include_dnit=True)
    if enable_is_shared_antt:
        files = files.filter(is_shared=True)
    files = files.order_by("datetime")
    return download_files_pictures(s3, path, files, width, height, low_quality)


def download_reporting_file_pictures(
    s3,
    path: str,
    reporting_file_uuid: Union[UUID, list],
    width: int = 0,
    height: int = 0,
    order_by="datetime",
    enable_is_shared_antt=False,
    enable_include_dnit=True,
    low_quality=True,
) -> dict:
    images = []
    queryset_filter_common_image_type = _get_query_upload_valid_img()
    if isinstance(reporting_file_uuid, list):
        query_filter = Q(uuid__in=reporting_file_uuid)
    else:
        query_filter = Q(uuid=reporting_file_uuid)
    files = ReportingFile.objects.filter(query_filter)
    if enable_include_dnit:
        files = files.filter(include_dnit=True)
    files = files.filter(queryset_filter_common_image_type).distinct()
    if enable_is_shared_antt:
        files = files.filter(is_shared=True)

    if order_by is not None:
        files = files.order_by(order_by)
    for index, file_obj in enumerate(files):
        if file_obj.upload:
            try:
                file_path = file_obj.upload.url.split("?")[0].split(".com/")[1]
                bucket_name = file_obj.upload.url.split(".s3")[0].split("/")[-1]
                image_format = file_path.split(".")[-1]
                image_name = str(index)
            except Exception:
                continue

            image_path = "{}{}.{}".format(
                path, image_name + str(file_obj.uuid), image_format
            )

            if low_quality:
                try:
                    uncoded_file_name = urllib.parse.unquote(file_path)
                    s3.download_file(
                        bucket_name + "-400px", uncoded_file_name, image_path
                    )
                except Exception:
                    try:
                        uncoded_file_name = urllib.parse.unquote(file_path)
                        s3.download_file(bucket_name, uncoded_file_name, image_path)
                    except Exception:
                        continue
            else:
                try:
                    uncoded_file_name = urllib.parse.unquote(file_path)
                    s3.download_file(bucket_name, uncoded_file_name, image_path)
                except Exception:
                    continue
            try:
                image_path = save_image_as_png(image_path)
                images.append(
                    {
                        "path": image_path,
                        "description": file_obj.description,
                        "uuid": str(file_obj.uuid),
                    }
                )
            except Exception:
                continue
    return {"status": True if images else False, "images": images}


def download_reporting_file_picture(
    file_list: List[ReportingFile],
    s3,
    path,
    width: int,
    height: int,
):
    if len(file_list) > 0:
        file_obj = file_list[0]
    else:
        raise Exception("No reporting file")

    if file_obj.upload:
        file_path = file_obj.upload.url.split("?")[0].split(".com/")[1]
        bucket_name = file_obj.upload.url.split(".s3")[0].split("/")[-1]
        image_format = file_path.split(".")[-1]

        if not check_image_file(file_path):
            raise Exception("Image check fail")

        image_path = "{}{}.{}".format(path, str(file_obj.uuid), image_format)

        uncoded_file_name = urllib.parse.unquote(file_path)
        s3.download_file(bucket_name, uncoded_file_name, image_path)
        image_path = save_image_as_png(image_path)
        resized = resize_image(image_path, width, height)
        if not resized:
            raise Exception("Image resize fail")
        return {
            "path": image_path,
            "description": file_obj.description,
            "uuid": str(file_obj.uuid),
        }
    else:
        raise Exception("Image without upload")


def result_photos(
    s3,
    temp_file: str,
    photo_id: UUID,
    width: int,
    height: int,
    enable_is_shared_antt=False,
    enable_include_dnit=True,
):
    reporting_files = download_reporting_file_pictures(
        s3,
        temp_file,
        photo_id,
        width,
        height,
        enable_is_shared_antt=enable_is_shared_antt,
        enable_include_dnit=enable_include_dnit,
    )
    photo_result = (
        [x["path"] for x in reporting_files["images"] if x["uuid"] == photo_id]
        if photo_id
        else ""
    )
    return photo_result


def insert_img(
    image: str,
    worksheet: Worksheet,
    row_init: int,
    row_end: int,
    col_init: int,
    col_end: int,
    width: int,
    height: int,
    scale_width: float = 1,
    scale_height: float = 1,
) -> None:
    img = Image(image)
    med_width = width / img.width
    med_height = height / img.height
    average = min(med_width, med_height)
    img.width = int(img.width * average) * scale_width
    img.height = int(img.height * average) * scale_height

    left = round((col_init + col_end) / 2)
    top = round((row_init + row_end) / 2)

    worksheet.add_image(img, f"{worksheet.cell(row=top, column=left).coordinate}")


def resize_image(image_path: str, width: int, height: int) -> bool:
    if width != 0 or height != 0:
        try:
            image = ImagePil.open(image_path)
            image = image.resize((width, height))
            image.save(image_path)
        except Exception:
            return False
        try:
            image.close()
        except Exception:
            pass
    return True


def save_image_as_png(image_path: str) -> str:
    base_name, current_ext = os.path.splitext(image_path)
    type_image = imghdr.what(image_path)
    types_ok = [".jpeg", ".jpg", ".png", ".gif", ".bmp"]
    image = ImagePil.open(image_path)
    try:
        transposed = ImageOps.exif_transpose(image)
        image.close()
        image = transposed
    except Exception:
        pass
    ext = ".png"
    new_image_path = base_name + ext
    if (
        str(current_ext).lower() not in types_ok
        or f".{str(type_image)}" not in types_ok
        or image.format not in ["BMP", "GIF", "JPEG", "JPEG 2000", "PNG"]
    ):
        i = 1
        while isfile(new_image_path):
            new_image_path = f"{base_name}({i}){ext}"
            i += 1
        image.save(new_image_path, "PNG")
        image_path = new_image_path

    try:
        image.close()
    except Exception:
        pass
    return image_path


def get_subcompany_logo_file(s3, temp_prefix: str, reporting: Reporting) -> str:
    firm: Firm = getattr(reporting, "firm", None)
    logo = ""
    if firm and firm.subcompany and firm.subcompany.logo:
        logo = firm.subcompany.logo
        try:
            file_path = logo.url.split("?")[0].split(".com/")[1]
            bucket_name = logo.url.split(".s3")[0].split("/")[-1]
            image_format = file_path.split(".")[-1]
            encoded_file_name = urllib.parse.quote(file_path)
            logo = f"{temp_prefix}{uuid4()}.{image_format}"
            s3.download_file(bucket_name, encoded_file_name, logo)
        except Exception:
            logo = ""
    return logo


def get_logo_file(s3, temp_prefix: str, reporting: Reporting) -> str:
    company = getattr(reporting, "company", None)
    logo = ""
    if company and company.logo:
        try:
            file_path = company.logo.url.split("?")[0].split(".com/")[1]
            bucket_name = company.logo.url.split(".s3")[0].split("/")[-1]
            image_format = file_path.split(".")[-1]
            encoded_file_name = urllib.parse.quote(file_path)
            logo = f"{temp_prefix}{uuid4()}.{image_format}"
            s3.download_file(bucket_name, encoded_file_name, logo)
        except Exception:
            logo = ""
    return logo


def get_provider_logo_file(s3, temp_prefix: str, reporting: Reporting) -> str:
    company = getattr(reporting, "company", None)
    provider_logo = ""
    if company and company.provider_logo:
        try:
            file_path = company.provider_logo.url.split("?")[0].split(".com/")[1]
            bucket_name = company.provider_logo.url.split(".s3")[0].split("/")[-1]
            image_format = file_path.split(".")[-1]
            encoded_file_name = urllib.parse.quote(file_path)
            provider_logo = f"{temp_prefix}{uuid4()}.{image_format}"
            s3.download_file(bucket_name, encoded_file_name, provider_logo)
        except Exception:
            provider_logo = ""
    return provider_logo


def get_image(
    s3,
    dir: str,
    image_name: str,
    reporting_file: ReportingFile,
    width: int = 0,
    height: int = 0,
) -> Image:
    try:
        picture_path = download_picture(
            s3,
            dir,
            image_name,
            reporting_file=reporting_file,
        )
        if width != 0 and height != 0:
            resize_image(picture_path, width, height)
        return Image(picture_path)
    except Exception:
        return None


def insert_img_in_cell(
    path_image: str,
    worksheet: Worksheet,
    row_init: int,
    col_init: int,
    width: int = None,
    height: int = None,
    horizontal_offset: float = 0,
    vertical_offset: float = 0,
):
    img = Image(path_image)
    coordinate = f"{worksheet.cell(row=row_init, column=col_init).coordinate}"
    if width:
        img.width = width

    if height:
        img.height = height

    if horizontal_offset or vertical_offset:
        # Convert the offsets to EMUs
        horizontal_offset_EMU = pixels_to_EMU(horizontal_offset)
        vertical_offset_EMU = pixels_to_EMU(vertical_offset)

        # Create the anchor marker
        marker = AnchorMarker(
            col=col_init - 1,
            colOff=horizontal_offset_EMU,
            row=row_init - 1,
            rowOff=vertical_offset_EMU,
        )
        size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))

        # Create an anchor with the marker
        anchor = OneCellAnchor(_from=marker, ext=size)

        # Set the anchor to the image
        img.anchor = anchor

        worksheet.add_image(img)

    else:
        worksheet.add_image(img, coordinate)


def insert_logo_and_provider_logo(
    worksheet,
    logo_company: dict = None,
    provider_logo: dict = None,
    target: SheetTarget = SheetTarget.DesktopExcel,
):
    """
    Inserts the company logo and provider logo into the specified worksheet if they are provided.

    This function checks if the `logo_company` and `provider_logo` dictionaries are not None and contain a path to an image. If so, it proceeds to insert these images into the worksheet at the specified locations with the given dimensions and resize method. The `insert_picture` function is used to handle the actual insertion of the images, which includes resizing according to the specified method and dimensions.

    :param worksheet: The worksheet object where the logos will be inserted.
    :param logo_company: A dictionary containing information about the company logo, including its path, range string, width, height, and resize method. If None, the company logo will not be inserted.
    :param provider_logo: A dictionary containing information about the provider logo, including its path, range string, width, height, and resize method. If None, the provider logo will not be inserted.
    :param target: The target application for the worksheet, which affects how the image is resized. Defaults to SheetTarget.DesktopExcel.
    """
    if logo_company and logo_company.get("path_image"):
        insert_picture_2(
            picture=Image(logo_company.get("path_image")),
            worksheet=worksheet,
            target=target,
            range_string=logo_company.get("range_string"),
            resize_method=logo_company.get(
                "resize_method", ResizeMethod.ProportionalCentered
            ),
            border_width=(2, 2, 2, 2),
        )

    if provider_logo and provider_logo.get("path_image"):
        insert_picture_2(
            picture=Image(provider_logo.get("path_image")),
            worksheet=worksheet,
            target=target,
            range_string=provider_logo.get("range_string"),
            resize_method=provider_logo.get(
                "resize_method", ResizeMethod.ProportionalCentered
            ),
            border_width=(2, 2, 2, 2),
        )


def download_picture_to_url(
    s3, dir: str, upload: str, image_name: str = str(uuid4())
) -> str:
    try:
        file_path = upload.split("?")[0].split(".com/")[1]
        bucket_name = upload.split(".s3")[0].split("/")[-1]
        image_format = file_path.split(".")[-1]
    except Exception:
        return ""

    if not check_image_file(file_path):
        return ""

    image_path = "{}/{}.{}".format(dir, image_name, image_format)
    temp_file = f"{image_path}"
    try:
        s3.download_file(bucket_name, unquote(file_path), temp_file)
        return temp_file
    except Exception:
        return ""


def insert_picture_to_path(
    path_picture,
    worksheet,
    range_string,
    target: SheetTarget = SheetTarget.DesktopExcel,
    width=0,
    height=0,
    border_width=0,
    resize_method=ResizeMethod.Stretch,
):
    if path_picture:
        insert_picture(
            picture=Image(path_picture),
            worksheet=worksheet,
            target=target,
            range_string=range_string,
            width=width,
            height=height,
            border_width=border_width,
            resize_method=resize_method,
        )
