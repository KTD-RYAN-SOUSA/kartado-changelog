from copy import copy
from typing import List

from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from helpers.strings import clean_latin_string


def save_workbook(file_name: str, workbook: Workbook) -> str:
    file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
    result_file = f"/tmp/{file_name}.xlsx"
    workbook.save(result_file)
    return result_file


def set_block_style(
    worksheet: Worksheet,
    row_begin: int,
    row_end: int,
    col_begin: str,
    col_end: str,
    height: float,
    font: Font,
    border: Border,
    alignment: Alignment,
):
    for row in range(row_begin, row_end + 1):
        worksheet.row_dimensions[row].height = height
    cells_range = "{}{}:{}{}".format(col_begin, row_begin, col_end, row_end)
    rows = worksheet[cells_range]
    for row in rows:
        for cell in row:
            cell: Cell = cell
            cell.font = font
            cell.border = border
            cell.alignment = alignment


def set_row_style(
    worksheet: Worksheet,
    row: int,
    height: float,
    font: Font,
    border: Border,
    alignment: Alignment,
):
    worksheet.row_dimensions[row].height = height
    for cell in worksheet["{}:{}".format(row, row)]:
        cell: Cell = cell
        cell.font = font
        cell.border = border
        cell.alignment = alignment


def append_row(worksheet: Worksheet, values: List[str], offset: int = 0) -> None:
    """
    Appends a new row with the style of the previous last row.
    """
    max_row = worksheet.max_row
    row = max_row + 1 + offset
    col = 1
    for value in values:
        template_cell: Cell = worksheet.cell(max_row, col)
        cell: Cell = worksheet.cell(row, col)
        cell.value = value
        copy_cell_style(cell, template_cell)
        col += 1
    worksheet.row_dimensions[row] = worksheet.row_dimensions[max_row]


def copy_cell_style(dst: Cell, src: Cell) -> None:
    """
    Copy style from a row to another within the same
    worksheet.
    """
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def set_zoom(workbook: Workbook, zoom: int, view: str = "normal"):
    """
    Sets zoom of all worksheets in the workbook to the given zoom level.

    :param workbook: workbook to set zoom level
    :param zoom: zoom level
    :param view: sets page view. Accepts 'normal', 'pageBreakPreview', 'pageLayout'
    """
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.zoomToFit = False
        worksheet.sheet_view.view = view
        worksheet.sheet_view.zoomScale = zoom
        worksheet.sheet_view.zoomScaleNormal = zoom
        worksheet.sheet_view.zoomScaleSheetLayoutView = zoom
        worksheet.sheet_view.zoomScalePageLayoutView = zoom


def set_active_cell(workbook: Workbook, cell: str, pane_to_active: bool = True) -> None:
    """
    Sets active cell of all worksheets in the workbook to the given cell

    :param workbook: workbook to set active cell
    :param cell: cell
    :param pane_to_active: when enabled, sets the view top-left cell to the active cell
    """
    for worksheet in workbook.worksheets:
        for sv in worksheet.views.sheetView:
            for s in sv.selection:
                s.activeCell = cell
            if pane_to_active:
                sv.topLeftCell = cell
