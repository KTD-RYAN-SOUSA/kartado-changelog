import json
import shutil
import tempfile
from ast import List

from django.core import serializers
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.export_utils import get_s3
from helpers.apps.ccr_report_utils.image import (
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
)


class BaseXlsxHandler(object):
    def __init__(
        self,
        query_set_serializer: str,
        path_file_xlsx: str,
        sheet_target=SheetTarget.DesktopExcel,
        is_logo: bool = True,
        is_sort_files: bool = True,
        split_xlsx_in: List = ["road_name"],
        orders_rows: dict = {
            "filters": [
                {
                    "field": "",
                    "value": "",
                }
            ],
            "sorts": [],
        },
        **kwargs,
    ):
        self.__temp_file: str = tempfile.mkdtemp()
        self.s3 = kwargs.get("s3", get_s3())
        self.__xlsx_file = path_file_xlsx
        self.workbook: Workbook = None
        self.worksheet: Worksheet = None
        self.is_logo: bool = is_logo
        self.data_logo_company: dict = dict(
            path_image="",
            range_string=kwargs.get("logo_company_range_string", "B1:B1"),
            resize_method=ResizeMethod.ProportionalRight,
        )
        self.data_provider_logo: dict = dict(
            path_image="",
            range_string=kwargs.get("provider_logo_range_string", "A1:A1"),
            resize_method=ResizeMethod.ProportionalLeft,
        )

        self.sheet_target = sheet_target
        self.reload_workbook(path_file_xlsx)
        self.queryset_reporting = self.deserializer_queryset(query_set_serializer)
        self.is_sort_files: bool = is_sort_files
        self._split_xlsx_in: list = []
        self.orders_rows: dict = {}
        if is_sort_files:
            self._split_xlsx_in = split_xlsx_in
            self.orders_rows: dict = orders_rows

    def insert_logos(
        self, logo_company_config: dict = {}, provider_logo_config: dict = {}
    ):
        insert_logo_and_provider_logo(
            worksheet=self.worksheet,
            target=self.sheet_target,
            logo_company=logo_company_config,
            provider_logo=provider_logo_config,
        )

    def get_temp_dir(self) -> str:
        return self.__temp_file

    def reload_workbook(self, file_path_xlsx=None):
        file_path = file_path_xlsx or self.__xlsx_file
        self.workbook = load_workbook(file_path)
        self.worksheet = self.workbook.active

    def deserializer_queryset(self, json_string):
        try:
            # Verifica se a string é um JSON válido
            json.loads(json_string)
            array = []
            gen = serializers.deserialize("json", json_string)
            for obj in gen:
                obj_instance = obj.object  # Obtenha a instância do modelo Django
                # Realize o processamento necessário no obj_instance
                array.append(obj_instance)
            return array
        except json.JSONDecodeError:
            return json_string

    def set_logos_basic(self, reporting):
        if not self.data_logo_company.get("path_image"):
            path_logo_company = get_logo_file(
                s3=self.s3,
                temp_prefix=self.get_temp_dir(),
                reporting=reporting,
            )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

        if not self.data_provider_logo.get("path_image"):
            path_provider_logo = get_provider_logo_file(
                s3=self.s3,
                temp_prefix=self.get_temp_dir(),
                reporting=reporting,
            )
            if path_provider_logo:
                self.data_provider_logo["path_image"] = path_provider_logo

    def get_sorted_data(self, array, sorts) -> list:
        """
        Ordena uma lista de dados com base em critérios de ordenação fornecidos.

        Esta função aceita uma lista de dados e um conjunto de critérios de ordenação (sorts). Ela utiliza uma função de chave personalizada para ordenar a lista de dados de acordo com os critérios fornecidos. A ordenação pode ser feita em múltiplas colunas, levando em conta a prioridade dos campos especificados na lista de ordenação.

        Parâmetros:
        - array (list): Lista de dados a serem ordenados.
        - sorts (list): Lista de tuplas, onde cada tupla contém o nome do campo pelo qual ordenar e uma flag indicando se a ordenação deve ser ascendente (True) ou descendente (False).

        Retorna:
        - list: A lista de dados ordenada de acordo com os critérios fornecidos.
        """

        def generate_key_function(priority_fields):
            def key_function(item):
                return tuple(item[field] for field in priority_fields)

            return key_function

        return (
            sorted(
                array,
                key=generate_key_function(sorts),
                reverse=False,
            )
            if array
            else array
        )

    def sort_data_work(self, data_work: dict, orders: dict) -> dict:
        """
        Ordena um dicionário de dados trabalhando, aplicando filtros e ordenações específicas.

        Esta função aceita um dicionário de dados trabalhando, onde cada chave representa um grupo de dados e cada valor é uma lista de dicionários contendo os dados desse grupo. Além disso, ela aceita um dicionário de ordens que contém critérios de filtragem e ordenação para os grupos de dados. A função aplica esses critérios ao dicionário de dados trabalhando, retornando um novo dicionário com os dados ordenados e filtrados conforme especificado.

        Parâmetros:
        - data_work (dict): Dicionário de dados trabalhando, onde cada chave representa um grupo de dados e cada valor é uma lista de dicionários contendo os dados desse grupo.
        - orders (dict): Dicionário contendo critérios de filtragem e ordenação para os grupos de dados. Deve conter chaves para 'filters', onde 'filters' define condições para filtrar os dados com base em campos específicos e valores exatos. A ordem dos filtros é crucial e determina a prioridade na aplicação dos critérios de filtragem.

        Retorna:
        - dict: Um dicionário com os dados ordenados e filtrados conforme especificado nas ordens.
        """
        sorts = orders.get("sorts", [])
        for key, _data_list in data_work.items():
            _orders = []
            _values_excludes = []
            for _filter in orders.get("filters", []):
                _field = _filter.get("field")
                _value = _filter.get("value")
                if _field and _value:
                    exec(
                        f"{_value} = [obj for obj in {_data_list} if obj[{_field}].lower() == {_value}]"
                    )
                    _data = []

                    locals()[_value] = self.get_sorted_data(locals()[_value], sorts)
                    _values_excludes.append(_field)
                    _data.extend(locals()[_field])
                    _rest_data = [
                        obj for obj in _data_list if obj[_field] not in _values_excludes
                    ]
                    _rest_data = self.get_sorted_data(_rest_data, sorts)
                    _data.extend(_rest_data)

                    _orders.extend(_data)

                else:
                    _orders = self.get_sorted_data(_data_list, sorts)

            data_work[key] = _orders

        return data_work

    def rules_files_data_work(
        self,
        data_list: list,
        split_xlsx_in: list = [],
        orders: dict = {},
    ) -> dict:
        """
        Organiza e agrupa dados de entrada em um dicionário estruturado, aplicando filtros e ordenações específicos.

        Esta função recebe uma lista de dados (`data_list`) e aplica uma lógica de agrupamento, filtragem e ordenação sobre esses dados, utilizando as chaves especificadas em `split_xlsx_in` para determinar como os dados devem ser divididos. Os critérios de filtragem e ordenação são definidos em `orders`.

        A função retorna um dicionário (`data_work`) onde cada chave representa um grupo único de dados, identificado pela combinação das chaves especificadas em `split_xlsx_in`. Os valores associados a cada chave são listas de dicionários contendo os dados pertencentes a esse grupo, já filtrados e ordenados de acordo com os critérios especificados em `orders`.

        Parâmetros:
        - data_list (list): Lista de dicionários contendo os dados a serem organizados.
        - split_xlsx_in (list, optional): Lista de chaves pelos quais os dados serão divididos. Se não fornecido, todos os dados serão considerados como parte de um único grupo.
        - orders (dict, optional): Dicionário contendo critérios de filtragem e ordenação para os grupos de dados. Deve conter chaves para 'filters', onde 'filters' define condições para filtrar os dados com base em campos específicos e valores exatos. A ordem dos filtros é crucial e determina a prioridade na aplicação dos critérios de filtragem.

        Retorna:
        - dict: Dicionário com a chave sendo uma string única que representa o grupo de dados e o valor sendo uma lista de dicionários contendo os dados desse grupo, já filtrados e ordenados de acordo com os critérios especificados em `orders`.

        Exemplo:
            >>> data_list = [{'direction': 'norte', 'road_name': "BR 101 RJ"}, {'direction': 'sul', 'road_name': "BR 101 RJ"}]
            >>> split_xlsx_in = ['road_name']
            >>> orders = {'filters': [{'field': 'direction', 'value': 'sul'}, {'field': 'direction', 'value': 'norte'}]}
            >>> handler.rules_files_data_work(data_list, split_xlsx_in, orders)
            {
                'BR 101 RJ': [
                    {'direction': 'sul', 'road_name': "BR 101 RJ"},
                    {'direction': 'norte', 'road_name': "BR 101 RJ"}
                ]
            }
        """
        data_work = {}

        for _d in data_list:
            try:
                tmp_key = ""
                for index, key in enumerate(split_xlsx_in):
                    if index > 0:
                        tmp_key += "|"
                    tmp_key += str(_d.get(key))
                key_to_xlsx = tmp_key
            except Exception:
                continue

            if key_to_xlsx not in data_work:
                data_work[key_to_xlsx] = []

            data_work[key_to_xlsx].append(_d)

        data_work = self.sort_data_work(data_work, orders)

        return data_work

    def merge_cells(
        self,
        start_row: int = 1,
        start_column: int = 1,
        end_row: int = 1,
        end_column: int = 1,
        alignment_horizontal="center",
        alignment_vertical="center",
    ):
        self.worksheet.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )
        merged_cell = self.worksheet.cell(row=start_row, column=start_column)
        merged_cell.alignment = Alignment(
            horizontal=alignment_horizontal, vertical=alignment_vertical
        )
        return merged_cell

    def insert_new_row(
        self,
        row: int,
        number_col: int,
        height_row: float = 18,
        alignment: Alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=0, wrap_text=True
        ),
        border: Border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    ):
        chars = [chr(ord("A") + i) for i in range(number_col)]
        if height_row:
            self.worksheet.row_dimensions[row].height = height_row

        for char in chars:
            if border:
                self.worksheet[f"{char}{row}"].border = border
            if alignment:
                self.worksheet[f"{char}{row}"].alignment = alignment

    def create_dict(self, reporting: Reporting) -> dict:
        """
        Cria um dicionário a partir dos dados fornecidos e da instância de Reporting.

        Este método deve ser implementado por subclasses para definir como os dados são transformados em um formato específico
        adequado para a geração da planilha XLSX.

        Parâmetros:
        - data (dict): Dicionário Python contendo as informações a serem inseridas na planilha.
        - reporting (Reporting): Instância de Reporting que contém informações adicionais necessárias para a criação do dicionário.

        Retorna:
        - dict: Um dicionário formatado que será usado para preencher a planilha XLSX.
        """
        raise NotImplementedError

    def fill_sheet(self, data_work: dict) -> List:
        """
        Preenche a planilha com base nos dados fornecidos no dicionário data_work, que é organizado e ordenado pelo método rules_files_data_work.

        O dicionário data_work é esperado a ser criado após a chamada do método create_dict para cada instância de Reporting, onde cada chave representa um arquivo de dados e seus valores são listas de dicionários contendo os dados a serem inseridos na planilha.

        Os dados dentro de data_work são então passados para o método rules_files_data_work, que organiza e ordena esses dados conforme especificado nas chaves 'keys' e 'orders'. Isso resulta em um dicionário data_work atualizado, onde cada chave ainda representa um arquivo de dados, mas agora os dados estão ordenados e organizados de acordo com as regras definidas.

        Após a organização, o método preenche a planilha com esses dados, garantindo que cada arquivo de dados seja corretamente inserido na posição apropriada na planilha.

        O ponto de saída deste método é uma lista de strings, onde cada string é um caminho (path) para o arquivo gerado. Esses caminhos representam os arquivos de dados que foram criados ou modificados durante o processo de preenchimento da planilha.

        Args:
            data_work (dict): Dicionário contendo os dados a serem inseridos na planilha, organizado por arquivo de dados.

        Returns:
            List[str]: Lista de caminhos (strings) para os arquivos gerados durante o processo de preenchimento da planilha.
        """
        raise NotImplementedError

    def execute(self) -> list:
        """
        Executa o processo de criação e preenchimento de uma planilha XLSX com dados de relatórios.

        Esta função coordena várias etapas para preparar e preencher uma planilha XLSX com dados relevantes extraídos de objetos `Reporting`. Primeiro, ela itera sobre cada objeto `Reporting` disponível, criando um dicionário para cada um usando a função `create_dict`. Se configurado, também configura os logotipos básicos para cada relatório. Em seguida, organiza e ordena os dados conforme especificado pelas configurações de `split_xlsx_in` e `orders_rows`, aplicando filtros e ordenações conforme necessário. Finalmente, preenche a planilha XLSX com os dados organizados e remove temporariamente o diretório temporário utilizado durante o processo.

        Retorna:
        - list: Uma lista de caminhos (strings) para os arquivos XLSX gerados durante o processo de preenchimento da planilha.

        Nota: Esta função assume que todas as funções auxiliares necessárias (como `create_dict`, `rules_files_data_work`, e `fill_sheet`) foram implementadas corretamente e podem manipular os dados conforme esperado.
        """
        data = []
        for reporting in self.queryset_reporting:
            data.append(self.create_dict(reporting=reporting))

            if self.is_logo:
                self.set_logos_basic(reporting)

        data = list(filter(None.__ne__, data))
        data.reverse()

        if self._split_xlsx_in:
            data_work = self.rules_files_data_work(
                data_list=data,
                split_xlsx_in=self._split_xlsx_in,
                orders=self.orders_rows,
            )
        else:
            data_work = data

        files = self.fill_sheet(data_work=data_work)
        shutil.rmtree(self.get_temp_dir(), ignore_errors=True)

        return files
