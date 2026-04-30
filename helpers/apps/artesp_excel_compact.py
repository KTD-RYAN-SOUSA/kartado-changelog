import math
import os
from copy import copy
from datetime import datetime, timedelta

import boto3
import pytz
from arrow import Arrow
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval, range_boundaries
from openpyxl.worksheet.cell_range import CellRange
from zappa.asynchronous import task

from apps.companies.models import Company
from apps.reportings.models import Reporting
from helpers.apps.inventory_schedule import InventoryScheduleEndpoint
from helpers.dates import utc_to_local
from RoadLabsAPI.settings import credentials


def datetime_to_date(datetime, clear_tzinfo=True):
    try:
        if clear_tzinfo:
            datetime = utc_to_local(datetime).replace(tzinfo=None)
        date = datetime.date()
    except Exception:
        date = None
    return date


class ArtespExcelEndpoint:
    temp_path = "/tmp/excel_artesp_export"
    # len_rows = 18

    first_data_boundaries = "D15:JE74"
    page_offset = 60
    cover_ws_index = 0
    schedule_ws_index = 1
    page_count = 1

    def __init__(
        self,
        schedule_data,
        excel_name,
        template_path="fixtures/reports/artesp_excel_compact.xlsm",
    ):
        self.template_path = template_path
        self.schedule_data = schedule_data
        self.path_excel = "{}/excel".format(self.temp_path)
        os.makedirs(self.path_excel, exist_ok=True)

        self.excel_name = excel_name
        self.file_path_name = "{}/{}.xlsm".format(self.path_excel, self.excel_name)

    def load_data(self):
        try:
            self.wb = load_workbook(filename=self.template_path, keep_vba=True)
        except Exception:
            self.wb = None

    def upload_file(self, path, name):
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        expires = datetime.now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)
        object_name = "media/private/{}".format(name)

        try:
            self.s3.upload_file(
                path, bucket_name, object_name, ExtraArgs={"Expires": expires}
            )
        except Exception:
            return False

        # Delete file
        os.remove(path)

        url_s3 = self.s3.generate_presigned_url(
            "get_object", Params={"Bucket": bucket_name, "Key": object_name}
        )
        return url_s3

    def add_page(self):

        ws = self.wb.worksheets[self.schedule_ws_index]

        min_col, min_row, max_col, max_row = range_boundaries(
            self.first_data_boundaries
        )

        for row, row_cells in enumerate(ws[min_row:max_row], min_row):
            for column, cell in enumerate(row_cells[min_col - 1 : max_col], min_col):
                #         # Copy Value from Copy.Cell to given Worksheet.Cell
                ws.cell(
                    row=row + self.page_offset * self.page_count, column=column
                ).value = cell.value
                ws.cell(
                    row=row + self.page_offset * self.page_count, column=column
                )._style = cell._style
                ws.row_dimensions[row + self.page_offset * self.page_count] = copy(
                    ws.row_dimensions[row]
                )

        for mcr in ws.merged_cells:
            if mcr.coord not in CellRange(self.first_data_boundaries):
                continue
            cr = CellRange(mcr.coord)
            cr.shift(row_shift=self.page_offset * self.page_count)
            ws.merge_cells(cr.coord)

        self.page_count += 1

    def set_print_area(self):
        ws = self.wb.worksheets[self.schedule_ws_index]

        end_row = self.page_count * self.page_offset + 3

        ws.print_area = "D12:JE{}".format(end_row)
        ws.print_title_rows = "1:3"

    def hide_inspection_columns(self):
        ws = self.wb.worksheets[self.schedule_ws_index]

        hide_start = 19
        hide_end = 74
        if self.schedule_data["routine_inspection_columns"] > 1:
            hide_start += (self.schedule_data["routine_inspection_columns"] - 1) * 4

        hide_interval = get_column_interval(hide_start, hide_end)

        for row_letter, dimension in ws.column_dimensions.items():
            if row_letter in hide_interval:
                dimension.hidden = True

        for year_index, inspection_year in enumerate(
            self.schedule_data["routine_inspection_years"]
        ):
            ws.cell(row=3, column=12 + year_index * 4).value = inspection_year

    def hide_date_columns(self):
        ws = self.wb.worksheets[self.schedule_ws_index]

        first_column = "CD"
        date_start = 82
        date_end = 261
        hide_start = 94

        start_year = self.schedule_data["start_date"].year
        end_year = self.schedule_data["end_date"].year
        year_count = end_year - start_year + 1

        if year_count > 1:
            hide_start += (year_count - 1) * 12
        if hide_start > 261:
            hide_start = 261

        for year_index, year in enumerate(range(start_year, end_year + 1)):
            ws.cell(row=14, column=82 + year_index * 12).value = year

        date_interval = get_column_interval(date_start, date_end)
        for column_index, date_column in enumerate(date_interval, date_start):
            ws.column_dimensions[date_column] = copy(ws.column_dimensions[first_column])
            ws.column_dimensions[date_column].min = column_index
            ws.column_dimensions[date_column].max = column_index
            if column_index >= hide_start:
                ws.column_dimensions[date_column].hidden = True

    def add_inventory_items(self):
        ws = self.wb.worksheets[self.schedule_ws_index]

        for index, inventory in enumerate(self.schedule_data["inventory"]):
            page_index = math.floor(index / 6)
            main_row = 15 + (page_index * self.page_offset) + (10 * (index % 6))
            ws.cell(row=main_row, column=4).value = index + 1
            ws.cell(row=main_row, column=5).value = inventory["number"]
            ws.cell(row=main_row, column=6).value = inventory["occurrence_type_name"]
            ws.cell(row=main_row, column=7).value = inventory["km"]
            ws.cell(row=main_row, column=8).value = inventory["direction"]

            try:
                initial_inspection = next(
                    a
                    for a in inventory["inspections"]
                    if "inicial" in a["type"].lower()
                )
            except Exception:
                initial_inspection = None

            try:
                special_inspection = next(
                    a
                    for a in inventory["inspections"]
                    if "especial" in a["type"].lower()
                )
            except Exception:
                special_inspection = None

            try:
                routine_inspections = [
                    a
                    for a in inventory["inspections"]
                    if "rotineira" in a["type"].lower()
                ]
            except Exception:
                routine_inspections = []

            if initial_inspection:
                ws.cell(row=main_row, column=9).value = datetime_to_date(
                    initial_inspection["date"]
                )
                ws.cell(row=main_row, column=11).value = initial_inspection[
                    "structural_classification"
                ]
                ws.cell(row=main_row, column=12).value = initial_inspection[
                    "functional_classification"
                ]
                ws.cell(row=main_row, column=13).value = initial_inspection[
                    "wear_classification"
                ]

            if special_inspection:
                ws.cell(row=main_row, column=75).value = special_inspection[
                    "structural_classification"
                ]
                ws.cell(row=main_row, column=76).value = special_inspection[
                    "functional_classification"
                ]
                ws.cell(row=main_row, column=77).value = special_inspection[
                    "wear_classification"
                ]

            for index, routine_inspection in enumerate(routine_inspections):
                if "structural_classification" in routine_inspection:
                    ws.cell(
                        row=main_row, column=(15 + index * 4)
                    ).value = routine_inspection["structural_classification"]
                if "functional_classification" in routine_inspection:
                    ws.cell(
                        row=main_row, column=(16 + index * 4)
                    ).value = routine_inspection["functional_classification"]
                if "wear_classification" in routine_inspection:
                    ws.cell(
                        row=main_row, column=(17 + index * 4)
                    ).value = routine_inspection["wear_classification"]

            job_match_conditions = [
                lambda a: "especial" in a["title"].lower(),
                lambda a: "projeto" in a["title"].lower(),
                lambda a: "fase" in a["title"].lower()
                and "1" in a["title"].split("-")[0],
                lambda a: "fase" in a["title"].lower()
                and "2" in a["title"].split("-")[0],
                lambda a: "fase" in a["title"].lower()
                and "3" in a["title"].split("-")[0],
                lambda a: "fase" in a["title"].lower()
                and "4" in a["title"].split("-")[0],
                lambda a: "fase" in a["title"].lower()
                and "5" in a["title"].split("-")[0],
                lambda a: "barreira" in a["title"].lower(),
                lambda a: "gabarito" in a["title"].lower(),
            ]

            for job_index in range(9):
                try:
                    job = next(
                        a
                        for a in inventory["jobs"]
                        if job_match_conditions[job_index](a)
                    )
                except Exception:
                    job = None

                if not job:
                    continue

                if job["start_date"]:
                    ws.cell(
                        row=main_row + 1 + job_index, column=80
                    ).value = datetime_to_date(job["start_date"])

                if job["end_date"]:
                    ws.cell(
                        row=main_row + 1 + job_index, column=81
                    ).value = datetime_to_date(job["end_date"])

                if job["start_date"] and job["end_date"]:
                    date_list = []
                    for month in Arrow.range(
                        "month", job["start_date"], job["end_date"]
                    ):
                        date_list.append(month._datetime)

                    year_offset = (
                        date_list[0].year - self.schedule_data["start_date"].year
                    )
                    month_offset = (
                        date_list[0].month - self.schedule_data["start_date"].month
                    )
                    start_column = 82 + year_offset * 12 + month_offset
                    end_column = start_column + len(date_list) - 1
                    row_style = ws.cell(row=main_row + 1 + job_index, column=79)._style
                    for column in range(start_column, end_column + 1):
                        ws.cell(
                            row=main_row + 1 + job_index, column=column
                        )._style = row_style

            jobs = {"pending": "", "done": ""}
            for job in inventory["jobs"]:
                job_key = "done" if job["is_done"] else "pending"
                jobs[job_key] += job["title"].split(" - ")[0] + "\n"
                for service in job["all_services"]:
                    service_name = (
                        service["description"]
                        if service["description"]
                        else service["occurrence_type_name"]
                    )
                    if service["date"]:
                        jobs[job_key] += "- ({}) {}\n".format(
                            service["date"].strftime("%d/%m/%Y"), service_name
                        )
                    else:
                        jobs[job_key] += "- {} \n".format(service_name)

            ws.cell(row=main_row + 1, column=262).value = jobs["done"]
            ws.cell(row=main_row + 1, column=264).value = jobs["pending"]

    def get_data(self):
        empty = {"url": "", "name": ""}

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )

        # Load templates
        self.load_data()
        if not self.wb:
            return empty

        self.hide_inspection_columns()
        self.hide_date_columns()

        for _ in range(math.floor(len(self.schedule_data["inventory"]) / 6)):
            self.add_page()

        self.add_inventory_items()
        self.set_print_area()

        self.wb.save(self.file_path_name)

        url = self.upload_file(self.file_path_name, self.file_path_name.split("/")[-1])

        if not url:
            return empty

        # Delete excel folder
        for file_name in os.listdir(self.path_excel):
            os.remove(self.path_excel + "/" + file_name)
        os.rmdir(self.path_excel)

        return True


def get_url_compact(excel_name):
    empty = {"url": "", "name": ""}

    s3 = boto3.client(
        "s3",
        aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
        aws_session_token=credentials.AWS_SESSION_TOKEN,
    )

    bucket_name = settings.AWS_STORAGE_BUCKET_NAME
    object_name = "media/private/{}".format(excel_name + ".xlsm")

    url = s3.generate_presigned_url(
        "get_object", Params={"Bucket": bucket_name, "Key": object_name}
    )

    if not url:
        return empty

    return {"url": url, "name": excel_name + ".xlsm"}


def get_excel_name(road_name):
    now = utc_to_local(datetime.now())
    return "Cronograma_{}_{}_{}_{}_{}_{}".format(
        road_name, now.day, now.month, now.year, now.hour, now.minute
    )


@task
def run_async_artesp_excel_export_compact(uuids, company, excel_name):

    queryset = Reporting.objects.filter(uuid__in=uuids)

    schedule_data = InventoryScheduleEndpoint(
        inventory=queryset, company=Company.objects.get(uuid=company)
    ).get_data()

    excel_endpoint = ArtespExcelEndpoint(schedule_data, excel_name)

    excel_endpoint.get_data()
