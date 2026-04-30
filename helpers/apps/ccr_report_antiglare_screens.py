import re
import tempfile
from datetime import datetime
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.companies.models import Firm
from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_picture,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.apps.ccr_resume_description_measures_antiglare_screens import (
    XlsxHandlerResumeReportDescriptionMeasuresAntGlareScreens,
)
from helpers.strings import clean_latin_string, format_km

from .ccr_report_resume_antiglare_screens import XlsxHandlerResumeReportAntiglareScreens


class AntiGlareScreenXlsxHandler(object):
    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.wb = load_workbook("./fixtures/reports/ccr_report_antiglare_screen.xlsx")
        self.s3 = s3
        self.__sheet_target = sheet_target
        self._worksheet: Worksheet = self.wb.active

        self.form = OccurrenceType.objects.filter(
            name="Monitoração de Tela Antiofuscante"
        )
        self.list_uuids = list_uuids
        self.reportings = Reporting.objects.filter(uuid=list_uuids[0]).prefetch_related(
            "company"
        )
        self.temp_file = tempfile.mkdtemp()
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="T1:W5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B5",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company

        self._worksheet.column_dimensions[
            "K"
        ].width = self._worksheet.column_dimensions["J"].width
        self.static_fields = {
            "year": "A",
            "index": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "side": "F",
            "latitude": "G",
            "longitude": "H",
            "notes": "I",
            "photo_start": "J",
            "photo_end": "K",
            "height": "L",
            "length": "M",
            "general_state": {"BOM": "O", "REGULAR": "O", "RUIM": "O"},
            "corrosion_state": {"BOM": "Q", "REGULAR": "Q", "RUIM": "Q"},
            "alignment": {"BOM": "S", "REGULAR": "S", "RUIM": "S"},
            "fix_state": {"BOM": "U", "REGULAR": "U", "RUIM": "U"},
            "screws_state": {"BOM": "W", "REGULAR": "W", "RUIM": "W"},
        }

    def __insert_new_rows(self, row: int):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        chars = [chr(ord("A") + i) for i in range(10)]
        second_chars = [chr(ord("P") + i) for i in range(8)]

        for char in chars + ["K", "L", "M"]:
            self._worksheet.merge_cells(f"{char}{row}:{char}{row + 2}")
        for count in range(0, 3):
            self._worksheet.row_dimensions[row + count].height = 20
            for char in chars + second_chars + ["K", "L", "M", "N", "O"]:
                self._worksheet[f"{char}{row + count}"].border = border

    def create_dict(self, reporting):
        photos = reporting.form_data.get("mandatory_pictures")
        photo_start = ""
        photo_end = ""
        eps_start_images = []
        eps_end_images = []
        if photos:
            for photo in photos:
                if "eps_image_start" in photo:
                    eps_start_images.extend(photo["eps_image_start"])
                if "eps_image_end" in photo:
                    eps_end_images.extend(photo["eps_image_end"])
            eps_start_rfs = ReportingFile.objects.filter(
                uuid__in=eps_start_images
            ).order_by("datetime", "uploaded_at")
            for eps_start_rf in eps_start_rfs:
                photo_start = download_picture(
                    self.s3,
                    self.temp_file,
                    str(eps_start_rf.uuid),
                    reporting_file=eps_start_rf,
                    enable_is_shared=True,
                )
                if photo_start != "":
                    break
            eps_end_rfs = ReportingFile.objects.filter(
                uuid__in=eps_end_images
            ).order_by("datetime", "uploaded_at")
            for eps_end_rf in eps_end_rfs:
                photo_end = download_picture(
                    self.s3,
                    self.temp_file,
                    str(eps_end_rf.uuid),
                    reporting_file=eps_end_rf,
                    enable_is_shared=True,
                )
                if photo_end != "":
                    break

        result_year = new_get_form_data(reporting, "inspectionCampaignYear", default="")
        result_index = new_get_form_data(reporting, "index", "idCcrAntt", default="")
        result_notes = new_get_form_data(reporting, "notes", default="")

        direction = get_custom_option(reporting, "direction")

        result_side: str = new_get_form_data(reporting, "side", default="")
        side = ""
        try:
            side = result_side.upper()
        except Exception:
            pass
        result_corrosion = new_get_form_data(reporting, "corrosionState", default="")
        corrosion_state = {
            "BOM": "X" if result_corrosion == "Bom" else "",
            "REGULAR": "X" if result_corrosion == "Regular" else "",
            "RUIM": "X" if result_corrosion == "Ruim" else "",
        }

        result_aligment = new_get_form_data(reporting, "alignmentState", default="")
        alignment = {
            "BOM": "X" if result_aligment == "Bom" else "",
            "REGULAR": "X" if result_aligment == "Regular" else "",
            "RUIM": "X" if result_aligment == "Ruim" else "",
        }

        result_fix_state = new_get_form_data(reporting, "fixState", default="")
        fix_state = {
            "BOM": "X" if result_fix_state == "Bom" else "",
            "REGULAR": "X" if result_fix_state == "Regular" else "",
            "RUIM": "X" if result_fix_state == "Ruim" else "",
        }

        result_screw_state = new_get_form_data(reporting, "screwState", default="")
        screw_state = {
            "BOM": "X" if result_screw_state == "Bom" else "",
            "REGULAR": "X" if result_screw_state == "Regular" else "",
            "RUIM": "X" if result_screw_state == "Ruim" else "",
        }

        result_general_state = new_get_form_data(reporting, "generalState", default="")

        general_state = {
            "BOM": "X" if result_general_state == "Bom" else "",
            "REGULAR": "X" if result_general_state == "Regular" else "",
            "RUIM": "X" if result_general_state == "Ruim" else "",
        }

        data = {
            "year": result_year,
            "index": result_index,
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "end_km", 3),
            "raw_initial_km": reporting.km,
            "direction": direction.upper(),
            "side": side,
            "latitude": reporting.form_data.get("lat"),
            "longitude": reporting.form_data.get("long"),
            "notes": result_notes,
            "photo_start": photo_start,
            "photo_end": photo_end,
            "length": reporting.form_data.get("length"),
            "height": reporting.form_data.get("height"),
            "general_state": general_state,
            "corrosion_state": corrosion_state,
            "alignment": alignment,
            "fix_state": fix_state,
            "screws_state": screw_state,
            "executed_at": reporting.executed_at,
            "team": str(reporting.firm.__dict__.get("uuid")) if reporting.firm else "",
            "road_name": reporting.__dict__.get("road_name"),
            "subcompany": (
                reporting.firm.subcompany.__dict__.get("name") if reporting.firm else ""
            ),
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def __insert_status_values(self, key, cell, value):
        cell_status_string = self._worksheet[cell].offset(0, -1)
        if key == "tipo_refletor":
            cell_status_string.value = value
        else:
            cell_status_string.value = key.replace("_", " ") if key else key
            if value in "X":
                self._worksheet[cell] = "X"
            self._worksheet[cell] = value
            AntiGlareScreenXlsxHandler.__format_fonts(
                cell=self._worksheet[cell], size=10
            )

        AntiGlareScreenXlsxHandler.__format_fonts(
            cell=cell_status_string, size=10, horizontal="left"
        )

    def fill_sheet(self, *, data_list: list):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 7
        team_list = []
        date_list = []
        road_name_list = []
        subcompany_list = []

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            target=self.__sheet_target,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

        for values in data_list:
            for key, value in values.items():
                self.__insert_new_rows(row=row)
                if key in [
                    "general_state",
                    "corrosion_state",
                    "alignment",
                    "fix_state",
                    "screws_state",
                ]:
                    sum_intern_rows = 0
                    for _key, _value in value.items():
                        if not sum_intern_rows >= len(_key):
                            key_value = f"{self.static_fields[key][_key]}{row + sum_intern_rows}"
                            self.__insert_status_values(
                                key=_key, cell=key_value, value=_value
                            )
                            sum_intern_rows += 1
                elif key in ["photo_start", "photo_end"] and value:
                    col = self.static_fields[key]
                    range_str = f"{col}{row}:{col}{row+2}"
                    if value != "":
                        insert_picture_2(
                            self._worksheet,
                            range_str,
                            Image(value),
                            self.__sheet_target,
                            (1, 1, 1, 1),
                            ResizeMethod.ProportionalCentered,
                        )
                elif key == "executed_at":
                    value_date = (
                        value.strftime("%Y-%m-%d")
                        if isinstance(value, datetime)
                        else ""
                    )
                    date_list.append(value_date)
                elif key == "road_name":
                    road_name_list.append(value)
                elif key == "subcompany":
                    subcompany_list.append(value)
                elif key == "team":
                    team_list.append(values["team"])
                elif key == "raw_initial_km":
                    continue

                else:
                    key_value = f"{self.static_fields[key]}{row}"
                    self._worksheet[key_value] = value
                    AntiGlareScreenXlsxHandler.__format_fonts(
                        cell=self._worksheet[key_value], size=10
                    )

            row += 3

        filtered_team_list = set([uuid for uuid in team_list if uuid != ""])
        query_set_teams = Firm.objects.filter(uuid__in=filtered_team_list).all()
        names = []
        for team in query_set_teams:
            users_query_set = team.users.all()
            intern_list = []
            for user in users_query_set:
                intern_list.append(user.full_name)
            names.append(intern_list)
        names = [", ".join(_) for _ in names]
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(names),
            cell="I3",
            horizontal="left",
        )
        date_list.sort()
        filtered_date_list = [date for date in date_list if date != ""]
        if len(filtered_date_list) == 0:
            date_text = ""
        elif len(filtered_date_list) == 1:
            date_text = filtered_date_list[0]
        else:
            date_text = f"{filtered_date_list[0]} até {filtered_date_list[-1]}"

        insert_centered_value(
            worksheet=self._worksheet, value=date_text, cell="I4", horizontal="left"
        )

        road_name = list(set(road_name_list))
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(road_name),
            cell="M2",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value="Elementos de Proteção e Segurança - Ficha de Monitoração de Telas Antiofuscantes",
            cell="M3",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value=self.company.name,
            cell="M4",
            horizontal="left",
        )

        subcompany = list(set(subcompany_list))
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(subcompany),
            cell="I2",
            horizontal="left",
        )

        total_length = sum(
            [t.value for t in self._worksheet["M"] if isinstance(t.value, (int, float))]
        )
        reference_total_legth = self._worksheet[f"{'M'}{self._worksheet.max_row + 1 }"]
        reference_total_legth.value = total_length
        reference_total_legth.border = border
        AntiGlareScreenXlsxHandler.__format_fonts(
            cell=reference_total_legth, size=10, bold=True
        )

        reference_total_string = reference_total_legth.offset(0, -1)
        reference_total_string.value = "TOTAL(m)"
        reference_total_string.border = border
        AntiGlareScreenXlsxHandler.__format_fonts(
            cell=reference_total_string, size=10, bold=True
        )

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo
        sorted_data = sorted(
            data,
            key=lambda x: (
                x["raw_initial_km"] if x["direction"] == "NORTE" else 0,
                -x["raw_initial_km"] if x["direction"] == "SUL" else 0,
                x["direction"] == "CANTEIRO CENTRAL",
                x["direction"] == "AMBOS",
                x["direction"] == "CRESCENTE",
                x["direction"] == "DECRESCENTE",
                x["direction"] == "LESTE",
                x["direction"] == "LESTE/OESTE",
                x["direction"] == "NORTE/SUL",
                x["direction"] == "TRANSVERSAL",
                x["direction"] == "OESTE",
                x["direction"],
            ),
            reverse=True,
        )

        self.fill_sheet(data_list=sorted_data)

        road_name = self.reportings.first().road_name

        file_name = "Ficha de Monitoração - {} - Telas Antiofuscantes".format(road_name)

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        result = f"/tmp/{file_name}.xlsx"
        self.wb.save(result)
        return result


class CCRAntiGlareScreens(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        qs_reporting = Reporting.objects.filter(pk__in=self.uuids)
        road_name = ""
        list_inspection_campaign_year = ""

        road_name = (" ").join(
            [
                re.sub(r"[- ]", "", road)
                for road in qs_reporting.order_by("road_name")
                .distinct("road_name")
                .values_list("road_name", flat=True)
            ]
        )
        list_inspection_campaign_year = list(
            set(
                [
                    str(reporting.form_data.get("inspection_campaign_year", ""))
                    for reporting in qs_reporting
                    if str(reporting.form_data.get("inspection_campaign_year", ""))
                ]
            )
        )

        if list_inspection_campaign_year:
            list_inspection_campaign_year.sort()

        inspection_campaign_year = (" ").join(list_inspection_campaign_year)

        file_name = "Relatórios ANTT de Tela Antiofuscante - {} - {}".format(
            road_name, inspection_campaign_year
        )

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = list()
        files.append(
            AntiGlareScreenXlsxHandler(s3, self.uuids, self.sheet_target()).execute()
        )
        files.append(
            XlsxHandlerResumeReportAntiglareScreens(
                s3, self.uuids, self.sheet_target()
            ).execute()
        )

        description_measures = (
            XlsxHandlerResumeReportDescriptionMeasuresAntGlareScreens(
                uuid=self.uuids[0],
                list_uuids=self.uuids,
                s3=s3,
                sheet_target=self.sheet_target(),
                report_format=self.report_format(),
            ).execute()
        )

        if description_measures:
            files.extend(description_measures)

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_antiglare_screen_async_handler(reporter_dict: dict):
    reporter = CCRAntiGlareScreens.from_dict(reporter_dict)
    reporter.export()
