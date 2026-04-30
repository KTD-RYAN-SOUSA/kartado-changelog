import shutil
import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Firm, Reporting, ReportingFile
from helpers.apps.ccr_embankments_retaining_structures import get_form_data
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_km_plus_meter,
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_picture,
    download_reporting_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_img,
    insert_logo_and_provider_logo,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_custom_option,
    get_km,
    get_previous_campaign_report,
)


class XlsxHandler:
    DEFAULT_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(
        self,
        s3,
        list_reporting: List[Reporting],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.__sheet_target = sheet_target
        self.list_reporting = list_reporting
        self.__xlsx_file = (
            "./fixtures/reports/ccr_report_device_horizontal_signage.xlsx"
        )
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Dispositivos"]

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="S1:U3",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B3",
            resize_method=ResizeMethod.ProportionalLeft,
        )

    @classmethod
    def __average_device(cls, values: list):
        values_int = []
        for value in values:
            try:
                values_int.append(float(value))
            except Exception:
                pass
        if values_int:
            _sum = sum(values_int)
            _min = min(values_int)
            _max = max(values_int)
            return (_sum - _min - _max) / 8
        return ""

    @classmethod
    def __func_border(cls, sheet: Worksheet, row: int, row_end: int) -> None:
        for _row in sheet.iter_rows(min_row=row, max_row=row_end):
            for cell in _row:
                try:
                    if cell.column_letter == "V":
                        break
                except AttributeError:
                    pass
                cell.border = XlsxHandler.DEFAULT_BORDER

    def fill_sheet(self, data_list: list):
        data_dict = {}
        order_dict = {}
        for data in data_list:
            __key = f"{data['road_name']} - {data['km']} - {data['end_km']}"
            if __key not in data_dict.keys():
                data_dict[__key] = [data]
                order_dict[__key] = [data["direction"]]
            else:
                data_dict[__key].append(data)
                order_dict[__key].append(data["direction"])

        for k, v in order_dict.items():
            __direction = list(set(v))
            if len(__direction) == 1:
                if str(__direction[0]).lower() == "sul":
                    new_list = sorted(data_dict[k], key=lambda x: (x["rame"], x["km"]))
                    data_dict[k] = new_list
                elif str(__direction[0]).lower() == "norte":
                    new_list = sorted(data_dict[k], key=lambda x: (x["rame"], -x["km"]))
                    data_dict[k] = new_list
                else:
                    new_list = sorted(
                        data_dict[k], key=lambda x: (x["rame"], x["direction"], x["km"])
                    )
                    data_dict[k] = new_list
            else:
                new_list = sorted(
                    data_dict[k], key=lambda x: (x["rame"], x["direction"], x["km"])
                )
                data_dict[k] = new_list

        bold = Font(bold=True)
        files = []
        # Para cada dispositivo (acho que vai sempre ter apenas um por Excel)
        for __key, datalist in data_dict.items():
            init_row = 8
            direction = []
            lanes = []
            all_roads = []
            team_list = []
            subcompany_list = []
            company_list = []
            dates = []
            km = ""
            km_init = ""
            croqui = ""
            # Para cada ramo do dispositivo
            for data in datalist:
                team_list.append(data["team"])
                subcompany_list.append(data["subcompany"])
                company_list.append(data["company"])
                if not croqui:
                    croqui = data["croqui"]
                if data["executed_at"]:
                    dates.append(data["executed_at"])
                else:
                    dates.append(data["found_at"])
                if data["km_init"]:
                    km_init = data["km_init"]
                if data.get("road_name") not in all_roads:
                    all_roads.append(data.get("road_name"))
                if data["lane"] not in lanes:
                    lanes.append(data["lane"].replace("/", ""))
                if data["direction"] not in direction:
                    direction.append(data["direction"].replace("/", ""))

                len_station_one = (
                    len(data["station_one"])
                    if len(data["station_one"]) > 3 or len(data["station_one"]) == 0
                    else 3
                )
                len_station_two = (
                    len(data["station_two"])
                    if len(data["station_two"]) > 3 or len(data["station_two"]) == 0
                    else 3
                )
                len_station_three = (
                    len(data["station_three"])
                    if len(data["station_three"]) > 3 or len(data["station_three"]) == 0
                    else 3
                )

                max_rows = len_station_one + len_station_two + len_station_three
                count = 0
                if len_station_one:
                    count += 1
                if len_station_two:
                    count += 1
                if len_station_three:
                    count += 1
                add = max_rows if max_rows >= (count * 3) else count * 3
                end_row = add + init_row

                if add > 0:
                    self._worksheet.merge_cells(f"A{init_row}:A{init_row + add-1}")
                    self._worksheet.merge_cells(f"I{init_row}:I{init_row + add-1}")

                # Preenche nome do ramo
                self._worksheet[f"A{init_row}"].alignment = Alignment(
                    textRotation=90, horizontal="center", vertical="center"
                )
                self._worksheet[f"A{init_row}"].font = bold
                self._worksheet[f"A{init_row}"].value = data["rame"]

                # Preenche observações
                insert_centered_value(
                    self._worksheet,
                    value=data["notes"],
                    cell=f"I{init_row}",
                    wrapText=True,
                )

                # Mescla células de Estação, Lat, Lng, Foto, Data da medição, e Observações, para cada uma das estações
                for intern_col in ["B", "C", "D", "G", "H", "I"]:
                    if len_station_one > 1:
                        num = init_row
                        add = len_station_one if len_station_one >= 3 else 3
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                    if len_station_two > 1:
                        num = init_row + len_station_one
                        add = len_station_two if len_station_two >= 3 else 3
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                    if len_station_three > 1:
                        num = init_row + len_station_one + len_station_two
                        add = len_station_three if len_station_three >= 3 else 3
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )

                offset = 0
                # Para cada estação do ramo
                for station in ["one", "two", "three"]:
                    len_station = 0
                    if station == "one":
                        len_station = len_station_one
                    if station == "two":
                        len_station = len_station_two
                    if station == "three":
                        len_station = len_station_three

                    if len_station == 0:
                        continue

                    # insere km da estação
                    insert_centered_value(
                        self._worksheet,
                        value=data["km_station_" + station],
                        cell=f"B{init_row + offset}",
                        wrapText=True,
                    )
                    self._worksheet[f"B{init_row + offset}"].font = bold

                    insert_centered_value(
                        self._worksheet,
                        value=data["lat_station_" + station],
                        cell=f"C{init_row + offset}",
                        number_format="General",
                        wrapText=True,
                    )
                    insert_centered_value(
                        self._worksheet,
                        value=data["long_station_" + station],
                        cell=f"D{init_row + offset}",
                        number_format="General",
                        wrapText=True,
                    )

                    images: dict = data.get("images")
                    if images.get("station_" + station):
                        range_str = f"G{init_row + offset}:G{init_row + offset + 2}"
                        insert_picture(
                            self._worksheet,
                            range_str,
                            Image(images.get("station_" + station)),
                            self.__sheet_target,
                        )

                    insert_centered_value(
                        self._worksheet,
                        value=data["executed_at"],
                        cell=f"H{init_row + offset}",
                        wrapText=True,
                    )

                    if len_station:
                        XlsxHandler.__func_border(
                            sheet=self._worksheet,
                            row=init_row + offset,
                            row_end=init_row + offset + len_station - 1,
                        )

                        init = init_row + offset
                        for item in data["station_" + station]:
                            insert_centered_value(
                                self._worksheet,
                                value=item["sinalization_lane"],
                                cell=f"E{init}",
                                bold=True,
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["station_color"],
                                cell=f"F{init}",
                                bold=True,
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_one"],
                                cell=f"J{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_two"],
                                cell=f"K{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_three"],
                                cell=f"L{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_four"],
                                cell=f"M{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_five"],
                                cell=f"N{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_six"],
                                cell=f"O{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_seven"],
                                cell=f"P{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_eight"],
                                cell=f"Q{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_nine"],
                                cell=f"R{init}",
                                wrapText=True,
                            )
                            insert_centered_value(
                                self._worksheet,
                                value=item["measure_ten"],
                                cell=f"S{init}",
                                wrapText=True,
                            )
                            __average = ""
                            if not item["average"]:
                                __average = XlsxHandler.__average_device(
                                    values=[
                                        item["measure_one"],
                                        item["measure_two"],
                                        item["measure_three"],
                                        item["measure_four"],
                                        item["measure_five"],
                                        item["measure_six"],
                                        item["measure_seven"],
                                        item["measure_eight"],
                                        item["measure_nine"],
                                        item["measure_ten"],
                                    ]
                                )
                            insert_centered_value(
                                self._worksheet,
                                value=__average,
                                cell=f"T{init}",
                                wrapText=True,
                            )
                            __font_bool = False
                            if __average and item["minimal_value"]:
                                try:
                                    __font_bool = (
                                        False
                                        if __average > float(item["minimal_value"])
                                        else True
                                    )
                                except ValueError:
                                    __font_bool = False
                            self._worksheet[f"T{init}"].font = Font(
                                color="ff0000" if __font_bool else "000000",
                                bold=True,
                            )
                            if data["current_average_last_year"].get("station_one"):
                                _data_ = data["current_average_last_year"].get(
                                    "station_one"
                                )
                                for _itern_data in _data_:
                                    _average = _itern_data.get(
                                        f"{item['sinalization_lane']}-{item['station_color']}"
                                    )
                                    if _average:
                                        break
                                insert_centered_value(
                                    self._worksheet,
                                    value=_average if _average else "",
                                    cell=f"U{init}",
                                    wrapText=True,
                                )
                            else:
                                self._worksheet[f"U{init}"].value = ""
                            init += 1

                    offset += len_station

                signal_strip = {}
                if data["left_border"] > 0:
                    signal_strip["BORDO ESQUERDO"] = data["left_border"]
                if data["axis"] > 0:
                    signal_strip["EIXO"] = data["axis"]
                if data["axis_two"] > 0:
                    signal_strip["EIXO 2"] = data["axis_two"]
                if data["axis_three"] > 0:
                    signal_strip["EIXO 3"] = data["axis_three"]
                if data["axis_four"] > 0:
                    signal_strip["EIXO 4"] = data["axis_four"]
                if data["right_border"] > 0:
                    signal_strip["BORDO DIREITO"] = data["right_border"]
                for signal, average in signal_strip.items():
                    self._worksheet.merge_cells(f"R{end_row}:S{end_row}")
                    self._worksheet.row_dimensions[end_row].height = 19.5
                    self._worksheet[f"R{end_row}"] = signal
                    self._worksheet[f"T{end_row}"] = average
                    self._worksheet[f"U{end_row}"] = (
                        data["current_average_last_year_location"].get(signal) or ""
                    )
                    color = "CCFFFF" if "BORDO" in signal else "FFFF99"
                    for col in ["R", "S", "T", "U"]:
                        self._worksheet[f"{col}{end_row}"].font = bold
                        self._worksheet[f"{col}{end_row}"].fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
                        self._worksheet[f"{col}{end_row}"].border = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000"),
                        )
                        self._worksheet[f"{col}{end_row}"].alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    end_row += 1

                for __row in range(init_row, end_row):
                    self._worksheet.row_dimensions[__row].height = 21.75

                init_row = end_row

            # Insere borda das células do croqui
            for __row in self._worksheet.iter_rows(
                min_row=init_row, max_row=init_row + 20
            ):
                self._worksheet.row_dimensions[__row[0].row].height = 21.75
                for cell in __row:
                    try:
                        if cell.column_letter == "V":
                            break
                    except AttributeError:
                        pass
                    if cell.row == init_row:
                        cell.border = Border(
                            top=Side(style="thin"),
                        )
                    elif cell.row == init_row + 20:
                        cell.border = Border(
                            bottom=Side(style="thin"),
                        )
                    if cell.column_letter == "U":
                        cell.border = Border(
                            right=Side(style="thin"),
                        )
                    if cell.row == init_row and cell.column_letter == "U":
                        cell.border = Border(
                            top=Side(style="thin"),
                            right=Side(style="thin"),
                        )
                    elif cell.row == init_row + 20 and cell.column_letter == "U":
                        cell.border = Border(
                            bottom=Side(style="thin"),
                            right=Side(style="thin"),
                        )

            if croqui:
                insert_img(
                    worksheet=self._worksheet,
                    image=croqui,
                    row_init=init_row + 2,
                    row_end=init_row + 2,
                    col_init=5,
                    col_end=5,
                    width=1074,
                    height=530,
                )
            team_list = list(set(team_list))
            query_set_teams = Firm.objects.filter(uuid__in=team_list).all()
            names = []
            for team in query_set_teams:
                users_query_set = team.users.all()
                intern_list = []
                for user in users_query_set:
                    intern_list.append(user.get_full_name())
                names.append(intern_list)
            names = [", ".join(_) for _ in names]
            team_text = f"Operador: {' - '.join(names)}"
            subcompany_text = f"Empresa: {' - '.join(list(set(subcompany_list)))}"
            company_text = f"Concessionária: {' - '.join(list(set(company_list)))}"
            road_text = f"Rodovia: {data['road_name']}"
            reporting_text = f"Monitoração: Sinalização Horizontal - Dispositivos {km}"
            if len(dates) > 1:
                dates_new = sorted(
                    dates,
                    key=lambda x: (x.split("/")[2], x.split("/")[1], x.split("/")[0]),
                )
                date_text = f"Data: {dates_new[0]} à {dates_new[-1]}"
            elif dates:
                date_text = f"Data: {dates[0]}"
            else:
                date_text = "Data: "
            self._worksheet["F1"] = subcompany_text
            self._worksheet["F2"] = team_text
            self._worksheet["F3"] = date_text
            self._worksheet["L1"] = road_text
            self._worksheet["L2"] = reporting_text
            self._worksheet["L3"] = company_text

            for row in self._worksheet.iter_rows(min_row=6):
                cell: Cell = row[0]
                curr_border: Border = cell.border
                cell.border = Border(
                    left=Side(style="thin"),
                    right=curr_border.right,
                    top=curr_border.top,
                    bottom=curr_border.bottom,
                )

            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
            )
            road_name: str = data["road_name"]
            __filename = f"Dispositivo {km_init} {road_name.replace('/', '-')}.xlsx"
            result = f"/tmp/{__filename}"
            files.append(result)

            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook["Dispositivos"]
        return {"files": files, "all_roads": all_roads}

    def __find_station(self, reporting, station: list, position: int) -> tuple:
        datalist = []
        data_id = {}
        if station:
            for obj in station:
                first_station = obj
                if position == 1:
                    _id = "1"
                    pos = ""
                    ordinal = ""
                elif position == 2:
                    _id = "2"
                    pos = "_two"
                    ordinal = "_second"
                else:
                    _id = "3"
                    pos = "_three"
                    ordinal = "_third"
                if first_station.get(f"sinalization_lane{pos}"):
                    minimal_value_list = [
                        __value
                        for __key, __value in first_station.items()
                        if "minimal_value" in __key
                    ]
                    minimal_value = (
                        "" if not minimal_value_list else minimal_value_list[0]
                    )
                    sinalization_lane = get_form_data(
                        reporting,
                        data_name="sinalizationLane",
                        subgroup="stationOne",
                        field_name="sinalization_lane",
                        value=first_station.get(f"sinalization_lane{pos}"),
                    )
                    station_color = get_form_data(
                        reporting,
                        data_name="stationColor",
                        subgroup="stationOne",
                        field_name=f"station_color{pos}",
                        value=first_station.get(f"station_color{pos}"),
                    )
                    data_id[
                        f"{first_station.get(f'sinalization_lane{pos}')}-{first_station.get(f'station_color{pos}')}"
                    ] = f"{sinalization_lane}-{station_color}"
                    datalist.append(
                        {
                            "sinalization_lane": sinalization_lane,
                            "station_color": station_color,
                            "id": _id,
                            "station_image": first_station.get(
                                f"station_image{ordinal}"
                            ),
                            "measure_one": first_station.get(f"measure_one{ordinal}"),
                            "measure_two": first_station.get(f"measure_two{ordinal}"),
                            "measure_three": first_station.get(
                                f"measure_three{ordinal}"
                            ),
                            "measure_four": first_station.get(f"measure_four{ordinal}"),
                            "measure_five": first_station.get(f"measure_five{ordinal}"),
                            "measure_six": first_station.get(f"measure_six{ordinal}"),
                            "measure_seven": first_station.get(
                                f"measure_seven{ordinal}"
                            ),
                            "measure_eight": first_station.get(
                                f"measure_eight{ordinal}"
                            ),
                            "measure_nine": first_station.get(f"measure_nine{ordinal}"),
                            "measure_ten": first_station.get(f"measure_ten{ordinal}"),
                            "average": first_station.get(f"average_station{ordinal}"),
                            "images": first_station.get(f"station_image{pos}"),
                            "attendance_status": str(
                                first_station.get(f"attendance_status{ordinal}")
                            ),
                            "minimal_value": minimal_value,
                        }
                    )
        return datalist, data_id

    def custom_key(self, item):
        if item["sinalization_lane"] == "BORDO ESQUERDO":
            return (0, item["sinalization_lane"])
        elif item["sinalization_lane"] == "BORDO DIREITO":
            return (2, item["sinalization_lane"])
        else:
            return (1, item["sinalization_lane"])

    def is_dict_empty(self, item):
        return 1 if not item else 0

    def create_dict(self, reporting: Reporting, s3) -> dict:
        data_id = {}
        station_one, data_id_one = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_one"),
            position=1,
        )
        data_id.update(data_id_one)
        station_two, data_id_two = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_two"),
            position=2,
        )
        data_id.update(data_id_two)
        station_three, data_id_three = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_three"),
            position=3,
        )
        data_id.update(data_id_three)
        station_one = sorted(station_one, key=self.custom_key)
        station_two = sorted(station_two, key=self.custom_key)
        station_three = sorted(station_three, key=self.custom_key)
        empty_obj = {
            "sinalization_lane": "",
            "station_color": "",
            "station_image": "",
            "measure_one": "",
            "measure_two": "",
            "measure_three": "",
            "measure_four": "",
            "measure_five": "",
            "measure_six": "",
            "measure_seven": "",
            "id": "",
            "measure_eight": "",
            "measure_nine": "",
            "measure_ten": "",
            "average": "",
            "images": "",
            "attendance_status": "",
        }
        if 0 < len(station_one) < 3:
            station_one.insert(1, empty_obj)
        if 0 < len(station_two) < 3:
            station_two.insert(1, empty_obj)
        if 0 < len(station_three) < 3:
            station_three.insert(1, empty_obj)
        allocation = [station_one, station_two, station_three]
        allocation = sorted(allocation, key=lambda x: self.is_dict_empty(x))
        station_one, station_two, station_three = (
            allocation[0],
            allocation[1],
            allocation[2],
        )
        keys = {
            "BORDO DIREITO": "right_border",
            "BORDO ESQUERDO": "left_border",
            "EIXO": "axis",
            "EIXO 2": "axis_two",
            "EIXO 3": "axis_three",
            "EIXO 4": "axis_four",
        }
        keys_list = []
        if station_one:
            for _lane in station_one:
                if _lane.get("sinalization_lane") in keys.keys():
                    keys_list.append(keys[_lane.get("sinalization_lane")])
        if station_two:
            for _lane in station_two:
                if _lane.get("sinalization_lane") in keys.keys():
                    keys_list.append(keys[_lane.get("sinalization_lane")])
        if station_three:
            for _lane in station_three:
                if _lane.get("sinalization_lane") in keys.keys():
                    keys_list.append(keys[_lane.get("sinalization_lane")])

        right_border = 0
        left_border = 0
        axis = 0
        axis_two = 0
        axis_three = 0
        axis_four = 0
        if keys_list:
            for __key in keys_list:
                if __key == "right_border":
                    right_border = reporting.form_data.get("right_border")
                elif __key == "left_border":
                    left_border = reporting.form_data.get("left_border")
                elif __key == "axis":
                    axis = reporting.form_data.get("axis")
                elif __key == "axis_two":
                    axis_two = reporting.form_data.get("axis_two")
                elif __key == "axis_three":
                    axis_three = reporting.form_data.get("axis_three")
                elif __key == "axis_four":
                    axis_four = reporting.form_data.get("axis_four")

        reporting_files = download_reporting_pictures(
            s3=s3,
            path=self.temp_file,
            reporting=reporting,
            width=337,
            height=242,
            enable_include_dnit=False,
            enable_is_shared_antt=True,
        )
        images = {"station_one": "", "station_two": "", "station_three": ""}

        station_pictures_one = reporting.form_data.get("station_pictures_one")
        station_pictures_two = reporting.form_data.get("station_pictures_two")
        station_pictures_three = reporting.form_data.get("station_pictures_three")

        if (
            station_pictures_one
            and len(station_pictures_one) > 0
            and ("foto_da_estacao_um" in station_pictures_one[0])
            and len(station_pictures_one[0].get("foto_da_estacao_um")) > 0
        ):
            for reporting_file_uuid in station_pictures_one[0].get(
                "foto_da_estacao_um"
            ):
                try:
                    path = download_picture(
                        self.s3,
                        self.temp_file,
                        reporting_file_uuid,
                        reporting_file_uuid,
                        enable_is_shared=True,
                    )
                    if path:
                        images["station_one"] = path
                        break
                except Exception:
                    continue
        if (
            station_pictures_two
            and len(station_pictures_two) > 0
            and ("foto_da_estacao_dois" in station_pictures_two[0])
            and len(station_pictures_two[0].get("foto_da_estacao_dois")) > 0
        ):
            for reporting_file_uuid in station_pictures_two[0].get(
                "foto_da_estacao_dois"
            ):
                try:
                    path = download_picture(
                        self.s3,
                        self.temp_file,
                        reporting_file_uuid,
                        reporting_file_uuid,
                        enable_is_shared=True,
                    )
                    if path:
                        images["station_two"] = path
                        break
                except Exception:
                    continue
        if (
            station_pictures_three
            and len(station_pictures_three) > 0
            and ("foto_da_estacao_tres" in station_pictures_three[0])
            and len(station_pictures_three[0].get("foto_da_estacao_tres")) > 0
        ):
            for reporting_file_uuid in station_pictures_three[0].get(
                "foto_da_estacao_tres"
            ):
                try:
                    path = download_picture(
                        self.s3,
                        self.temp_file,
                        reporting_file_uuid,
                        reporting_file_uuid,
                        enable_is_shared=True,
                    )
                    if path:
                        images["station_three"] = path
                        break
                except Exception:
                    continue
        croqui_list = reporting.form_data.get("croqui")
        croqui_uuid = ""
        croqui = ""
        if croqui_list:
            croqui_dict: dict = croqui_list[0]
            croqui_uuid_list = croqui_dict.get("croqui_image")
            if croqui_uuid_list:
                for reporting_file_uuid in croqui_uuid_list:
                    if ReportingFile.objects.filter(
                        uuid=reporting_file_uuid, is_shared=True
                    ).exists():
                        croqui_uuid = reporting_file_uuid
                        break
        for obj in reporting_files["images"]:
            for k, v in images.items():
                if obj["uuid"] in v:
                    images[k] = obj["path"]
            if obj["uuid"] == croqui_uuid:
                croqui = obj["path"]
        if not reporting_files["images"]:
            images = {"station_one": "", "station_two": "", "station_three": ""}
        direction = get_custom_option(reporting, "direction")

        current_average_last_year = {}
        current_average_last_year_location = {}
        if reporting.parent:

            obj = get_previous_campaign_report(
                reporting.occurrence_type, reporting, "form_data"
            )
            if not (obj is None) and obj.form_data:
                current_average_last_year_location = {
                    "BORDO DIREITO": obj.form_data.get("right_border"),
                    "BORDO ESQUERDO": obj.form_data.get("left_border"),
                    "EIXO": obj.form_data.get("axis"),
                    "EIXO 2": obj.form_data.get("axis_two"),
                    "EIXO 3": obj.form_data.get("axis_three"),
                    "EIXO 4": obj.form_data.get("axis_four"),
                }
                for _key, _value in obj.form_data.items():
                    __list_measure = []
                    if type(_value) is list and "station" in _key:
                        for item in _value:
                            try:
                                sinalization_lane = next(
                                    y
                                    for x, y in item.items()
                                    if "sinalization_lane" in x
                                )
                                station_color = next(
                                    y for x, y in item.items() if "station_color" in x
                                )
                                _key_name = (
                                    data_id.get(f"{sinalization_lane}-{station_color}")
                                    if data_id.get(
                                        f"{sinalization_lane}-{station_color}"
                                    )
                                    else f"{sinalization_lane}-{station_color}"
                                )
                                __list_measure.append(
                                    {
                                        _key_name: XlsxHandler.__average_device(
                                            values=[
                                                _v
                                                for _k, _v in item.items()
                                                if "measure" in _k
                                            ]
                                        )
                                    }
                                )
                            except StopIteration:
                                continue
                        current_average_last_year[_key] = __list_measure

        allocation_two = [
            current_average_last_year.get("station_one"),
            current_average_last_year.get("station_two"),
            current_average_last_year.get("station_three"),
        ]
        allocation_two = sorted(allocation_two, key=lambda x: self.is_dict_empty(x))
        current_average_last_year = {
            "station_one": allocation_two[0] or {},
            "station_two": allocation_two[1] or {},
            "station_three": allocation_two[2] or {},
        }

        km_station_one = get_km_plus_meter(km=reporting.form_data.get("km_station"))
        km_station_two = get_km_plus_meter(km=reporting.form_data.get("km_station_two"))
        km_station_three = get_km_plus_meter(
            km=reporting.form_data.get("km_station_three")
        )
        lat_station = reporting.form_data.get("lat_station")
        long_station = reporting.form_data.get("long_station")
        lat_station_two = reporting.form_data.get("lat_station_two")
        long_station_two = reporting.form_data.get("long_station_two")
        lat_station_three = reporting.form_data.get("lat_station_three")
        long_station_three = reporting.form_data.get("long_station_three")

        obj_ord = {
            "0": {"km": "", "lat": "", "long": "", "notes": ""},
            "1": {
                "km": km_station_one,
                "lat": lat_station,
                "long": long_station,
                # "notes": notes_one,
            },
            "2": {
                "km": km_station_two,
                "lat": lat_station_two,
                "long": long_station_two,
                # "notes": notes_station_two,
            },
            "3": {
                "km": km_station_three,
                "lat": lat_station_three,
                "long": long_station_three,
                # "notes": notes_station_three,
            },
        }

        data = {
            "road_name": reporting.road_name,
            "subcompany": reporting.firm.subcompany.__dict__.get("name"),
            "company": reporting.company.__dict__.get("name"),
            "team": str(reporting.firm.__dict__.get("uuid")),
            "executed_at": (
                reporting.executed_at.strftime("%d/%m/%Y")
                if reporting.executed_at
                else ""
            ),
            "found_at": (
                reporting.found_at.strftime("%d/%m/%Y") if reporting.found_at else ""
            ),
            "km": reporting.km,
            "km_init": get_km_plus_meter(km=reporting.km),
            "end_km": reporting.end_km,
            "direction": direction,
            "rame": reporting.form_data.get("rame"),
            "km_station_one": (
                obj_ord[station_one[0]["id"]]["km"] if station_one else ""
            ),
            "km_station_two": (
                obj_ord[station_two[0]["id"]]["km"] if station_two else ""
            ),
            "km_station_three": (
                obj_ord[station_three[0]["id"]]["km"] if station_three else ""
            ),
            "lat_station_one": (
                obj_ord[station_one[0]["id"]]["lat"] if station_one else ""
            ),
            "lat_station_two": (
                obj_ord[station_two[0]["id"]]["lat"] if station_two else ""
            ),
            "lat_station_three": (
                obj_ord[station_three[0]["id"]]["lat"] if station_three else ""
            ),
            "long_station_one": (
                obj_ord[station_one[0]["id"]]["long"] if station_one else ""
            ),
            "long_station_two": (
                obj_ord[station_two[0]["id"]]["long"] if station_two else ""
            ),
            "long_station_three": (
                obj_ord[station_three[0]["id"]]["long"] if station_three else ""
            ),
            "station_one": station_one,
            "station_two": station_two,
            "station_three": station_three,
            "notes": reporting.form_data.get("notes_two"),
            "right_border": right_border,
            "left_border": left_border,
            "axis": axis,
            "axis_two": axis_two,
            "axis_three": axis_three,
            "axis_four": axis_four,
            "reporting": reporting,
            "lane": get_custom_option(reporting, "lane"),
            "current_average_last_year": current_average_last_year,
            "current_average_last_year_location": current_average_last_year_location,
            "images": images,
            "croqui": croqui,
            "monitoring_kind": reporting.form_data.get("monitoring_kind"),
        }
        for k, v in data.items():
            if v is None:
                data[k] = ""

        return data

    def execute(self):
        data = []
        for reporting in self.list_reporting:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo
        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRReportDeviceHorizontalSignage(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""

        reporting = Reporting.objects.get(uuid=self.uuids[0])
        reportings = Reporting.objects.filter(
            occurrence_type=reporting.occurrence_type, uuid__in=self.uuids
        ).prefetch_related("road")
        all_roads = []
        datawork = {}
        for obj in reportings:
            if obj.road.name not in all_roads:
                all_roads.append(obj.road.name)
            key = f"{obj.road_name} - {obj.km} - {obj.end_km}"
            if key not in datawork.keys():
                datawork[key] = [obj]
            else:
                datawork[key].append(obj)
        if len(datawork.keys()) == 1:
            sample_reporting = datawork[list(datawork.keys())[0]][0]
            km = get_km(sample_reporting)
            road_name = reporting.road_name.replace("/", "-")
            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            else:
                extension = "xlsx"

            file_name = f"Dispositivo {km} {road_name}.{extension}"
        else:
            all_roads.sort()
            all_roads = [road.replace("/", "-") for road in all_roads]
            file_name = f"Monitoração de Dispositivo - {'-'.join(all_roads)}.zip"
        return file_name

    def __get_repotings_obj(self):
        form = OccurrenceType.objects.get(name="Retrorrefletância de Dispositivos")
        query_set = Reporting.objects.filter(
            occurrence_type=form, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def export(self):
        list_reporting = self.__get_repotings_obj()
        s3 = get_s3()
        obj = XlsxHandler(s3, list_reporting, self.sheet_target()).execute()
        files = obj["files"]

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_device_horizontal_signage_async_handler(
    reporter_dict: dict,
):
    reporter = CCRReportDeviceHorizontalSignage.from_dict(reporter_dict)
    reporter.export()
