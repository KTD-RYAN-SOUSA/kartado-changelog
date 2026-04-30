import re
import tempfile
from datetime import datetime
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from zappa.asynchronous import task

from apps.companies.models import Firm
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
    result_photos,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.apps.ccr_resume_metallic_fenders_oae_metal_defenses_oae import (
    XlsxHandlerResumeReportMetallicFendersOAE,
)
from helpers.strings import clean_latin_string, format_km

from .ccr_resume_report_defesas_oae import XlsxHandlerResumeReportDefensesOAE


class MetalDefensesOAEXlsxHandler(object):
    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.wb = load_workbook("fixtures/reports/crr_report_metal_defenses_oae.xlsx")
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self._worksheet = self.wb.active
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company"
        )
        self.temp_file = tempfile.mkdtemp()

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="W1:Y5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B5",
            resize_method=ResizeMethod.ProportionalLeft,
        )

        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company

        self.static_fields = {
            "year": "A",
            "index": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "latitude": "F",
            "longitude": "G",
            "notes": "H",
            "photo_start": "I",
            "photo_end": "J",
            "general_appearance": {"BOA": "L", "REGULAR": "L", "RUIM": "L"},
            "height": {"ATENDE_NORMA": "N", "NAO_ATENDE": "N"},
            "length": "O",
            "fixation": {"SOLO": "Q", "PARAFUSOS": "Q", "PONTEIRAS": "Q"},
            "oae_fixation": {"ENGASTADO": "S", "AEREO": "S"},
            "conversion": {
                "FERRUGEM": "U",
                "PINTURA_DEFICIENTE": "U",
                "AMASSADOS": "U",
            },
            "reflectors": {"AUSENTE": "W", "RUIM": "W", "REGULAR": "W", "BOM": "W"},
            "reflectors_color": {
                "tipo_refletor": "Y",
                "VERMELHO": "Y",
                "BRANCO": "Y",
                "LIMA-LIMAO": "Y",
            },
        }

    def __insert_new_rows(self, row: int):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        chars = [chr(ord("A") + i) for i in range(10)]
        second_chars = [chr(ord("P") + i) for i in range(10)]

        for char in chars + ["O"]:
            self._worksheet.merge_cells(f"{char}{row}:{char}{row + 3}")
        for count in range(0, 4):
            self._worksheet.row_dimensions[row + count].height = 20
            for char in chars + second_chars + ["K", "M", "L", "N", "O"]:
                self._worksheet[f"{char}{row + count}"].border = border

    def create_dict(self, reporting, s3=None):
        photos = reporting.form_data.get("mandatory_pictures")
        photo_start = None
        photo_end = None
        if photos:
            for photo in reporting.form_data.get("mandatory_pictures"):
                photo_start = (
                    result_photos(
                        s3=self.s3,
                        temp_file=tempfile.mkdtemp(),
                        photo_id=photo["eps_image_start"][0],
                        width=170,
                        height=100,
                        enable_is_shared_antt=True,
                        enable_include_dnit=False,
                    )
                    if photo["eps_image_start"]
                    else ""
                )
                photo_end = (
                    result_photos(
                        s3=self.s3,
                        temp_file=tempfile.mkdtemp(),
                        photo_id=photo["eps_image_end"][0],
                        width=170,
                        height=100,
                        enable_is_shared_antt=True,
                        enable_include_dnit=False,
                    )
                    if photo["eps_image_end"]
                    else ""
                )

        result_index = new_get_form_data(reporting, "idCcrAntt")

        result_year = new_get_form_data(
            reporting,
            "inspectionCampaignYear",
        )

        result_notes = new_get_form_data(reporting, "notes")

        direction = get_custom_option(reporting, "direction")

        result_type_reflector = new_get_form_data(reporting, "reflectiveKind")

        reflectors_color = self.__delineator_state(reporting)
        result_reflectors_color = {
            "tipo_refletor": (
                result_type_reflector.upper()
                if result_type_reflector in ("Monodirecional", "Bidirecional")
                else ""
            ),
            "VERMELHO": "X" if "Vermelho" in reflectors_color else "",
            "BRANCO": "X" if "Branco" in reflectors_color else "",
            "LIMA-LIMAO": "X" if "Lima-Limão" in reflectors_color else "",
        }

        reflector_state = new_get_form_data(reporting, "reflectionState")
        result_reflector_state = {
            "AUSENTE": "X" if reflector_state == "Ausente" else "",
            "BOM": "X" if reflector_state == "Bom" else "",
            "REGULAR": "X" if reflector_state == "Regular" else "",
            "RUIM": "X" if reflector_state == "Ruim" else "",
        }

        soil = new_get_form_data(reporting, "metalPost")

        pointer = new_get_form_data(reporting, "ponctualGadget")
        _fixation = {
            "SOLO": "Sim" if soil else "Não",
            "PARAFUSOS": "Sim",
            "PONTEIRAS": "Sim" if pointer in ("Tripla Onda", "Engastado") else "Não",
        }

        result_fixation_oae = new_get_form_data(reporting, "fixationOae")
        fixation_oae = {
            "ENGASTADO": "X" if result_fixation_oae == "Engastado" else "",
            "AEREO": "X" if result_fixation_oae == "Aéreo" else "",
        }
        result_deficient_paint = new_get_form_data(reporting, "defficientPainting")

        result_crumples = new_get_form_data(reporting, "crumples")

        result_rust = new_get_form_data(reporting, "rust")

        conservation = {
            "FERRUGEM": "Sim" if result_rust else "Não",
            "PINTURA_DEFICIENTE": "Sim" if result_deficient_paint else "Não",
            "AMASSADOS": "Sim" if result_crumples else "Não",
        }

        result_height = new_get_form_data(reporting, "heightCriteria")
        height = {
            "ATENDE_NORMA": "X" if result_height == "Atende" else "",
            "NAO_ATENDE": "X" if result_height == "Não Atende" else "",
        }

        result_general_appearance = new_get_form_data(reporting, "generalAppearance")

        general_appearance = {
            "BOA": "X" if result_general_appearance == "Bom" else "",
            "REGULAR": "X" if result_general_appearance == "Regular" else "",
            "RUIM": "X" if result_general_appearance == "Ruim" else "",
        }

        data = {
            "year": result_year,
            "index": result_index,
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "end_km", 3),
            "raw_initial_km": reporting.km,
            "direction": direction.upper(),
            "latitude": reporting.form_data.get("latitude"),
            "longitude": reporting.form_data.get("longitude"),
            "notes": result_notes,
            "photo_start": photo_start[0] if photo_start else "",
            "photo_end": photo_end[0] if photo_end else "",
            "general_appearance": general_appearance,
            "height": height,
            "length": reporting.form_data.get("length"),
            "fixation": _fixation,
            "oae_fixation": fixation_oae,
            "conversion": conservation,
            "reflectors": result_reflector_state,
            "reflectors_color": result_reflectors_color,
            "executed_at": reporting.executed_at,
            "team": str(reporting.firm.__dict__.get("uuid")) if reporting.firm else "",
            "road_name": reporting.__dict__.get("road_name"),
            "subcompany": (
                reporting.firm.subcompany.__dict__.get("name") if reporting.firm else ""
            ),
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""

        return data

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def __insert_status_values(self, key, cell, value):
        cell_status_string = self._worksheet[cell].offset(0, -1)
        if key == "tipo_refletor":
            cell_status_string.value = value
        else:
            cell_status_string.value = key.replace("_", " ") if key else key
            if value in "X":
                self._worksheet[cell] = "X"
            self._worksheet[cell] = value
            MetalDefensesOAEXlsxHandler.__format_fonts(
                cell=self._worksheet[cell], size=10
            )

        MetalDefensesOAEXlsxHandler.__format_fonts(
            cell=cell_status_string, size=10, horizontal="left"
        )

    def fill_sheet(self, *, data_list: list):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 7
        team_list = []
        date_list = []
        road_name_list = []
        subcompany_list = []

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            target=self.__sheet_target,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

        for values in data_list:
            for key, value in values.items():
                self.__insert_new_rows(row=row)
                if key in [
                    "general_appearance",
                    "height",
                    "fixation",
                    "oae_fixation",
                    "conversion",
                    "reflectors",
                    "reflectors_type",
                    "reflectors_color",
                ]:
                    sum_intern_rows = 0
                    for _key, _value in value.items():
                        if not sum_intern_rows >= len(_key):
                            key_value = f"{self.static_fields[key][_key]}{row + sum_intern_rows}"
                            self.__insert_status_values(
                                key=_key, cell=key_value, value=_value
                            )
                            sum_intern_rows += 1
                elif key in ["photo_start", "photo_end"] and value:
                    cell = f"{self.static_fields[key]}{row}:{self.static_fields[key]}{row+3}"
                    insert_picture_2(
                        self._worksheet,
                        cell,
                        Image(value),
                        self.__sheet_target,
                        (1, 1, 1, 1),
                        ResizeMethod.ProportionalCentered,
                    )
                elif key == "executed_at":
                    value_date = (
                        value.strftime("%Y-%m-%d")
                        if isinstance(value, datetime)
                        else ""
                    )
                    date_list.append(value_date)
                elif key == "road_name":
                    road_name_list.append(value)
                elif key == "subcompany":
                    subcompany_list.append(value)
                elif key == "team":
                    team_list.append(values["team"])
                elif key == "raw_initial_km":
                    continue

                else:
                    key_value = f"{self.static_fields[key]}{row}"
                    self._worksheet[key_value] = value
                    MetalDefensesOAEXlsxHandler.__format_fonts(
                        cell=self._worksheet[key_value], size=10
                    )

            row += 4

        filtered_team_list = set([uuid for uuid in team_list if uuid != ""])
        query_set_teams = Firm.objects.filter(uuid__in=filtered_team_list).all()
        names = []
        for team in query_set_teams:
            users_query_set = team.users.all()
            intern_list = []
            for user in users_query_set:
                intern_list.append(user.full_name)
            names.append(intern_list)
        names = [", ".join(_) for _ in names]
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(names),
            cell="I3",
            horizontal="left",
        )
        date_list.sort()
        filtered_date_list = [date for date in date_list if date != ""]
        if len(filtered_date_list) == 0:
            date_text = ""
        elif len(filtered_date_list) == 1:
            date_text = filtered_date_list[0]
        else:
            date_text = f"{filtered_date_list[0]} até {filtered_date_list[-1]}"

        insert_centered_value(
            worksheet=self._worksheet, value=date_text, cell="I4", horizontal="left"
        )

        road_name = list(set(road_name_list))
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(road_name),
            cell="N2",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value="Elementos de Proteção e Segurança - Monitoração de Defensas em OAE",
            cell="N3",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value=self.company.name,
            cell="N4",
            horizontal="left",
        )

        subcompany = list(set(subcompany_list))
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(subcompany),
            cell="I2",
            horizontal="left",
        )

        total_length = sum(
            [t.value for t in self._worksheet["O"] if isinstance(t.value, (int, float))]
        )

        reference_total_legth = self._worksheet[f"{'N'}{self._worksheet.max_row + 1}"]
        cells_for_merge = (
            f"{reference_total_legth.coordinate}:O{reference_total_legth.row}"
        )

        self._worksheet.merge_cells(cells_for_merge)
        reference_total_legth.value = total_length
        reference_total_legth.border = border
        MetalDefensesOAEXlsxHandler.__format_fonts(
            cell=reference_total_legth, size=10, bold=True
        )

        reference_total_string = reference_total_legth.offset(0, -1)
        reference_total_string.value = "TOTAL(m)"

        reference_total_string.border = border

        self._worksheet[cells_for_merge][0][1].border = border
        MetalDefensesOAEXlsxHandler.__format_fonts(
            cell=reference_total_string, size=10, bold=True
        )

    def __delineator_state(self, reporting_state):
        get_state_in_data = new_get_form_data(reporting_state, "reflectiveColor")
        if reporting_state.form_data.get("reflective_presence"):
            return get_state_in_data if get_state_in_data else ""
        return ""

    def execute(self):
        query_set = (
            Reporting.objects.filter(
                occurrence_type=self.form, uuid__in=self.list_uuids
            )
            .prefetch_related("occurrence_type", "firm", "firm__subcompany")
            .distinct()
        )

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        sorted_data = sorted(
            data,
            key=lambda x: (
                x["raw_initial_km"] if x["direction"] == "NORTE" else False,
                -x["raw_initial_km"] if x["direction"] == "SUL" else False,
                x["direction"] == "CANTEIRO CENTRAL",
                x["direction"] == "AMBOS",
                x["direction"] == "CRESCENTE",
                x["direction"] == "DECRESCENTE",
                x["direction"] == "LESTE",
                x["direction"] == "LESTE/OESTE",
                x["direction"] == "NORTE/SUL",
                x["direction"] == "TRANSVERSAL",
                x["direction"] == "OESTE",
                x["direction"],
            ),
            reverse=True,
        )

        self.fill_sheet(data_list=sorted_data)

        road_name = self.reportings.first().road_name
        file_name = "Ficha de Monitoração - {} - Defensas em OAE".format(road_name)

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        result = f"/tmp/{file_name}.xlsx"
        self.wb.save(result)
        return result


class CCRMetalDefensesOAE(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        qs_reporting = Reporting.objects.filter(pk__in=self.uuids)
        road_name = ""
        list_inspection_campaign_year = ""

        road_name = (" ").join(
            [
                re.sub(r"[- ]", "", road)
                for road in qs_reporting.order_by("road_name")
                .distinct("road_name")
                .values_list("road_name", flat=True)
            ]
        )
        list_inspection_campaign_year = list(
            set(
                [
                    str(reporting.form_data.get("inspection_campaign_year", ""))
                    for reporting in qs_reporting
                    if str(reporting.form_data.get("inspection_campaign_year", ""))
                ]
            )
        )

        if list_inspection_campaign_year:
            list_inspection_campaign_year.sort()

        inspection_campaign_year = (" ").join(list_inspection_campaign_year)

        file_name = "Relatórios ANTT de Defensas em OAE - {} - {}".format(
            road_name, inspection_campaign_year
        )

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = list()
        files.append(
            MetalDefensesOAEXlsxHandler(
                list_uuids=self.uuids,
                s3=s3,
                sheet_target=self.sheet_target(),
            ).execute()
        )
        files.append(
            XlsxHandlerResumeReportDefensesOAE(
                list_uuids=self.uuids,
                s3=s3,
                sheet_target=self.sheet_target(),
            ).execute()
        )

        metallic_fenders_oae = XlsxHandlerResumeReportMetallicFendersOAE(
            uuid=self.uuids[0],
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        for _result in metallic_fenders_oae:
            files.append(_result)

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_metal_defenses_oae_async_handler(reporter_dict: dict):
    reporter = CCRMetalDefensesOAE.from_dict(reporter_dict)
    reporter.export()
