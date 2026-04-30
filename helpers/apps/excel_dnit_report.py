import logging
import os
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from datetime import datetime
from io import BytesIO
from math import ceil
from tempfile import mkdtemp
from textwrap import wrap
from typing import Dict, List, Tuple, Union
from unicodedata import normalize
from urllib.parse import unquote

import boto3
import botocore.config
import sentry_sdk
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as ImagePIL
from PIL import ImageDraw, ImageFont, ImageOps
from zappa.asynchronous import task

from apps.companies.models import Company
from apps.reportings.models import Reporting, ReportingFile
from apps.templates.models import ExcelDnitReport
from helpers.apps.ccr_report_utils.export_utils import format_km
from helpers.apps.ccr_report_utils.image import (
    ResizeMethod,
    SheetTarget,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.workbook_utils import set_active_cell, set_zoom
from helpers.dates import utc_to_local
from helpers.images import build_text_dict
from helpers.strings import (
    check_image_file,
    clean_latin_string,
    get_obj_from_path,
    to_snake_case,
)
from RoadLabsAPI.settings import credentials


def download_picture(
    s3,
    dir: str,
    rf: ReportingFile,
    quality: str = None,
) -> str:
    """
    Downloads a picture and returns the image local absolute path

    :param s3: boto s3 client instance
    :param dir: directory to download picture
    :param quality: possible values: None (original size), 400px, 1000px
    :returns: image local absolute path
    """
    try:
        upload = rf.upload
        file_path = upload.url.split("?")[0].split(".com/")[1]
        bucket_name = upload.url.split(".s3")[0].split("/")[-1]
        image_format = file_path.split(".")[-1]
    except Exception:
        return None

    if not check_image_file(file_path):
        return None

    image_path = "{}/{}.{}".format(dir, str(rf.uuid), image_format)
    downloaded = False
    if quality is not None:
        try:
            s3.download_file(
                bucket_name + "-" + quality, unquote(file_path), image_path
            )
            downloaded = True
        except Exception:
            pass

    if not downloaded:
        try:
            s3.download_file(bucket_name, unquote(file_path), image_path)
            downloaded = True
        except Exception:
            pass

    if downloaded:
        return image_path
    return None


def get_size(font, in_string):
    left, top, right, bottom = font.getbbox(in_string)
    width = right - left
    height = bottom - top
    return (width, height)


def watermark_image_with_text(
    filename,
    text_list: List[Tuple[str, str]],
    font_size="medium",
    fontfamily="assets/fonts/Roboto-Regular.ttf",
) -> Union[str, None]:
    if not text_list:
        return None

    FOREGROUND = (255, 255, 255)
    KEEP_lINE_LENGTH = ["none0", "none1", "none2", "Sentido", "Serial", "Status"]

    # Convert to png
    try:
        if filename.split(".")[-1] != "png":
            img = ImagePIL.open(filename)
            try:
                img = ImageOps.exif_transpose(img)
            except Exception:
                pass
            new_filename = ".".join(filename.split(".")[:-1] + ["png"])
            img.save(new_filename)
            img.close()
            os.remove(filename)
            filename = new_filename
    except Exception:
        return None

    # open image
    try:
        image = ImagePIL.open(filename).convert("RGBA")
    except Exception:
        return None
    # properties in pixels
    width, height = image.size

    if width <= 1000:
        new_width = 1000
        aspect_ratio = height / width
        new_height = int(new_width * aspect_ratio)
        image = image.resize((new_width, new_height), ImagePIL.LANCZOS)
        width, height = image.size

    margin = 5

    # build text list
    max_lines_for_item = 5
    max_height = int(height * (2 / 3))
    max_text_width = int(width / 3)
    # font size
    if font_size == "large":
        max_text_height = int(height / 20)
    elif font_size == "small":
        max_text_height = int(height / 40)
    else:
        max_text_height = int(height / 30)

    font = ImageFont.truetype(fontfamily, max_text_height)
    text = []
    for key, value in text_list:
        if "none" in key:
            text_str = value
        else:
            text_str = "{}: {}".format(key, value)

        if key in KEEP_lINE_LENGTH:
            text.append(text_str)
            max_text_width = max(get_size(font, text_str)[0], max_text_width)
        else:
            # get_size returns in pixels
            if get_size(font, text_str)[0] > max_text_width:
                pixels_per_carac = int((get_size(font, text_str)[0]) / len(text_str))
                width = int(max_text_width / pixels_per_carac)

                # "gambiarra" to not break first words
                check_str = text_str.split(" ")
                if (len(check_str) > 1) and (
                    (len(check_str[0]) + len(check_str[1]) + 1) > width
                ):
                    text.append(check_str.pop(0))
                    text_str = " ".join(check_str)

                values = wrap(
                    text_str,
                    width=width,
                    break_long_words=False,
                    max_lines=max_lines_for_item,
                )
                for item in values:
                    text.append(item)
            else:
                text.append(text_str)

        # text cannot be bigger than max_height
        if (len(text) * max_text_height) > max_height:
            break

    # get coordinates to draw the retangle
    x_list = []
    y_list = []
    for line in text:
        x, y = get_size(font, line)
        x_list.append(x)
        y_list.append(y)

    shape_x0 = 0
    shape_y0 = height - (max(y_list) * len(y_list)) - margin
    shape_x1 = max(x_list) + int(2 * margin)
    shape_y1 = height - int(margin / 2)

    shape = [(shape_x0, shape_y0), (shape_x1, shape_y1)]

    # draw rectangle
    overlay = ImagePIL.new("RGBA", image.size, FOREGROUND + (0,))
    draw = ImageDraw.Draw(overlay)
    draw.rectangle(shape, fill=FOREGROUND + (100,))
    image = ImagePIL.alpha_composite(image, overlay)
    del draw

    # write lines inside rectangle
    draw = ImageDraw.Draw(image, mode="RGBA")
    y = max(y_list)
    for i, line in enumerate(text[::-1]):
        line = normalize("NFC", line)
        draw.text(
            (margin, (height - margin) - int((i + 1) * y)),
            line,
            font=font,
            fill="black",
        )
    del draw

    # save image
    background = ImagePIL.new("RGB", image.size, (255, 255, 255))
    background.paste(image, mask=image.split()[3])
    image.close()
    del image
    os.remove(filename)
    watermarked_file_name = ".".join(filename.split(".")[:-1] + ["jpg"])
    background.save(watermarked_file_name, "JPEG")
    background.close()

    return watermarked_file_name


class ExcelDnitReportExport:
    TEMPLATE_PATH = "templates/reports/template_dnit_report.xlsm"
    FONT_SIZE = {"small": 10, "medium": 16, "large": 24}
    TEMPLATE_SHEET_NAME = "template"

    HEADER_HEIGHT = 2

    PAGE_COLS = 11
    PAGE_ROWS = 23
    PAGE_COL_OFFSET = 12
    IMAGE_COORDS = [(2, 9), (6, 9), (2, 14), (6, 14), (2, 19), (6, 19)]
    IMAGE_COL_COVER = 4

    IMAGE_INFO_ROW_OFFSET = 2
    KM_COL_OFFSET = 0
    KIND_COL_OFFSET = 1
    OCC_TYPE_COL_OFFSET = 2

    UNAVAILABLE_KM = "INDISPONÍVEL"
    UNAVAILABLE_KIND = ""
    UNAVAILABLE_OCC_TYPE = "ERRO"

    PICS_THREADING_LIMIT = 30

    COMPANY_COORD = (3, 5)
    CONTRACT_COORD = (3, 6)
    SEGMENT_COORD = (3, 7)
    ROAD_COORD = (7, 5)
    MEASURMENT_COORD = (7, 6)
    START_DATE_COORD = (7, 7)
    END_DATE_COORD = (9, 7)

    def __init__(
        self,
        excel_dnit_report: ExcelDnitReport,
        company_uuid: str,
        filters: str,
        reporting_file_uuids: List[str],
        font_size: str,
        measurement_number: str,
        start_date: str,
        end_date: str,
        file_kinds: str,
        images_per_kind: str,
        fields_to_print_images: str,
        coordinates_format: str,
        include_hour: str,
        use_location: str,
        sort_photos_by: str,
    ):
        self.excel_dnit_report = excel_dnit_report
        self.company: Company = Company.objects.get(uuid=company_uuid)
        self.filters = filters
        self.reporting_file_uuids = reporting_file_uuids

        self.temp_dir = mkdtemp()
        self.url = None
        self.name = None
        self.file_path = None

        self.font_size = self.FONT_SIZE[font_size]
        self.measurement_number = measurement_number
        self.start_date = datetime.strptime(start_date, "%Y-%m-%d")
        self.end_date = datetime.strptime(end_date, "%Y-%m-%d")

        self.file_kinds = file_kinds.split("_") if file_kinds else []

        self.images_per_kind = int(images_per_kind) if images_per_kind else 0
        self.include_hour = include_hour == "true"

        self.use_location = use_location == "true"
        self.coordinates_format = coordinates_format
        self.order_list = []
        self.water_mark_fields = self.get_water_mark_fields(
            fields_to_print_images, self.include_hour, self.coordinates_format
        )
        self.sort_photos_by = sort_photos_by if sort_photos_by else "datetime"

    def get_water_mark_fields(
        self, fields_to_print_images: str, include_hour: bool, coordinates_format: str
    ) -> List[str]:
        fields = fields_to_print_images
        converted_fields = []
        self.order_list = []
        for field in fields:
            field = field.strip()
            if include_hour and field == "image_date":
                converted_fields.append("date_and_hour")
                self.order_list.append("Data da imagem")
            elif include_hour and field == "executed_at":
                self.order_list.append("Executado em")
                converted_fields.append("executed_at_with_hour")
            elif field == "image_date":
                self.order_list.append("Data da imagem")
                converted_fields.append("date")
            elif field == "patology":
                self.order_list.append("Classe")
                converted_fields.append("classe")
            elif field == "coordinates":
                self.order_list.append("none2")
                converted_fields.append(to_snake_case(coordinates_format))
            elif field:
                if field == "direction":
                    self.order_list.append("Sentido")
                elif field == "road":
                    self.order_list.append("none1")
                elif field == "number":
                    self.order_list.append("Serial")
                elif field == "status":
                    self.order_list.append("Status")
                elif field == "executed_at":
                    self.order_list.append("Executado em")
                elif field == "note":
                    self.order_list.append("Observações")
                converted_fields.append(field)
        return converted_fields

    def add_water_mark_to_images(
        self,
        reporting_files: Dict[str, ReportingFile],
        rf_uuid_to_file_path: Dict[str, str],
    ):
        for rf_uuid, rf in reporting_files.items():
            water_mark_text_dict = build_text_dict(
                rf, self.water_mark_fields, self.use_location, self.company
            )
            sorted_water_mark_text: List[Tuple[str, str]] = []
            for key in self.order_list:
                if key in water_mark_text_dict:
                    sorted_water_mark_text.append((key, water_mark_text_dict[key]))

            file_path = rf_uuid_to_file_path.get(str(rf.uuid))
            watermarked_file_name = watermark_image_with_text(
                file_path, sorted_water_mark_text, self.font_size
            )
            if watermarked_file_name:
                rf_uuid_to_file_path[rf_uuid] = watermarked_file_name

    def filter_reporting_files(
        self, reporting_files_list: List[ReportingFile]
    ) -> List[ReportingFile]:
        """
        Filters reporting files based on enabled kinds and limits by images_per_kind.

        :param reporting_files_list: List of ReportingFile objects to filter
        :return: Filtered list with at most images_per_kind files per enabled kind
        """
        filtered_files = []
        kind_counts = {}

        # Initialize counters for enabled kinds
        for kind in self.file_kinds:
            kind_counts[kind] = 0

        for rf in reporting_files_list:
            kind = rf.kind
            # If no kinds are selected, include all kinds
            if len(self.file_kinds) == 0:
                if kind not in kind_counts:
                    kind_counts[kind] = 0

            if kind in kind_counts and (
                self.images_per_kind == 0 or kind_counts[kind] < self.images_per_kind
            ):
                filtered_files.append(rf)
                kind_counts[kind] += 1

        return filtered_files

    def copy_cell_range_style(
        self,
        ws: Worksheet,
        src_initial_col: int,
        src_initial_row: int,
        cols: int,
        rows: int,
        dest_initial_col: int,
    ):
        """
        Copies all styling from a source range of cells to a destination range.
        The destination range uses the same initial_row, cols, and rows as the source.

        :param worksheet: The worksheet to operate on
        :param src_initial_col: Source starting column (1-indexed)
        :param src_initial_row: Source starting row (1-indexed)
        :param cols: Number of columns in the range
        :param rows: Number of rows in the range
        :param dest_initial_col: Destination starting column (1-indexed)
        """

        for row_offset in range(rows):
            row = src_initial_row + row_offset

            for col_offset in range(cols):
                src_col = src_initial_col + col_offset
                dest_col = dest_initial_col + col_offset

                src_cell = ws.cell(row, src_col)
                dest_cell = ws.cell(row, dest_col)
                if src_cell.font:
                    dest_cell.font = copy(src_cell.font)
                if src_cell.border:
                    dest_cell.border = copy(src_cell.border)
                if src_cell.fill:
                    dest_cell.fill = copy(src_cell.fill)
                if src_cell.alignment:
                    dest_cell.alignment = copy(src_cell.alignment)
                if src_cell.number_format:
                    dest_cell.number_format = src_cell.number_format

                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

        # Copy merged cells
        src_end_col = src_initial_col + cols - 1
        src_end_row = src_initial_row + rows - 1
        col_offset = dest_initial_col - src_initial_col

        merged_ranges_to_copy = []
        for merged_range in ws.merged_cells.ranges:
            # Check if merged range intersects with source range
            if (
                merged_range.min_col >= src_initial_col
                and merged_range.max_col <= src_end_col
                and merged_range.min_row >= src_initial_row
                and merged_range.max_row <= src_end_row
            ):
                merged_ranges_to_copy.append(merged_range)

        # Create new merged ranges in destination
        for merged_range in merged_ranges_to_copy:
            new_min_col = merged_range.min_col + col_offset
            new_max_col = merged_range.max_col + col_offset
            new_min_row = merged_range.min_row
            new_max_row = merged_range.max_row

            dest_range = f"{get_column_letter(new_min_col)}{new_min_row}:{get_column_letter(new_max_col)}{new_max_row}"
            ws.merge_cells(dest_range)

        for col_offset in range(cols):
            src_col = src_initial_col + col_offset
            dest_col = dest_initial_col + col_offset
            src_col_letter = get_column_letter(src_col)
            dest_col_letter = get_column_letter(dest_col)

            src_col_dim = ws.column_dimensions[src_col_letter]
            ws.column_dimensions[dest_col_letter].width = src_col_dim.width
            ws.column_dimensions[dest_col_letter].hidden = src_col_dim.hidden
            ws.column_dimensions[
                dest_col_letter
            ].outlineLevel = src_col_dim.outlineLevel
            ws.column_dimensions[dest_col_letter].collapsed = src_col_dim.collapsed

            if src_col_dim.min:
                min_offset = src_col - src_col_dim.min
                ws.column_dimensions[dest_col_letter].min = dest_col - min_offset

            if src_col_dim.max:
                max_offset = (src_col_dim.max or src_col) - src_col
                ws.column_dimensions[dest_col_letter].max = dest_col + max_offset

    def create_pages(self, ws: Worksheet, file_count: int):
        page_count = ceil(file_count / len(self.IMAGE_COORDS))
        for page in range(1, page_count):
            initial_col = page * self.PAGE_COL_OFFSET
            self.copy_cell_range_style(
                ws, 1, 1, self.PAGE_COLS, self.PAGE_ROWS, initial_col
            )

            for row_offset in range(self.PAGE_ROWS):
                row = row_offset + 1
                for col_offset in range(self.PAGE_COLS):
                    col = col_offset + 1
                    dst_col = initial_col + col - 1
                    try:
                        src_cell = ws.cell(row, col)
                        if src_cell.value:
                            ws.cell(row, dst_col).value = src_cell.value
                    except Exception:
                        pass

    def get_excel_name(self):
        date = utc_to_local(datetime.now())
        title = f"Relatório Dnit - {self.company.name} - {date.day}-{date.month}-{date.year} - {date.hour}-{date.minute}"
        title = clean_latin_string(title.replace(".", "").replace("/", ""))
        return f"{title}.xlsx"

    def get_image_range(self, image_col: int, image_row: int):
        return f"{get_column_letter(image_col)}{image_row}:{get_column_letter(image_col + self.IMAGE_COL_COVER-1)}{image_row}"

    def create_aditional_pages(
        self,
        wb: Workbook,
        reporting_files_count: int,
        reporting_idx: int,
    ):
        template_ws = wb.get_sheet_by_name(self.TEMPLATE_SHEET_NAME)
        ws = wb.copy_worksheet(template_ws)
        title = f"{reporting_idx}"
        ws.title = title

        self.create_pages(ws, reporting_files_count)

    def add_header_picture(self, wb: Workbook, ws: Worksheet, pages_count: int):
        template_ws = wb.get_sheet_by_name(self.TEMPLATE_SHEET_NAME)
        images = template_ws._images
        images[0].ref.seek(0)
        image_bytes = images[0].ref.read()

        for page in range(pages_count):
            image_stream = BytesIO(image_bytes)
            image = Image(image_stream)

            start_col = page * self.PAGE_COL_OFFSET
            if page == 0:
                start_col = 1
            end_col = start_col + self.PAGE_COLS - 2
            start_col_letter = get_column_letter(start_col)
            end_col_letter = get_column_letter(end_col)
            image_range = f"{start_col_letter}1:{end_col_letter}{self.HEADER_HEIGHT}"
            insert_picture_2(
                ws,
                image_range,
                image,
                target=SheetTarget.DesktopExcel,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalCentered,
            )

    def add_header(self, wb: Workbook, ws: Worksheet, pages_count: int):
        company = get_obj_from_path(
            self.company.custom_options, "reporting__contractDnit__company"
        )
        contract = get_obj_from_path(
            self.company.custom_options, "reporting__contractDnit__contract"
        )
        segment = get_obj_from_path(
            self.company.custom_options, "reporting__contractDnit__segment"
        )
        road = get_obj_from_path(
            self.company.custom_options, "reporting__contractDnit__road"
        )

        for page in range(pages_count):
            start_col = page * self.PAGE_COL_OFFSET
            if page == 0:
                start_col = 1

            company_col = self.COMPANY_COORD[0] + start_col - 1
            contract_col = self.CONTRACT_COORD[0] + start_col - 1
            segment_col = self.SEGMENT_COORD[0] + start_col - 1
            road_col = self.ROAD_COORD[0] + start_col - 1
            measurement_col = self.MEASURMENT_COORD[0] + start_col - 1
            start_date_col = self.START_DATE_COORD[0] + start_col - 1
            end_date_col = self.END_DATE_COORD[0] + start_col - 1

            company_cell = ws.cell(self.COMPANY_COORD[1], company_col)
            contract_cell = ws.cell(self.CONTRACT_COORD[1], contract_col)
            segment_cell = ws.cell(self.SEGMENT_COORD[1], segment_col)
            road_cell = ws.cell(self.ROAD_COORD[1], road_col)
            measurement_cell = ws.cell(self.MEASURMENT_COORD[1], measurement_col)
            start_date_cell = ws.cell(self.START_DATE_COORD[1], start_date_col)
            end_date_cell = ws.cell(self.END_DATE_COORD[1], end_date_col)

            if isinstance(company, str):
                company_cell.value = company.upper()
            if isinstance(contract, str):
                contract_cell.value = contract.upper()
            if isinstance(segment, str):
                segment_cell.value = segment.upper()
            if isinstance(road, str):
                road_cell.value = road.upper()
            measurement_cell.value = str(self.measurement_number).upper()
            start_date_cell.value = self.start_date
            end_date_cell.value = self.end_date

    def add_reporting_to_sheet(
        self,
        wb: Workbook,
        reporting: Reporting,
        reporting_files: List[ReportingFile],
        files: Dict[str, str],
        reporting_idx: int,
    ):
        title = f"{reporting_idx}"
        ws = wb.get_sheet_by_name(title)

        images_per_page = len(self.IMAGE_COORDS)

        page_count = ceil(len(reporting_files) / images_per_page)
        if page_count == 0:
            page_count = 1
        self.add_header(wb, ws, page_count)
        self.add_header_picture(wb, ws, page_count)

        reporting_files = sorted(
            reporting_files, key=lambda rf: getattr(rf, self.sort_photos_by)
        )
        for i, rf in enumerate(reporting_files, 0):
            page = i // images_per_page
            idx_in_page = i % (images_per_page)

            page_initial_col = page * self.PAGE_COL_OFFSET
            if page != 0:
                page_initial_col -= 1

            image_col, image_row = self.IMAGE_COORDS[idx_in_page]
            image_col += page_initial_col

            image_range = self.get_image_range(image_col, image_row)
            file_path = files.get(str(rf.uuid))
            if file_path:
                image = Image(file_path)
                insert_picture_2(
                    worksheet=ws,
                    range_string=image_range,
                    picture=image,
                    target=SheetTarget.DesktopExcel,
                    border_width=(1, 1, 1, 1),
                    resize_method=ResizeMethod.ProportionalCentered,
                )

            info_row = image_row + self.IMAGE_INFO_ROW_OFFSET
            km_col = image_col + self.KM_COL_OFFSET
            kind_col = image_col + self.KIND_COL_OFFSET
            occ_type_col = image_col + self.OCC_TYPE_COL_OFFSET

            km = (
                f"KM {format_km(rf.km, separator=',')}"
                if rf.km
                else self.UNAVAILABLE_KM
            )
            kind = rf.kind.upper() if rf.kind else self.UNAVAILABLE_KIND
            occ_type = (
                reporting.occurrence_type.name.upper()
                if reporting.occurrence_type
                else self.UNAVAILABLE_OCC_TYPE
            )

            ws.cell(info_row, km_col).value = km
            ws.cell(info_row, kind_col).value = kind
            ws.cell(info_row, occ_type_col).value = occ_type
        print_range = f"$A$1:${get_column_letter(self.PAGE_COLS-1)}${self.PAGE_ROWS}"
        ws.print_area = print_range

    @classmethod
    def append_file(cls, s3, uuid_to_file: Dict[str, str], dir: str, rf: ReportingFile):
        file_path = download_picture(s3, dir, rf)
        uuid_to_file[str(rf.uuid)] = file_path

    def get_files(
        self, s3, reporting_files: Dict[str, ReportingFile]
    ) -> Dict[str, str]:
        uuid_to_file: Dict[str, str] = {}
        executor = ThreadPoolExecutor(max_workers=self.PICS_THREADING_LIMIT)
        for rf in reporting_files.values():
            executor.submit(self.append_file, s3, uuid_to_file, self.temp_dir, rf)
        executor.shutdown()
        return uuid_to_file

    def create_file(self):
        client_config = botocore.config.Config(
            max_pool_connections=self.PICS_THREADING_LIMIT,
        )
        s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
            config=client_config,
        )

        wb = load_workbook(self.TEMPLATE_PATH)
        self.name = self.get_excel_name()
        self.file_path = f"{self.temp_dir}/{self.name}.xlsx"

        reportings_qs = Reporting.objects.filter(company=self.company)

        from apps.reportings.views import ReportingFilter

        reportings_qs = ReportingFilter(
            self.filters, queryset=reportings_qs
        ).qs.prefetch_related("reporting_files", "occurrence_type")

        reporting_files = {
            rf.uuid: rf
            for rf in ReportingFile.objects.filter(
                uuid__in=self.reporting_file_uuids
            ).prefetch_related("reporting")
        }

        reporting_to_rfs = {}
        for i, reporting in enumerate(reportings_qs, 1):
            reporting_files_list = [
                reporting_files.get(rf.uuid) for rf in reporting.reporting_files.all()
            ]
            reporting_files_list = [rf for rf in reporting_files_list if rf is not None]
            reporting_files_list = self.filter_reporting_files(reporting_files_list)
            reporting_to_rfs[reporting.uuid] = reporting_files_list
            self.create_aditional_pages(wb, len(reporting_files_list), i)

        wb.save(self.file_path)
        wb.close()
        wb = load_workbook(self.file_path)

        reporting_files: Dict[str, ReportingFile] = {}
        for rf_list in reporting_to_rfs.values():
            for rf in rf_list:
                reporting_files[str(rf.uuid)] = rf

        rf_uuid_to_file_path = self.get_files(s3, reporting_files)

        self.add_water_mark_to_images(reporting_files, rf_uuid_to_file_path)
        for i, reporting in enumerate(reportings_qs, 1):
            self.add_reporting_to_sheet(
                wb, reporting, reporting_to_rfs[reporting.uuid], rf_uuid_to_file_path, i
            )

        wb.remove(wb.get_sheet_by_name(self.TEMPLATE_SHEET_NAME))
        set_zoom(wb, 50, "pageBreakPreview")
        set_active_cell(wb, "A1")
        wb.active = wb.worksheets[0]
        wb.save(self.file_path)
        wb.close()


@task
def create_and_upload_excel_dnit_report(excel_dnit_report_id):
    excel_dnit_report = None
    error = True
    try:
        excel_dnit_report = ExcelDnitReport.objects.get(pk=excel_dnit_report_id)
        params = excel_dnit_report.extra_info
        filters = excel_dnit_report.filters
        company_uuid = None
        if excel_dnit_report.company:
            company_uuid = excel_dnit_report.company.uuid
        else:
            company_uuid = filters.get("company")
        excel_dnit_report_export = ExcelDnitReportExport(
            excel_dnit_report=excel_dnit_report,
            company_uuid=company_uuid,
            filters=filters,
            reporting_file_uuids=params.get("file_id"),
            font_size=params.get("font_size"),
            measurement_number=params.get("measurement_number"),
            start_date=params.get("start_date"),
            end_date=params.get("end_date"),
            file_kinds=params.get("file_kinds"),
            images_per_kind=params.get("images_per_kind"),
            fields_to_print_images=params.get("fields_to_print_images"),
            coordinates_format=params.get("coordinates_format"),
            include_hour=params.get("include_hour"),
            use_location=params.get("use_location"),
            sort_photos_by=params.get("sort_photos_by"),
        )

        excel_dnit_report_export.create_file()
        with open(excel_dnit_report_export.file_path, "rb") as excel_file:
            excel_dnit_report.exported_file.save(
                excel_dnit_report_export.name,
                ContentFile(excel_file.read()),
            )
            error = False

    except Exception as e:
        logging.error(f"Error creating and uploading Excel Dnit Report: {e}")
        sentry_sdk.capture_exception(e)
    finally:
        excel_dnit_report.error = error
        excel_dnit_report.done = True
        excel_dnit_report.save()
