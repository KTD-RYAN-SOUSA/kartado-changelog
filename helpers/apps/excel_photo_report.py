import json
import os
import shutil
from datetime import datetime, timedelta
from urllib.parse import unquote
from zipfile import ZipFile

import boto3
import pytz
from django.conf import settings
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D

# from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from PIL import Image as ImagePil

from apps.reportings.models import Reporting
from helpers.apps.json_logic import apply_json_logic
from helpers.strings import UF_CODE, check_image_file, format_km, get_direction_name
from RoadLabsAPI.settings import credentials

from ..kartado_excel.workbook import load_workbook


class ExcelPhotoEndpoint:
    temp_path = "/tmp/excel_photo_report"
    number_pictures = 4

    def __init__(
        self,
        reportings,
        template_path="fixtures/reports/excel_photo_report.xlsx",
        json_path="fixtures/reports/excel_photo_report.json",
    ):
        self.json_path = json_path
        self.template_path = template_path
        self.reportings = reportings.split(",")
        self.path_excel = "{}/excel".format(self.temp_path)
        os.makedirs(self.path_excel, exist_ok=True)

    def load_data(self):
        try:
            self.wb = load_workbook(filename=self.template_path)
            self.ws = self.wb.active
            with open(self.json_path) as json_file:
                self.data = json.load(json_file)
        except Exception:
            self.wb = None
            self.ws = None
            self.data = None
        return

    def get_reportings(self):
        try:
            reportings = Reporting.objects.filter(pk__in=self.reportings).distinct()
        except Exception:
            reportings = Reporting.objects.none()
        return reportings

    def get_attr(self, path, reporting):
        for simple_path in path.split("."):
            if isinstance(self.attr_result, dict):
                self.attr_result = self.attr_result.get(simple_path, "")
            else:
                self.attr_result = getattr(self.attr_result, simple_path, "")
        return self.attr_result

    def get_kms(self, reportings):
        reportings_kms = [
            "{} FX{} {}".format(
                format_km(item, "km"),
                item.lane,
                get_direction_name(self.company, item.direction, "short"),
            )
            if item.km == item.end_km
            else "{} A {} FX{} {}".format(
                format_km(item, "km"),
                format_km(item, "end_km"),
                item.lane,
                get_direction_name(self.company, item.direction, "short"),
            )
            for item in reportings
        ]
        return reportings_kms

    def fill_cell(self, json_field, reporting, images_path):
        self.attr_result = reporting
        start_cell = json_field.get("start_cell", None)
        source = json_field.get("source", "")
        logic = json_field.get("logic", False)
        field = json_field.get("field", "")

        if not start_cell:
            return

        cell = self.ws[start_cell]

        if field in [
            "number",
            "type",
            "title",
            "notes",
            "date",
            "hour",
            "location",
        ]:
            result = (
                apply_json_logic(logic, getattr(reporting, source, {}))
                if logic
                else self.get_attr(source, reporting)
            )
            if field in ["type", "notes"]:
                cell.value = result
            elif field == "number":
                year = reporting.found_at.year or datetime.now().year
                cell.value = "Relatório {} de {}".format(result, year)
            elif field == "title":
                initial = json_field.get("initial", "")
                cell.value = initial + UF_CODE.get(str(result), "")
            elif field == "date":
                initial = json_field.get("initial", "")
                date_str = (
                    result.strftime("%d/%m/%Y") if isinstance(result, datetime) else ""
                )
                cell.value = initial + date_str
            elif field == "hour":
                cell.value = (
                    "{} h {} min".format(result.hour, result.minute)
                    if isinstance(result, datetime)
                    else ""
                )
            elif field == "location":
                full_path = source + ".{}".format(result)
                form_data_value = self.get_attr(full_path, reporting)
                form_data_search = full_path.replace(".", "__")
                reportings = Reporting.objects.filter(
                    **{form_data_search: form_data_value}
                )
                road_name = "{} - KM'S: ".format(reporting.road_name)
                reportings_kms = self.get_kms(reportings)
                full_string = road_name + "; ".join(reportings_kms)
                current_km = self.get_kms([reporting])
                black_text = full_string.split(current_km[0])
                try:
                    last_str = black_text[1]
                except Exception:
                    last_str = ""
                cell.value = '__kartado_styled_string<is><r><t xml:space="preserve">{}</t></r><r><rPr><b val="true" /><u val="true" /><sz val="10" /><color rgb="FFFF0000" /></rPr><t xml:space="preserve">{}</t></r><r><t xml:space="preserve">{}</t></r></is>'.format(
                    black_text[0], current_km[0], last_str
                )
        elif field == "picture_title":
            cell.value = source
        elif field == "picture":
            images_names = os.listdir(images_path)
            try:
                name = list(filter(lambda name: source in name, images_names))[0]
                img = Image(images_path + name)
                cell_width_cm = json_field.get("cell_width_cm", 0)
                cell_height_cm = json_field.get("cell_height_cm", 0)

                margin_top_cm = json_field.get("margin_top_cm", 0)
                margin_left_cm = json_field.get("margin_left_cm", 0)

                if cell_height_cm and cell_width_cm:
                    # center image
                    height, width = img.height, img.width
                    size = XDRPositiveSize2D(
                        pixels_to_EMU(width), pixels_to_EMU(height)
                    )

                    width_EMU = pixels_to_EMU(width)
                    height_EMU = pixels_to_EMU(height)

                    cell_width_EMU = cm_to_EMU(cell_width_cm)
                    cell_height_EMU = cm_to_EMU(cell_height_cm)

                    rowoffset = int((cell_height_EMU - height_EMU) / 2)
                    coloffset = int((cell_width_EMU - width_EMU) / 2)

                    marker = AnchorMarker(
                        col=cell.column - 1,
                        colOff=coloffset,
                        row=cell.row - 1,
                        rowOff=rowoffset,
                    )
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                elif margin_top_cm or margin_left_cm:
                    height, width = img.height, img.width
                    size = XDRPositiveSize2D(
                        pixels_to_EMU(width), pixels_to_EMU(height)
                    )

                    margin_left_EMU = cm_to_EMU(margin_left_cm)
                    margin_top_EMU = cm_to_EMU(margin_top_cm)

                    marker = AnchorMarker(
                        col=cell.column - 1,
                        colOff=margin_left_EMU,
                        row=cell.row - 1,
                        rowOff=margin_top_EMU,
                    )
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                else:
                    img.anchor = start_cell
                self.ws.add_image(img)
            except Exception:
                return

        cell.alignment = Alignment(**json_field.get("alignment_style", {}))
        cell.font = Font(**json_field.get("font_style", {}))
        return

    def write_on_sheet(self, reporting, images_path):
        if isinstance(self.data, list):
            for json_field in self.data:
                if isinstance(json_field, dict):
                    self.fill_cell(json_field, reporting, images_path)
        return

    def resize_images(self, image_path):
        try:
            image = ImagePil.open(image_path)
        except Exception:
            return False

        # width 337px = 8.9cm / height 242px = 6.4cm
        width = 337
        height = 242

        try:
            image = image.resize((width, height))
            image.save(image_path)
        except Exception:
            return False
        else:
            return True

    def download_pictures(self, s3, path, reporting):
        files = reporting.reporting_files.filter(include_dnit=True).order_by("datetime")

        for index, file_obj in enumerate(files):
            if file_obj.upload:
                try:
                    file_path = file_obj.upload.url.split("?")[0].split(".com/")[1]
                    bucket_name = file_obj.upload.url.split(".s3")[0].split("/")[-1]
                    image_format = file_path.split(".")[-1]
                    image_name = str(index)
                except Exception:
                    return False

                if not check_image_file(file_path):
                    return False

                image_path = "{}{}.{}".format(path, image_name, image_format)

                try:
                    s3.download_file(bucket_name, unquote(file_path), image_path)
                except Exception:
                    return False
                else:
                    resized = self.resize_images(image_path)
                    if not resized:
                        return False
            else:
                return False
        return True

    def get_excel_name(self, reporting):
        logic = {"or": [{"var": ["tro_number"]}, {"var": ["codigo_engenho"]}]}
        result = apply_json_logic(logic, reporting.form_data)
        if result:
            result = str(result).replace("/", "_")
        year = reporting.found_at.year or datetime.now().year
        name = "{}_de_{}".format(result, year)

        reportings_kms = self.get_kms([reporting])
        kms = reportings_kms[0] if reportings_kms else ""

        return "TRO_{}_{}".format(name, kms.replace(" ", "_"))

    def upload_file(self, s3, path, name):
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        expires = datetime.now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)
        object_name = "media/private/{}".format(name)

        try:
            s3.upload_file(
                path, bucket_name, object_name, ExtraArgs={"Expires": expires}
            )
        except Exception:
            return False

        # Delete file
        os.remove(path)

        url_s3 = s3.generate_presigned_url(
            "get_object", Params={"Bucket": bucket_name, "Key": object_name}
        )
        return url_s3

    def get_data(self):
        empty = {"url": "", "name": ""}
        reportings = self.get_reportings()

        if not reportings:
            return empty

        self.company = reportings.first().company

        s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )

        # Create temporary folder for images
        path = "{}/images/".format(self.temp_path)
        os.makedirs(path, exist_ok=True)

        for reporting in reportings:
            # Load templates
            self.load_data()
            if not self.wb or not self.ws or not self.data:
                return empty

            # Download images
            downloaded = self.download_pictures(s3, path, reporting)
            if not downloaded or len(os.listdir(path)) != self.number_pictures:
                continue

            # Fill the template excel
            self.write_on_sheet(reporting, path)

            # Save the excel
            excel_name = self.get_excel_name(reporting)
            self.wb.save("{}/{}.xlsx".format(self.path_excel, excel_name))

            # Delete images
            for file_name in os.listdir(path):
                os.remove(path + file_name)

        # Delete temporary folder
        shutil.rmtree(path, ignore_errors=True)

        # Upload files
        if len(os.listdir(self.path_excel)) == 0:
            return empty
        elif len(os.listdir(self.path_excel)) > 1:
            final_filename = "Relatorio_ANTT_Excel.zip"
            path_zip_total = self.temp_path + "/" + final_filename
            with ZipFile(path_zip_total, "w") as zipObj:
                for folder_name, subfolders, filenames in os.walk(self.path_excel):
                    for filename in filenames:
                        file_path = os.path.join(folder_name, filename)
                        zipObj.write(file_path, filename)
            url = self.upload_file(s3, path_zip_total, final_filename)
        else:
            final_filename = (
                os.listdir(self.path_excel)[0] if os.listdir(self.path_excel) else ""
            )
            if final_filename:
                path_total = self.path_excel + "/" + final_filename
                url = self.upload_file(s3, path_total, final_filename)
            else:
                final_filename = ""
                url = False

        if not url or not final_filename:
            return empty

        # Delete excel folder
        for file_name in os.listdir(self.path_excel):
            os.remove(self.path_excel + "/" + file_name)
        shutil.rmtree(self.path_excel, ignore_errors=True)

        return {"url": url, "name": final_filename}
