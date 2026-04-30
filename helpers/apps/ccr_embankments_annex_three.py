import shutil
import tempfile
from datetime import datetime
from typing import Dict, List, Tuple
from uuid import UUID
from zipfile import ZipFile

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_conditions_date,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import ReportFormat, SheetTarget
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_direction,
    get_found_at,
    get_identification,
    get_km,
    get_parent_serial,
    get_previous_year_report,
    get_road_name,
    get_serial,
)
from helpers.apps.ccr_report_utils.workbook_utils import save_workbook, set_block_style
from helpers.strings import clean_latin_string


def get_found_at_after_condition(panel_uuid: str):
    return get_conditions_date(panel_uuid, "found_at__date__gt", r"\d+\-\d+\-\d+")


def get_found_at_before_condition(panel_uuid: str):
    return get_conditions_date(panel_uuid, "found_at__date__lt", r"\d+\-\d+\-\d+")


MONTH_LIST = [
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]


class EmbankmentsAnnexThreeXlsxHandler(object):
    __OCCURRENCE_TYPE_UUIDS: List[UUID] = []

    __RECUPERATION_NAMES: List[str] = [
        "Desassoreamento",
        "Esgotamento e destinação",
        "Hidrojateamento em drenagem",
        "Implantação Drenagem",
        "Limpeza/desobstrução manual",
        "Limpeza/desobstrução mecânica",
        "Pintura em elemento de drenagem",
        "Reconstrução Drenagem",
        "Recuperação de drenagem",
        "Reparo em drenagem",
    ]

    __EMPTY_MESSAGE = "Os apontamentos selecionados ou filtrados não atendem ao requisito de nível de risco igual a R2 ou R3 ou R4"

    __DEFAULT_ROW_HEIGHT = 19.5

    __DEFAULT_BORDER_COLOR = "bfbfbf"

    __DEFAULT_FONT = Font(name="Aptos", size=10)
    __DEFAULT_SIDE = Side(border_style="thin", color=__DEFAULT_BORDER_COLOR)

    __DEFAULT_BORDER = Border(
        left=__DEFAULT_SIDE,
        right=__DEFAULT_SIDE,
        top=__DEFAULT_SIDE,
        bottom=__DEFAULT_SIDE,
    )

    __DEFAULT_ALIGNMENT = Alignment(
        vertical="center", horizontal="center", wrap_text=True
    )

    __RISK_TEXT_LOOKUP = {
        "": 0,
        "R1 - Baixo": 1,
        "R2 - Moderado": 2,
        "R3 - Alto": 3,
        "R4 - Muito Alto": 4,
    }

    __RISK_RANGE_LOOKUP = {
        "": 0,
        "R2 - Moderado": 4,
        "R3 - Alto": 1,
        "R4 - Muito Alto": 1,
    }

    __RISK_COLOR_LOOKUP = {
        "": "000000",
        "R1 - Baixo": "00b050",
        "R2 - Moderado": "ffff00",
        "R3 - Alto": "ffc000",
        "R4 - Muito Alto": "ff0000",
    }

    __RISK_COLOR_LOOKUP = {
        "": "000000",
        "R1 - Baixo": "00b050",
        "R2 - Moderado": "ffff00",
        "R3 - Alto": "ffc000",
        "R4 - Muito Alto": "ff0000",
    }

    __LOWERING_COLOR = "002060"

    @classmethod
    def __get_risk_level_text(cls, reporting: Reporting) -> str:
        try:
            return new_get_form_data(reporting, "riskLevel")
        except Exception:
            return ""

    @classmethod
    def __get_risk_level_color(cls, reporting: Reporting) -> str:
        try:
            risk_level = new_get_form_data(reporting, "riskLevel")
            return EmbankmentsAnnexThreeXlsxHandler.__RISK_COLOR_LOOKUP[risk_level]
        except Exception:
            return ""

    @classmethod
    def __get_risk_level_previous_consecutive_years(
        cls, reporting: Reporting, risk_level: str
    ) -> int:
        previous_consecutive_years = 0
        search_range = EmbankmentsAnnexThreeXlsxHandler.__RISK_RANGE_LOOKUP[risk_level]

        risk_level_change = False
        i = 1
        while risk_level_change is False and i < search_range + 1:
            previous_reporting = get_previous_year_report(
                reporting, i, "form_data", "occurrence_type"
            )
            previous_risk_level = (
                EmbankmentsAnnexThreeXlsxHandler.__get_risk_level_text(
                    previous_reporting
                )
            )

            if previous_risk_level == risk_level:
                previous_consecutive_years += 1
            else:
                risk_level_change = True
            i += 1
        return previous_consecutive_years

    @classmethod
    def __get_lowering_offset(
        cls, risk_level: str, previous_consecutive_years: int
    ) -> int:
        lowering_offset = 1
        if risk_level == "R3 - Alto":
            lowering_offset = 2 - min(previous_consecutive_years, 1)
        if risk_level == "R2 - Moderado":
            lowering_offset = 5 - min(previous_consecutive_years, 4)

        return lowering_offset

    def __init__(
        self,
        found_at,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.__sheet_target = sheet_target
        self.__sheet_reporting: Dict[Tuple[str, int], List[Reporting]] = {}
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.occurrence_type = Reporting.objects.get(uuid=list_uuids[0]).occurrence_type

        self.reference_date = None
        try:
            self.reference_date = datetime.strptime(found_at["before"], "%Y-%m-%d")
        except Exception:
            self.reference_date = datetime.strptime(found_at["after"], "%Y-%m-%d")

        for e in OccurrenceType.objects.values("uuid", "name"):
            if e["name"] in EmbankmentsAnnexThreeXlsxHandler.__RECUPERATION_NAMES:
                EmbankmentsAnnexThreeXlsxHandler.__OCCURRENCE_TYPE_UUIDS.append(
                    e["uuid"]
                )

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __fill_years_headers(self, worksheet: Worksheet) -> None:
        column_range = range(9, 14)
        i = 1
        for column_number in column_range:
            index = get_column_letter(column_number) + "3"
            cell: Cell = worksheet[index]
            cell.value = str(self.reference_date.year + i)
            i += 1

    def __fill_months_headers(self, worksheet: Worksheet) -> None:
        first_year = (self.reference_date + relativedelta(months=1)).year
        second_year = 0
        first_year_end = -1
        column_range = range(8, 20)
        i = 1
        for column_number in column_range:
            index = get_column_letter(column_number) + "4"
            cell: Cell = worksheet[index]
            date = self.reference_date + relativedelta(months=i)
            if first_year_end == -1 and date.year > first_year:
                first_year_end = i
            cell.value = MONTH_LIST[date.month - 1]
            i += 1

        try:
            first_merge_area = "H3:" + get_column_letter(first_year_end + 6) + "3"
            worksheet.merge_cells(first_merge_area)
            first_year_cell: Cell = worksheet["H3"]
            first_year_cell.value = first_year
            second_year = first_year + 1
        except Exception:
            first_year_end = 1
            second_year = first_year

        second_merge_area = get_column_letter(first_year_end + 7) + "3:S3"
        worksheet.merge_cells(second_merge_area)
        second_year_cell: Cell = worksheet[get_column_letter(first_year_end + 7) + "3"]
        second_year_cell.value = second_year

    @classmethod
    def __get_lowering_year(
        cls, reporting: Reporting, previous_consecutive_years: int, risk_level: str
    ) -> int:
        found_at = get_found_at(reporting)
        current_year = found_at.year
        lowering_offset = cls.__get_lowering_offset(
            risk_level, previous_consecutive_years
        )
        lowering_year = current_year + lowering_offset
        return lowering_year

    def __append_yearly_entry(
        self,
        worksheet: Worksheet,
        reporting: Reporting,
        previous_consecutive_years: int,
        risk_level: str,
        lowering_year: int,
        item: int,
    ) -> None:
        identification = get_identification(reporting)
        road_name = get_road_name(reporting)
        km = get_km(reporting)
        direction = get_direction(reporting)
        serial = get_serial(reporting)
        inventory_serial = get_parent_serial(reporting)
        risk_level_color = EmbankmentsAnnexThreeXlsxHandler.__get_risk_level_color(
            reporting
        )
        risk_level_cell = Cell(worksheet, value=risk_level)
        risk_level_cell.fill = PatternFill(
            start_color=risk_level_color, fill_type="solid"
        )

        if risk_level == "R4 - Muito Alto" and previous_consecutive_years == 1:
            risk_level_cell.value = "SUJEITO A MULTA"

        lowering_list = [""] * 5
        lowering_pos = lowering_year - self.reference_date.year - 1
        if lowering_pos >= 0 and lowering_pos < 5:
            lowering_cell = Cell(worksheet, value="")
            lowering_cell.fill = PatternFill(
                start_color=EmbankmentsAnnexThreeXlsxHandler.__LOWERING_COLOR,
                fill_type="solid",
            )
            lowering_list[lowering_pos] = lowering_cell

        entry: List[str | Cell] = [
            item,
            road_name,
            identification,
            km,
            direction,
            risk_level_cell,
            "Recuperação",
            lowering_year,
        ]
        entry.extend(lowering_list)
        entry.extend([serial, inventory_serial])
        worksheet.append(entry)

        last_row_index = worksheet.max_row
        set_block_style(
            worksheet,
            row_begin=last_row_index,
            row_end=last_row_index,
            col_begin="A",
            col_end=get_column_letter(len(entry)),
            height=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_ROW_HEIGHT,
            font=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_FONT,
            border=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_BORDER,
            alignment=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_ALIGNMENT,
        )

    def __append_monthly_entry(
        self, worksheet: Worksheet, reporting: Reporting, risk_level: str, item: int
    ) -> None:
        identification = get_identification(reporting)
        road_name = get_road_name(reporting)
        km = get_km(reporting)
        direction = get_direction(reporting)
        serial = get_serial(reporting)
        inventory_serial = get_parent_serial(reporting)
        risk_level_color = EmbankmentsAnnexThreeXlsxHandler.__get_risk_level_color(
            reporting
        )
        risk_level_cell = Cell(worksheet, value=risk_level)
        risk_level_cell.fill = PatternFill(
            start_color=risk_level_color, fill_type="solid"
        )

        entry: List[str | Cell] = [
            item,
            road_name,
            identification,
            km,
            direction,
            risk_level_cell,
            "Recuperação",
        ]
        merging_range = len(entry)

        entry.extend([""] * 13)
        entry.extend([serial, inventory_serial])

        worksheet.append(entry)
        last_row_index = worksheet.max_row

        for i in range(1, merging_range + 1):
            column = get_column_letter(i)
            merged_cells_area = (
                column + str(last_row_index) + ":" + column + str(last_row_index + 1)
            )
            worksheet.merge_cells(merged_cells_area)

        for i in range(20, 23):
            column = get_column_letter(i)
            merged_cells_area = (
                column + str(last_row_index) + ":" + column + str(last_row_index + 1)
            )
            worksheet.merge_cells(merged_cells_area)

        set_block_style(
            worksheet,
            row_begin=last_row_index,
            row_end=last_row_index + 1,
            col_begin="A",
            col_end=get_column_letter(len(entry)),
            height=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_ROW_HEIGHT,
            font=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_FONT,
            border=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_BORDER,
            alignment=EmbankmentsAnnexThreeXlsxHandler.__DEFAULT_ALIGNMENT,
        )

    def __add_to_sheet_reporting(self, reporting: Reporting) -> None:
        found_at = get_found_at(reporting)
        if (
            found_at is not None
            and (reporting.road_name, found_at.year) not in self.__sheet_reporting
        ):
            self.__sheet_reporting[(reporting.road_name, found_at.year)] = []

        self.__sheet_reporting[(reporting.road_name, found_at.year)].append(reporting)

    def __create_empty_workbook_file(self) -> str:
        workbook = load_workbook(
            "./fixtures/reports/ccr_embankments_retaining_structures_annex_three.xlsx"
        )
        yearly_worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
        yearly_worksheet.merge_cells("B5:O5")
        warn_cell: Cell = yearly_worksheet["B5"]
        warn_cell.value = EmbankmentsAnnexThreeXlsxHandler.__EMPTY_MESSAGE
        workbook_name = "ANEXO III - CRONOGRAMA DE INTERVENÇÕES - PLANO DE AÇÃO"
        file = save_workbook(workbook_name, workbook)
        return file

    def __create_workbooks_files(self) -> List[str]:
        files: List[str] = []

        sorted_sheets = sorted(self.__sheet_reporting.items())
        for sheet_key, reportings in sorted_sheets:
            (road_name, year) = sheet_key
            reportings.sort(key=lambda reporting: reporting.km)

            workbook = load_workbook(
                "./fixtures/reports/ccr_embankments_retaining_structures_annex_three.xlsx"
            )
            yearly_worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
            self.__fill_years_headers(yearly_worksheet)
            yearly_worksheet.column_dimensions["N"].hidden = True
            yearly_worksheet.column_dimensions["O"].hidden = True

            monthly_worksheet = workbook.get_sheet_by_name(
                workbook.get_sheet_names()[1]
            )
            self.__fill_months_headers(monthly_worksheet)
            monthly_worksheet.column_dimensions["U"].hidden = True
            monthly_worksheet.column_dimensions["V"].hidden = True

            yearly_index = 1
            monthly_index = 1
            for reporting in reportings:
                risk_level = EmbankmentsAnnexThreeXlsxHandler.__get_risk_level_text(
                    reporting
                )
                previous_consecutive_years = EmbankmentsAnnexThreeXlsxHandler.__get_risk_level_previous_consecutive_years(
                    reporting, risk_level
                )
                lowering_year = EmbankmentsAnnexThreeXlsxHandler.__get_lowering_year(
                    reporting, previous_consecutive_years, risk_level
                )
                self.__append_yearly_entry(
                    yearly_worksheet,
                    reporting,
                    previous_consecutive_years,
                    risk_level,
                    lowering_year,
                    yearly_index,
                )
                yearly_index += 1

                if (
                    EmbankmentsAnnexThreeXlsxHandler.__RISK_TEXT_LOOKUP[risk_level]
                    in [2, 3, 4]
                    and lowering_year == self.reference_date.year + 1
                ):
                    self.__append_monthly_entry(
                        monthly_worksheet, reporting, risk_level, monthly_index
                    )
                    monthly_index += 1

            monthly_worksheet.sheet_state = Worksheet.SHEETSTATE_HIDDEN

            workbook_name = (
                "ANEXO III - CRONOGRAMA DE INTERVENÇÕES - PLANO DE AÇÃO {} {}".format(
                    road_name, year
                )
            )
            file = save_workbook(workbook_name, workbook)
            files.append(file)

        if len(files) == 0:
            file = self.__create_empty_workbook_file()
            files.append(file)

        return files

    def execute(self) -> List[str]:
        query_set = (
            Reporting.objects.filter(
                occurrence_type=self.occurrence_type,
                uuid__in=self.list_uuids,
                form_data__risk_level__in=["2", "3", "4"],
            )
            .only(
                "uuid",
                "number",
                "road_name",
                "found_at",
                "km",
                "direction",
                "form_data",
                "parent__number",
                "occurrence_type",
            )
            .prefetch_related("parent", "occurrence_type")
        )

        for report in query_set:
            self.__add_to_sheet_reporting(reporting=report)

        return self.__create_workbooks_files()


class CCREmbankmentsAnnexThree(CCRReport):
    def __init__(
        self,
        found_at=None,
        panel_uuid: str = None,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.found_at_filter = found_at

        self.found_at_filter: Dict = {}
        if found_at is None and panel_uuid is not None:
            try:
                self.found_at_filter["before"] = get_found_at_before_condition(
                    panel_uuid
                )
            except Exception:
                pass
            try:
                self.found_at_filter["after"] = get_found_at_after_condition(panel_uuid)
            except Exception:
                pass
        else:
            self.found_at_filter = found_at

        super().__init__(uuids, report_format)

    def get_file_name(self):
        road_names = list(
            Reporting.objects.filter(uuid__in=self.uuids)
            .only("uuid", "road_name")
            .order_by("road_name")
            .distinct("road_name")
            .values_list("road_name", flat=True)
        )
        file_name = "ANEXO III - CRONOGRAMA DE INTERVENÇÕES - PLANO DE AÇÃO {}".format(
            " - ".join(road_names)
        )
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"
        return file_name

    def export(self):
        s3 = get_s3()
        files = EmbankmentsAnnexThreeXlsxHandler(
            found_at=self.found_at_filter,
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_embankments_annex_three_async_handler(
    reporter_dict: dict,
):
    reporter = CCREmbankmentsAnnexThree.from_dict(reporter_dict)
    reporter.export()
