import json
import os
from copy import copy
from datetime import datetime, timedelta
from urllib.parse import unquote

import boto3
import pytz
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from PIL import Image as ImagePil

from apps.reportings.models import ReportingFile
from helpers.dates import utc_to_local
from helpers.strings import check_image_file, get_direction_name, keys_to_snake_case
from RoadLabsAPI.settings import credentials


class ExcelEloEndpoint:
    temp_path = "/tmp/excel_elo_report"
    len_rows = 18

    def __init__(
        self,
        reportings,
        template_path="fixtures/reports/elo_excel.xlsx",
        json_path="fixtures/reports/elo_excel.json",
    ):
        self.json_path = json_path
        self.template_path = template_path
        self.reportings = reportings.prefetch_related("reporting_files")
        self.custom_fill = ""
        self.thin = Side(border_style="thin", color="000000")
        self.borders = {
            "left": self.thin,
            "right": self.thin,
            "top": self.thin,
            "bottom": self.thin,
        }
        self.path_excel = "{}/excel".format(self.temp_path)
        os.makedirs(self.path_excel, exist_ok=True)

    def load_data(self):
        try:
            self.wb = load_workbook(filename=self.template_path)
            with open(self.json_path) as json_file:
                self.data = json.load(json_file)
        except Exception:
            self.wb = None
            self.data = None
        return

    def get_attr(self, path, reporting):
        if not path:
            return path
        for simple_path in path.split("."):
            if isinstance(self.attr_result, dict):
                self.attr_result = self.attr_result.get(simple_path, "")
            else:
                self.attr_result = getattr(self.attr_result, simple_path, "")
        return self.attr_result

    def filter_image_names(self, name, source_fields, reporting):
        return any(
            [
                "{}-{}".format(str(reporting["count"] - 1), kind) in name
                for kind in source_fields
            ]
        )

    def write_on_sheet(self, images_path):
        count = 0
        for reporting in self.reportings_data:
            downloaded = self.download_pictures(images_path, reporting, count)
            if not downloaded:
                continue
            count += 1
            reporting["count"] = count
            for json_field in self.data:
                if isinstance(json_field, dict):
                    self.fill_cell(json_field, reporting, images_path)
        return

    def fill_cell(self, json_field, reporting, images_path):
        self.attr_result = reporting
        sheet = json_field.get("sheet", None)
        start_cell = json_field.get("start_cell", None)
        end_cell = json_field.get("end_cell", None)
        source = json_field.get("source", "")
        field = json_field.get("field", "")
        initial = json_field.get("initial", "")

        if not start_cell or not end_cell or not isinstance(sheet, int):
            return

        ws = self.wb.worksheets[sheet]

        # Change print_area in the first iteration
        if reporting["count"] == 1:
            if sheet == 0:
                final_row = int(start_cell[1:]) + len(self.reportings_data) - 1
                ws.print_area = "B1:N{}".format(final_row)
            else:
                self.custom_fill = ws["C22"].fill
                final_row = 8 + (len(self.reportings_data) * self.len_rows) - 1
                ws.print_area = "B1:Z{}".format(final_row)

        # Increase row
        if sheet == 0:
            new_start_row = int(start_cell[1:]) + reporting["count"] - 1
            new_end_row = int(end_cell[1:]) + reporting["count"] - 1
        else:
            new_start_row = int(start_cell[1:]) + (
                (reporting["count"] - 1) * self.len_rows
            )
            new_end_row = int(end_cell[1:]) + ((reporting["count"] - 1) * self.len_rows)

        new_start_cell = start_cell[0] + str(new_start_row)
        new_end_cell = end_cell[0] + str(new_end_row)
        cell = ws[new_start_cell]

        if field in [
            "number",
            "text",
            "date",
            "km",
            "form_data",
            "road",
            "direction",
            "quantity",
            "unit",
        ]:
            result = self.get_attr(source, reporting)
            if field in ["number", "text"]:
                cell.value = initial + str(result).upper()
                if source == "count" and sheet == 1 and self.custom_fill:
                    cell.fill = copy(self.custom_fill)
            elif field == "road":
                cell.value = result.replace("-", " ").upper()
            elif field == "date":
                date_str = (
                    result.strftime("%d/%m/%Y") if isinstance(result, datetime) else ""
                )
                cell.value = initial + date_str
            elif field == "form_data":
                cell.value = (initial + str(result) if result else "").upper()
            elif field == "km":
                cell.value = format(round(float(result), 3), ".3f").replace(".", "+")
            elif field == "direction":
                cell.value = get_direction_name(self.company, result).upper()
            elif field in ["quantity", "unit"]:
                fields = [
                    keys_to_snake_case(item)
                    for item in reporting["occurrence_type_form"].get("fields", [])
                ]
                main_field = list(
                    filter(
                        lambda x: x.get("main_report", False) and x.get("unit", False),
                        fields,
                    )
                )
                if main_field:
                    if field == "quantity":
                        value = reporting["form_data"].get(
                            main_field[0]["api_name"], ""
                        )
                    else:
                        value = main_field[0]["unit"]
                    cell.value = str(value) if value else ""

        if field == "picture":
            images_names = os.listdir(images_path)
            try:
                name = list(
                    filter(
                        lambda image_name: self.filter_image_names(
                            image_name, source, reporting
                        ),
                        images_names,
                    )
                )[0]
                img = Image(images_path + name)
                cell_width_cm = json_field.get("cell_width_cm", 0)
                cell_height_cm = json_field.get("cell_height_cm", 0)

                margin_top_cm = json_field.get("margin_top_cm", 0)
                margin_left_cm = json_field.get("margin_left_cm", 0)

                if cell_height_cm and cell_width_cm:
                    # center image
                    height, width = img.height, img.width
                    size = XDRPositiveSize2D(
                        pixels_to_EMU(width), pixels_to_EMU(height)
                    )

                    width_EMU = pixels_to_EMU(width)
                    height_EMU = pixels_to_EMU(height)

                    cell_width_EMU = cm_to_EMU(cell_width_cm)
                    cell_height_EMU = cm_to_EMU(cell_height_cm)

                    rowoffset = int((cell_height_EMU - height_EMU) / 2)
                    coloffset = int((cell_width_EMU - width_EMU) / 2)

                    marker = AnchorMarker(
                        col=cell.column - 1,
                        colOff=coloffset,
                        row=cell.row - 1,
                        rowOff=rowoffset,
                    )
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                elif margin_top_cm or margin_left_cm:
                    height, width = img.height, img.width
                    size = XDRPositiveSize2D(
                        pixels_to_EMU(width), pixels_to_EMU(height)
                    )

                    margin_left_EMU = cm_to_EMU(margin_left_cm)
                    margin_top_EMU = cm_to_EMU(margin_top_cm)

                    marker = AnchorMarker(
                        col=cell.column - 1,
                        colOff=margin_left_EMU,
                        row=cell.row - 1,
                        rowOff=margin_top_EMU,
                    )
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                else:
                    img.anchor = new_start_cell
                ws.add_image(img)
            except Exception:
                pass

        if json_field.get("border_style", {}):
            cell.border = Border(
                **{
                    key: self.borders.get(key, None)
                    for key, value in json_field.get("border_style").items()
                    if value
                }
            )
        else:
            cell.border = Border(**self.borders)
        cell.alignment = Alignment(**json_field.get("alignment_style", {}))
        cell.font = Font(name="Calibri", size=11, **json_field.get("font_style", {}))
        ws.merge_cells("{}:{}".format(new_start_cell, new_end_cell))
        return

    def resize_images(self, image_path):
        try:
            image = ImagePil.open(image_path)
        except Exception:
            return False

        # width 356px = 9.4cm / height 267px = 7.05cm
        width = 356
        height = 267

        try:
            image = image.resize((width, height))
            image.save(image_path)
        except Exception:
            return False
        else:
            return True

    def download_pictures(self, path, reporting, count):
        files = ReportingFile.objects.filter(
            reporting_id=reporting["uuid"], include_dnit=True
        ).order_by("datetime")

        for index, file_obj in enumerate(files):
            if file_obj.upload:
                try:
                    file_path = file_obj.upload.url.split("?")[0].split(".com/")[1]
                    bucket_name = (
                        file_obj.upload.url.split(".s3")[0].split("/")[-1] + "-400px"
                    )
                    image_format = file_path.split(".")[-1]
                except Exception:
                    return False

                if file_obj.kind:
                    image_name = str(count) + "-{}".format(file_obj.kind)
                else:
                    image_name = str(count) + "-{}".format(str(index))

                if not check_image_file(file_path):
                    return False

                image_path = "{}{}.{}".format(path, image_name, image_format)

                try:
                    self.s3.download_file(bucket_name, unquote(file_path), image_path)
                except Exception:
                    return False
                else:
                    resized = self.resize_images(image_path)
                    if not resized:
                        return False
            else:
                return False
        return True

    def get_excel_name(self):
        now = utc_to_local(datetime.now())
        return "Relatorio_Fotografico_{}_{}_{}".format(now.day, now.month, now.year)

    def upload_file(self, path, name):
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        expires = datetime.now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)
        object_name = "media/private/{}".format(name)

        try:
            self.s3.upload_file(
                path, bucket_name, object_name, ExtraArgs={"Expires": expires}
            )
        except Exception:
            return False

        # Delete file
        os.remove(path)

        url_s3 = self.s3.generate_presigned_url(
            "get_object", Params={"Bucket": bucket_name, "Key": object_name}
        )
        return url_s3

    def get_data(self):
        empty = {"url": "", "name": ""}

        if not self.reportings.exists():
            return empty

        self.company = self.reportings.first().company

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )

        # Load templates
        self.load_data()
        if not self.wb or not self.data or not isinstance(self.data, list):
            return empty

        # Create temporary folder for images
        path = "{}/images/".format(self.temp_path)
        os.makedirs(path, exist_ok=True)

        # Get data from reportings
        self.reportings_data = [
            {
                "uuid": str(item.pk),
                "km": item.km,
                "end_km": item.end_km,
                "road_name": item.road_name or "Sem Rodovia",
                "direction": item.direction,
                "form_data": item.form_data,
                "occurrence_type": str(item.occurrence_type.name),
                "occurrence_type_form": item.occurrence_type.form_fields,
                "created_by": str(item.created_by.get_full_name()),
                "found_at": utc_to_local(item.found_at),
            }
            for item in self.reportings.select_related("occurrence_type", "created_by")
        ]

        # Fill the template excel
        self.write_on_sheet(path)

        # Save the excel
        excel_name = self.get_excel_name()
        self.wb.save("{}/{}.xlsx".format(self.path_excel, excel_name))

        # Delete images
        for file_name in os.listdir(path):
            os.remove(path + file_name)

        # Delete temporary folder
        os.rmdir(path)

        # Upload files
        if len(os.listdir(self.path_excel)) == 0:
            return empty
        else:
            final_filename = (
                os.listdir(self.path_excel)[0] if os.listdir(self.path_excel) else ""
            )
            if final_filename:
                path_total = self.path_excel + "/" + final_filename
                url = self.upload_file(path_total, final_filename)
            else:
                final_filename = ""
                url = False

        if not url or not final_filename:
            return empty

        # Delete excel folder
        for file_name in os.listdir(self.path_excel):
            os.remove(self.path_excel + "/" + file_name)
        os.rmdir(self.path_excel)

        return {"url": url, "name": final_filename}
