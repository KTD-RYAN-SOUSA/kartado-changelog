import shutil
import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Firm, Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_km_plus_meter,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data_selected_option
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string


class XlsxHandler:
    def __init__(
        self,
        s3,
        list_reporting: List[Reporting],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.__sheet_target = sheet_target
        self.__list_reporting = list_reporting
        self.__xlsx_file = (
            "./fixtures/reports/ccr_report_road_stud_horizontal_signage.xlsx"
        )
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Tachas e Tachoes"]

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="Q1:S3",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B3",
            resize_method=ResizeMethod.ProportionalLeft,
        )

    def fill_sheet(self, data_list: list):
        data_dict = {}
        order_dict = {}
        for data in data_list:
            if data["road_name"] not in data_dict.keys():
                data_dict[data["road_name"]] = [data]
                order_dict[data["road_name"]] = [data["direction"]]
            else:
                data_dict[data["road_name"]].append(data)
                order_dict[data["road_name"]].append(data["direction"])
        for k, v in order_dict.items():
            __direction = list(set(v))
            if len(__direction) == 1:
                if str(__direction[0]).lower() == "sul":
                    new_list = sorted(
                        data_dict[k], key=lambda x: x["km"], reverse=False
                    )
                    data_dict[k] = new_list
                elif str(__direction[0]).lower() == "norte":
                    new_list = sorted(data_dict[k], key=lambda x: x["km"], reverse=True)
                    data_dict[k] = new_list
                else:
                    new_list = sorted(
                        data_dict[k], key=lambda x: (x["direction"], x["km"])
                    )
                    data_dict[k] = new_list
            else:
                new_list = sorted(data_dict[k], key=lambda x: (x["direction"], x["km"]))
                data_dict[k] = new_list

        bold = Font(bold=True)
        files = []
        for road, datalist in data_dict.items():
            init_row = 7
            lanes = []
            direction = []
            team_list = []
            subcompany_list = []
            company_list = []
            dates = []
            logo = ""
            for data in datalist:
                team_list.append(data["team"])
                subcompany_list.append(data["subcompany"])
                company_list.append(data["company"])
                if not logo:
                    logo = get_logo_file(
                        self.s3, temp_prefix=self.temp_file, reporting=data["reporting"]
                    )
                if data["executed_at"]:
                    dates.append(data["executed_at"])
                else:
                    dates.append(data["found_at"])
                if data["lane"] not in lanes:
                    lanes.append(data["lane"])
                if data["direction"] not in direction:
                    direction.append(data["direction"])
                len_station_one = len(data["station_one"])
                len_station_two = len(data["station_two"])
                len_station_three = len(data["station_three"])
                segment_rows = len_station_one + len_station_two + len_station_three
                if segment_rows > 1:
                    self._worksheet.merge_cells(
                        f"A{init_row}:A{init_row + segment_rows - 1}"
                    )
                self._worksheet[f"A{init_row}"].alignment = Alignment(
                    textRotation=90, horizontal="center", vertical="center"
                )
                self._worksheet[f"A{init_row}"].font = bold
                self._worksheet[f"A{init_row}"].value = data["stretch"]

                for intern_col in ["B", "S"]:
                    if len_station_one > 1:
                        num = init_row
                        add = len_station_one
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                    if len_station_two > 1:
                        num = init_row + len_station_one
                        add = len_station_two
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                    if len_station_three > 1:
                        num = init_row + len_station_one + len_station_two
                        add = len_station_three
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                all_cols = [chr(i) for i in range(ord("A"), ord("T"))]
                for col in all_cols:
                    for row in range(
                        init_row,
                        init_row + segment_rows,
                    ):
                        self._worksheet[f"{col}{row}"].border = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000"),
                        )
                        self._worksheet.row_dimensions[row].height = 24.75
                if len_station_one:
                    self._worksheet[f"B{init_row}"].value = data["km_station_one"]
                    self._worksheet[f"S{init_row}"].value = data["notes_one"]
                    init = init_row
                    for item in data["station_one"]:
                        __general_appearance = (
                            str(item["general_appearance"]).upper() == "BOA"
                        )
                        self._worksheet[f"C{init}"].value = item["sinalization_lane"]
                        if __general_appearance:
                            self._worksheet[f"D{init}"].value = "X"
                        elif str(item["general_appearance"]).upper() == "REGULAR":
                            self._worksheet[f"E{init}"].value = "X"
                        elif str(item["general_appearance"]).upper() == "RUIM":
                            self._worksheet[f"F{init}"].value = "X"

                        if str(item["station_color"]).upper() == "BRANCA":
                            self._worksheet[f"G{init}"].value = "X"
                        elif str(item["station_color"]).upper() == "AMARELA":
                            self._worksheet[f"H{init}"].value = "X"
                        elif str(item["station_color"]).upper() == "VERMELHA":
                            self._worksheet[f"I{init}"].value = "X"

                        if str(item["body"]).upper() == "BOM" or __general_appearance:
                            self._worksheet[f"J{init}"].value = "X"
                        elif str(item["body"]).upper() == "TRINCADO":
                            self._worksheet[f"K{init}"].value = "X"
                        elif str(item["body"]).upper() == "QUEBRADO":
                            self._worksheet[f"L{init}"].value = "X"

                        if str(item["lens"]).upper() == "BOA" or __general_appearance:
                            self._worksheet[f"M{init}"].value = "X"
                        elif str(item["lens"]).upper() == "MANCHADA":
                            self._worksheet[f"N{init}"].value = "X"
                        elif str(item["lens"]).upper() == "OPACA/SUJA":
                            self._worksheet[f"O{init}"].value = "X"

                        if (
                            str(item["fixation"]).upper() == "BOA"
                            or __general_appearance
                        ):
                            self._worksheet[f"P{init}"].value = "X"
                        elif str(item["fixation"]).upper() == "SOLTA":
                            self._worksheet[f"Q{init}"].value = "X"
                        elif str(item["fixation"]).upper() == "AFUNDADA":
                            self._worksheet[f"R{init}"].value = "X"

                        init += 1
                if len_station_two:
                    count_station_two = init_row + len_station_one
                    self._worksheet[f"B{count_station_two}"].value = data[
                        "km_station_two"
                    ]
                    self._worksheet[f"S{count_station_two}"].value = data[
                        "notes_station_two"
                    ]
                    init = count_station_two
                    for item in data["station_two"]:
                        __general_appearance = (
                            str(item["general_appearance"]).upper() == "BOA"
                        )
                        self._worksheet[f"C{init}"].value = item["sinalization_lane"]
                        if str(item["general_appearance"]).upper() == "BOA":
                            self._worksheet[f"D{init}"].value = "X"
                        elif str(item["general_appearance"]).upper() == "REGULAR":
                            self._worksheet[f"E{init}"].value = "X"
                        elif str(item["general_appearance"]).upper() == "RUIM":
                            self._worksheet[f"F{init}"].value = "X"

                        if str(item["station_color"]).upper() == "BRANCA":
                            self._worksheet[f"G{init}"].value = "X"
                        elif str(item["station_color"]).upper() == "AMARELA":
                            self._worksheet[f"H{init}"].value = "X"
                        elif str(item["station_color"]).upper() == "VERMELHA":
                            self._worksheet[f"I{init}"].value = "X"

                        if str(item["body"]).upper() == "BOM" or __general_appearance:
                            self._worksheet[f"J{init}"].value = "X"
                        elif str(item["body"]).upper() == "TRINCADO":
                            self._worksheet[f"K{init}"].value = "X"
                        elif str(item["body"]).upper() == "QUEBRADO":
                            self._worksheet[f"L{init}"].value = "X"

                        if str(item["lens"]).upper() == "BOA" or __general_appearance:
                            self._worksheet[f"M{init}"].value = "X"
                        elif str(item["lens"]).upper() == "MANCHADA":
                            self._worksheet[f"N{init}"].value = "X"
                        elif str(item["lens"]).upper() == "OPACA/SUJA":
                            self._worksheet[f"O{init}"].value = "X"

                        if (
                            str(item["fixation"]).upper() == "BOA"
                            or __general_appearance
                        ):
                            self._worksheet[f"P{init}"].value = "X"
                        elif str(item["fixation"]).upper() == "SOLTA":
                            self._worksheet[f"Q{init}"].value = "X"
                        elif str(item["fixation"]).upper() == "AFUNDADA":
                            self._worksheet[f"R{init}"].value = "X"
                        init += 1
                if len_station_three:
                    count_station_three = init_row + len_station_one + len_station_two
                    self._worksheet[f"B{count_station_three}"].value = data[
                        "km_station_three"
                    ]
                    self._worksheet[f"S{count_station_three}"].value = data[
                        "notes_station_three"
                    ]
                    init = count_station_three
                    for item in data["station_three"]:
                        self._worksheet[f"C{init}"].value = item["sinalization_lane"]
                        __general_appearance = (
                            str(item["general_appearance"]).upper() == "BOA"
                        )
                        if __general_appearance:
                            self._worksheet[f"D{init}"].value = "X"
                        elif str(item["general_appearance"]).upper() == "REGULAR":
                            self._worksheet[f"E{init}"].value = "X"
                        elif str(item["general_appearance"]).upper() == "RUIM":
                            self._worksheet[f"F{init}"].value = "X"

                        if str(item["station_color"]).upper() == "BRANCA":
                            self._worksheet[f"G{init}"].value = "X"
                        elif str(item["station_color"]).upper() == "AMARELA":
                            self._worksheet[f"H{init}"].value = "X"
                        elif str(item["station_color"]).upper() == "VERMELHA":
                            self._worksheet[f"I{init}"].value = "X"

                        if str(item["body"]).upper() == "BOM" or __general_appearance:
                            self._worksheet[f"J{init}"].value = "X"
                        elif str(item["body"]).upper() == "TRINCADO":
                            self._worksheet[f"K{init}"].value = "X"
                        elif str(item["body"]).upper() == "QUEBRADO":
                            self._worksheet[f"L{init}"].value = "X"

                        if str(item["lens"]).upper() == "BOA" or __general_appearance:
                            self._worksheet[f"M{init}"].value = "X"
                        elif str(item["lens"]).upper() == "MANCHADA":
                            self._worksheet[f"N{init}"].value = "X"
                        elif str(item["lens"]).upper() == "OPACA/SUJA":
                            self._worksheet[f"O{init}"].value = "X"

                        if (
                            str(item["fixation"]).upper() == "BOA"
                            or __general_appearance
                        ):
                            self._worksheet[f"P{init}"].value = "X"
                        elif str(item["fixation"]).upper() == "SOLTA":
                            self._worksheet[f"Q{init}"].value = "X"
                        elif str(item["fixation"]).upper() == "AFUNDADA":
                            self._worksheet[f"R{init}"].value = "X"
                        init += 1
                specific_cols = [chr(i) for i in range(ord("B"), ord("T"))]
                for col in specific_cols:
                    for row in range(
                        init_row,
                        init_row + segment_rows,
                    ):
                        self._worksheet[f"{col}{row}"].alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                init_row += segment_rows

            team_list = list(set(team_list))
            query_set_teams = Firm.objects.filter(uuid__in=team_list).all()
            names = []
            for team in query_set_teams:
                users_query_set = team.users.all()
                intern_list = []
                for user in users_query_set:
                    intern_list.append(user.get_full_name())
                names.append(intern_list)
            names = [", ".join(_) for _ in names]
            team_text = f"Operador: {' - '.join(names)}"
            subcompany_text = f"Empresa: {' - '.join(list(set(subcompany_list)))}"
            company_text = f"Concessionária: {' - '.join(list(set(company_list)))}"
            road_text = f"Rodovia: {data['road_name']}"
            reporting_text = f"Monitoração: Tachas e Tachões - Expressa {' - '.join(list(set(direction)))}"
            if len(dates) > 1:
                dates_new = sorted(
                    dates,
                    key=lambda x: (x.split("/")[2], x.split("/")[1], x.split("/")[0]),
                )
                date_text = f"Data: {dates_new[0]} à {dates_new[-1]}"
            elif dates:
                date_text = f"Data: {dates[0]}"
            else:
                date_text = "Data: "

            self._worksheet["C1"] = subcompany_text
            self._worksheet["C2"] = team_text
            self._worksheet["C3"] = date_text
            self._worksheet["H1"] = road_text
            self._worksheet["H2"] = reporting_text
            self._worksheet["H3"] = company_text

            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                target=self.__sheet_target,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
            )

            result = f"Fichas de Tachas {road} {' - '.join(list(set(lanes)))} {' - '.join(direction)}.xlsx"
            result = "/tmp/" + result.replace("/", "-")
            files.append(result)
            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook["Tachas e Tachoes"]
        return {"files": files}

    def __find_station(self, reporting, station: list, position: int) -> list:
        datalist = []
        if station:
            for obj in station:
                first_station = obj
                if position == 1:
                    _id = "1"
                    pos = ""
                    ordinal = ""
                elif position == 2:
                    _id = "2"
                    pos = "_two"
                    ordinal = "_second"
                else:
                    _id = "3"
                    pos = "_three"
                    ordinal = "_third"
                if first_station.get(f"sinalization_lane{pos}"):
                    sinalization_lane = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__sinalizationLane",
                        first_station.get(f"sinalization_lane{pos}"),
                    )
                    station_color = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__stationColor",
                        first_station.get(f"station_color{pos}"),
                    )
                    body = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__body",
                        first_station.get(f"body{ordinal}"),
                    )
                    general_appearance = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__generalAppearance",
                        first_station.get(f"general_appearance{ordinal}"),
                    )
                    lens = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__lens",
                        first_station.get(f"lens{ordinal}"),
                    )
                    fixation = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__fixation",
                        first_station.get(f"fixation{ordinal}"),
                    )
                    datalist.append(
                        {
                            "sinalization_lane": sinalization_lane,
                            "station_color": station_color,
                            "general_appearance": general_appearance,
                            "body": body,
                            "lens": lens,
                            "fixation": fixation,
                            "id": _id,
                        }
                    )
        return datalist

    def custom_key(self, item):
        try:
            lane: str = item["sinalization_lane"]
            if lane.upper() == "BORDO ESQUERDO":
                return (0, item["sinalization_lane"])
            elif lane.upper() == "BORDO DIREITO":
                return (2, item["sinalization_lane"])
        except Exception:
            pass
        return (1, item["sinalization_lane"])

    def is_dict_empty(self, item):
        return 1 if not item else 0

    def create_dict(self, reporting: Reporting) -> dict:
        km = get_km_plus_meter(km=reporting.km)
        km_end = get_km_plus_meter(km=reporting.end_km)
        km_station_one = get_km_plus_meter(km=reporting.form_data.get("km_station"))
        km_station_two = get_km_plus_meter(km=reporting.form_data.get("km_station_two"))
        km_station_three = get_km_plus_meter(
            km=reporting.form_data.get("km_station_three")
        )
        notes_one = reporting.form_data.get("notes_station_one")
        notes_station_two = reporting.form_data.get("notes_station_two")
        notes_station_three = reporting.form_data.get("notes_station_three")

        station_one = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_one"),
            position=1,
        )
        station_two = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_two"),
            position=2,
        )
        station_three = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_three"),
            position=3,
        )
        station_one = sorted(station_one, key=self.custom_key)
        station_two = sorted(station_two, key=self.custom_key)
        station_three = sorted(station_three, key=self.custom_key)
        allocation = [station_one, station_two, station_three]
        allocation = sorted(allocation, key=lambda x: self.is_dict_empty(x))
        station_one, station_two, station_three = (
            allocation[0],
            allocation[1],
            allocation[2],
        )

        direction = get_custom_option(reporting, "direction")
        obj_ord = {
            "0": {"km": "", "lat": "", "long": "", "notes": ""},
            "1": {
                "km": km_station_one,
                "notes": notes_one,
            },
            "2": {
                "km": km_station_two,
                "notes": notes_station_two,
            },
            "3": {
                "km": km_station_three,
                "notes": notes_station_three,
            },
        }

        data = {
            "stretch": f"Trecho do KM  {km} ao KM {km_end}",
            "subcompany": reporting.firm.subcompany.__dict__.get("name"),
            "team": str(reporting.firm.__dict__.get("uuid")),
            "company": reporting.company.__dict__.get("name"),
            "direction": direction if direction else "Pista",
            "lane": get_custom_option(reporting, "lane"),
            "road_name": reporting.road_name,
            "km": reporting.km,
            "km_station_one": (
                obj_ord[station_one[0]["id"]]["km"] if station_one else ""
            ),
            "km_station_two": (
                obj_ord[station_two[0]["id"]]["km"] if station_two else ""
            ),
            "km_station_three": (
                obj_ord[station_three[0]["id"]]["km"] if station_three else ""
            ),
            "executed_at": (
                reporting.executed_at.strftime("%d/%m/%Y")
                if reporting.executed_at
                else ""
            ),
            "found_at": (
                reporting.found_at.strftime("%d/%m/%Y") if reporting.found_at else ""
            ),
            "station_one": station_one,
            "station_two": station_two,
            "station_three": station_three,
            "notes_one": obj_ord[station_one[0]["id"]]["notes"] if station_one else "",
            "notes_station_two": (
                obj_ord[station_two[0]["id"]]["notes"] if station_two else ""
            ),
            "notes_station_three": (
                obj_ord[station_three[0]["id"]]["notes"] if station_three else ""
            ),
            "reporting": reporting,
        }
        for k, v in data.items():
            if v is None and k not in ["station_one", "station_two", "station_three"]:
                data[k] = ""
        return data

    def execute(self):
        data = []
        for reporting in self.__list_reporting:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRReportRoadStudHorizontalSignage(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""
        road_names = []
        directions = []
        lanes = []
        reportings = Reporting.objects.filter(uuid__in=self.uuids).prefetch_related(
            "occurrence_type", "firm", "firm__subcompany", "company"
        )
        for reporting in reportings:
            lane = ""
            try:
                lane = get_custom_option(reporting, "lane")
            except Exception:
                lane = "Faixa"
            direction = ""
            try:
                direction = get_custom_option(reporting, "direction")
            except Exception:
                direction = "Pista"

            if reporting.road_name not in road_names:
                road_names.append(reporting.road_name)
            if lane not in lanes:
                lanes.append(lane)
            if direction not in directions:
                directions.append(direction)

        file_name = f"Fichas de Tachas {'-'.join(road_names)} {'-'.join(lanes)} {'-'.join(directions)}"
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        if len(road_names) > 1:
            file_name += ".zip"
        else:
            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{file_name}.{extension}"

        return file_name

    def __get_repotings_obj(self):
        form = OccurrenceType.objects.get(
            name="Retrorrefletância Horizontal Longitudinal"
        )
        query_set = Reporting.objects.filter(
            occurrence_type=form, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def export(self):
        list_reporting = self.__get_repotings_obj()
        s3 = get_s3()
        obj = XlsxHandler(s3, list_reporting, self.sheet_target()).execute()
        files = obj["files"]
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_road_stud_horizontal_signage_async_handler(reporter_dict: dict):
    reporter = CCRReportRoadStudHorizontalSignage.from_dict(reporter_dict)
    reporter.export()
