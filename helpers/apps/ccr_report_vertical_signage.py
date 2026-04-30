import shutil
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from threading import Lock
from typing import List, Tuple
from zipfile import ZipFile

from django.db.models import Prefetch
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from zappa.asynchronous import task

from apps.reportings.models import Firm, Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_files_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    form_data_images,
    get_custom_option,
)
from helpers.signals import DisableSignals
from helpers.strings import clean_latin_string, get_obj_from_path


class XlsxHandler:
    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        only_shared: bool = True,
        artesp: bool = False,
    ):
        self.s3 = s3
        self.artesp = artesp
        self.__list_reporting = list_reporting
        self.__sheet_target = sheet_target
        self.temp_file = tempfile.mkdtemp()

        if self.artesp:
            self.__xlsx_file = "./fixtures/reports/ccr_artesp_vertical_signage.xlsx"
        else:
            self.__xlsx_file = "./fixtures/reports/ccr_vertical_signage.xlsx"

        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Medição Vertical"]
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="Y1:Z3",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:A3",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        # Lock para thread-safety ao processar logos
        self._logo_lock = Lock()
        if self.artesp:
            self.__static_fields = {
                "indice_sv": "A",
                "id_intern": "B",
                "km": "C",
                "direction_text": "D",
                "lat": "E",
                "long": "F",
                "code_type": "G",
                "lane_side": "H",
                "photo": "I",
                "legend": "J",
                "fabrication_name": "K",
                "fabrication_date": "L",
                "plate_material": "M",
                "plate_state": "N",
                "support_material": "O",
                "support_state": "P",
                "notes": "Q",
                "width": "R",
                "height": "S",
                "area": "T",
                "pelicule_kind": "U",
                "pelicule_color": "V",
                "m1": "W",
                "m2": "X",
                "m3": "Y",
                "m4": "Z",
                "m5": "AA",
                "average": "AB",
                "average_last_year": "AC",
                "minimal_value": "AD",
                "executed_at_text": "AE",
            }
        else:
            self.__static_fields = {
                "identificacao": "A",
                "lat": "B",
                "long": "C",
                "code_type": "D",
                "photo": "E",
                "fabrication_name": "F",
                "fabrication_date": "G",
                "plate_material": "H",
                "plate_state": "I",
                "support_material": "J",
                "support_state": "K",
                "width": "L",
                "height": "M",
                "area": "N",
                "pelicule_kind": "O",
                "pelicule_color": "P",
                "m1": "Q",
                "m2": "R",
                "m3": "S",
                "m4": "T",
                "m5": "U",
                "average": "V",
                "average_last_year": "W",
                "minimal_value": "X",
                "found_at": "Y",
                "notes": "Z",
            }

        self.__only_shared = only_shared

    def __insert_new_rows(self, row: int):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        chars = [get_column_letter(i) for i in range(1, 21)]
        second_chars = [get_column_letter(i) for i in range(21, 31)]
        for char in chars + ["AD", "AE"]:
            self._worksheet.merge_cells(f"{char}{row}:{char}{row+3}")
        for count in range(0, 4):
            self._worksheet.row_dimensions[row + count].height = 30
            for char in chars + second_chars + ["AD", "AE"]:
                self._worksheet[f"{char}{row + count}"].border = border

    def __order_dict(self, current_dict, keys=["a", "b", "c", "d"]):
        default_dict = {k: None for k in keys}
        for k in default_dict.keys():
            for _k, _v in current_dict.items():
                if _v:
                    default_dict[k] = _v
                    current_dict.pop(_k)
                    break
        return default_dict

    def __update_var(self, current_value, new_value):
        return new_value if not current_value and new_value else current_value

    def fill_sheet(self, data_list: list, selected_direction: bool = True):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        data_dict = {}
        for data in data_list:
            key = f"{data.get('road_name')} - {data.get('implantation_place')} - {data.get('direction')} - {data.get('year')}"
            if key not in data_dict.keys():
                data_dict[key] = [data]
            else:
                data_dict[key].append(data)
        files = []
        filenames = []
        for _key, data in data_dict.items():
            if self.artesp:
                data = sorted(data, key=lambda x: x.get("km"))
            else:
                order_south = [
                    _ for _ in data if str(_.get("direction")).lower() == "sul"
                ]
                order_south = (
                    sorted(order_south, key=lambda x: x.get("km"), reverse=False)
                    if order_south
                    else order_south
                )
                order_north = [
                    _ for _ in data if str(_.get("direction")).lower() == "norte"
                ]
                order_north = (
                    sorted(order_north, key=lambda x: x.get("km"), reverse=True)
                    if order_north
                    else order_north
                )
                order_rest = [
                    _
                    for _ in data
                    if str(_.get("direction")).lower() not in ["sul", "norte"]
                ]
                order_rest = (
                    sorted(
                        order_rest,
                        key=lambda x: (
                            x.get("implantation_place"),
                            x.get("direction"),
                            x.get("km"),
                        ),
                    )
                    if order_rest
                    else order_rest
                )
                data = order_south + order_north + order_rest

            row = 7
            auto_sum = 0
            date_list = []
            team_list = []
            road_name_list = []
            subcompany_list = []
            change_color_list = []
            direction = ""
            implantation_place = ""
            company = ""
            for values in data:
                self.__insert_new_rows(row=row)
                team_list.append(values["team"])
                direction = self.__update_var(
                    current_value=direction, new_value=values.get("direction")
                )
                implantation_place = self.__update_var(
                    current_value=implantation_place,
                    new_value=values.get("implantation_place"),
                )
                company = self.__update_var(
                    current_value=company, new_value=values.get("company")
                )
                for key, value in values.items():
                    if value:
                        if key in [
                            "pelicule_kind",
                            "pelicule_color",
                            "m1",
                            "m2",
                            "m3",
                            "m4",
                            "m5",
                            "average",
                            "average_last_year",
                            "minimal_value",
                        ]:
                            intern_row = row
                            for intern_key, intern_value in value.items():
                                color = "ff0000"
                                if (
                                    "average" == intern_key
                                    and intern_value
                                    and values["minimal_value"]["minimal_value"]
                                ):
                                    change_color_list.append(
                                        f"{self.__static_fields[key]}{intern_row}"
                                        if float(intern_value)
                                        < float(
                                            values["minimal_value"]["minimal_value"]
                                        )
                                        else ""
                                    )
                                elif (
                                    "average_two" == intern_key
                                    and intern_value
                                    and values["minimal_value"]["minimal_value_two"]
                                ):
                                    change_color_list.append(
                                        f"{self.__static_fields[key]}{intern_row}"
                                        if float(intern_value)
                                        < float(
                                            values["minimal_value"]["minimal_value_two"]
                                        )
                                        else ""
                                    )
                                elif (
                                    "average_three" == intern_key
                                    and intern_value
                                    and values["minimal_value"]["minimal_value_three"]
                                ):
                                    change_color_list.append(
                                        f"{self.__static_fields[key]}{intern_row}"
                                        if float(intern_value)
                                        < float(
                                            values["minimal_value"][
                                                "minimal_value_three"
                                            ]
                                        )
                                        else ""
                                    )
                                elif (
                                    "average_four" == intern_key
                                    and intern_value
                                    and values["minimal_value"]["minimal_value_four"]
                                ):
                                    change_color_list.append(
                                        f"{self.__static_fields[key]}{intern_row}"
                                        if float(intern_value)
                                        < float(
                                            values["minimal_value"][
                                                "minimal_value_four"
                                            ]
                                        )
                                        else ""
                                    )
                                insert_centered_value(
                                    self._worksheet,
                                    value=intern_value,
                                    cell=f"{self.__static_fields[key]}{intern_row}",
                                )
                                intern_row += 1
                        elif key in ["plate_state", "support_state"]:
                            if key == "plate_state":
                                color = (
                                    "ff0000" if values["plate_state_id"] == "3" else ""
                                )
                            else:
                                color = (
                                    "ff0000"
                                    if values["support_state_id"] == "3"
                                    else ""
                                )
                            insert_centered_value(
                                self._worksheet,
                                value=str(value).upper(),
                                cell=f"{self.__static_fields[key]}{row}",
                            )
                            if color:
                                self._worksheet[
                                    f"{self.__static_fields[key]}{row}"
                                ].font = Font(color=color)
                        elif key in ["photo"] and value:
                            range_str = f"I{row}:I{row+3}"
                            insert_picture_2(
                                self._worksheet,
                                range_str,
                                Image(value),
                                self.__sheet_target,
                                (1, 1, 1, 1),
                                ResizeMethod.ProportionalCentered,
                            )
                        elif key in ["width", "height", "area"] and value:
                            self._worksheet[
                                f"{self.__static_fields[key]}{row}"
                            ].alignment = Alignment(
                                horizontal="center", vertical="center", wrapText=True
                            )
                            self._worksheet[f"{self.__static_fields[key]}{row}"] = value
                            self._worksheet[
                                f"{self.__static_fields[key]}{row}"
                            ].number_format = "0.00"
                            self._worksheet[
                                f"{self.__static_fields[key]}{row}"
                            ].data_type = "n"
                            auto_sum += float(value) if key == "area" else 0
                        elif key in self.__static_fields.keys():
                            self._worksheet[f"{self.__static_fields[key]}{row}"] = value
                            self._worksheet[
                                f"{self.__static_fields[key]}{row}"
                            ].alignment = Alignment(
                                horizontal="center", vertical="center", wrapText=True
                            )
                        elif key == "executed_at":
                            date_list.append(value)
                        elif key == "road_name":
                            road_name_list.append(value)
                        elif key == "subcompany":
                            subcompany_list.append(value)
                row += 4

            team_list = list(set(team_list))
            query_set_teams = Firm.objects.filter(uuid__in=team_list).all()
            names = []
            for team in query_set_teams:
                users_query_set = team.users.all()
                intern_list = []
                for user in users_query_set:
                    intern_list.append(user.get_full_name())
                names.append(intern_list)
            names = [", ".join(_) for _ in names]
            insert_centered_value(
                self._worksheet, value=" / ".join(names), cell="G2", horizontal="left"
            )
            date_list.sort()
            if len(date_list) == 0:
                date_text = ""
            elif len(date_list) == 1:
                date_text = date_list[0].strftime("%d/%m/%Y")
            else:
                date_text = f"{date_list[0].strftime('%d/%m/%Y')} até {date_list[-1].strftime('%d/%m/%Y')}"

            insert_centered_value(
                self._worksheet, value=date_text, cell="G3", horizontal="left"
            )
            road_name = list(set(road_name_list))
            insert_centered_value(
                self._worksheet,
                value=" / ".join(road_name),
                cell="O1",
                horizontal="left",
            )
            monitoroting_text = (
                f"Sinalização Vertical - Pista - {direction} - {implantation_place}"
            )
            insert_centered_value(
                self._worksheet, value=monitoroting_text, cell="O2", horizontal="left"
            )
            insert_centered_value(
                self._worksheet,
                value=company,
                cell="O3",
                horizontal="left",
            )
            subcompany = list(set(subcompany_list))
            insert_centered_value(
                self._worksheet,
                value=" / ".join(subcompany),
                cell="G1",
                horizontal="left",
            )
            self._worksheet.merge_cells("Y1:Z4")
            self._worksheet.merge_cells(f"L{row}:M{row}")
            insert_centered_value(self._worksheet, value="Total (m²)", cell=f"L{row}")
            self._worksheet[f"L{row}"].border = border
            self._worksheet[f"M{row}"].border = border
            self._worksheet.merge_cells(f"L{row+1}:M{row+1}")
            insert_centered_value(
                self._worksheet, value="Quantidade de placas", cell=f"L{row+1}"
            )
            self._worksheet[f"L{row+1}"].border = border
            self._worksheet[f"M{row+1}"].border = border
            self._worksheet[f"N{row}"].border = border
            insert_centered_value(self._worksheet, value=auto_sum, cell=f"N{row}")
            self._worksheet[f"N{row}"].number_format = "0.00"
            self._worksheet[f"N{row+1}"].border = border
            insert_centered_value(self._worksheet, value=len(data), cell=f"N{row+1}")
            for change_color in change_color_list:
                if change_color:
                    self._worksheet[change_color].font = Font(color="ff0000")
            name_list = []
            repeat = []
            lanes = []
            directions = []
            implantation_places = []
            for values in data:
                road_name = values["road_name"]
                lane = values["lane"]
                direction = values["direction"]
                implantation_place = values["implantation_place"]
                if values and f"{road_name} {lane} {direction}" not in repeat:
                    if lane:
                        lanes.append(lane)
                    if direction:
                        directions.append(direction)
                    if implantation_place:
                        implantation_places.append(implantation_place)
                    filenames.append(_key)
                    repeat.append(f"{road_name} {lane} {direction}")
            lanes = "-".join(list(set(lanes))) if lanes else ""
            directions = "-".join(list(set(directions))) if directions else ""
            direction_text = directions if selected_direction else "Pista"
            implantation_places = (
                "-".join(list(set(implantation_places))) if implantation_places else ""
            )
            file_name_options = [
                f"{road_name} {direction_text} {implantation_places} {values['year']}",
                values["number"],
                str(time.time()),
            ]
            file_name = clean_latin_string(
                next(a for a in file_name_options if a)
                .replace(".", "")
                .replace("/", "")
            )
            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                target=self.__sheet_target,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
            )

            name_list.append(file_name)
            file_name = " - ".join(name_list)
            result = f"/tmp/{file_name}.xlsx"
            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook["Medição Vertical"]
            files.append(result)
        filenames = list(set(filenames))
        filenames = (
            sorted(filenames, key=lambda x: x[0], reverse=False)
            if filenames
            else filenames
        )
        return {"files": files, "names": filenames}

    def create_dict(self, reporting: Reporting, s3) -> dict:
        sign_dict = {
            "1": {
                "field_name": "sign_code_indicacao",
                "data_name": "signCodeIndicacao",
            },
            "2": {
                "field_name": "sign_code_advertencia",
                "data_name": "signCodeAdvertencia",
            },
            "3": {
                "field_name": "sign_code_regulamentacao",
                "data_name": "signCodeRegulamentacao",
            },
            "5": {
                "field_name": "sign_code_dispositivos",
                "data_name": "signCodeDispositivos",
            },
        }
        form_data_rf_uuids = form_data_images(reporting)
        reporting_files_list = list(reporting.reporting_files.all())
        reporting_files_list = [
            rf for rf in reporting_files_list if str(rf.uuid) not in form_data_rf_uuids
        ]
        reporting_files = download_files_pictures(
            s3=s3,
            path=self.temp_file,
            files=reporting_files_list,
            width=0,
            height=0,
            limit=1,
            low_quality=True,
        )

        photo = next(iter(reporting_files["images"]), {}).get("path", [])
        second_sign = bool(reporting.form_data.get("second_sign"))
        sign_type = reporting.form_data.get("sign_type")

        if sign_type and (sign_type in sign_dict) and (second_sign is False):
            sign = new_get_form_data(
                reporting,
                api_path=sign_dict[sign_type].get("data_name"),
                name_path=sign_dict[sign_type].get("field_name"),
            )
        elif sign_type == "4":
            field = "Marcador Quilométrico"
            if second_sign is False:
                sign = field
            else:
                sign_type2 = reporting.form_data.get("sign_type2")

                if sign_type2 in sign_dict:
                    field_2 = new_get_form_data(
                        reporting,
                        api_path=sign_dict[sign_type2].get("data_name"),
                        name_path=f"{sign_dict[sign_type2].get('field_name')}2",
                    )
                elif sign_type2 == "4":
                    field_2 = field
                else:
                    field_2 = ""
                sign = f"{field} - {field_2}"
        elif sign_type and (sign_type in sign_dict) and (second_sign is True):
            field = "Marcador Quilométrico"
            sign_type2 = reporting.form_data.get("sign_type2")
            if sign_type2 in sign_dict:
                field_2 = new_get_form_data(
                    reporting,
                    api_path=sign_dict[sign_type2].get("data_name"),
                    name_path=f"{sign_dict[sign_type2].get('field_name')}2",
                )
            elif sign_type2 == "4":
                field_2 = field
            else:
                field_2 = ""
            field = new_get_form_data(
                reporting,
                api_path=sign_dict[sign_type].get("data_name"),
                name_path=sign_dict[sign_type].get("field_name"),
            )
            sign = f"{field} - {field_2}"
        else:
            sign = ""

        plate_material = new_get_form_data(reporting, "plateMaterial")
        plate_state = new_get_form_data(reporting, "plateState")

        support_material = new_get_form_data(reporting, "supportMaterial")
        support_state = new_get_form_data(reporting, "supportState")

        pelicule_kind = {
            "pelicule_kind_one": new_get_form_data(reporting, "peliculeKind"),
            "pelicule_kind_two": new_get_form_data(reporting, "peliculeKindTwo"),
            "pelicule_kind_three": new_get_form_data(reporting, "peliculeKindThree"),
            "pelicule_kind_four": new_get_form_data(reporting, "peliculeKindFour"),
        }
        pelicule_kind = self.__order_dict(
            current_dict=pelicule_kind, keys=list(pelicule_kind.keys())
        )

        peliculeColor = new_get_form_data(reporting, "peliculeColor")
        peliculeColorTwo = new_get_form_data(reporting, "peliculeColorTwo")
        peliculeColorThree = new_get_form_data(reporting, "peliculeColorThree")
        peliculeColorFour = new_get_form_data(reporting, "peliculeColorFour")

        pelicule_color = {
            "pelicule_color": peliculeColor,
            "pelicule_color_two": peliculeColorTwo,
            "pelicule_color_three": peliculeColorThree,
            "pelicule_color_four": peliculeColorFour,
        }
        color_ids = {
            peliculeColor: reporting.form_data.get("pelicule_color"),
            peliculeColorTwo: reporting.form_data.get("pelicule_color_two"),
            peliculeColorThree: reporting.form_data.get("pelicule_color_three"),
            peliculeColorFour: reporting.form_data.get("pelicule_color_four"),
        }
        pelicule_color = self.__order_dict(
            current_dict=pelicule_color, keys=list(pelicule_color.keys())
        )

        number = reporting.number

        found_at = reporting.found_at
        average_last_year = {}
        current_average_last_year = {}
        if reporting.parent:
            parent: Reporting = reporting.parent
            query_set = parent.children.all()
            objs = {}
            for obj in query_set:
                if obj.found_at.year == found_at.year - 1:
                    objs[f"{obj.found_at.month}-{obj.found_at.day}"] = obj
            list_objs = list(sorted(objs.items(), reverse=True))
            if list_objs:
                obj = list_objs[0][1]
                form_data = obj.form_data
                if form_data:
                    current_average_last_year = {
                        "1": [
                            form_data.get("pelicule_kind"),
                            form_data.get("pelicule_color"),
                            form_data.get("average"),
                        ],
                        "2": [
                            form_data.get("pelicule_kind_two"),
                            form_data.get("pelicule_color_two"),
                            form_data.get("average_second"),
                        ],
                        "3": [
                            form_data.get("pelicule_kind_three"),
                            form_data.get("pelicule_color_three"),
                            form_data.get("average_third"),
                        ],
                        "4": [
                            form_data.get("pelicule_kind_four"),
                            form_data.get("pelicule_color_four"),
                            form_data.get("average_fourth"),
                        ],
                    }

        if current_average_last_year:
            one = [
                x[-1]
                for x in current_average_last_year.values()
                if color_ids[pelicule_color["pelicule_color"]] == x[1]
            ]
            two = [
                x[-1]
                for x in current_average_last_year.values()
                if color_ids[pelicule_color["pelicule_color_two"]] == x[1]
            ]
            three = [
                x[-1]
                for x in current_average_last_year.values()
                if color_ids[pelicule_color["pelicule_color_three"]] == x[1]
            ]
            four = [
                x[-1]
                for x in current_average_last_year.values()
                if color_ids[pelicule_color["pelicule_color_four"]] == x[1]
            ]
            average_last_year[f"{pelicule_color['pelicule_color']}"] = (
                one[0] if one else 0
            )
            average_last_year[f"{pelicule_color['pelicule_color_two']}"] = (
                two[0] if two else 0
            )
            average_last_year[f"{pelicule_color['pelicule_color_three']}"] = (
                three[0] if three else 0
            )
            average_last_year[f"{pelicule_color['pelicule_color_four']}"] = (
                four[0] if four else 0
            )

        m1 = {
            "measure_one": reporting.form_data.get("measure_one"),
            "measure_one_two": reporting.form_data.get("measure_one_second"),
            "measure_one_three": reporting.form_data.get("measure_one_third"),
            "measure_one_four": reporting.form_data.get("measure_one_fourth"),
        }
        m1 = self.__order_dict(current_dict=m1, keys=list(m1.keys()))
        m2 = {
            "measure_two": reporting.form_data.get("measure_two"),
            "measure_two_two": reporting.form_data.get("measure_two_second"),
            "measure_two_three": reporting.form_data.get("measure_two_third"),
            "measure_two_four": reporting.form_data.get("measure_two_fourth"),
        }
        m2 = self.__order_dict(current_dict=m2, keys=list(m2.keys()))
        m3 = {
            "measure_three": reporting.form_data.get("measure_three"),
            "measure_three_two": reporting.form_data.get("measure_three_second"),
            "measure_three_three": reporting.form_data.get("measure_three_third"),
            "measure_three_four": reporting.form_data.get("measure_three_fourth"),
        }
        m3 = self.__order_dict(current_dict=m3, keys=list(m3.keys()))
        m4 = {
            "measure_four": reporting.form_data.get("measure_four"),
            "measure_four_two": reporting.form_data.get("measure_four_second"),
            "measure_four_three": reporting.form_data.get("measure_four_third"),
            "measure_four_four": reporting.form_data.get("measure_four_fourth"),
        }
        m4 = self.__order_dict(current_dict=m4, keys=list(m4.keys()))
        m5 = {
            "measure_five": reporting.form_data.get("measure_five"),
            "measure_five_two": reporting.form_data.get("measure_five_second"),
            "measure_five_three": reporting.form_data.get("measure_five_third"),
            "measure_five_four": reporting.form_data.get("measure_five_fourth"),
        }
        m5 = self.__order_dict(current_dict=m5, keys=list(m5.keys()))
        average = {
            "average": reporting.form_data.get("average"),
            "average_two": reporting.form_data.get("average_second"),
            "average_three": reporting.form_data.get("average_third"),
            "average_four": reporting.form_data.get("average_fourth"),
        }
        average = self.__order_dict(current_dict=average, keys=list(average.keys()))
        minimal_value = {
            "minimal_value": reporting.form_data.get("minimal_value"),
            "minimal_value_two": reporting.form_data.get("minimal_value_two"),
            "minimal_value_three": reporting.form_data.get("minimal_value_three"),
            "minimal_value_four": reporting.form_data.get("minimal_value_four"),
        }
        minimal_value = self.__order_dict(
            current_dict=minimal_value, keys=list(minimal_value.keys())
        )

        fabrication_date: str = reporting.form_data.get("fabrication_date")
        try:
            fabrication_date: datetime = datetime.fromisoformat(
                fabrication_date.replace("Z", "+00:00")
            )
        except Exception:
            fabrication_date = None

        lane = get_custom_option(reporting, "lane")
        direction = get_custom_option(reporting, "direction")
        implantation_place = new_get_form_data(reporting, "implantationPlace")
        implantation_place = (
            implantation_place
            if implantation_place
            else reporting.form_data.get("implantation_place")
        )
        width = reporting.form_data.get("width")
        height = reporting.form_data.get("height")

        year = ""
        if reporting.executed_at:
            year = reporting.executed_at.year
        elif reporting.created_at:
            year = reporting.created_at.year
        elif reporting.found_at:
            year = reporting.found_at.year

        executed_at_text = ""
        if reporting.executed_at is not None:
            executed_at_text = reporting.executed_at.strftime("%d/%m/%Y")
        data = {
            "indice_sv": new_get_form_data(reporting, "indiceSv"),
            "id_intern": new_get_form_data(reporting, "idIntern"),
            "lane_side": new_get_form_data(reporting, "laneSide"),
            "legend": new_get_form_data(reporting, "legend"),
            "executed_at_text": executed_at_text,
            "identificacao": number,
            "lat": reporting.form_data.get("lat"),
            "long": reporting.form_data.get("long"),
            "code_type": sign,
            "photo": photo,
            "fabrication_name": reporting.form_data.get("fabrication_name"),
            "fabrication_date": (
                fabrication_date.strftime("%d/%m/%Y") if fabrication_date else ""
            ),
            "plate_material": plate_material,
            "plate_state": plate_state,
            "support_material": support_material,
            "support_state": support_state,
            "width": (
                reporting.form_data.get("diameter")
                if reporting.form_data.get("diameter")
                and reporting.form_data.get("plate_format") == "1"
                else width
            ),
            "height": 0 if reporting.form_data.get("plate_format") == "1" else height,
            "area": reporting.form_data.get("area"),
            "pelicule_kind": pelicule_kind,
            "pelicule_color": pelicule_color,
            "m1": m1,
            "m2": m2,
            "m3": m3,
            "m4": m4,
            "m5": m5,
            "average": average,
            "average_last_year": average_last_year,
            "minimal_value": minimal_value,
            "found_at": reporting.found_at.strftime("%d/%m/%Y"),
            "executed_at": reporting.executed_at,
            "notes": reporting.form_data.get("notes"),
            "team": str(reporting.firm.__dict__.get("uuid")),
            "road_name": reporting.__dict__.get("road_name"),
            "subcompany": reporting.firm.subcompany.__dict__.get("name"),
            "lane": lane,
            "direction": direction,
            "direction_text": direction,
            "implantation_place": implantation_place,
            "number": reporting.number,
            "company": reporting.company.__dict__.get("name"),
            "km": reporting.__dict__.get("km"),
            "plate_state_id": reporting.form_data.get("plate_state"),
            "support_state_id": reporting.form_data.get("support_state"),
            "classification": reporting.form_data.get("classification"),
            "classification_second": reporting.form_data.get("classification_second"),
            "classification_third": reporting.form_data.get("classification_third"),
            "classification_fourth": reporting.form_data.get("classification_fourth"),
            "reporting": reporting,
            "year": year,
        }
        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def _process_single_reporting(
        self, reporting: Reporting, index: int
    ) -> Tuple[int, dict]:
        """
        Processa um único reporting e retorna seu índice e dados.
        Thread-safe para uso com ThreadPoolExecutor.

        Args:
            reporting: Objeto Reporting a ser processado
            index: Índice do reporting na lista original (para manter ordem)

        Returns:
            Tupla (índice, dados_dict)
        """
        # Processar dados do reporting
        reporting_data = self.create_dict(reporting=reporting, s3=self.s3)

        # Processar logos de forma thread-safe
        # Usar lock para evitar múltiplas tentativas simultâneas
        with self._logo_lock:
            # Logo da empresa
            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_logo_company:
                    self.data_logo_company["path_image"] = path_logo_company

            # Logo do fornecedor
            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        return (index, reporting_data)

    def execute(self, use_parallel: bool = False, batch_size: int = 10):
        data = []
        total_reportings = len(self.__list_reporting)

        if use_parallel and total_reportings > 1:
            # Processamento paralelo em batches
            max_workers = min(batch_size, total_reportings)

            # Processar em batches para evitar sobrecarga
            for batch_start in range(0, total_reportings, batch_size):
                batch_end = min(batch_start + batch_size, total_reportings)
                batch_reportings = self.__list_reporting[batch_start:batch_end]

                # Criar índices para manter ordem
                batch_indices = list(range(batch_start, batch_end))

                # Processar batch em paralelo
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # Submeter todas as tarefas do batch
                    future_to_index = {
                        executor.submit(
                            self._process_single_reporting, reporting, idx
                        ): idx
                        for reporting, idx in zip(batch_reportings, batch_indices)
                    }

                    # Coletar resultados conforme completam (mantendo ordem)
                    batch_results = {}
                    for future in as_completed(future_to_index):
                        try:
                            index, reporting_data = future.result()
                            batch_results[index] = reporting_data
                        except Exception as e:
                            # Log erro mas continue processando
                            original_index = future_to_index[future]
                            print(
                                f"Erro ao processar reporting no índice {original_index}: {e}"
                            )
                            raise  # Re-raise para não perder o erro

                # Adicionar resultados do batch em ordem
                for idx in sorted(batch_results.keys()):
                    data.append(batch_results[idx])

        else:
            # Processamento sequencial (fallback ou para debugging)
            for reporting in self.__list_reporting:
                reporting_data = self.create_dict(reporting=reporting, s3=self.s3)
                data.append(reporting_data)

                # Processar logos (sequencial)
                if not self.data_logo_company.get("path_image"):
                    path_logo_company = get_logo_file(
                        s3=self.s3,
                        temp_prefix=self.temp_file,
                        reporting=reporting,
                    )
                    if path_logo_company:
                        self.data_logo_company["path_image"] = path_logo_company

                if not self.data_provider_logo.get("path_image"):
                    path_provider_logo = get_provider_logo_file(
                        s3=self.s3,
                        temp_prefix=self.temp_file,
                        reporting=reporting,
                    )
                    if path_provider_logo:
                        self.data_provider_logo["path_image"] = path_provider_logo

        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRVerticalSignage(CCRReport):
    def __init__(
        self,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
        only_shared: bool = True,
        is_artesp: bool = False,
    ) -> None:
        self.__only_shared = only_shared
        self.__is_artesp = is_artesp
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""
        data_list = self.__obj_filter()

        reporting = Reporting.objects.get(uuid=self.uuids[0])

        file_name_options = [
            get_obj_from_path(reporting.form_data, "id_ccr_antt"),
            reporting.number,
            str(time.time()),
        ]
        file_name = clean_latin_string(
            next(a for a in file_name_options if a).replace(".", "").replace("/", "")
        )
        if len(data_list) == 1:
            extension = ""
            if self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"
            else:
                extension = "pdf"

            file_name = f"{file_name}.{extension}"

        else:
            file_name = f"Monitoração de Sinalização Vertical {str(time.time())}.zip"
            self.file_name = file_name

        return file_name

    def __get_repotings_obj(self):
        rf_queryset = ReportingFile.objects.all()
        if self.__only_shared:
            rf_queryset = rf_queryset.filter(is_shared=True)
        print(f"self.uuids: {self.uuids}")
        with DisableSignals():
            query_set = Reporting.objects.filter(uuid__in=self.uuids).prefetch_related(
                "occurrence_type",
                "firm",
                "firm__subcompany",
                "company",
                "parent",
                "parent__children",
                Prefetch(
                    "reporting_files",
                    queryset=rf_queryset.order_by("datetime", "uploaded_at"),
                ),
            )
            return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def __obj_filter(self):
        list_reporting = self.__get_repotings_obj()
        data_list = []
        for data in list_reporting:
            road_name = data.road_name
            implantation_place = data.form_data.get("implantation_place")
            direction = data.direction
            year = ""
            if data.executed_at:
                year = data.executed_at.year
            elif data.created_at:
                year = data.created_at.year
            elif data.found_at:
                year = data.found_at.year
            key = f"{road_name} - {implantation_place} - {direction} - {year}"
            if key not in data_list:
                data_list.append(key)
        return data_list

    def export(self):
        list_reporting = self.__get_repotings_obj()
        s3 = get_s3()
        obj = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
            only_shared=self.__only_shared,
            artesp=self.__is_artesp,
        ).execute(use_parallel=True, batch_size=10)
        files = obj["files"]
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_vertical_signage_async_handler(reporter_dict: dict):
    reporter = CCRVerticalSignage.from_dict(reporter_dict)
    reporter.export()
