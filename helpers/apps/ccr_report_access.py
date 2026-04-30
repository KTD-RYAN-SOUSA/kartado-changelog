import tempfile
from datetime import datetime
from os.path import isfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_direction_letter,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_picture_2,
    result_photos,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option, get_km
from helpers.strings import clean_latin_string


class XlsxHandlerAccess:

    _LOGO_CELL = "S1:U4"
    _PROVIDER_LOGO_CELL = "A1:G4"

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):

        self.wb: Workbook = None
        self._worksheet: Worksheet = None
        self.action_worksheet: Worksheet = None
        self.__init_wb()
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.list_uuids = list_uuids
        first_reporting = (
            Reporting.objects.filter(uuid=list_uuids[0])
            .prefetch_related("company")
            .first()
        )
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.photo_07 = None
        self.photo_06 = None

        self.static_fields = {
            "name": "F7",
            "delivery_address": "F8",
            "telephone_delivery": "F10",
            "titular_foil": "F11",
            "cartory": "F12",
            "telephone_access": "V8",
            "cep_delivery": "V9",
            "titular_phone": "V10",
            "matricule": "V11",
            "road_name": "E15",
            "city": "E16",
            "km": "N15",
            "direction": "V15",
            "uf": "V16",
            "uf_delivery": "R9",
            "delivery_city": "F9",
            "latitude": "I17",
            "longitude": "U17",
            "id_antt": "W4",
            "access_kind": {
                "Comercial": "F20",
                "Industrial": "K20",
                "Residencial": "O20",
                "Lavoura": "S20",
                "Público": "W20",
                "Outros": "F21",
            },
            "zone": {
                "Rural": "F22",
                "Urbana": "K22",
            },
            "access_sup": {"Pavimentado": "F23", "Não pavimentado": "K23"},
            "access_revest": {
                "Flexível": "F24",
                "Semi-rígido": "K24",
                "Rígido": "O24",
                "Sem revestimento": "S24",
            },
            "access_condition": {
                "Regulares/autorizados, em condições satisfatórias (nada a providenciar)": "F28",
                "Acesso Fechado/ Descaracterizado": "F27",
                "Irregulares/não autorizados, necessitando de analise do projeto perante normas técnicas vigentes (DNIT, ABNT e ANTT)": "F30",
                "Regulares/autorizados, porém necessitando adequação ou conservação (geometria, pavimento, sinalização, etc)": "F29",
                " Irregulares/não autorizados, e que devem ser fechados, pois apresentam problemas de segurança": "F31",
            },
            "observation": "A34",
            "photo_06": "A36",
            "photo_07": "O36",
            "access_kind_other": "I21",
        }

        self.action_static_fields = {
            "id_antt": "W4",
            "name": "F7",
            "road_name": "E10",
            "city": "E11",
            "km": "N10",
            "direction": "V10",
            "uf": "V11",
            "latitude": "I12",
            "longitude": "U12",
        }

    @classmethod
    def _insert_logos(
        cls,
        s3,
        temp_dir: str,
        sheet_target: SheetTarget,
        workbook: Workbook,
        reporting: Reporting,
    ) -> None:
        logo = get_logo_file(s3, temp_dir, reporting)
        provider_logo = get_provider_logo_file(s3, temp_dir, reporting)

        for worksheet in workbook.worksheets:
            try:
                insert_picture_2(
                    worksheet,
                    cls._LOGO_CELL,
                    Image(logo),
                    sheet_target,
                    border_width=(2, 2, 2, 2),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass
            try:
                insert_picture_2(
                    worksheet,
                    cls._PROVIDER_LOGO_CELL,
                    Image(provider_logo),
                    sheet_target,
                    border_width=(2, 2, 2, 2),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass

    def create_dict(self, reporting):
        self.photo_07 = None
        self.photo_06 = None

        uf_delivery = new_get_form_data(reporting, "ufDelivery", default="")

        delivery_city = new_get_form_data(reporting, "deliveryCity", default="")

        result_access_sup = new_get_form_data(reporting, "accessSup")
        access_sup = self.result_access_condition(
            self.static_fields["access_sup"], result_access_sup
        )

        name = new_get_form_data(reporting, "nameSocialReason", default="")

        id_antt = new_get_form_data(reporting, "idCcrAntt", default="")

        city = new_get_form_data(reporting, "city", default="")

        uf = new_get_form_data(reporting, "ufAcess", default="")

        latitude = new_get_form_data(reporting, "lat", default="")

        longitude = new_get_form_data(reporting, "long", default="")

        matricule = new_get_form_data(reporting, "matricule", default="")

        cartory = new_get_form_data(reporting, "cartory")

        titular_phone = new_get_form_data(reporting, "titularTelephone", default="")

        titular_foil = new_get_form_data(reporting, "titularFoil", default="")

        delivery_address = new_get_form_data(reporting, "deliveryAddress", default="")

        telephone_access = new_get_form_data(reporting, "telephoneAccess", default="")

        cep_delivery = new_get_form_data(reporting, "cepDelivery", default="")

        telephone_delivery = new_get_form_data(
            reporting, "telephoneDelivery", default=""
        )

        observations = new_get_form_data(reporting, "observations", default="")

        result_access_kind = new_get_form_data(reporting, "accessKind", default="")

        access_kind = self.result_access_condition(
            self.static_fields["access_kind"], result_access_kind
        )

        access_kind_other = (
            new_get_form_data(reporting, "accessKindOther", default="")
            if result_access_kind == "Outros"
            else ""
        )

        result_zone = new_get_form_data(reporting, "zone")
        zone = self.result_access_condition(self.static_fields["zone"], result_zone)

        result_access_revest = new_get_form_data(reporting, "accessRevest")
        access_revest = self.result_access_condition(
            self.static_fields["access_revest"], result_access_revest
        )

        result_access_items = new_get_form_data(reporting, "accessCondition")
        access_condition = self.result_access_condition(
            self.static_fields["access_condition"], result_access_items
        )

        images_reports = list(reporting.reporting_files.filter().order_by("datetime"))

        reporting_images = new_get_form_data(reporting, "therapy")
        if reporting_images:
            for k, v in reporting_images[0].items():
                if k == "treatment_images":
                    images_reports = [
                        images for images in images_reports if str(images.uuid) not in v
                    ]

        if images_reports:
            try:
                photo_06 = str(images_reports[-2].uuid)
                self.photo_06 = result_photos(
                    s3=self.s3,
                    temp_file=tempfile.mkdtemp(),
                    photo_id=photo_06,
                    width=337,
                    height=242,
                    enable_is_shared_antt=True,
                    enable_include_dnit=False,
                )[0]
            except (IndexError, KeyError):
                self.photo_06 = ""
            try:
                photo_07 = str(images_reports[-1].uuid)
                self.photo_07 = result_photos(
                    s3=self.s3,
                    temp_file=tempfile.mkdtemp(),
                    photo_id=photo_07,
                    width=337,
                    height=242,
                    enable_is_shared_antt=True,
                    enable_include_dnit=False,
                )[0]
            except (IndexError, KeyError):
                self.photo_07 = ""

        result_road_name = reporting.road_name
        road_name = result_road_name if result_road_name else ""

        direction = get_custom_option(reporting, "direction")

        km = get_km(reporting)

        result_therapy = reporting.form_data.get("therapy")
        therapy = result_therapy if result_therapy else ""

        data = {
            "name": name,
            "delivery_address": delivery_address,
            "telephone_delivery": telephone_delivery,
            "titular_foil": titular_foil,
            "cartory": cartory,
            "telephone_access": telephone_access,
            "cep_delivery": cep_delivery,
            "titular_phone": titular_phone,
            "matricule": matricule,
            "road_name": road_name,
            "uf": uf,
            "city": city,
            "km": km,
            "direction": direction,
            "uf_delivery": uf_delivery,
            "delivery_city": delivery_city,
            "latitude": latitude,
            "longitude": longitude,
            "id_antt": id_antt,
            "access_kind": access_kind,
            "access_condition": access_condition,
            "access_sup": access_sup,
            "access_revest": access_revest,
            "access_kind_other": access_kind_other,
            "zone": zone,
            "observation": observations,
            "photo_06": self.photo_06,
            "photo_07": self.photo_07,
            "therapy": therapy,
            "reporting": reporting,
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def fill_actions_date_and_description(self, data_actions: list):
        row = 16
        limit_rows = 0
        count = 0
        step_rows = 7
        for action in data_actions:
            sum_coor = row + count
            description_coordinate = (
                f"D{sum_coor}" if limit_rows <= 6 else f"Q{sum_coor}"
            )
            date_coordinate = f"A{sum_coor}" if limit_rows <= 6 else f"N{sum_coor}"
            column_letter_description = self.action_worksheet[
                description_coordinate
            ].column_letter
            column_letter_date = self.action_worksheet[date_coordinate].column_letter

            cell_merge_description = (
                f"{description_coordinate}:M{sum_coor}"
                if column_letter_description == "D"
                else f"{description_coordinate}:Z{sum_coor}"
            )
            self.action_worksheet.merge_cells(cell_merge_description)

            cell_merge_date = (
                f"{date_coordinate}:C{sum_coor}"
                if column_letter_date == "A"
                else f"{date_coordinate}:P{sum_coor}"
            )

            self.action_worksheet.merge_cells(cell_merge_date)

            occurrence = action.get("occurrence_type", "")
            if occurrence:
                occurrence = OccurrenceType.objects.filter(uuid=occurrence).first()

            self.action_worksheet[description_coordinate] = (
                occurrence.name if occurrence else ""
            )

            XlsxHandlerAccess.format_fonts(
                cell=self.action_worksheet[description_coordinate],
                size=11,
                color="0070C0",
                bold=True,
            )

            date = action.get("action_date", "")
            if date:
                date = datetime.strptime(date, "%Y-%m-%dT%H:%M:%S.%fZ").strftime(
                    "%d/%m/%Y"
                )
            self.action_worksheet[date_coordinate] = date
            XlsxHandlerAccess.format_fonts(
                cell=self.action_worksheet[date_coordinate],
                size=11,
                color="0070C0",
                bold=True,
            )

            if limit_rows == 12:
                limit_rows = 0
                row += step_rows
            limit_rows += 1
            count = 0 if count > 6 else +1

    def fill_action_worksheet(self, key, value):
        if key in [_ for _ in self.action_static_fields.keys()]:
            key_value = self.action_static_fields[key]
            self.action_worksheet[key_value] = value

            if key not in ("name", "latitude", "longitude"):
                return
            XlsxHandlerAccess.format_fonts(
                cell=self.action_worksheet[key_value], size=11, color="0070C0"
            )
            XlsxHandlerAccess.format_fonts(
                cell=self.action_worksheet[key_value],
                size=11,
                horizontal="left",
                color="0070C0",
                bold=True,
            )

    def __init_wb(self):
        self.wb = load_workbook("./fixtures/reports/ccr_access.xlsx")
        self._worksheet = self.wb["Acesso"]
        self.action_worksheet = self.wb["Acoes"]

    def fill_sheet(self, *, data_list: list):
        count = 0
        list_files = list()
        for values in data_list:

            reporting = values.pop("reporting")
            for key, value in values.items():
                if key in [
                    "access_kind",
                    "access_sup",
                    "access_revest",
                    "access_condition",
                    "zone",
                ]:
                    for _key, _value in value.items():
                        key_value = self.static_fields[key][_key]
                        self.insert_status_values(cell=key_value, value=_value)

                elif key == "therapy":
                    self.fill_actions_date_and_description(value)
                elif key in ["photo_06"] and value:
                    range_str = "A36:N36"
                    insert_picture_2(
                        self._worksheet,
                        range_str,
                        Image(value),
                        self.__sheet_target,
                        resize_method=ResizeMethod.ProportionalCentered,
                    )

                elif key in ["photo_07"] and value:
                    range_str = "O36:Z36"
                    insert_picture_2(
                        self._worksheet,
                        range_str,
                        Image(value),
                        self.__sheet_target,
                        resize_method=ResizeMethod.ProportionalCentered,
                    )
                else:
                    self.fill_action_worksheet(key, value)
                    key_value = self.static_fields[key]
                    self._worksheet[key_value] = value
                    self.custom_formats(key, self._worksheet[key_value])

            count += 1
            road_name = values["road_name"]
            uf_access = values["uf"]
            km = values["km"]
            direction = (
                get_direction_letter(values["direction"])
                if values["direction"]
                else "NA"
            )
            file_name_components = [uf_access, road_name, km, direction]
            file_name = " ".join([c for c in file_name_components if c])
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
            file_path = f"/tmp/{file_name}.xlsx"

            i = 1
            while isfile(file_path):
                file_path = f"/tmp/{file_name} ({i}).xlsx"
                i += 1
            temp_dir = tempfile.mkdtemp()
            self._insert_logos(
                self.s3, temp_dir, self.__sheet_target, self.wb, reporting
            )
            self.wb.save(file_path)
            list_files.append(file_path)
            self.__init_wb()
            if len(list_files) > 1:
                list_files = list(set(list_files))
                list_files.sort()
        return list_files

    def custom_formats(self, key, cell):
        if key not in (
            "name",
            "delivery_address",
            "titular_foil",
            "cartory",
            "telephone_delivery",
            "cep_delivery",
            "delivery_city",
            "latitude",
            "longitude",
            "observation",
            "telephone_access",
            "titular_phone",
            "matricule",
            "uf_delivery",
        ):
            XlsxHandlerAccess.format_fonts(
                cell=cell, size=11, color="0070C0", bold=True
            )

        elif key == "observation":
            XlsxHandlerAccess.format_fonts(
                cell=cell,
                size=11,
                vertical="top",
                horizontal="left",
                color="0070C0",
                bold=True,
            )
        else:
            XlsxHandlerAccess.format_fonts(
                cell=cell, size=11, horizontal="left", color="0070C0", bold=True
            )

    def insert_status_values(self, cell, value):
        if value:
            self._worksheet[cell] = "X"
            XlsxHandlerAccess.format_fonts(
                cell=self._worksheet[cell],
                size=11,
                horizontal="center",
                color="0070C0",
            )
        else:
            self._worksheet[cell] = ""

    @classmethod
    def format_fonts(
        cls,
        *,
        cell,
        name="Calibri",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        color="FF000000",
    ) -> None:

        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold, color=color)

    @classmethod
    def result_access_condition(cls, access_conditions: dict, result_get_form_data):
        conditions_dict = dict()

        for access_condition in access_conditions.keys():
            if result_get_form_data == access_condition:
                conditions_dict[access_condition] = "X"
            else:
                conditions_dict[access_condition] = ""
        return conditions_dict

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]
        data = [self.create_dict(reporting=reporting) for reporting in list_reporting]

        result_file = self.fill_sheet(data_list=data)

        return result_file


class CCRAccess(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = "Relatorios ANTT de Monitoracao de Acessos"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerAccess(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_access_async_handler(reporter_dict: dict):
    reporter = CCRAccess.from_dict(reporter_dict)
    reporter.export()
