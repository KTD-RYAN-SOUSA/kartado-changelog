import tempfile
from datetime import timedelta
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km


class XlsxHandlerReportOACAnnexFive(object):
    _LOGO_CELL = "K1"
    _PROVIDER_LOGO_CELL = "A1:B1"

    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.wb = load_workbook("./fixtures/reports/crr_drainage_annex_five.xlsx")
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self._worksheet = self.wb.active
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company"
        )
        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.resume_report_dict = []
        self.static_fields = {
            "id_ccr_annt": "A",
            "initial_km": "B",
            "end_km": "C",
            "direction": "D",
            "type_element": "E",
            "material": "F",
            "environment": "G",
            "monitoring_date": "H",
            "conservation_state": "I",
            "actions": "J",
            "scheduled_execution_date": "K",
        }

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

    @classmethod
    def _insert_logos(
        cls,
        s3,
        temp_dir: str,
        sheet_target: SheetTarget,
        worksheet: Worksheet,
        reporting: Reporting,
    ) -> None:
        logo = get_logo_file(s3, temp_dir, reporting)
        provider_logo = get_provider_logo_file(s3, temp_dir, reporting)

        try:
            insert_picture_2(
                worksheet,
                cls._LOGO_CELL,
                Image(logo),
                sheet_target,
                border_width=(5, 5, 5, 5),
                resize_method=ResizeMethod.ProportionalRight,
            )
        except Exception:
            pass
        try:
            insert_picture_2(
                worksheet,
                cls._PROVIDER_LOGO_CELL,
                Image(provider_logo),
                sheet_target,
                border_width=(5, 5, 5, 5),
                resize_method=ResizeMethod.ProportionalLeft,
            )
        except Exception:
            pass

    def __insert_new_rows(self, row: int):
        chars = [chr(ord("A") + i) for i in range(11)]

        for char in chars:
            self._worksheet[f"{char}{row}"].border = self.border
            self._worksheet.row_dimensions[row].height = 20

    def create_dict(self, reporting):

        result_id_ccr_annt = new_get_form_data(reporting, "idCcrAntt")

        result_type_element = new_get_form_data(reporting, "tipoelemento")

        result_type_environment = new_get_form_data(reporting, "ambiente")

        type_element = (
            XlsxHandlerReportOACAnnexFive.__set_name_type_environment(
                result_type_element
            )
            if result_type_environment
            else ""
        )

        if (
            (type_element is None)
            or (isinstance(type_element, str) and type_element == "")
        ) and result_type_element is not None:
            type_element = result_type_element

        result_material = new_get_form_data(reporting, "material")

        result_conservation_state = new_get_form_data(
            reporting,
            "generalConservationState",
        )

        if result_conservation_state in ["Bom", "Precário"]:
            data = {"Bom": "-", "Precário": "Atuado conforme Anexo VII"}
            scheduled_execution_date = data[result_conservation_state]
        else:
            scheduled_execution_date = self.__add_month(
                reporting, result_type_environment
            )

        direction = get_custom_option(reporting, "direction")
        actions = self.__set_conservation_state(
            reporting, conservation_state=result_conservation_state
        )

        monitoring_date = (
            reporting.executed_at.strftime("%d/%m/%Y") if reporting.executed_at else ""
        )
        found_at = reporting.found_at
        data = {
            "id_ccr_annt": result_id_ccr_annt,
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "end_km", 3),
            "direction": direction,
            "type_element": type_element,
            "material": result_material,
            "environment": result_type_environment,
            "monitoring_date": monitoring_date,
            "conservation_state": result_conservation_state,
            "actions": actions,
            "scheduled_execution_date": scheduled_execution_date,
            "road_name": reporting.road_name,
            "found_at": found_at,
            "reporting": reporting,
        }
        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def __add_month(self, reporting, conservation_state):
        meses_ingles_para_portugues = {
            "Jan": "Jan",
            "Feb": "Fev",
            "Mar": "Mar",
            "Apr": "Abr",
            "May": "Mai",
            "Jun": "Jun",
            "Jul": "Jul",
            "Aug": "Ago",
            "Sep": "Set",
            "Oct": "Out",
            "Nov": "Nov",
            "Dec": "Dez",
        }

        if conservation_state == "Bom":
            return "-"

        if reporting.executed_at:
            current_date = reporting.executed_at
            initial_date = current_date.strftime("%b/%Y")
            format_initial_month = initial_date.split("/")
            new_string_initial_month = meses_ingles_para_portugues[
                format_initial_month[0]
            ]

            new_date = current_date + timedelta(days=90)
            new_date = new_date.strftime("%b/%Y")
            format_month = new_date.split("/")
            new_string_month = meses_ingles_para_portugues[format_month[0]]
            new_date = f"{new_string_initial_month}/{format_initial_month[1]} a {new_string_month}/{format_month[1]}"

            return new_date

    @classmethod
    def __set_name_type_environment(cls, set_name):
        data = {
            "DI": "Dissipador",
            "DR": "Descida D'água",
            "DD": "Descida D'agua",
            "MF": "Meio Fio",
            "SJ": "Sarjeta",
            "VA": "Valeta",
            "TSS": "Transposição de Sarjeta",
        }
        name = [name for name in data.keys() if name == set_name]
        if name:
            return data[name[0]]

    def __set_conservation_state(self, reporting, conservation_state):
        join_result = conservation_state
        limpeza = new_get_form_data(reporting, "material")
        reparar = new_get_form_data(reporting, "reparar")
        implantar = new_get_form_data(reporting, "implantar")
        status = {"IMPLANTAR": implantar, "LIMPEZA": limpeza, "REPARAR": reparar}
        if join_result == "Bom":
            return "MONITORAR"

        result2 = []
        for state, b in status.items():
            if status[state]:
                result2.append(state)
        if result2:
            join_result = "/".join(result2)
        return join_result

    def fill_sheet(self, *, data_dict: dict):
        row = 5
        list_files = list()
        files_list = []
        for key, values_list in data_dict.items():
            files_list.clear()
            for values in values_list:
                count = 0
                for year_key, list_year_values in values.items():
                    count_key_years = 0
                    reporting = None
                    for values_in_year in list_year_values:
                        reporting = values_in_year.pop("reporting")
                        for internal_key, value in values_in_year.items():
                            if internal_key not in ["road_name", "found_at"]:
                                self.__insert_new_rows(row=row)
                                key_value = f"{self.static_fields[internal_key]}{row}"
                                self._worksheet[key_value] = value
                                XlsxHandlerReportOACAnnexFive.__format_fonts(
                                    cell=self._worksheet[key_value], size=10
                                )
                            self._worksheet[
                                "A1"
                            ] = f"Anexo V - Diagnóstico e Ações - Drenagem Superficial -{values_in_year['road_name']}"
                            XlsxHandlerReportOACAnnexFive.__format_fonts(
                                cell=self._worksheet["A1"], size=12, bold=True
                            )

                            self._worksheet[
                                "A2"
                            ] = f"DIAGNÓSTICO E AÇÕES - DRENAGEM SUPERFICIAL -{values_in_year['road_name']}"
                            XlsxHandlerReportOACAnnexFive.__format_fonts(
                                cell=self._worksheet["A2"], size=10
                            )
                        count += 1
                        row += 1
                    count_key_years += 1
                    file_name = "Anexo V - Diagnóstico e Ações Drenagem Superficial - {}_{}".format(
                        key, year_key
                    )
                    file_path = f"/tmp/{file_name}.xlsx"
                    temp_file = tempfile.mkdtemp()
                    self._insert_logos(
                        self.s3,
                        temp_file,
                        self.__sheet_target,
                        self._worksheet,
                        reporting,
                    )
                    self.wb.save(file_path)
                    files_list.append(file_path)

                    if count == len(list_year_values) and len(files_list) == len(
                        values_list
                    ):
                        saved_files = []
                        if self.__report_format == ReportFormat.PDF:
                            saved_files = convert_files_to_pdf(files_list)
                        else:
                            saved_files = files_list

                        self._worksheet.title = key
                        row = 5
                        temp_dict = tempfile.mkdtemp()
                        path_file = f"{temp_dict}/{key}.zip"
                        with ZipFile(path_file, "w") as zipObj:
                            for file in saved_files:
                                zipObj.write(file, file.split("/")[-1])
                        list_files.append(path_file)
                        XlsxHandlerReportOACAnnexFive.__clear_all_data(
                            files_list, self._worksheet, row
                        )
        return list_files

    @classmethod
    def __clear_all_data(cls, file_path, sheet_name, initial_row):
        for files in file_path:
            wb = load_workbook(files)
            sheet = sheet_name

            for row in sheet.iter_rows(
                min_row=initial_row,
                max_row=sheet.max_row,
                min_col=1,
                max_col=sheet.max_column,
            ):
                for cell in row:
                    cell.value = None
                    cell.border = None
            wb.save(files)

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:

        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]
        data = [self.create_dict(reporting=reporting) for reporting in list_reporting]

        for item in data:
            road_name = item.get("road_name", None)
            found_at = item.get("found_at", None)

            if road_name in self.dict_filtered_roads:
                year_dict = next(
                    (
                        d
                        for d in self.dict_filtered_roads[road_name]
                        if found_at.year in d
                    ),
                    None,
                )

                if year_dict:
                    year_dict[found_at.year].append(item)
                else:
                    year_dict = {found_at.year: [item]}
                    self.dict_filtered_roads[road_name].append(year_dict)
            else:
                self.dict_filtered_roads[road_name] = [{found_at.year: [item]}]

        self.dict_filtered_roads = {
            k: v
            for k, v in self.dict_filtered_roads.items()
            if v and any(item for sublist in v for item in sublist)
        }
        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)
        return result_file


class CrrSurfaceDrainageAnnexFive(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = "Anexo V - Diagnóstico e Ações Drenagem Superficial"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerReportOACAnnexFive(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()
        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_oac_annex_five_async_handler(reporter_dict: dict):
    reporter = CrrSurfaceDrainageAnnexFive.from_dict(reporter_dict)
    reporter.export()
