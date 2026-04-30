import logging
import os
import re
import uuid
from collections import OrderedDict, defaultdict
from copy import deepcopy

import sentry_sdk
from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework_json_api import serializers
from sentry_sdk import capture_exception
from zappa.asynchronous import task

from apps.companies.models import Firm
from apps.service_orders.const import status_types
from apps.users.models import User
from helpers.dates import parse_dict_dates_tz, utc_to_local
from helpers.strings import get_obj_from_path


def create_procedure_objects(record):
    from apps.service_orders.models import Procedure, ServiceOrderAction
    from apps.service_orders.serializers import (
        ProcedureSerializer,
        ServiceOrderActionSerializer,
    )

    mandatory_fields = [
        "firm_id",
        "to_do",
        "done_at",
        "service_order_action_status_id",
        "deadline",
    ]
    error_message = ""
    objs = []
    service_orders = []
    procedure_objects = record.form_data.get("procedure_objects", [])
    if procedure_objects and record.form_data.get("include_procedures", False):
        for procedure in procedure_objects:
            copy_procedure = deepcopy(procedure)
            # Validate procedure
            if not set(mandatory_fields).issubset(copy_procedure.keys()):
                error_message = "kartado.error.occurrence_record.no_procedure"
                return record, error_message

            # Validate service_order and action
            service_order = copy_procedure.pop("service_order", {})
            service_order_action = copy_procedure.pop("service_order_action", {})
            if not service_order or not service_order_action:
                error_message = "kartado.error.occurrence_record.no_service_and_action"
                return record, error_message

            if "id" in service_order:
                service_orders.append(service_order["id"])
            else:
                error_message = "kartado.error.occurrence_record.service_not_found"
                return record, error_message

            if "id" in service_order_action:
                copy_procedure["action_id"] = service_order_action.get("id", "")
            elif service_order_action:
                service_order_action["service_order_id"] = service_order.get("id", "")
                service_order_action["parent_record_id"] = str(record.uuid)
                action_parsed = parse_dict_dates_tz(
                    service_order_action, ["estimated_end_date", "opened_at"]
                )
                action_parsed["uuid"] = uuid.uuid4()
                action_serialized_item = ServiceOrderActionSerializer(
                    data=action_parsed
                )
                if action_serialized_item.is_valid():
                    objs.append(ServiceOrderAction(**action_parsed))
                    copy_procedure["action_id"] = action_parsed["uuid"]
                    procedure["service_order_action"]["id"] = str(action_parsed["uuid"])
                else:
                    error_message = "kartado.error.occurrence_record.action_error"
                    return record, error_message
            else:
                error_message = "kartado.error.occurrence_record.action_not_found"
                return record, error_message

            parsed_item = parse_dict_dates_tz(
                copy_procedure, ["created_at", "deadline", "done_at"]
            )
            parsed_item["uuid"] = uuid.uuid4()
            serialized_item = ProcedureSerializer(data=parsed_item)
            if serialized_item.is_valid():
                objs.append(Procedure(**parsed_item))
                procedure["uuid"] = str(parsed_item["uuid"])
            else:
                error_message = "kartado.error.occurrence_record.procedures_error"
                return record, error_message

    if not error_message:
        if procedure_objects:
            record.form_data["objects_created"] = True
        if service_orders:
            record.service_orders.add(*service_orders)
        for item in objs:
            model_name = item._meta.model_name
            print("{} : {}".format(model_name, item.uuid))
            item.save()
            if model_name == "procedure":
                item.occurrence_records.add(record)

    return record, error_message


def pending_procedures_data(procedures, extra_fields=[]):
    all_results = []

    # Get all necessary data of procedures
    for procedure in procedures:
        action = procedure.action
        service_order = action.service_order if action else None
        so_record_form_data = (
            service_order.so_records.first().form_data
            if service_order and service_order.so_records.exists()
            else {}
        )

        try:
            property_intersections = get_obj_from_path(
                so_record_form_data, "propertyintersections"
            )
        except Exception:
            continue
        attributes = (
            property_intersections[0].get("attributes", None)
            if property_intersections
            else None
        )

        construction = (
            get_obj_from_path(
                attributes,
                "obra",
            )
            if attributes
            else None
        )
        sequential = (
            get_obj_from_path(
                attributes,
                "sequencial",
            )
            if attributes
            else None
        )
        identifier = (
            get_obj_from_path(
                attributes,
                "identificador",
            )
            if attributes
            else None
        )

        def translate_kind(in_kind):
            lookup_table = {
                "ENVIRONMENT": "Ambiental",
                "LAND": "Patrimonial",
            }
            if in_kind in lookup_table:
                return lookup_table[in_kind]
            return in_kind

        result = {
            "id": str(procedure.uuid),
            "firm": str(procedure.firm_id),
            "firm_name": procedure.firm.name,
            "description": service_order.description,
            "service_order_id": str(service_order.uuid),
            "service_order_number": service_order.number,
            "service_order_description": service_order.description,
            "kind": translate_kind(service_order.kind),
            "construction": construction if construction else None,
            "sequential": sequential if sequential else None,
            "identifier": identifier if identifier else None,
            "occurrence_records": procedure.occurrence_records.values_list(
                "number", flat=True
            ),
            "action": action.name,
            "action_id": str(action.uuid),
            "to_do": procedure.to_do,
            "responsible": procedure.responsible.get_full_name()
            if procedure.responsible
            else "",
            "deadline": procedure.deadline,
            "delay": (timezone.now() - procedure.deadline).days
            if (timezone.now() - procedure.deadline).days > 1
            else None,
            "rel_url": f"{settings.FRONTEND_URL}/#/SharedLink/Procedure/{procedure.uuid}/show/?company={procedure.firm.company.uuid}",
        }

        if "company" in extra_fields:
            result["company"] = procedure.firm.company.name

        if "opened_at" in extra_fields:
            result["opened_at"] = service_order.opened_at

        all_results.append(result)

    return all_results


def extract_pending_procedures(request, permissions, query_params, user) -> dict:
    """
    Endpoint extract all results to pending procedures

    Raises:
        serializers.ValidationError: permission_not_found
        serializers.ValidationError: permission_denied
        serializers.ValidationError: company_not_found
        serializers.ValidationError: company_not_found

    Returns:
        dict: Returns results if the operation was successful
    """
    from apps.companies.models import Company
    from apps.service_orders.models import Procedure
    from apps.service_orders.views import ProcedureFilter

    results = []

    # Get permission value
    try:
        perm = permissions.get_permission("can_view_pending_procedures")[0]
    except Exception:
        raise serializers.ValidationError(
            "kartado.error.pending_procedures.permission_not_found"
        )

    if perm not in ["all", "firm", "none"]:
        raise serializers.ValidationError(
            "kartado.error.pending_procedures.permission_denied"
        )

    # Get company
    if "company" not in query_params.keys():
        raise serializers.ValidationError(
            "kartado.error.pending_procedures.company_not_found"
        )

    try:
        company = Company.objects.get(pk=query_params.get("company"))
    except Exception:
        raise serializers.ValidationError(
            "kartado.error.pending_procedures.company_not_found"
        )

    # Get firms
    if perm == "all":
        firms = company.company_firms.all()
    elif perm == "firm":
        firms = user.user_firms.all()
    else:
        return Response({"type": "PendingProcedures", "attributes": results})

    # Get pending procedures
    procedures = Procedure.objects.filter(
        action__service_order__company=company,
        service_order_action_status__kind=status_types.ACTION_STATUS,
        service_order_action_status__is_final=False,
        firm__in=firms,
        procedure_next__isnull=True,
    ).select_related("action", "action__service_order", "responsible", "firm")
    procedures = ProcedureFilter(request.GET, queryset=procedures, request=request).qs

    all_results = pending_procedures_data(procedures)

    # Format result
    firm_ids = list(set([item["firm"] for item in all_results]))
    for firm_id in firm_ids:
        procedures_by_firm = list(filter(lambda x: x["firm"] == firm_id, all_results))
        if "sort" in query_params and query_params["sort"] == "-deadline":
            procedures_by_firm.sort(reverse=True, key=lambda x: x["deadline"])
        else:
            procedures_by_firm.sort(key=lambda x: x["deadline"])
        if procedures_by_firm:
            results.append(
                {
                    "id": firm_id,
                    "firm_name": procedures_by_firm[0]["firm_name"],
                    "data": procedures_by_firm,
                }
            )

    # Sort results by firm_name
    results.sort(key=lambda x: x["firm_name"])

    return results


def report_pending_procedures(raw_filters, perm, company, user):
    from apps.service_orders.models import Procedure
    from apps.service_orders.views import ProcedureFilter

    if perm not in ["all", "firm", "none"]:
        raise serializers.ValidationError(
            "kartado.error.pending_procedures.permission_denied"
        )

    # Get firms
    if perm == "all":
        firms = company.company_firms.all()
    elif perm == "firm":
        firms = user.user_firms.all()
    else:
        firms = Firm.objects.none()

    procedures = Procedure.objects.filter(
        action__service_order__company=company,
        service_order_action_status__kind=status_types.ACTION_STATUS,
        service_order_action_status__is_final=False,
        procedure_next__isnull=True,
        firm__in=firms,
    ).select_related("action", "action__service_order", "responsible", "firm")

    # Prepare the filters when they can't be used directly
    proc_filters = {}
    for filter_name, filter_value in raw_filters.items():
        if isinstance(filter_value, list):
            str_values = [str(value) for value in filter_value]
            proc_filters[filter_name] = ",".join(str_values)
        elif isinstance(filter_value, dict):  # Ex: date ranges
            proc_filters.update(
                {
                    f"{filter_name}_{key}": str(value)
                    for key, value in filter_value.items()
                }
            )
        elif isinstance(filter_value, str):
            proc_filters[filter_name] = filter_value
        else:
            logging.error("Unsupported filter value was found")

    # Apply the filters
    procedures = ProcedureFilter(proc_filters, queryset=procedures).qs

    extra_fields = ["company", "opened_at"]
    all_results = pending_procedures_data(procedures, extra_fields)

    return all_results


@task
def generate_pending_procedures_excel_file(
    pending_procedure_export_id,
    user_id,
    permissions,
):
    """
    Gathers all the data needed for the export, fills the excel template
    """

    from apps.service_orders.models import PendingProceduresExport

    try:
        pending_procedure_export = PendingProceduresExport.objects.get(
            pk=pending_procedure_export_id
        )
    except PendingProceduresExport.DoesNotExist as e:
        logging.error("PendingProceduresExport not found")
        capture_exception(e)
    else:
        error = True  # Error until proven otherwise
        filters = pending_procedure_export.filters
        company = pending_procedure_export.company
        user = User.objects.get(uuid=user_id)

        def datetime_to_date(datetime, clear_tzinfo=True):
            try:
                if clear_tzinfo:
                    datetime = utc_to_local(datetime).replace(tzinfo=None)
                date = datetime.date()
            except Exception:
                date = None
            return date

        try:
            # Data gathering
            pending_procedure_datas = report_pending_procedures(
                filters, permissions, company, user
            )
            data_offset = None

            def get_cell(column_letter, list_pos):
                """
                Apply the data_offset to get the cell for the current
                group of data
                """
                if data_offset:
                    return "{}{}".format(column_letter, list_pos + data_offset)

            temp_dir = "/tmp/procedures/"
            os.makedirs(temp_dir, exist_ok=True)
            templ_wb = load_workbook(
                filename="apps/service_orders/templates/Pendências.xlsx",
                read_only=False,
                keep_vba=False,
            )
            source_ws = templ_wb["Equipe 3"]
            # templ_ws = templ_wb.active

            # all colums width must be fixed
            COLUMN_WIDTH = 45

            exported_data = timezone.now()
            company_name = pending_procedure_export.company.name

            data_offset = 2

            # {
            #     "IPT": [
            #         {firm_name: "IPT", ...},
            #         {firm_name: "IPT", ...},
            #         {firm_name: "IPT", ...},
            #     ],
            #     "TOPOGRAFIA": [
            #         {firm_name: "TOPOGRAFIA", ...},
            #     ]
            # }

            firm_pending_procedures = defaultdict(list)

            for dicts in pending_procedure_datas:
                if "firm_name" in dicts:
                    firm_pending_procedures[dicts["firm_name"]].append(dicts)

            # updanting tabs
            for (
                firm_name,
                pending_procedures,
            ) in firm_pending_procedures.items():
                templ_ws = templ_wb.copy_worksheet(source_ws)

                firm_name_pattern = r'[\\/:\*\?"<>\|\[\]]'
                clean_firm_name = re.sub(firm_name_pattern, "", firm_name[:30])
                templ_ws.title = clean_firm_name

                for i, pending_procedure in enumerate(pending_procedures):
                    templ_ws[get_cell("A", i)] = pending_procedure["company"]
                    templ_ws[get_cell("B", i)] = pending_procedure[
                        "service_order_number"
                    ]
                    templ_ws[get_cell("C", i)] = pending_procedure[
                        "service_order_description"
                    ]
                    templ_ws[get_cell("D", i)] = pending_procedure["kind"]
                    templ_ws[get_cell("E", i)] = pending_procedure["construction"]
                    templ_ws[get_cell("F", i)] = pending_procedure["sequential"]
                    templ_ws[get_cell("G", i)] = pending_procedure["identifier"]
                    templ_ws[get_cell("H", i)] = datetime_to_date(
                        pending_procedure["opened_at"]
                    )
                    templ_ws[get_cell("I", i)] = "; ".join(
                        pending_procedure["occurrence_records"]
                    )
                    templ_ws[get_cell("J", i)] = pending_procedure["action"]
                    templ_ws[get_cell("K", i)] = pending_procedure["to_do"]
                    templ_ws[get_cell("L", i)] = pending_procedure["firm_name"]
                    templ_ws[get_cell("M", i)] = pending_procedure["responsible"]
                    templ_ws[get_cell("N", i)] = datetime_to_date(
                        pending_procedure["deadline"]
                    )
                    templ_ws[get_cell("O", i)] = pending_procedure["delay"]
                    templ_ws[get_cell("P", i)] = pending_procedure["rel_url"]
                    templ_ws[get_cell("P", i)].value = '=HYPERLINK("{}", "{}")'.format(
                        pending_procedure["rel_url"], "Link"
                    )

                    if pending_procedure is None:
                        return ""

                    headers = [
                        "A",
                        "B",
                        "C",
                        "D",
                        "E",
                        "F",
                        "G",
                        "H",
                        "I",
                        "J",
                        "K",
                        "L",
                        "M",
                        "N",
                        "O",
                        "P",
                    ]

                    # Set column width
                    for column in headers:
                        templ_ws.column_dimensions[column].width = COLUMN_WIDTH

                    # Set row height
                    for row in range(2, templ_ws.max_row + 1):
                        templ_ws.row_dimensions[row].height = 30

            # sorted results when exist sort in filters
            if "sort" in filters:
                firm_ids = list(set([item["firm"] for item in pending_procedure]))
                for firm_id in firm_ids:
                    procedures_by_firm = list(
                        filter(
                            lambda x: x["firm"] == firm_id,
                            pending_procedure,
                        )
                    )
                    if "sort" in filters and filters["sort"] == "-deadline":
                        procedures_by_firm.sort(
                            reverse=True, key=lambda x: x["deadline"]
                        )
                    else:
                        procedures_by_firm.sort(key=lambda x: x["deadline"])
                    if procedures_by_firm:
                        pending_procedure.append(
                            {
                                "id": firm_id,
                                "firm_name": procedures_by_firm[0]["firm_name"],
                                "data": procedures_by_firm,
                            }
                        )

            del templ_wb["Kartado"]
            del templ_wb["Equipe 2"]
            del templ_wb["Equipe 3"]

            file_name = "{}{}.xlsx".format(temp_dir, "teste")

            ws = templ_wb.active if templ_wb.active else templ_wb.create_sheet()
            ws.sheet_state = "visible"
            templ_wb.save(file_name)

            # Create temp folder to save the excel files
            num_procedures = len(pending_procedure_datas)
            exported_file_name = "Pendências {} - {}.xlsx".format(
                company_name, exported_data
            )

            # Save final file
            if num_procedures >= 1:
                filename = os.listdir(temp_dir)[0]
                ws.sheet_state = "visible"
                with open("{}{}".format(temp_dir, filename), "rb") as excel_file:
                    pending_procedure_export.exported_file.save(
                        exported_file_name,
                        ContentFile(excel_file.read()),
                    )
                    error = False

                # Delete temp files
                os.remove("{}{}".format(temp_dir, filename))
                os.rmdir(temp_dir)
            else:
                filename = os.listdir(temp_dir)[0]
                ws.sheet_state = "visible"
                with open("{}{}".format(temp_dir, filename), "rb") as excel_file:
                    pending_procedure_export.exported_file.save(
                        exported_file_name,
                        ContentFile(excel_file.read()),
                    )
                    error = False

                # Delete temp files
                os.remove("{}{}".format(temp_dir, filename))
                os.rmdir(temp_dir)

        except Exception as e:
            logging.error(
                "Untreated exception found while exporting file. Check Sentry."
            )
            sentry_sdk.capture_exception(e)
            error = True

        # Finish and set flags
        pending_procedure_export.error = error
        pending_procedure_export.done = True
        pending_procedure_export.save()


def handle_resources(self, service_order_resource):
    """
    Handles creating and editing Resources using special fields on a serializer.
    """

    from apps.resources.serializers import ResourceSerializer

    serializer = None
    if "create_resource" in self.initial_data:
        item = self.initial_data["create_resource"]

        # Pass created_by
        if "request" in self.context:
            try:
                user = self.context["request"].user
                user_data = OrderedDict({"type": "User", "id": str(user.pk)})
                item["created_by"] = user_data
                self.initial_data["created_by"] = user_data
            except Exception:
                pass
        elif "created_by" in self.initial_data:
            item["created_by"] = self.initial_data["created_by"]

        # Pass company
        if "company" in self.initial_data:
            item["company"] = self.initial_data["company"]

        serializer = ResourceSerializer(data=item)
    elif "edit_resource" in self.initial_data:
        item = self.initial_data["edit_resource"]

        if "id" in item:
            item_id = item.pop("id")
        else:
            raise serializers.ValidationError(
                "kartado.error.resource.inform_id_when_using_edit_fields"
            )

        # Determine the model
        model = ResourceSerializer.Meta.model

        # Try to get the instance
        try:
            instance = model.objects.get(pk=item_id)
        except model.DoesNotExist:
            raise serializers.ValidationError(
                "kartado.error.resource.invalid_id_on_edit_resource"
            )

        # Determine if the update is partial
        is_partial = self.partial

        # Add instance and item data to serializer
        serializer = ResourceSerializer(
            instance=instance, data=item, partial=is_partial
        )

    # If not valid, errors are returned normally as JSON
    if serializer and serializer.is_valid(raise_exception=True):
        kwargs = {}
        if "created_by" in self.initial_data:
            kwargs["created_by_id"] = self.initial_data["created_by"]["id"]
        resource = serializer.save(**kwargs)

        service_order_resource.resource = resource
        service_order_resource.save()
