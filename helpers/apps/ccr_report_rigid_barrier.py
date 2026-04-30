import re
import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.pagebreak import Break
from zappa.asynchronous import task

from apps.companies.models import Firm
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_s3,
    insert_centered_value,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
    result_photos,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.apps.ccr_resume_description_measures_rigid_barrier import (
    XlsxHandlerResumeReportDescriptionMeasures,
)
from helpers.strings import clean_latin_string, format_km

from .ccr_resume_report_rigid_barrier import XlsxHandlerResumeReportRigidBarrier


class RigidBarrierXlsxHandler(object):
    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.__sheet_target = sheet_target
        self.photo_start = None
        self.photo_end = None
        self.wb = load_workbook("./fixtures/reports/ccr_report_rigid_barrier.xlsx")
        self.s3 = s3
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self._worksheet = self.wb.active
        self._worksheet.title = "Monit. Barreira New Jersey"
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company"
        )

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="V1:Y5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B5",
            resize_method=ResizeMethod.ProportionalLeft,
        )

        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company

        self.static_fields = {
            "year": "A",
            "index": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "latitude": "F",
            "longitude": "G",
            "notes": "H",
            "photo_start": "I",
            "photo_end": "J",
            "general_appearance": {"BOA": "L", "REGULAR": "L", "RUIM": "L"},
            "height": {"ATENDE_NORMA": "N", "NÃO_ATENDE": "N"},
            "length": "O",
            "exposed_armour": {"SIM": "Q", "NÃO": "Q"},
            "desagregation": {"SIM": "S", "NÃO": "S"},
            "cracks": {"SIM": "U", "NÃO": "U"},
            "reflectors": {"AUSENTE": "W", "RUIM": "W", "REGULAR": "W", "BOM": "W"},
            "reflectors_color": {
                "tipo_refletor": "Y",
                "VERMELHO": "Y",
                "BRANCO": "Y",
                "LIMA-LIMÃO": "Y",
            },
        }

    def __insert_new_rows(self, row: int):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        chars = [chr(ord("A") + i) for i in range(10)]
        second_chars = [chr(ord("P") + i) for i in range(10)]

        for char in chars + ["O"]:
            self._worksheet.merge_cells(f"{char}{row}:{char}{row + 3}")
        for count in range(0, 4):
            self._worksheet.row_dimensions[row + count].height = 20
            for char in chars + second_chars + ["K", "M", "L", "N", "O"]:
                self._worksheet[f"{char}{row + count}"].border = border

    def create_dict(self, reporting):
        photos = reporting.form_data.get("mandatory_pictures")
        photo_start = None
        photo_end = None

        if photos:
            for photo in reporting.form_data.get("mandatory_pictures"):
                try:
                    photo_start = (
                        result_photos(
                            s3=self.s3,
                            temp_file=tempfile.mkdtemp(),
                            photo_id=photo["foto_inicio"][0],
                            width=170,
                            height=100,
                            enable_is_shared_antt=True,
                            enable_include_dnit=False,
                        )
                        if photo["foto_inicio"]
                        else ""
                    )
                except Exception:
                    pass

                try:
                    photo_end = (
                        result_photos(
                            s3=self.s3,
                            temp_file=tempfile.mkdtemp(),
                            photo_id=photo["foto_fim"][0],
                            width=170,
                            height=100,
                            enable_is_shared_antt=True,
                            enable_include_dnit=False,
                        )
                        if photo["foto_fim"]
                        else ""
                    )
                except Exception:
                    pass

        result_year = new_get_form_data(reporting, "inspectionCampaignYear")
        result_index = new_get_form_data(reporting, "idCcrAntt")

        result_notes = new_get_form_data(reporting, "notesTwo")
        notes = result_notes if result_notes else ""
        _reflectors_color = new_get_form_data(
            reporting, "corRefletoresBarreiraNj", default=[]
        )
        direction = get_custom_option(reporting, "direction")

        result_type_reflector = new_get_form_data(reporting, "refletoresBarreiraNj")

        result_reflectors_color = {
            "tipo_refletor": (
                result_type_reflector.upper()
                if result_type_reflector in ("Monodirecional", "Bidirecional")
                else ""
            ),
            "VERMELHO": "X" if "Vermelha" in _reflectors_color else "",
            "BRANCO": "X" if "Branca" in _reflectors_color else "",
            "LIMA-LIMÃO": "X" if "Lima-Limão" in _reflectors_color else "",
        }

        result_exposed_armour = reporting.form_data.get("exposed_armour")
        exposed_armour = {
            "SIM": "X" if result_exposed_armour else "",
            "NÃO": "" if result_exposed_armour else "X",
        }
        result_desagregation = reporting.form_data.get("desagregation")

        desagregation = {
            "SIM": "X" if result_desagregation else "",
            "NÃO": "" if result_desagregation else "X",
        }

        result_cracks = reporting.form_data.get("cracks")
        cracks = {
            "SIM": "X" if result_cracks else "",
            "NÃO": "" if result_cracks else "X",
        }

        result_height = new_get_form_data(reporting, "heightBarrier")

        height = {
            "ATENDE_NORMA": "X" if result_height == "Atende" else "",
            "NÃO_ATENDE": "X" if result_height == "Não Atende" else "",
        }

        reflectors = {
            "AUSENTE": "X"
            if RigidBarrierXlsxHandler.__delineator_state(reporting) == "Ausente"
            else "",
            "BOM": "X"
            if RigidBarrierXlsxHandler.__delineator_state(reporting) == "Bom"
            else "",
            "REGULAR": "X"
            if RigidBarrierXlsxHandler.__delineator_state(reporting) == "Regular"
            else "",
            "RUIM": "X"
            if RigidBarrierXlsxHandler.__delineator_state(reporting) == "Ruim"
            else "",
        }

        result_general_appearance = new_get_form_data(reporting, "generalState")

        general_appearance = {
            "BOA": "X" if result_general_appearance == "Bom" else "",
            "REGULAR": "X" if result_general_appearance == "Regular" else "",
            "RUIM": "X" if result_general_appearance == "Ruim" else "",
        }
        data = {
            "year": result_year,
            "index": result_index,
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "km", 3),
            "raw_initial_km": reporting.km,
            "direction": direction.upper(),
            "latitude": reporting.form_data.get("latitude_barreira_nj"),
            "longitude": reporting.form_data.get("longitude_barreira_nj"),
            "notes": notes,
            "photo_start": photo_start[0] if photo_start else "",
            "photo_end": photo_end[0] if photo_end else "",
            "general_appearance": general_appearance,
            "height": height,
            "length": reporting.form_data.get("length_barreira_nj"),
            "exposed_armour": exposed_armour,
            "desagregation": desagregation,
            "cracks": cracks,
            "reflectors": reflectors,
            "reflectors_color": result_reflectors_color,
            "executed_at": reporting.executed_at,
            "team": str(reporting.firm.__dict__.get("uuid")) if reporting.firm else "",
            "road_name": reporting.__dict__.get("road_name"),
            "subcompany": (
                reporting.firm.subcompany.__dict__.get("name") if reporting.firm else ""
            ),
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def __insert_status_values(self, key, cell, value):

        cell_status_string = self._worksheet[cell].offset(0, -1)
        if key == "tipo_refletor":
            cell_status_string.value = value
        else:
            cell_status_string.value = key.replace("_", " ") if key else key
            if value:
                self._worksheet[cell] = "X"
                RigidBarrierXlsxHandler.__format_fonts(
                    cell=self._worksheet[cell], size=10
                )

        RigidBarrierXlsxHandler.__format_fonts(
            cell=cell_status_string, size=10, horizontal="left"
        )

    def fill_sheet(self, *, data_list: list):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 7
        team_list = []
        date_list = []
        road_name_list = []
        subcompany_list = []

        for values in data_list:
            for key, value in values.items():
                self.__insert_new_rows(row=row)
                if key in [
                    "general_appearance",
                    "height",
                    "exposed_armour",
                    "desagregation",
                    "cracks",
                    "reflectors",
                    "reflectors_type",
                    "reflectors_color",
                ]:
                    sum_intern_rows = 0
                    for _key, _value in value.items():
                        if not sum_intern_rows >= len(_key):
                            key_value = f"{self.static_fields[key][_key]}{row + sum_intern_rows}"
                            self.__insert_status_values(
                                key=_key, cell=key_value, value=_value
                            )
                            sum_intern_rows += 1
                elif key in ["photo_start", "photo_end"] and value:
                    cell = f"{self.static_fields[key]}{row}:{self.static_fields[key]}{row+3}"
                    insert_picture_2(
                        self._worksheet,
                        cell,
                        Image(value),
                        self.__sheet_target,
                        (1, 1, 1, 1),
                        ResizeMethod.ProportionalCentered,
                    )
                elif key == "executed_at":
                    date_list.append(value)
                elif key == "road_name":
                    road_name_list.append(value)
                elif key == "subcompany":
                    subcompany_list.append(value)
                elif key == "team":
                    team_list.append(values["team"])
                elif key == "raw_initial_km":
                    continue

                else:
                    key_value = f"{self.static_fields[key]}{row}"
                    self._worksheet[key_value] = value
                    RigidBarrierXlsxHandler.__format_fonts(
                        cell=self._worksheet[key_value], size=10
                    )

            row += 4

        filtered_team_list = set([uuid for uuid in team_list if uuid != ""])
        query_set_teams = Firm.objects.filter(uuid__in=filtered_team_list).all()
        names = []
        for team in query_set_teams:
            users_query_set = team.users.all()
            intern_list = []
            for user in users_query_set:
                intern_list.append(user.full_name)
            names.append(intern_list)
        names = [", ".join(_) for _ in names]
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(names),
            cell="H3",
            horizontal="left",
        )
        filtered_date_list = [date for date in date_list if date != ""]
        filtered_date_list.sort()
        if len(filtered_date_list) == 0:
            date_text = ""
        elif len(filtered_date_list) == 1:
            date_text = filtered_date_list[0].strftime("%d/%m/%Y")
        else:
            date_text = f"{filtered_date_list[0].strftime('%d/%m/%Y')} até {filtered_date_list[-1].strftime('%d/%m/%Y')}"

        insert_centered_value(
            worksheet=self._worksheet, value=date_text, cell="H4", horizontal="left"
        )

        road_name = list(set(road_name_list))
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(road_name),
            cell="L2",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value="Elementos de Proteção e Segurança - Ficha de Monitoração de Barreiras Rígidas",
            cell="L3",
            horizontal="left",
        )

        insert_centered_value(
            worksheet=self._worksheet,
            value=self.company.name,
            cell="L4",
            horizontal="left",
        )

        subcompany = list(set(subcompany_list))
        insert_centered_value(
            worksheet=self._worksheet,
            value=" / ".join(subcompany),
            cell="H2",
            horizontal="left",
        )

        total_length = sum(
            [t.value for t in self._worksheet["O"] if isinstance(t.value, (int, float))]
        )

        reference_total_legth = self._worksheet[f"{'N'}{self._worksheet.max_row + 1}"]
        cells_for_merge = (
            f"{reference_total_legth.coordinate}:O{reference_total_legth.row}"
        )

        self._worksheet.merge_cells(cells_for_merge)
        reference_total_legth.value = total_length
        reference_total_legth.border = border
        RigidBarrierXlsxHandler.__format_fonts(
            cell=reference_total_legth, size=10, bold=True
        )

        reference_total_string = reference_total_legth.offset(0, -1)
        reference_total_string.value = "TOTAL(m)"

        reference_total_string.border = border

        self._worksheet[cells_for_merge][0][1].border = border
        RigidBarrierXlsxHandler.__format_fonts(
            cell=reference_total_string, size=10, bold=True
        )

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            target=self.__sheet_target,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

        self.__insert_page_breaks(row)

    def __insert_page_breaks(self, rows):
        row = 62
        while row < rows:
            self._worksheet.row_breaks.append(Break(row - 4))
            row += 60

    @classmethod
    def __delineator_state(cls, reporting_state):
        get_state_in_data = new_get_form_data(reporting_state, "delineatorState")
        if not reporting_state.form_data.get("elemento_refletivo_barreira_nj"):
            return ""
        return get_state_in_data

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        sorted_data = sorted(
            data,
            key=lambda x: (
                x["raw_initial_km"] if x["direction"] == "NORTE" else False,
                -x["raw_initial_km"] if x["direction"] == "SUL" else False,
                x["direction"] == "CANTEIRO CENTRAL",
                x["direction"] == "AMBOS",
                x["direction"] == "CRESCENTE",
                x["direction"] == "DECRESCENTE",
                x["direction"] == "LESTE",
                x["direction"] == "LESTE/OESTE",
                x["direction"] == "NORTE/SUL",
                x["direction"] == "TRANSVERSAL",
                x["direction"] == "OESTE",
                x["direction"],
            ),
            reverse=True,
        )

        self.fill_sheet(data_list=sorted_data)
        road_name = self.reportings.first().road_name

        file_name = "Ficha de Monitoração - {} - Barreira Rígida".format(road_name)

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        result_file = f"/tmp/{file_name}.xlsx"
        self.wb.save(result_file)
        return result_file


class CCRRigidBarrier(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        qs_reporting = Reporting.objects.filter(pk__in=self.uuids)
        road_name = ""
        list_inspection_campaign_year = ""

        road_name = (" ").join(
            [
                re.sub(r"[- ]", "", road)
                for road in qs_reporting.order_by("road_name")
                .distinct("road_name")
                .values_list("road_name", flat=True)
            ]
        )
        list_inspection_campaign_year = list(
            set(
                [
                    str(reporting.form_data.get("inspection_campaign_year", ""))
                    for reporting in qs_reporting
                    if str(reporting.form_data.get("inspection_campaign_year", ""))
                ]
            )
        )

        if list_inspection_campaign_year:
            list_inspection_campaign_year.sort()

        inspection_campaign_year = (" ").join(list_inspection_campaign_year)

        file_name = "Relatórios ANTT de Barreira Rígida - {} - {}".format(
            road_name, inspection_campaign_year
        )

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = list()
        files.append(
            RigidBarrierXlsxHandler(
                list_uuids=self.uuids,
                s3=s3,
                sheet_target=self.sheet_target(),
            ).execute()
        )
        files.append(
            XlsxHandlerResumeReportRigidBarrier(
                list_uuids=self.uuids,
                s3=s3,
                sheet_target=self.sheet_target(),
            ).execute()
        )
        description_measures = XlsxHandlerResumeReportDescriptionMeasures(
            uuid=self.uuids[0],
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        for _file in description_measures:
            files.append(_file)

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_rigid_barrier_async_handler(reporter_dict: dict):
    reporter = CCRRigidBarrier.from_dict(reporter_dict)
    reporter.export()
