import shutil
import tempfile
import time
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Firm, Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_km_plus_meter,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data_selected_option
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_reporting_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_custom_option,
    get_custom_option_value,
    get_end_km,
    get_km,
)
from helpers.strings import clean_latin_string


class XlsxHandler:
    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.list_reporting = list_reporting
        self.__sheet_target = sheet_target
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp() + "/"
        self.__xlsx_file = (
            "./fixtures/reports/ccr_report_longitudinal_horizontal_signage.xlsx"
        )
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Longitudinal"]
        self._worksheet_data_graf_10km = self._workbook["Dados 10km"]
        self._worksheet_data_graf = self._workbook["Dados graficos"]

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="R1:U5",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B5",
            resize_method=ResizeMethod.ProportionalLeft,
        )

    @classmethod
    def __average_str(cls, values: list):
        avg = ""
        values_int = []
        for value in values:
            try:
                values_int.append(float(value))
            except Exception:
                pass
        if values_int:
            _sum = sum(values_int)
            _min = min(values_int)
            _max = max(values_int)
            avg = str((_sum - _min - _max) / 8)
        return avg

    def fill_sheet(self, data_list: list):

        data_graf_columns = {
            "BORDO DIREITO": "B",
            "EIXO": "C",
            "EIXO 2": "D",
            "EIXO 3": "E",
            "EIXO 4": "F",
            "BORDO ESQUERDO": "G",
        }

        data_dict = {}
        order_dict = {}
        for data in data_list:
            if data["road_name"] not in data_dict.keys():
                data_dict[data["road_name"]] = [data]
                order_dict[data["road_name"]] = [data["direction"]]
            else:
                data_dict[data["road_name"]].append(data)
                order_dict[data["road_name"]].append(data["direction"])
        for k, v in order_dict.items():
            __direction = list(set(v))
            if len(__direction) == 1:
                if str(__direction[0]).lower() == "sul":
                    new_list = sorted(
                        data_dict[k], key=lambda x: x["km"], reverse=False
                    )
                    data_dict[k] = new_list
                elif str(__direction[0]).lower() == "norte":
                    new_list = sorted(data_dict[k], key=lambda x: x["km"], reverse=True)
                    data_dict[k] = new_list
                else:
                    new_list = sorted(
                        data_dict[k], key=lambda x: (x["direction"], x["km"])
                    )
                    data_dict[k] = new_list
            else:
                new_list = sorted(data_dict[k], key=lambda x: (x["direction"], x["km"]))
                data_dict[k] = new_list

        bold = Font(bold=True)
        files = []
        for road, datalist in data_dict.items():
            init_row = 8
            init_data_graf = 2
            init_data_graf_10km = 2
            lanes = []
            direction = []
            team_list = []
            subcompany_list = []
            company_list = []
            dates = []
            for data in datalist:
                team_list.append(data["team"])
                subcompany_list.append(data["subcompany"])
                company_list.append(data["company"])
                if data["executed_at"]:
                    dates.append(data["executed_at"])
                else:
                    dates.append(data["found_at"])
                if data["lane"] not in lanes:
                    lanes.append(data["lane"])
                if data["direction"] not in direction:
                    direction.append(data["direction"])
                len_station_one = (
                    len(data["station_one"])
                    if len(data["station_one"]) > 3 or len(data["station_one"]) == 0
                    else 3
                )
                len_station_two = (
                    len(data["station_two"])
                    if len(data["station_two"]) > 3 or len(data["station_two"]) == 0
                    else 3
                )
                len_station_three = (
                    len(data["station_three"])
                    if len(data["station_three"]) > 3 or len(data["station_three"]) == 0
                    else 3
                )
                max_rows = len_station_one + len_station_two + len_station_three
                count = 0
                if len_station_one:
                    count += 1
                if len_station_two:
                    count += 1
                if len_station_three:
                    count += 1
                add = max_rows if max_rows >= (count * 3) else count * 3
                end_row = add + init_row
                self._worksheet.merge_cells(f"A{init_row}:A{init_row + add-1}")
                self._worksheet[f"A{init_row}"].alignment = Alignment(
                    textRotation=90, horizontal="center", vertical="center"
                )
                self._worksheet[f"A{init_row}"].font = bold
                self._worksheet[f"A{init_row}"].value = data["stretch"]
                for intern_col in ["B", "C", "D", "G", "H", "I"]:
                    if len_station_one > 1:
                        num = init_row
                        add = len_station_one if len_station_one >= 3 else 3
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                    if len_station_two > 1:
                        num = init_row + len_station_one
                        add = len_station_two if len_station_two >= 3 else 3
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                    if len_station_three > 1:
                        num = init_row + len_station_one + len_station_two
                        add = len_station_three if len_station_three >= 3 else 3
                        self._worksheet.merge_cells(
                            f"{intern_col}{num}:{intern_col}{num + add-1}"
                        )
                all_cols = [chr(i) for i in range(ord("A"), ord("V"))]
                for col in all_cols:
                    for row in range(
                        init_row,
                        init_row + (max_rows if max_rows >= (count * 3) else count * 3),
                    ):
                        self._worksheet[f"{col}{row}"].border = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000"),
                        )
                        self._worksheet.row_dimensions[row].height = 19.5
                if len_station_one:
                    self._worksheet[f"B{init_row}"].value = data["km_station_one"]
                    self._worksheet[f"C{init_row}"].value = data["lat_station"]
                    self._worksheet[f"D{init_row}"].value = data["long_station"]
                    range_str = f"G{init_row}:G{init_row+2}"
                    if data["images"]["station_one"]:
                        try:
                            insert_picture(
                                self._worksheet,
                                range_str,
                                Image(data["images"]["station_one"]),
                                self.__sheet_target,
                            )
                        except Exception as e:
                            print(e)
                    self._worksheet[f"H{init_row}"].value = data["executed_at"]
                    self._worksheet[f"I{init_row}"].value = data["notes_one"]
                    init = init_row
                    for item in data["station_one"]:
                        self._worksheet[f"E{init}"].value = item["sinalization_lane"]
                        self._worksheet[f"F{init}"].value = item["station_color"]
                        self._worksheet[f"J{init}"].value = item["measure_one"]
                        self._worksheet[f"K{init}"].value = item["measure_two"]
                        self._worksheet[f"L{init}"].value = item["measure_three"]
                        self._worksheet[f"M{init}"].value = item["measure_four"]
                        self._worksheet[f"N{init}"].value = item["measure_five"]
                        self._worksheet[f"O{init}"].value = item["measure_six"]
                        self._worksheet[f"P{init}"].value = item["measure_seven"]
                        self._worksheet[f"Q{init}"].value = item["measure_eight"]
                        self._worksheet[f"R{init}"].value = item["measure_nine"]
                        self._worksheet[f"S{init}"].value = item["measure_ten"]
                        self._worksheet[f"T{init}"].value = (
                            item["average"]
                            if item["average"]
                            else XlsxHandler.__average_str(
                                values=[
                                    item["measure_one"],
                                    item["measure_two"],
                                    item["measure_three"],
                                    item["measure_four"],
                                    item["measure_five"],
                                    item["measure_six"],
                                    item["measure_seven"],
                                    item["measure_eight"],
                                    item["measure_nine"],
                                    item["measure_ten"],
                                ]
                            )
                        )
                        self._worksheet[f"T{init}"].font = Font(
                            color=(
                                "ff0000"
                                if item["attendance_status"].lower() == "reprovado"
                                else "000000"
                            ),
                            bold=True,
                        )
                        if data["current_average_last_year"].get("station_one"):
                            _data_ = data["current_average_last_year"].get(
                                "station_one"
                            )
                            for _itern_data in _data_:
                                _average = _itern_data.get(
                                    f"{item['sinalization_lane']}-{item['station_color']}"
                                )
                                if _average:
                                    break
                            self._worksheet[f"U{init}"].value = (
                                _average if _average else ""
                            )
                        else:
                            self._worksheet[f"U{init}"].value = ""

                        init += 1

                        try:
                            sinalization_column = data_graf_columns[
                                item["sinalization_lane"]
                            ]
                        except Exception:
                            sinalization_column = None
                        else:
                            self._worksheet_data_graf[f"A{init_data_graf}"] = data[
                                "km_station_one"
                            ]
                            self._worksheet_data_graf[
                                f"{sinalization_column}{init_data_graf}"
                            ] = item["average"]
                            self._worksheet_data_graf[f"H{init_data_graf}"] = "80"
                            self._worksheet_data_graf[f"I{init_data_graf}"] = "100"

                    init_data_graf += 1

                if len_station_two:
                    count_station_two = init_row + (
                        3 if 0 < len_station_one < 3 else len_station_one
                    )
                    self._worksheet[f"B{count_station_two}"].value = data[
                        "km_station_two"
                    ]
                    self._worksheet[f"C{count_station_two}"].value = data[
                        "lat_station_two"
                    ]
                    self._worksheet[f"D{count_station_two}"].value = data[
                        "long_station_two"
                    ]
                    range_str = f"G{count_station_two}:G{count_station_two+2}"
                    if data["images"]["station_two"]:
                        try:
                            insert_picture(
                                self._worksheet,
                                range_str,
                                Image(data["images"]["station_two"]),
                                self.__sheet_target,
                            )
                        except Exception as e:
                            print(e)
                    self._worksheet[f"H{count_station_two}"].value = data["executed_at"]
                    self._worksheet[f"I{count_station_two}"].value = data[
                        "notes_station_two"
                    ]
                    init = count_station_two
                    for item in data["station_two"]:
                        self._worksheet[f"E{init}"].value = item["sinalization_lane"]
                        self._worksheet[f"F{init}"].value = item["station_color"]
                        self._worksheet[f"J{init}"].value = item["measure_one"]
                        self._worksheet[f"K{init}"].value = item["measure_two"]
                        self._worksheet[f"L{init}"].value = item["measure_three"]
                        self._worksheet[f"M{init}"].value = item["measure_four"]
                        self._worksheet[f"N{init}"].value = item["measure_five"]
                        self._worksheet[f"O{init}"].value = item["measure_six"]
                        self._worksheet[f"P{init}"].value = item["measure_seven"]
                        self._worksheet[f"Q{init}"].value = item["measure_eight"]
                        self._worksheet[f"R{init}"].value = item["measure_nine"]
                        self._worksheet[f"S{init}"].value = item["measure_ten"]
                        self._worksheet[f"T{init}"].value = item["average"]
                        self._worksheet[f"T{init}"].font = Font(
                            color=(
                                "ff0000"
                                if item["attendance_status"].lower() == "reprovado"
                                else "000000"
                            ),
                            bold=True,
                        )
                        if data["current_average_last_year"].get("station_two"):
                            _data_ = data["current_average_last_year"].get(
                                "station_two"
                            )
                            for _itern_data in _data_:
                                _average = _itern_data.get(
                                    f"{item['sinalization_lane']}-{item['station_color']}"
                                )
                                if _average:
                                    break
                            self._worksheet[f"U{init}"].value = (
                                _average if _average else ""
                            )
                        else:
                            self._worksheet[f"U{init}"].value = ""
                        init += 1

                        try:
                            sinalization_column = data_graf_columns[
                                item["sinalization_lane"]
                            ]
                        except Exception:
                            sinalization_column = None
                        else:
                            self._worksheet_data_graf[f"A{init_data_graf}"] = data[
                                "km_station_two"
                            ]
                            self._worksheet_data_graf[
                                f"{sinalization_column}{init_data_graf}"
                            ] = item["average"]
                            self._worksheet_data_graf[f"H{init_data_graf}"] = "80"
                            self._worksheet_data_graf[f"I{init_data_graf}"] = "100"

                    init_data_graf += 1

                if len_station_three:
                    count_station_three = (
                        init_row
                        + (3 if 0 < len_station_one < 3 else len_station_one)
                        + (3 if 0 < len_station_two < 3 else len_station_two)
                    )
                    self._worksheet[f"B{count_station_three}"].value = data[
                        "km_station_three"
                    ]
                    self._worksheet[f"C{count_station_three}"].value = data[
                        "lat_station_three"
                    ]
                    self._worksheet[f"D{count_station_three}"].value = data[
                        "long_station_three"
                    ]
                    range_str = f"G{count_station_three}:G{count_station_three+2}"
                    if data["images"]["station_three"]:
                        try:
                            insert_picture(
                                self._worksheet,
                                range_str,
                                Image(data["images"]["station_three"]),
                                self.__sheet_target,
                            )
                        except Exception as e:
                            print(e)
                    self._worksheet[f"H{count_station_three}"].value = data[
                        "executed_at"
                    ]
                    self._worksheet[f"I{count_station_three}"].value = data[
                        "notes_station_three"
                    ]
                    init = count_station_three
                    for item in data["station_three"]:
                        self._worksheet[f"E{init}"].value = item["sinalization_lane"]
                        self._worksheet[f"F{init}"].value = item["station_color"]
                        self._worksheet[f"J{init}"].value = item["measure_one"]
                        self._worksheet[f"K{init}"].value = item["measure_two"]
                        self._worksheet[f"L{init}"].value = item["measure_three"]
                        self._worksheet[f"M{init}"].value = item["measure_four"]
                        self._worksheet[f"N{init}"].value = item["measure_five"]
                        self._worksheet[f"O{init}"].value = item["measure_six"]
                        self._worksheet[f"P{init}"].value = item["measure_seven"]
                        self._worksheet[f"Q{init}"].value = item["measure_eight"]
                        self._worksheet[f"R{init}"].value = item["measure_nine"]
                        self._worksheet[f"S{init}"].value = item["measure_ten"]
                        self._worksheet[f"T{init}"].value = (
                            item["average"]
                            if item["average"]
                            else XlsxHandler.__average_str(
                                values=[
                                    item["measure_one"],
                                    item["measure_two"],
                                    item["measure_three"],
                                    item["measure_four"],
                                    item["measure_five"],
                                    item["measure_six"],
                                    item["measure_seven"],
                                    item["measure_eight"],
                                    item["measure_nine"],
                                    item["measure_ten"],
                                ]
                            )
                        )
                        self._worksheet[f"T{init}"].font = Font(
                            color=(
                                "ff0000"
                                if item["attendance_status"].lower() == "reprovado"
                                else "000000"
                            ),
                            bold=True,
                        )
                        if data["current_average_last_year"].get("station_three"):
                            _data_ = data["current_average_last_year"].get(
                                "station_three"
                            )
                            for _itern_data in _data_:
                                _average = _itern_data.get(
                                    f"{item['sinalization_lane']}-{item['station_color']}"
                                )
                                if _average:
                                    break
                            self._worksheet[f"U{init}"].value = (
                                _average if _average else ""
                            )
                        else:
                            self._worksheet[f"U{init}"].value = ""
                        init += 1

                        try:
                            sinalization_column = data_graf_columns[
                                item["sinalization_lane"]
                            ]
                        except Exception:
                            sinalization_column = None
                        else:
                            self._worksheet_data_graf[f"A{init_data_graf}"] = data[
                                "km_station_three"
                            ]
                            self._worksheet_data_graf[
                                f"{sinalization_column}{init_data_graf}"
                            ] = item["average"]
                            self._worksheet_data_graf[f"H{init_data_graf}"] = "80"
                            self._worksheet_data_graf[f"I{init_data_graf}"] = "100"

                    init_data_graf += 1

                specific_cols = [chr(i) for i in range(ord("B"), ord("V"))]
                for col in specific_cols:
                    for row in range(
                        init_row,
                        init_row + (max_rows if max_rows >= (count * 3) else count * 3),
                    ):
                        self._worksheet[f"{col}{row}"].alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                signal_strip = {}
                if data["left_border"] > 0:
                    signal_strip["BORDO ESQUERDO"] = data["left_border"]
                if data["axis"] > 0:
                    signal_strip["EIXO"] = data["axis"]
                if data["axis_two"] > 0:
                    signal_strip["EIXO 2"] = data["axis_two"]
                if data["axis_three"] > 0:
                    signal_strip["EIXO 3"] = data["axis_three"]
                if data["axis_four"] > 0:
                    signal_strip["EIXO 4"] = data["axis_four"]
                if data["right_border"] > 0:
                    signal_strip["BORDO DIREITO"] = data["right_border"]
                for signal, average in signal_strip.items():
                    self._worksheet.merge_cells(f"Q{end_row}:S{end_row}")
                    self._worksheet.row_dimensions[end_row].height = 19.5
                    self._worksheet[f"Q{end_row}"] = signal
                    self._worksheet[f"T{end_row}"] = average
                    self._worksheet[f"U{end_row}"] = (
                        data["current_average_last_year_location"].get(signal) or ""
                    )
                    color = "CCFFFF" if "BORDO" in signal else "FFFF99"
                    for col in ["Q", "R", "S", "T", "U"]:
                        self._worksheet[f"{col}{end_row}"].font = bold
                        self._worksheet[f"{col}{end_row}"].fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
                        self._worksheet[f"{col}{end_row}"].border = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000"),
                        )
                        self._worksheet[f"{col}{end_row}"].alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    end_row += 1

                init_row = end_row

                self._worksheet_data_graf_10km[f"A{init_data_graf_10km}"] = data[
                    "stretch"
                ]
                self._worksheet_data_graf_10km[f"B{init_data_graf_10km}"] = data[
                    "right_border"
                ]
                self._worksheet_data_graf_10km[f"C{init_data_graf_10km}"] = data["axis"]
                self._worksheet_data_graf_10km[f"D{init_data_graf_10km}"] = data[
                    "axis_two"
                ]
                self._worksheet_data_graf_10km[f"E{init_data_graf_10km}"] = data[
                    "axis_three"
                ]
                self._worksheet_data_graf_10km[f"F{init_data_graf_10km}"] = data[
                    "axis_four"
                ]
                self._worksheet_data_graf_10km[f"G{init_data_graf_10km}"] = data[
                    "left_border"
                ]
                self._worksheet_data_graf_10km[f"H{init_data_graf_10km}"] = "80"
                self._worksheet_data_graf_10km[f"I{init_data_graf_10km}"] = "100"

                init_data_graf_10km += 1

            team_list = list(set(team_list))
            query_set_teams = Firm.objects.filter(uuid__in=team_list).all()
            names = []
            for team in query_set_teams:
                users_query_set = team.users.all()
                intern_list = []
                for user in users_query_set:
                    intern_list.append(user.get_full_name())
                names.append(intern_list)
            names = [", ".join(_) for _ in names]
            team_text = f"Operador: {' - '.join(names)}"
            subcompany_text = f"Empresa: {' - '.join(list(set(subcompany_list)))}"
            company_text = f"Concessionária: {' - '.join(list(set(company_list)))}"
            road_text = f"Rodovia: {data['road_name']}"
            reporting_text = f"Monitoração: Sinalização Horizontal: Pista {' - '.join(list(set(lanes)))} {' - '.join(list(set(direction)))}"
            if len(dates) > 1:
                dates_new = sorted(
                    dates,
                    key=lambda x: (x.split("/")[2], x.split("/")[1], x.split("/")[0]),
                )
                date_text = f"Data: {dates_new[0]} à {dates_new[-1]}"
            elif dates:
                date_text = f"Data: {dates[0]}"
            else:
                date_text = "Data: "
            self._worksheet["E2"] = subcompany_text
            self._worksheet["E3"] = team_text
            self._worksheet["E4"] = date_text
            self._worksheet["J2"] = road_text
            self._worksheet["J3"] = reporting_text
            self._worksheet["J4"] = company_text
            result = f"Fichas de Horizontal {road} {'-'.join(sorted(list(set((lanes)))))} {'-'.join(sorted(list(set((direction)))))}"
            result = clean_latin_string(result.replace(".", "").replace("/", ""))
            result = f"/tmp/{result}.xlsx"
            files.append(result)
            insert_logo_and_provider_logo(
                worksheet=self._worksheet,
                target=self.__sheet_target,
                logo_company=self.data_logo_company,
                provider_logo=self.data_provider_logo,
            )
            self._workbook.save(result)
            self._workbook = load_workbook(self.__xlsx_file)
            self._worksheet = self._workbook["Longitudinal"]
            self._worksheet_data_graf_10km = self._workbook["Dados 10km"]
            self._worksheet_data_graf = self._workbook["Dados graficos"]

        return {"files": files}

    def __is_float(self, value):
        try:
            if float(value):
                return True
            else:
                return False
        except Exception:
            return False

    def __find_station(self, reporting, station: list, position: int) -> tuple:
        datalist = []
        data_id = {}
        if station:
            for obj in station:
                first_station = obj
                if position == 1:
                    _id = "1"
                    pos = ""
                    pos_pic = "_one"
                    ordinal = ""
                elif position == 2:
                    _id = "2"
                    pos = "_two"
                    pos_pic = "_two"
                    ordinal = "_second"
                else:
                    _id = "3"
                    pos = "_three"
                    pos_pic = "_three"
                    ordinal = "_third"
                if first_station.get(f"sinalization_lane{pos}"):
                    sinalization_lane = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__sinalizationLane",
                        first_station.get(f"sinalization_lane{pos}"),
                    )
                    station_color = new_get_form_data_selected_option(
                        reporting,
                        "stationOne__stationColor",
                        first_station.get(f"station_color{pos}"),
                    )
                    data_id[
                        f"{first_station.get(f'sinalization_lane{pos}')}-{first_station.get(f'station_color{pos}')}"
                    ] = f"{sinalization_lane}-{station_color}"

                    station_images = []
                    try:
                        for image_values in reporting.form_data.get(
                            f"station_pictures{pos_pic}", []
                        ):
                            station_images += list(image_values.values())
                    except Exception:
                        pass

                    data_obj = {
                        "sinalization_lane": sinalization_lane,
                        "station_color": station_color,
                        "id": _id,
                        "measure_one": first_station.get(f"measure_one{ordinal}"),
                        "measure_two": first_station.get(f"measure_two{ordinal}"),
                        "measure_three": first_station.get(f"measure_three{ordinal}"),
                        "measure_four": first_station.get(f"measure_four{ordinal}"),
                        "measure_five": first_station.get(f"measure_five{ordinal}"),
                        "measure_six": first_station.get(f"measure_six{ordinal}"),
                        "measure_seven": first_station.get(f"measure_seven{ordinal}"),
                        "measure_eight": first_station.get(f"measure_eight{ordinal}"),
                        "measure_nine": first_station.get(f"measure_nine{ordinal}"),
                        "measure_ten": first_station.get(f"measure_ten{ordinal}"),
                        "average": first_station.get(f"average_station{ordinal}"),
                        # "images": first_station.get(f"station_image{pos}"),
                        "images": station_images,
                        "attendance_status": str(
                            first_station.get(f"attendance_status{ordinal}")
                        ),
                    }
                    data_obj["average"] = XlsxHandler.__average_str(
                        values=[
                            data_obj["measure_one"],
                            data_obj["measure_two"],
                            data_obj["measure_three"],
                            data_obj["measure_four"],
                            data_obj["measure_five"],
                            data_obj["measure_six"],
                            data_obj["measure_seven"],
                            data_obj["measure_eight"],
                            data_obj["measure_nine"],
                            data_obj["measure_ten"],
                        ]
                    )
                    datalist.append(data_obj)
        return datalist, data_id

    def custom_key(self, item):
        if item["sinalization_lane"] == "BORDO ESQUERDO":
            return (0, item["sinalization_lane"])
        elif item["sinalization_lane"] == "BORDO DIREITO":
            return (2, item["sinalization_lane"])
        else:
            return (1, item["sinalization_lane"])

    def is_dict_empty(self, item):
        return 1 if not item else 0

    def create_dict(self, reporting: Reporting, s3) -> dict:
        km = get_km(reporting)
        km_end = get_end_km(reporting)
        km_station_one = get_km_plus_meter(reporting.form_data.get("km_station"))
        km_station_two = get_km_plus_meter(reporting.form_data.get("km_station_two"))
        km_station_three = get_km_plus_meter(
            km=reporting.form_data.get("km_station_three")
        )
        lat_station = reporting.form_data.get("lat_station")
        long_station = reporting.form_data.get("long_station")
        lat_station_two = reporting.form_data.get("lat_station_two")
        long_station_two = reporting.form_data.get("long_station_two")
        lat_station_three = reporting.form_data.get("lat_station_three")
        long_station_three = reporting.form_data.get("long_station_three")
        notes_one = reporting.form_data.get("notes_station_one")
        notes_station_two = reporting.form_data.get("notes_station_two")
        notes_station_three = reporting.form_data.get("notes_station_three")

        data_id = {}
        station_one, data_id_one = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_one"),
            position=1,
        )
        data_id.update(data_id_one)
        station_two, data_id_two = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_two"),
            position=2,
        )
        data_id.update(data_id_two)
        station_three, data_id_three = self.__find_station(
            reporting=reporting,
            station=reporting.form_data.get("station_three"),
            position=3,
        )
        data_id.update(data_id_three)

        station_one = sorted(station_one, key=self.custom_key)
        station_two = sorted(station_two, key=self.custom_key)
        station_three = sorted(station_three, key=self.custom_key)
        empty_obj = {
            "sinalization_lane": "",
            "station_color": "",
            "station_image": "",
            "measure_one": "",
            "measure_two": "",
            "measure_three": "",
            "measure_four": "",
            "measure_five": "",
            "measure_six": "",
            "measure_seven": "",
            "id": "",
            "measure_eight": "",
            "measure_nine": "",
            "measure_ten": "",
            "average": "",
            "images": "",
            "attendance_status": "",
        }
        if 0 < len(station_one) < 3:
            station_one.insert(1, empty_obj)
        if 0 < len(station_two) < 3:
            station_two.insert(1, empty_obj)
        if 0 < len(station_three) < 3:
            station_three.insert(1, empty_obj)
        allocation = [station_one, station_two, station_three]
        allocation = sorted(allocation, key=lambda x: self.is_dict_empty(x))
        station_one, station_two, station_three = (
            allocation[0],
            allocation[1],
            allocation[2],
        )

        try:
            right_border = float(reporting.form_data.get("right_border"))
        except Exception:
            right_border = 0
        try:
            left_border = float(reporting.form_data.get("left_border"))
        except Exception:
            left_border = 0
        try:
            axis = float(reporting.form_data.get("axis"))
        except Exception:
            axis = 0
        try:
            axis_two = float(reporting.form_data.get("axis_two"))
        except Exception:
            axis_two = 0
        try:
            axis_three = float(reporting.form_data.get("axis_three"))
        except Exception:
            axis_three = 0
        try:
            axis_four = float(reporting.form_data.get("axis_four"))
        except Exception:
            axis_four = 0

        reporting_files = download_reporting_pictures(
            s3,
            self.temp_file,
            reporting,
            width=337,
            height=242,
            enable_include_dnit=False,
            enable_is_shared_antt=True,
        )
        images = {"station_one": [], "station_two": [], "station_three": []}
        if station_one:
            for obj in station_one:
                if obj["images"]:
                    try:
                        for obj_images in obj["images"]:
                            for image in obj_images:
                                if ReportingFile.objects.filter(
                                    uuid=image, is_shared=True
                                ).exists():
                                    images["station_one"].append(image)
                            if len(images["station_one"]) > 0:
                                break
                    except Exception as e:
                        print(e)
        if station_two:
            for obj in station_two:
                if obj["images"]:
                    try:
                        for obj_images in obj["images"]:
                            for image in obj_images:
                                if ReportingFile.objects.filter(
                                    uuid=image, is_shared=True
                                ).exists():
                                    images["station_two"].append(image)
                            if len(images["station_two"]) > 0:
                                break
                    except Exception as e:
                        print(e)
        if station_three:
            for obj in station_three:
                if obj["images"]:
                    try:
                        for obj_images in obj["images"]:
                            for image in obj_images:
                                if ReportingFile.objects.filter(
                                    uuid=image, is_shared=True
                                ).exists():
                                    images["station_three"].append(image)
                            if len(images["station_three"]) > 0:
                                break
                    except Exception as e:
                        print(e)

        for obj in reporting_files["images"]:
            for k, v in images.items():
                if obj["uuid"] in v:
                    images[k] = obj["path"]
        direction = get_custom_option(reporting, "direction")
        _km = km.replace("+", ".")
        _km_end = km_end.replace("+", ".")
        if self.__is_float(value=_km) and self.__is_float(value=_km_end):
            graf_10km = True if (float(_km_end) - float(_km)) <= 10 else False
        else:
            graf_10km = False
        obj_ord = {
            "0": {"km": "", "lat": "", "long": "", "notes": ""},
            "1": {
                "km": km_station_one,
                "lat": lat_station,
                "long": long_station,
                "notes": notes_one,
            },
            "2": {
                "km": km_station_two,
                "lat": lat_station_two,
                "long": long_station_two,
                "notes": notes_station_two,
            },
            "3": {
                "km": km_station_three,
                "lat": lat_station_three,
                "long": long_station_three,
                "notes": notes_station_three,
            },
        }

        current_average_last_year = {}
        current_average_last_year_location = {}

        try:
            current_inspection_campaign_year = int(
                reporting.form_data.get("inspection_campaign_year")
            )
        except Exception:
            current_inspection_campaign_year = None

        if reporting.parent and current_inspection_campaign_year:
            parent: Reporting = reporting.parent
            query_set = Reporting.objects.filter(parent=str(parent.uuid)).only(
                "form_data", "found_at"
            )
            objs = {}
            for obj in query_set:
                try:
                    obj_inspection_campaign_year = int(
                        obj.form_data.get("inspection_campaign_year")
                    )
                except Exception:
                    obj_inspection_campaign_year = None

                if obj_inspection_campaign_year and (
                    obj_inspection_campaign_year == current_inspection_campaign_year - 1
                ):
                    objs[f"{obj.found_at.month}-{obj.found_at.day}"] = obj
            list_objs = list(sorted(objs.items(), reverse=True))
            if list_objs:
                obj = list_objs[0][1]
                form_data = obj.form_data
                if form_data:
                    current_average_last_year_location = {
                        "BORDO DIREITO": form_data.get("right_border"),
                        "BORDO ESQUERDO": form_data.get("left_border"),
                        "EIXO": form_data.get("axis"),
                        "EIXO 2": form_data.get("axis_two"),
                        "EIXO 3": form_data.get("axis_three"),
                        "EIXO 4": form_data.get("axis_four"),
                    }
                    for _key, _value in form_data.items():
                        __list_measure = []
                        if type(_value) is list and "station" in _key:
                            for item in _value:
                                try:
                                    sinalization_lane = next(
                                        y
                                        for x, y in item.items()
                                        if "sinalization_lane" in x
                                    )
                                    station_color = next(
                                        y
                                        for x, y in item.items()
                                        if "station_color" in x
                                    )
                                except StopIteration:
                                    continue
                                _key_name = (
                                    data_id.get(f"{sinalization_lane}-{station_color}")
                                    if data_id.get(
                                        f"{sinalization_lane}-{station_color}"
                                    )
                                    else f"{sinalization_lane}-{station_color}"
                                )
                                __list_measure.append(
                                    {
                                        _key_name: XlsxHandler.__average_str(
                                            values=[
                                                _v
                                                for _k, _v in item.items()
                                                if "measure" in _k
                                            ]
                                        )
                                    }
                                )
                            current_average_last_year[_key] = __list_measure

        allocation_two = [
            current_average_last_year.get("station_one"),
            current_average_last_year.get("station_two"),
            current_average_last_year.get("station_three"),
        ]
        allocation_two = sorted(allocation_two, key=lambda x: self.is_dict_empty(x))
        current_average_last_year = {
            "station_one": allocation_two[0] or {},
            "station_two": allocation_two[1] or {},
            "station_three": allocation_two[2] or {},
        }
        data = {
            "stretch": f"Trecho do KM  {km} ao KM {km_end}",
            "subcompany": reporting.firm.subcompany.__dict__.get("name"),
            "company": reporting.company.__dict__.get("name"),
            "team": str(reporting.firm.__dict__.get("uuid")),
            "graf_10km": graf_10km,
            "direction": direction if direction else "Pista",
            "lane": get_custom_option(reporting, "lane"),
            "road_name": reporting.road_name,
            "km": reporting.km,
            "km_station_one": (
                obj_ord[station_one[0]["id"]]["km"] if station_one else ""
            ),
            "km_station_two": (
                obj_ord[station_two[0]["id"]]["km"] if station_two else ""
            ),
            "km_station_three": (
                obj_ord[station_three[0]["id"]]["km"] if station_three else ""
            ),
            "lat_station": obj_ord[station_one[0]["id"]]["lat"] if station_one else "",
            "lat_station_two": (
                obj_ord[station_two[0]["id"]]["lat"] if station_two else ""
            ),
            "lat_station_three": (
                obj_ord[station_three[0]["id"]]["lat"] if station_three else ""
            ),
            "long_station": (
                obj_ord[station_one[0]["id"]]["long"] if station_one else ""
            ),
            "long_station_two": (
                obj_ord[station_two[0]["id"]]["long"] if station_two else ""
            ),
            "long_station_three": (
                obj_ord[station_three[0]["id"]]["long"] if station_three else ""
            ),
            "executed_at": (
                reporting.executed_at.strftime("%d/%m/%Y")
                if reporting.executed_at
                else ""
            ),
            "found_at": (
                reporting.found_at.strftime("%d/%m/%Y") if reporting.found_at else ""
            ),
            "station_one": station_one,
            "station_two": station_two,
            "station_three": station_three,
            "notes_one": obj_ord[station_one[0]["id"]]["notes"] if station_one else "",
            "notes_station_two": (
                obj_ord[station_two[0]["id"]]["notes"] if station_two else ""
            ),
            "notes_station_three": (
                obj_ord[station_three[0]["id"]]["notes"] if station_three else ""
            ),
            "right_border": right_border,
            "left_border": left_border,
            "axis": axis,
            "axis_two": axis_two,
            "axis_three": axis_three,
            "axis_four": axis_four,
            "images": images,
            "reporting": reporting,
            "current_average_last_year": current_average_last_year,
            "current_average_last_year_location": current_average_last_year_location,
        }
        for k, v in data.items():
            if v is None and k not in ["station_one", "station_two", "station_three"]:
                data[k] = ""
        return data

    def execute(self):

        data = []
        for reporting in self.list_reporting:
            data.append(self.create_dict(reporting=reporting, s3=self.s3))

        self.data_logo_company["path_image"] = get_logo_file(
            s3=self.s3,
            temp_prefix=self.temp_file,
            reporting=self.list_reporting[0],
        )
        self.data_provider_logo["path_image"] = get_provider_logo_file(
            s3=self.s3,
            temp_prefix=self.temp_file,
            reporting=self.list_reporting[0],
        )
        files = self.fill_sheet(data_list=data)
        shutil.rmtree(self.temp_file, ignore_errors=True)
        return files


class CCRLongitudinalHorizontalSignage(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""
        data_list = self.__obj_filter()

        reporting = Reporting.objects.get(uuid=self.uuids[0])

        direction_name = "Pista"
        direction_options = list(
            Reporting.objects.filter(uuid__in=self.uuids)
            .values_list("direction", flat=True)
            .distinct()
        )
        if len(direction_options) > 0:
            directions = []
            for direction_option in direction_options:
                direction = get_custom_option_value(
                    reporting, "direction", direction_option
                )
                directions.append(direction)
            direction_name = " - ".join(sorted(directions))

        lane_name = "Faixa"
        lane_options = list(
            Reporting.objects.filter(uuid__in=self.uuids)
            .values_list("lane", flat=True)
            .distinct()
        )
        if len(lane_options) > 0:
            lanes = []
            for lane_option in lane_options:
                lane = get_custom_option_value(reporting, "lane", lane_option)
                lanes.append(lane)
            lane_name = " - ".join(sorted(lanes))

        road_name = "Road"
        roads = list(
            Reporting.objects.filter(uuid__in=self.uuids)
            .values_list("road_name", flat=True)
            .distinct()
        )
        if len(roads) > 0:
            road_name = " - ".join(sorted(roads))

        road_name = reporting.road_name

        file_name = f"Fichas de Horizontal {road_name} {lane_name} {direction_name}"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))

        if len(data_list) == 1:
            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{file_name}.{extension}"
        else:
            file_name = f"Retrorrefletância Horizontal Longitudinal {str(time.time())}"
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
            file_name = f"{file_name}.zip"

        return file_name

    def __get_repotings_obj(self):
        form = OccurrenceType.objects.get(
            name="Retrorrefletância Horizontal Longitudinal"
        )
        query_set = Reporting.objects.filter(
            occurrence_type=form, uuid__in=self.uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
        return [_ for _ in query_set if str(_.uuid) in self.uuids]

    def __obj_filter(self):
        list_reporting = self.__get_repotings_obj()
        data_list = []
        for data in list_reporting:
            road_name = data.road_name
            key = road_name
            if key not in data_list:
                data_list.append(key)
        return data_list

    def export(self):
        list_reporting = self.__get_repotings_obj()
        s3 = get_s3()
        obj = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()
        files = obj["files"]

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_longitudinal_horizontal_signage_async_handler(reporter_dict: dict):
    reporter = CCRLongitudinalHorizontalSignage.from_dict(reporter_dict)
    reporter.export()
