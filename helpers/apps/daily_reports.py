import logging
import os
import re
import shutil
from collections import defaultdict
from copy import copy
from datetime import date, datetime, time
from tempfile import NamedTemporaryFile
from typing import DefaultDict, List

import requests
import sentry_sdk
from dateutil import parser
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.validators import EMPTY_VALUES
from django.db.models import OuterRef, Q, QuerySet, Subquery
from django.db.models.signals import post_save, pre_save
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.pagebreak import Break
from PIL import ImageFont
from rest_framework import status
from rest_framework.response import Response
from rest_framework_json_api import serializers
from sentry_sdk import capture_exception
from storages.utils import clean_name
from zappa.asynchronous import task

from apps.companies.models import Company, Firm, UserInCompany
from apps.daily_reports.const.export_formats import PDF as PDF_FORMAT
from apps.daily_reports.const.occurrence_origin import (
    TRANSLATE_OCCURRENCE_ORIGIN_CHOICES,
)
from apps.daily_reports.const.relation_fields import (
    FIELD_TO_MODEL_CLASS,
    RELATION_FIELDS,
)
from apps.daily_reports.models import (
    DailyReport,
    DailyReportContractUsage,
    DailyReportEquipment,
    DailyReportExport,
    DailyReportExternalTeam,
    DailyReportOccurrence,
    DailyReportRelation,
    DailyReportResource,
    DailyReportSignaling,
    DailyReportVehicle,
    DailyReportWorker,
    MultipleDailyReport,
    ProductionGoal,
)
from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting
from apps.reportings.serializers import LightReportingSerializer
from apps.resources.models import ContractPeriod
from apps.service_orders.models import MeasurementBulletin, ProcedureResource
from apps.services.models import ServiceUsage
from apps.users.models import User
from apps.work_plans.models import Job
from helpers.apps.json_logic import apply_json_logic
from helpers.apps.reportings import return_array_values, return_select_value
from helpers.const.holidays import HOLIDAYS
from helpers.dates import (
    format_minutes,
    format_minutes_decimal,
    minutes_between,
    utc_to_local,
)
from helpers.extra_hours import (
    _worker_result_to_decimal_cols,
    calculate_extra_hours_worker,
)
from helpers.serializers import get_obj_serialized
from helpers.signals import DisableSignals
from helpers.strings import (
    DAYS_PORTUGUESE,
    dict_to_casing,
    get_obj_from_path,
    is_valid_uuid,
    to_flatten_str,
    to_snake_case,
)
from RoadLabsAPI.settings.credentials import GOTENBERG_BASE_URL


def is_holiday_for_firm(company, firm_id, check_date):
    """
    Checks if a date is a holiday for a specific firm.
    Checks in order:
    1. National holidays (helpers/const/holidays.py)
    2. Custom holidays (company.custom_options.dailyReportHolidays)
    """
    if not check_date:
        return False

    # 1. National holidays (constants)
    year = check_date.year
    date_str = str(check_date)
    if date_str in HOLIDAYS.get(year, []):
        return True

    # 2. Custom holidays (custom_options)
    if not company:
        return False

    custom_options = company.custom_options or {}
    daily_report_holidays = custom_options.get("dailyReportHolidays", {})
    holidays = daily_report_holidays.get("fields", {}).get("holidays", [])

    firm_id_str = str(firm_id) if firm_id else None

    for holiday in holidays:
        holiday_date_str = holiday.get("date")
        if not holiday_date_str:
            continue

        try:
            holiday_date = datetime.strptime(holiday_date_str, "%Y-%m-%d").date()
        except ValueError:
            continue

        repeat = holiday.get("repeat", False)
        if repeat:
            date_matches = (
                holiday_date.month == check_date.month
                and holiday_date.day == check_date.day
                and check_date.year >= holiday_date.year
            )
        else:
            date_matches = holiday_date == check_date

        if not date_matches:
            continue

        firms = holiday.get("firms", [])
        if not firms:
            continue
        elif firm_id_str and firm_id_str in [str(f) for f in firms]:
            return True

    return False


def filter_board_item_contract_services(queryset, contract_service_ids):
    """
    Filters the provided board item queryset according to the contract_service IDs
    """

    return queryset.filter(
        contract_item_administration__contract_item_administration_services__uuid__in=contract_service_ids
    )


def create_and_update_contract_usage(instance):
    """
    Creates a DailyReportContractUsage according to the provided board item instance
    if the instance has a contract_item_administration

    Args:
        instance: The board item instance (DailyReportWorker/Equipment/Vehicle)
    """

    instance_has_contract_item = instance.contract_item_administration is not None

    if instance_has_contract_item:
        instance_has_usage = (
            getattr(instance, "worker_contract_usage", None)
            or getattr(instance, "equipment_contract_usage", None)
            or getattr(instance, "vehicle_contract_usage", None)
        )

        # Determine resource type and relation field
        relation_field = determine_relation_field_name(type(instance))

        def get_company_id():
            if relation_field == "worker" and instance.firm:
                return instance.firm.company_id
            else:
                return instance.company_id

        # Prepare denormalized fields (non-M2M)
        denormalized_fields = {
            "company_id": get_company_id(),
            "contract_item_administration_id": instance.contract_item_administration_id,
            "measurement_bulletin_id": getattr(
                instance, "measurement_bulletin_id", None
            ),
        }

        # Add firm only for Worker (Equipment and Vehicle don't have firm)
        if relation_field == "worker" and getattr(instance, "firm_id", None):
            denormalized_fields["firm_id"] = instance.firm_id

        if not instance_has_usage:
            # Create new ContractUsage with denormalized fields
            kwargs = {relation_field: instance, **denormalized_fields}
            contract_usage = DailyReportContractUsage.objects.create(**kwargs)
        else:
            # Get existing ContractUsage
            contract_usage = (
                getattr(instance, "worker_contract_usage", None)
                or getattr(instance, "equipment_contract_usage", None)
                or getattr(instance, "vehicle_contract_usage", None)
            )
            if not contract_usage:
                return  # Safety check

            # Update denormalized fields
            for field, value in denormalized_fields.items():
                setattr(contract_usage, field, value)

            contract_usage.save(update_fields=list(denormalized_fields.keys()))

        # Copy M2M relationships from resource to contract_usage
        # This needs to be done after the contract_usage is saved (has an ID)
        multiple_daily_reports = getattr(instance, "multiple_daily_reports", None)
        if multiple_daily_reports:
            multiple_daily_report_ids = multiple_daily_reports.values_list(
                "uuid", flat=True
            )
            contract_usage.multiple_daily_reports.set(multiple_daily_report_ids)

        daily_reports = getattr(instance, "daily_reports", None)
        if daily_reports:
            daily_report_ids = daily_reports.values_list("uuid", flat=True)
            contract_usage.daily_reports.set(daily_report_ids)


def update_resource_amounts(instance, changed_fields):
    """
    Updates used_price and remaining_amount of a ServiceOrderResource when a board
    item is approved, rejected, or has its price changed.
    Also check if the item was added to or removed from a MeasurementBulletin
    and if so, update the total_price of the MeasurementBulletin accordingly.
    """

    DEFAULT_WORK_DAY = 22

    instance_has_contract_item = instance.contract_item_administration is not None
    if instance_has_contract_item:
        try:
            if not instance._state.adding:
                service_order_resource = instance.contract_item_administration.resource
                for field, (old, new) in changed_fields.items():
                    old_value = str(old)
                    if field == "total_price" and instance.measurement_bulletin:
                        instance.total_price = old
                    if (
                        field == "total_price"
                        and not instance.measurement_bulletin
                        and instance.approval_status == "WAITING_APPROVAL"
                    ):
                        service_order_resource.used_price -= old
                        service_order_resource.used_price += new
                    if field == "measurement_bulletin":
                        if instance.measurement_bulletin:

                            # Since work_day is optional and can be zero, we'll default to 22
                            mb_work_day = instance.measurement_bulletin.work_day
                            work_day = (
                                mb_work_day
                                if mb_work_day and mb_work_day > 0
                                else DEFAULT_WORK_DAY
                            )

                            service_order_resource.remaining_amount -= (
                                instance.amount / work_day
                            )
                            service_order_resource.used_price += instance.total_price
                        else:
                            previous_measurement_bulletin = (
                                MeasurementBulletin.objects.get(uuid=old_value)
                            )

                            # Since work_day is optional and can be zero, we'll default to 22
                            mb_work_day = previous_measurement_bulletin.work_day
                            work_day = (
                                mb_work_day
                                if mb_work_day and mb_work_day > 0
                                else DEFAULT_WORK_DAY
                            )
                            service_order_resource.remaining_amount += (
                                instance.amount / work_day
                            )
                            service_order_resource.used_price -= instance.total_price
                            instance.total_price = 0
                with DisableSignals(disabled_signals=[post_save, pre_save]):
                    service_order_resource.save()
        except Exception as e:
            # TODO: Use specific exception
            print(e)


def get_uuids_jobs_user_firms(
    jobs_section: str, company: Company, user: User, use_reporting_limit: bool = True
) -> List[str]:
    """
    Retrieves a list of Job UUIDs based on user permissions and company settings.

    This function processes Jobs in two ways:
    1. Gets most recent Jobs based on a count (num_jobs)
    2. Gets specific Jobs based on provided UUIDs

    Args:
        jobs_section (str): A string containing Job information in format "num_jobs, uuid1, uuid2, ..."
                            where num_jobs is the number of recent jobs to fetch
        company (Company): The Company instance to filter Jobs
        user (User): The User instance to check permissions
        use_reporting_limit (bool, optional): Whether to apply the max_reportings_by_job limit.
                                              Defaults to True.

    Returns:
        List[UUID]: A unique list of Job UUIDs combining both recent and specifically requested Jobs

    Notes:
        - Uses company.metadata["num_jobs"] if available to override num_jobs from jobs_section
        - Uses company.metadata["max_reportings_by_job"] (default 250) as reporting limit
        - Only returns non-archived Jobs
        - Jobs are ordered by start_date in descending order
        - Removes duplicate UUIDs from final result
    """
    jobs_values = jobs_section.split(",")
    num_jobs = jobs_values.pop(0)
    user_firms = (
        user.user_firms.filter(company=company)
        .only("uuid")
        .values_list("uuid", flat=True)
    )

    if "num_jobs" in company.metadata:
        num_jobs = int(company.metadata["num_jobs"])

    max_reportings_by_job = int(company.metadata.get("max_reportings_by_job", 250))

    criteria = {"firm_id__in": user_firms, "archived": False, "reporting_count__gte": 1}
    if use_reporting_limit:
        criteria["reporting_count__lte"] = max_reportings_by_job

    jobs_by_count = [
        item.uuid
        for item in Job.objects.filter(**criteria)
        .only("uuid", "start_date")
        .order_by("-start_date")[0 : int(num_jobs)]
    ]

    criteria = {"uuid__in": jobs_values, "archived": False}
    if use_reporting_limit:
        criteria["reporting_count__lte"] = max_reportings_by_job
    jobs_by_ids = list(
        Job.objects.filter(**criteria).only("uuid").values_list("uuid", flat=True)
    )

    # Don't repeat the same uuids
    return list(set(jobs_by_count + jobs_by_ids))


def get_uuids_rdos_user_firms(
    rdos_section: str, company: Company, user: User
) -> List[str]:
    """
    Retrieves a list of MultipleDailyReport UUIDs based on user permissions and Company settings.

    This function processes in two ways:
    1. Gets most recent objects based on a count (num_rdos)
    2. Gets specific objects based on provided UUIDs

    Args:
        rdos_section (str): A string containing RDO information in format "num_rdos, uuid1, uuid2,..."
                            where num_rdos is the number of recent objects to fetch
        company (Company): The Company instance to filter reports
        user (User): The User instance to check permissions

    Returns:
        List[UUID]: A unique list of MultipleDailyReport UUIDs combining both recent and
                    specifically requested objects

    Notes:
        - Uses company.metadata["num_rdos"] if available to override num_rdos from rdos_section
        - Returns objects from firms where user is member, inspector or manager
        - Objects are ordered by date in descending order
        - Removes duplicate UUIDs from final result
    """
    rdos_values = rdos_section.split(",")
    num_rdos = rdos_values.pop(0)

    if "num_rdos" in company.metadata:
        num_rdos = int(company.metadata["num_rdos"])

    rdos_by_count = MultipleDailyReport.objects.none()
    firms_ids = list(
        set(
            company.company_firms.filter(
                (Q(users=user) | Q(inspectors=user) | Q(manager=user))
            )
            .only("uuid")
            .values_list("uuid", flat=True)
        )
    )

    for firm in firms_ids:
        rdos_by_count = rdos_by_count.union(
            MultipleDailyReport.objects.filter(firm_id=firm)
            .order_by("-date")[: int(num_rdos)]
            .only("uuid")
            .values_list("uuid", flat=True)
        )

    # Don't repeat the same uuids
    return list(set(list(rdos_by_count) + rdos_values))


def filter_jobs_rdos_user_firms(filter_input, queryset, user, request_data):
    """
    Filter models related to MultipleDailyReport according to a fixed number and provided uuids.
    Returns a queryset.
    """

    jobs_section, rdos_section = filter_input.split("|")

    if "company" not in request_data:
        return queryset
    else:
        company = Company.objects.get(uuid=request_data["company"])

    jobs_uuids = get_uuids_jobs_user_firms(jobs_section, company, user)
    rdos_uuids = get_uuids_rdos_user_firms(rdos_section, company, user)

    reps_mdr = list(
        Reporting.objects.filter(job_id__in=jobs_uuids).values_list(
            "reporting_multiple_daily_reports", flat=True
        )
    )

    return queryset.filter(
        multiple_daily_reports__in=set(rdos_uuids + reps_mdr)
    ).distinct()


def filter_num_jobs_only_user_firms(
    filter_input: str, queryset: QuerySet, user: User, request_data: dict
) -> QuerySet:
    """
    Filters a queryset based on Jobs associated with user's Firms and Company settings.

    This function processes Jobs in two ways:
    1. Gets most recent Jobs based on a count (num_Jobs)
    2. Gets specific Jobs based on provided UUIDs

    Args:
        filter_input (str): A string containing job information in format "num_jobs, uuid1, uuid2,..."
                             where num_jobs is the number of recent jobs to fetch
        queryset (QuerySet): The base queryset to filter
        user (User): The user instance to check firm memberships
        request_data (dict): Request data containing Company information

    Returns:
        QuerySet: Filtered queryset containing only items related to the specified Jobs

    Notes:
        - Uses company.metadata["num_jobs"] if available to override num_jobs from filter_input
        - Uses company.metadata["max_reportings_by_job"] (default 250) as reporting limit
        - Only includes non-archived Jobs
        - Jobs are ordered by start_date in descending order
    """
    if "company" not in request_data:
        return queryset
    else:
        company = Company.objects.get(uuid=request_data["company"])

    jobs_values = filter_input.split(",")
    num_jobs = jobs_values.pop(0)
    user_firms = user.user_firms.all()
    if "num_jobs" in company.metadata:
        num_jobs = int(company.metadata["num_jobs"])

    max_reportings_by_job = int(company.metadata.get("max_reportings_by_job", 250))

    jobs_by_count = (
        Job.objects.filter(
            firm__in=user_firms,
            archived=False,
            reporting_count__lte=max_reportings_by_job,
        )
        .order_by("-start_date")[0 : int(num_jobs)]
        .values_list("uuid", flat=True)
    )
    jobs_by_ids = Job.objects.filter(
        uuid__in=jobs_values,
        archived=False,
        reporting_count__lte=max_reportings_by_job,
    ).values_list("uuid", flat=True)

    return queryset.filter(
        Q(multiple_daily_reports__reportings__job_id__in=jobs_by_count)
        | Q(multiple_daily_reports__reportings__job_id__in=jobs_by_ids)
    ).distinct()


def filter_num_user_firms(
    filter_input: str, queryset: QuerySet, user: User, request_data: dict
) -> QuerySet:
    """
    Filters a queryset based on Firms where the user is a member and Company settings.

    This function processes Firms in two ways:
    1. Gets most recent Firms based on a count (num_firms)
    2. Gets specific Firms based on provided UUIDs

    Args:
        filter_input (str): A string containing firm information in format "num_firms, uuid1, uuid2,..."
                            where num_firms is the number of recent Firms to fetch
        queryset (QuerySet): The base queryset to filter
        user (User): The user instance to check firm memberships
        request_data (dict): Request data containing Company information

    Returns:
        QuerySet: Filtered queryset containing only items related to the specified Firms

    Notes:
        - Uses company.metadata["num_firms"] if available to override num_firms from filter_input
        - Only includes Firms where user is a member
        - Firms are ordered by name
        - Results are filtered through multiple_daily_reports__firm relationship
    """
    if "company" not in request_data:
        return queryset
    else:
        company = Company.objects.get(uuid=request_data["company"])

    firms_values = filter_input.split(",")
    num_firms = firms_values.pop(0)
    if "num_firms" in company.metadata:
        num_firms = int(company.metadata["num_firms"])

    firms_by_count = (
        Firm.objects.filter(company=company, users__in=[user])
        .order_by("name")[: int(num_firms)]
        .values_list("uuid", flat=True)
    )

    firms_by_ids = Firm.objects.filter(
        uuid__in=firms_values, company=company
    ).values_list("uuid", flat=True)

    return queryset.filter(
        Q(multiple_daily_reports__firm__in=firms_by_count)
        | Q(multiple_daily_reports__in=firms_by_ids)
    ).distinct()


def determine_relation_field_name(model_class):
    if model_class == DailyReportWorker:
        field_name = "worker"
    elif model_class == DailyReportExternalTeam:
        field_name = "external_team"
    elif model_class == DailyReportEquipment:
        field_name = "equipment"
    elif model_class == DailyReportVehicle:
        field_name = "vehicle"
    elif model_class == DailyReportSignaling:
        field_name = "signaling"
    elif model_class == ProductionGoal:
        field_name = "production_goal"
    elif model_class == DailyReportOccurrence:
        field_name = "occurrence"
    elif model_class == DailyReportResource:
        field_name = "resource"

    return field_name


def determine_report_type_and_field(report):
    """
    Checks if the provided report is supported and returns
    the proper report_field for the relation along with the type
    """

    if isinstance(report, DailyReport):
        report_field = "daily_report"
        report_type = "DailyReport"
    elif isinstance(report, MultipleDailyReport):
        report_field = "multiple_daily_report"
        report_type = "MultipleDailyReport"
    else:
        raise serializers.ValidationError("Modelo de RDO não suportado")

    return report_field, report_type


def destroy_report_board_items(
    daily_report: DailyReport = None, multiple_daily_report: MultipleDailyReport = None
) -> None:
    """
    Go through every DailyReportRelation that report had and make sure EVERY
    DailyReportRelation and respective board item is deleted.

    Only one of the arguments should be provided. Never both or neither.

    Args:
        daily_report (DailyReport, optional): If the report is a DailyReport. Defaults to None.
        multiple_daily_report (MultipleDailyReport, optional): If the report is a MultipleDailyReport. Defaults to None.
    """

    # A few guarantees
    assert bool(daily_report) != bool(
        multiple_daily_report
    ), "Please provide just one report"
    assert daily_report is None or isinstance(
        daily_report, DailyReport
    ), "daily_report needs to be a DailyReport instance"
    assert multiple_daily_report is None or isinstance(
        multiple_daily_report, MultipleDailyReport
    ), "multiple_daily_report needs to be a MultipleDailyReport instance"

    report = daily_report or multiple_daily_report
    report_field = "daily_report" if daily_report else "multiple_daily_report"

    # Filter every DailyReportRelation for that report
    kwargs = {report_field: report}
    relations_to_remove = DailyReportRelation.objects.filter(**kwargs)

    # Grouped fields
    field_to_uuids = {field_name: [] for field_name in RELATION_FIELDS}

    # Get the raw values of every item field in the DailyReportRelation
    raw_relations_data = relations_to_remove.values(*RELATION_FIELDS)

    # Go through the raw data and add each item UUID to the grouping dict
    for relation_data_dict in raw_relations_data:
        for field, board_item_uuid in relation_data_dict.items():
            # Check if part of accepted field and truthy
            if field in RELATION_FIELDS and board_item_uuid:
                field_to_uuids[field].append(board_item_uuid)

    # Go through the grouped UUIDs and delete the board items
    for field, uuid_list in field_to_uuids.items():
        if uuid_list:
            related_model = FIELD_TO_MODEL_CLASS[field]
            related_model.objects.filter(uuid__in=uuid_list).delete()

    # Finally, delete the DailyReportRelations
    relations_to_remove.delete()


def days_with_progress(reports, prod_goal, firm_id=None):
    """
    Returns how many days in a report queryset have progress.

    Supports DailyReport and MultipleDailyReport
    """
    days = 0
    for report in reports:
        kwargs = {
            "executed_at__date__gte": prod_goal.starts_at,
            "executed_at__date__lte": prod_goal.ends_at,
            "executed_at__date": report.date,
            "company": prod_goal.service.company_id,
        }

        # If firm_id is provided, use only reportings of that firm
        if firm_id:
            kwargs["firm"] = firm_id

        reportings = Reporting.objects.filter(**kwargs)
        service_usages = ServiceUsage.objects.filter(
            reporting__in=reportings, service=prod_goal.service
        )
        amount = sum(service_usages.values_list("amount", flat=True))

        if amount > 0:
            days += 1

    return days


def show_active_on_report_relationships(
    show_active, list_response, item_relation_field
):
    """
    Inject the active key to report relationships.
    """

    if show_active:
        list_data = list_response.data
        if "results" in list_data:
            all_multiple_daily_reports = [
                report["id"]
                for item in list_data["results"]
                for report in item["multiple_daily_reports"]
            ]
            all_daily_reports = [
                report["id"]
                for item in list_data["results"]
                for report in item["daily_reports"]
            ]
            all_items = [item["uuid"] for item in list_data["results"]]

            mdr_kwargs = {
                "multiple_daily_report__in": all_multiple_daily_reports,
                item_relation_field + "__in": all_items,
            }
            dr_kwargs = {
                "daily_report__in": all_daily_reports,
                item_relation_field + "__in": all_items,
            }

            prefetch_qs = DailyReportRelation.objects.filter(
                Q(**mdr_kwargs) | Q(**dr_kwargs)
            )
            prefetch_dict = defaultdict(dict)
            for dr_rel in prefetch_qs:
                dr_uuid = (
                    str(dr_rel.multiple_daily_report_id)
                    if dr_rel.multiple_daily_report_id
                    else str(dr_rel.daily_report_id)
                )
                prefetch_dict[dr_uuid][
                    str(getattr(dr_rel, item_relation_field + "_id"))
                ] = dr_rel

            for item in list_data["results"]:
                for report in item["multiple_daily_reports"]:
                    try:
                        relation = prefetch_dict[report["id"]][item["uuid"]]
                    except Exception:
                        report["active"] = True
                        continue
                    report["active"] = relation.active

                for report in item["daily_reports"]:
                    try:
                        relation = prefetch_dict[report["id"]][item["uuid"]]
                    except Exception:
                        report["active"] = True
                        continue
                    report["active"] = relation.active

        return Response(list_data)

    else:
        return list_response


def calculate_total_price(amount, unit_price, work_day):
    """
    Calculates the total_price according the the ServiceOrderResource
    if the board item has a relationship to ContractItemAdministration
    """

    try:
        return (amount * unit_price) / work_day
    except Exception:
        return 0.0


def calculate_board_item_total_price(instance):
    """
    Helper to facilitate signal usage of calculate_total_price
    """
    if instance.unit_price is not None:
        unit_price = instance.unit_price
    elif (
        instance.contract_item_administration
        and instance.contract_item_administration.resource
    ):
        unit_price = instance.contract_item_administration.resource.unit_price
    else:
        unit_price = 0.0

    try:
        return calculate_total_price(
            instance.amount,
            unit_price,
            instance.measurement_bulletin.work_day,
        )
    except Exception:
        return 0.0


def get_km_intervals_field(instance, only_query=True):
    intervals = []

    if instance.reportings.count() > 0:
        if only_query:
            reporting_values = [
                (rep.road_name, rep.km, rep.end_km)
                for rep in instance.reportings.all().only("road_name", "km", "end_km")
            ]
        else:
            reporting_values = [
                (rep.road_name, rep.km, rep.end_km) for rep in instance.reportings.all()
            ]

        interval_dict = {}  # For easy indexing
        for road_name, km, end_km in reporting_values:
            if road_name in interval_dict:
                # If there's a smaller km, set as the new start
                if km < interval_dict[road_name]["km"]:
                    interval_dict[road_name]["km"] = km

                # If there's a bigger or non None end_km, set as new end
                old_end_km = interval_dict[road_name]["end_km"]
                if end_km and (old_end_km is None or end_km > old_end_km):
                    interval_dict[road_name]["end_km"] = end_km
            else:
                interval_dict[road_name] = {"km": km, "end_km": end_km}

        # Turn dict into response list format
        for road_name, kms in interval_dict.items():
            entry = {"roadName": road_name, "km": kms["km"], "end_km": kms["end_km"]}
            intervals.append(entry)

    return intervals


def format_km(value, left_padding=3):
    numbers = format(round(float(value), 3), ".3f").split(".")
    zero_left = left_padding - len(numbers[0])
    try:
        return "{}{}+{:03d}".format("0" * zero_left, int(numbers[0]), int(numbers[1]))
    except Exception:
        return ""


def get_values_from_reporting_extra_columns(record, extra_columns, reference_values):

    new_val = {}
    if extra_columns:
        reporting_formatted = get_obj_serialized(record, is_reporting_bi=True)
        for item in extra_columns:
            json_logic = None
            key = item.get("key", "")
            logic = item.get("logic", False)
            is_date = item.get("isDate", False)
            is_select = item.get("isSelect", False)
            is_array = item.get("isArray", False)
            if key and logic:
                try:
                    json_logic = apply_json_logic(logic, reporting_formatted)
                except Exception:
                    pass
            if is_select:
                json_logic = return_select_value(key, record, reference_values)
            if is_array:
                json_logic = return_array_values(item, record, reference_values)
            if not json_logic:
                try:
                    json_logic = record.form_data[to_snake_case(key)]
                except Exception:
                    pass
                if json_logic in EMPTY_VALUES:
                    json_logic = ""
            if is_date and json_logic:
                try:
                    json_logic = parser.parse(json_logic)
                except Exception:
                    json_logic = None
            if isinstance(json_logic, dict):
                new_val.update(json_logic)
            else:
                new_val[key] = json_logic
    return new_val


def translate_weather(weather):
    lookup = {
        "SUNNY": "Aberto",
        "CLOUDY": "Nublado",
        "RAINY": "Chuvoso",
        "NOT_APPLIED": "Não se aplica",
    }
    return lookup.get(weather, "Não se aplica") if weather else None


def translate_condition(condition):
    lookup = {
        "FEASIBLE": "Praticável",
        "UNFEASIBLE": "Impraticável",
        "NOT_APPLIED": "Não se aplica",
    }
    return lookup.get(condition, "Não se aplica") if condition else None


def translate_kind(kind, company):
    options = get_obj_from_path(
        company.custom_options, "dailyreport__fields__kind__selectoptions__options"
    )
    try:
        return next(a["name"] for a in options if a["value"] == kind)
    except Exception:
        return ""


@task
def generate_exported_file(daily_report_export_id):
    """
    Gathers all the data needed for the export, fills the excel template
    and uploads it to S3
    """
    try:
        daily_report_export = DailyReportExport.objects.get(pk=daily_report_export_id)
    except DailyReportExport.DoesNotExist as e:
        logging.error("DailyReportExport not found")
        capture_exception(e)
    else:
        error = True  # Error until proven otherwise
        is_compiled = daily_report_export.is_compiled
        is_pdf = daily_report_export.format == PDF_FORMAT
        export_photos_mdr = daily_report_export.export_photos
        daily_reports = daily_report_export.daily_reports.all()
        multiple_daily_reports = daily_report_export.multiple_daily_reports.all()
        if daily_reports:
            reports = daily_reports
            field_prefix = "daily_report_"
        elif multiple_daily_reports:
            reports = multiple_daily_reports
            field_prefix = "multiple_daily_report_"
        else:
            return
        sort = daily_report_export.sort
        order = daily_report_export.order

        reports = reports.order_by("-date", "uuid")
        # if sort and order are present for the object, we order by it
        if daily_report_export.sort and daily_report_export.order:
            if order == "ASC":
                reports = reports.order_by(sort, "uuid")
            elif order == "DESC":
                reports = reports.order_by(f"-{sort}", "uuid")
        company = reports.first().company
        # Pre-cache occs for export
        reference_values = {
            str(a.uuid): a.name for a in OccurrenceType.objects.filter(company=company)
        }
        hide_reporting_location = get_obj_from_path(
            company.metadata, "hide_reporting_location", default_return=False
        )
        can_view_digital_signature = get_obj_from_path(
            company.metadata, "can_view_digital_signature", default_return=False
        )
        exporter_extra_columns = get_exporter_extra_columns(company)
        exporter_simple_excel_columns_order = get_obj_from_path(
            company.custom_options, "reporting__exporter__simple_excel_columns_order"
        )

        def get_rel_model_fields(
            report,
            model_suffix,
            rel_model_fields,
            pluralize=True,  # Defaults to True since it's the most common case
            select_options_fields=[],
            flat=False,
        ):
            """
            Accesses the M2M relationship using the field name built with the
            suffix (multiple_daily_report_workers for example) and retrieves
            the fields inside instances (listed in rel_model_fields)
            """

            if pluralize:
                report_rel_field_suffix = "{}s".format(model_suffix)
            else:
                report_rel_field_suffix = model_suffix

            rel_model_manager = getattr(report, field_prefix + report_rel_field_suffix)
            active_filter = {"{}_relations__active".format(model_suffix): True}

            values = (
                list(
                    rel_model_manager.filter(**active_filter)
                    .distinct()
                    .values_list(flat=flat, *rel_model_fields)
                    .order_by("starts_at")
                )
                if rel_model_manager
                and rel_model_manager.exists()
                and model_suffix == "occurrence"
                else (
                    list(
                        rel_model_manager.filter(**active_filter)
                        .distinct()
                        .values_list(flat=flat, *rel_model_fields)
                    )
                    if rel_model_manager and rel_model_manager.exists()
                    else []
                )
            )

            if select_options_fields and values:

                for option_field in select_options_fields:

                    options_model_suffix = to_flatten_str(model_suffix)
                    options_field_suffix = to_flatten_str(option_field)
                    options = get_obj_from_path(
                        company.custom_options,
                        "dailyreport{}__fields__{}__selectoptions__options".format(
                            options_model_suffix, options_field_suffix
                        ),
                    )

                    if options:
                        # Get the position of that field and substitute the reference value with
                        # the actual value if any options are found
                        field_position = rel_model_fields.index(option_field)
                        options_lookup = {
                            option["value"]: option["name"] for option in options
                        }

                        for i, value in enumerate(values):
                            try:
                                value_list = list(value)
                                value_list[field_position] = options_lookup[
                                    value_list[field_position]
                                ]
                                values[i] = tuple(value_list)
                            except Exception:
                                pass

            return values

        def translate_photo_value(company, field, value):
            options = get_obj_from_path(
                company.custom_options,
                "reporting_file__fields__{}__selectoptions__options".format(field),
            )
            try:
                return next(a["name"] for a in options if a["value"] == value)
            except Exception:
                return ""

        def get_multiple_daily_report_files(multiple_daily_report_instance):
            multiple_daily_report_files_data = []
            for (
                instance
            ) in multiple_daily_report_instance.multiple_daily_report_files.all():
                params = {}
                params["Bucket"] = "{}-{}px".format(
                    instance.upload.storage.bucket.name, 1000
                )
                params["Key"] = instance.upload.storage._normalize_name(
                    clean_name(instance.upload.name)
                )
                upload_url = (
                    instance.upload.storage.bucket.meta.client.generate_presigned_url(
                        "get_object", Params=params, ExpiresIn=3600
                    )
                )
                resp = requests.get(upload_url, stream=True)
                image_content = resp.content if resp.status_code == 200 else None
                multiple_daily_report_files_data.append(
                    {
                        "multiple_daily_report_id": str(
                            multiple_daily_report_instance.pk
                        ),
                        "content": image_content,
                        "kind": translate_photo_value(company, "kind", instance.kind),
                    }
                )
            return multiple_daily_report_files_data

        try:
            # Data gathering
            reports_data = {}

            # NOTE: We'll use report_order so the export can respect the order_by when creating PDFs
            for report_order, report in enumerate(reports):
                report_uuid = str(report.uuid)
                report_data = {}

                # BaseDailyReport
                report_data.update(
                    {
                        "report_order": report_order,
                        "number": report.number,
                        "date": report.date,
                        "morning_weather": translate_weather(report.morning_weather),
                        "afternoon_weather": translate_weather(
                            report.afternoon_weather
                        ),
                        "night_weather": translate_weather(report.night_weather),
                        "morning_conditions": translate_condition(
                            report.morning_conditions
                        ),
                        "afternoon_conditions": translate_condition(
                            report.afternoon_conditions
                        ),
                        "night_conditions": translate_condition(
                            report.night_conditions
                        ),
                        "morning_start": report.morning_start,
                        "morning_end": report.morning_end,
                        "afternoon_start": report.afternoon_start,
                        "afternoon_end": report.afternoon_end,
                        "night_start": report.night_start,
                        "night_end": report.night_end,
                        "notes": report.notes,
                        "day_without_work": "Não" if report.day_without_work else "Sim",
                        "created_by": (
                            report.created_by.get_full_name()
                            if report.created_by
                            else None
                        ),
                        "responsible": (
                            report.responsible.get_full_name()
                            if report.responsible
                            else None
                        ),
                        "compensation": "SIM" if report.compensation else "NÃO",
                    }
                )

                # ReportingFiles
                reporting_files_data = []
                if export_photos_mdr:
                    for report_file in report.reporting_files.all():
                        params = {}
                        params["Bucket"] = "{}-{}px".format(
                            report_file.upload.storage.bucket.name, 1000
                        )
                        params["Key"] = report_file.upload.storage._normalize_name(
                            clean_name(report_file.upload.name)
                        )
                        upload_url = report_file.upload.storage.bucket.meta.client.generate_presigned_url(
                            "get_object", Params=params, ExpiresIn=3600
                        )

                        # Fetch image data
                        resp = requests.get(upload_url, stream=True)
                        image_content = (
                            resp.content if resp.status_code == 200 else None
                        )

                        reporting_files_data.append(
                            {
                                "content": image_content,
                                "kind": translate_photo_value(
                                    company, "kind", report_file.kind
                                ),
                                "reporting_id": (
                                    str(report_file.reporting.pk)
                                    if report_file.reporting
                                    else ""
                                ),
                            }
                        )
                    report_data["reporting_files"] = reporting_files_data
                    report_data[
                        "multiple_daily_report_files"
                    ] = get_multiple_daily_report_files(report)
                else:
                    report_data["multiple_daily_report_files"] = []
                    report_data["reporting_files"] = reporting_files_data

                # MultipleDailyReport
                firm = report.firm if hasattr(report, "firm") else None
                subcompany = firm.subcompany if firm else None
                contract = report.contract if hasattr(report, "contract") else None
                km_intervals = ""
                if not hide_reporting_location:
                    left_paddings = [3]
                    km_intervals_list = sorted(
                        get_km_intervals_field(report),
                        key=lambda x: (x.get("roadName", ""), x.get("km", 0)),
                    )
                    for km_interval in km_intervals_list:
                        left_paddings.append(len(str(int(km_interval.get("km", 0)))))
                        left_paddings.append(
                            len(str(int(km_interval.get("end_km", 0))))
                        )
                    left_padding = max(left_paddings)
                    for km_interval in km_intervals_list:
                        km_intervals += "{}{} {} - {}".format(
                            "; " if km_intervals else "",
                            km_interval.get("roadName", ""),
                            format_km(km_interval.get("km", 0), left_padding),
                            format_km(km_interval.get("end_km", 0), left_padding),
                        )
                report_data.update(
                    {
                        "firm": firm.name if firm else None,
                        "firm_id": str(firm.uuid) if firm else None,
                        "contract": (
                            contract.extra_info.get("r_c_number")
                            if contract and contract.extra_info.get("r_c_number")
                            else (subcompany.contract if subcompany else None)
                        ),
                        "contract_id": str(contract.pk) if contract else None,
                        "construction_name": (
                            subcompany.construction_name if subcompany else None
                        ),
                        "contract_dates": (
                            "{} - {}".format(
                                contract.contract_start.strftime("%d/%m/%Y"),
                                contract.contract_end.strftime("%d/%m/%Y"),
                            )
                            if contract
                            and contract.contract_start
                            and contract.contract_end
                            else (
                                "{} - {}".format(
                                    subcompany.contract_start_date.strftime("%d/%m/%Y"),
                                    subcompany.contract_end_date.strftime("%d/%m/%Y"),
                                )
                                if subcompany
                                and subcompany.contract_start_date
                                and subcompany.contract_end_date
                                else None
                            )
                        ),
                        "subcompany": subcompany.name if subcompany else None,
                        "office": subcompany.office if subcompany else None,
                        "inspector": (
                            report.inspector.get_full_name()
                            if report.inspector
                            else None
                        ),
                        "header_info": report.header_info,
                        "km_intervals": km_intervals,
                    }
                )

                # Reportings
                reporting_manager = (
                    report.reportings if hasattr(report, "reportings") else None
                )
                report_data["reportings"] = []
                report_data["resources_total"] = 0.0
                if reporting_manager:
                    reportings = (
                        reporting_manager.order_by("road_name", "km")
                        .all()
                        .prefetch_related(
                            *LightReportingSerializer._PREFETCH_RELATED_FIELDS
                        )
                    )
                    procedure_resource_queryset = ProcedureResource.objects.filter(
                        reporting__in=reportings
                    ).prefetch_related(
                        "resource",
                        "service_order_resource",
                        "service_order_resource__resource_contract_unit_price_items",
                        "reporting",
                    )

                    # Process reporting data
                    for reporting in reportings:
                        reporting_data = get_reporting_data(
                            reporting, exporter_extra_columns, reference_values
                        )

                        # Handle KMs
                        road_name = reporting.road_name
                        km = reporting.km
                        end_km = reporting.end_km
                        if "stretches" not in report_data:
                            report_data["stretches"] = {}
                        if road_name in report_data["stretches"]:
                            # If there's a smaller km, set as the new start
                            if km < report_data["stretches"][road_name]["km"]:
                                report_data["stretches"][road_name]["km"] = km

                            # If there's a bigger or non None end_km, set as new end
                            old_end_km = report_data["stretches"][road_name]["end_km"]
                            if end_km and (old_end_km is None or end_km > old_end_km):
                                report_data["stretches"][road_name]["end_km"] = end_km
                        else:
                            report_data["stretches"][road_name] = {
                                "km": km,
                                "end_km": end_km,
                            }

                        # Handle resources
                        resources = {}
                        for procedure_resource in procedure_resource_queryset:
                            if procedure_resource.reporting_id != reporting.pk:
                                continue
                            resource_name = procedure_resource.resource.name
                            total_price = procedure_resource.total_price

                            if "total" in resources:
                                resources["total"] += total_price
                            else:
                                resources["total"] = total_price

                            if "proc_resources" not in resources:
                                resources["proc_resources"] = []

                            resources["proc_resources"].append(
                                {
                                    "resource_name": resource_name,
                                    "resource_unit": procedure_resource.resource.unit,
                                    "sort_string": (
                                        procedure_resource.service_order_resource.resource_contract_unit_price_items.first().sort_string
                                        if procedure_resource.service_order_resource
                                        and procedure_resource.service_order_resource.resource_contract_unit_price_items.exists()
                                        else ""
                                    ),
                                    "amount": procedure_resource.amount,
                                    "unit_price": procedure_resource.unit_price,
                                    "total_price": total_price,
                                }
                            )
                        reporting_data["resources"] = resources
                        report_data["resources_total"] += resources.get("total", 0.0)
                        reporting_data["id"] = str(reporting.pk)
                        report_data["reportings"].append(reporting_data)

                    # Compile KMs
                    road_names = ""
                    start_kms = ""
                    end_kms = ""
                    if "stretches" in report_data:
                        for road_name, km_data in report_data["stretches"].items():
                            road_names += (
                                "; {}".format(road_name)
                                if road_names
                                else "{}".format(road_name)
                            )
                            start_kms += (
                                "; {:.3f}".format(km_data["km"])
                                if start_kms
                                else "{:.3f}".format(km_data["km"])
                            )
                            end_kms += (
                                "; {:.3f}".format(km_data["end_km"])
                                if end_kms
                                else "{:.3f}".format(km_data["end_km"])
                            )
                            report_data["road_names"] = road_names
                            report_data["start_kms"] = start_kms
                            report_data["end_kms"] = end_kms
                    else:
                        report_data["road_names"] = None
                        report_data["start_kms"] = None
                        report_data["end_kms"] = None

                # SubCompany type logic
                if subcompany:
                    if subcompany.subcompany_type == "HIRING":
                        report_data["hiring"] = subcompany.name
                        report_data["hired"] = None

                        # Handle logos
                        try:
                            hiring_logo_url = subcompany.logo.url
                            resp = requests.get(hiring_logo_url, stream=True)
                            report_data["hiring_logo"] = (
                                resp.content if resp.status_code == 200 else None
                            )
                        except Exception:
                            report_data["hiring_logo"] = None

                        report_data["hired_logo"] = None
                    else:
                        report_data["hiring"] = (
                            subcompany.hired_by_subcompany.name
                            if subcompany.hired_by_subcompany
                            else None
                        )
                        report_data["hired"] = subcompany.name

                        # Handle logos
                        try:
                            hiring_logo_url = (
                                subcompany.hired_by_subcompany.logo.url
                                if subcompany.hired_by_subcompany
                                and subcompany.hired_by_subcompany.logo
                                else None
                            )
                            resp = (
                                requests.get(hiring_logo_url, stream=True)
                                if hiring_logo_url
                                else None
                            )
                            report_data["hiring_logo"] = (
                                resp.content
                                if resp and resp.status_code == 200
                                else None
                            )
                        except Exception:
                            report_data["hiring_logo"] = None

                        try:
                            hired_logo_url = subcompany.logo.url
                            resp = requests.get(hired_logo_url, stream=True)
                            report_data["hired_logo"] = (
                                resp.content if resp.status_code == 200 else None
                            )
                        except Exception:
                            report_data["hired_logo"] = None
                else:
                    report_data["hiring"] = None
                    report_data["hired"] = None

                # Related models
                report_data.update(
                    {
                        "workers": get_rel_model_fields(
                            report,
                            model_suffix="worker",
                            rel_model_fields=[
                                "role",
                                "amount",
                                "contract_item_administration__resource__resource__name",
                                "contract_item_administration__sort_string",
                                "extra_hours",
                                "contract_item_administration__resource__contract_id",
                            ],
                            select_options_fields=["role"],
                        ),
                        "equipment": get_rel_model_fields(
                            report,
                            model_suffix="equipment",
                            rel_model_fields=[
                                "description",
                                "amount",
                                "kind",
                                "contract_item_administration__resource__resource__name",
                                "contract_item_administration__sort_string",
                                "extra_hours",
                                "contract_item_administration__resource__contract_id",
                            ],
                            select_options_fields=["description"],
                            pluralize=False,
                        ),
                        "vehicles": get_rel_model_fields(
                            report,
                            model_suffix="vehicle",
                            rel_model_fields=[
                                "description",
                                "amount",
                                "kind",
                                "contract_item_administration__resource__resource__name",
                                "contract_item_administration__sort_string",
                                "extra_hours",
                                "contract_item_administration__resource__contract_id",
                            ],
                            select_options_fields=["description"],
                        ),
                        "signaling": get_rel_model_fields(
                            report,
                            model_suffix="signaling",
                            rel_model_fields=["kind"],
                            select_options_fields=["kind"],
                            pluralize=False,
                        ),
                        "occurrences": get_rel_model_fields(
                            report,
                            model_suffix="occurrence",
                            rel_model_fields=[
                                "origin",
                                "description",
                                "impact_duration",
                                "starts_at",
                                "ends_at",
                                "extra_info",
                            ],
                            select_options_fields=["description"],
                        ),
                        "resources": get_rel_model_fields(
                            report,
                            model_suffix="resource",
                            rel_model_fields=["amount", "resource__name", "kind"],
                        ),
                    }
                )

                reports_data[report_uuid] = report_data

            data_offset = None

            def get_cell(column_letter, list_pos):
                """
                Apply the data_offset to get the cell for the current
                group of data
                """
                if data_offset:
                    return "{}{}".format(column_letter, list_pos + data_offset)

            def fill_reportings_section(templ_ws, report_data):
                report_data_reportings = report_data["reportings"]
                new_row = len([cell for cell in list(templ_ws["A"]) if cell.value])
                HEADER_ROW = 2
                COLUMN_START_POS = 2
                PRICE_FORMAT = "R$ #,##0.00"
                DATE_FORMAT = "dd/mm/yyyy"
                static_columns = get_reporting_static_columns()
                extra_columns = get_exporter_extra_columns_parsed_infos(
                    exporter_extra_columns
                )
                fields_to_hide_reporting_location = (
                    get_fields_to_hide_reporting_location()
                )
                if hide_reporting_location:
                    static_columns = remove_fields_to_hide_reporting_location(
                        fields_to_hide_reporting_location, static_columns
                    )
                columns = {**static_columns, **extra_columns}
                static_first_cell_style = templ_ws["A2"]._style
                value_cell_style = templ_ws["A3"]._style
                header_font = Font(name="Calibri", size=11, bold=True, color="000000")

                columns_order = []
                if exporter_simple_excel_columns_order:
                    columns_order = [
                        column
                        for column in exporter_simple_excel_columns_order
                        if column in columns
                    ]
                else:
                    columns_order = [*columns.keys()]

                def cell_by_pos(column_pos, row_pos=None):
                    column_letter = get_column_letter(column_pos)
                    return cell(column_letter, row_pos)

                def cell(column_letter, row_pos=None):
                    return "{}{}".format(column_letter, row_pos or new_row)

                def fill_header_cell(pos, value):
                    cell_name = cell_by_pos(pos, HEADER_ROW)
                    templ_ws[cell_name] = value
                    templ_ws[cell_name]._style = copy(static_first_cell_style)
                    if value != static_columns.get("number"):
                        templ_ws[cell_name].font = header_font

                def fill_cell_reporting_section(pos, value, number_format=None):
                    cell_name = cell_by_pos(pos)
                    templ_ws[cell_name] = value
                    templ_ws[cell_name]._style = copy(value_cell_style)
                    if number_format:
                        templ_ws[cell_name].number_format = number_format

                def add_resource_headers():
                    try:
                        max_num_resources = max(
                            [
                                len(reporting_data["resources"]["proc_resources"])
                                for report_data in reports_data.values()
                                for reporting_data in report_data["reportings"]
                                if reporting_data["resources"]
                            ]
                        )
                    except Exception:
                        max_num_resources = 0

                    new_column_pos = COLUMN_START_POS + len(columns_order)
                    fill_header_cell(new_column_pos, "Valor total")

                    resource_fields = [
                        "Código",
                        "Recurso",
                        "Quantidade",
                        "Valor",
                        "Valor Unitário",
                    ]
                    for i in range(1, max_num_resources + 1):
                        for field in resource_fields:
                            new_column_pos += 1
                            fill_header_cell(new_column_pos, "{} {}".format(field, i))

                for pos, column_key in enumerate(columns_order, start=COLUMN_START_POS):
                    fill_header_cell(pos, columns.get(column_key, ""))
                add_resource_headers()

                for reporting_data in report_data_reportings:
                    new_row += 1

                    templ_ws[cell("A")] = report_data["number"]
                    for pos, column_key in enumerate(
                        columns_order, start=COLUMN_START_POS
                    ):
                        is_field_from_hide_reporting_location = (
                            column_key in fields_to_hide_reporting_location
                        )
                        if not is_field_from_hide_reporting_location or (
                            is_field_from_hide_reporting_location
                            and not hide_reporting_location
                        ):
                            is_extra_column = column_key in extra_columns
                            resource_value = (
                                reporting_data["extra_columns"].get(column_key)
                                if is_extra_column
                                else reporting_data.get(column_key)
                            )
                            if (
                                column_key
                                in [
                                    "km",
                                    "end_km",
                                    "km_reference",
                                    "project_km",
                                    "project_end_km",
                                ]
                                and resource_value
                            ):
                                cell_assignment = round(resource_value, 3)
                            elif type(resource_value) is datetime:
                                cell_assignment = datetime_to_date(resource_value)
                            else:
                                cell_assignment = resource_value

                            # NOTE: If we get a value that's not able to be converted for the excel file
                            # leave the cell empty
                            try:
                                templ_ws[cell_by_pos(pos)] = cell_assignment
                            except Exception as e:
                                if cell_assignment:
                                    sentry_sdk.capture_exception(e)
                                templ_ws[cell_by_pos(pos)] = None

                            if type(cell_assignment) is date:
                                templ_ws[cell_by_pos(pos)].number_format = DATE_FORMAT

                    # Dynamic part (resources)
                    resources = reporting_data["resources"]
                    if resources:
                        new_column_pos = COLUMN_START_POS + len(columns_order)
                        fill_cell_reporting_section(
                            new_column_pos, resources.pop("total"), PRICE_FORMAT
                        )
                        for proc_resource in resources["proc_resources"]:
                            resource_values = [
                                {
                                    "value": proc_resource["sort_string"],
                                    "number_format": None,
                                },
                                {
                                    "value": "{} ({})".format(
                                        proc_resource["resource_name"],
                                        proc_resource["resource_unit"],
                                    ),
                                    "number_format": None,
                                },
                                {
                                    "value": proc_resource["amount"],
                                    "number_format": None,
                                },
                                {
                                    "value": proc_resource["total_price"],
                                    "number_format": PRICE_FORMAT,
                                },
                                {
                                    "value": proc_resource["unit_price"],
                                    "number_format": PRICE_FORMAT,
                                },
                            ]
                            for resource_value in resource_values:
                                new_column_pos += 1
                                fill_cell_reporting_section(
                                    new_column_pos,
                                    resource_value.get("value"),
                                    resource_value.get("number_format"),
                                )

            def fill_rdo_photo_section(templ_wb, report_data, row_pos=0):
                templ_ws = templ_wb["Fotos"]
                data_offset = 3
                templ_ws[get_cell("A", row_pos)] = report_data["number"]
                templ_ws[get_cell("B", row_pos)] = ""
                templ_ws[get_cell("C", row_pos)] = report_data["firm"]
                templ_ws[get_cell("D", row_pos)] = report_data["date"]
                templ_ws[get_cell("D", row_pos)].number_format = "dd/mm/yyyy"

                cell_alignment = Alignment(horizontal="center", vertical="center")
                new_column_pos = column_index_from_string("E")
                removed_images = 0
                for multiple_daily_report_file in report_data[
                    "multiple_daily_report_files"
                ]:
                    if multiple_daily_report_file["content"]:
                        try:
                            temp_file = NamedTemporaryFile(delete=False)
                            temp_file.write(multiple_daily_report_file["content"])
                            image = Image(temp_file.name)
                            temp_file.close()
                            if image.format.lower() not in ["png", "jpeg", "jpg"]:
                                removed_images += 1
                        except Exception:
                            return removed_images
                        ROW_HEIGHT = 90
                        COLUMN_WIDTH = 30
                        if image.format.lower() in ["png", "jpeg", "jpg"]:
                            # Photo
                            from_anchor_marker = AnchorMarker(
                                col=new_column_pos - 1,
                                row=data_offset + row_pos - 1,
                            )
                            to_anchor_marker = AnchorMarker(
                                col=new_column_pos, row=data_offset + row_pos
                            )
                            anchor = TwoCellAnchor(
                                _from=from_anchor_marker, to=to_anchor_marker
                            )
                            templ_ws.add_image(image, anchor)
                            files_to_close.append(temp_file)
                            templ_ws.row_dimensions[
                                data_offset + row_pos
                            ].height = ROW_HEIGHT
                            templ_ws.column_dimensions[
                                get_column_letter(new_column_pos)
                            ].width = COLUMN_WIDTH
                            new_column_pos += 1
                            # Kind
                            kind_cell = "{}{}".format(
                                get_column_letter(new_column_pos),
                                data_offset + row_pos,
                            )
                            templ_ws[kind_cell].value = multiple_daily_report_file[
                                "kind"
                            ]
                            templ_ws[kind_cell].alignment = cell_alignment
                            templ_ws.column_dimensions[
                                get_column_letter(new_column_pos)
                            ].width = COLUMN_WIDTH
                            new_column_pos += 1

                    else:
                        removed_images += 1
                return removed_images

            def fill_photos_section(templ_wb, report_data, reporting_data, row_pos=0):
                templ_ws = templ_wb["Fotos"]
                data_offset = 3
                templ_ws[get_cell("A", row_pos)] = report_data["number"]
                templ_ws[get_cell("B", row_pos)] = reporting_data.get("number", "")
                templ_ws[get_cell("C", row_pos)] = report_data["firm"]
                templ_ws[get_cell("D", row_pos)] = report_data["date"]
                templ_ws[get_cell("D", row_pos)].number_format = "dd/mm/yyyy"

                cell_alignment = Alignment(horizontal="center", vertical="center")

                new_column_pos = column_index_from_string("E")
                removed_images = 0
                for reporting_file in report_data["reporting_files"]:
                    if reporting_file["reporting_id"] != reporting_data["id"]:
                        continue
                    photo_content = reporting_file["content"]
                    if photo_content:
                        try:
                            temp_file = NamedTemporaryFile(delete=False)
                            temp_file.write(photo_content)
                            image = Image(temp_file.name)
                            temp_file.close()
                            if image.format.lower() not in ["png", "jpeg", "jpg"]:
                                removed_images += 1
                        except Exception:
                            continue
                        ROW_HEIGHT = 90
                        COLUMN_WIDTH = 30
                        if image.format.lower() in ["png", "jpeg", "jpg"]:
                            # Photo
                            from_anchor_marker = AnchorMarker(
                                col=new_column_pos - 1, row=data_offset + row_pos - 1
                            )
                            to_anchor_marker = AnchorMarker(
                                col=new_column_pos, row=data_offset + row_pos
                            )
                            anchor = TwoCellAnchor(
                                _from=from_anchor_marker, to=to_anchor_marker
                            )
                            templ_ws.add_image(image, anchor)
                            files_to_close.append(temp_file)
                            templ_ws.row_dimensions[
                                data_offset + row_pos
                            ].height = ROW_HEIGHT
                            templ_ws.column_dimensions[
                                get_column_letter(new_column_pos)
                            ].width = COLUMN_WIDTH
                            new_column_pos += 1

                            # Kind
                            kind_cell = "{}{}".format(
                                get_column_letter(new_column_pos), data_offset + row_pos
                            )
                            templ_ws[kind_cell].value = reporting_file["kind"]
                            templ_ws[kind_cell].alignment = cell_alignment
                            templ_ws.column_dimensions[
                                get_column_letter(new_column_pos)
                            ].width = COLUMN_WIDTH
                            new_column_pos += 1

                    else:
                        removed_images += 1
                return removed_images

            def get_reporting_max_files_for_compiled():
                reportings_id = []
                duplicated_qntd = []
                for report_data in reports_data.values():
                    for reporting_file_data in (
                        report_data["reporting_files"]
                        + report_data["multiple_daily_report_files"]
                    ):
                        reportings_id.append(
                            reporting_file_data["reporting_id"]
                            if reporting_file_data.get("reporting_id")
                            else reporting_file_data["multiple_daily_report_id"]
                        )

                for id in list(set(reportings_id)):
                    duplicated_qntd.append(reportings_id.count(id))

                return max(duplicated_qntd)

            def get_reporting_max_files_():
                reportings_id = []
                duplicated_qntd = []
                for reporting_file_data in (
                    report_data["reporting_files"]
                    + report_data["multiple_daily_report_files"]
                ):
                    reportings_id.append(
                        reporting_file_data["reporting_id"]
                        if reporting_file_data.get("reporting_id")
                        else reporting_file_data["multiple_daily_report_id"]
                    )
                for id in list(set(reportings_id)):
                    duplicated_qntd.append(reportings_id.count(id))

                return max(duplicated_qntd)

            def add_photo_headers(templ_wb, removed_images):
                templ_ws = templ_wb["Fotos"]
                HEADER_ROW = 2
                next_column = column_index_from_string("D")
                header_style = templ_ws["C2"]._style
                try:
                    max_num_rep_files = (
                        get_reporting_max_files_for_compiled()
                        if is_compiled
                        else get_reporting_max_files_()
                    ) - removed_images
                except Exception:
                    max_num_rep_files = 0

                for num in range(max_num_rep_files):
                    next_column += 1
                    photo_cell = templ_ws.cell(column=next_column, row=HEADER_ROW)
                    photo_cell.value = "Foto {}".format(num + 1)
                    photo_cell._style = header_style

                    next_column += 1
                    photo_cell = templ_ws.cell(column=next_column, row=HEADER_ROW)
                    photo_cell.value = "Tipo {}".format(num + 1)
                    photo_cell._style = header_style

            def hide_reporting_location_compiled(templ_wb):
                templ_ws = templ_wb["RDO Compilado"]
                templ_ws.delete_cols(14, 3)

            def delete_row_with_merged_ranges(templ_ws, row):
                # Must be used at the end
                templ_ws.delete_rows(row)
                for mcr in templ_ws.merged_cells:
                    if row < mcr.min_row:
                        mcr.shift(row_shift=-1)
                    elif row < mcr.max_row:
                        mcr.shrink(bottom=1)

            def remove_items_rdo(report_data):
                report_data.update(
                    {
                        "morning_start": "",
                        "morning_end": "",
                        "afternoon_start": "",
                        "afternoon_end": "",
                        "night_start": "",
                        "night_end": "",
                        "morning_conditions": "",
                        "afternoon_conditions": "",
                        "night_conditions": "",
                        "morning_weather": "",
                        "afternoon_weather": "",
                        "night_weather": "",
                        "road_names": "",
                        "start_kms": "",
                        "end_kms": "",
                        "resources_total": "",
                        "workers": [],
                        "equipment": [],
                        "vehicles": [],
                        "resources": [],
                        "reportings": [],
                        "signaling": [],
                    }
                )

            # NOTE: Unfortunately the method .save() needs the image files open when saving the final Excel file
            # this means we can't use the proper way to close files (with as) and have to close the files after saving
            files_to_close = []
            if is_compiled:
                templ_wb = load_workbook(
                    filename="apps/daily_reports/templates/rdo_compilado.xlsx",
                    read_only=False,
                    keep_vba=True,
                )

                # Insert "Quadros de Controle" columns
                base_headers = ["Serial do RDO", "Equipe", "Data"]
                headers = DefaultDict(list)
                if report_data.get("day_without_work") == "Não":
                    remove_items_rdo(report_data)
                for report_data in reports_data.values():
                    headers["workers"] += [
                        (
                            resource_name
                            if (role and is_valid_uuid(str(role)))
                            else (role if role else resource_name)
                        )
                        for (role, _, resource_name, _, _, _) in report_data["workers"]
                    ]
                    headers["equipment"] += [
                        (
                            resource_name
                            if (description and is_valid_uuid(str(description)))
                            else (description if description else resource_name)
                        )
                        for (
                            description,
                            _,
                            _,
                            resource_name,
                            _,
                            _,
                            _,
                        ) in report_data["equipment"]
                    ]
                    headers["vehicles"] += [
                        (
                            resource_name
                            if (description and is_valid_uuid(str(description)))
                            else (description if description else resource_name)
                        )
                        for (
                            description,
                            _,
                            _,
                            resource_name,
                            _,
                            _,
                            _,
                        ) in report_data["vehicles"]
                    ]
                    headers["signaling"] += [
                        kind for (kind,) in report_data["signaling"]
                    ]

                # Remove duplicates and sort it
                headers = {
                    key: list(set(sorted(value, key=value.count, reverse=True)))
                    for key, value in headers.items()
                }

                # Add the headers in order
                templ_ws = templ_wb["Quadro de Controle"]
                HEADERS_ORDER = ["workers", "equipment", "vehicles", "signaling"]
                headers_flat = []
                # Used for merging cells of the main header
                section_lengths = []  # contains tuples (header_model, length)
                for header_model in HEADERS_ORDER:
                    headers_flat += headers[header_model]
                    section_lengths.append((header_model, len(headers[header_model])))

                all_headers = base_headers + headers_flat
                HEADER_ROW = 2
                header_style = (
                    templ_wb["Apontamentos e Recursos"]
                    .cell(column=1, row=HEADER_ROW)
                    ._style
                )

                # Fill sub headers
                for i, header in enumerate(all_headers):
                    cell = templ_ws.cell(column=i + 1, row=HEADER_ROW)
                    cell.value = header
                    cell._style = header_style

                # Fill main headers
                MAIN_HEADER_ROW = 1
                MODEL_TO_TITLE = {
                    "workers": "Função e Quantidade",
                    "equipment": "Tipo de Equipamento e Quantidade",
                    "vehicles": "Veículo",
                    "signaling": "Tipo de Sinalização",
                }
                start_column = len(base_headers) + 1
                for header_model, length in section_lengths:
                    if length:
                        # Change merged cells according to the first one
                        main_cell = templ_ws.cell(
                            row=MAIN_HEADER_ROW, column=start_column
                        )
                        main_cell.value = MODEL_TO_TITLE[header_model]
                        main_cell.fill = PatternFill("solid", fgColor="FCE4D6")
                        main_cell.font = Font(
                            name="Calibri", size=11, bold=True, color="000000"
                        )
                        main_cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrapText=True,
                            wrap_text=True,
                        )
                        main_cell.border = Border(
                            left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                        )

                        templ_ws.merge_cells(
                            "{0}{2}:{1}{2}".format(
                                get_column_letter(start_column),
                                get_column_letter(start_column + length - 1),
                                MAIN_HEADER_ROW,
                            )
                        )

                        # Set the new start (+1 to be outside of the merged cells)
                        start_column += length

                def fill_cell(worksheet, cell_column, cell_row, cell_data):
                    templ_style = templ_ws[get_cell(cell_column, 0)]._style
                    worksheet[get_cell(cell_column, cell_row)] = cell_data
                    worksheet[get_cell(cell_column, cell_row)]._style = templ_style

                report_iterator = 0
                rdos_without_work = 0
                for i, report_data in enumerate(reports_data.values()):
                    # Fill "RDO Compilado"
                    templ_ws = templ_wb["RDO Compilado"]
                    data_offset = 4
                    fill_cell(templ_ws, "A", i, report_data["number"])
                    fill_cell(templ_ws, "B", i, report_data["firm"])
                    fill_cell(templ_ws, "C", i, report_data["date"])
                    fill_cell(templ_ws, "D", i, report_data["created_by"])
                    fill_cell(templ_ws, "E", i, report_data["responsible"])
                    fill_cell(templ_ws, "F", i, report_data["day_without_work"])
                    fill_cell(templ_ws, "G", i, report_data["contract"])
                    fill_cell(templ_ws, "H", i, report_data["construction_name"])
                    fill_cell(templ_ws, "I", i, report_data["contract_dates"])
                    fill_cell(templ_ws, "J", i, report_data["subcompany"])
                    fill_cell(templ_ws, "K", i, report_data["hiring"])
                    fill_cell(templ_ws, "L", i, report_data["hired"])
                    fill_cell(templ_ws, "M", i, report_data["office"])
                    if not hide_reporting_location:
                        fill_cell(templ_ws, "N", i, report_data["road_names"])
                        fill_cell(templ_ws, "O", i, report_data["start_kms"])
                        fill_cell(templ_ws, "P", i, report_data["end_kms"])
                    fill_cell(templ_ws, "Q", i, report_data["morning_weather"])
                    fill_cell(templ_ws, "R", i, report_data["afternoon_weather"])
                    fill_cell(templ_ws, "S", i, report_data["night_weather"])
                    fill_cell(templ_ws, "T", i, report_data["morning_conditions"])
                    fill_cell(templ_ws, "U", i, report_data["afternoon_conditions"])
                    fill_cell(templ_ws, "V", i, report_data["night_conditions"])
                    fill_cell(templ_ws, "W", i, report_data["morning_start"])
                    fill_cell(templ_ws, "X", i, report_data["morning_end"])
                    fill_cell(templ_ws, "Y", i, report_data["afternoon_start"])
                    fill_cell(templ_ws, "Z", i, report_data["afternoon_end"])
                    fill_cell(templ_ws, "AA", i, report_data["night_start"])
                    fill_cell(templ_ws, "AB", i, report_data["night_end"])
                    fill_cell(templ_ws, "AC", i, report_data["resources_total"])
                    fill_cell(templ_ws, "AD", i, report_data["notes"])

                    # Fill "Quadro de Controle"
                    templ_ws = templ_wb["Quadro de Controle"]
                    row_data = []
                    if report_data.get("day_without_work") == "Sim":
                        row_data = [
                            report_data["number"],
                            report_data["firm"],
                            report_data["date"],
                        ]
                    for worker_header in headers["workers"]:
                        try:
                            amount = next(
                                amount
                                for (
                                    role,
                                    amount,
                                    resource_name,
                                    _,
                                    _,
                                    _,
                                ) in report_data["workers"]
                                if role == worker_header
                                or resource_name == worker_header
                            )
                        except StopIteration:
                            amount = 0

                        row_data.append(amount)

                    for equipment_header in headers["equipment"]:
                        try:
                            amount = next(
                                amount
                                for (
                                    description,
                                    amount,
                                    _,
                                    resource_name,
                                    _,
                                    _,
                                    _,
                                ) in report_data["equipment"]
                                if description == equipment_header
                                or resource_name == equipment_header
                            )
                        except StopIteration:
                            amount = 0

                        row_data.append(amount)

                    for vehicle_header in headers["vehicles"]:
                        try:
                            amount = next(
                                amount
                                for (
                                    description,
                                    amount,
                                    _,
                                    resource_name,
                                    _,
                                    _,
                                    _,
                                ) in report_data["vehicles"]
                                if description == vehicle_header
                                or resource_name == vehicle_header
                            )
                        except StopIteration:
                            amount = 0

                        row_data.append(amount)

                    for signaling_header in headers["signaling"]:
                        try:
                            present = next(
                                "Sim"
                                for (kind,) in report_data["signaling"]
                                if kind == signaling_header
                            )
                        except StopIteration:
                            present = "Não"

                        row_data.append(present)

                    CELLS_ROW = 3
                    data_offset = 3
                    cell_style = (
                        templ_wb["Apontamentos e Recursos"]
                        .cell(column=1, row=CELLS_ROW)
                        ._style
                    )

                    if report_data.get("day_without_work") == "Sim":
                        for j, data in enumerate(row_data):
                            cell = templ_ws.cell(
                                column=j + 1, row=i + data_offset - rdos_without_work
                            )
                            cell.value = data
                            cell._style = copy(cell_style)
                            if j == 2:
                                cell.number_format = "dd/mm/yyyy"
                            if type(data) is time:
                                cell.number_format = "hh:mm"
                    else:
                        rdos_without_work += 1

                    # Fill "Ocorrências"
                    templ_ws = templ_wb["Ocorrências"]
                    data_offset = 0
                    occurrence_option_custom_options = get_obj_from_path(
                        company.custom_options,
                        "daily_report_occurrence__fields__origin__select_options__options",
                    )
                    occurrence_option = (
                        occurrence_option_custom_options
                        if occurrence_option_custom_options
                        else TRANSLATE_OCCURRENCE_ORIGIN_CHOICES
                    )
                    occurrence_option = {
                        a["value"]: a["name"] for a in occurrence_option
                    }
                    occurrence_row = templ_ws.max_row + 1
                    for (
                        origin,
                        description,
                        impact_duration,
                        starts_at,
                        ends_at,
                        extra_info,
                    ) in report_data["occurrences"]:
                        templ_ws["A{}".format(occurrence_row)].value = report_data[
                            "number"
                        ]
                        templ_ws[
                            "B{}".format(occurrence_row)
                        ].value = occurrence_option.get(origin, "")
                        templ_ws["C{}".format(occurrence_row)].value = description
                        templ_ws["D{}".format(occurrence_row)].value = starts_at
                        templ_ws["D{}".format(occurrence_row)].number_format = "hh:mm"
                        templ_ws["E{}".format(occurrence_row)].value = ends_at
                        templ_ws["E{}".format(occurrence_row)].number_format = "hh:mm"
                        templ_ws["F{}".format(occurrence_row)].value = extra_info
                        templ_ws["G{}".format(occurrence_row)].value = impact_duration

                        occurrence_row += 1

                    # Fill "Apontamentos e Recursos"
                    templ_ws = templ_wb["Apontamentos e Recursos"]
                    data_offset = 3
                    fill_reportings_section(templ_ws, report_data)

                    if export_photos_mdr:
                        # Fill "Fotos"
                        removed_images = 0
                        for reporting_data in report_data["reportings"]:
                            if report_data["reporting_files"]:
                                removed_images = fill_photos_section(
                                    templ_wb,
                                    report_data,
                                    reporting_data,
                                    report_iterator,
                                )
                                report_iterator += 1

                        if report_data["multiple_daily_report_files"]:
                            removed_images = fill_rdo_photo_section(
                                templ_wb,
                                report_data,
                                report_iterator,
                            )
                            report_iterator += 1

                        # Add photo section headers
                        add_photo_headers(templ_wb, removed_images)
                    else:
                        if "Fotos" in templ_wb.sheetnames:
                            del templ_wb["Fotos"]

                # Add detailed control board sheet
                fill_detailed_control_board(
                    templ_wb, reports_data, company, daily_report_export
                )

                if hide_reporting_location:
                    hide_reporting_location_compiled(templ_wb)

                with NamedTemporaryFile() as temp_file:
                    templ_wb.save(temp_file.name)
                    daily_report_export.exported_file.save(
                        "rdo_compilado.xlsm", ContentFile(temp_file.read())
                    )
                    error = False
            else:
                # Create temp folder to save the excel files
                temp_dir = "/tmp/rdos/"
                os.makedirs(temp_dir, exist_ok=True)

                # NOTE: We won't need macros for the PDF export
                file_ext = ".xlsx" if is_pdf else ".xlsm"

                for i, (report_uuid, report_data) in enumerate(reports_data.items()):
                    report_iterator = 0
                    templ_wb = load_workbook(
                        filename="apps/daily_reports/templates/rdo_individual.xlsx",
                        read_only=False,
                        keep_vba=True,
                    )
                    templ_ws = templ_wb["RDO Individual"]

                    if report_data.get("day_without_work") == "Não":
                        remove_items_rdo(report_data)
                    # Logos
                    LOGO_ROW_START = 1
                    LOGO_ROW_END = 6
                    if "hiring_logo" in report_data and report_data["hiring_logo"]:
                        temp_file_hiring_logo = NamedTemporaryFile()
                        temp_file_hiring_logo.write(report_data["hiring_logo"])

                        from_anchor_marker = AnchorMarker(
                            col=column_index_from_string("B") - 1, row=LOGO_ROW_START
                        )
                        to_anchor_marker = AnchorMarker(
                            col=column_index_from_string("C"), row=LOGO_ROW_END
                        )
                        anchor = TwoCellAnchor(
                            _from=from_anchor_marker, to=to_anchor_marker
                        )

                        hiring_logo = Image(temp_file_hiring_logo)
                        templ_ws.add_image(hiring_logo, anchor)
                        files_to_close.append(temp_file_hiring_logo)

                    if "hired_logo" in report_data and report_data["hired_logo"]:
                        temp_file_hired_logo = NamedTemporaryFile()
                        temp_file_hired_logo.write(report_data["hired_logo"])

                        from_anchor_marker = AnchorMarker(
                            col=column_index_from_string("I") - 1, row=LOGO_ROW_START
                        )
                        to_anchor_marker = AnchorMarker(
                            col=column_index_from_string("J"), row=LOGO_ROW_END
                        )
                        anchor = TwoCellAnchor(
                            _from=from_anchor_marker, to=to_anchor_marker
                        )

                        hired_logo = Image(temp_file_hired_logo)
                        templ_ws.add_image(hired_logo, anchor)
                        files_to_close.append(temp_file_hired_logo)

                    # Basic Info
                    header_info = report_data["header_info"]
                    templ_ws["C8"] = header_info.get("hirer_name")
                    templ_ws["F8"] = report_data["date"]
                    templ_ws["F8"].number_format = "dd/mm/yyyy"
                    templ_ws["I8"] = report_data["day_without_work"]
                    templ_ws["C9"] = report_data["inspector"]
                    templ_ws["F9"] = report_data["firm"]
                    templ_ws["I9"] = header_info.get("responsibles_hirer")
                    templ_ws["C10"] = report_data["responsible"]
                    templ_ws["F10"] = header_info.get("hired_name")
                    templ_ws["I10"] = header_info.get("responsibles_hired")
                    templ_ws["B11"] = "{} {}".format(
                        "Endereço Contratada:", header_info.get("office_hirer", "")
                    )
                    templ_ws["I11"] = header_info.get("contract_number")
                    templ_ws["C12"] = header_info.get("contract_deadline")
                    templ_ws["F12"] = header_info.get("contract_starts_at")
                    templ_ws["I12"] = header_info.get("contract_execution_days")
                    construction_name = header_info.get("construction_name")
                    templ_ws["B13"] = "{} {}".format(
                        "Nome da Obra:",
                        construction_name if construction_name is not None else "",
                    )

                    if header_info.get("object_description"):
                        templ_ws["D13"] = "{} {}".format(
                            "Descrição do Objeto Contratado:",
                            header_info.get("object_description"),
                        )
                    if not hide_reporting_location and report_data["km_intervals"]:
                        templ_ws["B14"] = report_data["km_intervals"]

                    # Condições de Trabalho
                    templ_ws["C18"] = report_data["morning_start"]
                    templ_ws["D18"] = report_data["morning_end"]
                    templ_ws["C19"] = report_data["afternoon_start"]
                    templ_ws["D19"] = report_data["afternoon_end"]
                    templ_ws["C20"] = report_data["night_start"]
                    templ_ws["D20"] = report_data["night_end"]
                    templ_ws["F17"] = report_data["morning_conditions"]
                    templ_ws["F18"] = report_data["afternoon_conditions"]
                    templ_ws["F19"] = report_data["night_conditions"]
                    templ_ws["I17"] = report_data["morning_weather"]
                    templ_ws["I18"] = report_data["afternoon_weather"]
                    templ_ws["I19"] = report_data["night_weather"]
                    # Quadros de Controle
                    header_style = templ_ws["B7"]._style
                    subheader_font = Font(name="Calibri", size=11, bold=True)
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    cell_alignment = Alignment(vertical="center", wrapText=True)

                    def add_vertical_header(text):
                        """
                        Adds the text to a new line that's going to receive the
                        header style and be merged
                        """
                        header_row = templ_ws.max_row + 1
                        header_cell = templ_ws["B{}".format(header_row)]
                        header_cell.value = text
                        header_cell._style = header_style

                        templ_ws.merge_cells("B{0}:J{0}".format(header_row))

                    def apply_subheader_style(cell):
                        cell.font = subheader_font
                        cell.alignment = header_alignment

                    def apply_data_cell_style(cell):
                        cell.alignment = cell_alignment

                    add_vertical_header("FUNCIONÁRIOS")
                    current_row = templ_ws.max_row + 1
                    for (role, amount, resource_name, _, _, _) in report_data[
                        "workers"
                    ]:
                        subheader_cell = templ_ws["B{}".format(current_row)]
                        subheader_cell.value = "Função"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["C{}".format(current_row)]

                        if role and is_valid_uuid(str(role)):
                            final_value = resource_name
                        else:
                            final_value = role if role else resource_name

                        data_cell.value = final_value
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["I{}".format(current_row)]
                        subheader_cell.value = "Quantidade"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["J{}".format(current_row)]
                        data_cell.value = amount
                        apply_data_cell_style(data_cell)

                        templ_ws.merge_cells("C{0}:H{0}".format(current_row))

                        current_row += 1

                    add_vertical_header("EQUIPAMENTOS")
                    current_row = templ_ws.max_row + 1
                    for (
                        description,
                        amount,
                        kind,
                        resource_name,
                        _,
                        _,
                        _,
                    ) in report_data["equipment"]:
                        subheader_cell = templ_ws["B{}".format(current_row)]
                        subheader_cell.value = "Tipo"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["C{}".format(current_row)]
                        data_cell.value = translate_kind(kind, company)
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["D{}".format(current_row)]
                        subheader_cell.value = "Descrição"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["E{}".format(current_row)]

                        if description and is_valid_uuid(str(description)):
                            final_value = resource_name
                        else:
                            final_value = description if description else resource_name

                        data_cell.value = final_value
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["I{}".format(current_row)]
                        subheader_cell.value = "Quantidade"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["J{}".format(current_row)]
                        data_cell.value = amount
                        apply_data_cell_style(data_cell)

                        templ_ws.merge_cells("E{0}:H{0}".format(current_row))

                        current_row += 1

                    add_vertical_header("VEÍCULOS")
                    current_row = templ_ws.max_row + 1
                    for (
                        description,
                        amount,
                        kind,
                        resource_name,
                        _,
                        _,
                        _,
                    ) in report_data["vehicles"]:
                        subheader_cell = templ_ws["B{}".format(current_row)]
                        subheader_cell.value = "Tipo"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["C{}".format(current_row)]
                        data_cell.value = translate_kind(kind, company)
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["D{}".format(current_row)]
                        subheader_cell.value = "Descrição"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["E{}".format(current_row)]

                        if description and is_valid_uuid(str(description)):
                            final_value = resource_name
                        else:
                            final_value = description if description else resource_name

                        data_cell.value = final_value
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["I{}".format(current_row)]
                        subheader_cell.value = "Quantidade"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["J{}".format(current_row)]
                        data_cell.value = amount
                        apply_data_cell_style(data_cell)

                        templ_ws.merge_cells("E{0}:H{0}".format(current_row))

                        current_row += 1

                    add_vertical_header("RECURSOS")
                    current_row = templ_ws.max_row + 1
                    for amount, description, kind in report_data["resources"]:
                        subheader_cell = templ_ws["B{}".format(current_row)]
                        subheader_cell.value = "Tipo"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["C{}".format(current_row)]
                        data_cell.value = translate_kind(kind, company)
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["D{}".format(current_row)]
                        subheader_cell.value = "Descrição"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["E{}".format(current_row)]
                        data_cell.value = description
                        apply_data_cell_style(data_cell)

                        subheader_cell = templ_ws["I{}".format(current_row)]
                        subheader_cell.value = "Quantidade"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["J{}".format(current_row)]
                        data_cell.value = amount
                        apply_data_cell_style(data_cell)

                        templ_ws.merge_cells("E{0}:H{0}".format(current_row))

                        current_row += 1

                    add_vertical_header("SINALIZAÇÃO")
                    current_row = templ_ws.max_row + 1
                    for (kind,) in report_data["signaling"]:
                        subheader_cell = templ_ws["B{}".format(current_row)]
                        subheader_cell.value = "Tipo"
                        apply_subheader_style(subheader_cell)

                        data_cell = templ_ws["C{}".format(current_row)]
                        data_cell.value = kind
                        apply_data_cell_style(data_cell)

                        templ_ws.merge_cells("C{0}:J{0}".format(current_row))

                        current_row += 1

                    occurrence_option_custom_options = get_obj_from_path(
                        company.custom_options,
                        "daily_report_occurrence__fields__origin__select_options__options",
                    )
                    occurrence_option = (
                        occurrence_option_custom_options
                        if occurrence_option_custom_options
                        else TRANSLATE_OCCURRENCE_ORIGIN_CHOICES
                    )

                    for option in occurrence_option:
                        add_vertical_header(
                            "OCORRÊNCIAS DA {}".format(option.get("name", "").upper())
                        )
                        current_row = templ_ws.max_row + 1
                        occurrence_count = 0
                        for (
                            origin,
                            description,
                            impact_duration,
                            starts_at,
                            ends_at,
                            extra_info,
                        ) in report_data["occurrences"]:
                            if origin != option.get("value"):
                                continue

                            occurrence_count += 1
                            cell_value = "{}. {}".format(occurrence_count, description)

                            if starts_at and ends_at and impact_duration:
                                cell_value += " das {} até {} com impacto de {}".format(
                                    starts_at, ends_at, impact_duration
                                )

                            if extra_info:
                                cell_value += " - {}".format(extra_info)

                            data_cell = templ_ws["B{}".format(current_row)]
                            data_cell.value = cell_value
                            apply_data_cell_style(data_cell)

                            templ_ws.merge_cells("B{0}:J{0}".format(current_row))

                            current_row += 1

                    add_vertical_header("APONTAMENTOS")

                    def format_reporting_attributes(
                        attributes: list, reporting_instance: Reporting
                    ) -> list:
                        formated_attributes = get_reporting_data(
                            reporting_instance, exporter_extra_columns, reference_values
                        )
                        parsed_attributes = []
                        for attr in attributes:
                            if attr in extra_fields:
                                exported_value = formated_attributes[
                                    "extra_columns"
                                ].get(attr)
                            else:
                                exported_value = formated_attributes.get(attr)

                            if (
                                attr
                                in [
                                    "km",
                                    "end_km",
                                    "km_reference",
                                    "project_km",
                                    "project_end_km",
                                ]
                                and exported_value
                            ):
                                exported_value = round(exported_value, 3)
                            parsed_attributes.append(exported_value)

                        return parsed_attributes

                    try:
                        mdr = reports.get(uuid=report_uuid)
                        reporting_instances = mdr.reportings.all().prefetch_related(
                            *LightReportingSerializer._PREFETCH_RELATED_FIELDS
                        )
                    except Exception:
                        mdr = None
                        reporting_instances = None
                    if reporting_instances:
                        current_row = templ_ws.max_row + 1
                        hide_reporting_location = company.metadata.get(
                            "hide_reporting_location", False
                        )
                        static_fields = get_reporting_static_columns()
                        extra_fields = get_exporter_extra_columns(company)
                        extra_fields = get_exporter_extra_columns_parsed_infos(
                            extra_fields
                        )
                        if hide_reporting_location is True:
                            fields_to_hide = get_fields_to_hide_reporting_location()
                            static_fields = remove_fields_to_hide_reporting_location(
                                fields_to_hide, static_fields
                            )
                        all_fields = {**static_fields, **extra_fields}
                        reporting_attr_list = get_obj_from_path(
                            company.custom_options,
                            "reporting__exporter__reporting_section_fields_individual_RDO_export",
                        )

                        if not reporting_attr_list:
                            reporting_attr_list = [
                                "occurrence_kind",
                                "occurrence_type",
                                "status",
                            ]

                        columns_to_fill = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
                        cell_color = PatternFill("solid", fgColor="EFEFEF")
                        cell_alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrapText=True,
                            wrap_text=True,
                        )
                        index = 0
                        HEADER_HEIGHT = 50
                        templ_ws.row_dimensions[current_row].height = HEADER_HEIGHT
                        for reporting in reporting_attr_list:
                            translated_name = all_fields.get(reporting)
                            if translated_name:
                                cell = templ_ws[
                                    columns_to_fill[index] + str(current_row)
                                ]
                                cell.value = translated_name
                                cell.fill = cell_color
                                cell.alignment = cell_alignment
                                index += 1

                        current_row += 1
                        for report_instance in reporting_instances:
                            parsed_attributes = format_reporting_attributes(
                                attributes=reporting_attr_list,
                                reporting_instance=report_instance,
                            )
                            ROW_HEIGHT = 50
                            for index, parsed_attribute in enumerate(parsed_attributes):
                                # Use automatic row height for PDF exports to avoid text being cut off
                                templ_ws.row_dimensions[current_row].height = (
                                    None if is_pdf else ROW_HEIGHT
                                )
                                cell = templ_ws[
                                    columns_to_fill[index] + str(current_row)
                                ]
                                if type(parsed_attribute) is datetime:
                                    parsed_attribute = datetime_to_date(
                                        parsed_attribute
                                    )
                                cell.value = parsed_attribute
                                if type(parsed_attribute) is date:
                                    cell.number_format = "dd/mm/yyyy"
                                cell.alignment = cell_alignment
                            current_row += 1
                    add_vertical_header("OBSERVAÇÕES")
                    current_row = templ_ws.max_row + 1

                    def wrap_pixel_words(text, font, max_width):
                        lines = []
                        line = ""
                        words_with_next_spaces = re.findall(r"(\S+)(\s*)", text)
                        for word, spaces in words_with_next_spaces:
                            spaces = spaces.replace("\t", "    ")
                            if font.getlength(line + word + spaces) <= max_width:
                                line += word + spaces
                            else:
                                lines.append(line)
                                line = word + spaces

                        if line:
                            lines.append(line)

                        return lines

                    if is_pdf:
                        notes = report_data.get("notes", "")
                        font = ImageFont.truetype("assets/fonts/Arial.ttf", size=11)
                        max_width = font.getlength("W" * 52)
                        if isinstance(notes, str):
                            notelines = notes.splitlines()
                            for line in notelines:
                                line = line.strip()
                                if line:
                                    sublines = wrap_pixel_words(line, font, max_width)
                                else:
                                    sublines = [line]
                                for subline in sublines:
                                    data_cell = templ_ws["B{}".format(current_row)]
                                    data_cell.value = subline
                                    templ_ws.merge_cells(
                                        "B{0}:J{0}".format(current_row)
                                    )
                                    data_cell.alignment = Alignment(
                                        horizontal="left", vertical="justify"
                                    )
                                    templ_ws.row_dimensions[current_row].height = 15
                                    current_row += 1
                            templ_ws.row_breaks.append(Break(id=current_row))
                    else:
                        data_cell = templ_ws["B{}".format(current_row)]
                        data_cell.value = report_data["notes"]
                        templ_ws.merge_cells("B{0}:J{0}".format(current_row))
                        data_cell.alignment = Alignment(
                            horizontal="left", wrapText=True
                        )

                    if is_pdf:
                        # Remove unwanted sheets for PDF export
                        SHEETS_TO_KEEP = ["RDO Individual", "RDO Compilado"]
                        sheets_to_remove = [
                            sheet_name
                            for sheet_name in templ_wb.sheetnames
                            if sheet_name not in SHEETS_TO_KEEP
                        ]
                        for sheet_name in sheets_to_remove:
                            del templ_wb[sheet_name]
                    else:
                        # Apontamentos e Recursos
                        templ_ws = templ_wb["Apontamentos e Recursos"]
                        data_offset = 3
                        fill_reportings_section(templ_ws, report_data)

                        # Fill "Fotos"
                        removed_images = 0
                        for reporting_data in report_data["reportings"]:
                            if report_data["reporting_files"]:
                                removed_images = fill_photos_section(
                                    templ_wb,
                                    report_data,
                                    reporting_data,
                                    report_iterator,
                                )
                                report_iterator += 1

                        if report_data["multiple_daily_report_files"]:
                            removed_images = fill_rdo_photo_section(
                                templ_wb,
                                report_data,
                                report_iterator,
                            )
                            report_iterator += 1
                        # Add photo section headers
                        add_photo_headers(templ_wb, removed_images)

                    if hide_reporting_location or not report_data["km_intervals"]:
                        templ_ws = templ_wb["RDO Individual"]
                        if is_pdf:
                            templ_ws.row_dimensions[14].hidden = True
                        else:
                            delete_row_with_merged_ranges(templ_ws, 14)

                    if can_view_digital_signature and mdr:
                        templ_ws = templ_wb["RDO Individual"]

                        if not is_pdf:
                            add_vertical_header("ASSINATURA DIGITAL")

                        current_row = (
                            templ_ws.max_row + 40 if is_pdf else templ_ws.max_row + 6
                        )
                        ROW_HEIGHT = 70 if is_pdf else 75
                        PHOTO_COLUMN_WIDTH = 10 if is_pdf else 20
                        text_alignment = Alignment(
                            horizontal="left",
                            vertical="bottom",
                            wrapText=True,
                            wrap_text=True,
                        )
                        signature_fill = PatternFill("solid", fgColor="EBEBEB")
                        signature_font = Font(name="Calibri", size=7, color="000000")
                        subquery = UserInCompany.objects.filter(
                            company=company, user=OuterRef("created_by")
                        )

                        signature_queryset = (
                            mdr.multiple_daily_report_signatures.all()
                            .annotate(
                                permission_name=Subquery(
                                    subquery.values("permissions__name")[:1]
                                )
                            )
                            .prefetch_related("created_by")
                            .only(
                                "signature_name",
                                "signature_date",
                                "upload",
                                "created_by",
                            )
                        )
                        for index, signature in enumerate(signature_queryset):
                            # Add signature image
                            file_format = signature.upload.name.split(".")[-1]
                            if file_format.lower() not in ["jpg", "jpeg", "png"]:
                                continue

                            templ_ws.row_dimensions[current_row - 1].height = ROW_HEIGHT

                            # Add signature image
                            params = {}
                            params["Bucket"] = "{}-{}px".format(
                                signature.upload.storage.bucket.name, 1000
                            )
                            params["Key"] = signature.upload.storage._normalize_name(
                                clean_name(signature.upload.name)
                            )

                            upload_url = signature.upload.storage.bucket.meta.client.generate_presigned_url(
                                "get_object", Params=params, ExpiresIn=3600
                            )
                            resp = requests.get(upload_url, stream=True)
                            image_content = (
                                resp.content if resp.status_code == 200 else None
                            )

                            first_column = 2 * index + 2
                            second_column = 2 * index + 3
                            column_letter = get_column_letter(first_column)

                            if not is_pdf:
                                title_cell = templ_ws[
                                    "{}{}".format(column_letter, current_row - 3)
                                ]
                                title_cell.value = "Assinado eletronicamente por:"
                                title_cell.alignment = text_alignment
                                title_cell.fill = signature_fill
                                title_cell.font = signature_font
                            if is_pdf:
                                templ_ws.column_dimensions[
                                    column_letter
                                ].width = PHOTO_COLUMN_WIDTH

                            if image_content:
                                templ_ws.column_dimensions[
                                    column_letter
                                ].width = PHOTO_COLUMN_WIDTH
                                temp_file = NamedTemporaryFile(delete=False)
                                temp_file.write(image_content)
                                image = Image(temp_file.name)
                                temp_file.close()

                                from_anchor_marker = AnchorMarker(
                                    col=first_column - 1,
                                    row=current_row - 2,
                                )
                                to_anchor_marker = AnchorMarker(
                                    col=second_column - 1, row=current_row - 1
                                )
                                anchor = TwoCellAnchor(
                                    _from=from_anchor_marker, to=to_anchor_marker
                                )
                                templ_ws.add_image(image, anchor)
                                files_to_close.append(temp_file)

                            # Add signature data

                            signature_cell = templ_ws[
                                "{}{}".format(column_letter, current_row)
                            ]
                            signature_cell.value = "Nome: " + signature.signature_name
                            signature_cell.alignment = text_alignment
                            signature_cell.fill = signature_fill
                            signature_cell.font = signature_font

                            signature_date = templ_ws[
                                "{}{}".format(column_letter, current_row + 1)
                            ]
                            signature_date.value = "Data/hora: " + utc_to_local(
                                signature.signature_date
                            ).strftime("%d/%m/%Y, %H:%M")
                            signature_date.alignment = text_alignment
                            signature_date.fill = signature_fill
                            signature_date.font = signature_font

                            login_cell = templ_ws[
                                "{}{}".format(column_letter, current_row + 2)
                            ]
                            login_cell.value = "Login: " + signature.created_by.username
                            login_cell.alignment = text_alignment
                            login_cell.fill = signature_fill
                            login_cell.font = signature_font

                            permission_cell = templ_ws[
                                "{}{}".format(column_letter, current_row + 3)
                            ]
                            permission_cell.value = (
                                "Permissão: " + signature.permission_name
                            )
                            permission_cell.alignment = text_alignment
                            permission_cell.fill = signature_fill
                            permission_cell.font = signature_font

                        if is_pdf:
                            for row in range(current_row + 500, current_row + 4, -1):
                                templ_ws.delete_rows(idx=row, amount=1)
                    # Save file to temp folder
                    if is_pdf:
                        file_name = (
                            f"{temp_dir}{report_data['report_order']:04d}{file_ext}"
                        )
                    else:
                        file_name = f"{temp_dir}{report_data['number']}{file_ext}"

                    templ_wb.save(file_name)

                num_reports = len(reports_data.values())
                if is_pdf:
                    GOTENBERG_URL = f"{GOTENBERG_BASE_URL}/forms/libreoffice/convert"

                    # Convert the files to PDF
                    filenames = os.listdir(temp_dir)
                    filenames.sort()
                    files = [
                        ("files", (open(f"{temp_dir}/{file_name}", "rb")))
                        for file_name in filenames
                    ]
                    data = {"merge": "true"}
                    response = requests.post(GOTENBERG_URL, data=data, files=files)

                    if response.status_code == status.HTTP_200_OK:
                        exported_pdf_file_name = (
                            "rdos_individuais" if num_reports > 1 else "rdo_individual"
                        )
                        daily_report_export.exported_file.save(
                            f"{exported_pdf_file_name}.pdf",
                            ContentFile(response.content),
                        )
                        error = False
                    else:
                        logging.error(
                            "generate_exported_file: Error while trying to convert excel file to PDF with Gotenberg"
                        )
                    for file_name in os.listdir(temp_dir):
                        os.remove(temp_dir + file_name)
                    os.rmdir(temp_dir)
                else:
                    # Save final ZIP file
                    if num_reports > 1:
                        zip_file_name = "rdos_individuais.zip"
                        zip_base_dir = "/tmp/"
                        shutil.make_archive(
                            zip_base_dir + "rdos_individuais", "zip", temp_dir
                        )
                        with open(zip_base_dir + zip_file_name, "rb") as zip_file:
                            daily_report_export.exported_file.save(
                                zip_file_name, ContentFile(zip_file.read())
                            )
                            error = False

                        # Delete temp files
                        for file_name in os.listdir(temp_dir):
                            os.remove(temp_dir + file_name)
                        os.remove(zip_base_dir + zip_file_name)
                        os.rmdir(temp_dir)
                    elif num_reports == 1:
                        filename = os.listdir(temp_dir)[0]
                        with open(
                            "{}{}".format(temp_dir, filename), "rb"
                        ) as exported_file:
                            final_file_name = "rdo_individual.{}".format(
                                "pdf" if is_pdf else "xlsm"
                            )
                            daily_report_export.exported_file.save(
                                final_file_name, ContentFile(exported_file.read())
                            )
                            error = False

                        # Delete temp files
                        os.remove("{}{}".format(temp_dir, filename))
                        os.rmdir(temp_dir)

            # Close open image files
            for open_file in files_to_close:
                if not open_file.closed:
                    open_file.close()
                if os.path.exists(open_file.name):
                    os.remove(open_file.name)
        except Exception as e:
            logging.error(
                "Untreated exception found while exporting file. Check Sentry."
            )
            logging.error(e)
            sentry_sdk.capture_exception(e)
            error = True

        # Finish and set flags
        daily_report_export.error = error
        daily_report_export.done = True
        daily_report_export.save()


class ActiveFieldModelSerializer(serializers.ModelSerializer):
    """
    A ModelSerializer that detects if the user is using the daily_reports
    filter or not to determine if the fields should be shown.

    Override pop_fields() on child serializer if new fields are added.
    """

    active = serializers.SerializerMethodField()

    def get_report_id_and_field(self):
        """
        Returns the id of the report and the field that should be used in the
        DailyReportRelation. If no supported query_param is found returns None.
        """
        if "request" in self.context:
            query_params = self.context["request"].query_params
            if "daily_reports" in query_params:
                return query_params["daily_reports"], "daily_report"
            elif "multiple_daily_reports" in query_params:
                return (query_params["multiple_daily_reports"], "multiple_daily_report")

        # No request in context or no supported filter found
        return None

    def get_firm_id(self):
        if "request" in self.context:
            query_params = self.context["request"].query_params
            firm_id = query_params.get("firm", None)
            try:
                if firm_id and not Firm.objects.filter(uuid=firm_id).exists():
                    raise serializers.ValidationError(
                        "Faça uma escolha válida. {} não é uma das escolhas disponíveis.".format(
                            firm_id
                        )
                    )
            except ValidationError:
                raise serializers.ValidationError(
                    "'{}' não é um UUID válido.".format(firm_id)
                )

            return firm_id

        return None

    def pop_fields(self):
        """
        Pops the fields that are not being used
        """
        self.fields.pop("active")

    def __init__(self, *args, **kwargs):
        # Instanciate the class normally
        super(ActiveFieldModelSerializer, self).__init__(*args, **kwargs)

        if self.get_report_id_and_field() is None:
            self.pop_fields()

    def get_active(self, obj):
        obj_id = obj.uuid
        report_info = self.get_report_id_and_field()
        if report_info is not None:
            report_id, report_field = report_info
            target_field = determine_relation_field_name(type(obj))
            kwargs = {report_field: report_id, target_field: obj_id}
            relation = DailyReportRelation.objects.get(**kwargs)

            return relation.active
        else:
            return None


def has_permission(model_permissions, permission):
    try:
        return True if model_permissions[permission][0] is True else False
    except Exception:
        return False


def get_reporting_static_columns():
    return {
        "number": "Serial Apontamento",
        "road": "Rodovia",
        "km": "KM inicial",
        "end_km": "KM final",
        "lot": "Lote",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "occurrence_kind": "Natureza",
        "occurrence_type": "Classe",
        "length": "Comprimento",
        "width": "Largura",
        "height": "Espessura",
        "lane": "Faixa",
        "direction": "Sentido",
        "status": "Status",
        "created_by": "Criado por",
        "updated_by": "Atualizado por",
        "firm": "Equipe",
        "job": "Programação",
        "job_start_date": "Início da Programação",
        "job_end_date": "Fim da Programação",
        "created_at": "Criado em",
        "found_at": "Encontrado em",
        "updated_at": "Atualizado em",
        "executed_at": "Executado em",
        "notes": "Observações",
        "due_at": "Prazo",
    }


def get_fields_to_hide_reporting_location():
    return [
        "road",
        "km",
        "end_km",
        "lot",
        "lane",
        "direction",
        "track",
        "branch",
        "km_reference",
    ]


def get_exporter_extra_columns(company, is_inventory=False):
    return get_obj_from_path(
        company.custom_options,
        "{}__exporter__extra_columns".format(
            "inventory" if is_inventory else "reporting"
        ),
    )


def remove_fields_to_hide_reporting_location(
    fields_to_hide_reporting_location, static_columns
):
    for field in fields_to_hide_reporting_location:
        static_columns.pop(field, None)
    return static_columns


def create_array_columns(extra_column, reporting_export, extra_info):
    header = extra_column.get("header", "")
    key = extra_column.get("key", "")
    max_repetitions = extra_column.get("maxRepetitions", 5)
    fields = extra_column.get("fields", [])
    array_result = {}
    for i in range(0, max_repetitions):
        for item in fields:
            inner_header = item.get("header", "")
            inner_key = item.get("field", "")
            is_image = item.get("isImage", False)
            if is_image:
                export_photos = extra_info.get("export_photos", False)
                number_of_images = item.get("numberOfImages", 5)
                if reporting_export and export_photos:
                    image_columns = get_array_image_columns(
                        key,
                        header,
                        inner_key,
                        inner_header,
                        i,
                        number_of_images,
                        extra_info,
                    )
                    array_result.update(image_columns)
                else:
                    continue
            else:
                array_result.update(
                    {
                        f"{key}{str(i)}{inner_key}": f"{header} {str(i+1)}: {inner_header}"
                    }
                )
    return array_result


def get_array_image_columns(
    key, header, inner_key, inner_header, position, number_of_images, extra_info
):
    image_columns = {}
    export_kind = extra_info.get("export_kind", False)
    export_date = extra_info.get("export_date", False)
    export_description = extra_info.get("export_description", False)
    for i in range(number_of_images):
        image_columns.update(
            {
                f"{key}#{str(position)}#{inner_key}{str(i)}#content": f"{header} {str(position+1)}: {inner_header} {str(i+1)}"
            }
        )
        if export_kind:
            image_columns.update(
                {
                    f"{key}#{str(position)}#{inner_key}{str(i)}#kind": f"{header} {str(position+1)}: {inner_header} {str(i+1)} - Tipo"
                }
            )
        if export_date:
            image_columns.update(
                {
                    f"{key}#{str(position)}#{inner_key}{str(i)}#datetime": f"{header} {str(position+1)}: {inner_header} {str(i+1)} - Data"
                }
            )
        if export_description:
            image_columns.update(
                {
                    f"{key}#{str(position)}#{inner_key}{str(i)}#description": f"{header} {str(position+1)}: {inner_header} {str(i+1)} - Descrição"
                }
            )

    return image_columns


def get_exporter_extra_columns_parsed_infos(
    exporter_extra_columns,
    skip_array_fields=False,
    reporting_export=False,
    extra_info={},
):
    extra_columns = {}
    for extra_column in exporter_extra_columns:
        is_array = extra_column.get("isArray", False)
        if not is_array:
            column_key = extra_column.get("key", "")
            extra_columns[column_key] = extra_column.get("header", "")
        elif is_array and not skip_array_fields:
            array_columns = create_array_columns(
                extra_column, reporting_export, extra_info
            )
            extra_columns.update(array_columns)
        else:
            pass
    return extra_columns


def get_updated_by(instance):
    history = instance.historicalreporting.all()

    try:
        return history[0].history_user.get_full_name()
    except Exception:
        return None


def translate_reporting_value(company, field, value):
    options = get_obj_from_path(
        company.custom_options,
        "reporting__fields__{}__selectoptions__options".format(field),
    )
    try:
        return next(a["name"] for a in options if a["value"] == value)
    except Exception:
        return ""


def datetime_to_date(dt, clear_tzinfo=True):
    try:
        if clear_tzinfo:
            dt = utc_to_local(dt).replace(tzinfo=None)
        dt_date = dt.date()
    except Exception:
        dt_date = None
    return dt_date


def get_reporting_data(reporting, exporter_extra_columns, reference_values):
    reporting_data = {
        "number": reporting.number,
        "road": reporting.road.name if reporting.road else None,
        "km": reporting.km,
        "end_km": reporting.end_km if reporting.end_km else None,
        "lot": (
            translate_reporting_value(reporting.company, "lot", reporting.lot)
            if reporting.lot
            else None
        ),
        "longitude": reporting.point.coords[0] if reporting.point else None,
        "latitude": reporting.point.coords[1] if reporting.point else None,
        "occurrence_type": (
            reporting.occurrence_type.name if reporting.occurrence_type else None
        ),
        "length": reporting.form_data.get("length"),
        "width": reporting.form_data.get("width"),
        "height": reporting.form_data.get("height"),
        "lane": translate_reporting_value(reporting.company, "lane", reporting.lane),
        "direction": translate_reporting_value(
            reporting.company, "direction", reporting.direction
        ),
        "status": reporting.status.name if reporting.status else None,
        "created_by": (
            reporting.created_by.get_full_name() if reporting.created_by else None
        ),
        "updated_by": get_updated_by(reporting),
        "firm": reporting.firm.name if reporting.firm else None,
        "job": reporting.job.title if reporting.job else None,
        "job_start_date": (
            datetime_to_date(reporting.job.start_date) if reporting.job else None
        ),
        "job_end_date": (
            datetime_to_date(reporting.job.end_date)
            if reporting.job and reporting.job.end_date
            else None
        ),
        "created_at": datetime_to_date(reporting.created_at),
        "found_at": datetime_to_date(reporting.found_at),
        "updated_at": datetime_to_date(reporting.updated_at),
        "executed_at": (
            datetime_to_date(reporting.executed_at) if reporting.executed_at else None
        ),
        "notes": reporting.form_data.get("notes"),
        "due_at": datetime_to_date(reporting.due_at) if reporting.due_at else None,
        "extra_columns": get_values_from_reporting_extra_columns(
            reporting, exporter_extra_columns, reference_values
        ),
    }

    # Find OccurrenceKind name
    possible_path = "reporting__fields__occurrence_kind__selectoptions__options"
    try:
        reporting_data["occurrence_kind"] = next(
            a
            for a in get_obj_from_path(reporting.company.custom_options, possible_path)
            if a["value"] == reporting.occurrence_type.occurrence_kind
        )["name"]
    except BaseException:
        reporting_data["occurrence_kind"] = None

    return reporting_data


def parse_time_string(time_value):
    """
    Parse time value from string format "hh:mm" to time object.

    Args:
        time_value: Time value as string ("hh:mm"), time object, or None

    Returns:
        time object or None if parsing fails or value is empty
    """
    if time_value is None or time_value == "":
        return None

    # If already a time object, return as is
    if isinstance(time_value, time):
        return time_value

    # If it's a string, try to parse it
    if isinstance(time_value, str):
        try:
            # Parse string in format "hh:mm"
            time_parts = time_value.split(":")
            if len(time_parts) == 2:
                hours = int(time_parts[0])
                minutes = int(time_parts[1])
                return time(hours, minutes)
        except (ValueError, IndexError):
            pass

    return None


def normalize_time_fields(hours_dict, default_hours):
    """
    Normalize time field names from camelCase to snake_case and apply field-level fallback.
    Also parse string time values to time objects.

    Args:
        hours_dict: Dictionary with time fields (may use camelCase or snake_case)
        default_hours: Dictionary with default values for fallback

    Returns:
        Dictionary with snake_case field names, individual field fallback applied,
        and time values converted to time objects
    """
    # Mapping from camelCase to snake_case
    hours_field_mapping = {
        "morningStart": "morning_start",
        "morningEnd": "morning_end",
        "afternoonStart": "afternoon_start",
        "afternoonEnd": "afternoon_end",
        "nightStart": "night_start",
        "nightEnd": "night_end",
    }

    original_and_deleted_mapping = {
        "morning_start": "morning_start_is_deleted",
        "morning_end": "morning_end_is_deleted",
        "afternoon_start": "afternoon_start_is_deleted",
        "afternoon_end": "afternoon_end_is_deleted",
        "night_start": "night_start_is_deleted",
        "night_end": "night_end_is_deleted",
    }

    hours_value_keys = list(original_and_deleted_mapping.keys()) + ["description"]

    # Start with default hours, ensuring they are time objects
    normalized_hours = {}
    for key, value in default_hours.items():
        normalized_hours[key] = parse_time_string(value)

    # Override with values from hours_dict, handling both camelCase and snake_case
    for original_key, value in hours_dict.items():
        # Convert camelCase to snake_case if needed
        snake_case_key = hours_field_mapping.get(original_key, original_key)

        if snake_case_key not in hours_value_keys:
            continue

        if snake_case_key == "description":
            normalized_hours[snake_case_key] = value
            continue

        # Parse the time value and only update if it's valid
        parsed_time = parse_time_string(value)

        # Get if it's a deleted case
        deleted_key = original_and_deleted_mapping.get(snake_case_key)
        deleted_value = hours_dict.get(deleted_key, False)

        if parsed_time is not None:
            normalized_hours[snake_case_key] = parsed_time
        if deleted_value and snake_case_key in normalized_hours:
            del normalized_hours[snake_case_key]

    return normalized_hours


def parse_extra_hours_to_list(extra_hours, default_hours=None):
    """
    Parse extra_hours from different formats to a list.

    Handles:
    - dict with 'extraHours' key containing a list
    - dict with numeric string keys (e.g., {"0": {...}, "1": {...}})
    - list format

    Args:
        extra_hours: The extra hours data in various formats
        default_hours: Default hours dictionary for field-level fallback

    Returns a list of hour dictionaries with snake_case field names and individual fallback.
    """
    extra_hours_list = []

    if extra_hours and isinstance(extra_hours, dict):
        # If extra_hours is a dict with extraHours key
        if "extraHours" in extra_hours:
            extra_hours_list = extra_hours.get("extraHours", [])
        else:
            # Check if it's a dict with numeric string keys
            numeric_keys = [key for key in extra_hours.keys() if key.isdigit()]
            if numeric_keys:
                # Sort by numeric value and convert to list
                sorted_keys = sorted(numeric_keys, key=int)
                extra_hours_list = [extra_hours[key] for key in sorted_keys]
    elif extra_hours and isinstance(extra_hours, list):
        # If extra_hours is already a list
        extra_hours_list = extra_hours

    # Apply normalization and field-level fallback to each item
    if default_hours:
        extra_hours_list = [
            normalize_time_fields(hours_dict, default_hours)
            for hours_dict in extra_hours_list
        ]

    return extra_hours_list


def fill_detailed_control_board(templ_wb, reports_data, company, daily_report_export):
    """
    Creates a new worksheet with detailed control board information, including time periods
    and multiple rows per item based on quantity and extra hours.
    """

    try:
        permissions = daily_report_export.created_by.companies_membership.get(
            company=company
        ).permissions
        extra_hours_permissions = get_obj_from_path(
            permissions.permissions, "multipledailyreport__cansetextrahours"
        )

        if not extra_hours_permissions:
            return

        allowed_contracts = get_obj_from_path(
            company.custom_options,
            "multipledailyreport__extrahourscontractids",
        )

        export_contracts = list(
            set(
                [
                    str(a)
                    for a in daily_report_export.multiple_daily_reports.values_list(
                        "contract_id", flat=True
                    )
                ]
            )
        )

        if not any(contract in allowed_contracts for contract in export_contracts):
            return

    except Exception:
        return

    # Pre-check: collect unique (contract_id, firm_id) pairs from all reports
    pairs = set()
    for report in reports_data.values():
        contract_id = report.get("contract_id")
        firm_id = report.get("firm_id")
        if contract_id and firm_id:
            pairs.add((contract_id, firm_id))

    # Build ContractPeriod cache: {(contract_id, firm_id): working_schedules}
    contract_period_cache = {}
    for contract_id, firm_id in pairs:
        cp = (
            ContractPeriod.objects.filter(
                contract_id=contract_id,
                firms__uuid=firm_id,
            )
            .order_by("-created_at")
            .first()
        )
        if cp:
            contract_period_cache[(contract_id, firm_id)] = cp.working_schedules

    show_extra_hours_columns = bool(contract_period_cache)

    # Create new worksheet
    templ_ws = templ_wb.create_sheet("Quadros de Controle Detalhado")

    # Define headers
    extra_hours_headers = (
        [
            "Hora Extra 50% Dia",
            "Hora Extra 50% Noite",
            "Hora Extra 100% Dia",
            "Hora Extra 100% Noite",
            "Horas Adicionais",
            "Falta",
            "Horas Compensadas",
        ]
        if show_extra_hours_columns
        else []
    )
    headers = [
        "Serial do RDO",
        "Nº contrato",
        "Equipe",
        "Data",
        "Dia da Semana",
        "Feriado",
        "Compensação",
        "Código",
        "Recurso",
        "Nome do Funcionário",
        "Entrada (manhã)",
        "Saída (manhã)",
        "Entrada (tarde)",
        "Saída (tarde)",
        "Entrada (noite)",
        "Saída (noite)",
        "Total Diurno",
        "Total Diurno (decimal)",
        "Total Noturno",
        "Total Noturno (decimal)",
        "Total de Horas",
        "Total de Horas (decimal)",
        *extra_hours_headers,
        "Observações",
    ]

    # Define column widths
    extra_hours_widths = (
        [
            23,  # Hora Extra 50% Dia
            23,  # Hora Extra 50% Noite
            23,  # Hora Extra 100% Dia
            23,  # Hora Extra 100% Noite
            21,  # Horas Adicionais
            12,  # Falta
            21,  # Horas Compensadas
        ]
        if show_extra_hours_columns
        else []
    )
    column_widths = [
        15,  # Serial do RDO
        15,  # Nº contrato
        20,  # Equipe
        12,  # Data
        15,  # Dia da Semana
        15,  # Feriado
        15,  # Compensação
        12,  # Código
        30,  # Recurso
        30,  # Nome do Funcionário
        15,  # Entrada (manhã)
        15,  # Saída (manhã)
        15,  # Entrada (tarde)
        15,  # Saída (tarde)
        15,  # Entrada (noite)
        15,  # Saída (noite)
        15,  # Total Diurno
        20,  # Total Diurno (decimal)
        15,  # Total Noturno
        20,  # Total Noturno (decimal)
        15,  # Total de Horas
        20,  # Total de Horas (decimal)
        *extra_hours_widths,
        40,  # Observações
    ]

    night_total_text = "Considera-se noturno o trabalho executado entre as 22 horas de um dia e as 5 horas do dia seguinte, utilizando como base o DEL9666"

    # Set column widths
    for col, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col)
        templ_ws.column_dimensions[column_letter].width = width

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = templ_ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if header == "Total Noturno":
            cell.comment = Comment(night_total_text, "")

    # Track current row
    current_row = 2

    def calculate_total_hours(hours):
        """Calculate total hours worked in a day"""
        total_minutes = 0
        day_total = 0
        night_total = 0

        # Morning period
        if hours.get("morning_start") and hours.get("morning_end"):
            start = hours["morning_start"]
            end = hours["morning_end"]
            if isinstance(start, time) and isinstance(end, time):
                morning_minutes = (end.hour * 60 + end.minute) - (
                    start.hour * 60 + start.minute
                )
                if morning_minutes > 0:
                    total_minutes += morning_minutes
                day_mins, night_mins = minutes_between(start, end)
                day_total += day_mins
                night_total += night_mins

        # Afternoon period
        if hours.get("afternoon_start") and hours.get("afternoon_end"):
            start = hours["afternoon_start"]
            end = hours["afternoon_end"]
            if isinstance(start, time) and isinstance(end, time):
                afternoon_minutes = (end.hour * 60 + end.minute) - (
                    start.hour * 60 + start.minute
                )
                if afternoon_minutes > 0:
                    total_minutes += afternoon_minutes
                day_mins, night_mins = minutes_between(start, end)
                day_total += day_mins
                night_total += night_mins

        # Night period
        if hours.get("night_start") and hours.get("night_end"):
            start = hours["night_start"]
            end = hours["night_end"]
            if isinstance(start, time) and isinstance(end, time):
                start_minutes = start.hour * 60 + start.minute
                end_minutes = end.hour * 60 + end.minute

                if end_minutes >= start_minutes:
                    night_minutes = end_minutes - start_minutes
                else:
                    night_minutes = (24 * 60 - start_minutes) + end_minutes

                total_minutes += night_minutes
                day_mins, night_mins = minutes_between(start, end)
                day_total += day_mins
                night_total += night_mins

        if total_minutes > 0:
            return (
                format_minutes(total_minutes),
                format_minutes(day_total),
                format_minutes(night_total),
                format_minutes_decimal(total_minutes),
                format_minutes_decimal(day_total),
                format_minutes_decimal(night_total),
            )
        return "-", "-", "-", "-", "-", "-"

    def format_value(value):
        """Format value, replacing empty values with dash"""
        if value is None or value == "":
            return "-"
        # Format time objects as HH:MM strings
        if isinstance(value, time):
            return value.strftime("%H:%M")
        return value

    data_len = 30 if show_extra_hours_columns else 23

    # Process each report
    for report_data in reports_data.values():

        if report_data.get("day_without_work") == "Não":
            continue

        # Get contract number
        contract_number = report_data.get("contract", "-")
        mdr_contract_id = report_data.get("contract_id")

        # Get notes from the report
        notes = report_data.get("notes", "-")
        if not notes or notes.strip() == "":
            notes = "-"

        # Get day of week
        mdr_date = report_data.get("date")
        if mdr_date:
            day_of_week = mdr_date.strftime("%A")

            # Translate to Portuguese
            day_of_week = DAYS_PORTUGUESE.get(day_of_week, day_of_week)

            # Check if it's a holiday (national or custom)
            firm_id = report_data.get("firm_id")
            holiday = (
                "SIM" if is_holiday_for_firm(company, firm_id, mdr_date) else "NÃO"
            )
        else:
            day_of_week = "-"
            holiday = "-"

        compensation = report_data.get("compensation", "-")

        # Process workers
        for (
            role,
            amount,
            resource_name,
            sort_string,
            extra_hours,
            contract_id,
        ) in report_data["workers"]:

            if str(contract_id) != mdr_contract_id:
                continue

            if role and is_valid_uuid(str(role)):
                role_name = resource_name
            else:
                role_name = role if role else resource_name

            if not role_name:
                continue

            # Default hours from report data
            default_hours = {
                "morning_start": report_data.get("morning_start"),
                "morning_end": report_data.get("morning_end"),
                "afternoon_start": report_data.get("afternoon_start"),
                "afternoon_end": report_data.get("afternoon_end"),
                "night_start": report_data.get("night_start"),
                "night_end": report_data.get("night_end"),
            }

            # Parse extra_hours from the JSONField
            extra_hours_list = parse_extra_hours_to_list(extra_hours, default_hours)

            # If no extra hours, use default times (normalized)
            if not extra_hours_list:
                extra_hours_list = [normalize_time_fields({}, default_hours)]

            amount = int(amount)

            # Create rows based on amount and extra hours
            for i in range(amount):
                hours = (
                    extra_hours_list[i]
                    if i < len(extra_hours_list)
                    else extra_hours_list[-1]
                )

                # Get contract item code (sort_string) - now directly available
                contract_code = sort_string if sort_string else "-"

                (
                    total_hours,
                    day_hours,
                    night_hours,
                    total_hours_decimal,
                    day_hours_decimal,
                    night_hours_decimal,
                ) = calculate_total_hours(hours)

                if show_extra_hours_columns:
                    working_schedules = contract_period_cache.get(
                        (mdr_contract_id, firm_id), None
                    )
                    if working_schedules is not None:
                        extra = calculate_extra_hours_worker(
                            worked_periods_item=dict_to_casing(
                                hours, format_type="camelize"
                            ),
                            working_schedules=working_schedules,
                            day_of_week=mdr_date.isoweekday(),
                            is_holiday=holiday == "SIM",
                            is_compensation=compensation == "SIM",
                        )
                        worker_extra_cols = _worker_result_to_decimal_cols(
                            extra, is_worker=True
                        )
                    else:
                        worker_extra_cols = [0] * 7
                else:
                    worker_extra_cols = []

                row_data = [
                    format_value(report_data["number"]),
                    format_value(contract_number),
                    format_value(report_data["firm"]),
                    mdr_date,
                    day_of_week,
                    holiday,
                    compensation,
                    format_value(contract_code),
                    format_value(role_name),
                    format_value(hours.get("description", None)),
                    format_value(hours.get("morning_start")),
                    format_value(hours.get("morning_end")),
                    format_value(hours.get("afternoon_start")),
                    format_value(hours.get("afternoon_end")),
                    format_value(hours.get("night_start")),
                    format_value(hours.get("night_end")),
                    day_hours,
                    day_hours_decimal,
                    night_hours,
                    night_hours_decimal,
                    total_hours,
                    total_hours_decimal,
                    *worker_extra_cols,
                    format_value(notes),
                ]

                # Add row
                for col, value in enumerate(row_data, 1):
                    cell = templ_ws.cell(row=current_row, column=col)
                    cell.value = value
                    if isinstance(value, date):
                        cell.number_format = "dd/mm/yyyy"
                    elif isinstance(value, time):
                        cell.number_format = "hh:mm:ss"
                    # Set text wrapping for the Notes column
                    if col == data_len:  # Last column (Notes)
                        cell.alignment = Alignment(wrapText=True, vertical="top")

                current_row += 1

        # Process equipment
        for (
            description,
            amount,
            _,
            resource_name,
            sort_string,
            extra_hours,
            contract_id,
        ) in report_data["equipment"]:

            if str(contract_id) != mdr_contract_id:
                continue

            if description and is_valid_uuid(str(description)):
                equip_name = resource_name
            else:
                equip_name = description if description else resource_name

            if not equip_name:
                continue

            # Default hours from report data
            default_hours = {
                "morning_start": report_data.get("morning_start"),
                "morning_end": report_data.get("morning_end"),
                "afternoon_start": report_data.get("afternoon_start"),
                "afternoon_end": report_data.get("afternoon_end"),
                "night_start": report_data.get("night_start"),
                "night_end": report_data.get("night_end"),
            }

            # Parse extra_hours from the JSONField
            extra_hours_list = parse_extra_hours_to_list(extra_hours, default_hours)

            # If no extra hours, use default times (normalized)
            if not extra_hours_list:
                extra_hours_list = [normalize_time_fields({}, default_hours)]

            amount = int(amount)

            # Create rows based on amount and extra hours
            for i in range(amount):
                hours = (
                    extra_hours_list[i]
                    if i < len(extra_hours_list)
                    else extra_hours_list[-1]
                )

                # Get contract item code (sort_string) - now directly available
                contract_code = sort_string if sort_string else "-"

                (
                    total_hours,
                    day_hours,
                    night_hours,
                    total_hours_decimal,
                    day_hours_decimal,
                    night_hours_decimal,
                ) = calculate_total_hours(hours)

                if show_extra_hours_columns:
                    working_schedules = contract_period_cache.get(
                        (mdr_contract_id, firm_id), None
                    )
                    if working_schedules is not None:
                        extra = calculate_extra_hours_worker(
                            worked_periods_item=dict_to_casing(
                                hours, format_type="camelize"
                            ),
                            working_schedules=working_schedules,
                            day_of_week=mdr_date.isoweekday(),
                            is_holiday=holiday == "SIM",
                            is_compensation=compensation == "SIM",
                        )
                        equip_extra_cols = _worker_result_to_decimal_cols(
                            extra, is_worker=False
                        )
                    else:
                        equip_extra_cols = [0] * 7
                else:
                    equip_extra_cols = []

                row_data = [
                    format_value(report_data["number"]),
                    format_value(contract_number),
                    format_value(report_data["firm"]),
                    mdr_date,
                    day_of_week,
                    holiday,
                    compensation,
                    format_value(contract_code),
                    format_value(equip_name),
                    "-",
                    format_value(hours.get("morning_start")),
                    format_value(hours.get("morning_end")),
                    format_value(hours.get("afternoon_start")),
                    format_value(hours.get("afternoon_end")),
                    format_value(hours.get("night_start")),
                    format_value(hours.get("night_end")),
                    day_hours,
                    day_hours_decimal,
                    night_hours,
                    night_hours_decimal,
                    total_hours,
                    total_hours_decimal,
                    *equip_extra_cols,
                    format_value(notes),
                ]

                # Add row
                for col, value in enumerate(row_data, 1):
                    cell = templ_ws.cell(row=current_row, column=col)
                    cell.value = value
                    if isinstance(value, date):
                        cell.number_format = "dd/mm/yyyy"
                    elif isinstance(value, time):
                        cell.number_format = "hh:mm:ss"
                    # Set text wrapping for the Notes column
                    if col == data_len:  # Last column (Notes)
                        cell.alignment = Alignment(wrapText=True, vertical="top")

                current_row += 1

        # Process vehicles
        for (
            description,
            amount,
            _,
            resource_name,
            sort_string,
            extra_hours,
            contract_id,
        ) in report_data["vehicles"]:

            if str(contract_id) != mdr_contract_id:
                continue

            if description and is_valid_uuid(str(description)):
                vehicle_name = resource_name
            else:
                vehicle_name = description if description else resource_name

            if not vehicle_name:
                continue

            # Default hours from report data
            default_hours = {
                "morning_start": report_data.get("morning_start"),
                "morning_end": report_data.get("morning_end"),
                "afternoon_start": report_data.get("afternoon_start"),
                "afternoon_end": report_data.get("afternoon_end"),
                "night_start": report_data.get("night_start"),
                "night_end": report_data.get("night_end"),
            }

            # Parse extra_hours from the JSONField
            extra_hours_list = parse_extra_hours_to_list(extra_hours, default_hours)

            # If no extra hours, use default times (normalized)
            if not extra_hours_list:
                extra_hours_list = [normalize_time_fields({}, default_hours)]

            amount = int(amount)

            # Create rows based on amount and extra hours
            for i in range(amount):
                hours = (
                    extra_hours_list[i]
                    if i < len(extra_hours_list)
                    else extra_hours_list[-1]
                )

                # Get contract item code (sort_string) - now directly available
                contract_code = sort_string if sort_string else "-"

                (
                    total_hours,
                    day_hours,
                    night_hours,
                    total_hours_decimal,
                    day_hours_decimal,
                    night_hours_decimal,
                ) = calculate_total_hours(hours)

                if show_extra_hours_columns:
                    working_schedules = contract_period_cache.get(
                        (mdr_contract_id, firm_id), None
                    )
                    if working_schedules is not None:
                        extra = calculate_extra_hours_worker(
                            worked_periods_item=dict_to_casing(
                                hours, format_type="camelize"
                            ),
                            working_schedules=working_schedules,
                            day_of_week=mdr_date.isoweekday(),
                            is_holiday=holiday == "SIM",
                            is_compensation=compensation == "SIM",
                        )
                        vehicle_extra_cols = _worker_result_to_decimal_cols(
                            extra, is_worker=False
                        )
                    else:
                        vehicle_extra_cols = [0] * 7
                else:
                    vehicle_extra_cols = []

                row_data = [
                    format_value(report_data["number"]),
                    format_value(contract_number),
                    format_value(report_data["firm"]),
                    mdr_date,
                    day_of_week,
                    holiday,
                    compensation,
                    format_value(contract_code),
                    format_value(vehicle_name),
                    "-",
                    format_value(hours.get("morning_start")),
                    format_value(hours.get("morning_end")),
                    format_value(hours.get("afternoon_start")),
                    format_value(hours.get("afternoon_end")),
                    format_value(hours.get("night_start")),
                    format_value(hours.get("night_end")),
                    day_hours,
                    day_hours_decimal,
                    night_hours,
                    night_hours_decimal,
                    total_hours,
                    total_hours_decimal,
                    *vehicle_extra_cols,
                    format_value(notes),
                ]

                # Add row
                for col, value in enumerate(row_data, 1):
                    cell = templ_ws.cell(row=current_row, column=col)
                    cell.value = value
                    if isinstance(value, date):
                        cell.number_format = "dd/mm/yyyy"
                    elif isinstance(value, time):
                        cell.number_format = "hh:mm:ss"
                    # Set text wrapping for the Notes column
                    if col == data_len:  # Last column (Notes)
                        cell.alignment = Alignment(wrapText=True, vertical="top")

                current_row += 1
