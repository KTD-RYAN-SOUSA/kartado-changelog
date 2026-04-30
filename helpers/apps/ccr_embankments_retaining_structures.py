import os
import tempfile
import time
from os.path import isfile
from typing import List
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import (
    get_form_array_iterator,
    new_get_form_data,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_picture,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import ThreadExecutor, synchronized_request_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_custom_option,
    get_end_km,
    get_km,
    get_previous_found_at_reporting,
)
from helpers.strings import clean_latin_string, get_obj_from_path


def get_form_data(
    reporting: Reporting,
    field_name: str,
    data_name: str = None,
    subgroup: str = None,
    value: str = None,
) -> str:
    try:
        dataname = data_name if data_name else field_name
        form_fields = reporting.occurrence_type.form_fields["fields"]
        if not subgroup:
            form_field = [
                obj
                for obj in form_fields
                if get_obj_from_path(obj, "apiname") == dataname
            ]
        else:
            form_field = [
                obj
                for obj in form_fields
                if get_obj_from_path(obj, "apiname") == subgroup
            ]
            if form_field:
                form_field = [
                    obj
                    for obj in form_field[0]["innerFields"]
                    if get_obj_from_path(obj, "apiname") == dataname
                ]

        if form_field:
            field_type = get_obj_from_path(form_field[0], "datatype")

            data = (
                get_obj_from_path(reporting.form_data, field_name)
                if not value
                else value
            )

            if data is None:
                return data

            if field_type in ["string", "number", "float"]:
                return data

            elif field_type == "select":
                options = get_obj_from_path(form_field[0], "selectoptions__options")
                my_option = next(a for a in options if a["value"] == data)
                return my_option["name"]

            elif field_type == "selectMultiple":
                options = get_obj_from_path(form_field[0], "selectoptions__options")
                my_options = [i["name"] for i in options if i["value"] in data]
                return ", ".join(my_options) if my_options else ""

            elif field_type == "arrayOfObjects":
                options = get_obj_from_path(form_field[0], "innerfields")
                options = next(a for a in options if a["apiName"] == field_name)
                if options:
                    options = get_obj_from_path(options, "selectoptions__options")
                    my_option = [a for a in options if a["value"] == data]
                    return my_option[0]["name"] if my_option else None
    except Exception:
        pass
    return None


class XlsxHandler:
    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.list_uuids = list_uuids
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.__xlsx_file = (
            "./fixtures/reports/ccr_embankments_retaining_structures.xlsx"
        )
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet: Worksheet = self._workbook["Terrapleno_1"]
        self._worksheet_pg2 = self._workbook["Terrapleno_2"]
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="E1:F2",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B2",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        self.__static_fields_pg_1 = {
            "id_ccr_antt": "B5",
            "km": "D5",
            "road_name": "F5",
            "type": "B6",
            "executed_at": "F6",
            "init_km": "B8",
            "end_km": "B9",
            "direction": "D8",
            "x_utm": "F8",
            "y_utm": "F9",
            "extension": "B11",
            "height_terrapleno": "D11",
            "inclination": "F11",
            "distance": "B12",
            "tipo_terrapleno": "B14",
            "vegetation": "B15",
            "terrain_embossing": "F14",
            "vegetation_density": "F15",
            "structurekind": "B17",
            "extenstion": "B18",
            "inner_height": "B19",
            "anchor": "B20",
            "elements_concrete": "B21",
            "drenagem_superficial": "B23",
            "drenagem_subterranea": "B24",
            "condition_dre_sup": "E23",
            "kind_drenagem_subterranea": "D24",
            "condition_dre_prof": "F24",
            "water_presence": "B26",
            "occurrence_type": "B28",
            "probable_causes": "B30",
            "passivo_ambiental": "B31",
            "risk_level": "E33",
            "rupture_case": "E34",
            "notes": "A36",
        }
        self.__static_fields_pg_2 = {
            "img_1": [1, 6],
            "img_2": [2, 6],
            "img_3": [1, 19],
            "img_4": [2, 19],
            "croqui": [1, 32, 2, 47],
        }

    def fill_sheet(self, data: list):
        image_paths: List[str] = []
        for key, value in data.items():
            if key not in ["images", "reporting"]:
                self._worksheet[self.__static_fields_pg_1[key]] = (
                    value if value else "-"
                )
            elif key == "images":
                for k, img in value.items():
                    if img:
                        col = self.__static_fields_pg_2[k][0]
                        row = self.__static_fields_pg_2[k][1]
                        if k != "croqui":
                            col_letter = get_column_letter(col)
                            range_str = f"{col_letter}{row}:{col_letter}{row+11}"
                            try:
                                insert_picture_2(
                                    self._worksheet_pg2,
                                    range_str,
                                    Image(img),
                                    self.__sheet_target,
                                    (1, 1, 1, 1),
                                    ResizeMethod.ProportionalCentered,
                                )
                            except Exception as e:
                                print(e)
                        else:
                            col_letter = get_column_letter(col)
                            col_2_letter = get_column_letter(
                                self.__static_fields_pg_2[k][2]
                            )
                            row_2 = self.__static_fields_pg_2[k][3]
                            range_str = f"{col_letter}{row}:{col_2_letter}{row_2}"
                            try:
                                insert_picture_2(
                                    self._worksheet_pg2,
                                    range_str,
                                    Image(img),
                                    self.__sheet_target,
                                    (1, 1, 1, 1),
                                    ResizeMethod.ProportionalCentered,
                                )
                            except Exception as e:
                                print(e)
                        image_paths.append(img)
            else:
                pass
        _filename = data["id_ccr_antt"]
        file_name_options = [
            _filename if _filename else "TP SEM NOME",
            str(time.time()),
        ]
        file_name = clean_latin_string(
            next(a for a in file_name_options if a).replace(".", "").replace("/", "")
        )

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            target=self.__sheet_target,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

        worksheet_pg2_data_logo_company = self.data_logo_company.copy()
        worksheet_pg2_data_data_provider_logo = self.data_provider_logo.copy()

        worksheet_pg2_data_logo_company["range_string"] = "B1:B3"

        worksheet_pg2_data_data_provider_logo["range_string"] = "A1:A3"

        insert_logo_and_provider_logo(
            worksheet=self._worksheet_pg2,
            target=self.__sheet_target,
            logo_company=worksheet_pg2_data_logo_company,
            provider_logo=worksheet_pg2_data_data_provider_logo,
        )

        result = f"/tmp/{file_name} {str(time.time())}.xlsx"
        self._workbook.save(result)
        self._workbook.close()
        for image_path in image_paths:
            if isfile(image_path):
                os.remove(image_path)
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Terrapleno_1"]
        self._worksheet_pg2 = self._workbook["Terrapleno_2"]

        return result

    @classmethod
    def __get_panoramic_picture(cls, s3, temp_dir: str, reporting: Reporting) -> object:
        picture_path = None
        try:
            uuids = new_get_form_data(
                reporting, "mandatoryPictures__0__panoramicMandatory"
            )
            reporting_files = (
                ReportingFile.objects.filter(uuid__in=uuids, is_shared=True)
                .only("upload")
                .order_by("-datetime", "-uploaded_at")
            )
            for reporting_file in reporting_files:
                curr_path = download_picture(
                    s3,
                    temp_dir,
                    str(reporting_file.uuid),
                    reporting_file=reporting_file,
                    quality="400px",
                )
                if curr_path is not None and curr_path != "":
                    picture_path = curr_path
                    break
        except Exception as e:
            print(e)
        return picture_path

    def create_dict(self, reporting: Reporting, s3) -> dict:
        km = get_km(reporting)
        end_km = get_end_km(reporting)
        direction = get_custom_option(reporting, "direction", "")
        contention_text = (
            "Contenção"
            if reporting.form_data.get("is_contention") is True
            else "Terrapleno"
        )
        tipo_terrapleno = get_form_data(
            reporting,
            data_name="tipoTerrapleno",
            field_name="tipo_terrapleno",
        )
        vegetation = get_form_data(
            reporting,
            data_name="vegetation",
            field_name="vegetation",
        )
        terrain_embossing = get_form_data(
            reporting,
            data_name="terrainEmbossing",
            field_name="terrain_embossing",
        )
        vegetation_density = get_form_data(
            reporting,
            data_name="vegetationDensity",
            field_name="vegetation_density",
        )
        drenagem_superficial = get_form_data(
            reporting,
            data_name="drenagemSuperficial",
            field_name="drenagem_superficial",
        )
        drenagem_subterranea = get_form_data(
            reporting,
            data_name="drenagemSubterranea",
            field_name="drenagem_subterranea",
        )
        condition_dre_sup = get_form_data(
            reporting,
            data_name="conditionDreSup",
            field_name="condition_dre_sup",
        )
        kind_drenagem_subterranea = get_form_data(
            reporting,
            data_name="kindDrenagemSubterranea",
            field_name="kind_drenagem_subterranea",
        )
        condition_dre_prof = get_form_data(
            reporting,
            data_name="conditionDreProf",
            field_name="condition_dre_prof",
        )
        water_presence = get_form_data(
            reporting,
            data_name="waterPresence",
            field_name="water_presence",
        )
        risk_level = get_form_data(
            reporting,
            data_name="riskLevel",
            field_name="risk_level",
        )

        occurrence_type = "-"
        try:
            varOccurrenceType = new_get_form_data(
                reporting, "ocorrenciaTipo", default=[]
            )
            rawVarOccurrenceType = new_get_form_data(
                reporting, "ocorrenciaTipo", raw=True, default=[]
            )
            if len(rawVarOccurrenceType) == 1 and int(rawVarOccurrenceType[0]) == 15:
                occurrence_type = new_get_form_data(
                    reporting, "occurrenceTypeOther", default="-"
                )
            elif len(rawVarOccurrenceType) > 0:
                occurrence_type = "/".join(varOccurrenceType)
        except Exception:
            pass

        contention = reporting.form_data.get("contention")
        if contention:
            structurekind_list = [
                (
                    str(
                        get_form_data(
                            reporting,
                            data_name="contention",
                            field_name="structureKind",
                            value=get_obj_from_path(x, "structurekind"),
                        )
                    )
                    if get_obj_from_path(x, "structurekind")
                    else "-"
                )
                for x in contention
            ]
            structurekind = "/".join(structurekind_list)
            extenstion_list = [
                str(x.get("extenstion")) if x.get("extenstion") else "-"
                for x in contention
            ]
            extenstion = "/".join(extenstion_list)
            inner_height_list = [
                str(x.get("inner_height")) if x.get("inner_height") else "-"
                for x in contention
            ]
            inner_height = "/".join(inner_height_list)
            anchor_list = [
                (
                    str(
                        get_form_data(
                            reporting,
                            data_name="contention",
                            field_name="anchor",
                            value=x.get("anchor"),
                        )
                    )
                    if x.get("anchor")
                    else "-"
                )
                for x in contention
            ]
            anchor = "/".join(anchor_list)
            elements_concrete_list = [
                (
                    str(
                        get_form_data(
                            reporting,
                            data_name="contention",
                            field_name="elementsConcrete",
                            value=x.get("elements_concrete"),
                        )
                    )
                    if x.get("elements_concrete")
                    else "-"
                )
                for x in contention
            ]
            elements_concrete = "/".join(elements_concrete_list)
        else:
            structurekind = "-"
            extenstion = "-"
            inner_height = "-"
            anchor = "-"
            elements_concrete = "-"

        probable_causes_raw = new_get_form_data(reporting, "probableCauses", raw=True)
        if isinstance(probable_causes_raw, list):
            if (
                "10" in probable_causes_raw
                and len(probable_causes_raw) == 1
                and reporting.form_data.get("probable_causes_other")
            ):
                probable_causes = new_get_form_data(reporting, "probableCausesOther")
            else:
                probable_causes_list = new_get_form_data(reporting, "probableCauses")
                probable_causes = ", ".join(probable_causes_list)
        else:
            probable_causes = "-"

        if reporting.form_data.get("passivo_ambiental") is True:
            environment_description = reporting.form_data.get("environment_description")
            passivo_ambiental = f"SIM - {environment_description}"
        elif reporting.form_data.get("passivo_ambiental") is False:
            passivo_ambiental = "NÃO"
        else:
            passivo_ambiental = "-"

        if reporting.form_data.get("rupture_case") is True:
            rupture_case = "SIM"
        elif reporting.form_data.get("rupture_case") is False:
            rupture_case = "NÃO"
        else:
            rupture_case = "-"

        images_list = reporting.form_data.get("mandatory_pictures")
        picture_order = []
        foot_mandatory_list = []
        criste_mandatory_list = []
        panoramic_mandatory_list = []
        patology_mandatory_list = []
        left_mandatory_list = []
        right_mandatory_list = []
        details_mandatory_list = []
        therapy_image_list = []
        if images_list:
            for intern_list in images_list:
                foot_mandatory_list.extend(intern_list.get("foot_mandatory", []))
                criste_mandatory_list.extend(intern_list.get("criste_mandatory", []))
                panoramic_mandatory_list.extend(
                    intern_list.get("panoramic_mandatory", [])
                )
                patology_mandatory_list.extend(
                    intern_list.get("patology_mandatory", [])
                )
                left_mandatory_list.extend(intern_list.get("left_mandatory", []))
                right_mandatory_list.extend(intern_list.get("right_mandatory", []))
                details_mandatory_list.extend(intern_list.get("details_mandatory", []))

        inner_pictures_list = []
        if reporting.form_data.get("contention"):
            for contention_dict in reporting.form_data.get("contention"):
                if contention_dict.get("inner_pictures"):
                    for img in contention_dict.get("inner_pictures"):
                        inner_pictures_list.append(img)
        drainage_pictures_list = (
            reporting.form_data.get("drainage_pictures")
            if reporting.form_data.get("drainage_pictures")
            else []
        )
        if drainage_pictures_list:
            new_list = []
            for item in drainage_pictures_list:
                for key, value in item.items():
                    new_list.extend(value)
            drainage_pictures_list = new_list
        if reporting.form_data.get("therapy"):
            for therapy in reporting.form_data["therapy"]:
                if therapy.get("treatment_images"):
                    therapy_image_list = (
                        [item for item in therapy.get("treatment_images")]
                        if reporting.form_data.get("therapy")
                        else []
                    )

        relationships = []

        reporting_files = ReportingFile.objects.filter(
            reporting=reporting, is_shared=True
        ).only("uuid")
        reporting_file_uuids = [rf.uuid for rf in reporting_files]
        for uuid in reporting_file_uuids:
            check = []
            check.extend(
                [
                    str(uuid) not in foot_mandatory_list,
                    str(uuid) not in criste_mandatory_list,
                    str(uuid) not in panoramic_mandatory_list,
                    str(uuid) not in patology_mandatory_list,
                    str(uuid) not in left_mandatory_list,
                    str(uuid) not in right_mandatory_list,
                    str(uuid) not in details_mandatory_list,
                    str(uuid) not in drainage_pictures_list,
                    str(uuid) not in therapy_image_list,
                    str(uuid) not in inner_pictures_list,
                ]
            )
            if False not in check:
                relationships.append(str(uuid))

        inner_pictures_list = list(
            ReportingFile.objects.filter(uuid__in=inner_pictures_list, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        patology_mandatory_list = list(
            ReportingFile.objects.filter(
                uuid__in=patology_mandatory_list, is_shared=True
            )
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        drainage_pictures_list = list(
            ReportingFile.objects.filter(
                uuid__in=drainage_pictures_list, is_shared=True
            )
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        foot_mandatory_list = list(
            ReportingFile.objects.filter(uuid__in=foot_mandatory_list, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        criste_mandatory_list = list(
            ReportingFile.objects.filter(uuid__in=criste_mandatory_list, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        left_mandatory_list = list(
            ReportingFile.objects.filter(uuid__in=left_mandatory_list, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        right_mandatory_list = list(
            ReportingFile.objects.filter(uuid__in=right_mandatory_list, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        details_mandatory_list = list(
            ReportingFile.objects.filter(
                uuid__in=details_mandatory_list, is_shared=True
            )
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        panoramic_mandatory_list = list(
            ReportingFile.objects.filter(
                uuid__in=panoramic_mandatory_list, is_shared=True
            )
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        therapy_image_list = list(
            ReportingFile.objects.filter(uuid__in=therapy_image_list, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )
        relationships = list(
            ReportingFile.objects.filter(uuid__in=relationships, is_shared=True)
            .only("uuid")
            .values_list("uuid", flat=True, named=False)
        )

        inner_pictures_list = [str(uuid) for uuid in inner_pictures_list]
        patology_mandatory_list = [str(uuid) for uuid in patology_mandatory_list]
        drainage_pictures_list = [str(uuid) for uuid in drainage_pictures_list]
        foot_mandatory_list = [str(uuid) for uuid in foot_mandatory_list]
        criste_mandatory_list = [str(uuid) for uuid in criste_mandatory_list]
        left_mandatory_list = [str(uuid) for uuid in left_mandatory_list]
        right_mandatory_list = [str(uuid) for uuid in right_mandatory_list]
        details_mandatory_list = [str(uuid) for uuid in details_mandatory_list]
        panoramic_mandatory_list = [str(uuid) for uuid in panoramic_mandatory_list]
        therapy_image_list = [str(uuid) for uuid in therapy_image_list]
        relationships = [str(uuid) for uuid in relationships]

        if inner_pictures_list:
            picture_order.append(inner_pictures_list[0])
        if patology_mandatory_list:
            picture_order.append(patology_mandatory_list[0])
        if drainage_pictures_list:
            picture_order.append(drainage_pictures_list[0])
        if foot_mandatory_list:
            picture_order.append(foot_mandatory_list[0])
        if criste_mandatory_list:
            picture_order.append(criste_mandatory_list[0])
        if left_mandatory_list:
            picture_order.append(left_mandatory_list[0])
        if right_mandatory_list:
            picture_order.append(right_mandatory_list[0])
        if details_mandatory_list:
            picture_order.append(details_mandatory_list[0])
        if len(panoramic_mandatory_list) > 1:
            picture_order.append(panoramic_mandatory_list[1])
        if len(panoramic_mandatory_list) > 2:
            picture_order.append(panoramic_mandatory_list[2])
        if len(therapy_image_list) > 1:
            picture_order.append(therapy_image_list[1])
        if len(therapy_image_list) > 2:
            picture_order.append(therapy_image_list[2])
        if relationships and len(picture_order) <= 1:
            picture_order.append(relationships[0])
        if len(relationships) > 1 and len(picture_order) == 1:
            picture_order.append(relationships[1])

        previous_reporting = get_previous_found_at_reporting(reporting, "form_data")

        img_1 = ""
        if previous_reporting is not None:
            panoramic_mandatory = XlsxHandler.__get_panoramic_picture(
                self.s3, self.temp_file, previous_reporting
            )
            if panoramic_mandatory is not None:
                img_1 = panoramic_mandatory
        self.temp_file = tempfile.mkdtemp()

        img_2 = ""
        img_3 = ""
        img_4 = ""
        croqui = ""

        for rf_uuid in panoramic_mandatory_list:
            try:
                img_2 = download_picture(
                    self.s3,
                    self.temp_file,
                    rf_uuid,
                    picture_uuid=rf_uuid,
                    quality="400px",
                    enable_is_shared=True,
                )
            except Exception:
                pass
            if img_2 is not None and img_2 != "":
                break

        image_index = -1
        for i, rf_uuid in enumerate(picture_order):
            try:
                img_3 = download_picture(
                    self.s3,
                    self.temp_file,
                    rf_uuid,
                    picture_uuid=rf_uuid,
                    quality="400px",
                    enable_is_shared=True,
                )
            except Exception:
                pass
            if img_3 is not None and img_3 != "":
                image_index = i
                break
        if image_index != -1:
            picture_order = picture_order[image_index + 1 :]

        for rf_uuid in picture_order:
            try:
                img_4 = download_picture(
                    self.s3,
                    self.temp_file,
                    rf_uuid,
                    picture_uuid=rf_uuid,
                    quality="400px",
                    enable_is_shared=True,
                )
            except Exception:
                pass
            if img_4 is not None and img_4 != "":
                break

        croqui_image_uuids: List[UUID] = []
        it = get_form_array_iterator(reporting, "croquiImages")
        try:
            while True:
                croqui_image_array = it.get("croquiImage")
                for croqui_image_uuid in croqui_image_array:
                    try:
                        croqui_image_uuids.append(UUID(croqui_image_uuid))
                    except Exception:
                        continue
                it.inc()
        except Exception as e:
            print(e)

        croqui_rfs = ReportingFile.objects.filter(
            uuid__in=croqui_image_uuids,
            is_shared=True,
        ).order_by("datetime", "uploaded_at")

        for croqui_rf in croqui_rfs:
            try:
                croqui = download_picture(
                    self.s3,
                    self.temp_file,
                    rf_uuid,
                    reporting_file=croqui_rf,
                )
            except Exception:
                pass
            if croqui is not None:
                break

        notes = reporting.form_data.get("notes")
        length = reporting.form_data.get("length")
        _extension = reporting.form_data.get("extension")
        x_utm = reporting.form_data.get("x_utm")
        y_utm = reporting.form_data.get("y_utm")
        executed_at = ""
        if reporting.executed_at is not None:
            executed_at = reporting.executed_at.strftime("%d/%m/%Y")
        data = {
            "id_ccr_antt": reporting.form_data.get("id_ccr_antt"),
            "km": km,
            "road_name": reporting.road_name,
            "type": contention_text,
            "executed_at": executed_at,
            "init_km": km,
            "end_km": end_km,
            "direction": direction,
            "x_utm": f"X {x_utm}" if x_utm else None,
            "y_utm": f"Y {y_utm}" if y_utm else None,
            "extension": _extension if _extension else length,
            "height_terrapleno": reporting.form_data.get("height_terrapleno"),
            "inclination": reporting.form_data.get("inclination"),
            "distance": reporting.form_data.get("distance"),
            "tipo_terrapleno": tipo_terrapleno,
            "vegetation": vegetation,
            "terrain_embossing": terrain_embossing,
            "vegetation_density": vegetation_density,
            "structurekind": structurekind,
            "extenstion": extenstion,
            "inner_height": inner_height,
            "anchor": anchor,
            "elements_concrete": elements_concrete,
            "drenagem_superficial": drenagem_superficial,
            "drenagem_subterranea": drenagem_subterranea,
            "condition_dre_sup": condition_dre_sup,
            "kind_drenagem_subterranea": kind_drenagem_subterranea,
            "condition_dre_prof": condition_dre_prof,
            "water_presence": water_presence,
            "occurrence_type": occurrence_type,
            "probable_causes": probable_causes,
            "passivo_ambiental": passivo_ambiental,
            "risk_level": risk_level,
            "rupture_case": rupture_case,
            "notes": notes if notes else "-",
            "images": {
                "img_1": img_1,
                "img_2": img_2,
                "img_3": img_3,
                "img_4": img_4,
                "croqui": croqui,
            },
            "reporting": reporting,
        }

        return data

    def execute(self):
        query_set = (
            Reporting.objects.filter(uuid__in=self.list_uuids)
            .prefetch_related("occurrence_type", "firm", "firm__subcompany", "company")
            .distinct()
        )
        convert_executor: ThreadExecutor = None
        if self.__report_format == ReportFormat.PDF:
            convert_executor = ThreadExecutor(50)

        files = []
        for reporting in query_set:
            data = self.create_dict(reporting=reporting, s3=self.s3)

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_logo_company:
                    self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

            sheet_file = self.fill_sheet(data)
            if self.__report_format == ReportFormat.PDF:
                convert_executor.submit(synchronized_request_pdf, sheet_file)
            else:
                files.append(sheet_file)

        if self.__report_format == ReportFormat.PDF:
            files = list(set(convert_executor.get()))
            files.sort()

        return files


class CCREmbankmentsRetainingStructures(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""

        if len(self.uuids) == 1:
            reporting = Reporting.objects.get(uuid=self.uuids[0])

            file_name_options = [
                get_obj_from_path(reporting.form_data, "id_ccr_antt"),
                reporting.number,
                str(time.time()),
            ]
            file_name = clean_latin_string(
                next(a for a in file_name_options if a)
                .replace(".", "")
                .replace("/", "")
            )

            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{file_name}.{extension}"

        elif len(self.uuids) > 1:
            file_name = "Anexo IV Fichas.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandler(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_embankments_retaining_structures_async_handler(
    reporter_dict: dict,
):
    reporter = CCREmbankmentsRetainingStructures.from_dict(reporter_dict)
    reporter.export()
