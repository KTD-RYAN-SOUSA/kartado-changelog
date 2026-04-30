import shutil
import tempfile
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Tuple
from uuid import UUID
from zipfile import ZipFile

from django.db.models import (
    ExpressionWrapper,
    F,
    IntegerField,
    OuterRef,
    Subquery,
    Value,
)
from django.db.models.functions import Coalesce, ExtractYear
from django.utils.timezone import make_aware
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    SheetTarget,
    get_image,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    form_data_images_grouped,
    get_direction,
    get_identification,
    get_km,
    get_parent_serial,
    get_road_name,
    get_serial,
)
from helpers.apps.ccr_report_utils.workbook_utils import save_workbook, set_row_style
from helpers.strings import clean_latin_string


class EmbankmentsIIXlsxHandler(object):

    __PICTURE_FIELDS = [
        "mandatoryPictures__panoramicMandatory",
        "contention__innerPictures",
        "drainagePictures__drainagePicture",
        "mandatoryPictures__footMandatory",
        "mandatoryPictures__cristeMandatory",
        "mandatoryPictures__leftMandatory",
        "mandatoryPictures__rightMandatory",
        "mandatoryPictures__detailsMandatory",
        "therapy__treatmentImages",
    ]

    __OCCURRENCE_TYPES: Dict[str, str] = {}

    __DEFAULT_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    __DEFAULT_FONT = Font(name="Arial", size=9)

    __DEFAULT_ALIGNMENT = Alignment(
        vertical="center", horizontal="center", wrap_text=True
    )

    __RISK_TEXT_LOOKUP = {
        "": 0,
        "R1 - Baixo": 1,
        "R2 - Moderado": 2,
        "R3 - Alto": 3,
        "R4 - Muito Alto": 4,
    }

    __RISK_COLOR_LOOKUP = {
        "": "ffffff",
        "R1 - Baixo": "00b050",
        "R2 - Moderado": "ffff00",
        "R3 - Alto": "ffc000",
        "R4 - Muito Alto": "ff0000",
    }

    @classmethod
    def __get_passivo_ambiental(cls, reporting: Reporting) -> str:
        try:
            passivo_ambiental = new_get_form_data(reporting, "passivoAmbiental")
            return "SIM" if passivo_ambiental else "NÃO"
        except Exception:
            return ""

    @classmethod
    def __has_valid_risk_level(cls, reporting: Reporting) -> int:
        valid = False
        try:
            num_risk_level = EmbankmentsIIXlsxHandler.__RISK_TEXT_LOOKUP[
                new_get_form_data(reporting, "riskLevel")
            ]
            if num_risk_level >= 1 and num_risk_level <= 4:
                valid = True
        except Exception:
            pass

        return valid

    @classmethod
    def __get_risk_level(cls, reporting: Reporting) -> int:
        try:
            return EmbankmentsIIXlsxHandler.__RISK_TEXT_LOOKUP[
                new_get_form_data(reporting, "riskLevel")
            ]
        except Exception:
            return 0

    @classmethod
    def __get_risk_level_text(cls, reporting: Reporting) -> str:
        try:
            return new_get_form_data(reporting, "riskLevel")
        except Exception:
            return ""

    @classmethod
    def __get_risk_level_color(cls, reporting: Reporting) -> str:
        try:
            risk_level = new_get_form_data(reporting, "riskLevel")
            return EmbankmentsIIXlsxHandler.__RISK_COLOR_LOOKUP[risk_level]
        except Exception:
            return "ffffff"

    @classmethod
    def __get_therapies(cls, reporting: Reporting) -> str:
        try:
            therapy = new_get_form_data(reporting, "therapy")
            occurrence_types = sorted(
                list(
                    {
                        EmbankmentsIIXlsxHandler.__OCCURRENCE_TYPES[
                            anomaly["occurrence_type"]
                        ]
                        for anomaly in therapy
                        if "occurrence_type" in anomaly
                    }
                )
            )
            return ", ".join(occurrence_types)
        except Exception:
            return ""

    def __get_picture(
        self, reporting: Reporting, reporting_files: Dict[str, ReportingFile]
    ):
        image = None
        rfs_by_priority: List[ReportingFile] = []

        image_arrays = form_data_images_grouped(reporting)
        try:
            patology_rfs = [
                reporting_files[uuid]
                for uuid in image_arrays["mandatoryPictures__patologyMandatory"].uuids
                if uuid in reporting_files
            ]
            patology_rfs = sorted(
                patology_rfs, key=lambda rf: (rf.datetime, rf.uploaded_at)
            )
            rfs_by_priority.extend(patology_rfs)
        except Exception:
            pass

        try:
            form_data_picture_uuids: List[str] = []
            for _, image_array in image_arrays.items():
                form_data_picture_uuids.extend(image_array.uuids)
            files_rfs = [
                rf
                for rf in reporting.reporting_files.all()
                if rf.is_shared and str(rf.uuid) not in form_data_picture_uuids
            ]
            files_rfs = sorted(files_rfs, key=lambda rf: (rf.datetime, rf.uploaded_at))
            rfs_by_priority.extend(files_rfs)
        except Exception:
            pass

        for picture_field_key in EmbankmentsIIXlsxHandler.__PICTURE_FIELDS:
            if picture_field_key in image_arrays:
                try:
                    curr_field_rfs = [
                        reporting_files[uuid]
                        for uuid in image_arrays[picture_field_key].uuids
                        if uuid in reporting_files
                    ]
                    curr_field_rfs = sorted(
                        curr_field_rfs, key=lambda rf: (rf.datetime, rf.uploaded_at)
                    )
                    rfs_by_priority.extend(curr_field_rfs)
                except Exception:
                    pass

        try:
            non_croqui_rfs: List[ReportingFile] = []
            for key, image_array in image_arrays.items():
                if (
                    key not in EmbankmentsIIXlsxHandler.__PICTURE_FIELDS
                    and key != "croquiImages__croquiImage"
                ):
                    rfs = [
                        reporting_files[uuid]
                        for uuid in image_array.uuids
                        if uuid in reporting_files
                    ]
                    non_croqui_rfs.extend(rfs)
            non_croqui_rfs = sorted(
                non_croqui_rfs, key=lambda rf: (rf.datetime, rf.uploaded_at)
            )
            rfs_by_priority.extend(non_croqui_rfs)
        except Exception:
            pass

        for rf in rfs_by_priority:
            try:
                image = get_image(
                    self.s3,
                    self.temp_dir,
                    "temp-" + str(rf.uuid) + "-pic",
                    rf,
                )
                if image is not None:
                    break
            except Exception:
                pass
        return image

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.__sheet_target = sheet_target
        self.__sheet_reporting: Dict[Tuple[str, int], List[Reporting]] = {}
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.occurrence_type = Reporting.objects.get(uuid=list_uuids[0]).occurrence_type

        for e in OccurrenceType.objects.values("uuid", "name"):
            EmbankmentsIIXlsxHandler.__OCCURRENCE_TYPES[str(e["uuid"])] = e["name"]

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __append_entry(
        self,
        worksheet: Worksheet,
        curr_reporting: Reporting,
        prev_reporting: Reporting,
        reporting_files: Dict[str, ReportingFile],
    ) -> None:
        curr_risk_level = EmbankmentsIIXlsxHandler.__get_risk_level_text(curr_reporting)
        prev_risk_level = EmbankmentsIIXlsxHandler.__get_risk_level_text(prev_reporting)
        road_name = get_road_name(curr_reporting)
        km = get_km(curr_reporting)
        direction = get_direction(curr_reporting)[0]
        identification = get_identification(curr_reporting)
        passivo_ambiental = EmbankmentsIIXlsxHandler.__get_passivo_ambiental(
            curr_reporting
        )
        therapies = EmbankmentsIIXlsxHandler.__get_therapies(prev_reporting)
        serial = get_serial(curr_reporting)
        inventory_serial = get_parent_serial(curr_reporting)

        prev_color = EmbankmentsIIXlsxHandler.__get_risk_level_color(prev_reporting)
        prev_risk_cell = Cell(worksheet, value=prev_risk_level)
        prev_risk_cell.fill = PatternFill(start_color=prev_color, fill_type="solid")

        curr_color = EmbankmentsIIXlsxHandler.__get_risk_level_color(curr_reporting)
        curr_risk_cell = Cell(worksheet, value=curr_risk_level)
        curr_risk_cell.fill = PatternFill(start_color=curr_color, fill_type="solid")

        text_entry: List[str | Cell] = [
            road_name,
            km,
            direction,
            identification,
            passivo_ambiental,
            prev_risk_cell,
            "",
            curr_risk_cell,
            "",
            therapies,
            serial,
            inventory_serial,
        ]

        worksheet.append(text_entry)

        row_height = worksheet.row_dimensions[5].height
        last_row_index = worksheet.max_row

        set_row_style(
            worksheet,
            row=last_row_index,
            height=row_height,
            font=EmbankmentsIIXlsxHandler.__DEFAULT_FONT,
            border=EmbankmentsIIXlsxHandler.__DEFAULT_BORDER,
            alignment=EmbankmentsIIXlsxHandler.__DEFAULT_ALIGNMENT,
        )

        prev_picture = self.__get_picture(prev_reporting, reporting_files)
        try:
            insert_picture(
                worksheet,
                range_string=f"G{last_row_index}",
                picture=prev_picture,
                target=self.__sheet_target,
            )
        except Exception:
            pass

        curr_picture = self.__get_picture(curr_reporting, reporting_files)
        try:
            insert_picture(
                worksheet,
                range_string=f"I{last_row_index}",
                picture=curr_picture,
                target=self.__sheet_target,
            )
        except Exception as e:
            print(e)

    def __map_sheet_to_reporting(
        self, reportings: List[Reporting]
    ) -> Dict[Tuple[str, int], List[Reporting]]:
        sheet_reporting: Dict[Tuple[str, int], List[Reporting]] = {}
        for reporting in reportings:
            found_at = reporting.found_at
            time_delta = timedelta(hours=-3)
            time_zone = timezone(time_delta)
            found_at = found_at.astimezone(time_zone)

            sheet = (reporting.road_name, found_at.year)
            if sheet not in sheet_reporting:
                sheet_reporting[sheet] = []

            sheet_reporting[sheet].append(reporting)
        return sheet_reporting

    def __create_workbooks_files(
        self,
        sheet_to_reporting: Dict[Tuple[str, int], List[Reporting]],
        previous_reportings: Dict[UUID, Reporting],
        reporting_files: Dict[UUID, ReportingFile],
    ) -> List[str]:
        files: List[str] = []

        sorted_sheets = sorted(sheet_to_reporting.items())
        for sheet_key, reportings in sorted_sheets:
            (road_name, year) = sheet_key
            reportings.sort(key=lambda reporting: reporting.km)

            workbook = load_workbook(
                "./fixtures/reports/ccr_report_embankments_retaining_structures_annex_two.xlsx"
            )
            worksheet = workbook[workbook.sheetnames[0]]

            worksheet.column_dimensions["K"].hidden = True
            worksheet.column_dimensions["L"].hidden = True

            for curr_reporting in reportings:
                previous_reporting: Reporting = None
                if curr_reporting.previous_reporting_uuid is not None:
                    previous_reporting = previous_reportings[
                        curr_reporting.previous_reporting_uuid
                    ]

                if (
                    previous_reporting is None
                    or not (
                        EmbankmentsIIXlsxHandler.__has_valid_risk_level(curr_reporting)
                    )
                    or not (
                        EmbankmentsIIXlsxHandler.__has_valid_risk_level(
                            previous_reporting
                        )
                    )
                ):
                    continue

                curr_risk_level = EmbankmentsIIXlsxHandler.__get_risk_level(
                    curr_reporting
                )
                previous_risk_level = EmbankmentsIIXlsxHandler.__get_risk_level(
                    previous_reporting
                )
                if curr_risk_level < previous_risk_level:
                    self.__append_entry(
                        worksheet, curr_reporting, previous_reporting, reporting_files
                    )

            workbook_name = "ANEXO II - COMPARATIVO COM MONITORAÇÃO ANTERIOR EM RELAÇÃO AOS SERVIÇOS REALIZADOS {} {}".format(
                road_name, year
            )
            file = save_workbook(workbook_name, workbook)
            files.append(file)

        return files

    def execute(self) -> List[str]:

        reportings_qs = (
            Reporting.objects.filter(
                uuid__in=self.list_uuids,
                parent__isnull=False,
            )
            .annotate(
                found_at_year=ExpressionWrapper(
                    ExtractYear(
                        Coalesce(F("found_at"), Value(make_aware(datetime(1970, 1, 1))))
                    )
                    - Value(1),
                    output_field=IntegerField(),
                ),
                previous_reporting_uuid=Subquery(
                    Reporting.objects.filter(
                        occurrence_type=OuterRef("occurrence_type"),
                        parent=OuterRef("parent"),
                        found_at__year=OuterRef("found_at_year"),
                    )
                    .order_by("found_at")
                    .reverse()
                    .values("uuid")[:1]
                ),
            )
            .select_related("parent")
            .prefetch_related("occurrence_type", "company", "reporting_files")
            .only(
                "uuid",
                "road_name",
                "km",
                "direction",
                "form_data",
                "found_at",
                "parent__number",
                "occurrence_type",
                "company",
            )
        )
        reportings = []

        reporting_uuids = []
        previous_reporting_uuids = []
        for r in reportings_qs:
            reportings.append(r)
            reporting_uuids.append(r.uuid)
            if r.previous_reporting_uuid is not None:
                previous_reporting_uuids.append(r.previous_reporting_uuid)

        previous_reportings_qs = (
            Reporting.objects.filter(uuid__in=previous_reporting_uuids)
            .only("uuid", "form_data", "number", "occurrence_type")
            .select_related("occurrence_type")
        )
        previous_reportings = {pr.uuid: pr for pr in previous_reportings_qs}

        reporting_files_qs = (
            ReportingFile.objects.filter(
                reporting__uuid__in=reporting_uuids + previous_reporting_uuids,
                is_shared=True,
            )
            .order_by("uploaded_at")
            .only("upload", "is_shared")
        )

        reporting_files = {str(rf.uuid): rf for rf in reporting_files_qs}

        sheet_to_reporting = self.__map_sheet_to_reporting(reportings)

        return self.__create_workbooks_files(
            sheet_to_reporting, previous_reportings, reporting_files
        )


class CCREmbankmentsRetainingStructuresAnnexTwo(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        road_names = list(
            Reporting.objects.filter(uuid__in=self.uuids)
            .only("uuid", "road_name")
            .order_by("road_name")
            .distinct("road_name")
            .values_list("road_name", flat=True)
        )
        file_name = "ANEXO II - COMPARATIVO COM MONITORAÇÃO ANTERIOR EM RELAÇÃO AOS SERVIÇOS REALIZADOS {}".format(
            "-".join(road_names)
        )
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"
        return file_name

    def export(self):
        s3 = get_s3()
        files = EmbankmentsIIXlsxHandler(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_embankments_retaining_structures_annex_two_async_handler(
    reporter_dict: dict,
):
    reporter = CCREmbankmentsRetainingStructuresAnnexTwo.from_dict(reporter_dict)
    reporter.export()
