import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingInReporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture,
    result_photos,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km


class XlsxHandlerReportOACAnnexTwo(object):

    __GREEN_CELLS: List[str] = {
        "B19:B21",
        "H18:H23",
    }

    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.wb: Workbook = None
        self._worksheet: Worksheet = None
        self.__init_wb()

        self.s3 = s3
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.photo_result = None
        self.filename = None
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company"
        )
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="L1:N1",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B1",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        self.temp_file = f"{tempfile.mkdtemp()}/"
        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.static_fields = {
            "date_inspection": "K3",
            "id_ccr_antt": "B5",
            "extension": "B6",
            "initial_km": "H5",
            "end_km": "H6",
            "coord_amount_N": "B8",
            "coord_amount_E": "B9",
            "type_manhole": "B10",
            "coord_downstream_N": "I8",
            "coord_downstream_E": "I9",
            "shape": "I10",
            "dimension_amount": "B12",
            "side_amount": "B13",
            "entrance_structure": "B14",
            "material_amount": "B15",
            "conservation_state_amount": "B16",
            "dimension_downstream": "H12",
            "side_downstream": "H13",
            "exit_structure_downstream": "H14",
            "material_downstream": "H15",
            "conservation_state_downstream": "H16",
            "photo_mont_panoramic": "A25:F25",
            "photo_mont_detail": "A27:F27",
            "photo_downstream_panoramic": "G25:N25",
            "photo_downstream_detail": "G27:N27",
            "conservation_check_mon": {
                "OK": "D18",
                "cleaning": "D19",
                "unblocking": "D20",
                "drowned": "D21",
            },
            "conservation_check_jus": {
                "OK": "E18",
                "cleaning": "E19",
                "unblocking": "E20",
                "drowned": "E21",
            },
            "structural_check_mon": {
                "Testa ou ala danificada": "M18",
                "Tubulação danificada": "M19",
                "Caixa danificada": "M20",
                "Erosão": "M21",
                "Fissuras/trincas": "M22",
                "Tampa Danificada/Inexistente": "M23",
            },
            "structural_check_jus": {
                "Testa ou ala danificada": "N18",
                "Tubulação danificada": "N19",
                "Caixa danificada": "N20",
                "Erosão": "N21",
                "Fissuras/trincas": "N22",
                "Tampa Danificada/Inexistente": "N23",
            },
        }

        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

    def __init_wb(self):
        self.wb = load_workbook("./fixtures/reports/ccr_defenses_oac_annex_two.xlsx")
        self._worksheet = self.wb.active

    def create_dict(self, reporting):
        self.photo_result = None
        linked_reports = ReportingInReporting.objects.filter(parent=reporting.uuid)
        result_dict = {}
        if linked_reports:
            for data in linked_reports:
                child_report = data.child
                name_reporting_relation = data.reporting_relation.name
                if (
                    name_reporting_relation == "Jusante"
                    and data.reporting_relation.outward
                ):
                    result_dict["exit_structure_downstream"] = new_get_form_data(
                        child_report, "entryStruc"
                    )
                    result_dict["coord_downstream_N"] = new_get_form_data(
                        child_report, "latitude"
                    )
                    result_dict["coord_downstream_E"] = new_get_form_data(
                        child_report, "longitude"
                    )
                    dimension = ""
                    diameter = new_get_form_data(child_report, "diameter")
                    if diameter is not None:
                        dimension = diameter
                    else:
                        dimensions = []
                        width = new_get_form_data(child_report, "width")
                        height = new_get_form_data(child_report, "height")

                        if height is not None:
                            dimensions.append(str(height))
                        if width is not None:
                            dimensions.append(str(width))
                        if width is not None or height is not None:
                            dimension = "x".join(dimensions)
                    result_dict["dimension_downstream"] = dimension
                    result_dict["side_downstream"] = get_custom_option(
                        child_report, "direction"
                    )
                    result_dict["material_downstream"] = new_get_form_data(
                        child_report, "materialRevMont"
                    )
                    result_dict["conservation_state_downstream"] = new_get_form_data(
                        child_report, "holeClassification"
                    )
                    result_conservation_check = self.__set_conservation_check(
                        child_report, result_dict["conservation_state_downstream"]
                    )
                    result_dict["conservation_check_jus"] = result_conservation_check
                    result_structural_check = self.__set_structural_check(child_report)

                    result_dict["structural_check_jus"] = result_structural_check

                    photos = child_report.form_data.get("photos_mon")
                    if photos:
                        photo_jus = XlsxHandlerReportOACAnnexTwo.__get_photos(photos)
                        result_dict["photo_downstream_panoramic"] = photo_jus[
                            "panorama"
                        ]
                        result_dict["photo_downstream_detail"] = photo_jus["detail"]

                if (
                    name_reporting_relation == "Montante"
                    and data.reporting_relation.outward
                ):

                    result_dict["entrance_structure"] = new_get_form_data(
                        child_report, "entryStruc"
                    )
                    result_dict["conservation_state_amount"] = new_get_form_data(
                        child_report, "holeClassification"
                    )
                    result_dict["coord_amount_N"] = new_get_form_data(
                        child_report, "latitude"
                    )
                    result_dict["coord_amount_E"] = new_get_form_data(
                        child_report, "longitude"
                    )
                    dimension = ""
                    diameter = new_get_form_data(child_report, "diameter")
                    if diameter is not None:
                        dimension = diameter
                    else:
                        dimensions = []
                        width = new_get_form_data(child_report, "width")
                        height = new_get_form_data(child_report, "height")

                        if height is not None:
                            dimensions.append(str(height))
                        if width is not None:
                            dimensions.append(str(width))
                        if width is not None or height is not None:
                            dimension = "x".join(dimensions)
                    result_dict["dimension_amount"] = dimension
                    result_dict["side_amount"] = get_custom_option(
                        child_report, "direction"
                    )
                    result_dict["material_amount"] = new_get_form_data(
                        child_report, "materialRevMont"
                    )
                    result_dict["conservation_state_amount"] = new_get_form_data(
                        child_report, "holeClassification"
                    )

                    result_conservation_check = self.__set_conservation_check(
                        child_report, result_dict["conservation_state_amount"]
                    )
                    result_dict["conservation_check_mon"] = result_conservation_check

                    result_structural_check = self.__set_structural_check(
                        child_report,
                    )
                    result_dict["structural_check_mon"] = result_structural_check
                    photos = child_report.form_data.get("photos_mon")
                    if photos:
                        photo_mon = XlsxHandlerReportOACAnnexTwo.__get_photos(photos)
                        result_dict["photo_mont_panoramic"] = photo_mon["panorama"]
                        result_dict["photo_mont_detail"] = photo_mon["detail"]

        result_dict["id_ccr_antt"] = new_get_form_data(reporting, "idCcrAntt")
        result_dict["date_inspection"] = (
            reporting.executed_at.strftime("%d/%m/%Y") if reporting.executed_at else ""
        )
        photos_fields = {
            key: value for key, value in result_dict.items() if key.startswith("photo")
        }

        if photos_fields:
            for key, photo in photos_fields.items():
                temp_file = tempfile.mkdtemp()
                self.photo_result = None
                has_image = False
                if result_dict[key]:
                    for reporting_file_uuid in result_dict[key]:
                        self.photo_result = result_photos(
                            self.s3,
                            temp_file=temp_file,
                            photo_id=reporting_file_uuid,
                            width=337,
                            height=242,
                            enable_is_shared_antt=True,
                            enable_include_dnit=False,
                        )
                        if self.photo_result:
                            result_dict[key] = self.photo_result[0]
                            has_image = True
                            break
                if not has_image:
                    result_dict[key] = ""

        result_dict["shape"] = new_get_form_data(reporting, "format")

        if result_dict["shape"] == "Tubular":
            result_dict["type_manhole"] = new_get_form_data(
                reporting, "holeKindTubular"
            )
        elif result_dict["shape"] == "Celular":
            result_dict["type_manhole"] = new_get_form_data(
                reporting, "holeKindCelular"
            )
        elif result_dict["shape"] == "Ovóide":
            result_dict["type_manhole"] = new_get_form_data(reporting, "holeKindOvoid")

        result_dict["extension"] = new_get_form_data(reporting, "length")
        result_dict["initial_km"] = format_km(reporting, "km", 3)
        result_dict["end_km"] = format_km(reporting, "end_km", 3)
        result_dict["road_name"] = reporting.road_name
        result_dict["found_at"] = reporting.found_at

        for key, valor in result_dict.items():
            if isinstance(valor, dict):
                for sub_key, sub_value in valor.items():
                    if sub_value is None:
                        valor[sub_key] = ""
            elif valor is None:
                result_dict[key] = ""
        return result_dict

    def __set_conservation_check(self, report, conservation_state):
        conservation_check_dict = dict()
        cleaning = new_get_form_data(report, "cleaningmon")
        unblocking = new_get_form_data(report, "desassoreamentomon")

        drowned = new_get_form_data(report, "desobstrucaomon")
        conservation_check_dict["OK"] = (
            conservation_state if conservation_state == "Bom" else ""
        )
        if conservation_check_dict["OK"] == "Bom":
            conservation_check_dict["cleaning"] = ""
            conservation_check_dict["unblocking"] = ""
            conservation_check_dict["drowned"] = ""

            return conservation_check_dict

        conservation_check_dict["cleaning"] = cleaning
        conservation_check_dict["unblocking"] = unblocking
        conservation_check_dict["drowned"] = drowned
        return conservation_check_dict

    def __set_structural_check(self, report):
        structural_check_dict = dict()
        forehead_damage = new_get_form_data(report, "foreheadDamagemon")
        tube_damage = new_get_form_data(report, "tubeDamagemon")

        box_damage = new_get_form_data(report, "boxDamagemon")
        erosion = new_get_form_data(report, "erosionmon")
        broken = new_get_form_data(report, "brokenmon")

        cover = new_get_form_data(report, "coverDamagemon")

        structural_check_dict["Testa ou ala danificada"] = forehead_damage
        structural_check_dict["Tubulação danificada"] = tube_damage
        structural_check_dict["Caixa danificada"] = box_damage
        structural_check_dict["Erosão"] = erosion
        structural_check_dict["Fissuras/trincas"] = broken
        structural_check_dict["Tampa Danificada/Inexistente"] = cover
        return structural_check_dict

    @classmethod
    def __get_photos(cls, photos):
        dict_photos = {"panorama": "", "detail": ""}
        for photo in photos:
            try:
                dict_photos["panorama"] = (
                    photo["panoramic_photos_mon"]
                    if photo["panoramic_photos_mon"]
                    else ""
                )
            except (IndexError, KeyError):
                dict_photos["panorama"] = ""

            try:
                dict_photos["detail"] = (
                    photo["detail_photos_mon"] if photo["detail_photos_mon"] else ""
                )

            except (IndexError, KeyError):
                dict_photos["detail"] = ""
        return dict_photos

    def __insert_status_values(self, cell, value, key):
        if value:
            self._worksheet[cell] = "X"
            XlsxHandlerReportOACAnnexTwo.__format_fonts(
                cell=self._worksheet[cell], size=10, horizontal="center"
            )
        else:
            self._worksheet[cell] = ""

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
    ) -> None:

        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
        cell.font = Font(name=name, sz=size, bold=bold)

    def __create_filename_and_header(self, road_name, id_ccr_antt):
        self._worksheet.merge_cells("A4:N4")
        id_antt = id_ccr_antt.split(" ")[0] if id_ccr_antt else ""
        road_name = road_name.split(" ")[0] if road_name else ""

        self._worksheet["A4"].value = f"OBRAS DE ARTES CORRENTE.{road_name} {id_antt}"
        XlsxHandlerReportOACAnnexTwo.__format_fonts(
            cell=self._worksheet["B5"], bold=True, size=12
        )

        file_name = "/tmp/{}.xlsx".format(id_ccr_antt.replace("/", "-"))
        return file_name

    def fill_sheet(self, *, data_dict):
        list_files = []
        for key, values_list in data_dict.items():
            count = 0
            files_list = []
            for values in values_list:
                for year_key, list_year_values in values.items():
                    files_list.clear()
                    for values_in_year in list_year_values:
                        for internal_key, value in values_in_year.items():
                            if internal_key in [
                                "conservation_check_mon",
                                "conservation_check_jus",
                                "structural_check_mon",
                                "structural_check_jus",
                            ]:
                                for _key, _value in value.items():
                                    key_value = self.static_fields[internal_key][_key]
                                    self.__insert_status_values(
                                        cell=key_value, value=_value, key=internal_key
                                    )

                            elif (
                                internal_key
                                in ["photo_mont_panoramic", "photo_mont_detail"]
                                and value
                            ):
                                cell = f"{self.static_fields[internal_key]}"
                                insert_picture(
                                    worksheet=self._worksheet,
                                    range_string=cell,
                                    picture=Image(value),
                                    target=self.__sheet_target,
                                )

                            elif (
                                internal_key
                                in [
                                    "photo_downstream_detail",
                                    "photo_downstream_panoramic",
                                ]
                                and value
                            ):
                                cell = f"{self.static_fields[internal_key]}"
                                insert_picture(
                                    worksheet=self._worksheet,
                                    range_string=cell,
                                    picture=Image(value),
                                    target=self.__sheet_target,
                                )
                            else:
                                if internal_key not in ["road_name", "found_at"]:
                                    key_value = self.static_fields[internal_key]
                                    try:
                                        self._worksheet[key_value] = value
                                    except Exception as e:
                                        print(e)
                            self.filename = self.__create_filename_and_header(
                                values_in_year["road_name"],
                                values_in_year["id_ccr_antt"],
                            )
                        insert_logo_and_provider_logo(
                            worksheet=self._worksheet,
                            target=self.__sheet_target,
                            logo_company=self.data_logo_company,
                            provider_logo=self.data_provider_logo,
                        )
                        count += 1

                        if self.__report_format == ReportFormat.PDF:
                            no_fill = PatternFill(fill_type=None)
                            for coord in XlsxHandlerReportOACAnnexTwo.__GREEN_CELLS:
                                for cell in CellRange(coord).cells:
                                    self._worksheet.cell(
                                        cell[0], cell[1]
                                    ).fill = no_fill
                        self.wb.save(self.filename)
                        self.__init_wb()
                        files_list.append(self.filename)

                    if count == len(files_list):
                        files_list = list(set(files_list))

                        if self.__report_format == ReportFormat.PDF:
                            files_list = convert_files_to_pdf(files_list)

                        files_list.sort()
                        temp_dict = tempfile.mkdtemp()
                        testname = f"{key}{year_key}"
                        path_file = f"{temp_dict}/{testname}.zip"

                        with ZipFile(path_file, "w") as zipObj:
                            for file in files_list:
                                zipObj.write(file, file.split("/")[-1])
                        list_files.append(path_file)
                        count = 0
        return list_files

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        for item in data:
            road_name = item.get("road_name", None)
            found_at = item.get("found_at", None)

            if road_name in self.dict_filtered_roads:
                year_dict = next(
                    (
                        d
                        for d in self.dict_filtered_roads[road_name]
                        if found_at.year in d
                    ),
                    None,
                )

                if year_dict:
                    year_dict[found_at.year].append(item)
                else:
                    year_dict = {found_at.year: [item]}
                    self.dict_filtered_roads[road_name].append(year_dict)
            else:
                self.dict_filtered_roads[road_name] = [{found_at.year: [item]}]
        self.dict_filtered_roads = {
            k: v
            for k, v in self.dict_filtered_roads.items()
            if v and any(item for sublist in v for item in sublist)
        }
        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)
        return result_file


class CrrDeepDrainageAnnexTwo(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    @classmethod
    def get_file_name(cls):
        file_name = "Anexo II - Obra de Arte Corrente"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerReportOACAnnexTwo(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_oac_annex_two_async_handler(reporter_dict: dict):
    reporter = CrrDeepDrainageAnnexTwo.from_dict(reporter_dict)
    reporter.export()
