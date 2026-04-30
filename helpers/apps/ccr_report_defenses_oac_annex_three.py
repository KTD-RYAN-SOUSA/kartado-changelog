import tempfile
from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
from typing import List
from zipfile import ZipFile

from django.db.models import Prefetch
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import (
    get_form_array_iterator,
    get_occurrence_kind_reference_options,
    new_get_form_data,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture,
    result_photos,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km

THREADING_LIMIT = 30
MAX_POOL_CONNECTIONS = THREADING_LIMIT + 1


def safe_download_photos_from_list(
    s3, photos_list, width: int = 157, height: int = 127
):
    """
    Download fotos de uma lista de forma thread-safe.
    Itera sobre a lista e retorna a primeira foto válida ou None.

    Este método coloca o loop DENTRO da execução da thread, seguindo o padrão
    recomendado pelo projeto para maximizar o uso do thread pool.
    """
    if not photos_list:
        return None

    for photo in photos_list:
        try:
            photo_id = photo["fotos_detalhe"][0]
            result = result_photos(
                s3=s3,
                temp_file=tempfile.mkdtemp(),
                photo_id=photo_id,
                width=width,
                height=height,
                enable_is_shared_antt=True,
                enable_include_dnit=False,
            )
            if result:
                return result[0]
        except (IndexError, KeyError):
            continue

    return None


class XlsxHandlerReportOACAnnexThree(object):
    def __init__(
        self,
        s3,
        list_uuids: List[str],
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
        anomaly_array: bool = False,
    ) -> None:
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.anomaly_array = anomaly_array
        self.previous_reporting = None
        self.corrective_action = None
        self.previous_photo = None
        self.current_photo = None
        self.wb = load_workbook("./fixtures/reports/ccr_dainage_annex_three.xlsx")
        self.list_uuids = list_uuids
        self.uuid = self.list_uuids[0]
        self._worksheet = self.wb.active

        # ThreadPool seguindo o padrão do projeto
        self.thread_pool = ThreadPoolExecutor(max_workers=THREADING_LIMIT)

        # Otimização: usar prefetch_related para parent e children
        self.reportings = Reporting.objects.filter(uuid=self.uuid).prefetch_related(
            "company",
            "occurrence_type",
            "firm",
            "firm__subcompany",
            "parent",
            Prefetch(
                "parent__children",
                queryset=Reporting.objects.filter(
                    executed_at__isnull=False
                ).prefetch_related("occurrence_type", "company"),
            ),
        )
        first_reporting = self.reportings[0]
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.resume_report_dict = []
        self.static_fields = {
            "id_ccr_annt": "A",
            "road_name": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "type_element": "F",
            "previous_conservation_state": "G",
            "previous_photo": "H",
            "conservation_state": "I",
            "current_photo": "J",
            "corrective_actions": "K",
        }

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

        self.logo_config: dict = dict(
            range_string="J1:K1",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.provider_logo_config: dict = dict(
            range_string="A1:B1",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        self.logo_config["path_image"] = get_logo_file(
            s3=self.s3,
            temp_prefix="/tmp/",
            reporting=self.reportings[0],
        )
        self.provider_logo_config["path_image"] = get_provider_logo_file(
            s3=self.s3,
            temp_prefix="/tmp/",
            reporting=self.reportings[0],
        )

        self.anomaly_options = None

    def get_anomalies(self, reporting: Reporting) -> str:
        """
        Get concatenated anomalies from therapy arrayOfObjects field.
        Iterates through therapy items and concatenates occurrenceType values with "/".
        """
        if self.anomaly_options is None:
            self.anomaly_options = get_occurrence_kind_reference_options(
                reporting, "therapy__occurrenceType"
            )

        anomalies = []
        it = get_form_array_iterator(reporting, "therapy")
        try:
            while True:
                occ_type_value = str(it.get("occurrenceType", raw=True)).strip()
                if occ_type_value:
                    occ_type = self.anomaly_options.get(occ_type_value)
                    if occ_type:
                        anomalies.append(str(occ_type).strip())
                it.inc()
        except Exception:
            pass
        anomalies = set(anomalies)
        return " / ".join(anomalies)

    def __insert_new_rows(self, row: int):
        chars = [chr(ord("A") + i) for i in range(11)]

        for char in chars:
            if char in ["H", "J"]:
                self._worksheet.row_dimensions[row].width = 35
            self._worksheet[f"{char}{row}"].border = self.border
            self._worksheet.row_dimensions[row].height = 100

    def create_dict(self, reporting):
        self.previous_photo = None
        self.current_photo = None
        self.corrective_action = None
        self.previous_reporting = None

        # Futures para downloads paralelos
        previous_photo_future = None
        current_photo_future = None

        result_id_ccr_annt = new_get_form_data(reporting, "idCcrAntt")

        result_type_element = new_get_form_data(reporting, "tipoelemento")

        conservation_state = new_get_form_data(
            reporting,
            "generalConservationState",
        )

        # Otimização: usar parent.children já carregado via prefetch_related
        if reporting.executed_at and reporting.parent:
            # children já está pre-carregado via Prefetch
            related_reportings = list(reporting.parent.children.all())
            if related_reportings:
                # Filtra em Python ao invés de fazer query SQL
                related_reportings = [r for r in related_reportings if r.executed_at]
                related_reportings.sort(key=lambda x: x.executed_at, reverse=True)
                previous_date = reporting.executed_at - timedelta(days=365)

                for r in related_reportings:
                    if (
                        str(r.uuid) != str(reporting.uuid)
                        and r.executed_at < reporting.executed_at
                        and r.executed_at >= previous_date
                        and r.occurrence_type == self.form
                        and r.km == reporting.km
                        and r.end_km == reporting.end_km
                        and r.direction == reporting.direction
                    ):
                        self.previous_reporting = [r]
                        break

        previous_conservation_state = (
            new_get_form_data(
                self.previous_reporting[0],
                "generalConservationState",
            )
            if self.previous_reporting
            else ""
        )

        # Download de fotos em paralelo usando ThreadPool (padrão do projeto)
        # IMPORTANTE: O loop for está DENTRO da função safe_download_photos_from_list
        # Isso permite que múltiplas threads processem diferentes reportings simultaneamente
        # maximizando o uso do thread pool (ao invés de apenas 2 threads por vez)
        if self.previous_reporting:
            previous_photos = self.previous_reporting[0].form_data.get(
                "fotos_relatorio"
            )
            if previous_photos:
                # Submit para thread pool - não bloqueia
                previous_photo_future = self.thread_pool.submit(
                    safe_download_photos_from_list,
                    self.s3,
                    previous_photos,
                    157,
                    127,
                )

        current_photos = reporting.form_data.get("fotos_relatorio")
        if current_photos:
            # Submit para thread pool - não bloqueia
            current_photo_future = self.thread_pool.submit(
                safe_download_photos_from_list,
                self.s3,
                current_photos,
                157,
                127,
            )

        if self.anomaly_array:
            self.corrective_action = self.get_anomalies(reporting)
        elif conservation_state and previous_conservation_state:
            recuperation_60 = new_get_form_data(
                reporting,
                "recuperationSixty",
            )

            recuperation_90 = new_get_form_data(
                reporting,
                "recuperationNinety",
            )
            self.__set_result_corrective_action(
                conservation_state,
                previous_conservation_state,
                recuperation_60,
                recuperation_90,
                self.previous_reporting[0],
            )

        direction = get_custom_option(reporting, "direction")
        road_name = reporting.__dict__.get("road_name", "")
        data = {
            "id_ccr_annt": result_id_ccr_annt,
            "road_name": road_name,
            "initial_km": format_km(reporting, "km", 3),
            "end_km": format_km(reporting, "end_km", 3),
            "direction": direction,
            "type_element": result_type_element,
            "previous_conservation_state": previous_conservation_state,
            "previous_photo": previous_photo_future if previous_photo_future else "",
            "conservation_state": conservation_state,
            "current_photo": current_photo_future if current_photo_future else "",
            "corrective_actions": self.corrective_action,
        }

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def fill_sheet(self, *, data_dict: dict):
        row = 5
        list_files = list()
        for key, values_list in data_dict.items():
            count = 0
            for values in values_list:
                for internal_key, value in values.items():
                    if internal_key in ["previous_photo", "current_photo"] and value:
                        path = value.result()
                        if path:
                            cell = f"{self.static_fields[internal_key]}{row}"
                            insert_picture(
                                self._worksheet,
                                cell,
                                Image(path),
                                self.__sheet_target,
                                border_width=1,
                            )
                    else:
                        self.__insert_new_rows(row=row)
                        col = self.static_fields[internal_key]
                        key_value = f"{col}{row}"
                        self._worksheet[key_value] = value
                        XlsxHandlerReportOACAnnexThree.__format_fonts(
                            horizontal="center",
                            cell=self._worksheet[key_value],
                            size=10,
                            wrapText=internal_key == "corrective_actions",
                        )

                count += 1
                row += 1

            if count == len(values_list):
                insert_logo_and_provider_logo(
                    worksheet=self._worksheet,
                    target=self.__sheet_target,
                    logo_company=self.logo_config,
                    provider_logo=self.provider_logo_config,
                )

                self._worksheet[
                    "A1"
                ] = f"Anexo III - Quadro Comparativo - Drenagem Superficial - {values['road_name']}"
                XlsxHandlerReportOACAnnexThree.__format_fonts(
                    cell=self._worksheet["A1"], size=12, bold=True
                )

                self._worksheet[
                    "A2"
                ] = "COMPARATIVO COM A MONITORAÇÃO ANTERIOR - DRENAGEM SUPERFICIAL"
                XlsxHandlerReportOACAnnexThree.__format_fonts(
                    cell=self._worksheet["A2"], size=10
                )

                file_name = (
                    "Anexo III - Quadro Comparativo - Drenagem Superficial - {}".format(
                        key
                    )
                )
                file_name = clean_latin_string(
                    file_name.replace(".", "").replace("/", "")
                )
                file_path = f"/tmp/{file_name}.xlsx"
                row = 5
                self.wb.save(file_path)
                list_files.append(file_path)
                self.__clear_all_data()

        return list_files

    def __clear_all_data(self):
        self.wb = load_workbook("./fixtures/reports/ccr_dainage_annex_three.xlsx")
        self._worksheet = self.wb.active

    def execute(self):
        # Otimização: usar prefetch_related para todas as relações necessárias
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related(
            "occurrence_type",
            "firm",
            "firm__subcompany",
            "company",
            "parent",
            Prefetch(
                "parent__children",
                queryset=Reporting.objects.filter(
                    executed_at__isnull=False
                ).prefetch_related("occurrence_type", "company"),
            ),
        )
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]

        # Otimização: filtrar ANTES de processar fotos para evitar downloads desnecessários do S3
        filtered_reporting = []
        for reporting in list_reporting:
            conservation_state = new_get_form_data(
                reporting, "generalConservationState"
            )

            # Verificar se tem previous reporting para checar conservation state anterior
            previous_conservation_state = ""
            if reporting.executed_at and reporting.parent:
                related_reportings = list(reporting.parent.children.all())
                if related_reportings:
                    related_reportings = [
                        r for r in related_reportings if r.executed_at
                    ]
                    previous_date = reporting.executed_at - timedelta(days=365)
                    for r in related_reportings:
                        if (
                            str(r.uuid) != str(reporting.uuid)
                            and r.executed_at < reporting.executed_at
                            and r.executed_at >= previous_date
                            and r.occurrence_type == self.form
                            and r.km == reporting.km
                            and r.end_km == reporting.end_km
                            and r.direction == reporting.direction
                        ):
                            previous_conservation_state = new_get_form_data(
                                r, "generalConservationState"
                            )
                            break

            # Só processa se não for "Bom" em ambos
            if conservation_state != "Bom" or previous_conservation_state != "Bom":
                filtered_reporting.append(reporting)

        # Agora sim processa apenas os reportings filtrados
        data = [
            self.create_dict(reporting=reporting) for reporting in filtered_reporting
        ]
        filtered_data = data

        for item in filtered_data:
            road_name = item.get("road_name", None)
            if road_name in self.dict_filtered_roads:
                self.dict_filtered_roads[road_name].append(item)
        self.dict_filtered_roads = {
            k: v for k, v in self.dict_filtered_roads.items() if v
        }

        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)

        # Shutdown do thread pool (padrão do projeto)
        self.thread_pool.shutdown(wait=False)

        return result_file

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        wrapText=False,
    ) -> None:

        cell.alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrap_text=wrapText
        )
        cell.font = Font(name=name, sz=size, bold=bold)

    def __set_result_corrective_action(
        self,
        conservation_state: str,
        previous_conservation_state: str,
        recuperation_60: bool,
        recuperation_90: bool,
        previous_reporting: Reporting,
    ):
        state = (previous_conservation_state, conservation_state)
        if state in [("Regular", "Bom"), ("Precário", "Bom")]:
            self.corrective_action = XlsxHandlerReportOACAnnexThree.__get_diagnoses(
                previous_reporting
            )
        elif state in [("Bom", "Regular"), ("Regular", "Regular")]:
            if recuperation_60:
                self.corrective_action = "Cronograma conforme Anexo V"
            elif recuperation_90:
                self.corrective_action = "Cronograma Conforme Anexo VII"
        elif state in [("Bom", "Precário"), ("Regular", "Precário")]:
            self.corrective_action = "Cronograma Conforme Anexo VII"
        else:
            self.corrective_action = ""

        return self.corrective_action

    @classmethod
    def __get_diagnoses(cls, reporting: Reporting):
        diagnoses = []
        if new_get_form_data(reporting, "reparar", default=False):
            diagnoses.append("Reparar")
        if new_get_form_data(reporting, "limpeza", default=False):
            diagnoses.append("Limpeza")
        if new_get_form_data(reporting, "rocada", default=False):
            diagnoses.append("Roçada")
        if new_get_form_data(reporting, "implantar", default=False):
            diagnoses.append("Implantar")

        return "/".join(diagnoses)


class CrrSurfaceDrainageAnnexThree(CCRReport):
    def __init__(
        self,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
        anomaly_array: bool = False,
    ) -> None:
        super().__init__(uuids, report_format)
        self.anomaly_array = anomaly_array

    @classmethod
    def get_file_name(cls):
        file_name = "Anexo III - Comparativo Drenagem Superficial"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3(max_pool_connections=(MAX_POOL_CONNECTIONS))
        files = XlsxHandlerReportOACAnnexThree(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
            anomaly_array=self.anomaly_array,
        ).execute()

        # Validação: verificar se existem arquivos para criar o ZIP
        if not files:
            raise ValueError(
                "Nenhum arquivo foi gerado. Verifique se existem reportings com estado de conservação diferente de 'Bom'."
            )

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        # Validação: verificar se o arquivo ZIP foi criado
        import os

        if not os.path.exists(result_file):
            raise FileNotFoundError(f"Arquivo ZIP não foi criado: {result_file}")

        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_oac_annex_three_async_handler(reporter_dict: dict):
    reporter = CrrSurfaceDrainageAnnexThree.from_dict(reporter_dict)
    reporter.export()
