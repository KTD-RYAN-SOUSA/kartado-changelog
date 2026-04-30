import re
import shutil
import tempfile
from typing import Any, List
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import (
    get_form_array_iterator,
    new_get_form_data,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_image,
    get_logo_file,
    get_provider_logo_file,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_direction,
    get_km,
    get_lane,
    get_road_name,
)
from helpers.apps.ccr_report_utils.workbook_utils import save_workbook
from helpers.strings import clean_latin_string


class XlsxHandler(object):
    __LOGO_CELL = "B1:G3"
    __PROVIDER_LOGO_CELL = "W1:AC3"

    __REG_SIMPLE_APIS: dict = {
        "length": "F19",
        "sideMont": "V21",
        "zone": "I23",
        "xUtmMon": "R23",
        "yUtmMon": "X23",
        "xUtmJus": "R24",
        "yUtmJus": "X24",
    }
    __REG_X_RELATIONS: dict = {
        "width": ("F20", ["widthMont", "widthJus"]),
    }

    __REG_RELATIONS: dict = {
        "diameter": ("F18", ["diameterMont", "diameterJus"]),
        "ambient": ("V17", ["ambientMon", "ambientJus"]),
        "material": ("V20", ["materialRevMont", "revestMaterial"]),
    }

    __REG_LISTS: dict = {
        "device_type": (
            "F17",
            ["holeKindCelular", "holeKindTubular", "holeKindOvoide"],
        ),
        "esconsity": ("F21", ["esconsity", "esconsityTwo"]),
        "holeKindOther": ("V18", ["holeKindOther"]),
        "holeKindOtherExit": ("V19", ["holeKindOtherExit"]),
    }

    __REG_CELL: dict = {
        "road_name": "G10",
        "direction_lane": "G11",
        "km": "G12",
        "croqui": "A28:AD49",
    }

    __ANALYSIS_BOXES_APIS: dict = {
        "mouthClassification": "D32",
        "mouthObsMon": "B34",
        "drownMouthMon": "B36",
        "solDamageMon": "B38",
        "foreheadDamageTwoMon": "B40",
        "aleDamageMon": "B42",
        "bodyClassification": "M32",
        "bodyAssoread": "K34",
        "drownBody": "K36",
        "damageBody": "K38",
        "desalignmentBody": "K40",
        "jointInfiltrated": "K42",
        "boxClassification": "W32",
        "obsBoxAss": "U34",
        "drownBox": "U36",
        "grailCoverDamage": "U38",
        "coverDamageMissing": "U40",
        "wallDamageMon": "U42",
        "abroadClassification": "D44",
        "denseVegetation": "B46",
        "otherGadgetsObs": "B48",
        "otherGadgetsDamage": "B50",
        "otherGadgetsPav": "B52",
        "otherGadgetsEros": "B54",
    }

    __SOL_BOXES_APIS: dict = {
        "nothingToDo": "B62",
        "executeCleaning": "B64",
        "recoverMouths": "B66",
        "recoverBodies": "B68",
        "recoverOrSubstitute": "B70",
        "recoverWalls": "B72",
        "recoverGadgets": "B74",
        "channelWaterCourse": "Q62",
        "recoverSlope": "Q64",
        "dissipeSpeed": "Q66",
        "stretchHole": "Q68",
        "investigateInfluence": "Q70",
        "locateHole": "Q72",
        "requestStudies": "Q74",
    }

    __INSPEC_SIMPLE_APIS: dict = {
        "preliminary_solution_notes": "E78",
    }

    __INSPEC_CELL: dict = {
        "date": "D9",
        "subcompany": "D10",
        "notes": "E56",
        "generalClassification": "B84",
        "noClassification": "B87",
    }

    __ENTRY_BOXES: dict = {
        "1": "D14",
        "2": "D16",
        "3": "D18",
        "4": "D20",
        "5": "D22",
    }

    __EXIT_BOXES: dict = {
        "1": "S14",
        "2": "S16",
        "3": "S18",
        "4": "S20",
        "5": "S22",
    }
    __PHOTO_CELL: dict = {
        "road_name": "G10",
        "direction_lane": "G11",
        "km": "G12",
    }
    __PHOTO_PICTURES_CELL: dict = [
        "B17:N31",
        "P17:AD31",
        "B36:N50",
        "P36:AD50",
        "B55:N69",
        "P55:AD69",
    ]

    @classmethod
    def __get_joined_form_fields(
        cls,
        reporting: Reporting,
        fields: list,
        separator: str,
        default: str,
        placeholder: str = None,
    ) -> str:
        value_list = []
        for field in fields:
            value = new_get_form_data(reporting, field)
            if value is not None:
                if isinstance(value, list):
                    value_list.extend(value)
                else:
                    value_list.append(value)
            elif placeholder is not None:
                value_list.append(placeholder)

        values_str = default
        if len(value_list) > 0 and not all(
            value == placeholder or (not str(value).strip()) for value in value_list
        ):
            values_str = separator.join(map(str, value_list))
        return values_str

    @classmethod
    def __get_cod_monit(cls, reporting: Reporting) -> str:
        return new_get_form_data(reporting, "codMonit")

    @classmethod
    def __get_direction_lane(cls, reporting: Reporting) -> ReportingFile:
        try:
            direction_lane = f"{get_direction(reporting)} ({get_lane(reporting)})"
            return direction_lane
        except Exception:
            return ""

    @classmethod
    def __get_croqui(cls, s3, dir: str, reporting: Reporting) -> Image:
        image: Image = None
        picture_uuids = new_get_form_data(reporting, "croqui", default=[])
        for uuid in picture_uuids:
            try:
                reporting_file = ReportingFile.objects.filter(uuid=uuid).only("upload")[
                    0
                ]
                image = get_image(s3, dir, uuid, reporting_file)
                if image is not None:
                    break
            except Exception:
                continue
        return image

    @classmethod
    def __get_executed_at(cls, reporting: Reporting) -> str:
        executed_at = "-"
        try:
            executed_at = reporting.executed_at.strftime("%d/%m/%Y")
        except Exception:
            pass
        return executed_at

    @classmethod
    def __get_subcompany(cls, reporting: Reporting) -> str:
        subcompany = "-"
        try:
            subcompany = str(reporting.firm.subcompany.name)
        except Exception:
            pass
        return subcompany

    @classmethod
    def __get_entry(cls, reporting: Reporting) -> str:
        entry = 0
        it = get_form_array_iterator(reporting, "element")
        try:
            while True:
                inspected_element = it.get("inspectedElement", raw=True)
                place = it.get("place", raw=True)
                if inspected_element == "1" and place == "1":
                    if entry != 0:
                        entry = 0
                        break
                    entry = it.get("entry", raw=True)
                it.inc()
        except Exception as e:
            print(e)

        if entry is None:
            entry = 0
        return str(entry).strip()

    @classmethod
    def __get_exit(cls, reporting: Reporting) -> str:
        entry = 0
        it = get_form_array_iterator(reporting, "element") or get_form_array_iterator(
            reporting, "elementArray"
        )
        try:
            while True:
                inspected_element = it.get("inspectedElement", raw=True)
                place = it.get("place", raw=True)
                if inspected_element == "1" and place == "2":
                    if entry != 0:
                        entry = 0
                        break
                    entry = it.get("entry", raw=True)
                it.inc()
        except Exception as e:
            print(e)

        if entry is None:
            entry = 0
        return str(entry).strip()

    @classmethod
    def __get_analysis_box(cls, api_name: str, reporting: Reporting) -> str:
        box = new_get_form_data(reporting, api_name, default="-")
        return box if isinstance(box, int) else "-"

    @classmethod
    def __get_notes(cls, reporting: Reporting) -> str:
        notes: str = "-"
        notes_list = []
        it = get_form_array_iterator(reporting, "element") or get_form_array_iterator(
            reporting, "elementArray"
        )
        try:
            while True:
                note = it.get("notes")
                if note is not None:
                    notes_list.append(note)
                it.inc()
        except Exception as e:
            print(e)

        if len(notes_list) > 0:
            notes = "/".join(notes_list)
        return notes

    @classmethod
    def __get_solicitation_box(cls, api_name: str, reporting: Reporting) -> str:
        box = new_get_form_data(reporting, api_name, default=False)
        return "x" if box else ""

    @classmethod
    def __get_general_classification(cls, reporting: Reporting) -> Any:
        general_classification = new_get_form_data(
            reporting, "generalClassification", default="-"
        )
        try:
            general_classification = int(general_classification)
            if general_classification < 1 or general_classification > 3:
                general_classification = "-"
        except ValueError:
            general_classification = "-"
        except Exception as e:
            print(e)

        return general_classification

    @classmethod
    def __get_pictures_reporting_files(
        cls, reporting: Reporting
    ) -> List[ReportingFile]:
        croqui_values = new_get_form_data(reporting, "croqui", default=[])
        croqui_uuids: List[UUID] = []
        for croqui_value in croqui_values:
            try:
                croqui_uuids.append(UUID(croqui_value))
            except Exception:
                continue

        reporting_files = list(
            ReportingFile.objects.filter(reporting__uuid=reporting.uuid)
            .exclude(uuid__in=croqui_uuids)
            .only("uuid", "upload", "uploaded_at", "datetime", "description")
            .order_by("datetime", "uploaded_at")
        )
        return reporting_files

    @classmethod
    def __get_picture_name_prefix(cls, reporting: Reporting) -> str:
        executed_at_year: str = None
        try:
            executed_at_year = reporting.executed_at.strftime("%Y")
        except Exception:
            executed_at_year = "0000"
        road_name_num = re.split(r"-|\s", get_road_name(reporting))[1]
        km = get_km(reporting)
        return f"OAC{executed_at_year}{road_name_num}k{km[:3]}{km[4:]}F"

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.__sheet_target = sheet_target
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.occurrence_type = Reporting.objects.get(uuid=list_uuids[0]).occurrence_type

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __fill_registration(self, worksheet: Worksheet, reporting: Reporting) -> None:
        try:
            logo = Image(get_logo_file(self.s3, f"{self.temp_dir}/reg_logo", reporting))
            insert_picture_2(
                worksheet,
                XlsxHandler.__LOGO_CELL,
                logo,
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalLeft,
            )
        except Exception as e:
            print(e)
        try:
            provider_logo = Image(
                get_provider_logo_file(
                    self.s3, f"{self.temp_dir}/reg_prov_logo", reporting
                )
            )
            insert_picture_2(
                worksheet,
                XlsxHandler.__PROVIDER_LOGO_CELL,
                provider_logo,
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalRight,
            )
        except Exception as e:
            print(e)
        worksheet.title = f"OAC {get_km(reporting)} - Ficha1"

        worksheet[XlsxHandler.__REG_CELL["road_name"]] = get_road_name(reporting)
        worksheet[
            XlsxHandler.__REG_CELL["direction_lane"]
        ] = XlsxHandler.__get_direction_lane(reporting)
        worksheet[XlsxHandler.__REG_CELL["km"]] = get_km(reporting)

        for api_name in XlsxHandler.__REG_SIMPLE_APIS:
            worksheet[XlsxHandler.__REG_SIMPLE_APIS[api_name]] = new_get_form_data(
                reporting, api_name, default="-"
            )

        for field in XlsxHandler.__REG_LISTS:
            cell, api_names = XlsxHandler.__REG_LISTS[field]
            worksheet[cell] = XlsxHandler.__get_joined_form_fields(
                reporting, api_names, ", ", "-"
            )

        for field in XlsxHandler.__REG_RELATIONS:
            cell, api_names = XlsxHandler.__REG_RELATIONS[field]
            worksheet[cell] = XlsxHandler.__get_joined_form_fields(
                reporting, api_names, " / ", "-", "-"
            )

        for field in XlsxHandler.__REG_X_RELATIONS:
            cell, api_names = XlsxHandler.__REG_X_RELATIONS[field]
            worksheet[cell] = XlsxHandler.__get_joined_form_fields(
                reporting, api_names, " x ", "-", "-"
            )

        croqui_image: Image = XlsxHandler.__get_croqui(
            self.s3, self.temp_dir, reporting
        )
        if croqui_image is not None:
            insert_picture_2(
                worksheet,
                XlsxHandler.__REG_CELL["croqui"],
                croqui_image,
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalCentered,
            )

    def __fill_inspection(
        self, workbook: Workbook, worksheet: Worksheet, reporting: Reporting
    ) -> None:
        try:
            logo = Image(
                get_logo_file(self.s3, f"{self.temp_dir}/insp_logo", reporting)
            )
            insert_picture_2(
                worksheet,
                XlsxHandler.__LOGO_CELL,
                logo,
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalLeft,
            )
        except Exception as e:
            print(e)
        try:
            provider_logo = Image(
                get_provider_logo_file(
                    self.s3, f"{self.temp_dir}/insp_prov_logo", reporting
                )
            )
            insert_picture_2(
                worksheet,
                XlsxHandler.__PROVIDER_LOGO_CELL,
                provider_logo,
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalRight,
            )
        except Exception as e:
            print(e)
        worksheet.title = f"OAC {get_km(reporting)} - Ficha2"

        worksheet[XlsxHandler.__INSPEC_CELL["date"]] = XlsxHandler.__get_executed_at(
            reporting
        )
        worksheet[
            XlsxHandler.__INSPEC_CELL["subcompany"]
        ] = XlsxHandler.__get_subcompany(reporting)

        entry = XlsxHandler.__get_entry(reporting)
        if entry in XlsxHandler.__ENTRY_BOXES:
            worksheet[XlsxHandler.__ENTRY_BOXES[str(entry)]] = "x"

        exit = XlsxHandler.__get_exit(reporting)
        if exit in XlsxHandler.__EXIT_BOXES:
            worksheet[XlsxHandler.__EXIT_BOXES[str(exit)]] = "x"

        for api_name in XlsxHandler.__INSPEC_SIMPLE_APIS:
            worksheet[XlsxHandler.__INSPEC_SIMPLE_APIS[api_name]] = new_get_form_data(
                reporting, api_name, default="-"
            )

        for api_name in XlsxHandler.__ANALYSIS_BOXES_APIS:
            worksheet[
                XlsxHandler.__ANALYSIS_BOXES_APIS[api_name]
            ] = XlsxHandler.__get_analysis_box(api_name, reporting)
        worksheet[XlsxHandler.__INSPEC_CELL["notes"]] = XlsxHandler.__get_notes(
            reporting
        )

        for api_name in XlsxHandler.__SOL_BOXES_APIS:
            worksheet[
                XlsxHandler.__SOL_BOXES_APIS[api_name]
            ] = XlsxHandler.__get_solicitation_box(api_name, reporting)

        general_classification = XlsxHandler.__get_general_classification(reporting)
        worksheet[
            XlsxHandler.__INSPEC_CELL["generalClassification"]
        ] = general_classification

        no_classification = "x" if general_classification == "-" else "-"
        worksheet[XlsxHandler.__INSPEC_CELL["noClassification"]] = no_classification

    def __fill_pictures(self, workbook: Workbook, reporting: Reporting) -> None:
        fields = reporting.occurrence_type.form_fields.get("fields")
        template_worksheet = workbook[workbook.sheetnames[2]]
        template_worksheet[XlsxHandler.__PHOTO_CELL["road_name"]] = get_road_name(
            reporting
        )
        template_worksheet[
            XlsxHandler.__PHOTO_CELL["direction_lane"]
        ] = XlsxHandler.__get_direction_lane(reporting)
        template_worksheet[XlsxHandler.__PHOTO_CELL["km"]] = get_km(reporting)
        logo_file: str = None
        provider_logo_file: str = None
        try:
            logo_file = get_logo_file(self.s3, f"{self.temp_dir}/pic_logo", reporting)
            insert_picture_2(
                template_worksheet,
                XlsxHandler.__LOGO_CELL,
                Image(logo_file),
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalLeft,
            )
        except Exception as e:
            print(e)
        try:
            provider_logo_file = get_provider_logo_file(
                self.s3, f"{self.temp_dir}/pic_prov_logo", reporting
            )
            insert_picture_2(
                template_worksheet,
                XlsxHandler.__PROVIDER_LOGO_CELL,
                Image(provider_logo_file),
                self.__sheet_target,
                border_width=(1, 1, 1, 1),
                resize_method=ResizeMethod.ProportionalRight,
            )
        except Exception as e:
            print(e)
        picture_name_prefix = XlsxHandler.__get_picture_name_prefix(reporting)

        fields_element = None
        for _ in fields:
            if _.get("apiName") == "element" or _.get("apiName") == "elementArray":
                fields_element = _.get("innerFields")
                break

        def get_value_in_element(fields_element, api_name, value):
            for field in fields_element:
                if field.get("apiName") == api_name:
                    for _ in field.get("selectOptions", {}).get("options", []):
                        if _.get("value") == value:
                            return _.get("name")

        data_photos = []
        element = (
            reporting.form_data.get("element")
            or reporting.form_data.get("elementArray")
            or reporting.form_data.get("element_array")
            or []
        )
        temp_element = []
        if isinstance(element, list):
            for el in element:
                try:
                    int(el["order_of_element"])
                except Exception:
                    continue
                temp_element.append(el)

            element = temp_element
            element.sort(key=lambda x: int(x["order_of_element"]))
            for el in element:
                photo_pks = el.get("panorama_picture")
                if not photo_pks:
                    continue

                reporting_file = ReportingFile.objects.filter(
                    pk__in=photo_pks
                ).order_by("datetime", "uploaded_at")
                if not reporting_file:
                    continue
                for rf in reporting_file:
                    inspectedElement = get_value_in_element(
                        fields_element,
                        "inspectedElement",
                        el.get("inspected_element"),
                    )
                    place = get_value_in_element(
                        fields_element, "place", el.get("place")
                    )
                    order_of_element = el.get("order_of_element")
                    data_photos.append(
                        {
                            "description": f"{inspectedElement},{place},{order_of_element} - {rf.description}",
                            "reporting_file": rf,
                        }
                    )

        picture_counter: int = 0
        sheet_counter: int = 0
        curr_worksheet: Worksheet = None
        for obj in data_photos:
            picture_cell = XlsxHandler.__PHOTO_PICTURES_CELL[picture_counter % 6]
            min_col, min_row, _, _ = range_boundaries(picture_cell)
            if picture_counter % 6 == 0:
                sheet_counter += 1
                curr_worksheet = workbook.copy_worksheet(template_worksheet)
                curr_worksheet.title = f"OAC {get_km(reporting)} - Foto{sheet_counter}"
                try:
                    insert_picture_2(
                        curr_worksheet,
                        XlsxHandler.__LOGO_CELL,
                        Image(logo_file),
                        self.__sheet_target,
                        border_width=(1, 1, 1, 1),
                        resize_method=ResizeMethod.ProportionalLeft,
                    )
                except Exception as e:
                    print(e)
                try:
                    insert_picture_2(
                        curr_worksheet,
                        XlsxHandler.__PROVIDER_LOGO_CELL,
                        Image(provider_logo_file),
                        self.__sheet_target,
                        border_width=(1, 1, 1, 1),
                        resize_method=ResizeMethod.ProportionalRight,
                    )
                except Exception as e:
                    print(e)

            try:
                reporting_file: ReportingFile = obj.get("reporting_file")
                image = get_image(
                    self.s3, self.temp_dir, reporting_file.uuid, reporting_file
                )
                insert_picture_2(
                    curr_worksheet,
                    picture_cell,
                    image,
                    self.__sheet_target,
                    border_width=(1, 1, 1, 1),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception as e:
                print(e)
                if picture_counter % 6 == 0:
                    sheet_counter -= 1
                    workbook.remove(curr_worksheet)
                continue

            picture_counter += 1
            desc_col = min_col + (5 if picture_counter % 2 == 0 else 4)
            name_cell = f"{get_column_letter(min_col)}{min_row+15}"
            desc_cell = f"{get_column_letter(desc_col)}{min_row+15}"
            try:
                curr_worksheet[
                    name_cell
                ] = f"{picture_name_prefix}{picture_counter:03d}"
                curr_worksheet[desc_cell] = obj.get("description", "")
            except Exception:
                pass

            if picture_counter == 150:
                break

        if picture_counter > 0:
            workbook.remove(template_worksheet)
        else:
            template_worksheet.title = f"OAC {get_km(reporting)} - Foto1"

    def __create_workbook_file(self, reporting: Reporting) -> str:
        workbook = load_workbook("./fixtures/reports/ccr_report_artesp_oac.xlsx")

        registration_worksheet = workbook[workbook.sheetnames[0]]
        self.__fill_registration(registration_worksheet, reporting)

        inspection_worksheet = workbook[workbook.sheetnames[1]]
        self.__fill_inspection(workbook, inspection_worksheet, reporting)

        self.__fill_pictures(workbook, reporting)

        workbook.active = registration_worksheet

        cod_monit = XlsxHandler.__get_cod_monit(reporting)
        workbook_name = ""
        if cod_monit is not None:
            workbook_name = cod_monit
        else:
            workbook_name = reporting.number
        workbook_file = save_workbook(workbook_name, workbook)

        return workbook_file

    def execute(self) -> List[str]:
        query_set = (
            Reporting.objects.filter(
                occurrence_type=self.occurrence_type,
                uuid__in=self.list_uuids,
            )
            .prefetch_related("firm", "firm__subcompany", "company")
            .only(
                "uuid",
                "road_name",
                "km",
                "direction",
                "lane",
                "form_data",
                "executed_at",
                "firm__subcompany__name",
                "company__uuid",
            )
        )
        return [self.__create_workbook_file(reporting) for reporting in query_set]


class CCRArtespOAC(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self) -> str:
        file_name: str = None
        if len(self.uuids) > 1:
            road_names = list(
                Reporting.objects.filter(uuid__in=self.uuids)
                .only("uuid", "road_name")
                .order_by("road_name")
                .distinct("road_name")
                .values_list("road_name", flat=True)
            )
            road_names.sort()
            file_name = "Fichas de OAC ARTESP - {}".format("_".join(road_names))
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
            file_name = f"{file_name}.zip"
        else:
            reporting = Reporting.objects.filter(uuid=self.uuids[0]).only(
                "number", "form_data"
            )[0]
            file_name = new_get_form_data(
                reporting, "codMonit", default=reporting.number
            )
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
            extension = "xlsx" if self.report_format() == ReportFormat.XLSX else "pdf"
            file_name = f"{file_name}.{extension}"
        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandler(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        else:
            result_file = files[0]

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_artesp_oac_async_handler(
    reporter_dict: dict,
):
    reporter = CCRArtespOAC.from_dict(reporter_dict)
    reporter.export()
