import shutil
import tempfile
import time
from copy import copy
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import (
    new_get_form_data,
    new_get_form_data_selected_option,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_reporting_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_picture,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option, get_km
from helpers.strings import clean_latin_string, get_obj_from_path


class XlsxHandler:
    _LOGO_CELL = ["I1:J1", "I45:J48"]
    _PROVIDER_LOGO_CELL = ["B1", "B45:B48"]

    def __init__(
        self,
        s3,
        uuid: str,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.uuid = uuid
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.temp_file = tempfile.mkdtemp()
        self.__xlsx_file = "./fixtures/reports/ccr_electrical.xlsx"
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Gabarito - Ficha de Eletrica"]

        self._worksheet.column_dimensions[
            "G"
        ].width = 6.7109375  # Column with null width. Requires further investigation

        self.reporting = Reporting.objects.get(uuid=uuid)
        self.__static_fields = {
            "identificacaoinstalacoes": "C3",
            "tipoedificacao": "H3",
            "road_name": "C4",
            "km": "E4",
            "direction": "I4",
            "municipio": "C5",
            "latitude": "I5",
            "longitude": "I6",
            "found_at": "C6",
            "conservacao": {"1": "D8", "2": "J8", "3": "H8"},
            "building_description": "C11",
            "description_electrical_installations": "C13",
            "concessionariaenergia": "C15",
            "atendimentoabnt": "E19",
            "fase_rn": "D24",
            "fases_n": "D25",
            "fase_tn": "D26",
            "voltage_in_power_phases_norm": {
                "within_the_norm": "F24",
                "outside_the_norm": "F25",
            },
            "fase_rs": "D29",
            "fases_t": "D30",
            "fase_tr": "D31",
            "obs_fase": "F28",
            "medicao_spda": "D35",
            "spda_measurement_norm": {
                "within_the_norm": "F35",
                "outside_the_norm": "F36",
            },
            "obs_spda": "F39",
        }
        self.__photo_report_len = 0

    @classmethod
    def _insert_logos(
        cls,
        s3,
        temp_dir: str,
        sheet_target: SheetTarget,
        worksheet: Worksheet,
        reporting: Reporting,
    ) -> None:
        logo = get_logo_file(s3, temp_dir, reporting)
        provider_logo = get_provider_logo_file(s3, temp_dir, reporting)

        for logo_cell in cls._LOGO_CELL:
            try:
                insert_picture_2(
                    worksheet,
                    logo_cell,
                    Image(logo),
                    sheet_target,
                    border_width=(5, 5, 5, 5),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass
        for provider_logo_cell in cls._PROVIDER_LOGO_CELL:
            try:
                insert_picture_2(
                    worksheet,
                    provider_logo_cell,
                    Image(provider_logo),
                    sheet_target,
                    border_width=(5, 5, 5, 5),
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass

    def __arrangements(self, reporting_files: list) -> list:
        arrangements = [num for num in range(0, int(len(reporting_files) / 2) + 1)]
        start = 0
        new_arrangements = []
        for arrangement in arrangements:
            if start > len(reporting_files) - 1:
                break
            else:
                if start + 1 < len(reporting_files):
                    new_arrangements.append(
                        (reporting_files[start], reporting_files[start + 1])
                    )
                else:
                    new_arrangements.append((reporting_files[start],))
            start += 2
        return new_arrangements

    def __change_color_cols(
        self,
        start_col: int,
        end_col: int,
        start_row: int,
        end_row: int,
        color: str,
        remove: list,
    ):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(start_col, end_col):
            if col not in remove:
                for row in range(start_row, end_row):
                    cell = self._worksheet.cell(row=row, column=col)
                    cell.fill = fill

    def __change_border(
        self,
        start_col: int,
        end_col: int,
        start_row: int,
        end_row: int,
        sides: list,
        remove: dict = {"row": [], "col": []},
    ):
        border = Border(
            left=Side(style="thin") if "left" in sides else None,
            right=Side(style="thin") if "right" in sides else None,
            top=Side(style="thin") if "top" in sides else None,
            bottom=Side(style="thin") if "bottom" in sides else None,
        )
        for col in range(start_col, end_col):
            if col not in remove["col"]:
                for row in range(start_row, end_row):
                    if row not in remove["row"]:
                        cell = self._worksheet.cell(row=row, column=col)
                        cell.border = border

    def __insert_photographic_reports(self, reporting_files: list):
        rows = {
            "white-6px": {"n": 1, "s": 4.5},
            "red-15px": {"n": 1, "s": 12},
            "white-15px": {"n": 9, "s": 15},
        }
        constant = var = 51
        arrangements = self.__arrangements(
            reporting_files=[_dict["path"] for _dict in reporting_files]
        )
        arrangements = arrangements[:2]
        update = {}
        for _item in arrangements:
            for item, value in rows.items():
                for n_row in range(value["n"]):
                    self._worksheet.insert_rows(constant)
                    self._worksheet.row_dimensions[constant + 1].height = value["s"]
                    new_cell = self._worksheet[f"K{constant}"]
                    cell_above = self._worksheet[f"K{constant-1}"]
                    new_cell.border = copy(cell_above.border)
                    if "red" in item:
                        self.__change_color_cols(
                            start_col=1,
                            end_col=11,
                            start_row=constant,
                            end_row=constant + 1,
                            color="73031B",
                            remove=[1, 5],
                        )
                        self.__change_border(
                            start_col=2,
                            end_col=11,
                            start_row=constant,
                            end_row=constant + 1,
                            sides=["left", "top", "bottom"],
                            remove={"col": [4, 5, 7, 8, 9, 10, 11], "row": []},
                        )
                        self.__change_border(
                            start_col=3,
                            end_col=10,
                            start_row=constant,
                            end_row=constant + 1,
                            sides=["top", "bottom"],
                            remove={"col": [3, 4, 5, 6, 11], "row": []},
                        )
                        self.__change_border(
                            start_col=4,
                            end_col=11,
                            start_row=constant,
                            end_row=constant + 1,
                            sides=["right", "top", "bottom"],
                            remove={"col": [5, 6, 7, 8, 9, 11], "row": []},
                        )
            self._worksheet.merge_cells(f"B{var+9}:D{var+9}")
            self._worksheet.merge_cells(f"F{var+9}:J{var+9}")
            self._worksheet.merge_cells(f"B{var}:D{var+8}")
            self._worksheet.merge_cells(f"F{var}:J{var+8}")
            self._worksheet.row_dimensions[var - 1].height = 4.5
            left = True
            self.__change_border(
                start_col=2,
                end_col=11,
                start_row=var,
                end_row=var + 1,
                sides=["left", "top"],
                remove={"col": [4, 5, 7, 8, 9, 10, 11], "row": []},
            )
            self.__change_border(
                start_col=2,
                end_col=11,
                start_row=var + 1,
                end_row=var + 9,
                sides=["left"],
                remove={"col": [3, 5, 7, 8, 9, 10, 11], "row": []},
            )
            self.__change_border(
                start_col=3,
                end_col=10,
                start_row=var,
                end_row=var + 1,
                sides=["top"],
                remove={"col": [3, 4, 5, 6, 11], "row": []},
            )
            self.__change_border(
                start_col=4,
                end_col=11,
                start_row=var + 1,
                end_row=var + 9,
                sides=["right"],
                remove={"col": [5, 6, 7, 8, 9, 11], "row": []},
            )
            self.__change_border(
                start_col=4,
                end_col=11,
                start_row=var,
                end_row=var + 1,
                sides=["right", "top"],
                remove={"col": [5, 6, 7, 8, 9, 11], "row": []},
            )
            for image in _item:
                col = "B" if left else "F"
                end_col = "D" if left else "J"
                range_str = f"{col}{var}:{end_col}{var+8}"
                try:
                    insert_picture(
                        self._worksheet, range_str, Image(image), self.__sheet_target
                    )
                except Exception as e:
                    print(e)
                left = False
                text = [
                    text["description"]
                    for text in reporting_files
                    if text["path"] == image
                ][0]
                update[f"{col}{var+9}"] = text
                self._worksheet.row_dimensions[var + 9].height = 12
                self.__photo_report_len += 1
            var += 11
        for key, text in update.items():
            self._worksheet[key] = text
            self._worksheet[key].font = Font(color="FFFFFF")
            self._worksheet[key].alignment = Alignment(
                horizontal="center", vertical="center"
            )
        self._worksheet.row_dimensions[var - 1].height = 4.5
        self._worksheet[f"K{var-1}"].border = copy(self._worksheet[f"K{var-2}"].border)
        self._worksheet.row_dimensions[var].height = 12.75
        self._worksheet[f"K{var}"].border = copy(self._worksheet[f"K{var-1}"].border)
        return var

    def __insert_value_change_color(self, cell, value):
        if value is True:
            self._worksheet[cell] = "X"
            self._worksheet[cell].fill = PatternFill(
                start_color="808080", end_color="808080", fill_type="solid"
            )
            self._worksheet[cell].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            self._worksheet[cell].font = Font(bold=True)

    def insert_page_breaks(self, remaining_anomalies):
        box_rows_amount = 13
        initial_row = 64
        if self.__photo_report_len > 2:
            initial_row = 75

        row_height = self._worksheet.row_dimensions[initial_row].ht
        for i in range(remaining_anomalies):
            base = initial_row + (i * box_rows_amount)
            self._worksheet.row_dimensions[base].ht = row_height
            self._worksheet.row_dimensions[base + 2].hidden = True
            self._worksheet.row_dimensions[base + 9].ht /= 3 / 2

        self._worksheet.row_dimensions[initial_row - 2].ht /= 2
        hidden_row = initial_row + 2 * box_rows_amount
        self._worksheet.row_dimensions[hidden_row].hidden = True
        self._worksheet.row_dimensions[hidden_row + 1].hidden = True
        for i in range(2, remaining_anomalies, 4):
            break_row = initial_row + i * box_rows_amount
            self._worksheet.row_breaks.append(Break(break_row))

        hidden_row = initial_row + remaining_anomalies * box_rows_amount
        self._worksheet.row_dimensions[hidden_row].ht = row_height
        self._worksheet.row_dimensions[hidden_row + 5].hidden = True
        self._worksheet.row_dimensions[hidden_row + 6].ht = row_height

    def __anomalies_found(self, var: int, anomalies_found: list, reporting_files: list):
        _small_rows = [4, 17]
        _init = var
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for anomalies in anomalies_found:
            for _row in range(1, 14):
                self._worksheet.insert_rows(var)
                self._worksheet.row_dimensions[var + 1].height = (
                    4.5 if _row in _small_rows else 15
                )
                new_cell = self._worksheet[f"K{var}"]
                cell_above = self._worksheet[f"K{var-1}"]
                new_cell.border = copy(cell_above.border)
                var += 1
            _merge = {
                2: f"C{var-10}:D{var-10}",
                3: f"F{var-10}:G{var-10}",
                4: f"H{var-10}:J{var-10}",
                5: f"F{var-8}:J{var+1}",
                6: f"B{var-1}:B{var+1}",
                7: f"C{var-1}:D{var+1}",
            }
            for value in _merge.values():
                self._worksheet.merge_cells(value)
            _texts = {
                3: {
                    1: {
                        "text": "Local:",
                        "col": f"B{var-10}",
                        "horizontal": "right",
                        "color": "73031B",
                    },
                    2: {
                        "text": "Tipo",
                        "col": f"F{var-10}",
                        "horizontal": "center",
                        "color": "73031B",
                    },
                },
                5: {
                    1: {
                        "text": "Anomalias:",
                        "col": f"B{var-8}",
                        "horizontal": "right",
                        "color": "73031B",
                    },
                    2: {
                        "text": "Sem Identificação",
                        "col": f"C{var-8}",
                        "horizontal": "right",
                        "color": "",
                    },
                },
                6: {
                    1: {
                        "text": "Sem Acabamento",
                        "col": f"C{var-7}",
                        "horizontal": "right",
                        "color": "",
                    }
                },
                7: {
                    1: {
                        "text": "Reparar",
                        "col": f"C{var-6}",
                        "horizontal": "right",
                        "color": "",
                    }
                },
                8: {
                    1: {
                        "text": "Danificado",
                        "col": f"C{var-5}",
                        "horizontal": "right",
                        "color": "",
                    }
                },
                9: {
                    1: {
                        "text": f'Outros ({anomalies.get("outrodetalhe") if anomalies.get("outrodetalhe") else "____" })',
                        "col": f"C{var-4}",
                        "horizontal": "right",
                        "color": "",
                    }
                },
                12: {
                    1: {
                        "text": "Descrição da\n anomalia:",
                        "col": f"B{var-1}",
                        "horizontal": "center",
                        "color": "73031B",
                    }
                },
            }
            for key, value in _texts.items():
                for intenr_value in value.values():
                    merged_cell = self._worksheet[intenr_value["col"]]
                    merged_cell.value = intenr_value["text"]
                    merged_cell.alignment = Alignment(
                        horizontal=intenr_value["horizontal"], vertical="center"
                    )
                    merged_cell.alignment = merged_cell.alignment.copy(wrapText=True)
                    if intenr_value["color"]:
                        merged_cell.font = Font(color="FFFFFF")
                        fill = PatternFill(
                            start_color=intenr_value["color"],
                            end_color=intenr_value["color"],
                            fill_type="solid",
                        )
                        cell = self._worksheet[intenr_value["col"]]
                        cell.fill = fill
            self.__change_border(
                start_col=2,
                end_col=5,
                start_row=var - 10,
                end_row=var - 9,
                sides=["left", "right", "top", "bottom"],
            )
            self.__change_border(
                start_col=6,
                end_col=11,
                start_row=var - 10,
                end_row=var - 9,
                sides=["left", "right", "top", "bottom"],
            )
            self._worksheet[f"B{var-8}"].border = border
            self.__change_border(
                start_col=4,
                end_col=5,
                start_row=var - 8,
                end_row=var - 3,
                sides=["left", "right", "top", "bottom"],
            )
            self.__change_border(
                start_col=6,
                end_col=11,
                start_row=var - 8,
                end_row=var + 3,
                sides=["left", "right", "top", "bottom"],
            )
            self.__change_border(
                start_col=2,
                end_col=5,
                start_row=var - 1,
                end_row=var + 3,
                sides=["left", "right", "top", "bottom"],
            )

            self._worksheet[f"C{var-10}"] = (
                anomalies["instalacao"] if anomalies.get("instalacao") else ""
            )
            if anomalies.get("anomalia_instalacoes"):
                anomalia_instalacoes = new_get_form_data_selected_option(
                    self.reporting,
                    anomalies.get("anomalia_instalacoes", ""),
                    "therapy__anomaliaInstalacoes",
                    "",
                )
                self._worksheet[f"H{var-10}"] = anomalia_instalacoes
            self.__insert_value_change_color(
                cell=f"D{var-8}", value=anomalies.get("semidentificacao")
            )
            self.__insert_value_change_color(
                cell=f"D{var-7}", value=anomalies.get("sem_acabamento")
            )
            self.__insert_value_change_color(
                cell=f"D{var-6}", value=anomalies.get("reparar")
            )
            self.__insert_value_change_color(
                cell=f"D{var-5}", value=anomalies.get("danificado")
            )
            self.__insert_value_change_color(
                cell=f"D{var-4}", value=anomalies.get("outros")
            )
            self._worksheet[f"C{var-1}"] = (
                anomalies["description"] if anomalies.get("description") else ""
            )
            if anomalies.get("treatment_images"):
                list_image = [
                    obj["path"]
                    for obj in reporting_files
                    if obj["uuid"] == anomalies["treatment_images"][0]
                ]
                try:
                    image = list_image[0]
                    range_str = f"F{var-8}:J{var-8+9}"
                    insert_picture(
                        self._worksheet, range_str, Image(image), self.__sheet_target
                    )
                except Exception as e:
                    print(e)
            self._worksheet[f"C{var-1}"].alignment = Alignment(vertical="center")
            self._worksheet[f"C{var-1}"].alignment = self._worksheet[
                f"C{var-1}"
            ].alignment.copy(wrapText=True)
        var += 1
        for n_rows in range(1, 7):
            cell_above = self._worksheet[f"K{var-1}"]
            self._worksheet.insert_rows(var)
            self._worksheet[f"K{var}"].border = copy(cell_above.border)
            self._worksheet[f"K{var+1}"].border = copy(cell_above.border)
            var += 1
        self.__change_border(
            start_col=2,
            end_col=11,
            start_row=var - 4,
            end_row=var,
            sides=["left", "right", "top", "bottom"],
        )
        self._worksheet.merge_cells(f"B{var-4}:B{var}")
        self._worksheet.merge_cells(f"C{var-4}:J{var}")
        self.__change_color_cols(
            start_col=2,
            end_col=3,
            start_row=var - 4,
            end_row=var,
            color="73031B",
            remove=[],
        )

        self._worksheet[f"B{var-4}"].value = "Conclusão:"
        self._worksheet[f"B{var-4}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self._worksheet[f"B{var-4}"].font = Font(color="FFFFFF")

        for col in range(1, 12):
            cell = self._worksheet.cell(row=var + 1, column=col)
            cell.border = Border(bottom=Side(style="medium"))
        self._worksheet.cell(row=var + 1, column=11).border = Border(
            bottom=Side(style="medium"), right=Side(style="medium")
        )
        self._worksheet.cell(row=var + 1, column=1).border = Border(
            bottom=Side(style="medium"), left=Side(style="medium")
        )
        count = var - 6
        for _ in range(1, len(anomalies_found) + 1):
            for col in [2, 3, 4, 6, 7, 8, 9, 10]:
                cell = self._worksheet.cell(row=count, column=col)
                cell.border = border
                cell_sec = self._worksheet.cell(row=count - 1, column=col)
                cell_sec.border = border
            count -= 13

        self._worksheet.merge_cells(f"B{_init+1}:J{_init+1}")
        merged_cell = self._worksheet[f"B{_init+1}"]
        merged_cell.value = "Anomalias Encontradas"
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.font = Font(color="FFFFFF")
        merged_cell.fill = PatternFill(
            start_color="73031B",
            end_color="73031B",
            fill_type="solid",
        )
        self.__change_border(
            start_col=2,
            end_col=11,
            start_row=_init + 1,
            end_row=_init + 2,
            sides=["left", "right", "top", "bottom"],
        )
        self._worksheet.row_dimensions[_init + 2].height = 4.5

        border = copy(self._worksheet["A50"].border)
        for row in range(51, self._worksheet.max_row):
            self._worksheet[f"A{row}"].border = border
        return var - 4

    def fill_sheet(self, values: dict):
        var = 51
        for key, value in values.items():
            if key in [
                "conservacao",
                "voltage_in_power_phases_norm",
                "spda_measurement_norm",
            ]:
                for _key, _value in value.items():
                    self._worksheet[self.__static_fields[key][_key]] = _value
                    if _value == "X":
                        self._worksheet[
                            self.__static_fields[key][_key]
                        ].fill = PatternFill(
                            start_color="808080",
                            end_color="808080",
                            fill_type="solid",
                        )
            elif key == "reporting_files":
                if not value:
                    value = [{"path": "", "description": "", "uuid": ""}]
                var = self.__insert_photographic_reports(reporting_files=value)
            elif key == "anomalies_found":
                if not value:
                    value = [
                        {
                            "outros": "",
                            "reparar": "",
                            "danificado": "",
                            "instalacao": "",
                            "description": "",
                            "sem_acabamento": "",
                            "occurrence_type": "",
                            "semidentificacao": "",
                            "treatment_images": [],
                            "anomalia_instalacoes": "",
                        },
                        {
                            "outros": "",
                            "reparar": "",
                            "danificado": "",
                            "instalacao": "",
                            "description": "",
                            "sem_acabamento": "",
                            "occurrence_type": "",
                            "semidentificacao": "",
                            "treatment_images": [],
                            "anomalia_instalacoes": "",
                        },
                    ]
                elif len(value) == 1:
                    value = [
                        value[0],
                        {
                            "outros": "",
                            "reparar": "",
                            "danificado": "",
                            "instalacao": "",
                            "description": "",
                            "sem_acabamento": "",
                            "occurrence_type": "",
                            "semidentificacao": "",
                            "treatment_images": [],
                            "anomalia_instalacoes": "",
                        },
                    ]

                var = self.__anomalies_found(
                    var=var,
                    anomalies_found=value,
                    reporting_files=values["anomalies_reporting_files"],
                )

                # if self.__report_format == ReportFormat.PDF:
                self.insert_page_breaks(len(value))
            elif key == "observacoes":
                self._worksheet[f"C{var}"] = value
                self._worksheet[f"C{var}"].alignment = Alignment(vertical="top")
                self._worksheet[f"C{var}"].alignment = self._worksheet[
                    f"C{var}"
                ].alignment.copy(wrapText=True)
            elif key in self.__static_fields:
                self._worksheet[self.__static_fields[key]] = value

    def create_dict(self, s3) -> dict:
        conservacao = {"1": "", "2": "", "3": ""}
        reporting = self.reporting
        building_description = []
        description_electrical_installations = []
        atendenormafase = new_get_form_data(reporting, "atendenormafase")
        atendenormaneutro = new_get_form_data(reporting, "atendenormaneutro")
        tipoedificacao = new_get_form_data(reporting, "tipoedificacao")
        direction = get_custom_option(reporting, "direction")
        if reporting.form_data.get("general_conservation_state"):
            conservacao[reporting.form_data["general_conservation_state"]] = "X"
        building_description.append(tipoedificacao)
        if reporting.form_data.get("municipio"):
            building_description.append(reporting.form_data["municipio"])
        building_description.append(f"Rodovia {reporting.road_name}")
        description_electrical_installations.append(
            new_get_form_data(reporting, "tensaosistema")
        )
        description_electrical_installations.append(
            new_get_form_data(reporting, "sistemaenergia")
        )
        description_electrical_installations = [
            a for a in description_electrical_installations if a
        ]
        concessionariaenergia = new_get_form_data(reporting, "concessionariaenergia")
        atendimentoabnt = new_get_form_data(reporting, "atendimentoabnt")
        atendimentoabnt_translate = {"Adequado": "ATENDE", "Inadequado": "NÃO ATENDE"}

        atendimentoabnt = (
            atendimentoabnt_translate[atendimentoabnt]
            if atendimentoabnt in atendimentoabnt_translate
            else atendimentoabnt
        )

        voltage_in_power_phases_norm = (
            True
            if str(atendenormafase).lower() == "dentro da norma"
            and str(atendenormaneutro).lower() == "dentro da norma"
            else False
        )
        spda_measurement_norm = (
            False
            if str(reporting.form_data.get("atendenorma_spda")).lower()
            == "fora da norma"
            else True
        )
        reporting_files = download_reporting_pictures(
            s3,
            self.temp_file,
            reporting,
            width=337,
            height=242,
            enable_include_dnit=False,
            enable_is_shared_antt=True,
        )

        anomalies_images_uuids = []
        if "therapy" in reporting.form_data:
            for therapy in reporting.form_data.get("therapy", []):
                if "treatment_images" in therapy:
                    for image in therapy.get("treatment_images", []):
                        anomalies_images_uuids.append(image)

        general_reporting_files = [
            a
            for a in reporting_files["images"]
            if a["uuid"] not in anomalies_images_uuids
        ]
        anomalies_reporting_files = [
            a for a in reporting_files["images"] if a["uuid"] in anomalies_images_uuids
        ]

        building_description = [a for a in building_description if a]

        data = {
            "identificacaoinstalacoes": reporting.form_data.get("id_ccr_antt"),
            "tipoedificacao": tipoedificacao,
            "road_name": reporting.road_name,
            "km": get_km(reporting),
            "direction": direction,
            "municipio": reporting.form_data.get("municipio"),
            "latitude": reporting.form_data.get("latitude"),
            "longitude": reporting.form_data.get("longitude"),
            "found_at": reporting.found_at.strftime("%d/%m/%Y"),
            "conservacao": conservacao,
            "building_description": (
                " - ".join(building_description) if building_description else ""
            ),
            "description_electrical_installations": (
                f'{" - ".join(description_electrical_installations)}\nPadrão de energia em baixa tensão.\nLigação do padrão até o painel geral, derivado para os circuitos de cada setor,\nindependente'
                if description_electrical_installations
                else ""
            ),
            "concessionariaenergia": concessionariaenergia,
            "atendimentoabnt": atendimentoabnt,
            "fase_rn": reporting.form_data.get("fase_rn"),
            "fases_n": reporting.form_data.get("fases_n"),
            "fase_tn": reporting.form_data.get("fase_tn"),
            "voltage_in_power_phases_norm": {
                "within_the_norm": "X" if voltage_in_power_phases_norm is True else "",
                "outside_the_norm": "" if voltage_in_power_phases_norm is True else "X",
            },
            "fase_rs": reporting.form_data.get("fase_rs"),
            "fases_t": reporting.form_data.get("fases_t"),
            "fase_tr": reporting.form_data.get("fase_tr"),
            "obs_fase": reporting.form_data.get("obs_fase"),
            "medicao_spda": reporting.form_data.get("medicao_spda"),
            "spda_measurement_norm": {
                "within_the_norm": "X" if spda_measurement_norm is True else "",
                "outside_the_norm": "" if spda_measurement_norm is True else "X",
            },
            "obs_spda": reporting.form_data.get("obs_spda"),
            "reporting_files": general_reporting_files,
            "anomalies_reporting_files": anomalies_reporting_files,
            "anomalies_found": (
                reporting.form_data["therapy"]
                if reporting.form_data.get("therapy")
                else []
            ),
            "observacoes": reporting.form_data.get("observacoes"),
        }
        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def execute(self):
        data = self.create_dict(s3=self.s3)
        file_name_options = [
            get_obj_from_path(self.reporting.form_data, "id_ccr_antt"),
            self.reporting.number,
            str(time.time()),
        ]
        file_name = clean_latin_string(
            next(a for a in file_name_options if a).replace(".", "").replace("/", "")
        )

        result = f"/tmp/{file_name}.xlsx"

        self.fill_sheet(values=data)
        self._insert_logos(
            self.s3,
            self.temp_file,
            self.__sheet_target,
            self._worksheet,
            self.reporting,
        )
        self._workbook.save(result)
        shutil.rmtree(self.temp_file, ignore_errors=True)

        return result


class CCRElectrical(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""

        if len(self.uuids) == 1:
            reporting = Reporting.objects.get(uuid=self.uuids[0])

            file_name_options = [
                get_obj_from_path(reporting.form_data, "id_ccr_antt"),
                reporting.number,
                str(time.time()),
            ]
            file_name = clean_latin_string(
                next(a for a in file_name_options if a)
                .replace(".", "")
                .replace("/", "")
            )

            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{file_name}.{extension}"

        elif len(self.uuids) > 1:
            file_name = (
                "{} Relatorios - Instalacoes Eletricas de Edificacoes.zip".format(
                    len(self.uuids)
                )
            )
        self.file_name = file_name
        return file_name

    def export(self):
        s3 = get_s3()
        files = []
        for uuid in self.uuids:
            files.append(
                XlsxHandler(
                    uuid=uuid,
                    s3=s3,
                    sheet_target=self.sheet_target(),
                    report_format=self.report_format(),
                ).execute()
            )

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_electrical_async_handler(reporter_dict: dict):
    reporter = CCRElectrical.from_dict(reporter_dict)
    reporter.export()
