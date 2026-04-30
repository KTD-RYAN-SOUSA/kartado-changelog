import shutil
import tempfile
from typing import Dict, List, Tuple
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ValidationError

from apps.companies.models import Company
from apps.reportings.models import Reporting, ReportingFile, ReportingInReporting
from apps.service_orders.models import ServiceOrderActionStatusSpecs
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import (
    get_form_array_iterator,
    new_get_form_data,
)
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_image,
    get_logo_file,
    get_provider_logo_file,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_direction,
    get_end_km,
    get_km,
)
from helpers.apps.ccr_report_utils.workbook_utils import append_row, save_workbook
from helpers.strings import clean_latin_string


class XlsxHandler(object):
    ElementType = str

    __ELEMENT_OCCURRENCE_TYPE = "Monitoração Drenagem Profunda Elemento"
    __ELEMENT_RELATION_NAMES = ["Montante", "Jusante"]
    _ELEMENT_HOLE_CLASSIFICATION = None

    _TEMPLATE_FILE = None
    _TEMPLATE_EMPTY_FILE = None
    __HEADER_CELL = "A1"
    __LOGO_CELL = "K1:L1"
    __PROVIDER_LOGO_CELL = "A1:C1"

    # Columns
    __ID = 0
    __ROAD = 1
    __KM = 2
    __END_KM = 3
    __DIRECTION = 4
    __HOLE = 5
    __TYPE = 6
    __CURR_STATE = 7
    __CURR_PIC = 8
    __ACT_STATE = 9
    __ACT_PIC = 10
    __CORRECT_ACT = 11
    __REPORTING_SERIAL = 12
    __ELEMENT_SERIAL = 13
    __INVENTORY_SERIAL = 14

    __TEMPLATE_ROW = 5

    @classmethod
    def __get_reportings(cls, uuids: List[str]) -> List[Reporting]:
        return list(
            Reporting.objects.filter(
                uuid__in=uuids,
                occurrence_type__name=OACVII.OCCURRENCE_TYPE_NAME,
            )
            .only(
                "uuid",
                "number",
                "form_data",
                "parent__number",
                "company__logo",
                "company__provider_logo",
            )
            .prefetch_related("company")
            .order_by("form_data__id_ccr_antt")
        )

    @classmethod
    def __get_status_name_to_order(cls, company_uuid) -> Dict[str, int]:
        status_name_to_order: Dict[str, int] = {}
        status_orders = ServiceOrderActionStatusSpecs.objects.filter(
            company__uuid=company_uuid
        ).only("order", "status__name")
        for status_order in status_orders:
            status_name_to_order[status_order.status.name] = status_order.order
        return status_name_to_order

    @classmethod
    def __get_elements(cls, reporting: str) -> List[Tuple[ElementType, Reporting]]:
        rels = ReportingInReporting.objects.filter(
            parent=reporting,
            child__occurrence_type__name=XlsxHandler.__ELEMENT_OCCURRENCE_TYPE,
            reporting_relation__name__in=XlsxHandler.__ELEMENT_RELATION_NAMES,
        )
        if cls._ELEMENT_HOLE_CLASSIFICATION is not None:
            rels = rels.filter(
                child__form_data__hole_classification=cls._ELEMENT_HOLE_CLASSIFICATION
            )

        rels = (
            rels.only(
                "reporting_relation__name",
                "child__uuid",
                "child__number",
                "child__road_name",
                "child__km",
                "child__end_km",
                "child__direction",
                "child__form_data",
            )
            .prefetch_related(
                "child",
                "reporting_relation",
            )
            .order_by("child__km")
        )
        return [
            (rel.reporting_relation.name, rel.child)
            for rel in rels
            if "ocultar_anexo" not in rel.child.form_data
            or rel.child.form_data["ocultar_anexo"] is False
        ]

    @classmethod
    def __get_recoveries(
        cls, reporting_uuid: str, occurrence_kinds: List[str]
    ) -> List[Reporting]:
        reporting_recovery_relations = (
            ReportingInReporting.objects.filter(
                parent=reporting_uuid,
                reporting_relation__name="Recuperação",
            )
            .only(
                "reporting_relation__name",
                "child__uuid",
                "child__number",
                "child__status__name",
                "child__occurrence_type__name",
                "child__form_data",
                "child__executed_at",
                "child__updated_at",
                "child__created_at",
            )
            .prefetch_related(
                "child",
                "child__occurrence_type",
                "child__status",
                "reporting_relation",
            )
        )
        if len(occurrence_kinds) > 0:
            reporting_recovery_relations = reporting_recovery_relations.filter(
                child__occurrence_type__occurrence_kind__in=occurrence_kinds,
            )
        return [
            reporting_recovery_relation.child
            for reporting_recovery_relation in reporting_recovery_relations
        ]

    def __set_header(self, worksheet: Worksheet, road_name: str) -> None:
        header_cell: Cell = worksheet[XlsxHandler.__HEADER_CELL]
        header_cell.value = f"{header_cell.value} {road_name}"
        sample_reporting = XlsxHandler.__get_reportings(self.list_uuids[:1])[0]

        try:
            logo_file = get_logo_file(self.s3, self.temp_dir, sample_reporting)
            insert_picture(
                worksheet,
                XlsxHandler.__LOGO_CELL,
                Image(logo_file),
                self.__sheet_target,
                resize_method=ResizeMethod.ProportionalRight,
                border_width=2,
            )
        except Exception as e:
            print(e)
        try:
            provider_logo_file = get_provider_logo_file(
                self.s3, self.temp_dir, sample_reporting
            )
            insert_picture(
                worksheet,
                XlsxHandler.__PROVIDER_LOGO_CELL,
                Image(provider_logo_file),
                self.__sheet_target,
                resize_method=ResizeMethod.ProportionalLeft,
                border_width=2,
            )
        except Exception as e:
            print(e)

    @classmethod
    def __get_current_state_picture(
        cls, s3, temp_dir: str, element: Reporting
    ) -> Image:
        detail_photos_mon_uuids: List[str] = []
        try:
            photos_mon_it = get_form_array_iterator(element, "photosMon")
            while True:
                try:
                    detail_photos_mon = photos_mon_it.get("detailPhotosMon")
                    for detail_photo_mon in detail_photos_mon:
                        try:
                            detail_photos_mon_uuids.append(detail_photo_mon)
                        except Exception as e:
                            print(e)
                except Exception as e:
                    print(e)
                photos_mon_it.inc()
        except Exception as e:
            print(e)

        reporting_files = list(
            ReportingFile.objects.filter(
                uuid__in=detail_photos_mon_uuids, is_shared=True
            ).only("uuid", "upload")
        )
        picture: str = None
        for reporting_file in reporting_files:
            try:
                picture = get_image(s3, temp_dir, reporting_file.uuid, reporting_file)
                if picture is not None:
                    break
            except Exception as e:
                print(e)
        return picture

    def __get_action_picture(
        self, s3, temp_dir: str, recoveries: List[Reporting]
    ) -> str:
        gt_executed_at_recovery: Reporting = None
        gt_updated_at_recovery: Reporting = None
        executed_recovery: Reporting = None
        selected_recovery: Reporting = None

        for recovery in recoveries:
            if self.__is_status_name_gte(recovery.status.name, "Executado"):
                executed_recovery = recovery
            is_executed = (
                self.__is_status_name_gte(recovery.status.name, "Executado")
                or recovery.executed_at is not None
            )
            if is_executed:
                has_greatest_executed_at = (
                    gt_executed_at_recovery is None
                    or recovery.executed_at > gt_executed_at_recovery.executed_at
                )
                if has_greatest_executed_at:
                    gt_executed_at_recovery = recovery

            has_greatest_updated_at = (
                gt_updated_at_recovery is None
                or recovery.updated_at > gt_updated_at_recovery.updated_at
            )
            if has_greatest_updated_at:
                gt_updated_at_recovery = recovery

        if gt_executed_at_recovery:
            selected_recovery = gt_executed_at_recovery
        elif executed_recovery is not None:
            selected_recovery = executed_recovery
        else:
            selected_recovery = gt_updated_at_recovery

        reporting_files = list(
            ReportingFile.objects.filter(
                reporting=selected_recovery,
                is_shared=True,
                datetime__gte=selected_recovery.created_at,
            )
            .only("uuid", "upload")
            .order_by("datetime", "uploaded_at")
            .reverse()
        )
        picture: str = None
        for reporting_file in reporting_files:
            try:
                picture = get_image(s3, temp_dir, reporting_file.uuid, reporting_file)
                if picture is not None:
                    break
            except Exception as e:
                print(e)
        return picture

    @classmethod
    def __get_act_state(cls, recoveries: List[Reporting]) -> str:
        max_recovery: Reporting = recoveries[0]
        for recovery in recoveries[1:]:
            if recovery.executed_at is not None and (
                max_recovery.executed_at is None
                or recovery.executed_at > max_recovery.executed_at
            ):
                max_recovery = recovery
        if max_recovery.executed_at is None:
            for recovery in recoveries[1:]:
                if recovery.updated_at is not None and (
                    max_recovery.updated_at is None
                    or recovery.updated_at > max_recovery.updated_at
                ):
                    max_recovery = recovery

        if max_recovery.executed_at is not None or max_recovery.updated_at is not None:
            return new_get_form_data(
                max_recovery, "generalConservationState", default="-"
            )
        return "-"

    def __get_correct_acts(self, recoveries: List[Reporting]) -> str:
        action_list: List[str] = []
        action_string = "-"
        all_executed = True

        for recovery in recoveries:
            is_executed = (
                self.__is_status_name_gte(recovery.status.name, "Executado")
                or recovery.executed_at is not None
            )
            if is_executed:
                action_list.append(recovery.occurrence_type.name)
            else:
                all_executed = False
                break

        if all_executed:
            action_string = ", ".join(action_list)
        else:
            action_list.clear()
            for recovery in recoveries:
                try:
                    is_executed = (
                        self.__is_status_name_gte(recovery.status.name, "Executado")
                        or recovery.executed_at is not None
                    )
                    if not is_executed:
                        justificativa: str = new_get_form_data(
                            recovery, "justificativaNaoExecutado", default=""
                        )
                        justificativa = justificativa.strip()
                        if justificativa:
                            action_list.append(justificativa)
                except Exception as e:
                    print(e)
            if len(action_list) > 0:
                action_string = " / ".join(action_list)
        return action_string

    @classmethod
    def __get_recovery_occurrence_kinds(cls, company_uuid: UUID, report_name: str):
        recovery_occurrence_kinds = []
        try:
            metadata = (
                Company.objects.filter(uuid=company_uuid).only("metadata")[0].metadata
            )
            options: dict = metadata["extraReportingExports"][report_name]["options"]
            recovery_occurrence_kinds = options["recovery_occurrence_kinds"]
        except Exception as e:
            print(e)
        return recovery_occurrence_kinds

    def __init__(
        self,
        report_name: str,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.__sheet_target = sheet_target
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.occurrence_type = Reporting.objects.get(uuid=list_uuids[0]).occurrence_type
        company_uuid = (
            Reporting.objects.filter(uuid=list_uuids[0])
            .only("company__uuid")[0]
            .company.uuid
        )

        self.__recovery_occurrence_kinds = XlsxHandler.__get_recovery_occurrence_kinds(
            company_uuid, report_name
        )
        self.__status_name_to_order = XlsxHandler.__get_status_name_to_order(
            company_uuid
        )

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __is_status_name_gte(self, status_a: str, status_b: str) -> bool:
        return (
            self.__status_name_to_order[status_a]
            >= self.__status_name_to_order[status_b]
        )

    def __append_to_sheet(self, worksheet: Worksheet, reporting: Reporting) -> None:
        cls = type(self)
        identification = new_get_form_data(reporting, "idCcrAntt", default="-")
        for element_type, element in cls.__get_elements(reporting):
            recoveries = XlsxHandler.__get_recoveries(
                element, self.__recovery_occurrence_kinds
            )
            if len(recoveries) < 1:
                continue

            road_name = element.road_name if element.road_name else "-"
            inventory_number = reporting.parent.number if reporting.parent else "-"

            row: List[str] = [""] * 15
            row[XlsxHandler.__ID] = identification
            row[XlsxHandler.__ROAD] = road_name
            row[XlsxHandler.__KM] = get_km(element, "-")
            row[XlsxHandler.__END_KM] = get_end_km(element, "-")
            row[XlsxHandler.__DIRECTION] = get_direction(element, "-")
            row[XlsxHandler.__HOLE] = new_get_form_data(
                element, "entryStruc", default="-"
            )
            row[XlsxHandler.__TYPE] = element_type
            row[XlsxHandler.__CURR_STATE] = new_get_form_data(
                element, "holeClassification", default="-"
            )
            row[XlsxHandler.__ACT_STATE] = XlsxHandler.__get_act_state(recoveries)
            row[XlsxHandler.__CORRECT_ACT] = self.__get_correct_acts(recoveries)
            row[XlsxHandler.__REPORTING_SERIAL] = reporting.number
            row[XlsxHandler.__ELEMENT_SERIAL] = element.number
            row[XlsxHandler.__INVENTORY_SERIAL] = inventory_number

            append_row(worksheet, row)

            curr_state_picture = XlsxHandler.__get_current_state_picture(
                self.s3, self.temp_dir, element
            )

            max_row = worksheet.max_row - 1
            if curr_state_picture is not None:
                cell = f"{get_column_letter(XlsxHandler.__CURR_PIC+1)}{max_row}"
                insert_picture(
                    worksheet,
                    cell,
                    curr_state_picture,
                    self.__sheet_target,
                    resize_method=ResizeMethod.Stretch,
                    border_width=1,
                )

            action_picture = self.__get_action_picture(
                self.s3, self.temp_dir, recoveries
            )
            if action_picture is not None:
                cell = f"{get_column_letter(XlsxHandler.__ACT_PIC+1)}{max_row}"
                insert_picture(
                    worksheet,
                    cell,
                    action_picture,
                    self.__sheet_target,
                    resize_method=ResizeMethod.Stretch,
                    border_width=1,
                )

    def __create_workbook_file(
        self, road_name: str, reportings: List[Reporting]
    ) -> str:
        cls = type(self)
        workbook = load_workbook(cls._TEMPLATE_FILE)
        worksheet = workbook[workbook.sheetnames[0]]

        for reporting in reportings:
            self.__append_to_sheet(worksheet, reporting)

        if worksheet.max_row == XlsxHandler.__TEMPLATE_ROW:
            workbook = load_workbook(cls._TEMPLATE_EMPTY_FILE)
            worksheet = workbook[workbook.sheetnames[0]]
        else:
            worksheet.delete_rows(XlsxHandler.__TEMPLATE_ROW)

        self.__set_header(worksheet, road_name)

        workbook_name = f"Anexo VII - Serviços Realizados OAC - {road_name}"
        workbook_name = clean_latin_string(
            workbook_name.replace(".", "").replace("/", "")
        )
        workbook_file = save_workbook(workbook_name, workbook)

        return workbook_file

    def execute(self) -> List[str]:
        reportings: List[Reporting] = XlsxHandler.__get_reportings(self.list_uuids)
        road_name_to_reportings: Dict[str, List[Reporting]] = {}

        for reporting in reportings:
            if reporting.road_name not in road_name_to_reportings:
                road_name_to_reportings[reporting.road_name] = []
            road_name_to_reportings[reporting.road_name].append(reporting)

        workbook_files: List[str] = []
        for road_name, reportings in road_name_to_reportings.items():
            workbook_files.append(self.__create_workbook_file(road_name, reportings))

        return workbook_files


class OACVII(CCRReport):
    OCCURRENCE_TYPE_NAME = "Monitoração de Drenagem Profunda Ficha Poder Concedente"
    _CLASSIFICATION = None
    _XLSX_HANDLER = None

    def __init__(
        self,
        report_name: str = None,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.__report_name = report_name
        super().__init__(uuids, report_format)

    def get_file_name(self) -> str:
        file_name: str = None

        reportings_query_set = (
            Reporting.objects.filter(
                uuid__in=self.uuids,
            )
            .only("uuid", "road_name", "occurrence_type__name")
            .prefetch_related("occurrence_type")
        )

        if any(
            reporting.occurrence_type.name != OACVII.OCCURRENCE_TYPE_NAME
            for reporting in reportings_query_set
        ):
            raise ValidationError("Apontamentos de outra classe")

        road_names = list(
            reportings_query_set.order_by("road_name")
            .distinct("road_name")
            .values_list("road_name", flat=True)
        )

        road_names.sort()
        extension = "zip"
        if len(road_names) > 1:
            cls = type(self)
            file_name = f"Anexo VII - Serviços Realizados OAC - {cls._CLASSIFICATION}_{'_'.join(road_names)}"
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        else:
            file_name = "Anexo VII - Serviços Realizados OAC - {}".format(
                "_".join(road_names)
            )
            file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
            extension = "xlsx" if self.report_format() == ReportFormat.XLSX else "pdf"
        file_name = f"{file_name}.{extension}"
        return file_name

    def export(self):
        cls = type(self)
        s3 = get_s3()
        files = cls._XLSX_HANDLER(
            report_name=self.__report_name,
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        else:
            result_file = files[0]

        upload_file(s3, result_file, self.object_name)

        return True
