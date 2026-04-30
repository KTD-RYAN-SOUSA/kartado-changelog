import shutil
import tempfile
from dataclasses import dataclass
from math import floor
from typing import Dict, List, Tuple
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ValidationError
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile, ReportingInReporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    get_recovery_occurrence_kinds,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_picture,
    get_logo_file,
    get_provider_logo_file,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_direction, get_km
from helpers.apps.ccr_report_utils.workbook_utils import (
    append_row,
    copy_cell_style,
    save_workbook,
)
from helpers.strings import clean_latin_string


@dataclass
class ParamStatus:
    bad_guard: bool = False
    bad_drainage: bool = False
    bad_sign: bool = False
    bad_joints: bool = False
    bad_support: bool = False

    def __init__(self, reporting: Reporting):
        self.bad_guard = new_get_form_data(reporting, "guardaCorpo", raw=True) == "2"
        self.bad_drainage = (
            new_get_form_data(reporting, "drenagemLimpeza", raw=True) == "2"
        )
        self.bad_sign = (
            new_get_form_data(reporting, "placaGabaritoVertical", raw=True) == "3"
        )
        self.bad_joints = (
            new_get_form_data(
                reporting, "juntasDilatacaoVidaUtilRemanescente", raw=True
            )
            == "2"
        )
        self.bad_support = (
            new_get_form_data(reporting, "aparelhoApoioVidaUtilRemanescente", raw=True)
            == "2"
        )

    def is_bad(self):
        return (
            self.bad_guard
            or self.bad_drainage
            or self.bad_sign
            or self.bad_joints
            or self.bad_support
        )


def append_string(value, list: list) -> bool:
    if isinstance(value, str):
        value = value.strip()
        if len(value) > 0:
            list.append(value)
            return True
    return False


class XlsxHandler(object):

    __TEMPLATE_FILE = "./fixtures/reports/oae_iv.xlsx"
    __TEMPLATE_EMPTY_FILE = "./fixtures/reports/oae_iv_empty.xlsx"
    __HEADER_CELL = "A1"
    __LOGO_CELL = "I1:I5"
    __PROVIDER_LOGO_CELL = "A1:B5"

    # Columns
    __OAE_NUMBER = 0
    __AGENCY = 1
    __NAME = 2
    __ROAD = 3
    __KM = 4
    __DIRECTION = 5
    __PARAM = 6
    __ACTIONS = 7
    __PIC = 8
    __REPORTING_SERIAL = 9
    __THERAPY_SERIAL = 10
    __INVENTORY_SERIAL = 11

    __COLUMNS = 12
    __TEMPLATE_ROW = 7
    __PIC_COLUMN_LETTER = get_column_letter(__PIC + 1)

    @classmethod
    def __get_sorting_key(
        cls, reporting: Reporting
    ) -> Tuple[Tuple[bool, int], Tuple[bool, str], float]:
        oae_number = new_get_form_data(reporting, "oaeNumeroCodigoObra", default=None)
        agency = new_get_form_data(reporting, "codigoAgencia", default=None)
        return (
            (oae_number is None, oae_number),
            (agency is None, agency),
            reporting.km,
        )

    @classmethod
    def __get_reportings(cls, uuids: List[str]) -> List[Reporting]:
        return list(
            Reporting.objects.filter(uuid__in=uuids)
            .only(
                "uuid",
                "number",
                "form_data",
                "parent__number",
                "company__logo",
                "company__provider_logo",
            )
            .prefetch_related("company")
        )

    @classmethod
    def __get_therapies(
        cls, reporting_uuid: str, occurrence_kinds: List[str]
    ) -> List[Reporting]:
        reporting_recovery_relations = (
            ReportingInReporting.objects.filter(
                parent=reporting_uuid,
                reporting_relation__name="Recuperação",
            )
            .only(
                "child__uuid",
                "child__number",
            )
            .prefetch_related(
                "child",
            )
        )
        if len(occurrence_kinds) > 0:
            reporting_recovery_relations = reporting_recovery_relations.filter(
                child__occurrence_type__occurrence_kind__in=occurrence_kinds,
            )
        return [
            reporting_recovery_relation.child
            for reporting_recovery_relation in reporting_recovery_relations
        ]

    @classmethod
    def __get_pictures(cls, s3, dir: str, therapies: List[Reporting]) -> List[Image]:
        pictures = []
        reporting_files = (
            ReportingFile.objects.filter(reporting__in=therapies, is_shared=True)
            .order_by("datetime", "uploaded_at")
            .only("uuid", "upload")
        )
        for rf in reporting_files:
            try:
                file = download_picture(s3, dir, rf.uuid, reporting_file=rf)
                if file is not None:
                    pictures.append(Image(file))
            except Exception as e:
                print(e)
        return pictures

    @classmethod
    def __insert_pictures(
        cls, worksheet: Worksheet, pictures: List[Image], sheet_target: SheetTarget
    ) -> None:
        if len(pictures) > 0:
            max_row = worksheet.max_row
            row_dimensions = worksheet.row_dimensions
            length = len(pictures)
            for row in range(max_row + 1, max_row + length):
                row_dimensions[row].ht = row_dimensions[row - 1].ht

            if len(pictures) > 1:
                for col in range(XlsxHandler.__COLUMNS):
                    end_row = max_row + length - 1
                    if col != XlsxHandler.__PIC:
                        worksheet.merge_cells(
                            start_row=max_row,
                            end_row=end_row,
                            start_column=col + 1,
                            end_column=col + 1,
                        )
                        copy_cell_style(
                            worksheet.cell(end_row, col + 1),
                            worksheet.cell(max_row, col + 1),
                        )

            for i, picture in enumerate(pictures):
                cell = f"{XlsxHandler.__PIC_COLUMN_LETTER}{max_row+i}"
                insert_picture(
                    worksheet,
                    cell,
                    picture,
                    sheet_target,
                    border_width=4,
                )
                copy_cell_style(
                    worksheet.cell(max_row + i, XlsxHandler.__PIC + 1),
                    worksheet.cell(max_row, XlsxHandler.__PIC + 1),
                )

    @classmethod
    def __get_params(cls, reporting: Reporting, param_status: ParamStatus) -> str:
        params = []
        params_str: str = "-"

        if param_status.bad_guard:
            params.append("Guarda-corpo ruim ou ausente")
        if param_status.bad_drainage:
            params.append("Drenagem/limpeza ruim")
        if param_status.bad_sign:
            params.append("Ausência de placas de gabarito vertical")

        if param_status.bad_joints:
            local: str = new_get_form_data(reporting, "juntaDilatacaoLocal5PistaAcesso")
            amount: str = new_get_form_data(
                reporting, "juntaDilatacaoObservacoesQuantidade5PistaAcesso"
            )
            appended = append_string(local, params)
            appended = append_string(amount, params) or appended
            if not appended:
                params.append("Uma ou mais juntas de dilatação não atendem")

        if param_status.bad_support:
            local: str = new_get_form_data(
                reporting, "aparelhoApoioLocal3Mesoestrutura"
            )
            amount: str = new_get_form_data(
                reporting, "aparelhoApoioObservacoesQuantidade3Mesoestrutura"
            )
            appended = append_string(local, params)
            appended = append_string(amount, params) or appended
            if not appended:
                params.append("Aparelho de apoio não atende")

        if len(params) > 0:
            params_str = "/ ".join(params)

        return params_str

    @classmethod
    def __get_actions(cls, reporting: Reporting, param_status: ParamStatus) -> str:
        actions = []
        actions_str: str = "-"

        if param_status.bad_guard:
            guard = new_get_form_data(reporting, "guardaCorpoObservacoesQuantidade")
            append_string(guard, actions)
        if param_status.bad_drainage:
            drainage = new_get_form_data(
                reporting, "drenagemLimpezaObservacoesQuantidade"
            )
            append_string(drainage, actions)
        if param_status.bad_sign:
            sign = new_get_form_data(
                reporting, "placaGabaritoVerticalObservacoesQuantidade"
            )
            append_string(sign, actions)
        if param_status.bad_joints:
            joints = new_get_form_data(
                reporting, "juntasDilatacaoVidaUtilRemanescenteObservacoesQuantidade"
            )
            append_string(joints, actions)
        if param_status.bad_support:
            support = new_get_form_data(
                reporting, "aparelhoApoioVidaUtilRemanescenteLocal"
            )
            append_string(support, actions)

        if len(actions) > 0:
            actions_str = "/ ".join(actions)

        return actions_str

    def __set_header(
        self, worksheet: Worksheet, road_name: str, sample_reporting: Reporting
    ) -> None:
        header_cell: Cell = worksheet[XlsxHandler.__HEADER_CELL]
        header_cell.value = f"{header_cell.value} {road_name}"

        try:
            logo_file = get_logo_file(self.s3, self.temp_dir, sample_reporting)
            insert_picture(
                worksheet,
                XlsxHandler.__LOGO_CELL,
                Image(logo_file),
                self.__sheet_target,
                resize_method=ResizeMethod.ProportionalRight,
                border_width=2,
            )
        except Exception as e:
            print(e)
        try:
            provider_logo_file = get_provider_logo_file(
                self.s3, self.temp_dir, sample_reporting
            )
            insert_picture(
                worksheet,
                XlsxHandler.__PROVIDER_LOGO_CELL,
                Image(provider_logo_file),
                self.__sheet_target,
                resize_method=ResizeMethod.ProportionalLeft,
                border_width=2,
            )
        except Exception as e:
            print(e)

    def __init__(
        self,
        report_name: str,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self.__report_format = report_format
        self.__sheet_target = sheet_target
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.last_page = 0

        company_uuid = (
            Reporting.objects.filter(uuid=list_uuids[0])
            .only("company__uuid")[0]
            .company.uuid
        )

        self.__recovery_occurrence_kinds = get_recovery_occurrence_kinds(
            company_uuid, report_name
        )

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __append_to_sheet(
        self, worksheet: Worksheet, reporting: Reporting, delete_template: bool
    ) -> bool:
        param_status = ParamStatus(reporting)
        appended = False
        if param_status.is_bad():
            therapies = XlsxHandler.__get_therapies(
                reporting.uuid, self.__recovery_occurrence_kinds
            )
            if len(therapies) > 0:
                pictures = XlsxHandler.__get_pictures(self.s3, self.temp_dir, therapies)

                road_name = reporting.road_name if reporting.road_name else "-"
                inventory_number = reporting.parent.number if reporting.parent else "-"

                row: List[str] = [""] * XlsxHandler.__COLUMNS
                oae_number = new_get_form_data(reporting, "oaeNumeroCodigoObra")
                row[XlsxHandler.__OAE_NUMBER] = oae_number if oae_number else "-"
                agency = new_get_form_data(reporting, "codigoAgencia")
                row[XlsxHandler.__AGENCY] = agency if agency else "-"
                name = new_get_form_data(reporting, "denominacao")
                row[XlsxHandler.__NAME] = name if name else "-"
                row[XlsxHandler.__ROAD] = road_name
                row[XlsxHandler.__KM] = get_km(reporting, "-")
                row[XlsxHandler.__DIRECTION] = get_direction(reporting, "-")

                row[XlsxHandler.__PARAM] = XlsxHandler.__get_params(
                    reporting, param_status
                )
                row[XlsxHandler.__ACTIONS] = XlsxHandler.__get_actions(
                    reporting, param_status
                )

                row[XlsxHandler.__REPORTING_SERIAL] = reporting.number
                therapies_serials = "-"
                if len(therapies) > 0:
                    therapies_serials = ", ".join(
                        [therapy.number for therapy in therapies]
                    )
                row[XlsxHandler.__THERAPY_SERIAL] = therapies_serials
                row[XlsxHandler.__INVENTORY_SERIAL] = inventory_number

                max_row = worksheet.max_row
                new_rows = max(len(pictures), 1)
                must_break = False
                if max_row < 13:
                    if max_row + new_rows > 13 and new_rows <= 8:
                        must_break = True
                        self.last_page = max_row
                    elif new_rows > 8:
                        self.last_page = 13 + floor(((max_row + new_rows - 13) / 8)) * 8
                else:
                    if self.last_page == 0:
                        self.last_page = 13
                    in_last_page = max_row - self.last_page
                    remaining_in_page = 8 - in_last_page
                    if new_rows < 8 and new_rows > remaining_in_page:
                        must_break = True
                        self.last_page = max_row
                    elif new_rows > 8:
                        self.last_page = (
                            self.last_page
                            + floor(((new_rows - remaining_in_page) / 8)) * 8
                        )

                if must_break:
                    worksheet.row_breaks.append(Break(max_row))

                append_row(worksheet, row)
                if delete_template:
                    worksheet.delete_rows(XlsxHandler.__TEMPLATE_ROW)
                appended = True
                XlsxHandler.__insert_pictures(worksheet, pictures, self.__sheet_target)
        return appended

    def __create_workbook_file(
        self, road_name: str, reportings: List[Reporting]
    ) -> str:
        workbook = load_workbook(XlsxHandler.__TEMPLATE_FILE)
        worksheet = workbook[workbook.sheetnames[0]]

        sorted_reportings = sorted(
            reportings, key=lambda r: XlsxHandler.__get_sorting_key(r)
        )

        appended = False
        for reporting in sorted_reportings:
            appended = (
                self.__append_to_sheet(worksheet, reporting, not appended) or appended
            )

        if not appended:
            workbook = load_workbook(XlsxHandler.__TEMPLATE_EMPTY_FILE)
            worksheet = workbook[workbook.sheetnames[0]]

        if self.__report_format == ReportFormat.PDF:
            for i in range(1, 6):
                cell: Cell = worksheet[f"I{i}"]
                side = Side(style="thin")
                cell.border = Border(
                    right=side,
                )

        self.__set_header(worksheet, road_name, reportings[0])

        workbook_name = f"Anexo IV - {road_name}"
        workbook_name = clean_latin_string(
            workbook_name.replace(".", "").replace("/", "")
        )
        workbook_file = save_workbook(workbook_name, workbook)

        return workbook_file

    def execute(self) -> List[str]:
        reportings: List[Reporting] = XlsxHandler.__get_reportings(self.list_uuids)
        road_name_to_reportings: Dict[str, List[Reporting]] = {}

        for reporting in reportings:
            if reporting.road_name not in road_name_to_reportings:
                road_name_to_reportings[reporting.road_name] = []
            road_name_to_reportings[reporting.road_name].append(reporting)

        workbook_files: List[str] = []
        for road_name, reportings in road_name_to_reportings.items():
            workbook_files.append(self.__create_workbook_file(road_name, reportings))

        return workbook_files


class OAEIV(CCRReport):
    OCCURRENCE_TYPE_NAME = "Monitoração OAE Poder Concedente"

    def __init__(
        self,
        report_name: str = None,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        self.__report_name = report_name
        super().__init__(uuids, report_format)

    def get_file_name(self) -> str:
        file_name: str = None

        reportings_query_set = (
            Reporting.objects.filter(
                uuid__in=self.uuids,
            )
            .only("uuid", "road_name", "occurrence_type__name")
            .prefetch_related("occurrence_type")
        )

        if any(
            reporting.occurrence_type.name != OAEIV.OCCURRENCE_TYPE_NAME
            for reporting in reportings_query_set
        ):
            raise ValidationError("Apontamentos de outra classe")

        road_names = list(
            reportings_query_set.order_by("road_name")
            .distinct("road_name")
            .values_list("road_name", flat=True)
        )

        road_names.sort()
        extension = "zip"
        file_name = "Anexo IV - {}".format("_".join(road_names))
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        if len(road_names) == 1:
            extension = "xlsx" if self.report_format() == ReportFormat.XLSX else "pdf"
        file_name = f"{file_name}.{extension}"
        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandler(
            report_name=self.__report_name,
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        else:
            result_file = files[0]

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_oae_iv_async_handler(
    reporter_dict: dict,
):
    reporter: OAEIV = OAEIV.from_dict(reporter_dict)
    reporter.export()
