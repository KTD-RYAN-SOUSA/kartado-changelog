import shutil
import tempfile
from datetime import timedelta, timezone
from typing import Dict, List, Tuple
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting, ReportingFile, ReportingInReporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    SheetTarget,
    get_image,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_direction,
    get_end_km,
    get_identification,
    get_km,
    get_parent_serial,
    get_road_name,
    get_serial,
)
from helpers.apps.ccr_report_utils.workbook_utils import save_workbook, set_block_style
from helpers.strings import clean_latin_string


class EmbankmentsAnnexFiveXlsxHandler(object):

    __OCCURRENCE_TYPE_UUIDS: List[UUID] = []

    __RECUPERATION_NAMES: List[str] = [
        "Desassoreamento",
        "Esgotamento e destinação",
        "Hidrojateamento em drenagem",
        "Implantação Drenagem",
        "Limpeza/desobstrução manual",
        "Limpeza/desobstrução mecânica",
        "Pintura em elemento de drenagem",
        "Reconstrução Drenagem",
        "Recuperação de drenagem",
        "Reparo em drenagem",
    ]

    __DEFAULT_ROW_HEIGHT = 99.75

    __DEFAULT_BORDER_COLOR = "A5A5A5"

    __DEFAULT_FONT = Font(name="Calibri", size=11)
    __DEFAULT_SIDE = Side(border_style="thin", color=__DEFAULT_BORDER_COLOR)

    __DEFAULT_BORDER = Border(
        left=__DEFAULT_SIDE,
        right=__DEFAULT_SIDE,
        top=__DEFAULT_SIDE,
        bottom=__DEFAULT_SIDE,
    )

    __DEFAULT_ALIGNMENT = Alignment(
        vertical="center", horizontal="center", wrap_text=True
    )

    @classmethod
    def __get_linked_reportings(cls, reporting: Reporting) -> List[Reporting]:
        try:
            query_set = (
                ReportingInReporting.objects.filter(
                    parent=reporting.uuid,
                    child__occurrence_type__name__in=EmbankmentsAnnexFiveXlsxHandler.__RECUPERATION_NAMES,
                )
                .exclude(child__status__name="Executado")
                .only("child__status__name", "child__occurrence_type__name")
                .prefetch_related("child", "child__occurrence_type", "child__status")
            )
            return [link.child for link in query_set]
        except Exception:
            return []

    @classmethod
    def __get_condition_dre_sup(cls, reporting: Reporting) -> str:
        try:
            return new_get_form_data(reporting, "conditionDreSup")
        except Exception:
            return ""

    @classmethod
    def __get_condition_dre_prof(cls, reporting: Reporting) -> str:
        try:
            return new_get_form_data(reporting, "conditionDreProf")
        except Exception:
            return ""

    @classmethod
    def __get_drainage_pictures(
        cls, reporting: Reporting, limit: int
    ) -> List[ReportingFile]:
        try:
            drainage_pictures = new_get_form_data(reporting, "drainagePictures")
            drainage_picture_uuids = [
                UUID(uuid)
                for drainage_picture in drainage_pictures
                for uuid in drainage_picture["drainage_picture"]
            ][:limit]
            return list(
                ReportingFile.objects.filter(
                    uuid__in=drainage_picture_uuids, is_shared=True
                ).only("upload")
            )
        except Exception:
            return []

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ):
        self.__sheet_target = sheet_target
        self.__sheet_reporting: Dict[Tuple[str, int], List[Reporting]] = {}
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids
        self.occurrence_type = Reporting.objects.get(uuid=list_uuids[0]).occurrence_type

        for e in OccurrenceType.objects.values("uuid", "name"):
            if e["name"] in EmbankmentsAnnexFiveXlsxHandler.__RECUPERATION_NAMES:
                EmbankmentsAnnexFiveXlsxHandler.__OCCURRENCE_TYPE_UUIDS.append(
                    e["uuid"]
                )

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def __append_entry(
        self,
        worksheet: Worksheet,
        reporting: Reporting,
        recuperation_reportings: List[Reporting],
    ) -> None:
        identification = get_identification(reporting)
        road_name = get_road_name(reporting)
        km = get_km(reporting)
        end_km = get_end_km(reporting)
        direction = get_direction(reporting)
        serial = get_serial(reporting)
        inventory_serial = get_parent_serial(reporting)
        recuperations = " - ".join(
            sorted(
                list(
                    {
                        recuperation_reporting.occurrence_type.name
                        for recuperation_reporting in recuperation_reportings
                    }
                )
            )
        )
        deadline = ""
        statuses = " - ".join(
            sorted(
                list(
                    {
                        recuperation_reporting.status.name
                        for recuperation_reporting in recuperation_reportings
                    }
                )
            )
        )

        text_entry: List[str | Cell] = [
            "",
            identification,
            road_name,
            km,
            end_km,
            direction,
            recuperations,
            deadline,
            statuses,
            " ",
            " ",
            serial,
            inventory_serial,
        ]
        worksheet.append(text_entry)

        last_row_index = worksheet.max_row
        condition_dre_sup = EmbankmentsAnnexFiveXlsxHandler.__get_condition_dre_sup(
            reporting
        )
        condition_dre_prof = EmbankmentsAnnexFiveXlsxHandler.__get_condition_dre_prof(
            reporting
        )

        set_block_style(
            worksheet,
            row_begin=last_row_index,
            row_end=last_row_index,
            col_begin="B",
            col_end="M",
            height=EmbankmentsAnnexFiveXlsxHandler.__DEFAULT_ROW_HEIGHT,
            font=EmbankmentsAnnexFiveXlsxHandler.__DEFAULT_FONT,
            border=EmbankmentsAnnexFiveXlsxHandler.__DEFAULT_BORDER,
            alignment=EmbankmentsAnnexFiveXlsxHandler.__DEFAULT_ALIGNMENT,
        )

        if condition_dre_sup != "Satisfatória" or condition_dre_prof != "Satisfatória":
            pictures_reporting_files = (
                EmbankmentsAnnexFiveXlsxHandler.__get_drainage_pictures(reporting, 2)
            )
            images = [
                get_image(
                    self.s3,
                    self.temp_dir,
                    "temp" + str(pictures_reporting_file.uuid) + "-file",
                    pictures_reporting_file,
                    150,
                    150,
                )
                for pictures_reporting_file in pictures_reporting_files
            ]

            try:
                insert_picture(
                    worksheet,
                    range_string=f"J{last_row_index}",
                    picture=images[0],
                    target=self.__sheet_target,
                )
                insert_picture(
                    worksheet,
                    range_string=f"K{last_row_index}",
                    picture=images[1],
                    target=self.__sheet_target,
                )
            except Exception:
                pass

    def __add_to_sheet_reporting(self, reporting: Reporting) -> None:
        found_at = reporting.found_at
        time_delta = timedelta(hours=-3)
        time_zone = timezone(time_delta)
        found_at = found_at.astimezone(time_zone)

        if (reporting.road_name, found_at.year) not in self.__sheet_reporting:
            self.__sheet_reporting[(reporting.road_name, found_at.year)] = []

        self.__sheet_reporting[(reporting.road_name, found_at.year)].append(reporting)

    def __create_workbooks_files(self) -> List[str]:
        files: List[str] = []

        sorted_sheets = sorted(self.__sheet_reporting.items())
        for sheet_key, reportings in sorted_sheets:
            (road_name, year) = sheet_key
            reportings.sort(key=lambda reporting: reporting.km)

            workbook = load_workbook(
                "./fixtures/reports/ccr_embankments_retaining_structures_annex_five.xlsx"
            )
            worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])

            worksheet.column_dimensions["L"].hidden = True
            worksheet.column_dimensions["M"].hidden = True

            for reporting in reportings:

                linked_reportings = (
                    EmbankmentsAnnexFiveXlsxHandler.__get_linked_reportings(reporting)
                )
                if len(linked_reportings) == 0:
                    continue

                self.__append_entry(worksheet, reporting, linked_reportings)

            workbook_name = " Anexo V Protocolo- Cronograma de Drenagem {} {}".format(
                road_name, year
            )
            file = save_workbook(workbook_name, workbook)
            files.append(file)

        return files

    def execute(self) -> List[str]:

        query_set = (
            Reporting.objects.filter(
                occurrence_type=self.occurrence_type, uuid__in=self.list_uuids
            )
            .only(
                "uuid",
                "number",
                "road_name",
                "found_at",
                "km",
                "end_km",
                "direction",
                "form_data",
                "parent__number",
            )
            .prefetch_related("parent")
        )

        for report in query_set:
            self.__add_to_sheet_reporting(reporting=report)

        return self.__create_workbooks_files()


class CCREmbankmentsAnnexFive(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        road_names = list(
            Reporting.objects.filter(uuid__in=self.uuids)
            .only("uuid", "road_name")
            .order_by("road_name")
            .distinct("road_name")
            .values_list("road_name", flat=True)
        )
        file_name = "Anexo V Protocolo- Cronograma de Drenagem {}".format(
            " - ".join(road_names)
        )
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"
        return file_name

    def export(self):
        s3 = get_s3()
        files = EmbankmentsAnnexFiveXlsxHandler(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_embankments_annex_five_async_handler(
    reporter_dict: dict,
):
    reporter = CCREmbankmentsAnnexFive.from_dict(reporter_dict)
    reporter.export()
