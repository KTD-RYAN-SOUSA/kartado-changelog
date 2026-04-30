import shutil
import tempfile
from typing import Dict, List, Tuple
from uuid import UUID
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ValidationError
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_picture,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_direction, get_km
from helpers.apps.ccr_report_utils.workbook_utils import append_row, save_workbook
from helpers.strings import clean_latin_string


class XlsxHandler(object):
    __TEMPLATE_FILE = "./fixtures/reports/oae_i.xlsx"
    __HEADER_CELL = "A1"
    __LOGO_CELL = "R1:U4"
    __PROVIDER_LOGO_CELL = "A1:B4"

    # Columns
    __OAE_NUMBER = 0
    __AGENCY = 1
    __NAME = 2
    __ROAD = 3
    __KM = 4
    __DIRECTION = 5
    __LEN = 6
    __WIDTH = 7
    __AREA = 8
    __TYPE = 9
    __INFO = 10
    __CURR = 11
    __PREV_3 = 12
    __PREV_2 = 13
    __PREV_1 = 14
    __SCHED = 17
    __REPORTING_SERIAL = 21
    __INVENTORY_SERIAL = 22

    __COLUMNS = 23
    __TEMPLATE_ROW = 7

    __GRADING_HEADER_ROW = 6

    __SCHED_FILL = PatternFill(
        start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
    )

    @classmethod
    def __area_formula(cls, row: int) -> str:
        return f"=G{row}*H{row}"

    @classmethod
    def __get_sorting_key(
        cls, reporting: Reporting
    ) -> Tuple[Tuple[bool, int], Tuple[bool, str], float]:
        oae_number = new_get_form_data(reporting, "oaeNumeroCodigoObra", default=None)
        agency = new_get_form_data(reporting, "codigoAgencia", default=None)
        return (
            (oae_number is None, oae_number),
            (agency is None, agency),
            reporting.km,
        )

    @classmethod
    def __get_reportings(cls, uuids: List[str]) -> List[Reporting]:
        metadata = (
            Reporting.objects.filter(uuid=uuids[0])
            .only("company__metadata")[0]
            .company.metadata
        )
        occurrence_type_uuid = UUID(metadata["oae_i_reporting_type_id"])
        reportings = list(
            Reporting.objects.filter(uuid__in=uuids)
            .only(
                "uuid",
                "number",
                "road_name",
                "km",
                "direction",
                "form_data",
                "parent__number",
                "parent__form_data",
                "company__logo",
                "company__provider_logo",
                "company__metadata",
            )
            .prefetch_related("company", "parent")
            .filter(
                occurrence_type__uuid=occurrence_type_uuid,
                form_data__has_key="inspection_year_campaign",
            )
        )

        return [
            r
            for r in reportings
            if isinstance(r.form_data["inspection_year_campaign"], int)
        ]

    @classmethod
    def __get_grade(cls, reporting: Reporting) -> str:
        grade_str = "-"
        grade = new_get_form_data(
            reporting, "notaTecnicaComentariosGerais", default="-"
        )

        if grade:
            try:
                grade_int = int(grade)
                if grade_int >= 1 and grade_int <= 5:
                    grade_str = str(grade)
            except Exception as e:
                print(e)

        return grade_str

    @classmethod
    def __get_previous_grade(cls, reporting: Reporting, previous_year: int) -> str:
        grade_str = "-"
        inventory: Reporting = reporting.parent
        previous_query_set = Reporting.objects.filter(
            parent__uuid=inventory.uuid,
            occurrence_type=reporting.occurrence_type,
            form_data__inspection_year_campaign=previous_year,
        ).only("form_data")
        if len(previous_query_set) == 1:
            grade_str = XlsxHandler.__get_grade(previous_query_set[0])
        elif len(previous_query_set) > 1:
            grade_str = "?"

        return grade_str

    @classmethod
    def __set_grading_header(cls, worksheet: Worksheet, year: int):
        row = XlsxHandler.__GRADING_HEADER_ROW
        worksheet.cell(row, XlsxHandler.__PREV_1 + 1).value = year - 1
        worksheet.cell(row, XlsxHandler.__PREV_2 + 1).value = year - 2
        worksheet.cell(row, XlsxHandler.__PREV_3 + 1).value = year - 3

    def __set_header(self, worksheet: Worksheet, sample_reporting: Reporting) -> None:
        road_name = sample_reporting.road_name
        header_cell: Cell = worksheet[XlsxHandler.__HEADER_CELL]
        header_cell.value = f"{header_cell.value} {road_name}"

        year = int(sample_reporting.form_data["inspection_year_campaign"])
        XlsxHandler.__set_grading_header(worksheet, year)

        try:
            logo_file = get_logo_file(self.s3, self.temp_dir, sample_reporting)
            insert_picture(
                worksheet,
                XlsxHandler.__LOGO_CELL,
                Image(logo_file),
                self.__sheet_target,
                resize_method=ResizeMethod.ProportionalRight,
                border_width=2,
            )
        except Exception as e:
            print(e)
        try:
            provider_logo_file = get_provider_logo_file(
                self.s3, self.temp_dir, sample_reporting
            )
            insert_picture(
                worksheet,
                XlsxHandler.__PROVIDER_LOGO_CELL,
                Image(provider_logo_file),
                self.__sheet_target,
                resize_method=ResizeMethod.ProportionalLeft,
                border_width=2,
            )
        except Exception as e:
            print(e)

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self.__report_format = report_format
        self.__sheet_target = sheet_target
        self.s3 = s3
        self.temp_dir = tempfile.mkdtemp()

        self.list_uuids: List[str] = list_uuids

    def __del__(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    @classmethod
    def __set_schedule(cls, worksheet: Worksheet, grade_str: str) -> None:
        sched_begin = XlsxHandler.__SCHED
        max_row = worksheet.max_row

        for col in range(sched_begin, sched_begin + 5):
            cell = worksheet.cell(max_row, col)
            cell.fill = PatternFill(fill_type=None)

        grade_int = -1
        try:
            grade_int = int(grade_str)
        except Exception:
            print(f"Grade {grade_str} not valid for schedule")

        if grade_int >= 1 and grade_int <= 5:
            sched_offset = min(grade_int, 4)
            sched_col = sched_begin + sched_offset
            sched_cell = worksheet.cell(max_row, sched_col)
            sched_cell.fill = XlsxHandler.__SCHED_FILL

    def __append_to_sheet(
        self, worksheet: Worksheet, reporting: Reporting, delete_template: bool
    ) -> None:
        year = reporting.form_data["inspection_year_campaign"]
        inv: Reporting = reporting.parent
        inventory_type_id = UUID(reporting.company.metadata["oae_i_inventory_type_id"])

        road_name = reporting.road_name if reporting.road_name else "-"

        row: List[str] = [""] * XlsxHandler.__COLUMNS
        oae_number = new_get_form_data(reporting, "oaeNumeroCodigoObra")
        row[XlsxHandler.__OAE_NUMBER] = oae_number if oae_number else "-"
        agency = new_get_form_data(reporting, "codigoAgencia")
        row[XlsxHandler.__AGENCY] = agency if agency else "-"
        name = new_get_form_data(reporting, "denominacao")
        row[XlsxHandler.__NAME] = name if name else "-"
        row[XlsxHandler.__ROAD] = road_name
        row[XlsxHandler.__KM] = get_km(reporting, "-")
        row[XlsxHandler.__DIRECTION] = get_direction(reporting, "-")

        info = new_get_form_data(
            reporting, "registroAnomaliasInformacoesComplementares"
        )
        row[XlsxHandler.__INFO] = info if info else "-"

        grade_str = XlsxHandler.__get_grade(reporting)
        row[XlsxHandler.__CURR] = grade_str

        row[XlsxHandler.__REPORTING_SERIAL] = reporting.number

        inventory_number = "-"
        length = "-"
        width = "-"
        type = "-"
        prev_3 = "-"
        prev_2 = "-"
        prev_1 = "-"

        if inv and inv.occurrence_type.uuid == inventory_type_id:
            inventory_number = inv.number
            length = new_get_form_data(inv, "comprimentoTotalOae", default="-") or "-"
            width = new_get_form_data(inv, "larguraTabuleiroOae", default="-") or "-"
            type = new_get_form_data(inv, "tremTipoClasseOae", default="-") or "-"

            prev_3 = XlsxHandler.__get_previous_grade(reporting, year - 3)
            prev_2 = XlsxHandler.__get_previous_grade(reporting, year - 2)
            prev_1 = XlsxHandler.__get_previous_grade(reporting, year - 1)

        row[XlsxHandler.__LEN] = length
        row[XlsxHandler.__WIDTH] = width
        row[XlsxHandler.__TYPE] = type

        row[XlsxHandler.__PREV_3] = prev_3
        row[XlsxHandler.__PREV_2] = prev_2
        row[XlsxHandler.__PREV_1] = prev_1

        if length == "-" or width == "-":
            row[XlsxHandler.__AREA] = "-"
        else:
            new_row = worksheet.max_row
            if not delete_template:
                new_row += 1
            row[XlsxHandler.__AREA] = XlsxHandler.__area_formula(new_row)

        row[XlsxHandler.__INVENTORY_SERIAL] = inventory_number
        append_row(worksheet, row)

        XlsxHandler.__set_schedule(worksheet, grade_str)

        if delete_template:
            worksheet.delete_rows(XlsxHandler.__TEMPLATE_ROW)

    def __create_workbook_file(
        self, road_name_year: str, reportings: List[Reporting]
    ) -> str:
        workbook = load_workbook(XlsxHandler.__TEMPLATE_FILE)
        worksheet = workbook[workbook.sheetnames[0]]

        sorted_reportings = sorted(
            reportings, key=lambda r: XlsxHandler.__get_sorting_key(r)
        )

        appended = False
        for reporting in sorted_reportings:
            self.__append_to_sheet(worksheet, reporting, not appended)
            appended = True

        self.__set_header(worksheet, reportings[0])

        workbook_name = f"Anexo I - {road_name_year}"
        workbook_name = clean_latin_string(
            workbook_name.replace(".", "").replace("/", "")
        )
        workbook_file = save_workbook(workbook_name, workbook)

        return workbook_file

    def execute(self) -> List[str]:
        reportings: List[Reporting] = XlsxHandler.__get_reportings(self.list_uuids)
        road_name_year_to_reportings: Dict[str, List[Reporting]] = {}

        for reporting in reportings:
            year = reporting.form_data["inspection_year_campaign"]
            road_name_year = f"{reporting.road_name} - {year}"
            if road_name_year not in road_name_year_to_reportings:
                road_name_year_to_reportings[road_name_year] = []
            road_name_year_to_reportings[road_name_year].append(reporting)

        workbook_files: List[str] = []
        for road_name_year, reportings in road_name_year_to_reportings.items():
            workbook_files.append(
                self.__create_workbook_file(road_name_year, reportings)
            )

        return workbook_files


class OAEI(CCRReport):
    def __init__(
        self,
        uuids: List[str] = None,
        report_format: ReportFormat = ReportFormat.XLSX,
    ) -> None:
        super().__init__(uuids, report_format)

    @classmethod
    def __get_reportings(cls, uuids: List[str]) -> List[Reporting]:
        metadata = (
            Reporting.objects.filter(uuid=uuids[0])
            .only("company__metadata")[0]
            .company.metadata
        )
        occurrence_type_uuid = UUID(metadata["oae_i_reporting_type_id"])
        reportings = list(
            Reporting.objects.filter(
                uuid__in=uuids,
                occurrence_type__uuid=occurrence_type_uuid,
                form_data__has_key="inspection_year_campaign",
            ).only(
                "uuid",
                "road_name",
                "form_data",
            )
        )

        return [
            r
            for r in reportings
            if isinstance(r.form_data["inspection_year_campaign"], int)
        ]

    def get_file_name(self) -> str:
        file_name: str = None

        reportings = OAEI.__get_reportings(self.uuids)

        if len(reportings) == 0:
            raise ValidationError("Nenhum apontamento válido")

        road_names = sorted({f"{r.road_name}" for r in reportings})
        years = list({f"{r.form_data['inspection_year_campaign']}" for r in reportings})

        extension = "zip"
        if len(road_names) > 1 or len(years) > 1:
            file_name = "Anexo I - {}".format("_".join(road_names))
        else:
            file_name = "Anexo I - {} - {}".format(road_names[0], years[0])
            extension = "xlsx" if self.report_format() == ReportFormat.XLSX else "pdf"
        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.{extension}"
        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandler(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])
        else:
            result_file = files[0]

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_oae_i_async_handler(
    reporter_dict: dict,
):
    reporter: OAEI = OAEI.from_dict(reporter_dict)
    reporter.export()
