import os
import tempfile
from collections import Counter
from datetime import timedelta
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km


class XlsxHandlerComparativeBuilds(object):

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    FIELDS = {
        "building": "A",
        "id_ccr_antt": "B",
        "km": "C",
        "direction": "D",
        "previous_conservation_state": "E",
        "conservation_state": "F",
        "actions": "G",
    }

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.wb = load_workbook(
            "./fixtures/reports/ccr_report_building_comparative.xlsx"
        )
        self.good_status_counter = 0
        self.regular_status_counter = 0
        self.bad_status_counter = 0
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.list_name_files = list()
        self._worksheet = self.wb.active
        self.resume_worksheet = self.wb["Resumo"]
        self.list_uuids = list_uuids
        first_reporting = Reporting.objects.filter(uuid__in=list_uuids).first()
        self.form = first_reporting.occurrence_type

        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="G1",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:B1",
            resize_method=ResizeMethod.ProportionalLeft,
        )

    def __insert_new_rows(self, row: int):
        chars = [chr(ord("A") + i) for i in range(7)]

        for char in chars:
            self._worksheet[f"{char}{row}"].border = XlsxHandlerComparativeBuilds.BORDER
            self._worksheet.row_dimensions[row].height = 20

    def create_dict(self, reporting):

        data = dict()

        data["id_ccr_antt"] = new_get_form_data(reporting, "idCcrAntt")
        data["building"] = new_get_form_data(
            reporting,
            "kindedificacaoInventario",
        )
        data["km"] = format_km(reporting, "km", 3)
        data["direction"] = get_custom_option(reporting, "direction")
        data["conservation_state"] = new_get_form_data(
            reporting,
            "generalConservationState",
        )
        data["road_name"] = reporting.road_name
        data["found_at"] = reporting.found_at

        previous_reporting = None
        if reporting.executed_at and reporting.parent:
            related_reportings = reporting.parent.children.all()
            if related_reportings:
                related_reportings = [r for r in related_reportings if r.executed_at]
                related_reportings.sort(key=lambda x: x.executed_at, reverse=True)
                previous_date = reporting.executed_at - timedelta(days=365)
                for r in related_reportings:
                    if (
                        str(r.uuid) != str(reporting.uuid)
                        and r.executed_at < reporting.executed_at
                        and r.executed_at >= previous_date
                        and r.occurrence_type == self.form
                        and r.km == reporting.km
                        and r.end_km == reporting.end_km
                        and r.direction == reporting.direction
                    ):
                        previous_reporting = r
                        break
        data["previous_conservation_state"] = (
            new_get_form_data(
                previous_reporting,
                "generalConservationState",
            )
            if previous_reporting
            else ""
        )
        if data["conservation_state"] and data["previous_conservation_state"]:
            string_states = (
                f"{data['conservation_state']}/{data['previous_conservation_state']}"
            )
            data["actions"] = self.__result_corrective_action(
                string_states,
                previous_reporting=previous_reporting,
            )

        for k, v in data.items():
            if v is None:
                data[k] = ""

        return data

    def fill_sheet(self, *, data_dict: dict):
        key_year = ""
        row = 6
        count = 0
        list_files = list()
        for key, values_list in data_dict.items():
            for values in values_list:
                for key_year, values_in_year in values.items():
                    values_in_year = XlsxHandlerComparativeBuilds.__sorted_values(
                        values_in_year
                    )
                    for items in values_in_year:
                        for internal_key, value in items.items():
                            if internal_key not in ["road_name", "found_at"]:
                                self.__insert_new_rows(row=row)
                                key_value = f"{XlsxHandlerComparativeBuilds.FIELDS[internal_key]}{row}"
                                self._worksheet[key_value] = value
                                XlsxHandlerComparativeBuilds.__format_fonts(
                                    cell=self._worksheet[key_value],
                                    size=11,
                                )
                        row += 1
                        count += 1
                if count == len(values_in_year):
                    self.list_name_files.append(key_year)
                    file_name = (
                        "Relatório ANTT - Comparativo Edificações - {}_{}".format(
                            items["road_name"], key_year
                        )
                    )
                    file_path = f"/tmp/{file_name}.xlsx"
                    row = 6
                    count = 0
                    self._fill_sheet_resume(file_name=file_path)
                    insert_logo_and_provider_logo(
                        worksheet=self._worksheet,
                        target=self.__sheet_target,
                        logo_company=self.data_logo_company,
                        provider_logo=self.data_provider_logo,
                    )
                    if self.__report_format == ReportFormat.PDF:
                        for row in range(6, self._worksheet.max_row + 1):
                            for col in range(1, 8):
                                cell = self._worksheet.cell(row, col)
                                XlsxHandlerComparativeBuilds.__format_fonts(
                                    cell=cell,
                                    size=11,
                                    wrap_text=True,
                                )
                                dimension = self._worksheet.row_dimensions[row]
                                dimension.height = None

                    self.wb.save(file_path)
                    list_files.append(file_path)

                    XlsxHandlerComparativeBuilds.__clear_all_data(
                        file_path,
                        self._worksheet,
                        initial_row=row,
                        min_col=1,
                        del_border=True,
                    )
                    XlsxHandlerComparativeBuilds.__clear_all_data(
                        file_path, self.resume_worksheet, initial_row=5, min_col=3
                    )

        temp_dir = tempfile.mkdtemp()
        file_name = self.__format_name_zip_file(
            "Relatório ANTT - Comparativo Edificações", self.list_name_files
        )
        path_file = os.path.join(temp_dir, file_name)
        saved_files = []
        if self.__report_format == ReportFormat.PDF:
            saved_files = convert_files_to_pdf(list_files)
        else:
            saved_files = list_files
        with ZipFile(path_file, "w") as zipObj:
            for file_path in saved_files:
                zipObj.write(file_path, os.path.basename(file_path))
        return path_file

    @classmethod
    def __sorted_values(cls, values: list):
        sorted_data = sorted(values, key=lambda x: (x["id_ccr_antt"], x["km"]))
        return sorted_data

    @classmethod
    def __clear_all_data(
        cls, file_path, sheet_name, initial_row, min_col, del_border=False
    ):
        wb = load_workbook(file_path)
        sheet = sheet_name

        for row in sheet.iter_rows(
            min_row=initial_row,
            max_row=sheet.max_row,
            min_col=min_col,
            max_col=sheet.max_column,
        ):
            for cell in row:
                cell.value = None
                if del_border:
                    cell.border = None

        wb.save(file_path)

    def _fill_sheet_resume(self, file_name):
        sheet = self.wb.active
        desired_column = "E"

        column_values = [
            celula.value for celula in sheet[desired_column] if celula.value
        ]
        column_values = Counter(column_values)
        self.resume_worksheet["C5"] = str(column_values.get("Regular", 0))
        self.resume_worksheet["C6"] = str(column_values.get("Bom", 0))
        self.resume_worksheet["C7"] = str(column_values.get("Ruim", 0))

    def __format_name_zip_file(self, name_string, list_name):
        if len(list_name) > 1:
            unique_names = sorted(set(list_name))
            name_roads = "-".join(map(str, unique_names))
        else:
            name_roads = list_name[0]

        name_zip_file = f"{name_string}-{name_roads}.zip"
        return name_zip_file

    def __result_corrective_action(
        self,
        string_state,
        previous_reporting,
    ):
        valid_states = [
            "Bom/Regular",
            "Bom/Ruim",
            "Regular/Regular",
            "Regular/Ruim",
            "Ruim/Regular",
            "Ruim/Ruim",
        ]
        if string_state in valid_states:
            return "Prevista ação de Recuperação"

        elif string_state == "Bom/Bom":
            return "SEM AÇÃO"
        else:
            corrective_action = self.__set_conservation_state(previous_reporting)

        return corrective_action

    def __set_conservation_state(self, reporting):
        join_result = ""
        foundation_structures = new_get_form_data(
            reporting,
            "foundationStructures",
        )
        floor_coating = new_get_form_data(reporting, "floorCoating")
        tile_coating = new_get_form_data(reporting, "tileCoating")
        sidewalk = new_get_form_data(reporting, "sidewalk")
        external_wall = new_get_form_data(reporting, "externalWall")
        internal_wall = new_get_form_data(reporting, "internalWall")
        metal_structure = new_get_form_data(reporting, "metalStructure")
        cealing = new_get_form_data(reporting, "cealing")
        climatization = new_get_form_data(reporting, "climatization")
        doors = new_get_form_data(reporting, "doors")
        windows = new_get_form_data(reporting, "windows")
        ilumination = new_get_form_data(reporting, "ilumination")
        electric_instalation = new_get_form_data(
            reporting,
            "electricInstalation",
        )
        sink_hidro_instalation = new_get_form_data(
            reporting,
            "sinkHidroInstalation",
        )
        toilet_hidro_instalation = new_get_form_data(
            reporting,
            "toiletHidroInstalation",
        )
        faucet_hidro_instalation = new_get_form_data(
            reporting,
            "faucetHidroInstalation",
        )
        landscape = new_get_form_data(reporting, "landscape")
        water_box = new_get_form_data(reporting, "waterBox")
        telephone_instalation = new_get_form_data(
            reporting,
            "telephoneInstalation",
        )
        external_painting = new_get_form_data(reporting, "externalPainting")
        internal_painting = new_get_form_data(reporting, "internalPainting")
        spda_protection = new_get_form_data(reporting, "spdaProtection")
        fences = new_get_form_data(reporting, "fences")
        utilitys = new_get_form_data(reporting, "utilitys")
        abnt_requirement = new_get_form_data(reporting, "abntRequirement")

        result_dict = {
            "01 Fundações e Estruturas": foundation_structures,
            "02 Revestimento de piso (cerâmico, polímero)": floor_coating,
            "03 Revestimento de azulejo (cerâmico, polímero)": tile_coating,
            "04 Calçada": sidewalk,
            "05 Parede externa": external_wall,
            "06 Parede interna": internal_wall,
            "07 Estrutura metálica": metal_structure,
            "08 Cobertura / Forro": cealing,
            "09 Climatização": climatization,
            "10 Portas": doors,
            "11 Janelas (vidro e armação metálica)": windows,
            "12 Iluminação (interna e externa)": ilumination,
            "13 Instalação elétrica (interna e externa)": electric_instalation,
            "14 Instalação hidrossanitária - Pia / Tanque de Lavar": sink_hidro_instalation,
            "15 Instalação hidrossanitária - Vaso Sanitário / Mictório": toilet_hidro_instalation,
            "16 Instalação hidrossanitária - Torneira / Registros/ chuveiros": faucet_hidro_instalation,
            "17 Paisagismo": landscape,
            "18 Caixa d'água": water_box,
            "19 instalação e telefonia": telephone_instalation,
            "20 Pintura Externa": external_painting,
            "21 Pintura interna": internal_painting,
            "22 Sistema de proteção de descarga atmosférica (SPDA)": spda_protection,
            "23 Cercas e alambrados": fences,
            "24 Utilidades (armários, gavetas)": utilitys,
            "25 Atendimento aos padrões de acessibilidade exigidos na NBR 9 050/2015 da ABNT": abnt_requirement,
        }

        result2 = []
        for state, value in result_dict.items():
            if result_dict[state] and value != "Nenhuma":
                result2.append(state)
        if result2:
            join_result = "| \n".join(result2)

        return join_result

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Calibri",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        wrap_text=False,
    ) -> None:
        cell.alignment = Alignment(
            wrap_text=wrap_text, horizontal=horizontal, vertical=vertical
        )
        cell.font = Font(name=name, sz=size, bold=bold)

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related(
            "occurrence_type",
            "firm",
            "firm__subcompany",
            "company",
            "parent",
            "parent__children",
            "parent__children__occurrence_type",
            "parent__children__company",
        )
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]
        data = [self.create_dict(reporting=reporting) for reporting in list_reporting]
        for item in data:
            road_name = item.get("road_name", None)
            found_at = item.get("found_at", None)

            if road_name in self.dict_filtered_roads:
                year_dict = next(
                    (
                        d
                        for d in self.dict_filtered_roads[road_name]
                        if found_at.year in d
                    ),
                    None,
                )

                if year_dict:
                    year_dict[found_at.year].append(item)
                else:
                    year_dict = {found_at.year: [item]}
                    self.dict_filtered_roads[road_name].append(year_dict)
            else:
                self.dict_filtered_roads[road_name] = [{found_at.year: [item]}]
        self.dict_filtered_roads = {
            k: v
            for k, v in self.dict_filtered_roads.items()
            if v and any(item for sublist in v for item in sublist)
        }

        self.data_logo_company["path_image"] = get_logo_file(
            s3=self.s3,
            temp_prefix=self.temp_file,
            reporting=list_reporting[0],
        )

        self.data_provider_logo["path_image"] = get_provider_logo_file(
            s3=self.s3,
            temp_prefix=self.temp_file,
            reporting=list_reporting[0],
        )

        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)

        return result_file


class CCrBuildingComparative(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = "Relatório ANTT - Comparativo Edificações"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        return f"{file_name}.zip"

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerComparativeBuilds(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            zipObj.write(files, files.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_building_comparative_async_handler(
    reporter_dict: dict,
):
    reporter = CCrBuildingComparative.from_dict(reporter_dict)
    reporter.export()
