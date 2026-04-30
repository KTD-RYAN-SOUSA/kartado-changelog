import os
import tempfile
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import ReportFormat, SheetTarget
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string, format_km


class XlsxHandlerBuildsDiagnostics(object):

    FIELDS = {
        "id_ccr_antt": "A",
        "buildings": "B",
        "km": "C",
        "direction": "D",
        "conservation_state": "E",
        "anomaly_identified": "F",
        "actions": "G",
    }

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.wb = load_workbook("./fixtures/reports/ccr_building_diagnostics.xlsx")
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.list_name_files = list()
        self._worksheet = self.wb.active
        self.form = OccurrenceType.objects.get(name="Monitoração de Edificações SGE")
        self.list_uuids = list_uuids
        self.reportings = Reporting.objects.filter(uuid=list_uuids[0]).prefetch_related(
            "company"
        )
        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company

        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

    def __insert_new_rows(self, row: int):
        chars = [chr(ord("A") + i) for i in range(7)]

        for char in chars:
            if char in ["H", "J"]:
                self._worksheet.row_dimensions[row].width = 15
            self._worksheet[f"{char}{row}"].border = XlsxHandlerBuildsDiagnostics.BORDER
            self._worksheet.row_dimensions[row].height = 20

    def create_dict(self, reporting):
        data = dict()

        data["id_ccr_antt"] = new_get_form_data(reporting, "idCcrAntt")
        data["buildings"] = new_get_form_data(
            reporting,
            "kindedificacaoInventario",
        )
        data["km"] = format_km(reporting, "km", 3)
        data["direction"] = get_custom_option(reporting, "direction")
        data["conservation_state"] = new_get_form_data(
            reporting,
            "generalConservationState",
        )
        data["road_name"] = reporting.road_name
        data["found_at"] = reporting.found_at

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def fill_sheet(self, *, data_dict: dict):
        key_year = ""
        row = 4
        count = 0
        list_files = list()
        for key, values_list in data_dict.items():
            for values in values_list:
                for key_year, values_in_year in values.items():
                    values_in_year = XlsxHandlerBuildsDiagnostics.sorted_values(
                        values_in_year
                    )
                    for items in values_in_year:
                        for internal_key, value in items.items():
                            if internal_key not in ["road_name", "found_at"]:
                                self.__insert_new_rows(row=row)
                                key_value = f"{XlsxHandlerBuildsDiagnostics.FIELDS[internal_key]}{row}"
                                self._worksheet[key_value] = value
                                XlsxHandlerBuildsDiagnostics.__format_fonts(
                                    cell=self._worksheet[key_value],
                                    size=11,
                                )
                        row += 1
                        count += 1
                if count == len(values_in_year):
                    self.list_name_files.append(key_year)
                    file_name = (
                        "Relatório ANTT - Diagnóstico Edificações - {}-{}".format(
                            items["road_name"], key_year
                        )
                    )
                    file_name = clean_latin_string(
                        file_name.replace(".", "").replace("/", "")
                    )
                    file_path = f"/tmp/{file_name}.xlsx"
                    row = 4
                    count = 0

                    if self.__report_format == ReportFormat.PDF:
                        for row in range(4, self._worksheet.max_row + 1):
                            for col in range(1, 8):
                                cell = self._worksheet.cell(row, col)
                                XlsxHandlerBuildsDiagnostics.__format_fonts(
                                    cell=cell,
                                    size=11,
                                    wrap_text=True,
                                )
                                dimension = self._worksheet.row_dimensions[row]
                                dimension.height = None

                    self.wb.save(file_path)
                    list_files.append(file_path)
                    XlsxHandlerBuildsDiagnostics.clear_all_data(
                        file_path, self._worksheet, row
                    )

        temp_dir = tempfile.mkdtemp()
        file_name = self.__format_name_zip_file(
            "Relatório ANTT - Diagnóstico Edificações", self.list_name_files
        )
        path_file = os.path.join(temp_dir, file_name)
        saved_files = []
        if self.__report_format == ReportFormat.PDF:
            saved_files = convert_files_to_pdf(list_files)
        else:
            saved_files = list_files
        with ZipFile(path_file, "w") as zipObj:
            for file_path in saved_files:
                zipObj.write(file_path, os.path.basename(file_path))

        return path_file

    @classmethod
    def sorted_values(cls, values: list):
        sorted_data = sorted(values, key=lambda x: (x["buildings"], x["km"]))
        return sorted_data

    def __format_name_zip_file(self, name_string, list_name):
        if len(list_name) > 1:
            unique_names = sorted(set(list_name))
            name_roads = "-".join(map(str, unique_names))
        else:
            name_roads = list_name[0]

        name_zip_file = f"{name_string}-{name_roads}.zip"
        return name_zip_file

    @classmethod
    def clear_all_data(cls, file_path, sheet_name, initial_row):
        wb = load_workbook(file_path)
        sheet = sheet_name

        for row in sheet.iter_rows(
            min_row=initial_row,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        ):
            for cell in row:
                cell.value = None
                cell.border = None
        wb.save(file_path)

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        wrap_text=False,
    ) -> None:
        cell.alignment = Alignment(
            wrap_text=wrap_text, horizontal=horizontal, vertical=vertical
        )
        cell.font = Font(name=name, sz=size, bold=bold)

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form, uuid__in=self.list_uuids
        ).prefetch_related("company", "occurrence_type", "firm", "firm__subcompany")
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]
        data = [self.create_dict(reporting=reporting) for reporting in list_reporting]

        for item in data:
            road_name = item.get("road_name", None)
            found_at = item.get("found_at", None)

            if road_name in self.dict_filtered_roads:
                year_dict = next(
                    (
                        d
                        for d in self.dict_filtered_roads[road_name]
                        if found_at.year in d
                    ),
                    None,
                )

                if year_dict:
                    year_dict[found_at.year].append(item)
                else:
                    year_dict = {found_at.year: [item]}
                    self.dict_filtered_roads[road_name].append(year_dict)
            else:
                self.dict_filtered_roads[road_name] = [{found_at.year: [item]}]
        self.dict_filtered_roads = {
            k: v
            for k, v in self.dict_filtered_roads.items()
            if v and any(item for sublist in v for item in sublist)
        }

        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)
        return result_file


class CCrBuildingDiagnostics(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):

        file_name = "Relatório ANTT - Diagnóstico Edificações"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        return f"{file_name}.zip"

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerBuildsDiagnostics(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            zipObj.write(files, files.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_building_diagnostics_async_handler(
    reporter_dict: dict,
):
    reporter = CCrBuildingDiagnostics.from_dict(reporter_dict)
    reporter.export()
