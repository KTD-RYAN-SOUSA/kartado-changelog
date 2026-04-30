import math
import random
import shutil
import string
import tempfile
import time
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_reporting_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_picture,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import ThreadExecutor, synchronized_request_pdf
from helpers.apps.ccr_report_utils.reporting_utils import (
    get_custom_option,
    get_end_km,
    get_km,
)
from helpers.strings import clean_latin_string, get_obj_from_path
from RoadLabsAPI.storage_backends import PrivateMediaStorage

storage = PrivateMediaStorage()


class XlsxHandler:
    _LOGO_CELL = "N2:O2"
    _PROVIDER_LOGO_CELL = "B2"

    def __init__(
        self,
        s3,
        uuid: str,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        self.uuid = uuid
        self.s3 = s3
        self.__sheet_target = sheet_target
        self.temp_file = tempfile.mkdtemp()
        self.__xlsx_file = "./fixtures/reports/ccr_lighting.xlsx"
        self._workbook = load_workbook(self.__xlsx_file)
        self._worksheet = self._workbook["Iluminação"]
        self.reporting = Reporting.objects.get(uuid=uuid)
        self.__static_fields = {
            "identificacaoiluminacao": "C4",
            "tipoiluminacao_a": "J4",
            "road_name": "C5",
            "km": "I5",
            "end_km": "M5",
            "direction": "C6",
            "municipio": "I6",
            "latitude": "E9",
            "latitudefim": "L9",
            "longitude": "E10",
            "longitudefim": "L10",
            "tipoiluminacao_b": "E15",
            "tipoposte": {
                "CONCRETO": "F17",
                "MADEIRA": "I17",
                "METAL": "L17",
                "METAL CURVO": "O17",
            },
            "modeloposte": "D19",
            "modeloluminaria": {
                "petala": "F21",
                "braço com petala": "I21",
                "refletor": "L21",
                "outro": "O21",
            },
            "tipolampada": {
                "led": "F23",
                "sodio": "I23",
                "mercurio": "L23",
                "outro": "O23",
            },
            "potencialampada": {
                "400w": "F25",
                "250w": "I25",
                "150w": "L25",
                "outro": "O25",
            },
            "qtdluminarias": "D27",
            "qtdlampadas": "J27",
            "qtdpostes": "O27",
            "qtdpontosdefeito": "D29",
            "quantidadeanomaliasposte": "J29",
            "quantidadepontosapagados": "O29",
            "conservacaoiluminacao": {"bom": "I31", "regular": "L31", "ruim": "O31"},
            "atendimentoabnt": "F33",
            "images": {
                "diurna_1": [
                    "B39:G39",
                ],
                "diurna_2": [
                    "B41:G41",
                ],
                "noturna_1": [
                    "I39:O39",
                ],
                "noturna_2": [
                    "I41:O41",
                ],
                "anomalia_1": ["B45", "B46"],
                "anomalia_2": ["I45", "I46"],
            },
            "observacoes": "C48",
        }

    @classmethod
    def _insert_logos(
        cls,
        s3,
        temp_dir: str,
        sheet_target: SheetTarget,
        worksheet: Worksheet,
        reporting: Reporting,
    ) -> None:
        logo = get_logo_file(s3, temp_dir, reporting)
        provider_logo = get_provider_logo_file(s3, temp_dir, reporting)

        try:
            insert_picture_2(
                worksheet,
                cls._LOGO_CELL,
                Image(logo),
                sheet_target,
                border_width=(5, 5, 5, 5),
                resize_method=ResizeMethod.ProportionalCentered,
            )
        except Exception:
            pass
        try:
            insert_picture_2(
                worksheet,
                cls._PROVIDER_LOGO_CELL,
                Image(provider_logo),
                sheet_target,
                border_width=(5, 5, 5, 5),
                resize_method=ResizeMethod.ProportionalCentered,
            )
        except Exception:
            pass

    def __change_border(
        self,
        start_col: int,
        end_col: int,
        start_row: int,
        end_row: int,
        sides: list,
        remove: dict = {"row": [], "col": []},
        keep: bool = False,
    ):
        border: Border = None
        if not keep:
            border = Border(
                left=Side(style="thin") if "left" in sides else None,
                right=Side(style="thin") if "right" in sides else None,
                top=Side(style="thin") if "top" in sides else None,
                bottom=Side(style="thin") if "bottom" in sides else None,
            )
        for col in range(start_col, end_col):
            if col not in remove["col"]:
                for row in range(start_row, end_row):
                    if row not in remove["row"]:
                        cell = self._worksheet.cell(row=row, column=col)
                        if keep:
                            left_s = getattr(cell.border, "left")
                            right_s = getattr(cell.border, "right")
                            top_s = getattr(cell.border, "top")
                            bottom_s = getattr(cell.border, "bottom")
                            border = Border(
                                left=Side(style="thin")
                                if "left" in sides
                                else (left_s if left_s else None),
                                right=Side(style="thin")
                                if "right" in sides
                                else (right_s if right_s else None),
                                top=Side(style="thin")
                                if "top" in sides
                                else (top_s if top_s else None),
                                bottom=Side(style="thin")
                                if "bottom" in sides
                                else (bottom_s if bottom_s else None),
                            )
                        cell.border = border

    def __anomalies_found(self, anomalies_found: list) -> dict:
        n_boards = math.ceil(len(anomalies_found) / 2)
        init_row = index_value = 45
        anomalies_dict = {}
        for _ in range(0, n_boards):
            self._worksheet.row_dimensions[init_row].height = 150
            self._worksheet.row_dimensions[init_row + 1].height = 9.75
            self._worksheet.row_dimensions[init_row + 2].height = 4
            self._worksheet.merge_cells(f"B{init_row}:G{init_row}")
            self._worksheet.merge_cells(f"I{init_row}:O{init_row}")
            self._worksheet.merge_cells(f"B{init_row+1}:G{init_row+1}")
            self._worksheet.merge_cells(f"I{init_row+1}:O{init_row+1}")
            self.__change_border(
                start_col=2,
                end_col=16,
                start_row=init_row,
                end_row=init_row + 2,
                sides=["left", "top", "bottom", "right"],
                remove={"col": [8], "row": []},
            )
            self.__change_border(
                start_col=16,
                end_col=17,
                start_row=init_row,
                end_row=init_row + 3,
                sides=["right"],
                remove={"col": [], "row": []},
            )
            init_row += 3
        for i in range(0, len(anomalies_found), 2):
            key_a = f"B{index_value}"
            key_b = f"I{index_value}"
            if len(anomalies_found) > i + 1:
                anomalies_dict[anomalies_found[i]["uuid"]] = key_a
                anomalies_dict[anomalies_found[i + 1]["uuid"]] = key_b
            else:
                anomalies_dict[anomalies_found[i]["uuid"]] = key_a
            index_value += 3
        self._worksheet.row_dimensions[init_row].height = 30
        self._worksheet.merge_cells(f"C{init_row}:O{init_row}")
        self.__change_border(
            start_col=2,
            end_col=16,
            start_row=init_row,
            end_row=init_row + 1,
            sides=["left", "top", "bottom", "right"],
            remove={"col": [], "row": []},
        )
        self.__change_border(
            start_col=16,
            end_col=17,
            start_row=init_row,
            end_row=init_row + 1,
            sides=["right"],
            remove={"col": [], "row": []},
        )
        init_row += 1
        self._worksheet.row_dimensions[init_row].height = 4
        self.__change_border(
            start_col=1,
            end_col=16,
            start_row=init_row,
            end_row=init_row + 1,
            sides=[
                "bottom",
            ],
            remove={"col": [], "row": []},
        )
        self.__change_border(
            start_col=16,
            end_col=17,
            start_row=init_row,
            end_row=init_row + 1,
            sides=["bottom", "right"],
            remove={"col": [], "row": []},
        )
        self._worksheet[f"B{init_row-1}"] = "Observação:"
        self._worksheet[f"B{init_row-1}"].font = Font(size=8, italic=True)
        self._worksheet[f"B{init_row-1}"].alignment = Alignment(
            horizontal="left", vertical="top"
        )
        anomalies_dict["obs"] = init_row - 1
        return anomalies_dict

    def __insert_value_change_color(
        self, cell, value, color="808080", alignment="center", bold=True
    ):
        if value:
            self._worksheet[cell] = "X"
            self._worksheet[cell].fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            self._worksheet[cell].alignment = Alignment(
                horizontal=alignment, vertical=alignment
            )
            self._worksheet[cell].font = Font(bold=bold)

    def fill_sheet(self, values: dict):
        field_name_outros = {
            "modeloluminaria": "N21",
            "tipolampada": "N23",
            "potencialampada": "N25",
        }
        if len(values["anomalies_photos"]) == 0:
            values["anomalies_photos"] = [
                {"path": "", "description": "", "uuid": ""},
                {"path": "", "description": "", "uuid": ""},
            ]

        anomalies_dict = self.__anomalies_found(
            anomalies_found=values["anomalies_photos"]
        )
        for key, value in values.items():
            if key in [
                "tipoposte",
                "modeloluminaria",
                "tipolampada",
                "potencialampada",
                "conservacaoiluminacao",
            ]:
                for _key, _value in value.items():
                    self.__insert_value_change_color(
                        cell=self.__static_fields[key][_key], value=_value
                    )
            elif key == "images":
                for _key, _value in value.items():
                    if _value:
                        cell_str = self.__static_fields[key][_key][0]
                        insert_picture(
                            self._worksheet,
                            cell_str,
                            Image(_value["path"]),
                            self.__sheet_target,
                        )
            elif key == "anomalies_photos":
                for item in value:
                    if item.get("uuid") in anomalies_dict.keys():
                        uuid = item.get("uuid")
                        cell = str(anomalies_dict[uuid])
                        cell_row, cell_col = coordinate_to_tuple(cell)
                        cell_col += 6
                        range_str = f"{cell}:{get_column_letter(cell_col)}{cell_row}"
                        if item["path"]:
                            insert_picture(
                                self._worksheet,
                                range_str,
                                Image(item["path"]),
                                self.__sheet_target,
                            )
                        col = "".join(c for c in cell if c.isalpha())
                        row = "".join(c for c in cell if c.isdigit())
                        if row.isdigit() and col.isalpha():
                            row = int(row)
                            self._worksheet[f"{col}{row+1}"] = item["description"]
                            self._worksheet[f"{col}{row+1}"].alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                            self._worksheet[f"{col}{row+1}"].font = Font(
                                bold=True, size=10
                            )
            elif key == "observacoes":
                self._worksheet[f"C{anomalies_dict['obs']}"] = value
                self._worksheet[f"C{anomalies_dict['obs']}"].font = Font(size=10)
                self._worksheet[f"C{anomalies_dict['obs']}"].alignment = Alignment(
                    horizontal="left", vertical="top", wrapText=True
                )
            elif key == "field_name_outros":
                for _key, _value in value.items():
                    text_x = [
                        True
                        for k, v in values[_key].items()
                        if k == "outro" and v.upper() == "X"
                    ]
                    if True in text_x:
                        text = values[key][_key]
                        self._worksheet[field_name_outros[_key]] = f"Outros({text})"
            else:
                self._worksheet[self.__static_fields[key]] = value

        self.__change_border(
            1, 2, 1, self._worksheet.max_row + 1, sides=["left"], keep=True
        )

    def create_dict(self, s3) -> dict:
        reporting = Reporting.objects.filter(uuid=self.uuid).first()
        direction = get_custom_option(reporting, "direction")
        reporting_files = download_reporting_pictures(
            s3,
            self.temp_file,
            reporting,
            width=337,
            height=242,
            enable_include_dnit=False,
            enable_is_shared_antt=True,
        )
        if reporting_files.get("images"):
            anomalies_photos_uuids = []
            no_treatment_images = []
            descriptions = {}
            if new_get_form_data(reporting, "therapy"):
                for anomalies in new_get_form_data(reporting, "therapy"):
                    if "treatment_images" in anomalies:
                        anomalies_photos_uuids += anomalies["treatment_images"]
                        for uuid in anomalies["treatment_images"]:
                            descriptions[uuid] = anomalies.get("description", "")
                    else:
                        chars = string.ascii_letters + string.digits
                        id_part_1 = "".join(random.choice(chars) for _ in range(20))
                        id_part_2 = "".join(random.choice(chars) for _ in range(20))
                        id_part_3 = "".join(random.choice(chars) for _ in range(20))
                        no_treatment_images.append(
                            {
                                "path": "",
                                "description": anomalies.get("description", ""),
                                "uuid": f"{id_part_1}-{id_part_2}-{id_part_3}",
                            }
                        )

            anomalies_photos = [
                obj
                for obj in reporting_files["images"]
                if obj["uuid"] in anomalies_photos_uuids
            ]
            for item in anomalies_photos:
                if item["uuid"] in descriptions.keys():
                    item.update({"description": descriptions[item["uuid"]]})
            anomalies_photos = anomalies_photos + no_treatment_images
        else:
            anomalies_photos = []
        if new_get_form_data(reporting, "fotos_monitoracao"):
            (imagem_diurna_uuid, imagem_noturna_uuid) = [], []
            for obj in new_get_form_data(reporting, "fotos_monitoracao"):
                imagem_diurna_uuid += (
                    obj["imagem_diurna"] if "imagem_diurna" in obj else ""
                )
                imagem_noturna_uuid += (
                    obj["imagem_noturna"] if "imagem_noturna" in obj else ""
                )
            imagem_diurna = [
                obj
                for obj in reporting_files["images"]
                if obj["uuid"] in imagem_diurna_uuid
            ]
            imagem_noturna = [
                obj
                for obj in reporting_files["images"]
                if obj["uuid"] in imagem_noturna_uuid
            ]
        else:
            imagem_diurna = []
            imagem_noturna = []
        tipoiluminacao = new_get_form_data(reporting, "tipoiluminacao")
        _tipoposte = str(new_get_form_data(reporting, "tipoposte")).upper()
        _modeloluminaria = str(new_get_form_data(reporting, "modeloluminaria")).lower()
        _tipolampada = str(new_get_form_data(reporting, "tipolampada")).lower()
        _potencialampada = str(new_get_form_data(reporting, "potencialampada")).lower()
        _conservacaoiluminacao = str(
            new_get_form_data(reporting, "general_conservation_state")
        )
        tipoposte = {
            "CONCRETO": "X" if "CONCRETO" in _tipoposte else "",
            "MADEIRA": "X" if "MADEIRA" in _tipoposte else "",
            "METAL": "X" if "METAL" in _tipoposte else "",
            "METAL CURVO": "X" if "METAL CURVO" in _tipoposte else "",
        }
        modeloluminaria = {
            "petala": "X" if "pétala" in _modeloluminaria else "",
            "braço com petala": "X" if "braço com pétala" in _modeloluminaria else "",
            "refletor": "X" if "refletor" in _modeloluminaria else "",
            "outro": "X" if "outro" in _modeloluminaria else "",
        }
        tipolampada = {
            "led": "X" if "led" in _tipolampada else "",
            "sodio": "X" if "sódio" in _tipolampada else "",
            "mercurio": "X" if "mercurio" in _tipolampada else "",
            "outro": "X" if "outro" in _tipolampada else "",
        }
        potencialampada = {
            "400w": "X" if "400" in _potencialampada else "",
            "250w": "X" if "250" in _potencialampada else "",
            "150w": "X" if "150" in _potencialampada else "",
            "outro": "X" if "outro" in _potencialampada else "",
        }
        conservacaoiluminacao = {
            "bom": "X" if "Bom" == _conservacaoiluminacao else "",
            "regular": "X" if "Regular" == _conservacaoiluminacao else "",
            "ruim": "X" if "Ruim" == _conservacaoiluminacao else "",
        }
        _modeloluminariaoutro = new_get_form_data(reporting, "modeloluminariaoutro")
        _tipolampadaoutro = new_get_form_data(reporting, "tipolampadaoutro")
        _potencialampadaoutro = new_get_form_data(reporting, "potencialampadaoutro")
        field_name_outros = {
            "modeloluminaria": (
                _modeloluminariaoutro if _modeloluminariaoutro else "____"
            ),
            "tipolampada": _tipolampadaoutro if _tipolampadaoutro else "____",
            "potencialampada": (
                _potencialampadaoutro if _potencialampadaoutro else "____"
            ),
        }
        qtdpostes_monit = new_get_form_data(reporting, "qtdpostes_monit")
        qtdpostes = new_get_form_data(reporting, "qtdpostes")
        data = {
            "anomalies_photos": anomalies_photos,
            "field_name_outros": field_name_outros,
            "identificacaoiluminacao": new_get_form_data(reporting, "id_ccr_antt"),
            "tipoiluminacao_a": tipoiluminacao,
            "road_name": reporting.road_name,
            "km": get_km(reporting),
            "end_km": get_end_km(reporting),
            "direction": direction,
            "municipio": new_get_form_data(reporting, "city"),
            "latitude": new_get_form_data(reporting, "latitude"),
            "latitudefim": new_get_form_data(reporting, "latitude_end"),
            "longitude": new_get_form_data(reporting, "longitude"),
            "longitudefim": new_get_form_data(reporting, "longitude_end"),
            "tipoiluminacao_b": tipoiluminacao,
            "tipoposte": tipoposte,
            "modeloposte": new_get_form_data(reporting, "modeloposte"),
            "modeloluminaria": modeloluminaria,
            "tipolampada": tipolampada,
            "potencialampada": potencialampada,
            "qtdluminarias": new_get_form_data(reporting, "qtdluminarias"),
            "qtdlampadas": new_get_form_data(reporting, "qtdlampadas"),
            "qtdpostes": qtdpostes_monit or qtdpostes,
            "qtdpontosdefeito": new_get_form_data(reporting, "qtdpontosdefeito"),
            "quantidadeanomaliasposte": new_get_form_data(
                reporting, "quantidadeposteanomalia"
            ),
            "quantidadepontosapagados": new_get_form_data(
                reporting, "quantidadepostesapagados"
            ),
            "conservacaoiluminacao": conservacaoiluminacao,
            "atendimentoabnt": new_get_form_data(reporting, "atendimentoabnt"),
            "images": {
                "diurna_1": imagem_diurna[0] if len(imagem_diurna) > 0 else "",
                "diurna_2": imagem_diurna[1] if len(imagem_diurna) > 1 else "",
                "noturna_1": imagem_noturna[0] if len(imagem_noturna) > 0 else "",
                "noturna_2": imagem_noturna[1] if len(imagem_noturna) > 1 else "",
            },
            "observacoes": new_get_form_data(reporting, "observacoes"),
        }
        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def execute(self):
        data = self.create_dict(s3=self.s3)

        file_name_options = [
            get_obj_from_path(self.reporting.form_data, "id_ccr_antt"),
            self.reporting.number,
            str(time.time()),
        ]
        file_name = clean_latin_string(
            next(a for a in file_name_options if a).replace(".", "").replace("/", "")
        )

        result = f"/tmp/{file_name}.xlsx"

        self.fill_sheet(values=data)
        self._insert_logos(
            self.s3,
            self.temp_file,
            self.__sheet_target,
            self._worksheet,
            self.reporting,
        )
        self._workbook.save(result)
        shutil.rmtree(self.temp_file, ignore_errors=True)

        return result


class CCRLighting(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = ""

        if len(self.uuids) == 1:
            reporting = Reporting.objects.get(uuid=self.uuids[0])
            file_name_options = [
                get_obj_from_path(reporting.form_data, "id_ccr_antt"),
                reporting.number,
                str(time.time()),
            ]
            file_name = clean_latin_string(
                next(a for a in file_name_options if a)
                .replace(".", "")
                .replace("/", "")
            )
            extension: str = None
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{file_name}.{extension}"

        elif len(self.uuids) > 1:
            file_name = "{} Relatorios - Iluminacao.zip".format(len(self.uuids))

        file_name = clean_latin_string(file_name)

        return file_name

    def export(self):
        convert_executor: ThreadExecutor = None
        if self.report_format() == ReportFormat.PDF:
            convert_executor = ThreadExecutor(25)
        s3 = get_s3()
        files = []
        for uuid in self.uuids:
            xlsx_file = XlsxHandler(
                uuid=uuid, s3=s3, sheet_target=self.sheet_target()
            ).execute()
            if self.report_format() == ReportFormat.PDF:
                convert_executor.submit(synchronized_request_pdf, xlsx_file)
            else:
                files.append(xlsx_file)

        if self.report_format() == ReportFormat.PDF:
            files = convert_executor.get()

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_lighting_async_handler(reporter_dict: dict):
    reporter = CCRLighting.from_dict(reporter_dict)
    reporter.export()
