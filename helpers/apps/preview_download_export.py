import json
import os
import random
import shutil
import string
from datetime import timedelta

import boto3
import pytz
from django.conf import settings
from django.db.models import F, Prefetch
from django.utils.timezone import now
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from zappa.asynchronous import task

from apps.daily_reports.models import (
    DailyReportContractUsage,
    DailyReportEquipment,
    DailyReportVehicle,
    DailyReportWorker,
)
from apps.files.models import GenericFile
from apps.resources.models import Contract
from apps.service_orders.models import ProcedureResource
from helpers.dates import utc_to_local
from RoadLabsAPI.settings import credentials


class PreviewDownloadExport:
    def __init__(
        self,
        file_uuid="",
        filename=None,
        object_name=None,
        queryset=[],
        company_name=None,
        contract_uuid="",
        work_days=1,
    ):

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )
        if file_uuid:
            generic_file = GenericFile.objects.get(uuid=file_uuid)
            uuid_list = json.loads(generic_file.file.read())

            procedure_uuid_list = uuid_list[0]
            procedure_prefetch_fields_list = [
                "service_order_resource",
                "service_order_resource__contract",
                "service_order_resource__contract__firm",
                "service_order_resource__contract__subcompany",
                "service_order_resource__entity",
                "service_order_resource__resource_contract_unit_price_items",
                "service_order_resource__resource_contract_unit_price_items__contract_item_unit_price_services",
                "resource",
                "created_by",
                "measurement_bulletin",
                "approved_by",
                "reporting",
            ]
            daily_uuid_list = uuid_list[1]
            daily_prefetch_fields_list = [
                self.get_contract_usage_prefetch_object("worker", DailyReportWorker),
                self.get_contract_usage_prefetch_object(
                    "equipment", DailyReportEquipment
                ),
                self.get_contract_usage_prefetch_object("vehicle", DailyReportVehicle),
            ]

            self.procedure_queryset = ProcedureResource.objects.filter(
                pk__in=set(procedure_uuid_list)
            ).prefetch_related(*procedure_prefetch_fields_list)

            self.daily_queryset = DailyReportContractUsage.objects.filter(
                pk__in=set(daily_uuid_list)
            ).prefetch_related(*daily_prefetch_fields_list)

        self.company_name = company_name
        if filename:
            self.filename = filename
        else:
            self.filename = self.get_filename()

        if object_name:
            self.object_name = object_name
        else:
            self.object_name = self.get_object_name()

        if contract_uuid:
            self.contract = Contract.objects.get(uuid=contract_uuid)

        self.work_days = work_days

        self.approval_status_translation = {
            "APPROVED_APPROVAL": "Aprovado",
            "DENIED_APPROVAL": "Rejeitado",
            "WAITING_APPROVAL": "Pendente",
        }

    def get_random_string(self):
        return "".join(
            random.SystemRandom().choice(string.ascii_lowercase + string.digits)
            for _ in range(10)
        )

    def get_filename(self):
        filename = "[Kartado] Prévia de valores de medição"

        return f"{filename}.xlsx"

    def get_object_name(self):
        random_string = self.get_random_string()
        object_name = "media/private/{}_{}.xlsx".format(
            self.filename.split(".xlsx")[0], random_string
        )

        return object_name

    def get_s3_url(self):
        empty = {"url": "", "name": ""}

        url = self.s3.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                "Key": self.object_name,
            },
        )

        if not url:
            return empty

        return url

    def upload_file(self, path):
        expires = now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)

        try:
            self.s3.upload_file(
                path,
                settings.AWS_STORAGE_BUCKET_NAME,
                self.object_name,
                ExtraArgs={"Expires": expires},
            )
        except Exception:
            return

    def copy_and_rename(self, old_folder, new_folder, temp_file):
        os.makedirs(new_folder, exist_ok=True)
        shutil.copy(old_folder + temp_file, new_folder + temp_file)
        os.rename(
            new_folder + temp_file,
            new_folder + self.filename + ".xlsx",
        )
        return

    def return_file_path(self, new_folder):
        return new_folder + self.filename + ".xlsx"

    def load_file(self, new_file_path):
        wb = load_workbook(filename=new_file_path)
        return wb

    def get_contract_usage_prefetch_object(self, relation_name, model):

        return Prefetch(
            relation_name,
            queryset=model.objects.all()
            .annotate(
                resource_name=F(
                    "contract_item_administration__resource__resource__name"
                ),
                resource_unit=F(
                    "contract_item_administration__resource__resource__unit"
                ),
                resource_unit_price=F(
                    "contract_item_administration__resource__unit_price"
                ),
            )
            .prefetch_related(
                "multiple_daily_reports",
                "multiple_daily_reports__created_by",
                "measurement_bulletin",
                "contract_item_administration",
                "contract_item_administration__entity",
                "contract_item_administration__contract_item_administration_services",
                "created_by",
            ),
        )

    def get_board_instance(self, obj):

        if obj.worker:
            return obj.worker
        elif obj.equipment:
            return obj.equipment
        elif obj.vehicle:
            return obj.vehicle
        return None

    def get_contract_data(self):

        object_number = (
            self.contract.extra_info.get("r_c_number", "")
            if self.contract and self.contract.extra_info
            else ""
        )

        try:
            subcompany_name = self.contract.subcompany.name
        except Exception:
            try:
                subcompany_name = self.contract.firm.name
            except Exception:
                subcompany_name = ""

        contract_name = self.contract.name

        start_date = self.contract.contract_start

        end_date = self.contract.contract_end

        accounting = (
            self.contract.extra_info.get("accounting_classification", "")
            if self.contract and self.contract.extra_info
            else ""
        )

        contract_status = (
            self.contract.status.name if self.contract and self.contract.status else ""
        )

        performance_months = self.contract.performance_months

        return [
            object_number,
            subcompany_name,
            contract_name,
            start_date,
            end_date,
            accounting,
            contract_status,
            performance_months,
        ]

    def fill_contract_info(self):
        data_contract = self.get_contract_data()

        cabecalho = self.wb["OBJETO"]

        data_row = 2

        for column_index, data in enumerate(data_contract):
            data_cell = cabecalho.cell(row=data_row, column=column_index + 1)
            data_cell.value = data
            if get_column_letter(column_index + 1) in ["D", "E"]:
                data_cell.number_format = "dd/mm/yyyy"

        return

    def get_procedure_resource_data(self, procedure_resource):
        resource = procedure_resource.resource
        service_order_resource = procedure_resource.service_order_resource
        items_list = list(
            procedure_resource.service_order_resource.resource_contract_unit_price_items.all()
        )
        contract_item_unit_price = items_list[0] if items_list else None

        section = ""
        if contract_item_unit_price:
            services_list = list(
                contract_item_unit_price.contract_item_unit_price_services.all()
            )
            if services_list:
                section = services_list[0].description
            elif contract_item_unit_price.contract_item_unit_price_services.exists():
                section = (
                    contract_item_unit_price.contract_item_unit_price_services.first().description
                )

        sort_string = (
            contract_item_unit_price.sort_string if contract_item_unit_price else ""
        )

        name = resource.name if resource else ""
        unit = resource.unit if resource else ""
        entity_name = (
            service_order_resource.entity.name
            if service_order_resource and service_order_resource.entity
            else ""
        )
        creation_date = utc_to_local(procedure_resource.creation_date).strftime(
            "%d/%m/%Y"
        )
        created_by = (
            procedure_resource.created_by.get_full_name()
            if procedure_resource.created_by
            else ""
        )
        reporting_number = (
            procedure_resource.reporting.number if procedure_resource.reporting else ""
        )

        approval_status = self.approval_status_translation.get(
            procedure_resource.approval_status, ""
        )
        approval_date = (
            utc_to_local(procedure_resource.approval_date).strftime("%d/%m/%Y %H:%M")
            if procedure_resource.approval_date
            else None
        )
        bulletin_number = (
            procedure_resource.measurement_bulletin.number
            if procedure_resource.measurement_bulletin
            else ""
        )
        return [
            "PREÇO UNITÁRIO",
            section,
            sort_string,
            name,
            procedure_resource.amount,
            unit,
            entity_name,
            creation_date,
            created_by,
            procedure_resource.unit_price,
            procedure_resource.total_price,
            reporting_number,
            approval_status,
            approval_date,
            bulletin_number,
        ]

    def get_daily_report_data(self, daily_report):
        obj = self.get_board_instance(daily_report)
        contract_item_administration = obj.contract_item_administration if obj else None

        mdr = None
        if obj and obj.multiple_daily_reports.exists():
            mdr_list = list(obj.multiple_daily_reports.all())
            mdr = mdr_list[0] if mdr_list else obj.multiple_daily_reports.all()[0]

        section = ""
        if contract_item_administration:
            services_list = list(
                contract_item_administration.contract_item_administration_services.all()
            )
            if services_list:
                section = services_list[0].description
            elif (
                contract_item_administration.contract_item_administration_services.exists()
            ):
                section = (
                    contract_item_administration.contract_item_administration_services.first().description
                )

        sort_string = (
            contract_item_administration.sort_string
            if contract_item_administration
            else ""
        )

        name = obj.resource_name if obj else ""

        amount = obj.amount if obj else None

        unit = obj.resource_unit if obj else None

        entity_name = (
            contract_item_administration.entity.name
            if contract_item_administration and contract_item_administration.entity
            else ""
        )

        creation_date = (
            utc_to_local(obj.creation_date).strftime("%d/%m/%Y") if obj else None
        )

        try:
            if obj.unit_price is None:
                unit_price = obj.resource_unit_price
            else:
                unit_price = obj.unit_price
        except Exception:
            unit_price = None

        try:
            if obj.total_price not in [None, 0]:
                total_price = obj.total_price
            else:
                total_price = (amount * unit_price) / self.work_days
        except Exception:
            total_price = None

        try:
            if obj.created_by:
                created_by = obj.created_by.get_full_name()
            else:
                created_by = mdr.created_by.get_full_name()
        except Exception:
            created_by = ""

        mdr_number = mdr.number if mdr else ""

        approval_status = (
            self.approval_status_translation.get(obj.approval_status, "") if obj else ""
        )

        approval_date = (
            utc_to_local(obj.approval_date).strftime("%d/%m/%Y %H:%M")
            if obj and obj.approval_date
            else None
        )

        bulletin_number = (
            obj.measurement_bulletin.number if obj and obj.measurement_bulletin else ""
        )

        return [
            "ADMINISTRAÇÃO",
            section,
            sort_string,
            name,
            amount,
            unit,
            entity_name,
            creation_date,
            created_by,
            unit_price,
            total_price,
            mdr_number,
            approval_status,
            approval_date,
            bulletin_number,
        ]

    def fill_workbook(self):
        border = Border(
            top=Side(border_style="thin", color="000000"),
        )
        procedure_resource_pattern = PatternFill("solid", fgColor="DCE6F1")
        daily_report_pattern = PatternFill("solid", fgColor="C5D9F1")

        relatorio = self.wb["PRÉVIA DE VALOR ADM e P.U."]

        start_line = 2

        for line_index, item in enumerate(self.procedure_queryset):
            procedure_resource_data = self.get_procedure_resource_data(item)
            for column_index, data in enumerate(procedure_resource_data):
                data_cell = relatorio.cell(
                    row=line_index + start_line, column=column_index + 1
                )
                data_cell.value = data
                if line_index == 0:
                    data_cell.border = border
                data_cell.fill = procedure_resource_pattern
                if get_column_letter(column_index + 1) in ["H", "N"]:
                    data_cell.number_format = "dd/mm/yyyy"
                if get_column_letter(column_index + 1) in ["J", "K"]:
                    data_cell.number_format = "R$ #,##0.00"

        start_line = relatorio.max_row + 1

        for line_index, item in enumerate(self.daily_queryset):
            daily_report_data = self.get_daily_report_data(item)
            for column_index, data in enumerate(daily_report_data):
                data_cell = relatorio.cell(
                    row=line_index + start_line, column=column_index + 1
                )
                data_cell.value = data
                data_cell.fill = daily_report_pattern
                if line_index == 0:
                    data_cell.border = border
                if get_column_letter(column_index + 1) in ["H", "N"]:
                    data_cell.number_format = "dd/mm/yyyy"
                if get_column_letter(column_index + 1) in ["J", "K"]:
                    data_cell.number_format = "R$ #,##0.00"
        return

    def generate_file(self):
        old_folder = "apps/resources/templates/"
        new_folder = "/tmp/preview_download/"
        temp_file = "previa_medicao.xlsx"

        self.copy_and_rename(old_folder, new_folder, temp_file)
        new_file_path = self.return_file_path(new_folder)
        self.wb = self.load_file(new_file_path)
        self.fill_contract_info()
        self.fill_workbook()

        self.wb.save(new_file_path)
        self.upload_file(new_file_path)
        os.remove(new_file_path)
        os.rmdir(new_folder)


@task
def preview_download_export_async(
    file_uuid, filename, object_name, company_name, contract_uuid, work_days
):
    preview_download_export = PreviewDownloadExport(
        file_uuid=file_uuid,
        filename=filename,
        object_name=object_name,
        company_name=company_name,
        contract_uuid=contract_uuid,
        work_days=work_days,
    )
    preview_download_export.generate_file()
