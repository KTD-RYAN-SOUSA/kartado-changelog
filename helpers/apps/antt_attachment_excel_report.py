import logging
import os
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from tempfile import mkdtemp
from typing import Dict, List
from urllib.parse import unquote

import boto3
import botocore.config
import pytz
from django.conf import settings
from django.db.models import Prefetch
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.companies.models import Company
from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.export_utils import format_km
from helpers.apps.ccr_report_utils.image import (
    ResizeMethod,
    SheetTarget,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.workbook_utils import set_active_cell, set_zoom
from helpers.edit_export.edit_export_commons import get_custom_options
from helpers.strings import (
    check_image_file,
    clean_latin_string,
    deep_keys_to_snake_case,
    to_snake_case,
)
from RoadLabsAPI.settings import credentials

logger = logging.getLogger(__name__)


def download_picture(s3, dir: str, rf: ReportingFile, quality: str = None) -> str:
    try:
        upload = rf.upload
        file_path = upload.url.split("?")[0].split(".com/")[1]
        bucket_name = upload.url.split(".s3")[0].split("/")[-1]
        image_format = file_path.split(".")[-1]
    except Exception:
        return None

    if not check_image_file(file_path):
        return None

    os.makedirs(dir, exist_ok=True)
    image_path = "{}/{}.{}".format(dir, str(rf.uuid), image_format)
    downloaded = False
    if quality is not None:
        try:
            s3.download_file(
                bucket_name + "-" + quality, unquote(file_path), image_path
            )
            downloaded = True
        except Exception:
            pass

    if not downloaded:
        try:
            s3.download_file(bucket_name, unquote(file_path), image_path)
            downloaded = True
        except Exception:
            pass
    return image_path if downloaded else None


class AnttAttachmentExcelReport:
    TEMPLATE_SHEET_NAME = "template"
    UNAVAILABLE = "N/A"
    PICS_THREADING_LIMIT = 30

    # Coordenadas padrão
    COORDS_DEFAULT = {
        "IDENTIFICATION": (5, 2),  # B
        "ROAD": (5, 3),  # C
        "KM_INITIAL": (5, 4),  # D
        "KM_FINAL": (5, 5),  # E
        "SENSE": (5, 6),  # F
        "TYPE": (5, 7),  # G
        "PREVIOUS_STATE": (5, 8),  # H
        "PREVIOUS_PHOTO": (5, 9),  # I
        "CURRENT_STATE": (5, 10),  # J
        "CURRENT_PHOTO": (5, 11),  # K
        "CORRECTIVE_ACTION": (5, 12),
    }

    # Coordenadas para dissipador (Anexo III.1)
    COORDS_DISSIPADOR = {
        "IDENTIFICATION": (5, 2),  # B
        "ROAD": (5, 3),  # C
        "KM_INITIAL": (5, 4),  # D
        "KM_FINAL": (5, 5),  # E
        "SENSE": (5, 6),  # F
        "TYPE": (5, 7),  # G
        "FLOW": (5, 8),  # H (Montante/Jusante)
        "PREVIOUS_STATE": (5, 9),  # I
        "PREVIOUS_PHOTO": (5, 10),  # J
        "CURRENT_STATE": (5, 11),  # K
        "CURRENT_PHOTO": (5, 12),  # L
        "CORRECTIVE_ACTION": (5, 13),
    }

    def __init__(
        self,
        company_uuid: str,
        reporting_uuids: List[str],
        occurrence_type_uuids: List[str] = None,
    ):
        self.company: Company = Company.objects.get(uuid=company_uuid)
        self.reporting_uuids = reporting_uuids.split(",")
        metadata: dict = deep_keys_to_snake_case(self.company.metadata)
        self.allowed_occurrence_types_antt_attachment = metadata[
            "allowed_occurrence_types_antt_attachment"
        ]

        self.occurrence_type_uuids = occurrence_type_uuids

        self._validate_occurrence_type_uuids()

        self.temp_dir = mkdtemp()
        self.url = None
        self.name = None
        self.file_path = None

    @staticmethod
    def _normalize_id_occurrence_type(id_occurrence_type) -> list:
        """Normaliza id_occurrence_type para sempre retornar uma lista de UUIDs."""
        if isinstance(id_occurrence_type, list):
            return id_occurrence_type
        if isinstance(id_occurrence_type, str):
            return [id_occurrence_type]
        return []

    @staticmethod
    def _is_class_based_type(occ_snake: dict) -> bool:
        """Detecta se a entrada 'type' classifica por classe (sem apiName) ou por campo."""
        form_fields = occ_snake.get("form_fields", []) or []
        for field in form_fields:
            field_snake = {to_snake_case(k): v for k, v in field.items()}
            if field_snake.get("name") == "type":
                return not bool(field_snake.get("api_name"))
        return False

    def _validate_occurrence_type_uuids(self):
        """Valida se os occurrence_type_uuids fornecidos estão nos tipos permitidos"""
        if not self.occurrence_type_uuids:
            return

        # Extrai os IDs permitidos do metadata (suporta string ou lista)
        allowed_ids = set()
        for occ in self.allowed_occurrence_types_antt_attachment:
            ids = self._normalize_id_occurrence_type(occ.get("id_occurrence_type"))
            allowed_ids.update(ids)
        allowed_ids.discard("")

        # Normaliza occurrence_type_uuids para set de UUIDs individuais
        requested_ids = set()

        if isinstance(self.occurrence_type_uuids, list):
            # Se for lista, processa cada item
            for item in self.occurrence_type_uuids:
                if isinstance(item, str):
                    # Se o item contém vírgulas, divide
                    if "," in item:
                        requested_ids.update(uuid.strip() for uuid in item.split(","))
                    else:
                        requested_ids.add(item.strip())
                else:
                    requested_ids.add(str(item).strip())
        elif isinstance(self.occurrence_type_uuids, str):
            # Se for string, divide por vírgula
            requested_ids = {
                uuid.strip() for uuid in self.occurrence_type_uuids.split(",")
            }
        else:
            requested_ids.add(str(self.occurrence_type_uuids).strip())

        # Remove strings vazias
        requested_ids.discard("")

        # Valida se todos os IDs solicitados estão nos permitidos
        invalid_ids = requested_ids - allowed_ids

        if invalid_ids:
            raise ValueError(
                f"IDs de occurrence_type não permitidos: {', '.join(invalid_ids)}. "
                f"IDs permitidos: {', '.join(sorted(allowed_ids))}"
            )

        # Atualiza self.occurrence_type_uuids para ser uma lista limpa de UUIDs
        self.occurrence_type_uuids = list(requested_ids)

        return True

    def get_excel_name(self, type_label: str) -> str:
        """Retorna o nome do arquivo Excel baseado no type_label.
        Retorna None se o type_label não for mapeado."""
        type_label_lower = type_label.lower().strip()

        if type_label_lower in ["dissipador", "anexo iii.1"]:
            return "Anexo III.1 - Comparativo Drenagem - Bueiros.xlsx"
        elif type_label_lower in ["descida", "anexo iii.2"]:
            return "Anexo III.2 - Comparativo Drenagem - Descidas d'água.xlsx"
        elif type_label_lower in ["meio fio", "anexo iii.3"]:
            return "Anexo III.3 - Comparativo Drenagem - Meio fio.xlsx"
        elif type_label_lower in ["sarjeta", "valeta", "anexo iii.4"]:
            return "Anexo III.4 - Comparativo Drenagem - Sarjetas.xlsx"

        # Type label não mapeado - retorna None para descartar
        logger.warning(
            f"Type label não mapeado em get_excel_name: '{type_label}'. "
            f"Reporting será descartado."
        )
        return None

    def get_excel_template_name(self, type_label: str) -> str:
        """Retorna o caminho do template Excel baseado no type_label.
        Retorna None se o type_label não for mapeado."""
        type_label_lower = type_label.lower().strip()

        if type_label_lower in ["dissipador", "anexo iii.1"]:
            return "fixtures/reports/template_antt_anexo_iii_drenagem_bueiros.xlsx"
        elif type_label_lower in ["descida", "anexo iii.2"]:
            return "fixtures/reports/template_antt_anexo_iii_descidas_dagua.xlsx"
        elif type_label_lower in ["meio fio", "anexo iii.3"]:
            return "fixtures/reports/template_antt_anexo_iii_meio_fio.xlsx"
        elif type_label_lower in ["sarjeta", "valeta", "anexo iii.4"]:
            return "fixtures/reports/template_antt_anexo_iii_sarjetas.xlsx"

        # Type label não mapeado - retorna None para descartar
        logger.warning(
            f"Type label não mapeado em get_excel_template_name: '{type_label}'. "
            f"Reporting será descartado."
        )
        return None

    def upload_file(self, s3):
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        expires = datetime.now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)
        object_name = f"{settings.AWS_PRIVATE_MEDIA_LOCATION}/{self.name}"

        s3.upload_file(
            self.file_path, bucket_name, object_name, ExtraArgs={"Expires": expires}
        )

        try:
            os.remove(self.file_path)
        except Exception:
            pass

        url_s3 = s3.generate_presigned_url(
            "get_object", Params={"Bucket": bucket_name, "Key": object_name}
        )
        self.url = url_s3

    def create_sheet(self, wb: Workbook, sheet_title: str):
        template_ws = wb[self.TEMPLATE_SHEET_NAME]
        ws = wb.copy_worksheet(template_ws)
        ws.title = sheet_title

    def _resolve_amount_type_label(self, reporting, type_name: str) -> str:
        form_data = getattr(reporting, "form_data", {}) or {}
        allowed_list = (
            getattr(self, "allowed_occurrence_types_antt_attachment", []) or []
        )

        # Obtém o occurrence_type_uuid do reporting atual
        occurrence_type_uuid = (
            str(getattr(reporting.occurrence_type, "uuid", None))
            if reporting and hasattr(reporting, "occurrence_type")
            else None
        )

        for occ in allowed_list:
            occ_snake = {to_snake_case(k): v for k, v in occ.items()}

            # Filtra apenas pelo occurrence_type do reporting atual
            id_list = self._normalize_id_occurrence_type(
                occ_snake.get("id_occurrence_type")
            )
            if occurrence_type_uuid and occurrence_type_uuid not in id_list:
                continue

            form_fields = occ_snake.get("form_fields")
            for field in form_fields:
                field_snake = {to_snake_case(k): v for k, v in field.items()}
                api_name = field_snake.get("api_name")
                name = field_snake.get("name")
                value = field_snake.get("value")
                if not api_name:
                    # Modo classe: resolve "type" pelo UUID do OccurrenceType
                    if type_name == "type" and name == "type":
                        options = field_snake.get("options", []) or []
                        for opt in options:
                            opt_snake = {to_snake_case(k): v for k, v in opt.items()}
                            opt_value = opt_snake.get("value")
                            if (
                                isinstance(opt_value, list)
                                and occurrence_type_uuid in opt_value
                            ):
                                return opt_snake.get("name") or self.UNAVAILABLE
                    continue

                if type_name == "previous-photo":
                    if type_name == name:
                        return value

                if type_name == "current-photo":
                    if type_name == name:
                        return value

                if type_name == "flow":
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                options = field_snake.get("options", []) or []
                                for opt in options:
                                    opt_snake = {
                                        to_snake_case(k): v for k, v in opt.items()
                                    }
                                    opt_value = opt_snake.get("value")
                                    if opt_value == form_data[key]:
                                        return opt_snake.get("name") or self.UNAVAILABLE

                if type_name == "identification":
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                return str(form_data[key]) or self.UNAVAILABLE

                if type_name == "previous-device-status":
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                options = field_snake.get("options", []) or []
                                for opt in options:
                                    opt_snake = {
                                        to_snake_case(k): v for k, v in opt.items()
                                    }
                                    opt_value = opt_snake.get("value")
                                    if opt_value == form_data[key]:
                                        return opt_snake.get("name") or self.UNAVAILABLE

                if type_name == "current-device-status":
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                options = field_snake.get("options", []) or []
                                for opt in options:
                                    opt_snake = {
                                        to_snake_case(k): v for k, v in opt.items()
                                    }
                                    opt_value = opt_snake.get("value")
                                    if opt_value == form_data[key]:
                                        return opt_snake.get("name") or self.UNAVAILABLE

                if type_name == "corrective-actions":
                    names = []
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                options = field_snake.get("options", []) or []
                                for opt in options:
                                    opt_snake = {
                                        to_snake_case(k): v for k, v in opt.items()
                                    }
                                    opt_value = opt_snake.get("value")
                                    if isinstance(form_data[key], list):
                                        if opt_value in form_data[key]:
                                            names.append(opt_snake.get("name"))
                                    else:
                                        if opt_value == form_data[key]:
                                            names.append(opt_snake.get("name"))
                                if names:
                                    return ", ".join(sorted(names))
                                else:
                                    return self.UNAVAILABLE
                if type_name == "type":
                    options = field_snake.get("options", []) or []
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                for opt in options:
                                    opt_snake = {
                                        to_snake_case(k): v for k, v in opt.items()
                                    }
                                    opt_value = opt_snake.get("value")
                                    if opt_value == form_data[key]:
                                        return opt_snake.get("name") or str(opt_value)

                if type_name == "type_column":
                    for key in form_data:
                        if to_snake_case(name) == to_snake_case(type_name):
                            if to_snake_case(key) == to_snake_case(api_name):
                                options = field_snake.get("options", []) or []
                                for opt in options:
                                    opt_snake = {
                                        to_snake_case(k): v for k, v in opt.items()
                                    }
                                    opt_value = opt_snake.get("value")
                                    if opt_value == form_data[key]:
                                        return opt_snake.get("name") or self.UNAVAILABLE
        return self.UNAVAILABLE

    def _copy_row_styles(self, ws: Worksheet, source_row: int, target_row: int):
        """Copia estilos e dimensões da linha de origem para a linha de destino"""
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(source_row, col)
            target_cell = ws.cell(target_row, col)

            if source_cell.has_style:
                target_cell.font = source_cell.font.copy()
                target_cell.border = source_cell.border.copy()
                target_cell.fill = source_cell.fill.copy()
                target_cell.number_format = source_cell.number_format
                target_cell.protection = source_cell.protection.copy()
                target_cell.alignment = source_cell.alignment.copy()

        # Copia altura da linha
        if source_row in ws.row_dimensions:
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    def add_reporting_to_sheet(
        self,
        wb: Workbook,
        reporting: Reporting,
        reporting_files: List[ReportingFile],
        files: Dict[str, str],
        sheet_title: str,
        previous_reporting: Reporting = None,
        row_offset: int = 0,
        type_label: str = None,  # NOVO PARÂMETRO
    ):
        ws = wb[sheet_title]
        # Usa type_label para determinar qual coordenada usar
        coords = (
            self.COORDS_DISSIPADOR
            if type_label and type_label.lower() in ["dissipador", "anexo iii.1"]
            else self.COORDS_DEFAULT
        )

        # Ajusta as coordenadas com o offset da linha
        base_row = coords["IDENTIFICATION"][0]
        adjusted_coords = {
            key: (row + row_offset, col) for key, (row, col) in coords.items()
        }

        # Se não for a primeira linha, copia os estilos da linha base
        if row_offset > 0:
            target_row = base_row + row_offset
            self._copy_row_styles(ws, base_row, target_row)

        ws.cell(
            *adjusted_coords["IDENTIFICATION"]
        ).value = self._resolve_amount_type_label(reporting, "identification")
        ws.cell(*adjusted_coords["ROAD"]).value = (
            getattr(reporting, "road_name", "") or self.UNAVAILABLE
        )
        km_value = getattr(reporting, "km", None)
        ws.cell(*adjusted_coords["KM_INITIAL"]).value = (
            format_km(km_value) if km_value is not None else self.UNAVAILABLE
        )
        end_km_value = getattr(reporting, "end_km", None) or getattr(
            reporting, "project_end_km", None
        )
        ws.cell(*adjusted_coords["KM_FINAL"]).value = (
            format_km(end_km_value) if end_km_value is not None else self.UNAVAILABLE
        )
        direction_value = getattr(reporting, "direction", "")
        custom_directions = get_custom_options(self.company, "direction")

        if (
            direction_value
            and custom_directions
            and direction_value in custom_directions
        ):
            direction_label = custom_directions[direction_value]
        else:
            direction_label = direction_value or self.UNAVAILABLE

        ws.cell(*adjusted_coords["SENSE"]).value = direction_label
        ws.cell(*adjusted_coords["TYPE"]).value = self._resolve_amount_type_label(
            reporting, "type_column"
        )

        # Só para dissipador (Anexo III.1)
        if coords is self.COORDS_DISSIPADOR:
            ws.cell(*adjusted_coords["FLOW"]).value = self._resolve_amount_type_label(
                reporting, "flow"
            )

        ws.cell(*adjusted_coords["PREVIOUS_STATE"]).value = (
            self._resolve_amount_type_label(
                previous_reporting, "previous-device-status"
            )
            if previous_reporting
            else self.UNAVAILABLE
        )
        ws.cell(
            *adjusted_coords["CURRENT_STATE"]
        ).value = self._resolve_amount_type_label(reporting, "current-device-status")
        corrective_action_value = self._resolve_amount_type_label(
            reporting, "corrective-actions"
        )
        ws.cell(*adjusted_coords["CORRECTIVE_ACTION"]).value = (
            corrective_action_value or self.UNAVAILABLE
        )

        # Fotos
        def insert_rf_image(coord, rf):
            if not rf:
                return
            img_path = files.get(str(rf.uuid))
            if not img_path or not os.path.exists(img_path):
                return
            row, col = coord
            range_string = f"{get_column_letter(col)}{row}"
            try:
                insert_picture_2(
                    worksheet=ws,
                    range_string=range_string,
                    picture=Image(img_path),
                    target=SheetTarget.DesktopExcel,
                    resize_method=ResizeMethod.ProportionalCentered,
                )
            except Exception:
                pass

        # Busca o valor do "kind" para previous-photo e current-photo do metadata
        previous_photo_kind = self._resolve_amount_type_label(
            reporting, "previous-photo"
        )
        current_photo_kind = self._resolve_amount_type_label(reporting, "current-photo")

        # Busca a foto "previous" (antes) do apontamento anterior
        previous_rf = None
        if previous_reporting:
            # Busca os reporting_files do previous_reporting
            previous_reporting_files = (
                list(getattr(previous_reporting, "reporting_files", []).all())
                if hasattr(previous_reporting, "reporting_files")
                else []
            )

            # Busca a foto com kind="antes" (ou o valor configurado para previous-photo)
            previous_rf = next(
                (
                    rf
                    for rf in previous_reporting_files
                    if getattr(rf, "kind", None) == previous_photo_kind
                ),
                None,
            )

        # Busca a foto "current" (depois) do apontamento atual
        current_rf = next(
            (
                rf
                for rf in reporting_files
                if getattr(rf, "kind", None) == current_photo_kind
            ),
            None,
        )

        # Insere as fotos ou "Sem foto"
        if previous_rf:
            insert_rf_image(adjusted_coords["PREVIOUS_PHOTO"], previous_rf)
        else:
            # Se não tem foto (com ou sem apontamento anterior), coloca "Sem foto"
            ws.cell(*adjusted_coords["PREVIOUS_PHOTO"]).value = "Sem foto"

        if current_rf:
            insert_rf_image(adjusted_coords["CURRENT_PHOTO"], current_rf)
        else:
            ws.cell(*adjusted_coords["CURRENT_PHOTO"]).value = "Sem foto"

    def get_files(
        self, s3, reporting_files: Dict[str, ReportingFile]
    ) -> Dict[str, str]:
        uuid_to_file: Dict[str, str] = {}

        def download_and_store(rf):
            file_path = download_picture(s3, self.temp_dir, rf)
            if file_path:
                uuid_to_file[str(rf.uuid)] = file_path

        with ThreadPoolExecutor(max_workers=self.PICS_THREADING_LIMIT) as executor:
            for rf in reporting_files.values():
                executor.submit(download_and_store, rf)
        return uuid_to_file

    def create_file(self, s3):
        reportings = (
            Reporting.objects.filter(
                company=self.company, uuid__in=self.reporting_uuids
            )
            .order_by("-created_at")
            .prefetch_related(
                Prefetch(
                    "reporting_files",
                    queryset=ReportingFile.objects.all().order_by(
                        "datetime", "uploaded_at"
                    ),
                ),
                "occurrence_type",
                "status",
                "parent",
                "parent__children",
                Prefetch(
                    "parent__children__reporting_files",
                    queryset=ReportingFile.objects.all().order_by(
                        "datetime", "uploaded_at"
                    ),
                ),
            )
        )

        reporting_to_rfs = {}
        type_to_reportings = {}
        seen_in_type_grouping = set()

        for reporting in reportings:
            reporting_uuid = str(reporting.uuid)

            # Validação: Se já processou este reporting, pula
            if reporting_uuid in seen_in_type_grouping:
                logger.warning(
                    f"Reporting duplicado detectado no agrupamento por type_label: {reporting_uuid}"
                )
                continue

            seen_in_type_grouping.add(reporting_uuid)

            if reporting.parent is not None:
                # Obtém o occurrence_type_id do reporting atual
                current_occurrence_type_id = (
                    str(reporting.occurrence_type.uuid)
                    if reporting.occurrence_type
                    else None
                )

                # Obtém o type_label do reporting atual
                current_type_label = self._resolve_amount_type_label(
                    reporting, "type"
                ).lower()

                # Filtra apenas os children com o mesmo occurrence_type_id E type_label
                reportings_list = [
                    r
                    for r in reporting.parent.children.all()
                    if r.occurrence_type
                    and str(r.occurrence_type.uuid) == current_occurrence_type_id
                    and self._resolve_amount_type_label(r, "type").lower()
                    == current_type_label
                ]

                # Ordena por created_at crescente (mais antigo primeiro)
                sorted_reportings = sorted(
                    reportings_list,
                    key=lambda r: r.created_at or datetime.min,
                )

                # Busca o imediatamente anterior (o que vem antes na lista ordenada)
                try:
                    current_index = next(
                        i
                        for i, r in enumerate(sorted_reportings)
                        if str(r.uuid) == str(reporting.uuid)
                    )
                    # Pega o anterior na ordem cronológica (índice - 1)
                    previous_reporting = (
                        sorted_reportings[current_index - 1]
                        if current_index > 0
                        else None
                    )
                except (StopIteration, IndexError):
                    previous_reporting = None
            else:
                reportings_list = []
                previous_reporting = None

            reporting_files_list = list(reporting.reporting_files.all())
            reporting_to_rfs[reporting.uuid] = reporting_files_list

            # Adiciona também os reporting_files do previous_reporting se existir
            if previous_reporting and previous_reporting.uuid not in reporting_to_rfs:
                previous_reporting_files_list = list(
                    previous_reporting.reporting_files.all()
                )
                reporting_to_rfs[
                    previous_reporting.uuid
                ] = previous_reporting_files_list

            type_label = self._resolve_amount_type_label(reporting, "type").lower()
            if type_label not in type_to_reportings:
                type_to_reportings[type_label] = []
            type_to_reportings[type_label].append((reporting, previous_reporting))

        # Logging para debug
        logger.info(
            f"Agrupamento por type_label: {dict((k, len(v)) for k, v in type_to_reportings.items())}"
        )

        if (
            len(type_to_reportings) == 1
            and len(list(type_to_reportings.values())[0]) == 1
        ):
            type_label = list(type_to_reportings.keys())[0]
            reportings_with_previous = list(type_to_reportings.values())[0]

            # Valida se type_label é mapeado
            excel_name = self.get_excel_name(type_label)
            excel_template = self.get_excel_template_name(type_label)

            if excel_name is None or excel_template is None:
                logger.warning(
                    f"Type label '{type_label}' não mapeado. Nenhum arquivo será gerado."
                )
                return

            excel_path = os.path.join(self.temp_dir, excel_name)
            wb = load_workbook(excel_template)
            reporting, previous_reporting = reportings_with_previous[0]
            sheet_title = str(getattr(reporting, "road_name", "Rodovia"))[:31]
            sheet_title = clean_latin_string(sheet_title).replace("/", "-")
            if sheet_title not in wb.sheetnames:
                self.create_sheet(wb, sheet_title)

            # Baixa fotos do reporting atual
            rf_dict = {str(rf.uuid): rf for rf in reporting_to_rfs[reporting.uuid]}

            # Baixa também fotos do previous_reporting se existir
            if previous_reporting and previous_reporting.uuid in reporting_to_rfs:
                rf_dict.update(
                    {
                        str(rf.uuid): rf
                        for rf in reporting_to_rfs[previous_reporting.uuid]
                    }
                )

            rf_uuid_to_file_path = self.get_files(s3, rf_dict)

            self.add_reporting_to_sheet(
                wb,
                reporting,
                reporting_to_rfs.get(reporting.uuid, []),
                rf_uuid_to_file_path,
                sheet_title,
                previous_reporting,
                row_offset=0,
                type_label=type_label,
            )

            if self.TEMPLATE_SHEET_NAME in wb.sheetnames:
                wb.remove(wb[self.TEMPLATE_SHEET_NAME])
            set_zoom(wb, 100, "normal")
            set_active_cell(wb, "A1")
            wb.active = wb.worksheets[0]
            wb.save(excel_path)
            wb.close()
            self.name = excel_name
            self.file_path = excel_path
            self.upload_file(s3)
            try:
                os.remove(excel_path)
            except Exception:
                pass
            return

        excel_paths = []
        for type_label, reportings_with_previous in type_to_reportings.items():
            # Valida se type_label é mapeado
            excel_name = self.get_excel_name(type_label)
            excel_template = self.get_excel_template_name(type_label)

            if excel_name is None or excel_template is None:
                logger.warning(
                    f"Type label '{type_label}' não mapeado. Reportings deste tipo serão ignorados."
                )
                continue  # Pula para o próximo type_label

            excel_path = os.path.join(self.temp_dir, excel_name)
            wb = load_workbook(excel_template)
            self.name = excel_name

            # Agrupa por rodovia
            road_to_reportings = {}
            seen_reportings = set()  # Tracking de reportings já processados

            for reporting, previous_reporting in reportings_with_previous:
                reporting_uuid = str(reporting.uuid)

                # Validação: Se já processou este reporting, pula
                if reporting_uuid in seen_reportings:
                    logger.warning(
                        f"Reporting duplicado detectado no agrupamento por road (type={type_label}): {reporting_uuid}"
                    )
                    continue

                seen_reportings.add(reporting_uuid)

                # Valida road_name não vazio
                road_name = getattr(reporting, "road_name", None)

                if not road_name or not str(road_name).strip():
                    # Usa occurrence_type ou UUID como fallback
                    road_name = f"Rodovia - {reporting.occurrence_type.name if reporting.occurrence_type else str(reporting.uuid)[:8]}"
                else:
                    road_name = str(road_name)[:31]
                road_name = clean_latin_string(road_name).replace("/", "-")

                if road_name not in road_to_reportings:
                    road_to_reportings[road_name] = []
                road_to_reportings[road_name].append((reporting, previous_reporting))

            # Ordena reportings dentro de cada rodovia por data de criação
            for road_name in road_to_reportings:
                road_to_reportings[road_name].sort(
                    key=lambda x: x[0].created_at if x[0].created_at else datetime.min
                )

            # Logging para debug
            logger.info(
                f"Agrupamento por road (type={type_label}): "
                f"{dict((k, len(v)) for k, v in road_to_reportings.items())}"
            )

            # Processa cada rodovia
            for road_name, road_reportings in road_to_reportings.items():
                sheet_title = road_name
                if sheet_title not in wb.sheetnames:
                    self.create_sheet(wb, sheet_title)

                # Adiciona cada reporting na mesma aba
                for idx, (reporting, previous_reporting) in enumerate(road_reportings):
                    # Baixa fotos do reporting atual
                    rf_dict = {
                        str(rf.uuid): rf for rf in reporting_to_rfs[reporting.uuid]
                    }

                    # Baixa também fotos do previous_reporting se existir
                    if (
                        previous_reporting
                        and previous_reporting.uuid in reporting_to_rfs
                    ):
                        rf_dict.update(
                            {
                                str(rf.uuid): rf
                                for rf in reporting_to_rfs[previous_reporting.uuid]
                            }
                        )

                    rf_uuid_to_file_path = self.get_files(s3, rf_dict)

                    self.add_reporting_to_sheet(
                        wb,
                        reporting,
                        reporting_to_rfs.get(reporting.uuid, []),
                        rf_uuid_to_file_path,
                        sheet_title,
                        previous_reporting,
                        row_offset=idx,
                        type_label=type_label,
                    )

            if self.TEMPLATE_SHEET_NAME in wb.sheetnames:
                wb.remove(wb[self.TEMPLATE_SHEET_NAME])
            set_zoom(wb, 100, "normal")
            set_active_cell(wb, "A1")
            wb.active = wb.worksheets[0]
            wb.save(excel_path)
            wb.close()
            excel_paths.append(excel_path)

        # Verifica se há arquivos válidos para criar o ZIP
        if not excel_paths:
            logger.warning(
                "Nenhum arquivo Excel foi gerado. Todos os type_labels foram descartados."
            )
            return

        zip_name = "Anexo III - Comparativo Drenagem e OAC.zip"
        zip_path = os.path.join(self.temp_dir, zip_name)
        with zipfile.ZipFile(zip_path, "w") as zipf:
            for excel_path in excel_paths:
                zipf.write(excel_path, os.path.basename(excel_path))

        self.name = zip_name
        self.file_path = zip_path
        self.upload_file(s3)
        for excel_path in excel_paths:
            try:
                os.remove(excel_path)
            except Exception:
                pass

    def get_url_and_name(self):
        client_config = botocore.config.Config(
            max_pool_connections=self.PICS_THREADING_LIMIT,
        )
        s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
            config=client_config,
        )
        if not self.name:
            self.create_file(s3)
        if not self.url:
            self.upload_file(s3)
        return {"url": self.url, "name": self.name}
