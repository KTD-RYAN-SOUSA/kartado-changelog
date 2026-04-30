import re
from pathlib import Path
from typing import Dict, List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from zappa.asynchronous import task

from apps.reportings.models import Reporting
from helpers.apps.ccr_report_monitoring_oae import CCRReportMonitoringOAE
from helpers.apps.ccr_report_monitoring_oae import XlsxHandler as MonitoringXlsxHandler
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_provider_logo_file,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.apps.ccr_report_utils.workbook_utils import set_zoom
from helpers.strings import keys_to_snake_case


def get_file_name(reporting):
    inspection_year_campaign = keys_to_snake_case(reporting.form_data).get(
        "inspection_year_campaign", ""
    )
    road_number = re.sub(r"\D", "", reporting.road_name)
    n_oae = new_get_form_data(reporting, "oaeNumeroCodigoObra", default="")
    try:
        n_oae = f"{int(n_oae):03}"
    except Exception:
        pass
    uf = reporting.road_name[-2:]
    file_name = f"OAE{inspection_year_campaign}{road_number}{n_oae}{uf}"
    return file_name


class XlsxHandler(MonitoringXlsxHandler):
    __OAE_SUMMARY_FIELDS: Dict[str, str] = {
        "superestruturaAntt": "H12",
        "mesoestruturaAntt": "K12",
        "infraestruturaAntt": "O12",
        "rail": "R12",
        "notaTecnicaNotaEstrutural": "W12",
        "resumOaeAtualSuperestrutura": "D14",
        "resumoOaeAtualJuntas": "D17",
        "resumoOaeAtualApoio": "D20",
        "resumoOaeAtualMesoestrutura": "D23",
        "resumoOaeAtualEncontros": "D26",
        "resumoOaeAtualPavimento": "D29",
        "resumoOaeAtualAcostamento": "D32",
        "resumoOaeAtualDrenagem": "D35",
        "resumoOaeAtualGuardaCorpos": "D38",
        "resumoOaeAtualGuardaRodas": "D41",
        "resumoOaeAtualTaludes": "D44",
        "resumoOaeAtualIluminacao": "D47",
        "resumoOaeAtualSinalizacao": "D50",
        "resumoOaeAtualGabaritos": "D53",
        "resumoOaeAtualPilares": "D56",
        "registroAnomaliasInformacoesComplementares": "D59",
    }
    __TUNNEL_SUMMARY_FIELDS: Dict[str, str] = {
        "superestruturaAntt": "H12",
        "mesoestruturaAntt": "K12",
        "infraestruturaAntt": "O12",
        "rail": "R12",
        "notaTecnicaNotaEstrutural": "W12",
        "resumoOaeAtualAbobada": "D14",
        "resumoOaeAtualParedesLaterais": "D17",
        "resumoOaeAtualJuntass": "D20",
        "resumoOaeAtualEmboques": "D23",
        "resumoOaeAtualPavimento": "D26",
        "resumoOaeAtualAcostamento": "D29",
        "resumoOaeAtualDrenagem": "D32",
        "resumoOaeAtualGuardaCorpos": "D35",
        "resumoOaeAtualGuardaRodas": "D38",
        "resumoOaeAtualContencao": "D41",
        "resumoOaeAtualIluminacao": "D44",
        "resumoOaeAtualSinalizacao": "D47",
        "resumoOaeAtualGabaritos": "D50",
        "registroAnomaliasInformacoesComplementares": "D53",
    }
    __FOOTBRIDGE_SUMMARY_FIELDS: Dict[str, str] = {
        "superestruturaAntt": "H12",
        "mesoestruturaAntt": "K12",
        "infraestruturaAntt": "O12",
        "rail": "R12",
        "notaTecnicaNotaEstrutural": "W12",
        "resumOaeAtualSuperestrutura": "D14",
        "resumoOaeAtualJuntas": "D17",
        "resumoOaeAtualApoio": "D20",
        "resumoOaeAtualMesoestrutura": "D23",
        "resumoOaeAtualEncontros": "D26",
        "resumoOaeAtualTiposAcesso": "D29",
        "resumoOaeAtualPiso": "D32",
        "resumoOaeAtualDrenagem": "D35",
        "resumoOaeAtualGuardaCorpos": "D38",
        "resumoOaeAtualTelamento": "D41",
        "resumoOaeAtualTaludes": "D44",
        "resumoOaeAtualIluminacao": "D47",
        "resumoOaeAtualSinalizacao": "D50",
        "resumoOaeAtualGabaritos": "D53",
        "resumoOaeAtualPilares": "D56",
        "registroAnomaliasInformacoesComplementares": "D59",
    }

    def __init__(
        self,
        list_reporting: List[Reporting],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
    ) -> None:
        super().__init__(list_reporting, s3, sheet_target)
        self._xlsx_file = "./fixtures/reports/ccr_report_oae_management.xlsx"
        self._workbook = load_workbook(self._xlsx_file)
        self._sheetnames = self._workbook.sheetnames
        self._worksheet = self._workbook["FICHA 1"]
        self._worksheets = self._workbook.worksheets
        self._copy_antt_logo = False

    def __set_summary_sheet(self, wb: Workbook, reporting: Reporting) -> None:
        summary_sheet = wb["RESUMO"]
        oae_type = new_get_form_data(reporting, "tipoDeObrasOae", raw=True)

        summary_fields: Dict[str, str] = None
        try:
            if int(oae_type) == 3:
                summary_fields = XlsxHandler.__FOOTBRIDGE_SUMMARY_FIELDS
            elif int(oae_type) == 5:
                summary_fields = XlsxHandler.__TUNNEL_SUMMARY_FIELDS
        except Exception:
            pass

        if summary_fields is None:
            summary_fields = XlsxHandler.__OAE_SUMMARY_FIELDS

        for form_field, coord in summary_fields.items():
            val_str = "-"
            val = new_get_form_data(reporting, form_field)
            if val is not None:
                try:
                    val_str = str(val)
                    val_str = val_str.strip()
                    val_str = val_str if val_str else "-"
                except Exception:
                    pass
            summary_sheet[coord] = val_str

    def __insert_bottom_logos(
        self, logo: str, logo_team: str, provider_logo: str
    ) -> None:
        max_row = self._worksheet.max_row
        logo_row = max_row - 1
        max_col = self._worksheet.max_column

        logo_image = Image(logo) if logo else None
        team_image = Image(logo_team) if logo_team else None
        provider_image = Image(provider_logo) if provider_logo else None
        if logo_image:
            range_str = f"A{logo_row}:C{max_row}"
            margins = (1, 1, 1, 1)
            insert_picture_2(
                self._worksheet,
                range_str,
                logo_image,
                self._sheet_target,
                margins,
                ResizeMethod.ProportionalLeft,
            )

        if team_image:
            range_str = f"A{logo_row}:{get_column_letter(max_col)}{max_row}"
            insert_picture_2(
                self._worksheet,
                range_str,
                team_image,
                self._sheet_target,
                margins,
                ResizeMethod.ProportionalCentered,
            )

        if provider_image:
            range_str = f"{get_column_letter(max_col-3)}{logo_row}:{get_column_letter(max_col)}{max_row}"
            insert_picture_2(
                self._worksheet,
                range_str,
                provider_image,
                self._sheet_target,
                margins,
                ResizeMethod.ProportionalRight,
            )

    def fill_sheet(self, data_list: list):
        files = []

        for i, data in enumerate(data_list, 1):

            oae_type = new_get_form_data(data["reporting"], "tipoDeObrasOae", raw=True)

            xlsx_file: str = None
            try:
                if int(oae_type) == 3:
                    xlsx_file = (
                        "./fixtures/reports/ccr_report_oae_footbridge_management.xlsx"
                    )
                elif int(oae_type) == 5:
                    xlsx_file = (
                        "./fixtures/reports/ccr_report_oae_tunnel_management.xlsx"
                    )
            except Exception:
                pass

            if xlsx_file is None:
                xlsx_file = "./fixtures/reports/ccr_report_oae_management.xlsx"
            self._xlsx_file = xlsx_file

            self._workbook = load_workbook(self._xlsx_file)

            provider_logo = get_provider_logo_file(
                self.s3, self.temp_file, data["reporting"]
            )
            logo = data.get("logo", "")
            logo_team = data.get("logo_team")

            self.__set_ficha_1(data)

            self.__set_ficha_2(data)

            self.__set_summary_sheet(self._workbook, data["reporting"])
            temp = self._worksheet
            self._worksheet = self._workbook["RESUMO"]
            self._worksheet = temp

            self.__set_croqui(data)

            self.__set_photos(data)

            temp = self._worksheet
            for worksheet in self._workbook.worksheets:
                self._worksheet = worksheet
                self.__insert_bottom_logos(logo, logo_team, provider_logo)
            self._worksheet = temp

            file_name = data.get("file_name")
            file_name += "_Ficha-Gerencial"
            folder = f"{data['road_name']}-Fichas Gerenciais OAE"
            result = f"/tmp/{folder}/{file_name}.xlsx"
            count = 0
            while result in files:
                count += 1
                result = f"/tmp/{folder}/{file_name}({count}).xlsx"
            Path(f"/tmp/{folder}").mkdir(parents=True, exist_ok=True)
            set_zoom(self._workbook, 50, "pageBreakPreview")
            self._workbook.save(result)

            files.append(result)
        return files

    def execute(self):
        data = [
            self.create_dict(reporting=reporting, s3=self.s3)
            for reporting in self.list_reporting
        ]
        files = self.fill_sheet(data_list=data)
        return files


class CCRReportOAEManagement(CCRReportMonitoringOAE):
    def get_file_name(self):
        file_name = ""
        occurrence_type_uuid = (
            Reporting.objects.filter(uuid=self.uuids[0])
            .only("occurrence_type__uuid")[0]
            .occurrence_type.uuid
        )
        reportings = Reporting.objects.filter(
            occurrence_type__uuid=occurrence_type_uuid, pk__in=self.uuids
        ).prefetch_related("road")

        road_names = sorted({r.road_name for r in reportings})

        if reportings.count() == 1:
            reporting = reportings.first()
            extension = ""
            if self.report_format() == ReportFormat.PDF:
                extension = "pdf"
            elif self.report_format() == ReportFormat.XLSX:
                extension = "xlsx"

            file_name = f"{get_file_name(reporting)}_Ficha-Gerencial.{extension}"
        else:
            file_name = f"{'_'.join(road_names)} - Fichas Gerenciais OAE.zip"

        return file_name

    def export(self):
        list_reporting = self._get_repotings_obj()
        s3 = get_s3()
        files = XlsxHandler(
            list_reporting=list_reporting,
            s3=s3,
            sheet_target=self.sheet_target(),
        ).execute()
        road_names = sorted({r.road_name for r in list_reporting})
        result_file = ""

        if self.report_format() == ReportFormat.PDF:
            files = convert_files_to_pdf(files)

        result_file = ""
        if len(files) == 1:
            result_file = files[0]
        elif len(files) > 1:
            self.file_name = f"{'_'.join(road_names)} - Fichas Gerenciais OAE"
            result_file = f"/tmp/{self.file_name}"
            with ZipFile(result_file, "w") as zipObj:
                for file in files:
                    zipObj.write(file, "/".join(file.split("/")[-2:]))
        upload_file(s3, result_file, self.object_name)
        return True


@task
def ccr_report_oae_management_async_handler(reporter_dict: dict):
    reporter = CCRReportOAEManagement.from_dict(reporter_dict)
    reporter.export()
