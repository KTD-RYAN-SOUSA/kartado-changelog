import os
import tempfile
from concurrent.futures import ThreadPoolExecutor
from typing import List
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting, ReportingFile, ReportingInReporting
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import get_s3, upload_file
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    SheetTarget,
    download_picture,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import convert_files_to_pdf
from helpers.strings import clean_latin_string, format_km, get_obj_from_path


def safe_download_picture(s3, temp_file, photo, reporting_file, i):
    try:
        file = download_picture(
            s3,
            temp_file,
            str(photo) + str(i),
            reporting_file=reporting_file,
        )
        return file
    except Exception:
        return ""


class XlsxHandlerBuildsDiagnostics2025(object):

    FIELDS = {
        "id_ccr_antt": "A",
        "buildings": "B",
        "city": "C",
        "km": "D",
        "conservation_state": "E",
        "anomaly_identified": "F",
        "anomaly_degree": "G",
        "anomaly_action": "H",
        "photos": "I",
    }

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(
        self,
        list_uuids: List[str],
        s3,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.wb = load_workbook("./fixtures/reports/ccr_building_diagnostics_2025.xlsx")
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.list_name_files = list()
        self._worksheet = self.wb.active
        self.list_uuids = list_uuids
        first_reporting = Reporting.objects.filter(uuid__in=list_uuids).first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.occurrence_types = {
            str(item.uuid): item.name
            for item in OccurrenceType.objects.filter(company=self.company).only(
                "uuid", "name"
            )
        }
        self.thread_pool = ThreadPoolExecutor(max_workers=10)
        self.reporting_relation_metadata = get_obj_from_path(
            self.company.metadata,
            "recuperation_reporting_relation",
            default_return=None,
        )

        if self.reporting_relation_metadata:
            rep_in_reps = ReportingInReporting.objects.filter(
                parent__uuid__in=self.list_uuids,
                reporting_relation_id=self.reporting_relation_metadata,
            )
            self.recuperation_reportings = rep_in_reps.prefetch_related(
                "child"
            ).values_list("child_id", flat=True)
            self.recuperation_reporting_files = list(
                ReportingFile.objects.filter(
                    reporting__in=self.recuperation_reportings, is_shared=True
                ).prefetch_related(
                    "reporting",
                    "reporting__occurrence_type",
                    "reporting__reporting_relation_child",
                    "reporting__reporting_relation_child__parent",
                )
            )

            self.recuperation_reporting_files.sort(
                key=lambda x: x.datetime, reverse=True
            )
        self.img_count = 0
        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

    def __insert_new_rows(self, row: int, has_image: bool = False):
        chars = [chr(ord("A") + i) for i in range(9)]

        for char in chars:
            if char in ["H"]:
                self._worksheet.column_dimensions[char].width = 50
            self._worksheet[
                f"{char}{row}"
            ].border = XlsxHandlerBuildsDiagnostics2025.BORDER
            self._worksheet.row_dimensions[row].height = 100 if has_image else 15

    def create_dict(self, reporting):
        data = dict()

        data["id_ccr_antt"] = new_get_form_data(reporting, "idCcrAntt")
        data["buildings"] = new_get_form_data(
            reporting,
            "kindedificacaoInventario",
        )
        data["city"] = new_get_form_data(
            reporting,
            "city",
        )
        data["km"] = format_km(reporting, "km", 3)
        data["conservation_state"] = new_get_form_data(
            reporting,
            "generalConservationState",
        )

        data["anomalies"] = self.__get_anomalies(reporting)

        data["anomalies_degree"] = self.__set_conservation_state(reporting)

        # File aggregation fields
        data["road_name"] = reporting.road_name
        data["found_at"] = reporting.found_at

        for k, v in data.items():
            if v is None:
                data[k] = ""
        return data

    def __get_anomalies(self, reporting):

        therapy = reporting.form_data.get("therapy", [])

        if not therapy:
            return []

        anomalies = []
        therapy_len = len(therapy)

        for index in range(therapy_len):
            element = (
                new_get_form_data(
                    reporting, "therapy__{}__anomaly_element".format(index)
                )
                or ""
            )
            action_id = new_get_form_data(
                reporting, "therapy__{}__occurrence_type".format(index), raw=True
            )
            action = self.occurrence_types.get(action_id, "")

            file_future = None
            if action_id and hasattr(self, "recuperation_reporting_files"):
                photo = None
                rfs = self.recuperation_reporting_files
                for rf in rfs:
                    relations = rf.reporting.reporting_relation_child.all()
                    if (
                        relations
                        and any(relation.parent == reporting for relation in relations)
                        and str(rf.reporting.occurrence_type.uuid) == str(action_id)
                    ):
                        photo = rf
                        break
                if photo:
                    file_future = self.thread_pool.submit(
                        safe_download_picture,
                        self.s3,
                        self.temp_file,
                        photo.uuid,
                        photo,
                        self.img_count,
                    )
                    self.img_count += 1

            anomalies.append((element, action, file_future))

        return anomalies

    def __set_conservation_state(self, reporting):

        foundation_structures = new_get_form_data(
            reporting,
            "foundationStructures",
        )
        floor_coating = new_get_form_data(reporting, "floorCoating")
        tile_coating = new_get_form_data(reporting, "tileCoating")
        sidewalk = new_get_form_data(reporting, "sidewalk")
        external_wall = new_get_form_data(reporting, "externalWall")
        internal_wall = new_get_form_data(reporting, "internalWall")
        metal_structure = new_get_form_data(reporting, "metalStructure")
        cealing = new_get_form_data(reporting, "cealing")
        climatization = new_get_form_data(reporting, "climatization")
        doors = new_get_form_data(reporting, "doors")
        windows = new_get_form_data(reporting, "windows")
        ilumination = new_get_form_data(reporting, "ilumination")
        electric_instalation = new_get_form_data(
            reporting,
            "electricInstalation",
        )
        sink_hidro_instalation = new_get_form_data(
            reporting,
            "sinkHidroInstalation",
        )
        toilet_hidro_instalation = new_get_form_data(
            reporting,
            "toiletHidroInstalation",
        )
        faucet_hidro_instalation = new_get_form_data(
            reporting,
            "faucetHidroInstalation",
        )
        landscape = new_get_form_data(reporting, "landscape")
        water_box = new_get_form_data(reporting, "waterBox")
        telephone_instalation = new_get_form_data(
            reporting,
            "telephoneInstalation",
        )
        external_painting = new_get_form_data(reporting, "externalPainting")
        internal_painting = new_get_form_data(reporting, "internalPainting")
        spda_protection = new_get_form_data(reporting, "spdaProtection")
        fences = new_get_form_data(reporting, "fences")
        utilitys = new_get_form_data(reporting, "utilitys")
        abnt_requirement = new_get_form_data(reporting, "abntRequirement")

        result_dict = {
            "Fundações e Estruturas": foundation_structures,
            "Revestimento de piso (cerâmico, polímero)": floor_coating,
            "Revestimento de azulejo (cerâmico, polímero)": tile_coating,
            "Calçada": sidewalk,
            "Parede Externa": external_wall,
            "Parede Interna": internal_wall,
            "Estrutura metálica": metal_structure,
            "Cobertura/ Forro": cealing,
            "Climatização": climatization,
            "Portas": doors,
            "Janelas (vidro e armação metálica)": windows,
            "Iluminação (interna e externa)": ilumination,
            "Instalação elétrica (interna e externa)": electric_instalation,
            "Instalação hidrossanitária - Pia / Tanque de Lavar": sink_hidro_instalation,
            "Instalação hidrossanitária - Vaso Sanitário / Mictório": toilet_hidro_instalation,
            "Instalação hidrossanitária - Torneira / Registros/ chuveiros": faucet_hidro_instalation,
            "Paisagismo": landscape,
            "Caixa D'água": water_box,
            "instalação e telefonia": telephone_instalation,
            "Pintura Externa": external_painting,
            "Pintura interna": internal_painting,
            "Sistema de proteção de descarga atmosférica (SPDA)": spda_protection,
            "Cercas e alambrados": fences,
            "Utilidades (armários, gavetas)": utilitys,
            "Atendimento aos padrões de acessibilidade exigidos na NBR 9 050/2015 da ABNT": abnt_requirement,
        }

        result2 = []
        for state, value in result_dict.items():
            if value and value != "Nenhuma":
                result2.append((state, value))

        return result2

    def fill_sheet(self, *, data_dict: dict):
        key_year = ""
        row = 4
        count = 0
        list_files = list()
        for key, values_list in data_dict.items():
            for values in values_list:
                for key_year, values_in_year in values.items():
                    values_in_year = XlsxHandlerBuildsDiagnostics2025.sorted_values(
                        values_in_year
                    )
                    for items in values_in_year:
                        anomalies_list = items.get("anomalies", [])
                        anomalies_degrees = items.get("anomalies_degree", [])
                        if anomalies_list:
                            for anomaly in anomalies_list:
                                (
                                    anomaly_identified,
                                    anomaly_action,
                                    photo_future,
                                ) = anomaly
                                key_value = f"{XlsxHandlerBuildsDiagnostics2025.FIELDS['anomaly_identified']}{row}"
                                self._worksheet[key_value] = anomaly_identified
                                XlsxHandlerBuildsDiagnostics2025.__format_fonts(
                                    cell=self._worksheet[key_value],
                                    size=11,
                                )
                                key_value = f"{XlsxHandlerBuildsDiagnostics2025.FIELDS['anomaly_action']}{row}"
                                self._worksheet[key_value] = anomaly_action
                                XlsxHandlerBuildsDiagnostics2025.__format_fonts(
                                    cell=self._worksheet[key_value],
                                    size=11,
                                    wrap_text=True,
                                )
                                anomaly_degree = next(
                                    (
                                        degree[1]
                                        for degree in anomalies_degrees
                                        if degree[0].lower()
                                        == anomaly_identified.lower()
                                    ),
                                    "",
                                )
                                key_value = f"{XlsxHandlerBuildsDiagnostics2025.FIELDS['anomaly_degree']}{row}"
                                self._worksheet[key_value] = anomaly_degree
                                XlsxHandlerBuildsDiagnostics2025.__format_fonts(
                                    cell=self._worksheet[key_value],
                                    size=11,
                                )
                                photo = ""
                                if photo_future is not None:
                                    photo = photo_future.result()
                                if photo:
                                    key_value = f"{XlsxHandlerBuildsDiagnostics2025.FIELDS['photos']}{row}"
                                    insert_picture_2(
                                        self._worksheet,
                                        key_value,
                                        Image(photo),
                                        self.__sheet_target,
                                        (5, 20, 5, 20),
                                    )

                                for internal_key, value in items.items():
                                    if internal_key not in [
                                        "road_name",
                                        "found_at",
                                        "anomalies",
                                        "anomalies_degree",
                                        "photos",
                                    ]:
                                        key_value = f"{XlsxHandlerBuildsDiagnostics2025.FIELDS[internal_key]}{row}"
                                        self._worksheet[key_value] = value
                                        XlsxHandlerBuildsDiagnostics2025.__format_fonts(
                                            cell=self._worksheet[key_value],
                                            size=11,
                                        )
                                self.__insert_new_rows(
                                    row=row, has_image=True if photo else False
                                )
                                row += 1
                        else:
                            for internal_key, value in items.items():
                                if internal_key not in [
                                    "road_name",
                                    "found_at",
                                    "anomalies",
                                    "anomalies_degree",
                                ]:
                                    key_value = f"{XlsxHandlerBuildsDiagnostics2025.FIELDS[internal_key]}{row}"
                                    self._worksheet[key_value] = value
                                    XlsxHandlerBuildsDiagnostics2025.__format_fonts(
                                        cell=self._worksheet[key_value],
                                        size=11,
                                    )
                            self.__insert_new_rows(row=row)
                            row += 1
                        count += 1

                if count == len(values_in_year):
                    self.list_name_files.append(key_year)
                    file_name = "Relatório ANTT - Tabela de Intervenções - Classificação Ruim - {} - {}".format(
                        items["road_name"], key_year
                    )
                    file_name = clean_latin_string(
                        file_name.replace(".", "").replace("/", "")
                    )
                    file_path = f"/tmp/{file_name}.xlsx"
                    row = 4
                    count = 0

                    if self.__report_format == ReportFormat.PDF:
                        for row in range(4, self._worksheet.max_row + 1):
                            for col in range(1, 8):
                                cell = self._worksheet.cell(row, col)
                                XlsxHandlerBuildsDiagnostics2025.__format_fonts(
                                    cell=cell,
                                    size=11,
                                    wrap_text=True,
                                )
                                dimension = self._worksheet.row_dimensions[row]
                                dimension.height = None

                    self.wb.save(file_path)
                    list_files.append(file_path)
                    XlsxHandlerBuildsDiagnostics2025.clear_all_data(
                        file_path, self._worksheet, row
                    )

        temp_dir = tempfile.mkdtemp()
        file_name = self.__format_name_zip_file(
            "Relatório ANTT - Tabela de Intervenções - Classificação Ruim",
            self.list_name_files,
        )
        path_file = os.path.join(temp_dir, file_name)
        saved_files = []
        if self.__report_format == ReportFormat.PDF:
            saved_files = convert_files_to_pdf(list_files)
        else:
            saved_files = list_files
        with ZipFile(path_file, "w") as zipObj:
            for file_path in saved_files:
                zipObj.write(file_path, os.path.basename(file_path))

        return path_file

    @classmethod
    def sorted_values(cls, values: list):
        sorted_data = sorted(values, key=lambda x: (x["buildings"], x["km"]))
        return sorted_data

    def __format_name_zip_file(self, name_string, list_name):
        if len(list_name) > 1:
            unique_names = sorted(set(list_name))
            name_roads = "-".join(map(str, unique_names))
        else:
            name_roads = list_name[0]

        name_zip_file = f"{name_string} - {name_roads}.zip"
        return name_zip_file

    @classmethod
    def clear_all_data(cls, file_path, sheet_name, initial_row):
        wb = load_workbook(file_path)
        sheet = sheet_name

        for row in sheet.iter_rows(
            min_row=initial_row,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        ):
            for cell in row:
                cell.value = None
                cell.border = None
        wb.save(file_path)

    @classmethod
    def __format_fonts(
        cls,
        *,
        cell,
        name="Cabrini",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        wrap_text=False,
    ) -> None:
        cell.alignment = Alignment(
            wrap_text=wrap_text, horizontal=horizontal, vertical=vertical
        )
        cell.font = Font(name=name, sz=size, bold=bold)

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type=self.form,
            uuid__in=self.list_uuids,
            form_data__general_conservation_state="3",
        ).prefetch_related("company", "occurrence_type", "firm", "firm__subcompany")
        list_reporting = [_ for _ in query_set if str(_.uuid) in self.list_uuids]
        data = [self.create_dict(reporting=reporting) for reporting in list_reporting]

        for item in data:
            road_name = item.get("road_name", None)
            found_at = item.get("found_at", None)

            if road_name in self.dict_filtered_roads:
                year_dict = next(
                    (
                        d
                        for d in self.dict_filtered_roads[road_name]
                        if found_at.year in d
                    ),
                    None,
                )

                if year_dict:
                    year_dict[found_at.year].append(item)
                else:
                    year_dict = {found_at.year: [item]}
                    self.dict_filtered_roads[road_name].append(year_dict)
            else:
                self.dict_filtered_roads[road_name] = [{found_at.year: [item]}]
        self.dict_filtered_roads = {
            k: v
            for k, v in self.dict_filtered_roads.items()
            if v and any(item for sublist in v for item in sublist)
        }

        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)
        return result_file


class CCrBuildingDiagnostics2025(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):

        file_name = "Relatório ANTT - Tabela de Intervenções - Classificação Ruim"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        return f"{file_name}.zip"

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerBuildsDiagnostics2025(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            zipObj.write(files, files.split("/")[-1])
        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_building_diagnostics_2025_async_handler(
    reporter_dict: dict,
):
    reporter = CCrBuildingDiagnostics2025.from_dict(reporter_dict)
    reporter.export()
