import math
import re
import tempfile
from concurrent.futures import ThreadPoolExecutor
from typing import List
from zipfile import ZipFile

from django.db.models import Prefetch
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from PIL import UnidentifiedImageError
from zappa.asynchronous import task

from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.ccr_report import CCRReport
from helpers.apps.ccr_report_utils.export_utils import (
    formatted_m_area,
    get_direction_letter,
    get_s3,
    upload_file,
)
from helpers.apps.ccr_report_utils.form_data import new_get_form_data
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    download_files_pictures,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.pdf import ThreadExecutor, synchronized_request_pdf
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option, get_km
from helpers.strings import clean_latin_string

THREADING_LIMIT = 30


def download_picture(
    s3,
    temp_dir: str,
    id_pictures,
    uuid_to_rf: dict,
    id_pictures_to_path: dict,
    low_quality=None,
):
    rfs = []
    if isinstance(id_pictures, list):
        rfs = [uuid_to_rf.get(str(id_picture)) for id_picture in id_pictures]
        rfs = [rf for rf in rfs if rf is not None]
    else:
        rfs = [uuid_to_rf.get(str(id_pictures))]

    try:
        rfs = sorted(rfs, key=lambda rf: (rf.datetime, rf.uploaded_at))
        result = download_files_pictures(
            s3, temp_dir, rfs, 0, 0, limit=1, low_quality=low_quality
        )
        if result.get("status"):
            id_pictures_to_path[id_pictures] = result["images"][0]
    except Exception:
        pass


class XlsxHandlerBuilds:
    def __init__(
        self,
        list_uuids: List[str] = None,
        s3=None,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self.__sheet_target = sheet_target
        self.__report_format = report_format
        self.s3 = s3
        self.count = 0
        self.count_sheets = 0
        self.filename = " "
        self.list_photo_builds = list()
        self.temp_file = tempfile.mkdtemp()
        self.list_uuids = list_uuids
        self.data_logo_company: dict = dict(
            path_image="",
            range_string="Y1:AC3",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:E3",
            resize_method=ResizeMethod.ProportionalLeft,
        )
        first_reporting = Reporting.objects.get(uuid=list_uuids[0])
        self.form_uuid = first_reporting.occurrence_type.uuid

        self.photo_count = 1

        self.static_fields = {
            "id_ccr_antt": "E5",
            "road_name": "E7",
            "km": "E8",
            "direction": "E9",
            "executed_at": "S5",
            "local": "S7",
            "type": "S8",
            "side": "S9",
            "built_area": "I11",
            "parking_area": "I12",
            "covered_parking_area": "I13",
            "building_description": "F15",
            "type_construction": "I18",
            "conservation_state": {"Bom": "L20", "Regular": "S20", "Ruim": "Y20"},
            "anomalies": {
                "01 Fundações e Estruturas": "V8",
                "02 Revestimento de piso (cerâmico, polímero)": "V9",
                "03 Revestimento de azulejo (cerâmico, polímero)": "V10",
                "04 Calçada": "V11",
                "05 Parede externa": "V12",
                "06 Parede interna": "V13",
                "07 Estrutura metálica": "V14",
                "08 Cobertura / Forro": "V15",
                "09 Climatização": "V16",
                "10 Portas": "V17",
                "11 Janelas (vidro e armação metálica)": "V18",
                "12 Iluminação (interna e externa)": "V19",
                "13 Instalação elétrica (interna e externa)": "V20",
                "14 Instalação hidrossanitária - Pia / Tanque de Lavar": "V21",
                "15 Instalação hidrossanitária - Vaso Sanitário / Mictório": "V22",
                "16 Instalação hidrossanitária - Torneira / Registros/ chuveiros": "V23",
                "17 Paisagismo": "V24",
                "18 Caixa d'água": "V25",
                "19 instalação e telefonia": "V26",
                "20 Pintura Externa": "V27",
                "21 Pintura interna": "V28",
                "22 Sistema de proteção de descarga atmosférica (SPDA)": "V29",
                "23 Cercas e alambrados": "V30",
                "24 Utilidades (armários, gavetas)": "V31",
                "25 Atendimento aos padrões de acessibilidade exigidos na NBR 9 050/2020 da ABNT": "V32",
                "notes": "A35",
            },
        }

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.dict_filtered_roads = {
            "BR-116 SP": [],
            "BR-101 SP": [],
            "BR-116 RJ": [],
            "BR-101 RJ": [],
        }

        self.plan_drawing_amount = 0
        self.elevation_drawing_amount = 0
        self.anomalies_photos_amount = 0
        self.general_photos_amount = 0
        self.init_worksheet()

    def init_worksheet(self):
        self.wb = load_workbook("./fixtures/reports/ccr_report_builds.xlsx")
        self._worksheet = self.wb["FICHA 01"]
        insert_logo_and_provider_logo(
            worksheet=self.wb["FICHA 01"],
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        self.anomalies = self.wb["ANOMALIAS"]
        insert_logo_and_provider_logo(
            worksheet=self.anomalies,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        self.general_photos_worksheet = self.wb["FICHA 02"]
        self.plan_photos = self.wb["PLANTA 01"]
        self.elevations_photos = self.wb["ELEVAÇÕES 01"]
        self.anomalies_photos = self.wb["FOTO 01"]

    def insert_new_rows(self, row: int):
        chars = [chr(ord("A") + i) for i in range(13)]

        for char in chars:
            self._worksheet[f"{char}{row}"].border = self.border
            self._worksheet.row_dimensions[row].height = 20

    def insert_status_values(self, cell, value, key, internal_key, specific_worksheet):
        value = value if value else ""
        if internal_key == "conservation_state":
            self._worksheet[cell] = value
            self.format_fonts(
                cell=self._worksheet[cell],
                size=11,
                horizontal="center",
                bold=True,
                italic=True,
            )

        elif key == "notes":
            self.format_fonts(
                cell=specific_worksheet[cell],
                size=11,
                vertical="top",
                horizontal="left",
                name="Arial",
                italic=True,
            )
            specific_worksheet[cell] = value
        else:
            self.format_fonts(
                cell=specific_worksheet[cell], size=11, horizontal="center", italic=True
            )
            specific_worksheet[cell] = value

    def create_dict(self, reporting):
        self.photo_count = 1
        dict_images = dict()
        id_ccr_antt = new_get_form_data(reporting, "idCcrAntt")
        result_road_name = reporting.road_name

        result_km = reporting.km

        km = result_km if result_km else ""

        direction = get_custom_option(reporting, "direction")

        side = new_get_form_data(reporting, "sideedificacaoInventario")

        built_area = new_get_form_data(
            reporting,
            "edifiedAreaedificacaoInventario",
        )
        parking_area = new_get_form_data(
            reporting,
            "parkingAreaedificacaoInventario",
        )

        covered_parking_area = new_get_form_data(
            reporting,
            "coveredParkingAreaedificacaoInventario",
        )

        building_description = new_get_form_data(
            reporting,
            "descriptionedificacaoInventario",
        )

        type_construction = new_get_form_data(
            reporting,
            "constructionKindedificacaoInventario",
        )

        local = new_get_form_data(reporting, "city")

        type_build = new_get_form_data(reporting, "kindedificacaoInventario")

        build_pictures = reporting.form_data.get("edification_pictures")
        uf = reporting.form_data.get("uf", None)
        if build_pictures:
            build_pictures = self.get_build_pictures(
                reporting,
                build_pictures,
                key_one="inner_edification_pictures_fotos_externas",
                key_two="inner_edification_pictures_fotos_internas",
                road_name=result_road_name,
                km=km,
                uf=uf,
                direction=direction,
                needs_code=True,
            )
            external_pictures = build_pictures[
                "inner_edification_pictures_fotos_externas"
            ]
            internal_pictures = build_pictures[
                "inner_edification_pictures_fotos_internas"
            ]

            dict_images["general_pictures"] = (
                external_pictures + internal_pictures
                if external_pictures and internal_pictures
                else external_pictures or internal_pictures
            )

        build_drawings_pictures = reporting.form_data.get("edification_drawings")
        if build_drawings_pictures:
            build_drawings_pictures = self.get_build_pictures(
                reporting,
                build_drawings_pictures,
                key_one="inner_edification_elevation_drawings",
                key_two="inner_edification_plant_drawings",
            )

            dict_images["elevation_pictures"] = build_drawings_pictures[
                "inner_edification_elevation_drawings"
            ]
            dict_images["plant_pictures"] = build_drawings_pictures[
                "inner_edification_plant_drawings"
            ]

        therapy = reporting.form_data.get("therapy")
        dict_images["treatment_images"] = []
        if therapy:
            try:
                therapy_images_uuids = []
                for images in therapy:
                    if images["treatment_images"]:
                        therapy_images_uuids.extend(images["treatment_images"])

                uuid_to_rf = {
                    str(rf.uuid): rf for rf in reporting.reporting_files.all()
                }
                rfs = [uuid_to_rf.get(uuid) for uuid in therapy_images_uuids]
                rfs = [rf for rf in rfs if rf is not None]
                uuid_to_rf = {str(rf.uuid): rf for rf in rfs}
                uuid_to_path = {}
                s3 = get_s3(max_pool_connections=THREADING_LIMIT)

                thread_pool = ThreadPoolExecutor(max_workers=THREADING_LIMIT)
                for uuid in therapy_images_uuids:
                    thread_pool.submit(
                        download_picture,
                        s3,
                        self.temp_file,
                        uuid,
                        uuid_to_rf,
                        uuid_to_path,
                        low_quality=True,
                    )

                thread_pool.shutdown()
                for images in therapy:
                    if images["treatment_images"]:
                        treatment_images = self.get_treatment_pictures(
                            images,
                            uuid_to_path,
                            road_name=result_road_name,
                            km=km,
                            uf=uf,
                            direction=direction,
                        )
                        dict_images["treatment_images"] += treatment_images
            except KeyError:
                pass

        anomalies = self.get_anomalies(reporting)

        result_conservation_state = new_get_form_data(
            reporting,
            "generalConservationState",
        )

        conservation_state = {
            "Bom": "x" if result_conservation_state == "Bom" else "",
            "Regular": "x" if result_conservation_state == "Regular" else "",
            "Ruim": "x" if result_conservation_state == "Ruim" else "",
        }

        data = {
            "number": reporting.number,
            "id_ccr_antt": id_ccr_antt,
            "road_name": result_road_name,
            "km": get_km(reporting),
            "direction": direction,
            "local": local,
            "type": type_build,
            "side": side,
            "built_area": formatted_m_area(built_area) if built_area else "-",
            "parking_area": formatted_m_area(parking_area) if parking_area else "-",
            "covered_parking_area": (
                formatted_m_area(covered_parking_area) if covered_parking_area else "-"
            ),
            "building_description": building_description,
            "type_construction": type_construction,
            "executed_at": reporting.executed_at,
            "found_at": reporting.found_at,
            "anomalies": anomalies,
            "conservation_state": conservation_state,
        }
        data.update(dict_images)
        for k, v in data.items():
            if v is None:
                data[k] = ""

        return data

    def get_anomalies(self, reporting):
        foundation_structures = new_get_form_data(
            reporting,
            "foundationStructures",
        )
        floor_coating = new_get_form_data(reporting, "floorCoating")
        tile_coating = new_get_form_data(reporting, "tileCoating")
        sidewalk = new_get_form_data(reporting, "sidewalk")
        external_wall = new_get_form_data(reporting, "externalWall")
        internal_wall = new_get_form_data(reporting, "internalWall")
        metal_structure = new_get_form_data(reporting, "metalStructure")
        cealing = new_get_form_data(reporting, "cealing")
        climatization = new_get_form_data(reporting, "climatization")
        doors = new_get_form_data(reporting, "doors")
        windows = new_get_form_data(reporting, "windows")
        ilumination = new_get_form_data(reporting, "ilumination")
        electric_instalation = new_get_form_data(
            reporting,
            "electricInstalation",
        )
        sink_hidro_instalation = new_get_form_data(
            reporting,
            "sinkHidroInstalation",
        )
        toilet_hidro_instalation = new_get_form_data(
            reporting,
            "toiletHidroInstalation",
        )
        faucet_hidro_instalation = new_get_form_data(
            reporting,
            "faucetHidroInstalation",
        )
        landscape = new_get_form_data(reporting, "landscape")
        water_box = new_get_form_data(reporting, "waterBox")
        telephone_instalation = new_get_form_data(
            reporting,
            "telephoneInstalation",
        )
        external_painting = new_get_form_data(reporting, "externalPainting")
        internal_painting = new_get_form_data(reporting, "internalPainting")
        spda_protection = new_get_form_data(reporting, "spdaProtection")
        fences = new_get_form_data(reporting, "fences")
        utilitys = new_get_form_data(reporting, "utilitys")
        abnt_requirement = new_get_form_data(reporting, "abntRequirement")
        notes = new_get_form_data(reporting, "notes")

        result_dict = {
            "01 Fundações e Estruturas": foundation_structures,
            "02 Revestimento de piso (cerâmico, polímero)": floor_coating,
            "03 Revestimento de azulejo (cerâmico, polímero)": tile_coating,
            "04 Calçada": sidewalk,
            "05 Parede externa": external_wall,
            "06 Parede interna": internal_wall,
            "07 Estrutura metálica": metal_structure,
            "08 Cobertura / Forro": cealing,
            "09 Climatização": climatization,
            "10 Portas": doors,
            "11 Janelas (vidro e armação metálica)": windows,
            "12 Iluminação (interna e externa)": ilumination,
            "13 Instalação elétrica (interna e externa)": electric_instalation,
            "14 Instalação hidrossanitária - Pia / Tanque de Lavar": sink_hidro_instalation,
            "15 Instalação hidrossanitária - Vaso Sanitário / Mictório": toilet_hidro_instalation,
            "16 Instalação hidrossanitária - Torneira / Registros/ chuveiros": faucet_hidro_instalation,
            "17 Paisagismo": landscape,
            "18 Caixa d'água": water_box,
            "19 instalação e telefonia": telephone_instalation,
            "20 Pintura Externa": external_painting,
            "21 Pintura interna": internal_painting,
            "22 Sistema de proteção de descarga atmosférica (SPDA)": spda_protection,
            "23 Cercas e alambrados": fences,
            "24 Utilidades (armários, gavetas)": utilitys,
            "25 Atendimento aos padrões de acessibilidade exigidos na NBR 9 050/2020 da ABNT": abnt_requirement,
            "notes": notes,
        }
        for key, value in result_dict.items():
            if not value and key != "notes":
                result_dict[key] = "Não se aplica"
        return result_dict

    def get_treatment_pictures(
        self,
        treatment_pictures,
        uuid_to_path,
        road_name=None,
        km=None,
        uf=None,
        direction=None,
    ):
        lista_values = list()
        for images in treatment_pictures["treatment_images"]:
            photo = uuid_to_path.get(images, None)
            if photo:
                count = self.get_photo_count()
                values = [
                    {
                        "photo": photo["path"],
                        "photo_id": self.create_photo_id(
                            road_name, km, uf, count, direction
                        ),
                        "description": photo["description"],
                        "count": count,
                    }
                ]
                lista_values += values

        return lista_values

    def get_build_pictures(
        self,
        reporting,
        build_pictures,
        key_one,
        road_name=None,
        km=None,
        uf=None,
        direction=None,
        key_two=None,
        key_three=None,
        key_four=None,
        needs_code=False,
    ):
        dict_picture = dict()
        dict_picture[key_one] = []
        optional_key = [key_two, key_three, key_four]

        for op in optional_key:
            if op is not None:
                dict_picture[op] = []

        id_pictures_list = []
        rf_uuids = []

        for item in build_pictures:
            for key, values in item.items():
                if not values:
                    dict_picture[key] = ""
                for index, id_pictures in enumerate(values):
                    if key in dict_picture.keys():
                        id_pictures_list.append(id_pictures)
                        if isinstance(id_pictures, list):
                            rf_uuids.extend(id_pictures)
                        else:
                            rf_uuids.append(id_pictures)

        rfs = ReportingFile.objects.filter(uuid__in=rf_uuids)
        uuid_to_rf = {str(rf.uuid): rf for rf in reporting.reporting_files.all()}
        rfs = [uuid_to_rf.get(uuid) for uuid in rf_uuids]
        rfs = [rf for rf in rfs if rf is not None]
        uuid_to_rf = {str(rf.uuid): rf for rf in rfs}

        id_pictures_to_path = {}

        thread_pool = ThreadPoolExecutor(max_workers=THREADING_LIMIT)

        s3 = get_s3(max_pool_connections=THREADING_LIMIT)
        for id_picture in id_pictures_list:
            thread_pool.submit(
                download_picture,
                s3,
                self.temp_file,
                id_picture,
                uuid_to_rf,
                id_pictures_to_path,
                low_quality=False,
            )

        thread_pool.shutdown()

        for item in build_pictures:
            for key, values in item.items():
                if not values:
                    dict_picture[key] = ""
                for index, id_pictures in enumerate(values):
                    if key in dict_picture.keys():
                        photo = id_pictures_to_path.get(id_pictures, None)
                        if photo:
                            if needs_code:
                                count = self.get_photo_count()
                            else:
                                count = 0
                            values = [
                                {
                                    "photo": photo["path"],
                                    "photo_id": self.create_photo_id(
                                        road_name, km, uf, count, direction
                                    ),
                                    "description": photo["description"],
                                    "count": count,
                                }
                            ]
                            dict_picture[key] += values

        return dict_picture

    def get_photo_count(self):
        previous_value = self.photo_count
        self.photo_count += 1
        return previous_value

    def create_photo_id(self, road_name, km, uf, index, direction):
        num_road_name = re.search(r"\d+", road_name).group() if road_name else ""
        km_mt = str(km).split(".")
        km = km_mt[0].rjust(3, "0")

        try:
            mt = km_mt[1].ljust(3, "0")
        except IndexError:
            mt = "000"

        uf = uf if uf else ""
        index = str(index).zfill(3)
        direction = get_direction_letter(direction=direction) if direction else ""
        photo_id = f"EDF{num_road_name}{uf}{km}{mt}{direction}F{index}:"

        return photo_id

    def insert_img(
        self,
        active_worksheet,
        image: str,
        row_init: str,
        scale_width: float,
        scale_height: float,
        width: int = 137,
        height: int = 120,
        fill_cell=False,
    ):
        try:
            if image:
                img = Image(image)
                if fill_cell:

                    xy = coordinate_from_string(row_init)
                    col = column_index_from_string(xy[0])
                    row = xy[1]

                    cell = active_worksheet.cell(row=row, column=col)

                    offset = 0
                    _from = AnchorMarker(
                        col=col - 1,
                        row=row - 1,
                        colOff=offset,
                        rowOff=offset,
                    )

                    merged_ranges_start_cells = [
                        a.start_cell.coordinate
                        for a in active_worksheet.merged_cells.ranges
                    ]
                    _to = None

                    if cell.coordinate in merged_ranges_start_cells:
                        merged_range = next(
                            a
                            for a in active_worksheet.merged_cells.ranges
                            if a.start_cell.coordinate == cell.coordinate
                        )
                        _to = AnchorMarker(
                            col=merged_range.max_col,
                            row=merged_range.max_row,
                            colOff=-offset,
                            rowOff=-offset,
                        )
                    else:
                        _to = AnchorMarker(
                            col=col + 1,
                            row=row + 1,
                            colOff=-offset,
                            rowOff=-offset,
                        )
                    img.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=_to)
                    active_worksheet.add_image(img)
                else:
                    med_width = width / img.width
                    med_height = height / img.height
                    average = min(med_width, med_height)
                    img.width = int(img.width * average) * scale_width
                    img.height = int(img.height * average) * scale_height

                    active_worksheet.add_image(img, row_init)

                return True
            return False

        except UnidentifiedImageError:
            return False

    def fill_sheet(self, *, data_dict: dict):
        list_files = []
        convert_executor: ThreadExecutor = None
        if self.__report_format == ReportFormat.PDF:
            convert_executor = ThreadExecutor(25)
        for key, values_list in data_dict.items():
            count = 0
            files_list = []
            for values in values_list:
                for year_key, list_year_values in values.items():
                    files_list.clear()
                    for values_in_year in list_year_values:
                        del values_in_year["found_at"]
                        self.init_worksheet()
                        for internal_key, value in values_in_year.items():
                            if internal_key in ["anomalies", "conservation_state"]:
                                self.list_photo_builds = []
                                for _key, _value in value.items():
                                    key_value = self.static_fields[internal_key][_key]
                                    self.insert_status_values(
                                        cell=key_value,
                                        value=_value,
                                        internal_key=internal_key,
                                        key=_key,
                                        specific_worksheet=self.anomalies,
                                    )

                            elif internal_key in ["general_pictures"]:
                                self.insert_build_photo(
                                    list_photos=value,
                                    row_init=26,
                                    amount_images=4,
                                    active_worksheet=self._worksheet,
                                )
                                if value:
                                    new_values = self.remove_values_inserted(
                                        value, self.list_photo_builds
                                    )
                                    self.remaining_photos(
                                        remaining_photos_list=new_values,
                                        row_init=9,
                                        name_new_sheet=self.general_photos_worksheet,
                                    )

                            elif internal_key == "treatment_images":
                                self.remaining_photos(
                                    remaining_photos_list=value,
                                    row_init=9,
                                    name_new_sheet=self.anomalies_photos,
                                )

                            elif internal_key in [
                                "elevation_pictures",
                                "plant_pictures",
                            ]:
                                self.sketch_photos(
                                    list_photos_sketch=value,
                                    row_init=6,
                                    name_new_sheet=(
                                        self.plan_photos
                                        if internal_key == "plant_pictures"
                                        else self.elevations_photos
                                    ),
                                )
                            else:
                                if internal_key in ["executed_at"]:
                                    try:
                                        value = value.date().strftime("%d/%m/%Y")
                                    except Exception:
                                        value = "-"
                                key_value = self.static_fields.get(internal_key)
                                if key_value:
                                    self._worksheet[key_value] = value
                        self.filename = (
                            values_in_year["id_ccr_antt"]
                            if values_in_year["id_ccr_antt"]
                            else values_in_year["number"]
                        )
                        count += 1
                        filename = f"/tmp/{self.filename}.xlsx"
                        sheets = self.wb._sheets
                        new_sheets = []

                        for sheet in sheets:
                            title = sheet.title.split(" ")[0]
                            search = [x for x in new_sheets if title in x.title]
                            if any(search):
                                index = new_sheets.index(search[-1])
                                new_sheets.insert(index + 1, sheet)
                                continue
                            new_sheets.append(sheet)

                        self.wb._sheets = new_sheets
                        if self.plan_drawing_amount == 0:
                            self.wb.remove(self.plan_photos)
                        if self.elevation_drawing_amount == 0:
                            self.wb.remove(self.elevations_photos)
                        if self.anomalies_photos_amount == 0:
                            self.wb.remove(self.anomalies_photos)
                        if self.general_photos_amount == 0:
                            self.wb.remove(self.general_photos_worksheet)

                        self.wb.save(filename)

                        self.plan_drawing_amount = 0
                        self.elevation_drawing_amount = 0
                        self.anomalies_photos_amount = 0
                        self.general_photos_amount = 0

                        if self.__report_format == ReportFormat.PDF:
                            convert_executor.submit(
                                synchronized_request_pdf, filename, filename
                            )
                        files_list.append(filename)
                    if count == len(files_list):
                        if self.__report_format == ReportFormat.PDF:
                            files_list = convert_executor.get()
                        files_list = list(set(files_list))

                        sorted_list = sorted(files_list)
                        temp_dict = tempfile.mkdtemp()
                        dir_tmp_name = f"{key}{year_key}"
                        path_file = f"{temp_dict}/{dir_tmp_name}.zip"

                        with ZipFile(path_file, "w") as zipObj:
                            for file in sorted_list:
                                zipObj.write(file, file.split("/")[-1])
                        list_files.append(path_file)
                        count = 0
        return list_files

    def remove_values_inserted(self, list_value, list_value_inserted):
        list_value_copy = list_value.copy()

        new_list_value = [
            dict(item) for item in list_value_copy if item not in list_value_inserted
        ]
        return new_list_value

    def insert_build_photo(
        self, list_photos, row_init, amount_images, active_worksheet
    ):
        current_worksheet = active_worksheet
        insert_logo_and_provider_logo(
            worksheet=current_worksheet,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        sorted_list = sorted(list_photos, key=lambda x: x["count"])
        for value in sorted_list[:4]:
            left_side = value["count"] % 2
            row = row_init + math.floor((value["count"] - 1) / 2) * 22
            coord = f"B{row}:N{row+15}" if left_side else f"P{row}:AB{row+15}"
            desc_coord = f"B{row}" if left_side else f"P{row}"

            image = False
            try:
                insert_picture_2(
                    current_worksheet,
                    coord,
                    Image(value["photo"]),
                    self.__sheet_target,
                    (5, 5, 5, 5),
                    ResizeMethod.ProportionalCentered,
                )
                image = True
            except Exception as e:
                print(e)

            if image and value["photo"]:
                self.description_photo(
                    desc_coord,
                    value["photo_id"],
                    current_worksheet=current_worksheet,
                    photo_string=value["description"],
                )

            if value:
                self.list_photo_builds.append(value)

    def description_photo(self, coord, photo_id, photo_string, current_worksheet):
        coord_for_description = current_worksheet[coord].offset(row=16)
        self.format_fonts(
            cell=coord_for_description, size=8, vertical="top", horizontal="left"
        )
        coord_for_description.value = photo_id
        coord_for_description.offset(column=4).value = photo_string
        self.format_fonts(
            cell=coord_for_description, size=8, vertical="top", horizontal="left"
        )

    def remaining_photos(self, remaining_photos_list, row_init, name_new_sheet):
        current_worksheet = name_new_sheet
        index = 0
        photo_count = 0

        insert_logo_and_provider_logo(
            worksheet=current_worksheet,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        if remaining_photos_list:
            remaining_photos_list = sorted(
                remaining_photos_list, key=lambda x: x["count"]
            )

            for photo in remaining_photos_list:
                if photo_count == 6:
                    index += 1
                    count_fichas = (
                        index + 2
                        if name_new_sheet.title.split(" ")[0] == "FICHA"
                        else index + 1
                    )
                    current_worksheet = self.create_worksheet(
                        title=name_new_sheet, index=count_fichas
                    )
                    current_worksheet.print_area = "A1:AC74"
                    photo_count = 0

                right_side = photo_count % 2
                row = row_init + math.floor((photo_count) / 2) * 22

                coord = f"P{row}:AB{row+15}" if right_side else f"B{row}:N{row+15}"
                desc_coord = f"P{row}" if right_side else f"B{row}"
                image = False
                try:
                    insert_picture_2(
                        current_worksheet,
                        coord,
                        Image(photo["photo"]),
                        self.__sheet_target,
                        (5, 5, 5, 5),
                        ResizeMethod.ProportionalCentered,
                    )
                    image = True
                except Exception as e:
                    print(e)

                if self.anomalies_photos == current_worksheet:
                    self.anomalies_photos_amount += 1
                elif self.general_photos_worksheet == current_worksheet:
                    self.general_photos_amount += 1
                if image and photo["photo"]:
                    self.description_photo(
                        desc_coord,
                        photo["photo_id"],
                        current_worksheet=current_worksheet,
                        photo_string=photo["description"],
                    )

                photo_count += 1

    def sketch_photos(self, list_photos_sketch, row_init, name_new_sheet):
        current_worksheet = name_new_sheet
        insert_logo_and_provider_logo(
            worksheet=current_worksheet,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        row_init = f"A{row_init}"
        for index, photo_sketch in enumerate(list_photos_sketch, 1):
            if index > 1:
                current_worksheet = self.create_worksheet_for_plant_elevations(
                    sheet=name_new_sheet, index=index
                )
                current_worksheet.print_area = "A1:AC73"

            try:
                insert_picture_2(
                    current_worksheet,
                    "A6:AC73",
                    Image(photo_sketch["photo"]),
                    self.__sheet_target,
                    (5, 5, 5, 5),
                    ResizeMethod.ProportionalCentered,
                )
            except Exception as e:
                print(e)

            if name_new_sheet == self.plan_photos:
                self.plan_drawing_amount += 1
            elif name_new_sheet == self.elevations_photos:
                self.elevation_drawing_amount += 1

    def create_worksheet_for_plant_elevations(self, sheet, index=None):
        worksheet = self.wb.copy_worksheet(sheet)

        insert_logo_and_provider_logo(
            worksheet=worksheet,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        title = (sheet.title).split(" ")[0]
        worksheet.title = f"{title} {str(index).zfill(2)}"
        self.clear_all_data(worksheet.title, initial_row=9)
        worksheet.HeaderFooter.oddFooter.left.text = " &F"
        worksheet.HeaderFooter.oddFooter.left.font = "Arial,Regular"
        worksheet.HeaderFooter.oddFooter.left.size = 8
        return worksheet

    def create_worksheet(self, title, index=None):
        worksheet = self.wb.copy_worksheet(title)
        title = "FICHA" if worksheet.title.split(" ")[0] == "FICHA" else "FOTO"
        insert_logo_and_provider_logo(
            worksheet=worksheet,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
            target=self.__sheet_target,
        )
        worksheet.title = f"{title} {str(index).zfill(2)}"
        self.clear_all_data(worksheet.title, initial_row=9)
        worksheet.HeaderFooter.oddFooter.left.text = " &F"
        worksheet.HeaderFooter.oddFooter.left.font = "Arial,Regular"
        worksheet.HeaderFooter.oddFooter.left.size = 8
        return worksheet

    def clear_all_data(self, sheet_name, initial_row):
        sheet = self.wb[sheet_name]

        for row in sheet.iter_rows(
            min_row=initial_row,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        ):
            for cell in row:
                cell.value = None
                # cell.border = None
        while len(sheet._images) > 0:
            del sheet._images[0]

    @classmethod
    def format_fonts(
        cls,
        *,
        cell,
        name="Arial",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        color="FF000000",
        italic=False,
        wrap_text=True,
    ) -> None:
        cell.alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrapText=wrap_text
        )
        cell.font = Font(italic=italic, name=name, sz=size, bold=bold, color=color)

    def execute(self):
        query_set = (
            Reporting.objects.filter(uuid__in=self.list_uuids)
            .filter(occurrence_type__uuid=self.form_uuid)
            .prefetch_related(
                "company",
                "occurrence_type",
                "firm",
                "firm__subcompany",
                Prefetch(
                    "reporting_files",
                    queryset=ReportingFile.objects.filter(is_shared=True),
                ),
            )
        )

        data = []
        for reporting in query_set:
            data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        for item in data:
            road_name = item.get("road_name", None)
            found_at = item.get("found_at", None)
            if item:
                if road_name in self.dict_filtered_roads:
                    year_dict = next(
                        (
                            d
                            for d in self.dict_filtered_roads[road_name]
                            if found_at.year in d
                        ),
                        None,
                    )

                    if year_dict:
                        year_dict[found_at.year] = sorted(
                            year_dict[found_at.year] + [item],
                            key=lambda x: x["id_ccr_antt"],
                        )
                    else:
                        year_dict = {found_at.year: [item]}
                        self.dict_filtered_roads[road_name].append(year_dict)
                else:
                    self.dict_filtered_roads[road_name] = [{found_at.year: [item]}]
        self.dict_filtered_roads = {
            k: v
            for k, v in self.dict_filtered_roads.items()
            if v and any(item for sublist in v for item in sublist)
        }
        sorted(self.dict_filtered_roads)

        result_file = self.fill_sheet(data_dict=self.dict_filtered_roads)
        return result_file


class CCRBuilds(CCRReport):
    def __init__(
        self, uuids: List[str] = None, report_format: ReportFormat = ReportFormat.XLSX
    ) -> None:
        super().__init__(uuids, report_format)

    def get_file_name(self):
        file_name = "Relatório - Edificações"

        file_name = clean_latin_string(file_name.replace(".", "").replace("/", ""))
        file_name = f"{file_name}.zip"

        return file_name

    def export(self):
        s3 = get_s3()
        files = XlsxHandlerBuilds(
            list_uuids=self.uuids,
            s3=s3,
            sheet_target=self.sheet_target(),
            report_format=self.report_format(),
        ).execute()

        result_file = f"/tmp/{self.file_name}"
        with ZipFile(result_file, "w") as zipObj:
            for file in files:
                zipObj.write(file, file.split("/")[-1])

        upload_file(s3, result_file, self.object_name)

        return True


@task
def ccr_report_builds_async_handler(reporter_dict: dict):
    reporter = CCRBuilds.from_dict(reporter_dict)
    reporter.export()
