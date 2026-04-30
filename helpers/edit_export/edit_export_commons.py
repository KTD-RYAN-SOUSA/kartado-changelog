import json
import re
from typing import List, Union

import requests
from django.conf import settings
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from apps.companies.models import Company
from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import RecordMenu
from helpers.permissions import PermissionManager
from helpers.strings import get_obj_from_path, to_snake_case

BOOL_OPTIONS = ["SIM", "NÃO"]
INNER_PICTURE_LIMIT = 3
ARRAY_LIMIT = 5
RESOURCE_LIMIT = 10
PICTURE_LIMIT = 10
VALIDATION_SHEET = "_kartado_validation"


def snake_case_inner_api_names(inner_fields: List[dict]):
    """DFS snake casing inner fields api_names"""
    for field in inner_fields:
        field["apiName"] = to_snake_case(field["apiName"])
        if field["dataType"] == "arrayOfObjects":
            snake_case_inner_api_names(field["innerFields"])


def get_deduplicated_fields(
    occ_type: OccurrenceType, snake_case_api: bool = False
) -> List[dict]:
    """Returns deduplicated fields of occurrence type preserving order.
    First field with a given api name or occurrence type is included
    in the deduped list
    When snake_case_api, apiNames are snake_cased
    """
    fields: list = occ_type.form_fields["fields"]
    api_names = set()
    display_names = set()
    deduped_fields = []
    for field in fields:
        display_name = field.get("displayName")
        data_type = field.get("dataType")
        api_name = field.get("apiName")
        if None in (display_name, data_type, api_name):
            continue
        if api_name not in api_names and display_name not in display_names:
            if snake_case_api:
                field["apiName"] = to_snake_case(api_name)
            if data_type == "arrayOfObjects":
                snake_case_inner_api_names(field["innerFields"])
            deduped_fields.append(field)
    return deduped_fields


def get_custom_options(company: Company, field: str):
    """Dictionary of custom option value to name, given a field"""
    options = {}
    try:
        options_path = f"reporting__fields__{field}__selectOptions__options"
        options = get_obj_from_path(company.custom_options, options_path)
        options = {
            option["value"]: option["name"] for option in options if "name" in option
        }
    except Exception:
        pass
    return options


def hide_inventory(inventory_permission: PermissionManager):
    try:
        return not bool(inventory_permission.get_permission("can_view"))
    except Exception:
        return True


def hide_local(company: Company) -> bool:
    opt = "hide_reporting_location"
    return get_obj_from_path(company.metadata, opt, default_return=False)


def show_coordinate(company: Company) -> bool:
    opt = "show_coordinate_input"
    return get_obj_from_path(company.metadata, opt, default_return=False)


def requires_track(company: Company) -> bool:
    opt = "is_track_required"
    return get_obj_from_path(company.metadata, opt, default_return=False)


def show_track(company: Company) -> bool:
    opt = "show_track"
    return get_obj_from_path(company.metadata, opt, default_return=False)


def get_record_menus(company: Company) -> List[str]:
    return (
        RecordMenu.objects.filter(company=company)
        .exclude(system_default=True)
        .only("name")
        .values_list("name", flat=True)
    )


SHEET_NAME_LIMIT = 28


def get_clean_occ_type_name(occ_type: OccurrenceType):
    title = occ_type.name
    title = re.sub(r"[*:/\\?\[\]]", "_", title)
    return title[:SHEET_NAME_LIMIT]


def get_sheet_title(occ_type: OccurrenceType, occ_types: List[OccurrenceType]) -> str:
    title = get_clean_occ_type_name(occ_type)
    idx = 0
    for other in occ_types:
        if occ_type == other:
            break
        other_title = get_clean_occ_type_name(other)
        if other_title == title:
            idx += 1

    if idx > 0:
        title = f"{title}_{idx}"
    return title


def get_reference_options(
    resource: str, filters: Union[str, dict], auth: str, company: Company
) -> list:
    if isinstance(filters, str):
        filters = json.loads(filters)
    headers = {
        "Accept": "application/vnd.api+json",
        "Content-Type": "application/vnd.api+json",
        "Authorization": auth,
    }
    params = {
        "page": 1,
        "page_size": 99999,
        "company": str(company.uuid),
    }
    if filters is not None:
        for k, v in filters.items():
            if isinstance(v, list):
                v = ",".join(v)
            params[k] = v

    response = requests.get(
        url=f"{settings.BACKEND_URL}/{resource}/", headers=headers, params=params
    )
    return [entry["attributes"] for entry in response.json()["data"]]


def add_list_validation(ws: Worksheet, col: int, values_list: List[str]):
    dv = DataValidation(
        type="list",
        formula1=f"\"{','.join(values_list)}\"",
        allow_blank=True,
        showErrorMessage=True,
        error="Selecione um item da lista.",
        errorTitle="Valor inválido",
    )
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col)
    dv.add("{0}2:{0}1048576".format(col_letter))


def add_cell_list_validation(
    ws: Worksheet,
    list_ws: Worksheet,
    list_col: int,
    list_row: int,
    col: int,
    row: int,
    values_list: List[str],
):
    list_col_letter = get_column_letter(list_col)
    list_row_end = list_row + len(values_list)
    for cell, value in zip(
        list_ws[list_col_letter][list_row:list_row_end], values_list
    ):
        cell.value = value

    quoted_ws_name = quote_sheetname(list_ws.title)
    formula = "{0}!${1}${2}:${1}${3}".format(
        quoted_ws_name, list_col_letter, list_row + 1, list_row_end
    )
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        error="Selecione um item da lista.",
        errorTitle="Valor inválido",
    )
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}{row}")


def is_map_features_select(field: dict):
    return field.get("selectOptions", {}).get("mapFeatures", False) is True
