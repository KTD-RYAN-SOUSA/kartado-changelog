import os
from tempfile import mkdtemp
from typing import List
from uuid import UUID

from dateutil import parser
from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting
from apps.users.models import User
from helpers.apps.ccr_report_utils.export_utils import (
    format_km,
    get_random_string,
    get_s3,
    upload_file,
)
from helpers.dates import utc_to_local
from helpers.edit_export.edit_export_commons import (
    ARRAY_LIMIT,
    BOOL_OPTIONS,
    INNER_PICTURE_LIMIT,
    PICTURE_LIMIT,
    RESOURCE_LIMIT,
    VALIDATION_SHEET,
    add_cell_list_validation,
    get_custom_options,
    get_deduplicated_fields,
    get_record_menus,
    get_reference_options,
    get_sheet_title,
    hide_inventory,
    hide_local,
    is_map_features_select,
    show_coordinate,
    show_track,
)
from helpers.edit_export.edit_export_template import EditTemplate
from helpers.permissions import PermissionManager


class EditExport:
    def __init__(
        self, object_name: str, uuid_strs: List[str], user_uuid_str: str, auth: str
    ):
        self.__object_name = object_name
        self.__s3 = get_s3()
        self.__wb = Workbook()
        self.__user = User.objects.filter(uuid=UUID(user_uuid_str))[0]
        self.__auth = auth

        reporting_uuids = [UUID(uuid) for uuid in uuid_strs]
        self.__reportings = self.__get_reportings(reporting_uuids)
        self.__occ_types = self.__get_occ_types()
        self.__company = self.__reportings[0].company

        inventory_permission = PermissionManager(
            user=self.__user, company_ids=self.__company, model="Inventory"
        )
        self.__hide_inventory = hide_inventory(inventory_permission)
        self.__show_coordinate = show_coordinate(self.__company)
        self.__hide_local = hide_local(self.__company)
        self.__show_track = show_track(self.__company)
        self.__record_menus = get_record_menus(self.__company)

        self.__direction = get_custom_options(self.__company, "direction")
        self.__lane = get_custom_options(self.__company, "lane")
        self.__track = get_custom_options(self.__company, "track")
        self.__branch = get_custom_options(self.__company, "branch")

        self.__row_count = {occ_type: 1 for occ_type in self.__occ_types}
        """Row count for each occurrence type/sheet
            cannot rely on ws.max_row because empty rows were inserted
            to insert formating in cells. Starts in one because of header.
        """

        self.__occ_type_fields = {
            occ_type: get_deduplicated_fields(occ_type, True)
            for occ_type in self.__occ_types
        }

        template = EditTemplate(
            self.__company, self.__occ_types, self.__user, self.__auth
        )
        self.__wb = template.get_wb()

        self.additional_values_col = 1
        self.additional_values_row = 1
        self.__set_additional_values_position()

    def __set_additional_values_position(self):
        ws = self.__wb[VALIDATION_SHEET]
        max_col = ws.max_column
        self.additional_values_col = max_col + 2
        self.additional_values_row = 1

    def __get_reportings(self, reporting_uuids: List[UUID]) -> QuerySet[Reporting]:
        return Reporting.objects.filter(uuid__in=reporting_uuids).prefetch_related(
            "occurrence_type", "company", "firm", "status", "parent", "menu"
        )

    def __get_occ_types(self) -> List[OccurrenceType]:
        occ_types = set()
        for reporting in self.__reportings:
            occ_types.add(reporting.occurrence_type)

        return sorted(occ_types, key=lambda t: t.name)

    def __save_workbook(self, dir: str, name: str) -> str:
        """Saves workbook file in the directory
        and returns file path
        """
        wb_file = os.sep.join((dir, name))
        self.__wb.save(wb_file)
        return wb_file

    def __get_worksheet(self, occ_type: OccurrenceType) -> Worksheet:
        title = get_sheet_title(occ_type, self.__occ_types)
        return self.__wb[title]

    def __get_row(self, occ_type: OccurrenceType) -> Worksheet:
        row = self.__row_count[occ_type] + 1
        self.__row_count[occ_type] = row
        return row

    def __insert_fixed_fields_data(
        self, ws: Worksheet, row: int, reporting: Reporting, occ_type: OccurrenceType
    ) -> int:
        """Inserts fixed fields data. Returns amount of used cols"""
        col = 1
        ws.cell(row, col).value = reporting.number
        col += 1

        if not self.__hide_inventory:
            if reporting.parent is not None:
                inv_num = reporting.parent.number
                ws.cell(row, col).value = inv_num
            col += 1

        if self.__show_coordinate or self.__hide_local:
            if reporting.point is not None:
                latitude = reporting.point.coords[1]
                longitude = reporting.point.coords[0]
                ws.cell(row, col).value = latitude
                ws.cell(row, col + 1).value = longitude
            col += 2

        if not self.__hide_local:
            if reporting.km is not None:
                ws.cell(row, col).value = format_km(reporting.km, 3)
            if reporting.end_km is not None:
                ws.cell(row, col + 1).value = format_km(reporting.end_km, 3)
            if reporting.project_km is not None:
                ws.cell(row, col + 2).value = format_km(reporting.project_km, 3)
            if reporting.project_end_km is not None:
                ws.cell(row, col + 3).value = format_km(reporting.project_end_km, 3)
            col += 4

        status_name = reporting.status.name if reporting.status else None
        firm_name = reporting.firm.name if reporting.firm else None
        ws.cell(row, col).value = status_name
        ws.cell(row, col + 1).value = firm_name
        if reporting.found_at is not None:
            found_at = utc_to_local(reporting.found_at)
            ws.cell(row, col + 2).value = found_at.replace(tzinfo=None)
        if reporting.executed_at is not None:
            executed_at = utc_to_local(reporting.executed_at)
            ws.cell(row, col + 3).value = executed_at.replace(tzinfo=None)
        col += 4

        if len(self.__record_menus) > 1:
            if reporting.menu is not None:
                ws.cell(row, col).value = reporting.menu.name
            col += 1

        if reporting.due_at is not None:
            due_at = utc_to_local(reporting.due_at)
            ws.cell(row, col).value = due_at.replace(tzinfo=None)
        col += 1

        if not self.__hide_local:
            ws.cell(row, col).value = self.__direction.get(reporting.direction)
            col += 1

        ws.cell(row, col).value = occ_type.name
        col += 1

        if not self.__hide_local:
            ws.cell(row, col).value = self.__lane.get(reporting.lane)
            col += 1

        if self.__show_track:
            ws.cell(row, col).value = self.__track.get(reporting.track)
            if reporting.km_reference is not None:
                ws.cell(row, col + 1).value = format_km(reporting.km_reference, 3)
            ws.cell(row, col + 2).value = self.__branch.get(reporting.branch)
            col += 3

        if not self.__hide_local:
            ws.cell(row, col).value = reporting.road_name
            col += 1

        return col

    def __solved_reference(self, field: dict):
        """Given a select form field with reference, gets a dictionary indexing 'optionText's by 'optionValue's."""
        reference = field["selectOptions"]["reference"]
        filter = reference.get("filter")
        resource = reference.get("resource")
        option_text = reference["optionText"]
        option_value = reference["optionValue"]
        options = None
        if resource is not None:
            resource = reference.get("resource")
            ref_options = get_reference_options(
                resource, filter, self.__auth, self.__company
            )
            options = {opt[option_value]: opt[option_text] for opt in ref_options}
        else:
            options = dict(
                list(
                    OccurrenceType.objects.filter(company=self.__company).values_list(
                        option_value, option_text, flat=True
                    )
                )
            )
        return options

    def __index_select_options(self, field: dict):
        """Given a select or select multiple field, creates a 'indexedOptions'
        entry with a dictionary indexing options 'name's by 'values'
        """
        select_options = field.get("selectOptions")
        indexed = {}
        if select_options is not None and "reference" in select_options:
            try:
                indexed = self.__solved_reference(field)
            except Exception:
                pass
        else:
            try:
                indexed = {
                    opt["value"]: opt["name"]
                    for opt in field["selectOptions"]["options"]
                }
            except Exception:
                pass
        del field["selectOptions"]
        field["indexedOptions"] = indexed

    def __get_selected_value(self, field: dict, value) -> str:
        if "selectOptions" in field:
            self.__index_select_options(field)
        if value is not None:
            return field["indexedOptions"].get(value)
        return None

    def __get_multiselected_value(self, field: dict, values: list) -> str:
        if "selectOptions" in field:
            self.__index_select_options(field)
        if values is not None:
            selected = [field["indexedOptions"].get(value) for value in values]
            return ",".join(selected)
        return None

    def __add_map_feature_validation(
        self, ws: Worksheet, col, row, reporting: Reporting
    ):
        features = [p.get("label") for p in reporting.properties if p.get("label")]
        add_cell_list_validation(
            ws,
            self.__wb[VALIDATION_SHEET],
            self.additional_values_col,
            self.additional_values_row,
            col,
            row,
            features,
        )
        self.additional_values_row += len(features)

    def __insert_form_data_value(
        self, ws, row, col, field: dict, data: dict, reporting: Reporting
    ) -> int:
        """Inserts form data value. Recurring when array of objects. Returns updated col state."""
        data_type = field.get("dataType")
        api_name = field.get("apiName")
        if data_type == "innerImagesArray":
            col += INNER_PICTURE_LIMIT * 4
        elif data_type == "arrayOfObjects":
            inner_data = data.get(api_name)
            col = self.__insert_array_values(
                ws, row, col, field.get("innerFields", []), inner_data, reporting
            )
        else:
            value = None
            if data_type == "select":
                if is_map_features_select(field):
                    self.__add_map_feature_validation(ws, col, row, reporting)
                    value = data.get(api_name)
                else:
                    value = self.__get_selected_value(field, data.get(api_name))
            elif data_type == "selectMultiple":
                value = self.__get_multiselected_value(field, data.get(api_name))
            elif data_type == "boolean":
                selected = data.get(api_name)
                if selected is not None:
                    value = BOOL_OPTIONS[int(not bool(selected))]
            elif data_type == "timestamp":
                date_string = data.get(api_name, None)
                try:
                    date = parser.isoparse(date_string)
                    value = date.replace(tzinfo=None)
                except (ValueError, TypeError):
                    pass
            else:
                value = data.get(api_name)
            ws.cell(row, col).value = value
            col += 1
        return col

    def __insert_array_values(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        inner_fields: list,
        inner_array: list,
        reporting: Reporting,
    ) -> int:
        """Inserts arrays inner fields data DFS. Returns updated col state"""
        inner_arr_it = None
        if inner_array is not None:
            inner_arr_it = iter(inner_array)

        for i in range(ARRAY_LIMIT):
            curr_data = {}
            try:
                curr_data = next(inner_arr_it)
            except Exception:
                pass

            for field in inner_fields:
                col = self.__insert_form_data_value(
                    ws, row, col, field, curr_data, reporting
                )

        return col

    def __insert_form_data(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        reporting: Reporting,
        occ_type: OccurrenceType,
    ):
        for field in self.__occ_type_fields[occ_type]:
            data_type = field.get("dataType")
            if data_type != "arrayOfObjects":  # Only non arrays
                col = self.__insert_form_data_value(
                    ws, row, col, field, reporting.form_data, reporting
                )

        col += (
            RESOURCE_LIMIT * 2 + PICTURE_LIMIT * 4
        )  # Skips resource and picture cells

        for field in self.__occ_type_fields[occ_type]:
            data_type = field.get("dataType")
            if data_type == "arrayOfObjects":  # Only arrays
                col = self.__insert_form_data_value(
                    ws, row, col, field, reporting.form_data, reporting
                )

    def __insert_reportings(self):
        for reporting in self.__reportings:
            occ_type = reporting.occurrence_type
            ws = self.__get_worksheet(occ_type)
            row = self.__get_row(occ_type)
            col = self.__insert_fixed_fields_data(ws, row, reporting, occ_type)
            self.__insert_form_data(ws, row, col, reporting, occ_type)

    def export(self):
        edit_mark_sheet = self.__wb.create_sheet("edit_export")
        edit_mark_sheet.sheet_state = Worksheet.SHEETSTATE_VERYHIDDEN

        self.__insert_reportings()

        temp_dir = mkdtemp()
        wb_file = self.__save_workbook(
            temp_dir, f"edit_export_{get_random_string()}.xlsx"
        )
        upload_file(self.__s3, wb_file, self.__object_name)


@task
def create_edit_export(
    object_name: str, reporting_uuids: List[str], user_uuid_str: str, auth: str
):
    EditExport(object_name, reporting_uuids, user_uuid_str, auth).export()
