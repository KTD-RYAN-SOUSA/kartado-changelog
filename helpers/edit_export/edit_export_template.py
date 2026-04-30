from copy import copy
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Color
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Pane
from openpyxl.worksheet.worksheet import Worksheet

from apps.companies.models import Company, Firm
from apps.occurrence_records.models import OccurrenceType
from apps.roads.models import Road
from apps.service_orders.const.status_types import REPORTING_STATUS
from apps.service_orders.models import ServiceOrderActionStatusSpecs
from apps.users.models import User
from helpers.edit_export.edit_export_commons import (
    ARRAY_LIMIT,
    BOOL_OPTIONS,
    INNER_PICTURE_LIMIT,
    PICTURE_LIMIT,
    RESOURCE_LIMIT,
    VALIDATION_SHEET,
    add_list_validation,
    get_deduplicated_fields,
    get_record_menus,
    get_reference_options,
    get_sheet_title,
    hide_inventory,
    hide_local,
    is_map_features_select,
    requires_track,
    show_coordinate,
    show_track,
)
from helpers.permissions import PermissionManager
from helpers.strings import get_obj_from_path, to_camel_case


class EditTemplate:
    __DEFAULT_COL_WIDTH = 15

    __SELECT_MULTIPLE_COMMENT = "Este é um campo de seleção múltipla. Separe os valores com vírgula. As opções são: {}"
    __TEXT_COMMENT = "Este é um campo de texto aberto."
    __SELECT_COMMENT = (
        "Este é um campo de seleção única. Selecione apenas um item da lista."
    )
    __TYPE_TO_COMMENT = {
        "select": __SELECT_COMMENT,
        "float": "Este é um campo numérico com casa decimal.",
        "number": "Este é um campo numérico para valores inteiros. Números com casas decimais serão arredondados.",
        "timestamp": "Este é um campo de data e hora. Insira no formato de data do Excel.",
        "boolean": "Este é um campo de SIM ou NÃO. Escolha apenas uma das opções.",
        "cpf": "Preencha com um CPF válido. Este campo aceita apenas números e tem um limite de 11 dígitos.",
        "phone": "Preencha com um número de telefone válido. Este campo aceita apenas números e tem um limite de 10 a 11 dígitos, contando com o DDD.",
        "textArea": __TEXT_COMMENT,
        "licensePlate": __TEXT_COMMENT,
        "string": __TEXT_COMMENT,
    }
    __PICTURE_TYPE_OPTIONS = ["Antes", "Durante", "Depois", "Outro"]
    __TYPE_TO_FMT = {
        "float": "0.0000",
        "number": "0",
        "timestamp": "dd/mm/yyyy hh:mm",
    }

    __COMMON_COLS_FORMAT = {
        "km": "0.000",
        "km final": "0.000",
        "km de Projeto": "0.000",
        "km final de Projeto": "0.000",
        "Encontrado em": "dd/mm/yyyy hh:mm",
        "Executado em": "dd/mm/yyyy hh:mm",
        "Prazo": "dd/mm/yyyy hh:mm",
    }

    def __init__(
        self,
        company: Company,
        occ_types: List[OccurrenceType],
        user: User,
        auth: str,
        rows: int = 200,
    ):
        self.__wb = load_workbook("./fixtures/reports/edit_export_template.xlsx")
        self.__occ_types = sorted(occ_types, key=lambda t: t.name)
        self.__user = user
        self.__company = company
        self.__rows = rows
        self.__auth = auth

        inventory_permission = PermissionManager(
            user=self.__user, company_ids=self.__company, model="Inventory"
        )
        self.__hide_inventory = hide_inventory(inventory_permission)
        self.__show_coordinate = show_coordinate(self.__company)
        self.__hide_local = hide_local(self.__company)
        self.__show_track = show_track(self.__company)
        self.__record_menus = get_record_menus(self.__company)

        self.create_template()

    def get_wb(self) -> Workbook:
        return self.__wb

    def __get_occ_type_names(self) -> List[str]:
        return [occ_type.name for occ_type in self.__occ_types]

    @classmethod
    def __get_form_data_comment(self, field: dict) -> str:
        """Given a formdata field, returns the comment related to the field type"""
        note = ""
        if field["dataType"] == "selectMultiple":
            options = field["selectOptions"]["options"]
            option_names = [opt["name"] for opt in options]
            options_text = ", ".join(option_names)
            note = self.__SELECT_MULTIPLE_COMMENT.format(options_text)
        else:
            note = self.__TYPE_TO_COMMENT.get(field["dataType"])
        return note

    def __get_field_custom_options(self, field: str) -> List[str]:
        """Given a fixed form field name, returns the list of options names"""
        options_names = []
        try:
            options_path = f"reporting__fields__{field}__selectOptions__options"
            options = get_obj_from_path(self.__company.custom_options, options_path)
            options_names = [option["name"] for option in options if "name" in option]
        except Exception:
            pass

        return options_names

    def __get_road_names(self) -> List[str]:
        """Returns this company road names options"""
        roads = (
            Road.objects.filter(company=self.__company).only("name").order_by("name")
        )
        return list(set(road.name for road in roads))

    def __get_teams_names(self) -> List[str]:
        """Returns this company team names options"""
        teams = (
            Firm.objects.filter(
                company=self.__company, is_company_team=True, active=True
            )
            .only("name")
            .order_by("name")
        )
        return [team.name for team in teams]

    def __get_status_names(self) -> List[str]:
        """Returns this company available reporting status options sorted by order"""
        status_orders = (
            ServiceOrderActionStatusSpecs.objects.filter(
                company=self.__company, status__kind=REPORTING_STATUS
            )
            .only("order", "status__name")
            .prefetch_related("status")
            .order_by("order")
        )
        return [s_order.status.name for s_order in status_orders]

    def __add_valid_options(self, field_name: str, values: List[str]):
        """Adds valid options of a given field (used to create the combobox)"""
        ws = self.__wb[VALIDATION_SHEET]
        col_idx = ws.max_column + 1
        ws.cell(1, col_idx, field_name)
        for row_idx, value in enumerate(values, 2):
            ws.cell(row_idx, col_idx, value)

    @classmethod
    def __get_validation_name(cls, occ_type: OccurrenceType, api_name: str) -> str:
        """given occurrence type and api name
        returns the header text of the column with the possible values
        for that field in the validation sheet
        """
        return f"{occ_type.name} {to_camel_case(api_name)}"

    @classmethod
    def __filter_inline_options(cls, field: dict) -> List[str]:
        """Validates and applies inline optionsFilter for select fields.
        Returns filtered option names based on optionsFilter configuration.
        """
        try:
            select_options = field.get("selectOptions", {})
            options = select_options.get("options", [])
            options_filter = select_options.get("optionsFilter")

            if not options_filter:
                return [opt.get("name", "") for opt in options]

            # optionsFilter must only contain the key 'filter'
            if (
                not isinstance(options_filter, dict)
                or isinstance(options_filter, list)
                or len(options_filter.keys()) != 1
                or "filter" not in options_filter
            ):
                return [opt.get("name", "") for opt in options]

            filter_arr = options_filter["filter"]
            # filter must be an array of exactly two elements
            if not isinstance(filter_arr, list) or len(filter_arr) != 2:
                return [opt.get("name", "") for opt in options]

            first, second = filter_arr
            # first must be exactly { var: 'options' }
            if (
                not isinstance(first, dict)
                or first.get("var") != "options"
                or len(first.keys()) != 1
            ):
                return [opt.get("name", "") for opt in options]

            # second must be { in: [ { var: 'value' }, [whitelist...] ] }
            if not isinstance(second, dict) or "in" not in second:
                return [opt.get("name", "") for opt in options]

            in_arr = second["in"]
            if not isinstance(in_arr, list) or len(in_arr) != 2:
                return [opt.get("name", "") for opt in options]

            left, whitelist = in_arr
            if (
                not isinstance(left, dict)
                or left.get("var") != "value"
                or len(left.keys()) != 1
            ):
                return [opt.get("name", "") for opt in options]

            if not isinstance(whitelist, list):
                return [opt.get("name", "") for opt in options]

            # Convert whitelist to set of strings for efficient lookup
            whitelist_set = set(str(v) for v in whitelist)

            # Filter options based on whitelist
            filtered_options = []
            for opt in options:
                if str(opt.get("value", "")) in whitelist_set:
                    filtered_options.append(opt.get("name", ""))

            return filtered_options

        except Exception:
            # Return all option names if any error occurs
            try:
                return [
                    opt.get("name", "")
                    for opt in field.get("selectOptions", {}).get("options", [])
                ]
            except Exception:
                return []

    def __get_reference_select_options(self, field: dict):
        """Given a select form field with reference, returns the referenced data
        If resource is not set, returns occurrence_type names
        """
        reference = field["selectOptions"]["reference"]
        filter = reference.get("filter")
        resource = reference.get("resource")
        option_text = reference["optionText"]
        if resource is not None:
            resource = reference.get("resource")
            ref_options = get_reference_options(
                resource, filter, self.__auth, self.__company
            )
            return [opt[option_text] for opt in ref_options]
        return list(
            OccurrenceType.objects.filter(company=self.__company).values_list(
                option_text, flat=True
            )
        )

    def __add_form_data_validation(self, occ_type: OccurrenceType, field: dict):
        """Adds validation column for a form data select field"""
        option_names = None
        if "reference" in field["selectOptions"]:
            option_names = self.__get_reference_select_options(field)
        elif "optionsFilter" in field["selectOptions"]:
            option_names = type(self).__filter_inline_options(field)
        else:
            try:
                options = field["selectOptions"]["options"]
                option_names = [opt["name"] for opt in options]
            except Exception:
                option_names = []
        field_name = self.__get_validation_name(occ_type, field["apiName"])
        self.__add_valid_options(field_name, option_names)

    def __add_form_data_valid_options(
        self, occ_type: OccurrenceType, fields: dict = None
    ):
        """Adds validation column for given occurrence type.
        DFS in array of objects.
        """
        if fields is None:
            fields = get_deduplicated_fields(occ_type)
        for field in fields:
            display_name = field.get("displayName")
            data_type = field.get("dataType")
            api_name = field.get("apiName")
            if None in (display_name, data_type, api_name):
                continue
            if data_type == "select":
                self.__add_form_data_validation(occ_type, field)
            elif data_type == "arrayOfObjects":
                self.__add_form_data_valid_options(
                    occ_type, field.get("innerFields", [])
                )

    def __create_valid_options(self):
        """Add validation sheet"""
        validation_sheet = self.__wb[VALIDATION_SHEET]
        self.__add_valid_options("Status", self.__get_status_names())
        self.__add_valid_options("Equipe", self.__get_teams_names())
        self.__add_valid_options(
            "Sentido", self.__get_field_custom_options("direction")
        )
        self.__add_valid_options("Classe", self.__get_occ_type_names())
        self.__add_valid_options("Faixa", self.__get_field_custom_options("lane"))
        self.__add_valid_options("Pista", self.__get_field_custom_options("track"))
        self.__add_valid_options("Ramo", self.__get_field_custom_options("branch"))
        self.__add_valid_options("Rodovia", self.__get_road_names())
        self.__add_valid_options("Menu", self.__record_menus)

        for occ_type in self.__occ_types:
            self.__add_form_data_valid_options(occ_type)
        validation_sheet.sheet_state = Worksheet.SHEETSTATE_VERYHIDDEN
        validation_sheet.delete_cols(1)

    def __add_form_field_header(self, ws: Worksheet, field: dict):
        """Given form field, adds header with comment"""
        cell = ws.cell(1, ws.max_column + 1, field["displayName"])
        comment = self.__get_form_data_comment(field)
        if comment is not None:
            cell.comment = Comment(comment, "")

    def __add_array_field_header(
        self, ws: Worksheet, inner_field: dict, parent: str = ""
    ):
        """Given array inner field, adds header with comment"""
        col = ws.max_column + 1
        cell = ws.cell(1, col, f"{parent}: {inner_field['displayName']}")
        comment = self.__get_form_data_comment(inner_field)
        if comment is not None:
            cell.comment = Comment(comment, "")

    def __add_image_field_header(
        self, ws: Worksheet, inner_field: dict, parent: str = None
    ):
        """Given inner images field, adds headers with comment"""
        prefix = ""
        if parent is not None:
            prefix = f"{parent}: "
        prefix += inner_field["displayName"]
        for i in range(1, INNER_PICTURE_LIMIT + 1):
            ws.cell(1, ws.max_column + 1, f"{prefix} - Foto_{i}")
            ws.cell(1, ws.max_column + 1, f"{prefix} - Descrição Foto_{i}")

            type_cell = ws.cell(1, ws.max_column + 1, f"{prefix} - Tipo Foto_{i}")
            type_cell.comment = Comment(self.__SELECT_COMMENT, "")
            self.__add_picture_type_validation(ws, ws.max_column)

            ws.cell(1, ws.max_column + 1, f"{prefix} - Data Foto_{i}")

    def __add_array_headers(
        self,
        ws: Worksheet,
        occ_type: OccurrenceType,
        fields: dict,
        parent_name: str,
        field_to_validation_col: dict,
    ):
        for inner_field in fields:
            display_name = inner_field.get("displayName")
            data_type = inner_field.get("dataType")
            api_name = inner_field.get("apiName")
            if None in (display_name, data_type, api_name):
                continue
            if data_type == "arrayOfObjects":
                for i in range(1, ARRAY_LIMIT + 1):
                    name = f"{parent_name} {display_name}_{i}"
                    self.__add_array_headers(
                        ws, occ_type, inner_field, name, field_to_validation_col
                    )
            elif data_type == "innerImagesArray":
                self.__add_image_field_header(ws, inner_field, parent_name)
            else:
                self.__add_array_field_header(ws, inner_field, parent_name)
                self.__set_form_field_validations(
                    ws,
                    occ_type,
                    inner_field,
                    ws.max_column,
                    field_to_validation_col,
                )

    def __set_form_field_validations(
        self,
        ws: Worksheet,
        occ_type: OccurrenceType,
        field: dict,
        col: int,
        field_to_valid_col: dict,
    ):
        data_type = field["dataType"]
        if data_type == "boolean":
            self.__add_bool_validation(ws, col)
        elif not is_map_features_select(field):
            validation_name = self.__get_validation_name(occ_type, field["apiName"])
            valid_options_idx = field_to_valid_col.get(validation_name)
            if valid_options_idx is not None:
                self.__add_validation(ws, col, valid_options_idx)
            fmt = self.__TYPE_TO_FMT.get(data_type, None)
            if fmt is not None:
                self.__add_number_format(ws, col, fmt)

    def __add_headers(self, ws: Worksheet, occ_type: OccurrenceType):
        fields = get_deduplicated_fields(occ_type)

        validation_ws: Worksheet = self.__wb[VALIDATION_SHEET]
        field_to_validation_col = {
            cell.value: i for i, cell in enumerate(validation_ws[1], 1)
        }

        for field in fields:
            display_name = field.get("displayName")
            data_type = field.get("dataType")
            api_name = field.get("apiName")
            if None in (display_name, data_type, api_name):
                continue
            if data_type == "innerImagesArray":
                self.__add_image_field_header(ws, field)
            elif data_type != "arrayOfObjects":
                self.__add_form_field_header(ws, field)
                self.__set_form_field_validations(
                    ws, occ_type, field, ws.max_column, field_to_validation_col
                )

        for i in range(1, RESOURCE_LIMIT + 1):
            ws.cell(1, ws.max_column + 1, f"Recurso_{i}")
            ws.cell(1, ws.max_column + 1, f"Quantidade_{i}")

        for i in range(1, PICTURE_LIMIT + 1):
            ws.cell(1, ws.max_column + 1, f"Foto_{i}")
            ws.cell(1, ws.max_column + 1, f"Data Foto_{i}")
            ws.cell(1, ws.max_column + 1, f"Tipo Foto_{i}")
            self.__add_picture_type_validation(ws, ws.max_column)
            ws.cell(1, ws.max_column + 1, f"Descrição Foto_{i}")

        for field in fields:
            display_name = field.get("displayName")
            data_type = field.get("dataType")
            api_name = field.get("apiName")
            if None in (display_name, data_type, api_name):
                continue
            if data_type == "arrayOfObjects":
                for i in range(1, ARRAY_LIMIT + 1):
                    name = f"{display_name}_{i}"
                    self.__add_array_headers(
                        ws,
                        occ_type,
                        field.get("innerFields", []),
                        name,
                        field_to_validation_col,
                    )

    def __add_validation(self, ws: Worksheet, col: int, valid_options_idx: int):
        valid_options_col = get_column_letter(valid_options_idx)
        quoted_ws_name = quote_sheetname(VALIDATION_SHEET)
        options_last_row = 0
        validation_ws: Worksheet = self.__wb[VALIDATION_SHEET]
        for cell in validation_ws[valid_options_col]:
            options_last_row += 1
            if cell.value is None:
                break

        formula = "{0}!${1}$2:${1}${2}".format(
            quoted_ws_name, valid_options_col, options_last_row
        )
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showErrorMessage=True,
            error="Selecione um item da lista.",
            errorTitle="Valor inválido",
        )
        ws.add_data_validation(dv)
        col_letter = get_column_letter(col)
        dv.add("{0}2:{0}1048576".format(col_letter))

    def __add_picture_type_validation(self, ws: Worksheet, col: int):
        add_list_validation(ws, col, self.__PICTURE_TYPE_OPTIONS)

    def __add_bool_validation(self, ws: Worksheet, col: int):
        add_list_validation(ws, col, BOOL_OPTIONS)

    def __add_number_format(self, ws: Worksheet, col: int, number_format: str):
        col_letter = get_column_letter(col)
        cells: Iterable[Cell] = ws[col_letter]
        cell_it = iter(cells)
        next(cell_it)
        for col_cell in cell_it:
            col_cell.number_format = number_format

    def __add_fixed_field_validations(self, ws: Worksheet):
        validation_ws: Worksheet = self.__wb[VALIDATION_SHEET]
        field_to_validation_col = {
            cell.value: i for i, cell in enumerate(validation_ws[1], 1)
        }

        for cell in ws[1]:
            valid_options_idx = field_to_validation_col.get(cell.value)
            fmt = self.__COMMON_COLS_FORMAT.get(cell.value)
            if valid_options_idx is not None:
                self.__add_validation(ws, cell.col_idx, valid_options_idx)
            if fmt is not None:
                self.__add_number_format(ws, cell.col_idx, fmt)

    def __set_header_style(self, ws: Worksheet, common_cols: int):
        ws.sheet_view.pane = Pane(ySplit=1, topLeftCell="A1")
        ws.print_title_rows = "1:1"
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )

        ws.row_dimensions[1].height = 30
        for col in range(common_cols, ws.max_column + 1):
            col_width = self.__DEFAULT_COL_WIDTH
            v = str(ws.cell(1, col).value)
            if (
                v.startswith("Recurso")
                or v.startswith("Foto_")
                or v.startswith("Data Foto_")
                or v.startswith("Descrição Foto_")
            ):
                col_width = 30
            elif v.startswith("Quantidade") or v.startswith("Tipo Foto_"):
                col_width = 20
            elif len(v) > 20:
                col_width = (len(v) * 3.0) / 4.0
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = col_width

    def __remove_disabled_columns(self, ws: Worksheet):
        delete_offset = 0
        for col in range(1, ws.max_column + 1):
            value = ws.cell(1, col - delete_offset).value
            if (
                (
                    value == "Código do Inventário para vinculo com apontamento"
                    and self.__hide_inventory
                )
                or (
                    value in ["Latitude", "Longitude"]
                    and not (self.__show_coordinate or self.__hide_local)
                )
                or (
                    value
                    in [
                        "km",
                        "km final",
                        "km de Projeto",
                        "km final de Projeto",
                        "Sentido",
                        "Faixa",
                        "Rodovia",
                    ]
                    and self.__hide_local
                )
                or (
                    value in ["Pista", "km de referência", "Ramo"]
                    and not self.__show_track
                )
                or (value == "Menu" and len(self.__record_menus) <= 1)
            ):
                ws.delete_cols(col - delete_offset)
                delete_offset += 1

    def __color_required_fixed_fields(self, ws: Worksheet):
        if requires_track(self.__company):
            for col in range(1, ws.max_column + 1):
                cell: Cell = ws.cell(1, col)
                if cell.value in ["km de referência", "Pista", "Ramo"]:
                    cell.fill = copy(ws.cell(1, 1).fill)

    @classmethod
    def __set_sheet_color(self, ws: Worksheet, occ_type: OccurrenceType):
        color = occ_type.color
        if color is None or color == "":
            color = "FFFFFFFF"
        else:
            color = "FF" + color[1:7]

        ws.sheet_properties.tabColor = Color(color)

    def __create_sheet(self, occ_type: OccurrenceType):
        """Creates sheet given occurrence type"""
        template_ws = self.__wb["Template"]
        ws = self.__wb.copy_worksheet(template_ws)
        ws.title = get_sheet_title(occ_type, self.__occ_types)

        self.__remove_disabled_columns(ws)
        self.__set_sheet_color(ws, occ_type)
        ws.insert_rows(2, self.__rows)

        common_cols = ws.max_column
        self.__color_required_fixed_fields(ws)
        self.__add_fixed_field_validations(ws)
        self.__add_headers(ws, occ_type)
        self.__set_header_style(ws, common_cols)

    def create_template(self):
        self.__wb.create_sheet(VALIDATION_SHEET)

        self.__create_valid_options()
        for occ_type in self.__occ_types:
            self.__create_sheet(occ_type)

        self.__wb.remove(self.__wb["Template"])
