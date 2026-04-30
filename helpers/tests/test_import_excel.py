import os
import tempfile
import uuid
from datetime import datetime
from unittest.mock import Mock, patch

from django.test import TestCase
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.companies.models import Firm
from helpers.import_excel.read_excel import ImportExcel
from helpers.testing.fixtures import TestBase


class TestImportExcel(TestBase, TestCase):
    model = "ExcelImport"

    def setUp(self):
        """Setup test environment"""
        self.temp_dir = tempfile.mkdtemp()

        # Mock excel import
        self.excel_import = Mock()
        self.excel_import.company = self.company
        self.excel_import.company_id = self.company.uuid
        self.excel_import.pk = uuid.uuid4()

        # Create test Excel file
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Test Sheet"
        ws["A1"] = "km"
        ws["B1"] = "Direção"
        ws["C1"] = "Faixa"
        ws["D1"] = "Rodovia"
        ws["A2"] = 145
        ws["B2"] = "Norte"
        ws["C2"] = "Faixa 1"
        ws["D2"] = "BR101-SC"
        self.wb.create_sheet("_validation")
        self.wb["_validation"].sheet_state = Worksheet.SHEETSTATE_VERYHIDDEN

        self.excel_path = os.path.join(self.temp_dir, "test.xlsx")
        self.wb.save(self.excel_path)

        # Add S3 client mock
        self.s3_client = Mock()

        # Create ImportExcel instance
        self.importer = ImportExcel(self.excel_import, self.user, "number")
        self.importer.file_name = self.excel_path
        self.importer.preload_inventory_cache(["APS-IT-2019.07179"])

    def tearDown(self):
        """Cleanup after tests"""
        if os.path.exists(self.excel_path):
            os.remove(self.excel_path)
        if os.path.exists(self.temp_dir):
            for file in os.listdir(self.temp_dir):
                try:
                    os.remove(os.path.join(self.temp_dir, file))
                except Exception:
                    pass
            try:
                os.rmdir(self.temp_dir)
            except Exception:
                pass

    def test_get_lane_and_direction(self):
        """Test lane and direction translations"""
        self.importer.get_lane_and_direction()

        assert "faixa 1" in self.importer.lane_translation
        assert self.importer.lane_translation["faixa 1"] == "1"
        assert "norte" in self.importer.direction_translation
        assert self.importer.direction_translation["norte"] == "0"

    @patch("openpyxl.load_workbook")
    def test_load_data(self, mock_load_wb):
        """Test loading Excel data"""
        mock_load_wb.return_value = self.wb
        self.importer.file_name = self.excel_path
        self.importer.load_data()
        assert hasattr(self.importer, "wb")

    def test_parse_km(self):
        """Test km parsing"""
        assert self.importer.parse_km("123.45") == 123.45
        assert self.importer.parse_km("123,45") == 123.45
        assert self.importer.parse_km("123+45") == 123.45
        assert self.importer.parse_km("invalid") is None
        assert self.importer.parse_km("") is None

    def test_parse_coordinates(self):
        """Test coordinate parsing"""
        assert self.importer.parse_coordinates("45°30'30\"N") == 45.508333333333333
        assert self.importer.parse_coordinates("45°30'30\"S") == -45.508333333333333
        assert self.importer.parse_coordinates("123.45") == 123.45
        assert self.importer.parse_coordinates("invalid") == "erro"

    @patch("apps.roads.models.Road.objects.filter")
    def test_parse_road(self, mock_road_filter):
        """Test road name parsing"""
        mock_road = Mock()
        mock_road.name = "BR101-SC"
        mock_road_filter.return_value.first.return_value = mock_road

        result = self.importer.parse_road("BR101-SC")
        assert result == "BR101-SC"
        assert "BR101-SC" in self.importer.road_names

    @patch("openpyxl.load_workbook")
    def test_get_provided_numbers(self, mock_load_wb):
        """Test loading Excel data"""
        mock_load_wb.return_value = self.wb
        self.importer.load_data()
        ws = self.wb.create_sheet("Numbers")
        ws["A1"] = "serial do apontamento ou do item de inventario que sera editado"
        ws["A2"] = "TEST-001"
        ws["A3"] = "TEST-002"

        with patch.object(self.importer, "wb", self.wb):
            numbers = self.importer.get_provided_numbers()
            assert "TEST-001" in numbers
            assert "TEST-002" in numbers
            assert len(numbers) == 2
            assert self.importer.is_edit_import() is True

    @patch("apps.occurrence_records.models.OccurrenceType.objects.filter")
    def test_parse_occurrence_type(self, mock_occ_filter):
        """Test occurrence type parsing"""
        mock_occ = Mock()
        mock_occ.uuid = uuid.uuid4()
        mock_occ_filter.return_value.first.return_value = mock_occ

        row_dict = {"occurrence_type_id": "Test Type"}
        result = self.importer.parse_occurrence_type(row_dict)
        assert result == mock_occ
        assert "Test Type" in self.importer.occ_types

    def test_parse_form_data(self):
        """Test form data parsing"""
        occ_type = Mock()
        occ_type.form_fields = {
            "fields": [
                {
                    "api_name": "test_field",
                    "data_type": "text_area",
                    "display_name": "Test Field",
                }
            ]
        }

        row_dict = {"Test Field": "test value"}
        header = ["Test Field"]

        result = self.importer.parse_form_data(row_dict, occ_type, header)
        assert "form_data" in result
        assert result["form_data"]["test_field"] == "test value"

    def test_parse_menu(self):
        menu, is_visible = self.importer.parse_menu("b")

        assert menu is not None
        assert is_visible

    def test_is_hidden_sheet(self):
        is_hidden = self.importer.is_hidden_sheet(self.wb["Test Sheet"])

        assert not is_hidden

    @patch("openpyxl.load_workbook")
    def test_is_edit_import(self, mock_load_wb):
        mock_load_wb.return_value = self.wb
        self.importer.load_data()
        is_edit_import = self.importer.is_edit_import()

        assert not is_edit_import

    def test_update_column_errors(self):
        """Test updating column errors"""
        row_dict = {"column_errors": []}
        column_errors = ["firm"]
        expected_value = {"column_errors": ["firm"]}

        return_value = self.importer.update_column_errors(row_dict, column_errors)
        assert return_value == expected_value

    def test_format_row_location(self):
        """Test formatting row location"""
        row_name = "C"
        row_number = 5
        result = self.importer.format_row_location(row_name, row_number)
        assert result == "C:5"

    def test_parse_firm(self):
        """Test parsing firm data"""
        firm = Firm.objects.filter(company=self.company).first()

        row_dict = {"firm_id": firm.name}
        result = self.importer.parse_firm(row_dict)
        assert result == firm
        assert firm.name in self.importer.firms

    def test_translate_image_kind(self):
        """Test image kind translation"""
        test_cases = [
            ("Antes", "antes"),
            ("", ""),
        ]

        for input_kind, expected in test_cases:
            result = self.importer.translate_image_kind(input_kind)
            assert result == expected

    def test_is_image_column(self):
        """Test image column detection"""
        test_cases = [
            ("Foto_1", True),
            ("normal_column", False),
        ]

        for column_name, expected in test_cases:
            result = self.importer.is_image_column(column_name)
            assert result == expected

    def test_parse_resources(self):
        """Test parsing resources from Excel row data"""
        # Setup firm and resources
        firm = Mock()
        firm.uuid = uuid.uuid4()
        reporting_id = str(uuid.uuid4())

        # Mock resource and service order
        mock_resource = Mock()
        mock_resource.uuid = uuid.uuid4()
        mock_resource.name = "Test Resource"

        mock_so_resource = Mock()
        mock_so_resource.uuid = uuid.uuid4()
        mock_so_resource.unit_price = 100.0

        mock_resource.resource_service_orders = Mock()
        mock_resource.resource_service_orders.all.return_value = [mock_so_resource]

        # Setup resources query mock
        mock_resources_qs = Mock()
        mock_resources_qs.filter.return_value = [mock_resource]

        # Setup input row dict with resources
        row_dict = {"resources": {"A1": {"name": "test resource", "quantity": 5}}}

        # Add resources to cache
        self.importer.resources[firm] = mock_resources_qs

        # Execute
        result = self.importer.parse_resources(row_dict, firm, reporting_id)

        # Verify
        assert "resources" in result
        assert len(result["resources"]) == 1

        created_resource = result["resources"][0]
        assert created_resource["total_price"] == 500.0  # quantity * unit_price
        assert created_resource["unit_price"] == 100.0
        assert created_resource["amount"] == 5
        assert created_resource["reporting_id"] == reporting_id
        assert created_resource["resource_id"] == str(mock_resource.uuid)
        assert created_resource["service_order_resource_id"] == str(
            mock_so_resource.uuid
        )
        assert created_resource["created_by_id"] == self.importer.user_id
        assert created_resource["firm_id"] == str(firm.uuid)

        # Test invalid quantity
        row_dict = {
            "resources": {
                "A1": {"name": "test resource", "quantity": None}  # Invalid quantity
            }
        }
        result = self.importer.parse_resources(row_dict, firm, reporting_id)
        assert "Recurso_A1" in result.get("column_errors", [])

        # Test resource not found
        mock_resources_qs.filter.return_value = []
        row_dict = {
            "resources": {"A1": {"name": "nonexistent resource", "quantity": 5}}
        }
        result = self.importer.parse_resources(row_dict, firm, reporting_id)
        assert "Recurso_A1" in result.get("column_errors", [])

        # Test section filled but no resource name
        mock_resources_qs.filter.return_value = [mock_resource]
        row_dict = {
            "resources": {
                "A1": {"section": "Test Section", "name": None, "quantity": 5}
            }
        }
        result = self.importer.parse_resources(row_dict, firm, reporting_id)
        assert "Recurso_A1" in result.get("column_errors", [])

        # Test resource with incompatible section
        mock_resources_qs.filter.return_value = []
        row_dict = {
            "resources": {
                "A1": {
                    "section": "Incompatible Section",
                    "name": "test resource",
                    "quantity": 5,
                }
            }
        }
        result = self.importer.parse_resources(row_dict, firm, reporting_id)
        assert "Recurso_A1" in result.get("column_errors", [])

        # Test section does not exist in system
        mock_resources_qs.filter.return_value = []
        row_dict = {
            "resources": {
                "A1": {
                    "section": "Nonexistent Section",
                    "name": "test resource",
                    "quantity": 5,
                }
            }
        }
        result = self.importer.parse_resources(row_dict, firm, reporting_id)
        assert "Recurso_A1" in result.get("column_errors", [])

    def test_add_value(self):
        """Test adding values to dictionary"""
        test_cases = [
            # Test normal string
            (
                {"key": "value"},
                "new_key",
                "new_value",
                {"key": "value", "new_key": "new_value"},
            ),
            # Test None value
            ({"key": "value"}, "lot", None, {"key": "value"}),
        ]
        column_errors = []
        for initial_dict, key, value, expected in test_cases:
            _, column_errors = self.importer.add_value(
                key, value, initial_dict, column_errors
            )
            assert initial_dict == expected

    def test_parse_obj(self):
        """Test parsing of Excel row data into reporting object"""
        # Setup test data
        test_uuid = uuid.uuid4()
        mock_occ_type = Mock()
        mock_occ_type.uuid = test_uuid
        mock_occ_type.occurrence_kind = "1"
        mock_occ_type.form_fields = {"fields": []}

        mock_firm = Mock()
        mock_firm.uuid = test_uuid

        mock_road = Mock()
        mock_road.name = "Test Road"
        mock_road.lot_logic = {"==": [{"var": "data.km"}, 100]}
        self.road = mock_road

        # Create row dict with all possible fields
        row_dict_initial = {
            "km": "123.45",
            "end_km": "124.45",
            "project_km": "123.00",
            "project_end_km": "124.00",
            "km_reference": "124.00",
            "road_name": "Test Road",
            "direction": "Norte",
            "lane": "Faixa 1",
            "track": "Principal",
            "branch": "Branch 1",
            "menu_id": "Test Menu",
            "occurrence_type_id": "Test Type",
            "firm_id": "Test Firm",
            "status_id": "Active",
            "executed_at": datetime(2023, 1, 2),
            "due_at": "22/01/2025",
            "latitude": "27°35'42\"S",
            "longitude": "48°32'39\"W",
            "inventory_value": "APS-IT-2019.07179",
            "form_data": {"field1": "value1"},
            "point": {},
        }

        # Setup method mocks
        with patch.object(
            self.importer, "parse_occurrence_type"
        ) as mock_parse_occ, patch.object(
            self.importer, "parse_firm"
        ) as mock_parse_firm, patch.object(
            self.importer, "parse_status"
        ) as mock_parse_status, patch.object(
            self.importer, "parse_menu"
        ) as mock_parse_menu:

            # Setup returns
            mock_parse_occ.return_value = mock_occ_type
            mock_parse_firm.return_value = mock_firm
            mock_parse_status.return_value = str(test_uuid)
            mock_parse_menu.return_value = (str(test_uuid), True)

            # Setup translations
            self.importer.direction_translation = {"norte": "0"}
            self.importer.lane_translation = {"faixa 1": "1"}
            self.importer.company.metadata = {
                "show_track": True,
                "show_coordinate_input": True,
            }

            # Execute
            row_dict = row_dict_initial.copy()
            basic_columns = ["row", "column_errors", "formula_errors"]
            result = self.importer.parse_obj(row_dict, basic_columns, [])

            # Verify
            row_dict = row_dict_initial.copy()
            assert result is not None
            assert "uuid" in result

            # Test coordinate error handling
            row_dict = row_dict_initial.copy()
            row_dict["latitude"] = "invalid"
            result = self.importer.parse_obj(row_dict, basic_columns, [])
            assert "point" in result
            assert "error" in result["point"]

            # Test inventory linking
            row_dict = row_dict_initial.copy()
            mock_occ_type.occurrence_kind = "2"
            result = self.importer.parse_obj(row_dict, basic_columns, [])
            assert "import_inventory_code" in result

            # Test edit mode
            row_dict = row_dict_initial.copy()
            self.importer.is_edit = True
            row_dict["number"] = "APS-IT-2019.02792"
            self.importer.number_to_reporting = {
                "APS-IT-2019.02792": [(uuid.uuid4(), True)]
            }
            result = self.importer.parse_obj(row_dict, basic_columns, [])
            assert result["created_by_id"] is None
            assert result["created_at"] is None

    def test_get_data(self):
        """Test Excel data extraction and parsing"""
        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"

        # Add headers
        headers = [
            "Serial do apontamento ou do item de inventário que será editado",  # number
            "km",
            "km Final",
            "km de Projeto",
            "km Final de Projeto",
            "Status",
            "Equipe",
            "Encontrado em",
            "Executado em",
            "Prazo",
            "Sentido",
            "Classe",
            "Faixa",
            "Rodovia",
            "Pista",
            "Menu",
            "Latitude",
            "Longitude",
            "Código do Inventário para vinculo com apontamento",
            "Foto_1",
            "Descrição Foto_1",
            "Tipo Foto_1",
            "Data Foto_1",
            "Recurso_1",  # Resource name
            "Quantidade_1",  # Resource quantity
        ]
        ws.append(headers)

        # Add a complete row of data
        row_data = [
            "TEST-001",  # number
            "123.45",  # km
            "124.45",  # end_km
            "123.00",  # project_km
            "124.00",  # project_end_km
            "Active",  # status
            "Team 1",  # firm
            "01/01/2023",  # found_at
            "02/01/2023",  # executed_at
            "03/01/2023",  # due_at
            "Norte",  # direction
            "Type 1",  # occurrence_type
            "Faixa 1",  # lane
            "Road 1",  # road_name
            "Principal",  # track
            "Menu 1",  # menu
            "27°35'42\"S",  # latitude
            "48°32'39\"W",  # longitude
            "INV-001",  # inventory_value
            "image1.jpg",  # foto
            "Test photo",  # description
            "Foto",  # kind
            "04/01/2023",  # date
            "Resource 1",  # resource name
            5,  # resource quantity
        ]
        ws.append(row_data)

        # Save workbook
        excel_path = os.path.join(self.temp_dir, "test.xlsx")
        wb.save(excel_path)

        # Setup mocks
        mock_status = Mock()
        mock_status.uuid = uuid.uuid4()

        mock_firm = Mock()
        mock_firm.uuid = uuid.uuid4()
        mock_firm.name = "Team 1"

        mock_occ_type = Mock()
        mock_occ_type.uuid = uuid.uuid4()
        mock_occ_type.name = "Type 1"
        mock_occ_type.occurrence_kind = "1"
        mock_occ_type.form_fields = {"fields": []}

        mock_road = Mock()
        mock_road.name = "Road 1"
        mock_road.lot_logic = {"==": [{"var": "data.km"}, 123.45]}

        mock_menu = Mock()
        mock_menu.uuid = uuid.uuid4()

        # Setup resource mocks
        mock_resource = Mock()
        mock_resource.uuid = uuid.uuid4()
        mock_resource.name = "Resource 1"

        mock_so_resource = Mock()
        mock_so_resource.uuid = uuid.uuid4()
        mock_so_resource.unit_price = 100.0

        mock_resource.resource_service_orders = Mock()
        mock_resource.resource_service_orders.all.return_value = [mock_so_resource]

        # Setup resource queryset
        mock_resource_qs = Mock()
        mock_resource_qs.filter.return_value = [mock_resource]

        # Setup company metadata and options
        self.importer.company.metadata = {
            "show_track": True,
            "show_coordinate_input": True,
        }

        # Setup patches
        with patch(
            "apps.service_orders.models.ServiceOrderActionStatus.objects.filter"
        ) as mock_status_filter, patch(
            "apps.companies.models.Firm.objects.filter"
        ) as mock_firm_filter, patch(
            "apps.occurrence_records.models.OccurrenceType.objects.filter"
        ) as mock_occ_filter, patch(
            "apps.roads.models.Road.objects.filter"
        ) as mock_road_filter, patch(
            "apps.reportings.models.RecordMenuRelation.objects.filter"
        ) as mock_menu_filter, patch(
            "apps.resources.models.Resource.objects"
        ) as mock_resource_objects:

            # Configure mock returns
            mock_status_filter.return_value.first.return_value = mock_status
            mock_firm_filter.return_value.first.return_value = mock_firm
            mock_occ_filter.return_value.first.return_value = mock_occ_type
            mock_road_filter.return_value.first.return_value = mock_road
            mock_menu_filter.return_value.values_list.return_value = [
                ("menu 1", mock_menu.uuid, False)
            ]
            mock_resource_objects.filter.return_value = mock_resource_qs

            # Set workbook
            self.importer.file_name = excel_path

            # Load data
            self.importer.load_data()
            # Execute
            result = self.importer.get_data()

            # Verify result structure
            assert result is not None
            assert "reportings" in result
            assert "images" in result
            assert "is_array_edit" in result

            # Verify reporting data
            reporting = result["reportings"][0]
            assert reporting["km"] == 123.45
            assert reporting["direction"] == "0"
            assert reporting["lane"] == "1"

    def test_download_excel_file(self):
        """Test downloading Excel file from S3"""
        # Setup mock excel file with proper URL
        mock_file = Mock()
        mock_file.url = (
            "https://test-bucket.s3.amazonaws.com/media/private/test_file.xlsx"
        )

        # Setup excel import with mock file
        self.excel_import.excel_file = mock_file

        # Setup S3 client
        self.importer.s3 = self.s3_client

        # Create temp directory
        os.makedirs(self.importer.temp_path, exist_ok=True)

        # Test download
        result = self.importer.download_excel_file()

        # Verify S3 download was called with correct parameters
        self.s3_client.download_file.assert_called_once_with(
            "test-bucket",
            "media/private/test_file.xlsx",
            f"{self.importer.temp_path}test_file{self.importer.uuid}.xlsx",
        )

        # Verify result path format
        expected_path = f"{self.importer.temp_path}test_file{self.importer.uuid}.xlsx"
        assert result == expected_path

        # Cleanup
        try:
            os.remove(expected_path)
        except Exception:
            pass
