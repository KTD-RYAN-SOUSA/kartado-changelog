import json
import uuid
from datetime import date, datetime, time, timedelta
from unittest.mock import Mock, patch

import pytest
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone

from apps.reportings.models import Reporting
from helpers.signals import DisableSignals
from helpers.testing.fixtures import TestBase

# Mock sentry_sdk before importing daily_reports
with patch.dict("sys.modules", {"sentry_sdk": Mock()}):
    from helpers.apps.daily_reports import (
        generate_exported_file,
        translate_weather,
        translate_condition,
        translate_kind,
        format_km,
        calculate_total_price,
        calculate_board_item_total_price,
        has_permission,
        get_reporting_static_columns,
        get_fields_to_hide_reporting_location,
        get_exporter_extra_columns,
        datetime_to_date,
        parse_time_string,
        normalize_time_fields,
        parse_extra_hours_to_list,
        get_updated_by,
        translate_reporting_value,
        fill_detailed_control_board,
        filter_board_item_contract_services,
        determine_relation_field_name,
        determine_report_type_and_field,
        get_km_intervals_field,
        get_values_from_reporting_extra_columns,
        remove_fields_to_hide_reporting_location,
        create_array_columns,
        get_array_image_columns,
        get_exporter_extra_columns_parsed_infos,
        get_reporting_data,
    )

from apps.companies.models import Company, Firm, UserInCompany
from apps.daily_reports.models import (
    DailyReport,
    DailyReportEquipment,
    DailyReportExport,
    DailyReportExternalTeam,
    DailyReportOccurrence,
    DailyReportResource,
    DailyReportSignaling,
    DailyReportVehicle,
    DailyReportWorker,
    MultipleDailyReport,
    ProductionGoal,
)
from apps.permissions.models import UserPermission
from apps.resources.models import Contract

User = get_user_model()

pytestmark = pytest.mark.django_db


class TestGenerateExportedFile(TestCase):
    """Tests for the main generate_exported_file function"""

    def setUp(self):
        """Set up test fixtures"""
        # Generate valid UUIDs for the test
        self.valid_contract_uuid = str(uuid.uuid4())
        self.different_contract_uuid = str(uuid.uuid4())

        # Create a test company
        self.company = Company.objects.create(
            name="Test Company",
            custom_options={
                "multipledailyreport": {
                    "extrahourscontractids": [
                        self.valid_contract_uuid,
                        self.different_contract_uuid,
                    ]
                }
            },
        )

        # Create a test user
        self.user = User.objects.create_user(
            username="testuser", email="test@example.com", password="testpass"
        )

        # Create user permissions with required permissions for extra hours
        self.user_permissions = UserPermission.objects.create(
            name="Test Permissions",
            permissions={"multipledailyreport": {"cansetextrahours": True}},
        )
        self.user_permissions.companies.add(self.company)

        # Create membership relation with permissions
        UserInCompany.objects.create(
            user=self.user,
            company=self.company,
            level=1,  # Admin level
            is_active=True,
            permissions=self.user_permissions,
        )

    @pytest.mark.django_db
    @patch("helpers.apps.daily_reports.logging.error")
    @patch("helpers.apps.daily_reports.capture_exception")
    def test_generate_exported_file_not_found(
        self, mock_capture_exception, mock_logging_error
    ):
        """Test generate_exported_file with non-existent export ID"""
        non_existent_id = uuid.uuid4()

        generate_exported_file(str(non_existent_id))

        # Verify error handling
        mock_logging_error.assert_called_once_with("DailyReportExport not found")
        mock_capture_exception.assert_called_once()

    @pytest.mark.django_db
    @patch("helpers.apps.daily_reports.OccurrenceType.objects.filter")
    @patch("helpers.apps.daily_reports.get_obj_from_path")
    @patch("helpers.apps.daily_reports.get_exporter_extra_columns")
    @patch("apps.daily_reports.signals.auto_add_daily_report_number")  # Disable signal
    def test_generate_exported_file_success_daily_reports(
        self, mock_signal, mock_get_exporter, mock_get_obj, mock_occ_filter
    ):
        """Test successful generation with daily reports"""
        # Setup mocks
        mock_get_obj.side_effect = lambda obj, path, default_return=None: {
            "hide_reporting_location": False,
            "can_view_digital_signature": True,
        }.get(path.split(".")[-1], default_return)

        mock_get_exporter.return_value = []
        mock_occ_filter.return_value.values_list.return_value = []

        # Create a firm for MultipleDailyReport
        from apps.companies.models import Firm

        firm = Firm.objects.create(name="Test Firm", company=self.company)

        # Create test data using MultipleDailyReport instead of DailyReport
        from apps.daily_reports.models import MultipleDailyReport

        multiple_daily_report = MultipleDailyReport.objects.create(
            date=timezone.now().date(),
            company=self.company,
            firm=firm,
            number="001",  # Add number manually since signal is disabled
        )

        daily_report_export = DailyReportExport.objects.create(
            format="EXCEL"  # Use format instead of export_type
        )
        daily_report_export.multiple_daily_reports.add(multiple_daily_report)

        # Mock reporting_files.all() to return empty list for simplicity
        with patch.object(
            multiple_daily_report.reporting_files, "all", return_value=[]
        ):
            # Mock the complex export process
            with patch("openpyxl.Workbook") as mock_workbook:
                mock_wb = Mock()
                mock_workbook.return_value = mock_wb
                mock_ws = Mock()
                mock_wb.active = mock_ws

                # Call the function under test
                generate_exported_file(str(daily_report_export.uuid))

                # Refresh DailyReportExport to get updated values
                daily_report_export.refresh_from_db()

                # Assertions - check that the export was processed successfully
                # The function doesn't return a value, but updates the export object
                self.assertTrue(daily_report_export.done)

    @pytest.mark.django_db
    def test_generate_exported_file_no_reports(self):
        """Test generate_exported_file with no reports attached"""
        # Create export instance without reports
        daily_report_export = DailyReportExport.objects.create(
            created_by=self.user,
            done=False,
            error=False,
            is_compiled=False,
            format="XLSX",
        )

        generate_exported_file(str(daily_report_export.uuid))

        # Should return early without processing
        daily_report_export.refresh_from_db()
        assert not daily_report_export.done


class TestTranslationFunctions(TestCase):
    """Tests for translation utility functions"""

    def test_translate_weather_valid_values(self):
        """Test translate_weather with valid weather codes"""
        self.assertEqual(translate_weather("SUNNY"), "Aberto")
        self.assertEqual(translate_weather("CLOUDY"), "Nublado")
        self.assertEqual(translate_weather("RAINY"), "Chuvoso")
        self.assertEqual(translate_weather("NOT_APPLIED"), "Não se aplica")

    def test_translate_weather_invalid_value(self):
        """Test translate_weather with invalid weather code"""
        self.assertEqual(translate_weather("INVALID"), "Não se aplica")
        self.assertIsNone(translate_weather(None))

    def test_translate_condition_valid_values(self):
        """Test translate_condition with valid condition codes"""
        self.assertEqual(translate_condition("FEASIBLE"), "Praticável")
        self.assertEqual(translate_condition("UNFEASIBLE"), "Impraticável")
        self.assertEqual(translate_condition("NOT_APPLIED"), "Não se aplica")

    def test_translate_condition_invalid_value(self):
        """Test translate_condition with invalid condition code"""
        self.assertEqual(translate_condition("INVALID"), "Não se aplica")
        self.assertIsNone(translate_condition(None))

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_translate_kind_success(self, mock_get_obj):
        """Test translate_kind with valid options"""
        mock_company = Mock()
        mock_get_obj.return_value = [
            {"value": "KIND1", "name": "Tipo 1"},
            {"value": "KIND2", "name": "Tipo 2"},
        ]

        result = translate_kind("KIND1", mock_company)
        self.assertEqual(result, "Tipo 1")

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_translate_kind_not_found(self, mock_get_obj):
        """Test translate_kind when kind is not found"""
        mock_company = Mock()
        mock_get_obj.return_value = [
            {"value": "KIND1", "name": "Tipo 1"},
        ]

        result = translate_kind("KIND2", mock_company)
        self.assertEqual(result, "")

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_translate_kind_exception(self, mock_get_obj):
        """Test translate_kind when an exception occurs"""
        mock_company = Mock()
        mock_get_obj.return_value = None

        result = translate_kind("KIND1", mock_company)
        self.assertEqual(result, "")


class TestCalculationFunctions(TestCase):
    """Tests for calculation utility functions"""

    def test_format_km_default_padding(self):
        """Test format_km with default padding"""
        self.assertEqual(format_km(5.5), "005+500")
        self.assertEqual(format_km(123.456), "123+456")
        self.assertEqual(format_km(0), "000+000")

    def test_format_km_custom_padding(self):
        """Test format_km with custom padding"""
        self.assertEqual(format_km(5.5, left_padding=5), "00005+500")
        self.assertEqual(format_km(123.456, left_padding=2), "123+456")

    def test_format_km_none_value(self):
        """Test format_km with None value - should raise TypeError"""
        with self.assertRaises(TypeError):
            format_km(None)

    def test_calculate_total_price_valid_inputs(self):
        """Test calculate_total_price with valid inputs"""
        # Formula is (amount * unit_price) / work_day
        result = calculate_total_price(10, 5.0, 22)
        self.assertEqual(result, 50.0 / 22)  # (10 * 5.0) / 22

        result = calculate_total_price(8, 12.5, 20)
        self.assertEqual(result, 100.0 / 20)  # (8 * 12.5) / 20

    def test_calculate_total_price_zero_work_day(self):
        """Test calculate_total_price with zero work_day (should handle exception)"""
        result = calculate_total_price(44, 5.0, 0)
        self.assertEqual(result, 0.0)  # Exception handling returns 0.0

    def test_calculate_board_item_total_price(self):
        """Test calculate_board_item_total_price function"""
        mock_instance = Mock()
        mock_instance.amount = 10
        mock_instance.unit_price = 5.0
        mock_measurement_bulletin = Mock()
        mock_measurement_bulletin.work_day = 22
        mock_instance.measurement_bulletin = mock_measurement_bulletin

        with patch(
            "helpers.apps.daily_reports.calculate_total_price", return_value=50.0
        ) as mock_calc:
            result = calculate_board_item_total_price(mock_instance)

            mock_calc.assert_called_once_with(10, 5.0, 22)
            self.assertEqual(result, 50.0)


class TestUtilityFunctions(TestCase):
    """Tests for utility functions"""

    def test_has_permission_valid_permission(self):
        """Test has_permission with valid permission"""
        model_permissions = {"read": [True], "write": [False]}

        self.assertTrue(has_permission(model_permissions, "read"))
        self.assertFalse(has_permission(model_permissions, "write"))

    def test_has_permission_missing_permission(self):
        """Test has_permission with missing permission"""
        model_permissions = {"read": [True]}

        self.assertFalse(has_permission(model_permissions, "delete"))

    def test_has_permission_invalid_structure(self):
        """Test has_permission with invalid permission structure"""
        model_permissions = {"read": True}  # Should be a list

        self.assertFalse(has_permission(model_permissions, "read"))

    def test_get_reporting_static_columns(self):
        """Test get_reporting_static_columns returns expected columns"""
        columns = get_reporting_static_columns()

        self.assertIsInstance(columns, dict)
        self.assertIn("number", columns)
        self.assertIn("road", columns)
        self.assertIn("km", columns)
        self.assertEqual(columns["number"], "Serial Apontamento")
        self.assertEqual(columns["road"], "Rodovia")

    def test_get_fields_to_hide_reporting_location(self):
        """Test get_fields_to_hide_reporting_location returns expected fields"""
        fields = get_fields_to_hide_reporting_location()

        expected_fields = [
            "road",
            "km",
            "end_km",
            "lot",
            "lane",
            "direction",
            "track",
            "branch",
            "km_reference",
        ]

        self.assertEqual(fields, expected_fields)

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_get_exporter_extra_columns_reporting(self, mock_get_obj):
        """Test get_exporter_extra_columns for reporting"""
        mock_company = Mock()
        mock_get_obj.return_value = [{"header": "Extra Column"}]

        result = get_exporter_extra_columns(mock_company, is_inventory=False)

        mock_get_obj.assert_called_once_with(
            mock_company.custom_options, "reporting__exporter__extra_columns"
        )
        self.assertEqual(result, [{"header": "Extra Column"}])

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_get_exporter_extra_columns_inventory(self, mock_get_obj):
        """Test get_exporter_extra_columns for inventory"""
        mock_company = Mock()
        mock_get_obj.return_value = [{"header": "Inventory Column"}]

        result = get_exporter_extra_columns(mock_company, is_inventory=True)

        mock_get_obj.assert_called_once_with(
            mock_company.custom_options, "inventory__exporter__extra_columns"
        )
        self.assertEqual(result, [{"header": "Inventory Column"}])


class TestDateTimeFunctions(TestCase):
    """Tests for date and time utility functions"""

    def test_datetime_to_date_with_timezone(self):
        """Test datetime_to_date with timezone info"""
        import pytz

        dt = datetime(2023, 1, 15, 12, 30, 0, tzinfo=pytz.UTC)

        result = datetime_to_date(dt, clear_tzinfo=True)

        self.assertEqual(result, date(2023, 1, 15))

    def test_datetime_to_date_without_timezone(self):
        """Test datetime_to_date without timezone info"""
        dt = datetime(2023, 1, 15, 12, 30, 0)

        result = datetime_to_date(dt, clear_tzinfo=False)

        self.assertEqual(result, date(2023, 1, 15))

    def test_datetime_to_date_none_value(self):
        """Test datetime_to_date with None"""
        result = datetime_to_date(None)
        self.assertIsNone(result)

    def test_parse_time_string_valid_format(self):
        """Test parse_time_string with valid time string"""
        result = parse_time_string("14:30")
        self.assertEqual(result, time(14, 30))

    def test_parse_time_string_seconds_format(self):
        """Test parse_time_string with seconds format (should return None)"""
        result = parse_time_string("14:30:45")
        self.assertIsNone(result)

    def test_parse_time_string_invalid_format(self):
        """Test parse_time_string with invalid format"""
        result = parse_time_string("invalid")
        self.assertIsNone(result)

    def test_parse_time_string_none_value(self):
        """Test parse_time_string with None"""
        result = parse_time_string(None)
        self.assertIsNone(result)

    def test_normalize_time_fields_complete_hours(self):
        """Test normalize_time_fields with complete hour entries"""
        hours_dict = {
            "morning_start": "08:00",
            "morning_end": "12:00",
            "afternoon_start": "13:00",
            "afternoon_end": "17:00",
        }
        default_hours = {
            "morning_start": "07:00",
            "morning_end": "11:00",
            "afternoon_start": "12:00",
            "afternoon_end": "16:00",
        }

        result = normalize_time_fields(hours_dict, default_hours)

        expected = {
            "morning_start": time(8, 0),
            "morning_end": time(12, 0),
            "afternoon_start": time(13, 0),
            "afternoon_end": time(17, 0),
        }
        self.assertEqual(result, expected)

    def test_normalize_time_fields_missing_hours(self):
        """Test normalize_time_fields with missing hour entries"""
        hours_dict = {
            "morning_start": "08:00",
            "afternoon_end": "17:00",
            "morning_start_is_deleted": True,
        }
        default_hours = {
            "morning_start": "07:00",
            "morning_end": "11:00",
            "afternoon_start": "12:00",
            "afternoon_end": "16:00",
            "night_start": "21:00",
            "night_end": "3:00",
        }

        result = normalize_time_fields(hours_dict, default_hours)

        expected = {
            # "morning_start": time(8, 0),  # Deleted key is True
            "morning_end": time(11, 0),  # From default
            "afternoon_start": time(12, 0),  # From default
            "afternoon_end": time(17, 0),
            "night_start": time(21, 0),
            "night_end": time(3, 0),
        }

        self.assertEqual(result, expected)

    def test_parse_extra_hours_to_list_valid_json_string(self):
        """Test parse_extra_hours_to_list with JSON string"""
        extra_hours_dict = {"morning_start": "08:00", "afternoon_end": "17:00"}
        extra_hours = json.dumps(extra_hours_dict)
        default_hours = {
            "morning_start": "07:00",
            "morning_end": "11:00",
            "afternoon_start": "12:00",
            "afternoon_end": "16:00",
        }

        # The function expects a dict/list format, not JSON string
        # It returns [] if not dict or list
        result = parse_extra_hours_to_list(extra_hours, default_hours)

        self.assertEqual(result, [])

    def test_parse_extra_hours_to_list_dict_format(self):
        """Test parse_extra_hours_to_list with dict format"""
        extra_hours = {"extraHours": [{"morning_start": "08:00"}]}
        default_hours = {
            "morning_start": "07:00",
            "morning_end": "11:00",
            "afternoon_start": "12:00",
            "afternoon_end": "16:00",
        }

        result = parse_extra_hours_to_list(extra_hours, default_hours)

        expected = [
            {
                "morning_start": time(8, 0),
                "morning_end": time(11, 0),
                "afternoon_start": time(12, 0),
                "afternoon_end": time(16, 0),
            }
        ]
        self.assertEqual(result, expected)

    def test_parse_extra_hours_to_list_none_value(self):
        """Test parse_extra_hours_to_list with None"""
        result = parse_extra_hours_to_list(None)

        self.assertEqual(result, [])


class TestReportingFunctions(TestCase):
    """Tests for reporting utility functions"""

    def test_get_updated_by_with_user(self):
        """Test get_updated_by with user object"""
        mock_instance = Mock()
        mock_user = Mock()
        mock_user.get_full_name.return_value = "John Doe"

        mock_history = Mock()
        mock_history.history_user = mock_user
        mock_instance.historicalreporting.all.return_value = [mock_history]

        result = get_updated_by(mock_instance)
        self.assertEqual(result, "John Doe")

    def test_get_updated_by_without_user(self):
        """Test get_updated_by without user object"""
        mock_instance = Mock()
        mock_instance.historicalreporting.all.return_value = []

        result = get_updated_by(mock_instance)
        self.assertIsNone(result)

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_translate_reporting_value_with_options(self, mock_get_obj):
        """Test translate_reporting_value with select options"""
        mock_company = Mock()
        mock_get_obj.return_value = [
            {"value": "VALUE1", "name": "Option 1"},
            {"value": "VALUE2", "name": "Option 2"},
        ]

        result = translate_reporting_value(mock_company, "test_field", "VALUE1")

        self.assertEqual(result, "Option 1")

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_translate_reporting_value_without_options(self, mock_get_obj):
        """Test translate_reporting_value without select options"""
        mock_company = Mock()
        mock_get_obj.return_value = None

        result = translate_reporting_value(mock_company, "test_field", "VALUE1")

        self.assertEqual(result, "")

    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_translate_reporting_value_value_not_found(self, mock_get_obj):
        """Test translate_reporting_value when value is not found in options"""
        mock_company = Mock()
        mock_get_obj.return_value = [
            {"value": "VALUE1", "name": "Option 1"},
        ]

        result = translate_reporting_value(mock_company, "test_field", "VALUE2")

        self.assertEqual(result, "")


class TestFillDetailedControlBoard(TestCase):
    """Tests for the fill_detailed_control_board function"""

    def setUp(self):
        """Set up test data"""
        # Create a test company with proper custom options
        self.valid_contract_uuid = "1cede63e-8dd7-45b0-a11a-c45e89c87874"  # Use existing contract from fixtures
        self.different_contract_uuid = (
            "339fc8c2-3351-4509-af8a-aa7c519d89ee"  # Another existing contract
        )

        self.company = Company.objects.create(
            name="Test Company",
            custom_options={
                "multipledailyreport": {
                    "extrahourscontractids": [
                        self.valid_contract_uuid,
                        self.different_contract_uuid,
                    ]
                }
            },
            metadata={
                "MDR_name_format": {
                    "default": {
                        "type": "RDO",
                        "format": "{tipo}-{nome}-{ano_completo}.{serial_ano:05}",
                    }
                }
            },
        )

        # Create a test user with proper permissions
        self.user = User.objects.create_user(
            username="testuser", email="test@example.com", password="testpass"
        )

        # Create user permissions with required permissions for extra hours
        self.user_permissions = UserPermission.objects.create(
            name="Test Permissions",
            permissions={"multipledailyreport": {"cansetextrahours": True}},
        )
        self.user_permissions.companies.add(self.company)

        # Create membership relation with permissions
        UserInCompany.objects.create(
            user=self.user,
            company=self.company,
            level=1,  # Admin level
            is_active=True,
            permissions=self.user_permissions,
        )

        # Create a test firm
        self.firm = Firm.objects.create(
            name="Test Firm", company=self.company, is_company_team=True
        )

        # Create MultipleDailyReport instances for testing
        self.report1 = MultipleDailyReport.objects.create(
            company=self.company,
            firm=self.firm,
            date=date.today(),
            created_by=self.user,
            notes="Test Report 1",
        )

        # Use different dates to avoid unique constraint violation
        self.report2 = MultipleDailyReport.objects.create(
            company=self.company,
            firm=self.firm,
            date=date.today() - timedelta(days=1),  # Yesterday
            created_by=self.user,
            notes="Test Report 2",
        )

        # Create one that should NOT be included (different contract_id)
        self.report_excluded = MultipleDailyReport.objects.create(
            company=self.company,
            firm=self.firm,
            date=date.today() - timedelta(days=2),  # Day before yesterday
            created_by=self.user,
            notes="Excluded Report",
        )

    def test_fill_detailed_control_board_with_equipment_and_vehicles(self):
        """Test with equipment and vehicles data"""
        # Get the existing contract from fixtures
        contract = Contract.objects.get(uuid=self.valid_contract_uuid)

        # Create export instance
        daily_report_export = DailyReportExport.objects.create(
            created_by=self.user,
            done=False,
            error=False,
            is_compiled=False,
            format="XLSX",
        )

        # Create multiple daily report with matching contract
        multiple_daily_report = MultipleDailyReport.objects.create(
            company=self.company,
            firm=self.firm,
            date=date.today(),
            contract=contract,
            number="002",
        )
        daily_report_export.multiple_daily_reports.add(multiple_daily_report)

        # Mock workbook
        from openpyxl import Workbook

        mock_wb = Workbook()

        # Sample reports data with equipment and vehicles
        reports_data = {
            "report2": {
                "number": "RDO002",
                "contract": self.valid_contract_uuid,
                "firm": "Test Firm",
                "date": date(2023, 1, 16),
                "notes": "Test with equipment",
                "morning_start": time(7, 0),
                "morning_end": time(11, 0),
                "afternoon_start": time(12, 0),
                "afternoon_end": time(16, 0),
                "night_start": None,
                "night_end": None,
                "workers": [],
                "equipment": [
                    (
                        "Excavator",
                        1,
                        None,
                        "Heavy Equipment",
                        "EX001",
                        None,
                        None,
                    ),  # description, amount, _, resource_name, sort_string, extra_hours, contract_id
                ],
                "vehicles": [
                    (
                        "Truck",
                        1,
                        None,
                        "Transport Vehicle",
                        "TR001",
                        None,
                        None,
                    ),  # description, amount, _, resource_name, sort_string, extra_hours, contract_id
                ],
            }
        }

        # Execute function
        fill_detailed_control_board(
            mock_wb, reports_data, self.company, daily_report_export
        )

        # Verify worksheet was created
        self.assertIn("Quadros de Controle Detalhado", mock_wb.sheetnames)

    def test_fill_detailed_control_board_no_permission(self):
        """Test function returns early when user has no permission"""
        # Create user with no extra hours permission
        user_without_permissions = User.objects.create_user(
            username="nopermuser", email="noperm@example.com", password="testpass"
        )

        # Create permissions without the required permission
        permissions_no_extra_hours = UserPermission.objects.create(
            name="No Extra Hours Permissions",
            permissions={
                "multipledailyreport": {"cansetextrahours": False}  # No permission
            },
        )
        permissions_no_extra_hours.companies.add(self.company)

        UserInCompany.objects.create(
            user=user_without_permissions,
            company=self.company,
            level=2,
            is_active=True,
            permissions=permissions_no_extra_hours,
        )

        # Create export instance with user without permissions
        daily_report_export = DailyReportExport.objects.create(
            created_by=user_without_permissions,
            done=False,
            error=False,
            is_compiled=False,
            format="XLSX",
        )

        # Mock workbook
        from openpyxl import Workbook

        mock_wb = Workbook()

        # Execute function
        fill_detailed_control_board(mock_wb, {}, self.company, daily_report_export)

        # Verify worksheet was NOT created
        self.assertNotIn("Quadros de Controle Detalhado", mock_wb.sheetnames)

    def test_fill_detailed_control_board_contract_not_allowed(self):
        """Test function returns early when contract is not in allowed list"""
        # Get the existing contract from fixtures
        contract = Contract.objects.get(uuid=self.valid_contract_uuid)

        # Create export instance
        daily_report_export = DailyReportExport.objects.create(
            created_by=self.user,
            done=False,
            error=False,
            is_compiled=False,
            format="XLSX",
        )

        # Create multiple daily report with contract
        multiple_daily_report = MultipleDailyReport.objects.create(
            company=self.company,
            firm=self.firm,
            date=date.today(),
            contract=contract,
            number="003",
        )
        daily_report_export.multiple_daily_reports.add(multiple_daily_report)

        # Update company custom options to NOT include this contract
        self.company.custom_options = {
            "multipledailyreport": {
                "extrahourscontractids": [
                    self.different_contract_uuid
                ]  # Different contract
            }
        }
        self.company.save()

        # Mock workbook
        from openpyxl import Workbook

        mock_wb = Workbook()

        # Execute function
        fill_detailed_control_board(mock_wb, {}, self.company, daily_report_export)

        # Verify worksheet was NOT created
        self.assertNotIn("Quadros de Controle Detalhado", mock_wb.sheetnames)

    def test_fill_detailed_control_board_exception_handling(self):
        """Test function handles exceptions gracefully"""
        # Create user without company membership to trigger exception
        user_no_membership = User.objects.create_user(
            username="nomembership", email="nomem@example.com", password="testpass"
        )

        # Create export instance with user that has no company membership
        daily_report_export = DailyReportExport.objects.create(
            created_by=user_no_membership,
            done=False,
            error=False,
            is_compiled=False,
            format="XLSX",
        )

        # Mock workbook
        from openpyxl import Workbook

        mock_wb = Workbook()

        # Execute function - should handle exception gracefully
        fill_detailed_control_board(mock_wb, {}, self.company, daily_report_export)

        # Verify worksheet was NOT created due to exception
        self.assertNotIn("Quadros de Controle Detalhado", mock_wb.sheetnames)


class TestFilterBoardItemContractServices(TestCase):
    """Tests for filter_board_item_contract_services function"""

    def test_filter_board_item_contract_services(self):
        """Test filtering board items by contract service IDs"""
        # Create a mock queryset
        mock_queryset = Mock()
        contract_service_ids = [uuid.uuid4(), uuid.uuid4()]

        # Call the function
        result = filter_board_item_contract_services(
            mock_queryset, contract_service_ids
        )

        # Verify the filter was called with correct parameters
        mock_queryset.filter.assert_called_once_with(
            contract_item_administration__contract_item_administration_services__uuid__in=contract_service_ids
        )
        assert result == mock_queryset.filter.return_value


class TestRelationFunctions(TestCase):
    """Tests for relation field determination functions"""

    def test_determine_relation_field_name_worker(self):
        """Test determine_relation_field_name for DailyReportWorker"""
        result = determine_relation_field_name(DailyReportWorker)
        self.assertEqual(result, "worker")

    def test_determine_relation_field_name_external_team(self):
        """Test determine_relation_field_name for DailyReportExternalTeam"""
        result = determine_relation_field_name(DailyReportExternalTeam)
        self.assertEqual(result, "external_team")

    def test_determine_relation_field_name_equipment(self):
        """Test determine_relation_field_name for DailyReportEquipment"""
        result = determine_relation_field_name(DailyReportEquipment)
        self.assertEqual(result, "equipment")

    def test_determine_relation_field_name_vehicle(self):
        """Test determine_relation_field_name for DailyReportVehicle"""
        result = determine_relation_field_name(DailyReportVehicle)
        self.assertEqual(result, "vehicle")

    def test_determine_relation_field_name_signaling(self):
        """Test determine_relation_field_name for DailyReportSignaling"""
        result = determine_relation_field_name(DailyReportSignaling)
        self.assertEqual(result, "signaling")

    def test_determine_relation_field_name_production_goal(self):
        """Test determine_relation_field_name for ProductionGoal"""
        result = determine_relation_field_name(ProductionGoal)
        self.assertEqual(result, "production_goal")

    def test_determine_relation_field_name_occurrence(self):
        """Test determine_relation_field_name for DailyReportOccurrence"""
        result = determine_relation_field_name(DailyReportOccurrence)
        self.assertEqual(result, "occurrence")

    def test_determine_relation_field_name_resource(self):
        """Test determine_relation_field_name for DailyReportResource"""
        result = determine_relation_field_name(DailyReportResource)
        self.assertEqual(result, "resource")

    def test_determine_report_type_and_field_daily_report(self):
        """Test determine_report_type_and_field for DailyReport"""
        mock_daily_report = Mock(spec=DailyReport)

        report_field, report_type = determine_report_type_and_field(mock_daily_report)

        self.assertEqual(report_field, "daily_report")
        self.assertEqual(report_type, "DailyReport")

    def test_determine_report_type_and_field_multiple_daily_report(self):
        """Test determine_report_type_and_field for MultipleDailyReport"""
        mock_mdr = Mock(spec=MultipleDailyReport)

        report_field, report_type = determine_report_type_and_field(mock_mdr)

        self.assertEqual(report_field, "multiple_daily_report")
        self.assertEqual(report_type, "MultipleDailyReport")

    def test_determine_report_type_and_field_invalid_report(self):
        """Test determine_report_type_and_field with invalid report type"""
        invalid_report = Mock()

        with self.assertRaises(Exception):  # Should raise ValidationError
            determine_report_type_and_field(invalid_report)


class TestKmIntervalsFunctions(TestCase, TestBase):
    """Tests for KM intervals related functions"""

    model = "Reporting"

    def test_get_km_intervals_field_no_reportings(self):
        """Test get_km_intervals_field with no reportings"""
        mock_instance = Mock()
        mock_instance.reportings.count.return_value = 0

        result = get_km_intervals_field(mock_instance)

        self.assertEqual(result, [])

    def test_get_km_intervals_field_with_reportings(self):
        """Test get_km_intervals_field with reportings"""

        # Create instances in test database without overriding the end_km value
        with DisableSignals():
            reporting1 = Reporting.objects.create(
                road_name="BR-101",
                km=10.5,
                end_km=15.2,
                company=self.company,
                end_km_manually_specified=True,
            )

            reporting2 = Reporting.objects.create(
                road_name="BR-101",
                km=8.0,
                end_km=20.0,
                company=self.company,
                end_km_manually_specified=True,
            )

            reporting3 = Reporting.objects.create(
                road_name="BR-116",
                km=50.0,
                end_km=None,
                company=self.company,
                end_km_manually_specified=True,
            )

        queryset = Reporting.objects.filter(
            pk__in=[reporting1.pk, reporting2.pk, reporting3.pk]
        )

        mock_instance = Mock()
        mock_instance.reportings.count.return_value = 3
        mock_instance.reportings.all.return_value = queryset

        result = get_km_intervals_field(mock_instance)

        # Should have 2 roads with merged intervals
        self.assertEqual(len(result), 2)

        # Find BR-101 entry
        br101_entry = next(entry for entry in result if entry["roadName"] == "BR-101")
        self.assertEqual(br101_entry["km"], 8.0)  # Minimum km
        self.assertEqual(br101_entry["end_km"], 20.0)  # Maximum end_km

        # Find BR-116 entry
        br116_entry = next(entry for entry in result if entry["roadName"] == "BR-116")
        self.assertEqual(br116_entry["km"], 50.0)
        self.assertIsNone(br116_entry["end_km"])


class TestReportingDataFunctions(TestCase):
    """Tests for reporting data processing functions"""

    @patch("helpers.apps.daily_reports.get_obj_serialized")
    @patch("helpers.apps.daily_reports.apply_json_logic")
    @patch("helpers.apps.daily_reports.return_select_value")
    @patch("helpers.apps.daily_reports.return_array_values")
    @patch("helpers.apps.daily_reports.to_snake_case")
    def test_get_values_from_reporting_extra_columns_with_logic(
        self,
        mock_to_snake,
        mock_return_array,
        mock_return_select,
        mock_apply_logic,
        mock_get_serialized,
    ):
        """Test get_values_from_reporting_extra_columns with JSON logic"""
        mock_record = Mock()
        mock_record.form_data = {"test_field": "test_value"}

        extra_columns = [
            {
                "key": "test_column",
                "logic": {"==": [{"var": "field1"}, "value1"]},
                "isDate": False,
                "isSelect": False,
                "isArray": False,
            }
        ]

        mock_get_serialized.return_value = {"field1": "value1"}
        mock_apply_logic.return_value = "logic_result"

        result = get_values_from_reporting_extra_columns(mock_record, extra_columns, {})

        self.assertEqual(result["test_column"], "logic_result")
        mock_apply_logic.assert_called_once()

    def test_get_values_from_reporting_extra_columns_empty(self):
        """Test get_values_from_reporting_extra_columns with no extra columns"""
        mock_record = Mock()

        result = get_values_from_reporting_extra_columns(mock_record, [], {})

        self.assertEqual(result, {})

    def test_remove_fields_to_hide_reporting_location(self):
        """Test remove_fields_to_hide_reporting_location"""
        fields_to_hide = ["road", "km", "end_km"]
        static_columns = {
            "number": "Serial Apontamento",
            "road": "Rodovia",
            "km": "KM inicial",
            "end_km": "KM final",
            "notes": "Observações",
        }

        result = remove_fields_to_hide_reporting_location(
            fields_to_hide, static_columns
        )

        expected = {"number": "Serial Apontamento", "notes": "Observações"}
        self.assertEqual(result, expected)

    def test_create_array_columns_basic(self):
        """Test create_array_columns with basic fields"""
        extra_column = {
            "header": "Test Header",
            "key": "test_key",
            "maxRepetitions": 2,
            "fields": [
                {"header": "Field 1", "field": "field1", "isImage": False},
                {"header": "Field 2", "field": "field2", "isImage": False},
            ],
        }

        result = create_array_columns(extra_column, False, {})

        expected = {
            "test_key0field1": "Test Header 1: Field 1",
            "test_key0field2": "Test Header 1: Field 2",
            "test_key1field1": "Test Header 2: Field 1",
            "test_key1field2": "Test Header 2: Field 2",
        }
        self.assertEqual(result, expected)

    def test_get_array_image_columns(self):
        """Test get_array_image_columns"""
        result = get_array_image_columns(
            "photo",
            "Photo",
            "image",
            "Image",
            0,
            2,
            {"export_kind": True, "export_date": True, "export_description": False},
        )

        expected_keys = [
            "photo#0#image0#content",
            "photo#0#image0#kind",
            "photo#0#image0#datetime",
            "photo#0#image1#content",
            "photo#0#image1#kind",
            "photo#0#image1#datetime",
        ]

        for key in expected_keys:
            self.assertIn(key, result)

    def test_get_exporter_extra_columns_parsed_infos_basic(self):
        """Test get_exporter_extra_columns_parsed_infos with basic columns"""
        exporter_extra_columns = [
            {"key": "column1", "header": "Column 1", "isArray": False},
            {"key": "column2", "header": "Column 2", "isArray": False},
        ]

        result = get_exporter_extra_columns_parsed_infos(exporter_extra_columns)

        expected = {"column1": "Column 1", "column2": "Column 2"}
        self.assertEqual(result, expected)

    def test_get_exporter_extra_columns_parsed_infos_with_arrays(self):
        """Test get_exporter_extra_columns_parsed_infos with array columns"""
        exporter_extra_columns = [
            {
                "key": "array_col",
                "header": "Array Column",
                "isArray": True,
                "maxRepetitions": 1,
                "fields": [{"header": "Field", "field": "field", "isImage": False}],
            }
        ]

        result = get_exporter_extra_columns_parsed_infos(exporter_extra_columns)

        self.assertIn("array_col0field", result)
        self.assertEqual(result["array_col0field"], "Array Column 1: Field")

    @patch("helpers.apps.daily_reports.get_values_from_reporting_extra_columns")
    @patch("helpers.apps.daily_reports.get_updated_by")
    @patch("helpers.apps.daily_reports.datetime_to_date")
    @patch("helpers.apps.daily_reports.translate_reporting_value")
    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_get_reporting_data(
        self,
        mock_get_obj,
        mock_translate,
        mock_datetime_to_date,
        mock_get_updated,
        mock_get_values,
    ):
        """Test get_reporting_data function"""
        # Setup mocks
        mock_reporting = Mock()
        mock_reporting.number = "12345"
        mock_reporting.road.name = "BR-101"
        mock_reporting.km = 10.5
        mock_reporting.end_km = 15.2
        mock_reporting.lot = "LOT001"
        mock_reporting.point.coords = [-45.123, -23.456]
        mock_reporting.occurrence_type.name = "Buraco"
        mock_reporting.occurrence_type.occurrence_kind = "POTHOLE"
        mock_reporting.form_data = {
            "length": 2.0,
            "width": 1.5,
            "height": 0.3,
            "notes": "Test note",
        }
        mock_reporting.lane = "LEFT"
        mock_reporting.direction = "NORTH"
        mock_reporting.status.name = "Aberto"
        mock_reporting.created_by.get_full_name.return_value = "John Doe"
        mock_reporting.firm.name = "Test Firm"
        mock_reporting.job.title = "Test Job"
        mock_reporting.job.start_date = datetime.now()
        mock_reporting.job.end_date = datetime.now()
        mock_reporting.created_at = datetime.now()
        mock_reporting.found_at = datetime.now()
        mock_reporting.updated_at = datetime.now()
        mock_reporting.executed_at = datetime.now()
        mock_reporting.due_at = datetime.now()
        mock_reporting.company = Mock()

        # Setup mock returns
        mock_translate.side_effect = lambda company, field, value: f"translated_{value}"
        mock_datetime_to_date.return_value = date.today()
        mock_get_updated.return_value = "Updated By User"
        mock_get_values.return_value = {"extra1": "value1"}
        mock_get_obj.return_value = [{"value": "POTHOLE", "name": "Buraco"}]

        result = get_reporting_data(mock_reporting, [], {})

        # Verify basic fields
        self.assertEqual(result["number"], "12345")
        self.assertEqual(result["road"], "BR-101")
        self.assertEqual(result["km"], 10.5)
        self.assertEqual(result["end_km"], 15.2)
        self.assertEqual(result["occurrence_type"], "Buraco")
        self.assertEqual(result["occurrence_kind"], "Buraco")
        self.assertEqual(result["extra_columns"], {"extra1": "value1"})

    @patch("helpers.apps.daily_reports.get_values_from_reporting_extra_columns")
    @patch("helpers.apps.daily_reports.get_updated_by")
    @patch("helpers.apps.daily_reports.datetime_to_date")
    @patch("helpers.apps.daily_reports.translate_reporting_value")
    @patch("helpers.apps.daily_reports.get_obj_from_path")
    def test_get_reporting_data_with_none_values(
        self,
        mock_get_obj,
        mock_translate,
        mock_datetime_to_date,
        mock_get_updated,
        mock_get_values,
    ):
        """Test get_reporting_data with None values"""
        # Setup mocks with None values
        mock_reporting = Mock()
        mock_reporting.number = "12345"
        mock_reporting.road = None
        mock_reporting.km = 10.5
        mock_reporting.end_km = None
        mock_reporting.lot = None
        mock_reporting.point = None
        mock_reporting.occurrence_type = None
        mock_reporting.form_data = {}
        mock_reporting.lane = None
        mock_reporting.direction = None
        mock_reporting.status = None
        mock_reporting.created_by = None
        mock_reporting.firm = None
        mock_reporting.job = None
        mock_reporting.created_at = datetime.now()
        mock_reporting.found_at = datetime.now()
        mock_reporting.updated_at = datetime.now()
        mock_reporting.executed_at = None
        mock_reporting.due_at = None
        mock_reporting.company = Mock()

        # Setup mock returns
        mock_translate.return_value = ""
        mock_datetime_to_date.return_value = date.today()
        mock_get_updated.return_value = None
        mock_get_values.return_value = {}
        mock_get_obj.side_effect = BaseException("Not found")

        result = get_reporting_data(mock_reporting, [], {})

        # Verify None values are handled correctly
        self.assertEqual(result["number"], "12345")
        self.assertIsNone(result["road"])
        self.assertIsNone(result["end_km"])
        self.assertIsNone(result["lot"])
        self.assertIsNone(result["longitude"])
        self.assertIsNone(result["latitude"])
        self.assertIsNone(result["occurrence_type"])
        self.assertIsNone(result["occurrence_kind"])
        self.assertIsNone(result["status"])
        self.assertIsNone(result["created_by"])
        self.assertIsNone(result["firm"])
        self.assertIsNone(result["job"])
        self.assertIsNone(result["executed_at"])
        self.assertIsNone(result["due_at"])
