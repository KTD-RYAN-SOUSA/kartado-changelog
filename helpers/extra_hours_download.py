import json
import os
import random
import shutil
import string
from collections import defaultdict
from datetime import timedelta

import arrow
import boto3
import pytz
from django.conf import settings
from django.db.models import F, Prefetch
from django.utils.timezone import now
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_GENERAL
from zappa.asynchronous import task

from apps.companies.models import Company, Firm
from apps.daily_reports.models import (
    DailyReportContractUsage,
    DailyReportEquipment,
    DailyReportVehicle,
    DailyReportWorker,
)
from apps.files.models import GenericFile
from apps.resources.models import Contract, ContractItemAdministration, ContractPeriod
from helpers.apps.daily_reports import is_holiday_for_firm
from helpers.dates import format_minutes, format_minutes_decimal
from helpers.extra_hours import calculate_extra_hours_worker, parse_time_to_minutes
from helpers.strings import DAYS_PORTUGUESE, MAPS_MONTHS_ENG_TO_PT_SHORT
from RoadLabsAPI.settings import credentials


class ExtraHoursExport:
    def __init__(
        self,
        file_uuid="",
        filename=None,
        object_name=None,
        contract_uuid="",
        company=None,
        creation_date_after=None,
        creation_date_before=None,
    ):

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )
        if file_uuid:
            generic_file = GenericFile.objects.get(uuid=file_uuid)
            uuid_list = json.loads(generic_file.file.read())

            daily_prefetch_fields_list = [
                self.get_contract_usage_prefetch_object("worker", DailyReportWorker),
                self.get_contract_usage_prefetch_object(
                    "equipment", DailyReportEquipment
                ),
                self.get_contract_usage_prefetch_object("vehicle", DailyReportVehicle),
                "multiple_daily_reports",
                "multiple_daily_reports__firm",
                "contract_item_administration",
            ]

            self.daily_queryset = (
                DailyReportContractUsage.objects.filter(pk__in=set(uuid_list))
                .annotate(
                    resource_unit_price=F(
                        "contract_item_administration__resource__unit_price"
                    ),
                )
                .prefetch_related(*daily_prefetch_fields_list)
            )

        if filename:
            self.filename = filename
        else:
            self.filename = self.get_filename()

        if object_name:
            self.object_name = object_name
        else:
            self.object_name = self.get_object_name()

        if contract_uuid:
            self.contract = Contract.objects.get(uuid=contract_uuid)
            self.company = Company.objects.get(uuid=company)

            self.creation_date_after = "/".join(creation_date_after.split("-")[::-1])
            self.creation_date_before = "/".join(creation_date_before.split("-")[::-1])

            date_after = arrow.get(self.creation_date_after, "DD/MM/YYYY")
            date_before = arrow.get(self.creation_date_before, "DD/MM/YYYY")

            self.date_range = [
                d.date() for d in arrow.Arrow.range("day", date_after, date_before)
            ]

            contract_periods = ContractPeriod.objects.filter(
                contract_id=contract_uuid
            ).order_by("created_at")
            self.contract_periods = {
                firm.uuid: period
                for period in contract_periods
                for firm in period.firms.all()
            }

            contract_item_set = defaultdict(set)
            services = self.contract.administration_services.all().prefetch_related(
                Prefetch("firms", queryset=Firm.objects.all()),
                Prefetch(
                    "contract_item_administration",
                    queryset=ContractItemAdministration.objects.all().prefetch_related(
                        "content_type",
                        "contract_item_administration_services",
                        "resource",
                        "resource__resource",
                    ),
                ),
            )
            for service in services:
                firms = service.firms.all()
                items = service.contract_item_administration.all()
                for firm in firms:
                    items_to_add = [
                        (item, item.content_type.model == "dailyreportworker")
                        for item in items
                    ]
                    contract_item_set[firm].update(items_to_add)

            self.contract_item_list_ordered = {
                firm: sorted(items, key=lambda x: x[1], reverse=True)
                for firm, items in contract_item_set.items()
            }

            self.cell_font_bold = Font(name="Arial", size=10, color="000000", bold=True)
            self.cell_font_italic = Font(
                name="Arial", size=9, color="000000", italic=True
            )
            self.cell_font_bold_and_italic = Font(
                name="Arial", size=10, color="000000", bold=True, italic=True
            )
            self.rdo_color = Font(name="Arial", size=10, color="888888")
            self.red_font = Font(name="Arial", size=9, color="FF0000")
            self.red_and_bold_font = Font(
                name="Arial", size=10, color="FF0000", bold=True
            )
            self.green_font = Font(name="Arial", size=9, color="00B050")
            self.blue_font = Font(name="Arial", size=9, color="0070C0")
            self.red_fill = PatternFill("solid", fgColor="FF0000")
            self.yellow_fill = PatternFill("solid", fgColor="FFFF00")
            self.horizontal_center = Alignment(horizontal="center")
            self.wrap_alignment = Alignment(vertical="top", wrap_text=True)
            self.qtd_comment = Comment(
                "Quantidade cadastrada no objeto no momento da exportação.",
                "",
                width=225,
                height=50,
            )

            self.unit_price_comment = Comment(
                "Valor unitário do item no objeto, no momento da exportação.",
                "",
                width=225,
                height=50,
            )

            self.sum_columns = [
                "FALTAS",
                "H.E. 50%",
                "H.E. 100%",
                "H.E. 50% Noturna",
                "H.E. 100% Noturna",
                "H.E.EQUIP",
                "COMP.",
            ]

            self.firm_information_columns = [
                "Dias",
                "Entrada",
                "Saída",
                "Turno",
                "Total decimal (h)",
            ]

            self.period_translation = {
                "morning": "Manhã",
                "afternoon": "Tarde",
                "night": "Noite",
            }

            self.number_to_day = {
                1: "seg",
                2: "ter",
                3: "qua",
                4: "qui",
                5: "sex",
                6: "sáb",
                7: "dom",
            }

            self.item_columns = ["ITEM", "QTD.", "SEÇÃO"]

            self.summary_columns = [
                "Valor Unitário (R$)",
                "Valor previsto (R$)",
                "R$ FALTA (20%)",
                "R$ H.E.50% Diurna",
                "R$ H.E.100% Diurna",
                "R$ H.E.50% Noturna",
                "R$ H.E.100% Noturna",
                "R$ H.E.EQUIP",
                "TOTAL",
            ]

            self.money_number_format = "R$ #,##0.00"

    def get_random_string(self):
        return "".join(
            random.SystemRandom().choice(string.ascii_lowercase + string.digits)
            for _ in range(10)
        )

    def get_filename(self):
        filename = "[Kartado] Memória de cálculo"

        return f"{filename}.xlsx"

    def get_object_name(self):
        random_string = self.get_random_string()
        object_name = "media/private/{}_{}.xlsx".format(
            self.filename.split(".xlsx")[0], random_string
        )

        return object_name

    def get_s3_url(self):
        empty = {"url": "", "name": ""}

        url = self.s3.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                "Key": self.object_name,
            },
        )

        if not url:
            return empty

        return url

    def upload_file(self, path):
        expires = now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)

        try:
            self.s3.upload_file(
                path,
                settings.AWS_STORAGE_BUCKET_NAME,
                self.object_name,
                ExtraArgs={"Expires": expires},
            )
        except Exception:
            return

    def copy_and_rename(self, old_folder, new_folder, temp_file):
        os.makedirs(new_folder, exist_ok=True)
        shutil.copy(old_folder + temp_file, new_folder + temp_file)
        os.rename(
            new_folder + temp_file,
            new_folder + self.filename + ".xlsx",
        )
        return

    def return_file_path(self, new_folder):
        return new_folder + self.filename + ".xlsx"

    def load_file(self, new_file_path):
        wb = load_workbook(filename=new_file_path)
        return wb

    def get_contract_usage_prefetch_object(self, relation_name, model):

        return Prefetch(
            relation_name,
            queryset=model.objects.all(),
        )

    def get_board_instance(self, obj):

        if obj.worker:
            return obj.worker
        elif obj.equipment:
            return obj.equipment
        elif obj.vehicle:
            return obj.vehicle
        return None

    def set_cell(
        self,
        relatorio,
        column,
        row,
        value,
        font=None,
        alignment=None,
        number_format=None,
        comment=None,
        fill=None,
    ):
        cell = relatorio.cell(column=column, row=row)
        cell.value = value
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format
        if comment is not None:
            cell.comment = comment
        if fill is not None:
            cell.fill = fill
        return cell

    def organize_queryset(self):
        mdr_dict = defaultdict(lambda: defaultdict(dict))
        item_dict = dict()
        hours_sum_dict = defaultdict(dict)

        for item in self.daily_queryset:
            mdr = item.multiple_daily_reports.all()[0]
            firm = mdr.firm
            is_holiday = is_holiday_for_firm(self.company, firm.uuid, mdr.date)
            mdr_data = {
                "number": mdr.number,
                "compensation": mdr.compensation,
                "date": mdr.date,
                "is_holiday": is_holiday,
            }
            mdr_dict[firm][mdr.date] = mdr_data

            obj = self.get_board_instance(item)
            is_worker = bool(item.worker)
            extra_hours_list = obj.extra_hours
            working_schedules = getattr(
                self.contract_periods.get(firm.uuid), "working_schedules", []
            )
            total_results = {"total": 0, "compensation": 0}
            contract_item = item.contract_item_administration
            if extra_hours_list == {}:
                extra_hours_list = [{}]
            if extra_hours_list and working_schedules:
                default_hours = {
                    "morning_start": mdr.morning_start,
                    "morning_end": mdr.morning_end,
                    "afternoon_start": mdr.afternoon_start,
                    "afternoon_end": mdr.afternoon_end,
                    "night_start": mdr.night_start,
                    "night_end": mdr.night_end,
                }

                for extra_hour in extra_hours_list:
                    worker_result = calculate_extra_hours_worker(
                        worked_periods_item=extra_hour,
                        working_schedules=working_schedules,
                        day_of_week=mdr.date.isoweekday(),
                        is_holiday=is_holiday,
                        is_compensation=mdr.compensation,
                        default_hours=default_hours,
                    )

                    if is_worker:
                        calculated = worker_result
                    else:
                        total_extra = sum(
                            parse_time_to_minutes(worker_result[k]) or 0
                            for k in (
                                "extra_hours_50_day",
                                "extra_hours_50_night",
                                "extra_hours_100_day",
                                "extra_hours_100_night",
                            )
                        )
                        calculated = {
                            "extra_hours": format_minutes(total_extra),
                            "absence": worker_result["absence"],
                            "compensation": worker_result["compensation"],
                        }
                    for key, value in calculated.items():
                        if key in [
                            "extra_hours_50_day",
                            "extra_hours_50_night",
                            "extra_hours_100_day",
                            "extra_hours_100_night",
                            "extra_hours",
                        ]:
                            total_results["total"] += format_minutes_decimal(
                                parse_time_to_minutes(value)
                            )
                        elif key == "absence":
                            total_results["total"] -= format_minutes_decimal(
                                parse_time_to_minutes(value)
                            )
                        else:
                            total_results[key] += format_minutes_decimal(
                                parse_time_to_minutes(value)
                            )

                        # Add item grouped by contract_item to hours_sum_dict
                        hours_sum_dict[(contract_item, firm)][key] = hours_sum_dict[
                            (contract_item, firm)
                        ].get(key, 0) + format_minutes_decimal(
                            parse_time_to_minutes(value)
                        )

            item_dict[(firm, contract_item, mdr.date)] = total_results

        mdr_dict_ordered = {
            firm: {
                date_key: date_value
                for date_key, date_value in sorted(inner_dict.items())
            }
            for firm, inner_dict in mdr_dict.items()
        }

        return mdr_dict_ordered, item_dict, hours_sum_dict

    def fill_firm_information(self, relatorio, firm, start_column, start_row):

        self.set_cell(
            relatorio,
            start_column,
            start_row,
            "Horas Cadastradas da Equipe",
            font=self.cell_font_bold,
        )

        start_row += 2

        for index, item in enumerate(self.firm_information_columns):
            self.set_cell(
                relatorio,
                start_column + index,
                start_row,
                item,
                font=self.cell_font_bold,
            )

        period_data = self.contract_periods.get(firm.uuid)
        working_schedules = getattr(period_data, "working_schedules", [])

        for index, schedule in enumerate(working_schedules):

            days_of_week = schedule.get("days_of_week", [])
            if days_of_week == [1, 2, 3, 4, 5]:
                days_value = "seg-sex"
            elif days_of_week == [6, 7]:
                days_value = "sáb-dom"
            else:
                days_value = "-".join(
                    list(map(lambda x: self.number_to_day.get(int(x)), days_of_week))
                )

            row = start_row + index + 1
            self.set_cell(relatorio, start_column, row, days_value)
            self.set_cell(
                relatorio, start_column + 1, row, schedule.get("start_time", "")
            )
            self.set_cell(
                relatorio, start_column + 2, row, schedule.get("end_time", "")
            )
            self.set_cell(
                relatorio,
                start_column + 3,
                row,
                self.period_translation.get(schedule.get("period", ""), ""),
            )

            start_time = schedule.get("start_time", "")
            end_time = schedule.get("end_time", "")
            if start_time and end_time:
                end_minutes = parse_time_to_minutes(end_time)
                start_minutes = parse_time_to_minutes(start_time)
                if end_minutes <= start_minutes:
                    end_minutes += 24 * 60
                total_time = format_minutes_decimal(end_minutes - start_minutes)
                self.set_cell(relatorio, start_column + 4, row, total_time)

        start_row += len(working_schedules) + 2

        self.set_cell(
            relatorio,
            start_column,
            start_row,
            "Horas por mês:",
            font=self.cell_font_bold,
        )
        self.set_cell(
            relatorio, start_column, start_row + 1, f"{period_data.hours} horas"
        )

    def fill_matrix_1(
        self,
        relatorio,
        start_line,
        firm,
        mdr_data,
        item_dict,
        hours_sum_dict,
        contract_list_ordered,
    ):

        self.set_cell(
            relatorio, 1, start_line, f"EQUIPE: {firm.name}", font=self.cell_font_bold
        )
        self.set_cell(
            relatorio, 3, start_line, "Compensação", font=self.cell_font_italic
        )
        self.set_cell(
            relatorio, 3, start_line + 1, "Feriados", font=self.cell_font_italic
        )
        self.set_cell(
            relatorio, 3, start_line + 2, "Dia da semana", font=self.cell_font_italic
        )
        self.set_cell(
            relatorio,
            3,
            start_line + 3,
            "Relatório Diário de Obra",
            font=self.cell_font_italic,
        )
        self.set_cell(
            relatorio,
            1,
            start_line + 3,
            "Compilado de horas",
            font=self.cell_font_bold_and_italic,
        )

        for index, title_value in enumerate(self.item_columns):
            comment = self.qtd_comment if title_value == "QTD." else None
            self.set_cell(
                relatorio,
                index + 1,
                start_line + 4,
                title_value,
                font=self.cell_font_bold,
                alignment=self.horizontal_center,
                comment=comment,
            )

        # Pre-calculation of metadata to avoid repeated loops
        date_metadata = {}
        for date in self.date_range:
            date_metadata[date] = {
                "mdr_data_for_date": mdr_data.get(date, {}),
                "is_holiday": is_holiday_for_firm(self.company, firm.pk, date),
                "is_weekend": date.isoweekday() in [6, 7],
            }

        for index, date in enumerate(self.date_range):
            metadata = date_metadata[date]
            mdr_data_for_date = metadata["mdr_data_for_date"]
            is_holiday = metadata["is_holiday"]
            is_weekend = metadata["is_weekend"]
            is_compensation = mdr_data_for_date.get("compensation", False)

            self.set_cell(
                relatorio,
                4 + index,
                start_line + 2,
                DAYS_PORTUGUESE.get(date.strftime("%A"), "").lower(),
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                4 + index,
                start_line + 3,
                mdr_data_for_date.get("number", ""),
                font=self.rdo_color,
            )

            month_value = MAPS_MONTHS_ENG_TO_PT_SHORT[date.strftime("%B")].lower()
            fill = (
                (self.red_fill if is_holiday else self.yellow_fill)
                if (is_holiday or is_weekend)
                else None
            )
            self.set_cell(
                relatorio,
                4 + index,
                start_line + 4,
                date.strftime(f"%d/{month_value}"),
                font=self.cell_font_bold,
                alignment=self.horizontal_center,
                fill=fill,
            )

            if is_holiday:
                self.set_cell(
                    relatorio,
                    4 + index,
                    start_line + 1,
                    "SIM",
                    alignment=self.horizontal_center,
                )

            if is_compensation:
                self.set_cell(
                    relatorio,
                    4 + index,
                    start_line,
                    "SIM",
                    alignment=self.horizontal_center,
                )

        max_column = len(self.date_range) + 4

        for index, column_value in enumerate(self.sum_columns):
            self.set_cell(
                relatorio,
                max_column + index,
                start_line + 4,
                column_value,
                font=self.cell_font_bold,
                alignment=self.horizontal_center,
            )

        for index, (contract_item, is_worker) in enumerate(contract_list_ordered):

            current_line = start_line + 5 + index

            self.set_cell(
                relatorio, 1, current_line, contract_item.resource.resource.name
            )
            # Amount cell, fixed to 1
            self.set_cell(relatorio, 2, current_line, 1)
            self.set_cell(
                relatorio,
                3,
                current_line,
                contract_item.contract_item_administration_services.all()[
                    0
                ].description,
            )

            for inner_index, date in enumerate(self.date_range):
                metadata = date_metadata[date]
                mdr_data_for_date = metadata["mdr_data_for_date"]
                item_data = item_dict.get((firm, contract_item, date))
                mdr = mdr_data.get(date, {})

                if item_data is None:
                    self.set_cell(
                        relatorio,
                        4 + inner_index,
                        current_line,
                        "-" if not mdr else "",
                        alignment=self.horizontal_center,
                    )
                    continue

                is_compensation = mdr_data_for_date.get("compensation", False)
                hours_value = (
                    item_data.get("compensation")
                    if is_compensation
                    else item_data.get("total")
                )

                font = (
                    self.green_font
                    if is_compensation
                    else self.red_font
                    if hours_value < 0
                    else self.blue_font
                )

                if hours_value == 0:
                    hours_value = ""
                alignment = self.horizontal_center
                self.set_cell(
                    relatorio,
                    4 + inner_index,
                    current_line,
                    hours_value,
                    font=font,
                    alignment=alignment,
                )

            total_hours = hours_sum_dict.get((contract_item, firm), {})

            self.set_cell(
                relatorio,
                max_column,
                current_line,
                total_hours.get("absence", 0) * -1,
                font=self.red_font,
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                max_column + 1,
                current_line,
                total_hours.get("extra_hours_50_day", 0) if is_worker else "N/A",
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                max_column + 2,
                current_line,
                total_hours.get("extra_hours_100_day", 0) if is_worker else "N/A",
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                max_column + 3,
                current_line,
                total_hours.get("extra_hours_50_night", 0) if is_worker else "N/A",
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                max_column + 4,
                current_line,
                total_hours.get("extra_hours_100_night", 0) if is_worker else "N/A",
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                max_column + 5,
                current_line,
                total_hours.get("extra_hours", 0) if not is_worker else "N/A",
                alignment=self.horizontal_center,
            )
            self.set_cell(
                relatorio,
                max_column + 6,
                current_line,
                total_hours.get("compensation", 0),
                alignment=self.horizontal_center,
            )

        self.fill_firm_information(relatorio, firm, max_column + 9, start_line + 4)

    def fill_matrix_2(
        self,
        relatorio,
        start_line,
        firm,
        hours_sum_dict,
        contract_list_ordered,
    ):

        self.set_cell(
            relatorio,
            1,
            start_line,
            "Resumo Financeiro",
            font=self.cell_font_bold_and_italic,
        )
        current_line = start_line + 1

        for index, title_value in enumerate(self.item_columns + self.summary_columns):
            comment = (
                self.qtd_comment
                if title_value == "QTD."
                else self.unit_price_comment
                if title_value == "Valor Unitário (R$)"
                else None
            )
            self.set_cell(
                relatorio,
                index + 1,
                current_line,
                title_value,
                font=self.cell_font_bold,
                alignment=self.horizontal_center,
                comment=comment,
            )

        current_line += 1

        period_data_hours = self.contract_periods.get(firm.uuid).hours

        total_provisioned_price = 0
        total_absence = 0
        total_hours_50_day = 0
        total_hours_100_day = 0
        total_hours_50_night = 0
        total_hours_100_night = 0
        total_hours_equip = 0
        all_items_total = 0

        for index, (contract_item, is_worker) in enumerate(contract_list_ordered):

            amount = 1
            unit_price = contract_item.resource.unit_price
            provisioned_price = amount * unit_price
            total_provisioned_price += provisioned_price

            self.set_cell(
                relatorio, 1, current_line, contract_item.resource.resource.name
            )
            self.set_cell(relatorio, 2, current_line, amount)
            self.set_cell(
                relatorio,
                3,
                current_line,
                contract_item.contract_item_administration_services.all()[
                    0
                ].description,
            )
            self.set_cell(
                relatorio,
                4,
                current_line,
                contract_item.resource.unit_price,
                number_format=self.money_number_format,
            )
            self.set_cell(
                relatorio,
                5,
                current_line,
                provisioned_price,
                number_format=self.money_number_format,
            )

            total_hours = hours_sum_dict.get((contract_item, firm), {})

            # Pre-calculate values and sum total values

            absence_value = (
                provisioned_price * total_hours.get("absence", 0) * -1.2
            ) / period_data_hours
            total_absence += absence_value

            hours_50_day_value = (
                (provisioned_price * total_hours.get("extra_hours_50_day", 0) * 1.5)
                / period_data_hours
                if is_worker
                else "N/A"
            )
            total_hours_50_day += hours_50_day_value if is_worker else 0

            hours_100_day_value = (
                (provisioned_price * total_hours.get("extra_hours_100_day", 0) * 2)
                / period_data_hours
                if is_worker
                else "N/A"
            )
            total_hours_100_day += hours_100_day_value if is_worker else 0

            hours_50_night_value = (
                (
                    provisioned_price
                    * total_hours.get("extra_hours_50_night", 0)
                    * 1.5
                    * 1.2
                )
                / period_data_hours
                if is_worker
                else "N/A"
            )
            total_hours_50_night += hours_50_night_value if is_worker else 0

            hours_100_night_value = (
                (
                    provisioned_price
                    * total_hours.get("extra_hours_100_night", 0)
                    * 2
                    * 1.2
                )
                / period_data_hours
                if is_worker
                else "N/A"
            )
            total_hours_100_night += hours_100_night_value if is_worker else 0

            hours_equip_value = (
                (provisioned_price * total_hours.get("extra_hours", 0))
                / period_data_hours
                if not is_worker
                else "N/A"
            )
            total_hours_equip += hours_equip_value if not is_worker else 0

            contract_item_total_value = sum(
                list(
                    filter(
                        lambda value: isinstance(value, (int, float)),
                        [
                            provisioned_price,
                            absence_value,
                            hours_50_day_value,
                            hours_100_day_value,
                            hours_50_night_value,
                            hours_100_night_value,
                            hours_equip_value,
                        ],
                    )
                )
            )
            all_items_total += contract_item_total_value

            start_column = 6

            self.set_cell(
                relatorio,
                start_column,
                current_line,
                absence_value,
                font=self.red_font,
                number_format=self.money_number_format,
            )
            self.set_cell(
                relatorio,
                start_column + 1,
                current_line,
                hours_50_day_value,
                alignment=self.horizontal_center if not is_worker else None,
                number_format=self.money_number_format if is_worker else FORMAT_GENERAL,
            )
            self.set_cell(
                relatorio,
                start_column + 2,
                current_line,
                hours_100_day_value,
                alignment=self.horizontal_center if not is_worker else None,
                number_format=self.money_number_format if is_worker else FORMAT_GENERAL,
            )
            self.set_cell(
                relatorio,
                start_column + 3,
                current_line,
                hours_50_night_value,
                alignment=self.horizontal_center if not is_worker else None,
                number_format=self.money_number_format if is_worker else FORMAT_GENERAL,
            )
            self.set_cell(
                relatorio,
                start_column + 4,
                current_line,
                hours_100_night_value,
                alignment=self.horizontal_center if not is_worker else None,
                number_format=self.money_number_format if is_worker else FORMAT_GENERAL,
            )
            self.set_cell(
                relatorio,
                start_column + 5,
                current_line,
                hours_equip_value,
                alignment=self.horizontal_center if is_worker else None,
                number_format=self.money_number_format
                if not is_worker
                else FORMAT_GENERAL,
            )
            self.set_cell(
                relatorio,
                start_column + 6,
                current_line,
                contract_item_total_value,
                font=self.cell_font_bold,
                number_format=self.money_number_format,
            )
            current_line += 1

        self.set_cell(relatorio, 3, current_line, "TOTAL", font=self.cell_font_bold)
        self.set_cell(
            relatorio,
            5,
            current_line,
            total_provisioned_price,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            6,
            current_line,
            total_absence,
            font=self.red_and_bold_font,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            7,
            current_line,
            total_hours_50_day,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            8,
            current_line,
            total_hours_100_day,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            9,
            current_line,
            total_hours_50_night,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            10,
            current_line,
            total_hours_100_night,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            11,
            current_line,
            total_hours_equip,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )
        self.set_cell(
            relatorio,
            12,
            current_line,
            all_items_total,
            font=self.cell_font_bold,
            number_format=self.money_number_format,
        )

    def fill_workbook(self):

        mdr_dict, item_dict, hours_sum_dict = self.organize_queryset()

        relatorio = self.wb["Medição por Horas"]

        start_line = 2

        self.set_cell(
            relatorio,
            1,
            start_line,
            f"CALENDÁRIO RETROATIVO: {self.creation_date_after} a {self.creation_date_before}",
            font=self.cell_font_bold,
        )
        start_line += 2

        for firm, mdr_data in mdr_dict.items():

            contract_items = self.contract_item_list_ordered.get(firm)

            self.fill_matrix_1(
                relatorio,
                start_line,
                firm,
                mdr_data,
                item_dict,
                hours_sum_dict,
                contract_items,
            )
            start_line = relatorio.max_row + 2
            self.fill_matrix_2(
                relatorio,
                start_line,
                firm,
                hours_sum_dict,
                contract_items,
            )
            start_line = relatorio.max_row + 3

    def generate_file(self):
        old_folder = "apps/resources/templates/"
        new_folder = "/tmp/extra_hours_download/"
        temp_file = "previa_horas_extras.xlsx"

        self.copy_and_rename(old_folder, new_folder, temp_file)
        new_file_path = self.return_file_path(new_folder)
        self.wb = self.load_file(new_file_path)
        self.fill_workbook()

        self.wb.save(new_file_path)
        self.upload_file(new_file_path)
        os.remove(new_file_path)
        os.rmdir(new_folder)


@task
def extra_hours_export_async(
    file_uuid,
    filename,
    object_name,
    contract_uuid,
    company,
    creation_date_after,
    creation_date_before,
):
    extra_hours_export = ExtraHoursExport(
        file_uuid=file_uuid,
        filename=filename,
        object_name=object_name,
        contract_uuid=contract_uuid,
        company=company,
        creation_date_after=creation_date_after,
        creation_date_before=creation_date_before,
    )
    extra_hours_export.generate_file()
