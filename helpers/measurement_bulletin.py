import logging
import os
import shutil
from collections import defaultdict

import sentry_sdk
from django.core.files.base import ContentFile
from django.db.models import Sum, prefetch_related_objects
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from sentry_sdk import capture_exception
from zappa.asynchronous import task

from apps.daily_reports.models import (
    DailyReportEquipment,
    DailyReportVehicle,
    DailyReportWorker,
)
from apps.resources.models import MeasurementBulletinExport
from apps.service_orders.models import MeasurementBulletin, ProcedureResource
from helpers.apps.contract_utils import get_provisioned_price
from helpers.apps.performance_calculations import (
    ContractItemPerformanceScope,
    MeasurementBulletinScope,
)
from helpers.dates import utc_to_local


@task
def generate_bulletin(bulletin_export_id):
    try:
        bulletin_export = MeasurementBulletinExport.objects.get(pk=bulletin_export_id)

    except Exception as e:
        logging.error("MeasurementBulletin not found")
        capture_exception(e)

    else:
        bulletin = bulletin_export.measurement_bulletin
        number = bulletin.number

        error = True
        old_folder = "apps/resources/templates/"
        new_folder = "/tmp/bulletin/"
        temp_file = "temp_bulletin.xlsx"

        CONTRACT_ROW = 2
        BULLETIN_ROW = 6
        SURVEY_HEADER_ROW = 9
        PERFORMANCE_HEADER_ROW = 3

        field_survey_style = NamedStyle(name="field_survey_style")
        field_survey_style.font = Font(name="Calibri", size=10, color="000000")
        field_survey_style.alignment = Alignment(horizontal="left")
        field_survey_style.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        field_survey_style.protection = Protection(hidden=True)

        unit_style = NamedStyle(name="unit_style")
        unit_style.fill = PatternFill("solid", fgColor="DDEBF7")
        unit_style.font = Font(name="Calibri", size=10, color="000000")
        unit_style.alignment = Alignment(horizontal="left")
        unit_style.protection = Protection(hidden=True)

        admin_style = NamedStyle(name="admin_style")
        admin_style.fill = PatternFill("solid", fgColor="9BC2E6")
        admin_style.font = Font(name="Calibri", size=10, color="000000")
        admin_style.alignment = Alignment(horizontal="left")
        admin_style.protection = Protection(hidden=True)

        performance_style = NamedStyle(name="performance_style")
        performance_style.fill = PatternFill("solid", fgColor="DDEBF7")
        performance_style.font = Font(name="Calibri", size=10, color="000000")
        performance_style.alignment = Alignment(horizontal="left")
        performance_style.protection = Protection(hidden=True)

        def move_and_rename(bulletin_number):
            os.makedirs(new_folder, exist_ok=True)
            shutil.copy(old_folder + temp_file, new_folder + temp_file)
            os.rename(
                new_folder + temp_file,
                new_folder + "Boletim de Medição " + bulletin_number + ".xlsx",
            )
            return

        def return_file_path(bulletin_number):
            return new_folder + "Boletim de Medição " + bulletin_number + ".xlsx"

        def load_file(bulletin_number):
            wb = load_workbook(filename=return_file_path(bulletin_number))
            return wb

        def get_prices(obj):
            try:
                provisioned = get_provisioned_price(obj)
                total = obj.total_price
                spent = obj.spent_price
                return [provisioned, total, spent, total - spent]
            except Exception:
                raise Exception(
                    "Alguns valores monetários do contrato podem não estar preenchidos"
                )

        def get_contract_data(contract):
            try:
                object_number = contract.extra_info["r_c_number"]
            except Exception:
                raise Exception('Campo  "Nº do Objeto" não está preenchido')

            try:
                subcompany_name = contract.subcompany.name
            except Exception:
                try:
                    subcompany_name = contract.firm.name
                except Exception:
                    raise Exception('Campo "Fornecedor" não está preenchido')

            try:
                contract_status = contract.status.name
            except Exception:
                contract_status = None

            try:
                hirer_list = ", ".join(
                    sorted(
                        [
                            hirer.get_full_name()
                            for hirer in contract.responsibles_hirer.all()
                        ]
                    )
                )
            except Exception:
                raise Exception(
                    'Erro no preenchimento do campo "Prepostos da Contratante"'
                )

            try:
                hired_list = ", ".join(
                    sorted(
                        [
                            hired.get_full_name()
                            for hired in contract.responsibles_hired.all()
                        ]
                    )
                )
            except Exception:
                raise Exception(
                    'Erro no preenchimento do campo "Prepostos do Contratado"'
                )

            try:
                contract_name = contract.name
            except Exception:
                raise Exception('Campo "Descrição" não está preenchido')

            try:
                start_date = contract.contract_start
            except Exception:
                raise Exception('Campo de data "Início" não está preenchido')

            try:
                end_date = contract.contract_end
            except Exception:
                raise Exception('Campo de data "Fim" não está preenchido')

            prices = get_prices(contract)

            return [
                object_number,
                subcompany_name,
                contract_status,
                hirer_list,
                hired_list,
                *prices,
                contract_name,
                start_date,
                end_date,
            ]

        def get_bulletin_data(bulletin):
            total_price = bulletin.total_price

            performance_total_price = 0.0
            measurement_bulletin_scope = MeasurementBulletinScope(
                bulletin.contract, measurement_bulletin=bulletin
            )
            measurement_bulletin_scope.calculate_mb_average_grade_percent()
            average_grade_percent = measurement_bulletin_scope.average_grade_percent
            for (
                contract_service
            ) in bulletin.contract.contract_services_bulletins.filter(
                measurement_bulletins=bulletin
            ):
                try:
                    performance_total_price += (
                        contract_service.price / bulletin.contract.performance_months
                    ) * average_grade_percent
                except Exception:
                    pass

            creation_date = utc_to_local(bulletin.creation_date).replace(tzinfo=None)
            description = bulletin.contract.name

            try:
                status = bulletin.approval_step.name
            except Exception:
                status = None

            try:
                start_date = utc_to_local(bulletin.period_starts_at).replace(
                    tzinfo=None
                )
            except Exception:
                start_date = None

            try:
                end_date = utc_to_local(bulletin.period_ends_at).replace(tzinfo=None)
            except Exception:
                end_date = None

            try:
                months_count = bulletin.contract.performance_months
            except Exception:
                months_count = None

            try:
                accounting_classification = bulletin.extra_info[
                    "accounting_classification"
                ]
            except Exception:
                accounting_classification = None

            if bulletin.description != "":
                try:
                    comments = bulletin.description
                except Exception:
                    comments = None
            else:
                comments = None

            return [
                number,
                total_price + performance_total_price,
                creation_date,
                status,
                description,
                start_date,
                end_date,
                months_count,
                accounting_classification,
                comments,
            ]

        def get_survey_data(bulletin):
            def get_signature_status(obj, hire_type):
                if hire_type == "hirer":
                    signed_list = obj.signatures.filter(hirer__isnull=False)
                elif hire_type == "hired":
                    signed_list = obj.signatures.filter(hired__isnull=False)
                signed_status = [
                    True if item.signed_at else False for item in signed_list
                ]

                return "Assinado" if all(signed_status) else "Aguardando assinatura"

            def get_average_grade(obj):
                total_section = 0
                for contract_service in (
                    obj.contract.contract_services_bulletins.filter(
                        measurement_bulletins=bulletin
                    )
                    .order_by("created_at", "description")
                    .prefetch_related("contract_item_performance")
                ):
                    total_item = 0
                    for (
                        contract_item_performance
                    ) in contract_service.contract_item_performance.all():
                        field_surveys = [obj]
                        contr_item_p_scope = ContractItemPerformanceScope(
                            contract_item_performance, field_surveys
                        )
                        average = contr_item_p_scope.calculate_field_surveys_average()
                        total_item += (average * contract_item_performance.weight) / 100
                    total_section += total_item * contract_service.weight / 100

                return total_section / 10

            def join_people_names(field_manager) -> str:
                if field_manager.exists():
                    return ", ".join(
                        sorted(
                            [person.get_full_name() for person in field_manager.all()]
                        )
                    )
                else:
                    return ""

            data = []
            total_average_grade = 0
            if bulletin.bulletin_surveys.exists():
                for item in bulletin.bulletin_surveys.prefetch_related(
                    "responsibles_hired", "responsibles_hirer"
                ).all():
                    average_grade = get_average_grade(item)
                    total_average_grade += average_grade
                    rounded_avr_grade = round(average_grade, 4)

                    # Extract item info
                    # WARN: Remember to properly handle optional fields

                    item_number = item.number or ""
                    item_name = item.name or ""
                    created_at = item.created_at.strftime("%d/%m/%Y")
                    manual = item.manual

                    # Hirer and hired info
                    hirers = join_people_names(item.responsibles_hirer)
                    hirer_sign_status = get_signature_status(item, "hirer")
                    hired = join_people_names(item.responsibles_hired)
                    hired_sign_status = get_signature_status(item, "hired")

                    # Bulletin info
                    bulletin_description = bulletin.description or ""
                    entity_name = ""
                    if bulletin.contract:
                        company = None
                        if bulletin.contract.firm:
                            company = bulletin.contract.firm.company
                        elif bulletin.contract.subcompany:
                            # Company is not optional in SubCompany
                            company = bulletin.contract.subcompany.company

                        if company and company.entity_set.exists():
                            entity_name = company.entity_set.first().name

                    data.append(
                        [
                            item_number,
                            item_name,
                            hirers,
                            hirer_sign_status,
                            hired,
                            hired_sign_status,
                            created_at,
                            entity_name,
                            bulletin_description,
                            rounded_avr_grade,
                            manual,
                        ]
                    )

                total_average_grade = round(
                    total_average_grade / bulletin.bulletin_surveys.count(), 4
                )

            return [total_average_grade, data]

        def fill_contract_info(bulletin, workbook):
            data_contract = get_contract_data(bulletin.contract)

            try:
                cabecalho = workbook["Cabeçalho"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            for index, column in enumerate(cabecalho[CONTRACT_ROW:CONTRACT_ROW]):
                column.value = data_contract[index]
                if get_column_letter(index + 1) in ["F", "G", "H", "I"]:
                    column.number_format = "R$ #,##0.00"
                if get_column_letter(index + 1) in ["K", "L"]:
                    column.number_format = "dd/mm/yyyy"

        def fill_bulletin_info(bulletin, workbook):
            data_bulletin = get_bulletin_data(bulletin)

            try:
                cabecalho = workbook["Cabeçalho"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            for index, column in enumerate(
                cabecalho[BULLETIN_ROW:BULLETIN_ROW][: len(data_bulletin)]
            ):
                column.value = data_bulletin[index]
                if get_column_letter(index + 1) in ["B"]:
                    column.number_format = "R$ #,##0.00"
                if get_column_letter(index + 1) in ["C", "F", "G"]:
                    column.number_format = "dd/mm/yyyy"

        def fill_survey_info(bulletin, workbook):
            total_average_grade, data_survey = get_survey_data(bulletin)
            survey_header = [
                "Serial",
                "Avaliação",
                "Avaliadores da Contratante",
                "Status",
                "Avaliadores da Contratada",
                "Status",
                "Data",
                "Entidade",
                "Boletim de Medição",
                "Performance por Avaliação",
                "Performance Final",
            ]
            column_indexes_to_hide_info = [3, 4, 5, 6]
            try:
                cabecalho = workbook["Cabeçalho"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            if data_survey != []:
                for index, item in enumerate(survey_header):
                    cell = cabecalho.cell(row=SURVEY_HEADER_ROW, column=index + 1)
                    cell.value = item
                    cell.style = field_survey_style
                    cell.fill = PatternFill("solid", fgColor="9BC2E6")

                survey_info_row = 10
                cell = cabecalho.cell(row=survey_info_row, column=11)
                cell.value = total_average_grade
                cell.style = field_survey_style
                cell.number_format = FORMAT_PERCENTAGE_00

                for line_index, line_item in enumerate(data_survey):
                    for column_index, column_item in enumerate(line_item[:-1]):
                        cell = cabecalho.cell(
                            row=survey_info_row, column=column_index + 1
                        )
                        if (
                            line_item[-1] is True
                            and (column_index + 1) in column_indexes_to_hide_info
                        ):
                            cell.value = ""
                        else:
                            cell.value = column_item
                        cell.style = field_survey_style
                        if get_column_letter(column_index + 1) in ["G"]:
                            cell.number_format = "dd/mm/yyyy"
                        if get_column_letter(column_index + 1) in ["J"]:
                            cell.number_format = FORMAT_PERCENTAGE_00
                    survey_info_row += 1
            else:
                return

        def get_old_bulletins(bulletin):
            old_bulletins = (
                MeasurementBulletin.objects.filter(
                    contract=bulletin.contract,
                    creation_date__lte=bulletin.creation_date,
                )
                .exclude(uuid=bulletin.uuid)
                .prefetch_related(
                    "contract",
                    "contract__performance_services",
                    "contract__contract_services_bulletins",
                )
            )

            return old_bulletins

        def get_unit_data(bulletin):
            consumed_dict = {}
            consumed_qs = (
                ProcedureResource.objects.filter(measurement_bulletin=bulletin)
                .values("service_order_resource_id")
                .annotate(
                    total_amount=Sum("amount"), total_price_sum=Sum("total_price")
                )
            )
            for row in consumed_qs:
                consumed_dict[row["service_order_resource_id"]] = {
                    "amount": row["total_amount"] or 0,
                    "total_price": row["total_price_sum"] or 0,
                }

            old_dict = {}
            old_qs = (
                ProcedureResource.objects.filter(measurement_bulletin__in=old_bulletins)
                .values("resource_id")
                .annotate(
                    total_amount=Sum("amount"), total_price_sum=Sum("total_price")
                )
            )
            for row in old_qs:
                old_dict[row["resource_id"]] = {
                    "amount": row["total_amount"] or 0,
                    "total_price": row["total_price_sum"] or 0,
                }

            data_unit = []

            for service in (
                bulletin.contract.unit_price_services.all()
                .prefetch_related("contract_item_unit_prices")
                .order_by("created_at", "description")
            ):
                for item in (
                    service.contract_item_unit_prices.all()
                    .prefetch_related(
                        "resource",
                        "resource__resource",
                    )
                    .order_by("order", "uuid")
                ):
                    try:
                        service_order_resource_id = item.resource.pk
                        resource_id = item.resource.resource.pk

                        consumed = consumed_dict.get(
                            service_order_resource_id, {"amount": 0, "total_price": 0}
                        )
                        old_values = old_dict.get(
                            resource_id, {"amount": 0, "total_price": 0}
                        )

                        amount = consumed["amount"]
                        total_price = consumed["total_price"]
                        old_amount = old_values["amount"]
                        old_total_price = old_values["total_price"]
                    except Exception:
                        total_price = 0
                        old_amount = 0
                        old_total_price = 0
                        amount = 0

                    data_unit.append(
                        [
                            "Preço Unitário",
                            service.description,
                            item.sort_string,
                            item.resource.resource.name,
                            item.resource.resource.unit,
                            item.resource.unit_price,
                            None,
                            None,
                            item.resource.unit_price * item.resource.amount,
                            old_amount,
                            old_total_price,
                            amount,
                            total_price,
                            total_price + old_total_price,
                            (item.resource.unit_price * item.resource.amount)
                            - (total_price + old_total_price),
                        ]
                    )
            return data_unit

        def fill_unit_items(bulletin, workbook):
            data_start_row = 3

            data_unit = get_unit_data(bulletin)

            try:
                measurement = workbook["Medições (ADM e P.U.)"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            for line_index, line_item in enumerate(data_unit):
                for column_index, column_item in enumerate(line_item):
                    cell = measurement.cell(row=data_start_row, column=column_index + 1)
                    cell.value = column_item
                    cell.style = unit_style

                    if get_column_letter(column_index + 1) in [
                        "F",
                        "I",
                        "K",
                        "M",
                        "N",
                        "O",
                    ]:
                        cell.number_format = "R$ #,##0.00"

                data_start_row += 1

        def get_admin_data(bulletin):
            months_count = (
                bulletin.contract.performance_months if bulletin.contract else 0
            )
            work_day = bulletin.work_day

            vehicles_current_dict = {}
            vehicles_current_qs = bulletin.bulletin_vehicles.values(
                "contract_item_administration_id"
            ).annotate(total_amount=Sum("amount"), total_price_sum=Sum("total_price"))
            for row in vehicles_current_qs:
                contract_item_id = row["contract_item_administration_id"]

                if contract_item_id in vehicles_current_dict:
                    vehicles_current_dict[contract_item_id]["amount"] += (
                        row["total_amount"] or 0
                    )
                    vehicles_current_dict[contract_item_id]["total_price"] += (
                        row["total_price_sum"] or 0
                    )
                else:
                    vehicles_current_dict[contract_item_id] = {
                        "amount": row["total_amount"] or 0,
                        "total_price": row["total_price_sum"] or 0,
                    }

            workers_current_dict = {}
            workers_current_qs = bulletin.bulletin_workers.values(
                "contract_item_administration_id"
            ).annotate(total_amount=Sum("amount"), total_price_sum=Sum("total_price"))
            for row in workers_current_qs:
                contract_item_id = row["contract_item_administration_id"]

                if contract_item_id in workers_current_dict:
                    workers_current_dict[contract_item_id]["amount"] += (
                        row["total_amount"] or 0
                    )
                    workers_current_dict[contract_item_id]["total_price"] += (
                        row["total_price_sum"] or 0
                    )
                else:
                    workers_current_dict[contract_item_id] = {
                        "amount": row["total_amount"] or 0,
                        "total_price": row["total_price_sum"] or 0,
                    }

            equipment_current_dict = {}
            equipment_current_qs = bulletin.bulletin_equipments.values(
                "contract_item_administration_id"
            ).annotate(total_amount=Sum("amount"), total_price_sum=Sum("total_price"))
            for row in equipment_current_qs:
                contract_item_id = row["contract_item_administration_id"]

                if contract_item_id in equipment_current_dict:
                    equipment_current_dict[contract_item_id]["amount"] += (
                        row["total_amount"] or 0
                    )
                    equipment_current_dict[contract_item_id]["total_price"] += (
                        row["total_price_sum"] or 0
                    )
                else:
                    equipment_current_dict[contract_item_id] = {
                        "amount": row["total_amount"] or 0,
                        "total_price": row["total_price_sum"] or 0,
                    }

            vehicles_old_dict = defaultdict(list)
            vehicles_old_qs = DailyReportVehicle.objects.filter(
                measurement_bulletin__in=old_bulletins
            ).values(
                "contract_item_administration_id",
                "amount",
                "total_price",
                "measurement_bulletin__work_day",
            )
            for row in vehicles_old_qs:
                vehicles_old_dict[row["contract_item_administration_id"]].append(row)

            workers_old_dict = defaultdict(list)
            workers_old_qs = DailyReportWorker.objects.filter(
                measurement_bulletin__in=old_bulletins
            ).values(
                "contract_item_administration_id",
                "amount",
                "total_price",
                "measurement_bulletin__work_day",
            )
            for row in workers_old_qs:
                workers_old_dict[row["contract_item_administration_id"]].append(row)

            equipment_old_dict = defaultdict(list)
            equipment_old_qs = DailyReportEquipment.objects.filter(
                measurement_bulletin__in=old_bulletins
            ).values(
                "contract_item_administration_id",
                "amount",
                "total_price",
                "measurement_bulletin__work_day",
            )
            for row in equipment_old_qs:
                equipment_old_dict[row["contract_item_administration_id"]].append(row)

            data_admin = []

            for service in (
                bulletin.contract.administration_services.all()
                .prefetch_related("contract_item_administration")
                .order_by("created_at", "description")
            ):
                for item in (
                    service.contract_item_administration.all()
                    .prefetch_related("content_type", "resource", "resource__resource")
                    .order_by("order", "uuid")
                ):
                    # Set default values to ensure there's no access before assignment
                    average = None
                    total_price = 0
                    old_average = 0
                    old_total_price = 0

                    item_id = item.uuid

                    if item.content_type.model == "dailyreportvehicle":
                        try:
                            if item_id in vehicles_current_dict:
                                current_data = vehicles_current_dict[item_id]
                                if work_day:
                                    average = current_data["amount"] / work_day
                                    total_price = current_data["total_price"]

                            if item_id in vehicles_old_dict:
                                old_items = vehicles_old_dict[item_id]
                                old_average = sum(
                                    filter(
                                        None,
                                        (
                                            (row["amount"] or 0)
                                            / (
                                                row["measurement_bulletin__work_day"]
                                                or 1
                                            )
                                            for row in old_items
                                        ),
                                    )
                                )
                                old_total_price = sum(
                                    filter(
                                        None, (row["total_price"] for row in old_items)
                                    )
                                )
                        except Exception:
                            pass
                    elif item.content_type.model == "dailyreportworker":
                        try:
                            if item_id in workers_current_dict:
                                current_data = workers_current_dict[item_id]
                                if work_day:
                                    average = current_data["amount"] / work_day
                                    total_price = current_data["total_price"]

                            if item_id in workers_old_dict:
                                old_items = workers_old_dict[item_id]
                                old_average = sum(
                                    filter(
                                        None,
                                        (
                                            (row["amount"] or 0)
                                            / (
                                                row["measurement_bulletin__work_day"]
                                                or 1
                                            )
                                            for row in old_items
                                        ),
                                    )
                                )
                                old_total_price = sum(
                                    filter(
                                        None, (row["total_price"] for row in old_items)
                                    )
                                )
                        except Exception:
                            pass
                    elif item.content_type.model == "dailyreportequipment":
                        try:
                            if item_id in equipment_current_dict:
                                current_data = equipment_current_dict[item_id]
                                if work_day:
                                    average = current_data["amount"] / work_day
                                    total_price = current_data["total_price"]

                            if item_id in equipment_old_dict:
                                old_items = equipment_old_dict[item_id]
                                old_average = sum(
                                    filter(
                                        None,
                                        (
                                            (row["amount"] or 0)
                                            / (
                                                row["measurement_bulletin__work_day"]
                                                or 1
                                            )
                                            for row in old_items
                                        ),
                                    )
                                )
                                old_total_price = sum(
                                    filter(
                                        None, (row["total_price"] for row in old_items)
                                    )
                                )
                        except Exception:
                            pass

                    heading = "Administração"
                    service_description = service.description
                    sort_string = item.sort_string

                    # Handle resource fields being optional
                    service_order_resource = item.resource
                    resource = (
                        service_order_resource.resource
                        if service_order_resource
                        else None
                    )

                    so_unit_price = (
                        service_order_resource.unit_price
                        if service_order_resource
                        else 0
                    )
                    so_resource_amount = (
                        service_order_resource.amount if service_order_resource else 0
                    )
                    amount_by_months = (
                        round(so_resource_amount / months_count, 2)
                        if months_count
                        else so_resource_amount
                    )
                    unit_price_by_amount = so_unit_price * so_resource_amount
                    total_with_old_price = total_price + old_total_price

                    resource_name = resource.name if resource else ""
                    resource_unit = resource.unit if resource else ""

                    data_admin.append(
                        [
                            heading,
                            service_description,
                            sort_string,
                            resource_name,
                            resource_unit,
                            so_unit_price,
                            amount_by_months,
                            so_resource_amount,
                            unit_price_by_amount,
                            round(old_average, 2),
                            old_total_price,
                            average,
                            total_price,
                            total_with_old_price,
                            unit_price_by_amount - total_with_old_price,
                        ]
                    )

            return data_admin

        def fill_admin_items(bulletin, workbook):
            data_admin = get_admin_data(bulletin)

            try:
                measurement = workbook["Medições (ADM e P.U.)"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            data_start_row = (
                len([cell for cell in list(measurement["A"]) if cell.value]) + 2
            )

            for line_index, line_item in enumerate(data_admin):
                for column_index, column_item in enumerate(line_item):
                    cell = measurement.cell(row=data_start_row, column=column_index + 1)
                    cell.value = column_item
                    cell.style = admin_style
                    if line_index == 0:
                        cell.border = Border(
                            top=Side(border_style="thin", color="000000")
                        )

                    if get_column_letter(column_index + 1) in [
                        "F",
                        "I",
                        "K",
                        "M",
                        "N",
                        "O",
                    ]:
                        cell.number_format = "R$ #,##0.00"
                data_start_row += 1

        def get_performance_data(bulletin, old_bulletins):
            def get_old_bulletin_performance_data(bulletin, old_bulletins):
                contract = bulletin.contract

                old_bulletin_count = old_bulletins.count()

                old_spend_value = defaultdict(int)

                if old_bulletins.exists():
                    for old_bulletin in old_bulletins:
                        old_bulletin_data = MeasurementBulletinScope(
                            contract, measurement_bulletin=old_bulletin
                        )
                        old_bulletin_data.calculate_mb_average_grade_percent()
                        old_contract = old_bulletin.contract
                        for (
                            old_service
                        ) in old_contract.contract_services_bulletins.filter(
                            measurement_bulletins=bulletin
                        ):
                            old_spend_value[str(old_service.uuid)] += (
                                old_bulletin_data.average_grade_percent
                                * old_service.price
                            ) / contract.performance_months

                return [old_bulletin_count, old_spend_value]

            def get_average_value(service, results):
                average = results.get((str(service.uuid)), 0)
                return average

            data_performance = []
            total_value = 0
            total_spend_value = 0
            total_percentage = 0
            total_old_spend_value = 0
            contract = bulletin.contract

            bulletin_performance_data = MeasurementBulletinScope(
                contract, measurement_bulletin=bulletin
            )
            _ = bulletin_performance_data.calculate_mb_average_grade_percent()

            result_performance_data = {
                str(item.contract_service.uuid): item.weighted_average
                for item in bulletin_performance_data.contract_services_scope
            }

            (
                old_bulletin_count,
                old_spend_value,
            ) = get_old_bulletin_performance_data(bulletin, old_bulletins)

            for service in contract.contract_services_bulletins.filter(
                measurement_bulletins=bulletin
            ).order_by("created_at", "description"):
                total_value += service.price
                total_spend_value += (
                    bulletin_performance_data.average_grade_percent * service.price
                ) / contract.performance_months
                average_value = get_average_value(service, result_performance_data)
                total_percentage += round(
                    (average_value * service.weight) / 100,
                    4,
                )
                try:
                    total_old_spend_value += old_spend_value[str(service.uuid)]
                except Exception:
                    pass
                data_performance.append(
                    [
                        "Performance",
                        service.description,
                        service.weight / 100,
                        None,
                        None,
                        round(
                            service.price / contract.performance_months,
                            2,
                        ),
                        1,
                        contract.performance_months,
                        service.price,
                        old_bulletin_count,
                        old_spend_value[str(service.uuid)],
                        round(
                            average_value,
                            4,
                        ),
                        round(
                            (
                                bulletin_performance_data.average_grade_percent
                                * service.price
                            )
                            / contract.performance_months,
                            2,
                        ),
                        round(
                            (average_value * service.weight) / 100,
                            4,
                        ),
                        round(
                            (
                                bulletin_performance_data.average_grade_percent
                                * service.price
                            )
                            / contract.performance_months
                            + old_spend_value[str(service.uuid)],
                            2,
                        ),
                        round(
                            service.price
                            - (
                                (
                                    bulletin_performance_data.average_grade_percent
                                    * service.price
                                )
                                / contract.performance_months
                                + old_spend_value[str(service.uuid)]
                            ),
                            2,
                        ),
                    ]
                )

            data_perforamance_header = [
                "Performance",
                None,
                1,
                None,
                None,
                None,
                None,
                None,
                round(total_value, 2),
                None,
                round(total_old_spend_value, 2),
                None,
                round(total_spend_value, 2),
                round(total_percentage, 4),
                round(total_spend_value + total_old_spend_value, 2),
                round(total_value - (total_spend_value + total_old_spend_value), 2),
            ]

            return [data_perforamance_header, data_performance]

        def fill_performance_items(bulletin, old_bulletins, workbook):
            data_perforamance_header, data_performance = get_performance_data(
                bulletin, old_bulletins
            )

            try:
                performance = workbook["Medição (Performance)"]
            except Exception:
                raise Exception("Favor contatar a equipe de Suporte da Kartado")

            performance.row_dimensions[PERFORMANCE_HEADER_ROW].height = 25

            for index, item in enumerate(data_perforamance_header):
                cell = performance.cell(row=PERFORMANCE_HEADER_ROW, column=index + 1)
                cell.value = item
                cell.style = performance_style
                cell.font = Font(name="Calibri", size=10, color="000000", bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if index == 0:
                    cell.border = Border(left=Side(border_style="thin", color="000000"))
                if index == len(data_perforamance_header) - 1:
                    cell.border = Border(
                        right=Side(border_style="thin", color="000000")
                    )
                if get_column_letter(index + 1) in ["I", "K", "M", "O", "P"]:
                    cell.number_format = "R$ #,##0.00"
                if get_column_letter(index + 1) in ["C", "N"]:
                    cell.number_format = FORMAT_PERCENTAGE_00

            performance_info_row = 4

            for line_index, line_item in enumerate(data_performance):
                for column_index, column_item in enumerate(line_item):
                    cell = performance.cell(
                        row=performance_info_row, column=column_index + 1
                    )
                    cell.value = column_item
                    cell.style = performance_style
                    if line_index == 0:
                        cell.border = Border(
                            top=Side(border_style="thin", color="000000")
                        )
                    if get_column_letter(column_index + 1) in [
                        "F",
                        "I",
                        "K",
                        "M",
                        "O",
                        "P",
                    ]:
                        cell.number_format = "R$ #,##0.00"
                    if get_column_letter(column_index + 1) in ["C", "L", "N"]:
                        cell.number_format = FORMAT_PERCENTAGE_00
                performance_info_row += 1

        try:
            old_bulletins = get_old_bulletins(bulletin)

            prefetch_related_objects(
                [bulletin],
                "contract__firm",
                "contract__subcompany",
                "contract__status",
                "contract__responsibles_hirer",
                "contract__responsibles_hired",
                "contract__unit_price_services",
                "contract__administration_services",
                "contract__performance_services",
                "contract__contract_services_bulletins",
                "approval_step",
                "bulletin_vehicles",
                "bulletin_workers",
                "bulletin_equipments",
                "bulletin_surveys",
            )

            move_and_rename(number)
            workbook = load_file(number)
            fill_contract_info(bulletin, workbook)
            fill_bulletin_info(bulletin, workbook)
            fill_survey_info(bulletin, workbook)
            fill_unit_items(bulletin, workbook)
            fill_admin_items(bulletin, workbook)
            fill_performance_items(bulletin, old_bulletins, workbook)
            workbook.save(return_file_path(number))

            filename = os.listdir(new_folder)[0]

            with open(new_folder + filename, "rb") as excel_file:
                bulletin_export.exported_file.save(
                    filename, ContentFile(excel_file.read())
                )
            os.remove(return_file_path(number))
            os.rmdir(new_folder)

            error = False
        except Exception as e:
            error = True
            logging.error(
                "Untreated exception found while exporting file. Check Sentry."
            )
            sentry_sdk.capture_exception(e)
            pass

        finally:
            bulletin_export.error = error
            bulletin_export.done = True
            bulletin_export.save()
