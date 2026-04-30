import tempfile
from uuid import uuid4

from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import Reporting, ReportingFile
from helpers.apps.ccr_report_utils.image import (
    ReportFormat,
    ResizeMethod,
    SheetTarget,
    get_logo_file,
    get_provider_logo_file,
    insert_logo_and_provider_logo,
    insert_picture_2,
)
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import COMMON_IMAGE_TYPE, clean_latin_string


def _therapy_format_km(km: float, left_padding=0):
    try:
        numbers = format(round(km, 3), ".3f").split(".")
        return f"{numbers[0].zfill(left_padding)}+{numbers[1].zfill(left_padding)}"
    except Exception:
        return ""


def _get_query_upload_valid_img():
    q = Q()
    for extension_file in COMMON_IMAGE_TYPE:
        q |= Q(upload__icontains=extension_file)
    return q


def _get_data_work(data_list: list) -> dict:
    data_work = {}

    for _d in data_list:
        try:
            road_name = _d[0].get("road_name")
        except Exception:
            continue

        if road_name not in data_work:
            data_work[road_name] = []
        for _ in _d:
            data_work[road_name].append(_)

    for key, _data_list in data_work.items():
        south = [obj for obj in _data_list if obj["direction"].lower() == "sul"]
        south = (
            sorted(south, key=lambda x: x.get("km_float"), reverse=False)
            if south
            else south
        )
        north = [obj for obj in _data_list if obj["direction"].lower() == "norte"]
        north = (
            sorted(north, key=lambda x: x.get("km_float"), reverse=True)
            if north
            else north
        )
        canteiro = [
            obj for obj in _data_list if obj["direction"].lower() == "canteiro central"
        ]
        canteiro = (
            sorted(canteiro, key=lambda x: x.get("km_float"), reverse=True)
            if canteiro
            else canteiro
        )
        others = [
            obj
            for obj in _data_list
            if obj["direction"].lower() not in ["sul", "norte", "canteiro central"]
        ]

        data_work[key] = north + south + canteiro + others

    return data_work


class XlsxHandlerBaseEPS(object):
    def __init__(
        self,
        uuid: str,
        list_uuids: list,
        s3,
        path_file: str,
        file_name: str,
        class_name: str,
        title: str,
        dispositive: str,
        sheet_target: SheetTarget = SheetTarget.DesktopExcel,
        report_format: ReportFormat = ReportFormat.XLSX,
    ):
        self._path_file = path_file
        self._sheet_target = sheet_target
        self._report_format = report_format
        self.wb: Workbook = None
        self.__worksheet: Worksheet = None
        self.file_name = file_name
        self.class_name = class_name
        self.title = title
        self.dispositive = dispositive
        self.s3 = s3
        self.temp_file = tempfile.mkdtemp()
        self.uuid = uuid
        self.__init_wb()
        self.list_uuids = list_uuids
        self.reportings = Reporting.objects.filter(uuid=uuid).prefetch_related(
            "company"
        )

        self.data_logo_company: dict = dict(
            path_image="",
            range_string="J1:J4",
            resize_method=ResizeMethod.ProportionalRight,
        )

        self.data_provider_logo: dict = dict(
            path_image="",
            range_string="A1:A4",
            resize_method=ResizeMethod.ProportionalLeft,
        )

        first_reporting = self.reportings.first()
        self.form = first_reporting.occurrence_type
        self.company = first_reporting.company
        self.resume_report_dict = []

        self.static_fields = {
            "data_sheet": "A",
            "dispositive": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "latitude": "F",
            "longitude": "G",
            "action": "H",
            "photo_therapy": "I",
            "photo_solution": "J",
            "serial_monitoring": "K",
            "serial_recovery": "L",
            "serial_inventory": "M",
        }

    def _get_download_pictures(self, url):
        path_image = ""
        if url:
            try:
                file_path = url.split("?")[0].split(".com/")[1]
                bucket_name = url.split(".s3")[0].split("/")[-1]
                image_format = file_path.split(".")[-1]
                path_image = f"{self.temp_file}{uuid4()}.{image_format}"
                self.s3.download_file(bucket_name, file_path, path_image)
            except Exception:
                path_image = ""
        return path_image

    def _get_first_photo(self, reporting_file_pk) -> str:
        photo = ""

        if isinstance(reporting_file_pk, list):
            query = {"pk__in": reporting_file_pk}
        else:
            query = {"pk": reporting_file_pk}

        q = _get_query_upload_valid_img()

        _file = (
            ReportingFile.objects.filter(**query)
            .filter(q)
            .filter(is_shared=True)
            .order_by("datetime")
            .first()
        )

        if _file and _file.upload:
            photo = self._get_download_pictures(url=_file.upload.url)

        return photo

    def _insert_new_rows(self, row: int = None):
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        chars = [chr(ord("A") + i) for i in range(13)]

        for char in chars:
            self._worksheet[f"{char}{row}"].border = border
            self._worksheet.row_dimensions[row].height = 81.75

    def create_dict(self, reporting):
        data = []

        form_data_display = reporting.get_form_data_display()

        DATA_SHEET = form_data_display.get("id_ccr_antt", "-")
        KM_FLOAT = reporting.km

        DIRECTION = get_custom_option(reporting, "direction") or "-"
        SERIAL_MONITORING = reporting.number or "-"
        SERIAL_INVENTORY = reporting.parent.number if reporting.parent else "-"
        ROAD_NAME = reporting.road_name or "-"

        therapy = reporting.form_data.get("therapy")

        for _trp in therapy:
            key = _trp.get("occurrence_type")
            # Algumas instancias estão com bug get() mesmo vindo nulo não esta pegando o "-"
            LATITUDE = _trp.get("latitude", "-") or "-"
            LONGITUDE = _trp.get("longitude", "-") or "-"
            photo_therapy = ""
            action = "-"
            photo_solution = ""
            serial_recovery = "-"
            try:
                km_initial = _therapy_format_km(float(_trp.get("kilometer")), 3) or "-"
            except Exception:
                km_initial = "-"
            try:
                end_km = _therapy_format_km(float(_trp.get("kilometer_end")), 3) or "-"
            except Exception:
                end_km = "-"

            if key:
                occurrence_type = OccurrenceType.objects.filter(pk=key).first()
                photo_therapy = self._get_first_photo(
                    reporting_file_pk=_trp.get("treatment_images", [])
                )

                if occurrence_type:
                    action = occurrence_type.name
                    recovery = reporting.self_relations.filter(
                        occurrence_type__pk=key
                    ).first()

                    if recovery:
                        q = _get_query_upload_valid_img()
                        recovery_file = (
                            recovery.reporting_files.filter(kind__iexact="depois")
                            .filter(q)
                            .order_by("-datetime")
                            .first()
                        )
                        if recovery_file:
                            photo_solution = self._get_first_photo(
                                reporting_file_pk=recovery_file.pk
                            )
                        serial_recovery = recovery.number

            data.append(
                {
                    "data_sheet": DATA_SHEET,
                    "dispositive": self.dispositive,
                    "initial_km": km_initial,
                    "end_km": end_km,
                    "direction": DIRECTION,
                    "latitude": LATITUDE,
                    "longitude": LONGITUDE,
                    "road_name": ROAD_NAME,
                    "action": action,
                    "photo_therapy": photo_therapy,
                    "photo_solution": photo_solution,
                    "serial_monitoring": SERIAL_MONITORING,
                    "serial_recovery": serial_recovery,
                    "serial_inventory": SERIAL_INVENTORY,
                    "km_float": KM_FLOAT,
                }
            )
        return data

    @classmethod
    def format_fonts(
        cls,
        *,
        cell,
        name="Calibri",
        size: int,
        bold=False,
        horizontal="center",
        vertical="center",
        wrap_text: bool = True,
    ) -> None:
        cell.alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrap_text=wrap_text
        )
        cell.font = Font(name=name, sz=size, bold=bold)

    def fill_sheet(self, *, data_list: list, road_name: str):
        self._worksheet.freeze_panes = "A5"
        initial_row = 6

        cell_title = "A1"

        self.format_fonts(
            cell=self._worksheet[cell_title],
            size=16,
            bold=True,
        )

        title = self.title % road_name
        self._worksheet[cell_title] = title

        col_photo_therapy = 9
        col_photo_solution = 10

        for values in data_list:
            for key, value in values.items():
                self._insert_new_rows(row=initial_row)
                if key not in self.static_fields:
                    continue

                col = self.static_fields[key]
                cell = f"{col}{initial_row}"

                if key in ["photo_therapy", "photo_solution"]:
                    if value:
                        if key == "photo_therapy":
                            column = col_photo_therapy
                        elif key == "photo_solution":
                            column = col_photo_solution

                        range_string = f"{get_column_letter(column)}{initial_row}"
                        insert_picture_2(
                            self._worksheet,
                            range_string,
                            Image(value),
                            self._sheet_target,
                            (1, 1, 1, 1),
                            ResizeMethod.ProportionalCentered,
                        )

                else:
                    data_format = dict(cell=self._worksheet[cell], size=11)
                    self._worksheet[cell] = value
                    if key in [
                        "serial_monitoring",
                        "serial_recovery",
                        "serial_inventory",
                    ]:
                        data_format["wrap_text"] = False

                    self.format_fonts(**data_format)

            initial_row += 1

        last_col = list(self.static_fields.values())[-4]
        self._worksheet.print_area = f"A1:{last_col}{initial_row-1}"

        if self._report_format == ReportFormat.PDF:
            self._worksheet.delete_cols(11, 3)

        insert_logo_and_provider_logo(
            worksheet=self._worksheet,
            logo_company=self.data_logo_company,
            provider_logo=self.data_provider_logo,
        )

    def execute(self):
        query_set = Reporting.objects.filter(
            occurrence_type__name=self.class_name,
            uuid__in=self.list_uuids,
        ).prefetch_related("occurrence_type", "firm", "firm__subcompany")

        data = []
        for reporting in query_set:
            if reporting.form_data.get("therapy"):
                data.append(self.create_dict(reporting=reporting))

            if not self.data_logo_company.get("path_image"):
                path_logo_company = get_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
            if path_logo_company:
                self.data_logo_company["path_image"] = path_logo_company

            if not self.data_provider_logo.get("path_image"):
                path_provider_logo = get_provider_logo_file(
                    s3=self.s3,
                    temp_prefix=self.temp_file,
                    reporting=reporting,
                )
                if path_provider_logo:
                    self.data_provider_logo["path_image"] = path_provider_logo

        data.reverse()

        data_work = _get_data_work(data)

        files = []

        for road_name, data_list in data_work.items():
            if data_list:
                self.fill_sheet(data_list=data_list, road_name=road_name)
                file_name = f"{self.file_name}{road_name}"
                file_name = clean_latin_string(
                    file_name.replace(".", "").replace("/", "")
                )
                result = f"/tmp/{file_name}.xlsx"
                self.wb.save(result)
                self.__init_wb()
                files.append(result)

        return files

    def __init_wb(self):
        self.wb = load_workbook(self._path_file)
        self._worksheet = self.wb.active
        self._worksheet.title = "Anomalias"
