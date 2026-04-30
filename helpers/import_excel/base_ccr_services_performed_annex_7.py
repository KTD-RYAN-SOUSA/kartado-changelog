import re
import tempfile

from openpyxl.utils import column_index_from_string

from helpers.apps.ccr_report_utils.base_ccr_xlsx_handler import BaseXlsxHandler
from helpers.apps.ccr_report_utils.export_utils import get_km_plus_meter
from helpers.apps.ccr_report_utils.image import (
    ResizeMethod,
    download_reporting_file_pictures,
    insert_picture_to_path,
)
from helpers.apps.ccr_report_utils.reporting_utils import get_custom_option
from helpers.strings import clean_latin_string


class XlsxHandlerBaseServicePerformedAnnex7(BaseXlsxHandler):
    def __init__(
        self,
        s3,
        query_set_serializer: str,
        path_file_xlsx,
        file_name_single,
        title,
        sheet_target,
        **kwargs,
    ):
        self.temp_file = tempfile.mkdtemp()
        self.s3 = s3
        self.file_name_single = file_name_single
        self.title = title
        self.executed_at_after = kwargs.pop("executed_at_after", None)
        self.executed_at_before = kwargs.pop("executed_at_before", None)
        self.text_fix = kwargs.get("text_fix", "-")
        self.height_row = 81.75
        self.initial_row = 5
        self.pk_invalid = kwargs.pop("pk_invalid", None)

        self.static_fields = {
            "data_sheet": "A",
            "road_name": "B",
            "initial_km": "C",
            "end_km": "D",
            "direction": "E",
            "type_element": "F",
            "text_fix": "G",
            "photo": "H",
            "action_general_conservation_statue": "I",
            "action_photo": "J",
            "action_names": "K",
            "hidden_serial_monitoring": "L",
            "hidden_serial_inventory": "M",
        }

        keys_sort = ["road_name", "km"]
        orders = {
            "filters": [
                {
                    "field": "",
                    "value": "",
                }
            ],
            "sorts": keys_sort,
        }
        split_xlsx_in = ["is_valid", "road_name"]
        super().__init__(
            query_set_serializer=query_set_serializer,
            path_file_xlsx=path_file_xlsx,
            split_xlsx_in=split_xlsx_in,
            sheet_target=sheet_target,
            orders_rows=orders,
            logo_company_range_string=kwargs.get("logo_company_range_string"),
            provider_logo_range_string=kwargs.get("provider_logo_range_string"),
        )

    def create_dict(self, reporting):
        data = {
            "km": reporting.km,
            "road_name": reporting.road_name,
        }

        IS_INVALID = str(reporting.pk) in self.pk_invalid
        if IS_INVALID:
            data.update({"is_invalid": True})
            return data

        form_data = reporting.form_data
        KM_INITAL = get_km_plus_meter(km=reporting.km) or "-"
        KM_END = get_km_plus_meter(km=reporting.end_km) or "-"
        DATA_SHEET = form_data.get("id_ccr_antt", "-")
        DIRECTION = get_custom_option(reporting, "direction") or "-"
        photos_report = form_data.get("fotos_relatorio", {})
        photo = ""
        if photos_report:
            for photos in photos_report:
                photos_detail = photos.get("fotos_detalhe", [])
                if photos_detail:
                    data_photo: dict = download_reporting_file_pictures(
                        s3=self.s3,
                        path=self.temp_file,
                        reporting_file_uuid=photos_detail,
                        order_by="datetime",
                        enable_is_shared_antt=True,
                        enable_include_dnit=False,
                    )
                    if data_photo.get("status"):
                        photo = data_photo.get("images")[0]["path"]
                        break

        action_photo = ""
        action_names = "-"
        action_general_conservation_statue = "-"
        TYPE_ELEMENT = form_data.get("tipoelemento") or "-"
        SERIAL_MONITORING = reporting.number or "-"
        instance_inventory = reporting.get_inventory()
        qs_children_raw = reporting.get_children()
        qs_children = qs_children_raw.filter(
            executed_at__gte=self.executed_at_after,
            executed_at__lte=self.executed_at_before,
        ).order_by("-executed_at")
        if qs_children.exists():
            currency_action = qs_children.first()
            action_names = (
                ", ".join(
                    list(
                        set(
                            (
                                qs_children.values_list(
                                    "occurrence_type__name", flat=True
                                ).distinct()
                            )
                        )
                    )
                )
                or "-"
            )
            pks_reporting_files = list(
                currency_action.reporting_files.values_list("pk", flat=True).filter(
                    datetime__gte=self.executed_at_after,
                    datetime__lte=self.executed_at_before,
                )
            )
            currency_action_form_data: dict = currency_action.get_form_data_display()
            action_general_conservation_statue = currency_action_form_data.get(
                "general_conservation_state", "-"
            )
            if pks_reporting_files:
                data_photos_action = download_reporting_file_pictures(
                    s3=self.s3,
                    path=self.temp_file,
                    reporting_file_uuid=pks_reporting_files,
                    order_by="-datetime",
                    enable_is_shared_antt=True,
                    enable_include_dnit=False,
                )
                if data_photos_action.get("status"):
                    action_photo = data_photos_action["images"][0]["path"]

        SERIAL_INVENTORY = instance_inventory.number if instance_inventory else "-"

        text_fix = self.text_fix

        data.update(
            {
                "data_sheet": DATA_SHEET,
                "initial_km": KM_INITAL,
                "end_km": KM_END,
                "direction": DIRECTION,
                "type_element": TYPE_ELEMENT,
                "text_fix": text_fix,
                "photo": photo,
                "action_photo": action_photo,
                "action_names": action_names,
                "action_general_conservation_statue": action_general_conservation_statue,
                "hidden_serial_monitoring": SERIAL_MONITORING,
                "hidden_serial_inventory": SERIAL_INVENTORY,
            }
        )
        return data

    def fill_sheet(self, data_work: dict):
        files = []
        for _, datas in data_work.items():
            road_name = ""
            initial_row = self.initial_row
            col_valid = [
                _v for _k, _v in self.static_fields.items() if "hidden" not in _k
            ]
            col_valid.sort()
            last_col = col_valid[-1]

            skip_invalids = any(set([not x.get("is_invalid") for x in datas]))
            for data in datas:
                if not road_name:
                    road_name = data.get("road_name")
                    self.worksheet["A1"] = self.title + road_name.replace(" ", "/")
                    self.worksheet.title = road_name

                is_invalid = data.get("is_invalid")
                if is_invalid:
                    if skip_invalids:
                        continue

                    merged_cell = self.merge_cells(
                        start_row=initial_row,
                        start_column=1,
                        end_row=initial_row,
                        end_column=column_index_from_string(last_col),
                    )
                    merged_cell.value = "Não foram selecionados apontamentos com os critérios necessários.."
                    break

                self.insert_new_row(
                    row=initial_row,
                    number_col=len(self.static_fields),
                    height_row=self.height_row,
                )
                for key, value in data.items():
                    if key not in self.static_fields:
                        continue

                    col = self.static_fields[key]
                    cell = f"{col}{initial_row}"

                    if key in ["photo", "action_photo"]:
                        if value:
                            insert_picture_to_path(
                                path_picture=value,
                                worksheet=self.worksheet,
                                target=self.sheet_target,
                                range_string=f"{cell}:{cell}",
                                resize_method=ResizeMethod.Stretch,
                            )

                    else:
                        self.worksheet[cell] = value

                initial_row += 1

            self.insert_logos(
                logo_company_config=self.data_logo_company,
                provider_logo_config=self.data_provider_logo,
            )

            self.worksheet.print_area = f"A1:{last_col}{self.worksheet.max_row}"

            file_name = self.file_name_single % re.sub(r"[-\ ]", "", road_name)
            file_name = clean_latin_string(re.sub(r"[./]", "", file_name))

            result_file = f"/tmp/{file_name}.xlsx"
            self.workbook.save(result_file)
            files.append(result_file)
            self.reload_workbook()

        return files
