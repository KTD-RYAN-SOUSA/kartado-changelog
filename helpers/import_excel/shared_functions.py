import logging
import os
from urllib import parse

import boto3
import openpyxl
from openpyxl import load_workbook

from RoadLabsAPI.settings import credentials


def shared_update_column_errors(item_dict, column_errors, column_mapping=None):
    """
    Updates the column errors in the item dictionary
    """
    if column_mapping:
        column_errors = [column_mapping.get(item, item) for item in column_errors]

    errors = item_dict.pop("column_errors", [])
    if column_errors:
        errors.extend(column_errors)
    item_dict["column_errors"] = list(set(errors))

    return item_dict


def shared_download_excel_file(excel_import, temp_path, s3_client=None):
    if not excel_import.excel_file:
        return ""

    if s3_client is None:
        s3_client = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )

    try:
        unquoted_file_path = parse.unquote(excel_import.excel_file.url)
        file_path = unquoted_file_path.split("?")[0].split(".com/")[1]
        bucket_name = unquoted_file_path.split(".s3")[0].split("/")[-1]
        file_format = file_path.split(".")[-1]
        file_name = file_path.split("/")[-1].split(".")[0]
    except Exception as e:
        logging.error(f"Error parsing S3 URL: {str(e)}")
        return ""

    os.makedirs(temp_path, exist_ok=True)
    file_temp_path = f"{temp_path}{file_name}_{excel_import.pk}.{file_format}"
    try:
        s3_client.download_file(bucket_name, file_path, file_temp_path)
        return file_temp_path
    except Exception as e:
        logging.error(f"Error downloading file from S3: {str(e)}")
        return ""


def shared_load_data(file_name, use_openpyxl=False):
    try:
        if use_openpyxl:
            return openpyxl.load_workbook(filename=file_name)
        else:
            return load_workbook(filename=file_name)
    except Exception:
        return None


def shared_is_hidden_sheet(worksheet):
    return worksheet.sheet_state in ["hidden", "veryHidden"]


def shared_clean_up(file_name, temp_path):
    try:
        if file_name and os.path.exists(file_name):
            os.remove(file_name)

        if os.path.exists(temp_path) and not os.listdir(temp_path):
            os.rmdir(temp_path)
    except Exception as e:
        logging.error(f"Error cleaning up: {str(e)}")
