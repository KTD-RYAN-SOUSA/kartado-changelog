import hashlib
import io
import json
import logging
import os
import re
import tempfile
import uuid
import zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
from urllib import parse

import boto3
import botocore.config
import pytz
import sentry_sdk
from django.conf import settings
from django.contrib.gis.geos import GeometryCollection, Point
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction
from django.db.models import F, OuterRef, Subquery
from django.db.models.signals import post_init, pre_init
from django.utils import timezone
from django.utils.timezone import make_aware
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from zappa.asynchronous import task

from apps.approval_flows.models import ApprovalStep
from apps.companies.models import Company, Firm
from apps.occurrence_records.models import OccurrenceType
from apps.reportings.models import RecordMenuRelation, Reporting, ReportingFile
from apps.reportings.serializers import ReportingFileSerializer, ReportingSerializer
from apps.resources.models import ContractService, Resource
from apps.roads.models import Road
from apps.service_orders.models import ProcedureResource, ServiceOrderActionStatus
from apps.service_orders.serializers import ProcedureResourceSerializer
from apps.templates.models import ExcelImport, ExcelReporting
from apps.templates.serializers import ExcelReportingSerializer
from apps.users.models import User
from apps.work_plans.models import Job
from helpers.apps.ccr_report_utils.export_utils import get_random_string, upload_file
from helpers.apps.json_logic import apply_json_logic
from helpers.apps.reporting_export import get_inventory_queryset, get_reporting_queryset
from helpers.dates import parse_dict_dates, to_utc_string
from helpers.histories import bulk_update_with_history
from helpers.import_excel.contract_item_administration_import import (
    execute_import as execute_administration_import,
)
from helpers.import_excel.contract_item_administration_import import (
    generate_preview as generate_administration_preview,
)
from helpers.import_excel.contract_item_unit_price_import import (
    execute_import,
    generate_preview,
)
from helpers.import_excel.contract_items_import import (
    execute_import as execute_contract_items_import,
)
from helpers.import_excel.contract_items_import import (
    generate_preview as generate_contract_items_preview,
)
from helpers.import_excel.shared_functions import (
    shared_clean_up,
    shared_download_excel_file,
    shared_is_hidden_sheet,
    shared_load_data,
    shared_update_column_errors,
)
from helpers.input_masks import (
    format_cpf_brazilin,
    format_mobile_number_brazilin,
    format_phone_number_brazilin,
)
from helpers.km_converter import get_road_coordinates
from helpers.permissions import PermissionManager
from helpers.road_defaults import create_default_segment_road, should_add_default_marks
from helpers.signals import DisableSignals
from helpers.strings import (
    COMMON_IMAGE_TYPE,
    clean_invalid_characters,
    clean_latin_string,
    dict_to_casing,
    get_obj_from_path,
    keys_to_snake_case,
    to_snake_case,
)
from helpers.validators.brazilian_documents import validate_CPF
from helpers.validators.brazilian_phone_number import phone_validation
from RoadLabsAPI.settings import credentials

ZIP_PICTURES_THREADING_LIMIT = 30
UPLOAD_THREADING_LIMIT = 30
REPORTING_LIMIT = 5000
ROW_LIMITS = 5000
THREADING_LIMIT = 5
MIN_RF_BATCH_SIZE = 100
RF_BATCHES = 10
PHOTO_COLUMNS = ["foto", "descricao foto", "tipo foto", "data foto"]
RESOURCE_COLUMNS = ["secao do recurso", "recurso", "quantidade"]
ARRAY_COLUMNS = {}
ARRAY_PHOTO_COLUMNS = {}
REQUIRED_FIELDS = [
    "lane",
    "direction",
    "firm_id",
    "status_id",
    "company_id",
    "occurrence_type_id",
]

IMG_PATTERN = re.compile(r"(^Foto_\d+$)|(- Foto_\d+$)")

NAME_TO_PROPERTY = {
    "km": "km",
    "km final": "end_km",
    "km de projeto": "project_km",
    "km final de projeto": "project_end_km",
    "status": "status_id",
    "equipe": "firm_id",
    "empreiteira": "firm_id",
    "equipe/empreiteira": "firm_id",
    "encontrado em": "found_at",
    "executado em": "executed_at",
    "prazo": "due_at",
    "sentido": "direction",
    "direcao": "direction",
    "unidade": "company_id",
    "obra": "job_id",
    "classe": "occurrence_type_id",
    "faixa": "lane",
    "rodovia": "road_name",
    "pista": "track",
    "km de referencia": "km_reference",
    "ramo": "branch",
    "lote": "lot",
    "codigo do inventario para vinculo com apontamento": "inventory_value",
    "latitude": "latitude",
    "longitude": "longitude",
    "menu": "menu_id",
    "serial do apontamento ou do item de inventario que sera editado": "number",
}

PROPERTY_TO_NAME = {
    "km": "km",
    "end_km": "km final",
    "project_km": "km de Projeto",
    "project_end_km": "km final de Projeto",
    "status_id": "Status",
    "firm_id": "Equipe",
    "found_at": "Encontrado em",
    "executed_at": "Executado em",
    "due_at": "Prazo",
    "direction": "Sentido",
    "company_id": "Unidade",
    "job_id": "Obra",
    "occurrence_type_id": "Classe",
    "lane": "Faixa",
    "road_name": "Rodovia",
    "track": "Pista",
    "km_reference": "km de referência",
    "branch": "Ramo",
    "lot": "Lote",
    "inventory_value": "Código do Inventário para vinculo com apontamento",
    "import_inventory_code": "Código do Inventário para vinculo com apontamento",  # translation added to error handling functions and methods
    "latitude": "Latitude",
    "longitude": "Longitude",
    "point": "Coordenadas",
    "km_error": "km",
    "road_error": "Rodovia",
    "menu_id": "Menu",
    "menu_error": "Menu",
    "number": "Serial do apontamento ou do item de inventário que será editado",
}


def get_real_max_rows_per_sheet(file_path: str) -> Dict[str, int]:
    """
    Calculate real max_row per sheet using read_only mode + reset_dimensions().
    Fixes incorrect max_row metadata (1,048,576) in some Excel files.
    """
    result = {}
    EMPTY_ROWS_THRESHOLD = 10
    wb_readonly = None

    try:
        wb_readonly = load_workbook(filename=file_path, read_only=True)

        for ws in wb_readonly.worksheets:
            if ws.sheet_state == "visible":
                ws.reset_dimensions()
                real_max = 1
                empty_rows_count = 0

                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if any(cell is not None for cell in row):
                        real_max = i
                        empty_rows_count = 0
                    else:
                        empty_rows_count += 1
                        if empty_rows_count >= EMPTY_ROWS_THRESHOLD:
                            break

                result[ws.title] = real_max

    except Exception as e:
        logging.warning(f"Error calculating real_max_rows: {e}")
        return {}
    finally:
        if wb_readonly:
            wb_readonly.close()

    return result


def detect_excel_import_type(excel_file_path):
    """
    Detects the type of Excel import based on its contents
    """
    try:
        wb = load_workbook(filename=excel_file_path, read_only=True)

        has_contract_id = False
        has_admin_section = False
        has_unit_price_section = False

        for sheet in wb.worksheets:
            first_row = next(sheet.rows, [])
            headers = [cell.value.lower() if cell.value else "" for cell in first_row]

            if "identificador do objeto" in headers:
                has_contract_id = True

            if any(
                "secao de administracao" in header or "seção de administração" in header
                for header in headers
            ):
                has_admin_section = True

            if any(
                "secao de preco unitario" in header
                or "seção de preço unitário" in header
                for header in headers
            ):
                has_unit_price_section = True

        if has_contract_id:
            if has_admin_section and has_unit_price_section:
                return "CONTRACT_ITEMS"
            elif has_admin_section:
                return "CONTRACT_ITEM_ADMINISTRATION"
            elif has_unit_price_section:
                return "CONTRACT_ITEM_UNIT_PRICE"

        return "REPORTING"
    except Exception:
        return


def get_object_path(id: str, object_name):
    return f"media/private/{id}_{object_name}"


def upload_progress(s3, id: str, data_file: str, data: dict) -> str:
    """Uploads partial importing data and returns file name"""
    temp_dir = tempfile.mkdtemp()
    if data_file is None:
        data_file = f"import_{get_random_string()}.json"
    local_data_file = f"{temp_dir}/{data_file}"
    with open(local_data_file, "w") as file:
        json.dump(data, file, indent=None)
    upload_file(s3, local_data_file, get_object_path(id, data_file))
    return data_file


def load_progress(s3, id: str, data_file: str):
    """Load partial importing data"""
    progress_data = None
    temp_dir = tempfile.mkdtemp()
    local_data_file = f"{temp_dir}/{data_file}"
    s3.download_file(
        settings.AWS_STORAGE_BUCKET_NAME,
        get_object_path(id, data_file),
        local_data_file,
    )
    with open(local_data_file, "r") as file:
        progress_data = json.load(file)
    return progress_data


def upload_image_from_zip(
    s3,
    excel_import: ExcelImport,
    file_path: str,
    object_name: str,
    expiration: datetime,
) -> None:
    bucket = settings.AWS_STORAGE_BUCKET_NAME
    try:
        s3.upload_file(
            file_path, bucket, object_name, ExtraArgs={"Expires": expiration}
        )
    except Exception as e:
        logging.error(f"Failed to upload {file_path} to {bucket}/{object_name}: {e}")
        sentry_sdk.capture_exception(e)

        excel_import.error = True
        excel_import.uploading_zip_images = False
        excel_import.save()
        raise e


def upload_zip_images(images_dir: str, excel_import: ExcelImport) -> None:
    client_config = botocore.config.Config(
        max_pool_connections=ZIP_PICTURES_THREADING_LIMIT,
    )
    s3 = boto3.client(
        "s3",
        aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
        aws_session_token=credentials.AWS_SESSION_TOKEN,
        config=client_config,
    )
    thread_pool = ThreadPoolExecutor(ZIP_PICTURES_THREADING_LIMIT)
    expiration = datetime.now().replace(tzinfo=pytz.UTC) + timedelta(hours=6)
    for root, _, files in os.walk(images_dir):
        for file in files:
            file_path = os.path.join(root, file)
            object_name = f"media/private/{str(excel_import.uuid)}_{file}"
            thread_pool.submit(
                upload_image_from_zip,
                s3,
                excel_import,
                file_path,
                object_name,
                expiration,
            )
    thread_pool.shutdown()


def download_excel_import_zip(excel_import: ExcelImport, zip_path: str) -> bool:
    s3 = boto3.client(
        "s3",
        aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
        aws_session_token=credentials.AWS_SESSION_TOKEN,
    )

    try:
        unquoted_file_path = parse.unquote(excel_import.zip_file.url)
        key = unquoted_file_path.split("?")[0].split(".com/")[1]
        bucket_name = unquoted_file_path.split(".s3")[0].split("/")[-1]
    except Exception as e:
        logging.error(f"Error parsing S3 URL: {str(e)}")
        sentry_sdk.capture_exception(e)
        return False

    try:
        s3.download_file(bucket_name, key, zip_path)
    except Exception as e:
        logging.error(f"Error downloading from S3: {str(e)}")
        sentry_sdk.capture_exception(e)
        return False
    return True


@task
def upload_zip_import_images(excel_import_id: str) -> bool:
    excel_import = None
    error = True
    zip_temp_dir = tempfile.mkdtemp()
    images_temp_dir = tempfile.mkdtemp()

    try:
        excel_import = ExcelImport.objects.get(pk=excel_import_id)
    except Exception as e:
        sentry_sdk.capture_exception(e)
        logging.error("Error getting ExcelImport object")
        return False

    zip_path = os.path.join(zip_temp_dir, f"{excel_import.uuid}.zip")

    downloaded = download_excel_import_zip(excel_import, zip_path)
    if downloaded:
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(images_temp_dir)

        upload_zip_images(images_temp_dir, excel_import)

        error = False

    if excel_import:
        excel_import.error = error
        excel_import.uploading_zip_images = False
        excel_import.save()
        logging.info("Zip Upload Done")

    return True


def use_old_import(company_id):
    try:
        company = Company.objects.get(pk=company_id)
        use_old_import = get_obj_from_path(
            company.metadata, "use_old_import", default_return=False
        )
        return use_old_import
    except Exception as e:
        sentry_sdk.capture_exception(e)
        logging.error("Error getting company")
        return False


@task
def parse_excel_to_json(excel_import_id, user_id, inventory_code):
    try:
        excel_import = ExcelImport.objects.get(pk=excel_import_id)
        user = User.objects.get(pk=user_id)

        if excel_import.excel_file:
            file_temp_path = shared_download_excel_file(
                excel_import, "/tmp/excel_imports/"
            )

            if file_temp_path:
                detected_type = detect_excel_import_type(file_temp_path)
                if detected_type == "CONTRACT_ITEM_UNIT_PRICE":
                    if os.path.exists(file_temp_path):
                        os.remove(file_temp_path)

                    generate_preview(excel_import_id, user_id)
                    return True
                elif detected_type == "CONTRACT_ITEM_ADMINISTRATION":
                    if os.path.exists(file_temp_path):
                        os.remove(file_temp_path)
                    generate_administration_preview(excel_import_id, user_id)
                    return True
                elif detected_type == "CONTRACT_ITEMS":
                    if os.path.exists(file_temp_path):
                        os.remove(file_temp_path)
                    generate_contract_items_preview(excel_import_id, user_id)
                    return True

                if os.path.exists(file_temp_path):
                    os.remove(file_temp_path)

        if use_old_import(excel_import.company_id):
            excel_import = ImportExcel(
                excel_import, user, inventory_code
            ).get_excel_import()
        else:
            ImportExcel(excel_import, user, inventory_code).import_excel()
            return
    except Exception as e:
        sentry_sdk.capture_exception(e)
        logging.error("Error getting objects")

    excel_import.generating_preview = False
    excel_import.save()
    logging.info("Parse Done")
    return True


@task
def parse_excel_part_to_json(
    excel_import_id,
    user_id,
    inventory_code,
    sheet_title,
    starting_row,
    idx,
    parts_count,
    rows_per_part,
):
    excel_import = ExcelImport.objects.get(pk=excel_import_id)
    user = User.objects.get(pk=user_id)
    excel_import = ImportExcel(excel_import, user, inventory_code).import_excel_part(
        sheet_title, starting_row, idx, parts_count, rows_per_part
    )


@task
def run_save_with_signals(reporting_id, is_edit=False):
    item: Reporting = Reporting.objects.get(uuid=reporting_id)

    if not is_edit:
        item._history_user = item.created_by
        item.history.bulk_history_create([item])

    # NOTE: Since we are not running the reporting_create() signal, it's important
    # to run here the logic relevant for the import (setting the Reporting.road)
    if item.road_name and not item.road:
        _, item.road = get_road_coordinates(
            item.road_name,
            item.km,
            item.direction,
            item.company,
        )

        # Se não encontrou road e existe road_name, cria road clone com trecho padrão
        if not item.road and item.road_name:
            # Busca rodovias com o mesmo nome (sem validação de KM)
            roads = Road.objects.filter(
                name=item.road_name,
                direction=int(item.direction),
                company=item.company,
            )

            if not roads.exists():
                # Busca sem direção específica
                roads = Road.objects.filter(
                    name=item.road_name, company=item.company
                ).order_by("direction")

            if roads.exists():
                # Encontrou rodovia mas o KM está fora do range
                road = roads.first()

                # Verifica se a rodovia não tem lot_logic e precisa de trecho padrão
                if should_add_default_marks(road):
                    # Cria nova road clone com marcos padrão
                    item.road = create_default_segment_road(road, item.company)

    item.save()

    # Auto-scheduling para novos reportings importados
    if not is_edit:
        from helpers.apps.auto_scheduling import process_auto_scheduling

        process_auto_scheduling(item)


@task
def update_job_after_bulk_edit(job_uuid):
    try:
        job = Job.objects.get(uuid=job_uuid)
        with DisableSignals(disabled_signals=[pre_init, post_init]):
            job.save()
    except Job.DoesNotExist:
        pass


def update_reporting_file_upload(
    s3,
    uuid,
    img_url,
    dir,
    errors: List[Exception],
    rfs: List[ReportingFile],
):
    try:
        parsed = parse.urlparse(img_url)
        bucket = settings.AWS_STORAGE_BUCKET_NAME
        basename = parse.urlparse(img_url).path.split("/")[-1]

        key = parsed.path.lstrip("/")
        if key.startswith(bucket + "/"):
            key = key[len(bucket) + 1 :]
        os.makedirs(f"{dir}/{uuid}", exist_ok=True)
        file = io.BytesIO()
        s3.download_fileobj(bucket, key, file)
        file.seek(0)

        reporting_file = ReportingFile.objects.get(pk=uuid)
        if reporting_file.md5 == "":
            md5 = hashlib.md5(file.read()).hexdigest()
            file.seek(0)
            reporting_file.md5 = md5

        reporting_file.upload = SimpleUploadedFile(basename, file.read())
        rfs.append(reporting_file)
    except Exception as e:
        errors.append(e)


def create_excel_reportings(excel_reportings: List[ExcelReporting], operation: str):
    if not excel_reportings:
        return

    keys_to_check = [
        (str(er.reporting.uuid), str(er.excel_import.uuid), operation)
        for er in excel_reportings
    ]

    existing_keys = set(
        ExcelReporting.objects.filter(
            reporting__uuid__in=[k[0] for k in keys_to_check],
            excel_import__uuid__in=[k[1] for k in keys_to_check],
            operation=operation,
        ).values_list("reporting__uuid", "excel_import__uuid", "operation")
    )

    existing_keys = [
        (str(existing_r), str(existing_import), existing_op)
        for existing_r, existing_import, existing_op in existing_keys
    ]
    to_create = []
    for er in excel_reportings:
        key = (er.reporting.uuid, er.excel_import.uuid, operation)
        if key not in existing_keys:
            to_create.append(er)

    if to_create:
        with DisableSignals():
            ExcelReporting.objects.bulk_create(to_create, batch_size=500)


def update_form_data(
    is_array_edit: bool,
    curr_form_data: dict,
    incominng_form_data: dict,
    form_metadata: dict,
):
    for form_k, form_v in incominng_form_data.items():
        # form_data won't be a direct assignment sometimes
        curr_v = curr_form_data.get(form_k)

        # Deal with array of objects cases
        are_lists = isinstance(form_v, list) and isinstance(curr_v, list)
        is_array_of_objs = (
            form_v
            and isinstance(form_v, list)
            and len(form_v)
            and isinstance(form_v[0], dict)
        )
        if are_lists and is_array_of_objs:
            if is_array_edit:
                v_it = iter(form_v)
                for curr_value in curr_form_data[form_k]:
                    new_values = next(v_it, {})
                    for (
                        api_name,
                        value,
                    ) in new_values.items():
                        curr_value[api_name] = value

                for new_values in v_it:
                    if new_values != {}:
                        curr_form_data[form_k].append(new_values)

            else:
                curr_form_data[form_k].extend(form_v)
        # Otherwise we'll do direct assignment
        else:
            if is_array_edit and isinstance(form_v, list):
                form_v = list(
                    filter(
                        lambda fields: fields != {},
                        form_v,
                    )
                )
            if form_metadata and form_k in form_metadata:
                old_value = curr_form_data.get(form_k)
                if old_value != form_v:
                    form_metadata[form_k] = {"manually_specified": True}
            curr_form_data[form_k] = form_v


def update_reporting_instance(reporting: Reporting, data: dict):
    for field, value in data.items():
        if value is not None and field != "form_data":
            setattr(reporting, field, value)


def create_reportings(reportings: List[Reporting], is_edit: bool, imported_by: User):
    if not is_edit:
        null_uuid_reportings = []
        uuid_to_reporting = {}
        for r in reportings:
            if r.uuid is None:
                null_uuid_reportings.append(r)
            else:
                uuid_to_reporting[str(r.uuid)] = r
        null_uuid_reportings = set(null_uuid_reportings)
        uuids = uuid_to_reporting.keys()
        existing_uuids = set(
            [
                str(uuid)
                for uuid in Reporting.objects.filter(uuid__in=uuids).values_list(
                    "uuid", flat=True
                )
            ]
        )
        missing_uuids = uuids - existing_uuids
        missing_reportings = set(
            [
                uuid_to_reporting.get(str(uuid))
                for uuid in missing_uuids
                if uuid_to_reporting.get(str(uuid)) is not None
            ]
        )
        reportings = missing_reportings.union(null_uuid_reportings)

        with DisableSignals():
            if len(missing_reportings) > 0:
                Reporting.objects.bulk_create(missing_reportings, batch_size=500)

    else:
        with DisableSignals():
            bulk_update_with_history(
                objs=reportings,
                model=reportings[0]._meta.model,
                user=imported_by,
                use_django_bulk=True,
            )

        unique_job_ids = list({str(r.job.uuid) for r in reportings if r.job})

        if unique_job_ids:
            for job_id in unique_job_ids:
                update_job_after_bulk_edit(job_id)

    for r in reportings:
        run_save_with_signals(str(r.uuid), is_edit=is_edit)


def create_reporting_files(reporting_files: List[ReportingFile]):
    missing_rfs = reporting_files

    null_uuid_rfs = set(rf for rf in reporting_files if rf.uuid is None)
    uuid_to_rf = {str(rf.uuid): rf for rf in reporting_files if rf.uuid is not None}
    uuids = uuid_to_rf.keys()
    existing_uuids = set(
        [
            str(uuid)
            for uuid in ReportingFile.objects.filter(uuid__in=uuids).values_list(
                "uuid", flat=True
            )
        ]
    )
    missing_uuids = uuids - existing_uuids
    missing_rfs = set(
        [
            uuid_to_rf.get(str(uuid))
            for uuid in missing_uuids
            if uuid_to_rf.get(str(uuid)) is not None
        ]
    )
    missing_rfs = missing_rfs.union(null_uuid_rfs)

    with DisableSignals():
        if len(missing_rfs) > 0:
            ReportingFile.objects.bulk_create(missing_rfs, batch_size=500)


def create_procedure_resources(procedure_resources: List[ProcedureResource]):
    procedure_resources = [pr for pr in procedure_resources if pr is not None]
    with DisableSignals():
        ProcedureResource.objects.bulk_create(procedure_resources, batch_size=500)


@task
def save_reporting_files(excel_import_id, data_file, processed=0):
    # Add images to ReportingFiles
    error = False
    try:
        excel_import = ExcelImport.objects.get(pk=excel_import_id)
    except Exception as e:
        sentry_sdk.capture_exception(e)
        logging.error("Error getting objects")
        error = True
    else:
        try:
            s3 = boto3.client(
                "s3",
                aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
                aws_session_token=credentials.AWS_SESSION_TOKEN,
            )
            images_urls = load_progress(s3, excel_import_id, data_file)
            batch_size = int(len(images_urls) / RF_BATCHES)
            batch_size = max(batch_size, MIN_RF_BATCH_SIZE)
            uuids = sorted([uuid for uuid in images_urls])[
                processed : processed + batch_size
            ]
            if len(uuids) > 0:
                thread_pool = ThreadPoolExecutor(THREADING_LIMIT)
                errors: List[Exception] = []
                rfs: List[ReportingFile] = []
                dir = tempfile.mkdtemp()
                for file_id in uuids:
                    img_url = images_urls[file_id]
                    if len(errors) > 0:
                        raise errors[0]
                    thread_pool.submit(
                        update_reporting_file_upload,
                        s3,
                        file_id,
                        img_url,
                        dir,
                        errors,
                        rfs,
                    )

                thread_pool.shutdown()

                if len(errors) > 0:
                    raise errors[0]

                fields = [
                    f.name for f in ReportingFile._meta.fields if not f.primary_key
                ]
                ReportingFile.objects.bulk_update(rfs, fields, batch_size=500)

                processed += len(uuids)
                if processed < len(images_urls):
                    save_reporting_files(excel_import_id, data_file, processed)
                    return True

        except Exception as e:
            logging.error(str(e))
            sentry_sdk.capture_exception(e)
            error = True

    # Set done to True
    excel_import.error = error
    excel_import.done = True
    excel_import.save()

    logging.info("Parse JSON Done")
    return True


@task
def parse_json_to_objs(excel_import_id):
    try:
        excel_import = ExcelImport.objects.get(pk=excel_import_id)
        if excel_import.preview_file:
            excel_import.preview_file.seek(0)
            preview_content = excel_import.preview_file.read()
            preview_data = json.loads(preview_content)
            if (
                isinstance(preview_data, dict)
                and "contract_id" in preview_data
                and "contract_items_unit_price" in preview_data
                and "contract_items_administration" in preview_data
            ):
                return execute_contract_items_import(excel_import_id)
            elif (
                isinstance(preview_data, dict)
                and "contract_id" in preview_data
                and "contract_items_unit_price" in preview_data
            ):
                return execute_import(excel_import_id)
            elif (
                isinstance(preview_data, dict)
                and "contract_id" in preview_data
                and "contract_items_administration" in preview_data
            ):
                return execute_administration_import(excel_import_id)
            excel_import.preview_file.seek(0)
    except Exception as e:
        sentry_sdk.capture_exception(e)
        logging.error("Error getting objects")
    else:
        date_fields = [
            "found_at",
            "created_at",
            "executed_at",
            "due_at",
            "datetime",
            "uploaded_at",
            "creation_date",
        ]
        error = True
        images_urls = {}
        reporting_instances: List[Reporting] = []
        excel_reportings: List[ExcelReporting] = []
        processed_items = 0
        company = excel_import.company
        imported_by = excel_import.created_by
        show_track = get_obj_from_path(company.metadata, "show_track")
        if show_track:
            branch_display = get_obj_from_path(
                company.custom_options, "reporting__fields__branch__display_if__track"
            )
            km_reference_display = get_obj_from_path(
                company.custom_options,
                "reporting__fields__km_reference__display_if__track",
            )
        # Get data
        try:
            data = json.loads(excel_import.preview_file.read())
        except Exception as e:
            sentry_sdk.capture_exception(e)
            data = {}

        # Clean the json data
        data = clean_invalid_characters(data)

        if data:
            # Change all keys to snake_case
            data = dict_to_casing(data, format_type="underscore")
            is_array_edit = data["is_array_edit"]
            is_edit = data.get("is_edit", False)
            operation = "EDIT" if is_edit else "CREATE"

            try:
                reportings = data.get("reportings", [])

                # Prefetch deadlines if they are going to be used
                form_ids = set(
                    item["occurrence_type_id"]
                    for item in reportings
                    if item.get("occurrence_type_id") and item.get("found_at")
                )
                form_deadlines = {}
                if form_ids:
                    forms = OccurrenceType.objects.filter(
                        uuid__in=form_ids
                    ).values_list("uuid", "deadline")
                    form_deadlines = {str(a): b for a, b in forms}

                # Prefetch the necessary reportings for an edit
                id_to_rep = {}
                if is_edit:
                    rep_ids = set(item.get("uuid") for item in reportings)
                    rep_instances = Reporting.objects.filter(uuid__in=rep_ids)
                    id_to_rep = {str(rep.pk): rep for rep in rep_instances}

                len_objs = len(reportings) * 2
                for item in reportings:
                    row = str(item.pop("row", ""))

                    # Initial data prep
                    parsed_item = {
                        k: v
                        for k, v in item.items()
                        # NOTE: This will skip fields with None for edits
                        if k != "column_errors" and (is_edit is False or v is not None)
                    }
                    parsed_item = parse_dict_dates(parsed_item, date_fields)

                    # Handle km inputs
                    end_km = parsed_item.get("end_km")
                    if end_km is not None:
                        parsed_item["end_km_manually_specified"] = True
                    elif "end_km" in parsed_item and end_km is None:
                        parsed_item["end_km"] = parsed_item["km"]

                    project_km = parsed_item.get("project_km")
                    project_end_km = parsed_item.get("project_end_km")
                    if project_km is not None:
                        if project_end_km is not None:
                            parsed_item["project_end_km_manually_specified"] = True
                        elif is_edit is False and project_end_km is None:
                            parsed_item["project_end_km"] = parsed_item["project_km"]
                    elif "project_km" in parsed_item and project_km is None:
                        try:
                            del parsed_item["project_km"]
                            del parsed_item["project_end_km"]
                        except Exception:
                            pass

                    # Handle date inputs
                    if parsed_item.get("due_at"):
                        parsed_item["due_at_manually_specified"] = True

                    # Autofill due_at if possible
                    elif parsed_item.get("found_at") and is_edit is False:
                        try:
                            form_deadline = form_deadlines[
                                parsed_item["occurrence_type_id"]
                            ]
                            if form_deadline:
                                parsed_item["due_at"] = (
                                    parsed_item["found_at"] + form_deadline
                                )
                        except Exception as e:
                            logging.error("erro ao buscar formulário " + str(e))

                    # Handle track requirements
                    if "track" in parsed_item and "branch" in parsed_item:
                        if parsed_item["track"] not in branch_display:
                            del parsed_item["branch"]
                    if "track" in parsed_item and "km_reference" in parsed_item:
                        if parsed_item["track"] not in km_reference_display:
                            del parsed_item["km_reference"]

                    # Handle form data with array of objects
                    if parsed_item.get("form_data"):
                        form_data_cp = parsed_item["form_data"].copy()
                        for k, v in form_data_cp.items():
                            if (
                                isinstance(v, list)
                                and v != []
                                and isinstance(v[0], dict)
                            ):
                                for inner_position in v:
                                    inner_position.pop("index_array")
                                    for (
                                        inner_key,
                                        inner_value,
                                    ) in inner_position.copy().items():
                                        if inner_value == []:
                                            inner_position.pop(inner_key)
                                if not is_array_edit:
                                    v = list(filter(lambda x: x != {}, v))
                                parsed_item["form_data"][k] = v
                                if v == []:
                                    parsed_item["form_data"].pop(k)
                            elif isinstance(v, list) and v == []:
                                parsed_item["form_data"].pop(k)

                    # Handle inventory data
                    if "import_inventory_code" in parsed_item:
                        inv_data = parsed_item.get("import_inventory_code", {})
                        inv_uuid = inv_data.get("uuid")

                        if inv_uuid:
                            parsed_item["parent_id"] = inv_uuid
                        del parsed_item["import_inventory_code"]
                    if "inventory_value" in parsed_item:
                        del parsed_item["inventory_value"]

                    # Handle coordinates
                    if "latitude" in parsed_item:
                        del parsed_item["latitude"]
                    if "longitude" in parsed_item:
                        del parsed_item["longitude"]
                    if "point" in parsed_item:
                        if parsed_item["point"]:
                            try:
                                point = Point(**parsed_item["point"], srid=4326)
                                parsed_item["point"] = point
                            except Exception:
                                del parsed_item["point"]
                            else:
                                parsed_item["manual_geometry"] = True
                                parsed_item["geometry"] = GeometryCollection(point)
                        else:
                            del parsed_item["point"]

                    # Instantiate and validate the ReportingSerializer and its respective ExcelReporting
                    instance = None
                    if is_edit:
                        try:
                            item_uuid = item.get("uuid")
                            instance = id_to_rep[item_uuid]
                        except Exception:
                            # NOTE: We break since there's no point processing the next items
                            # because we require every instance to be valid with len_objs == len(objs)
                            break

                        serialized_item = ReportingSerializer(
                            instance=instance, data=parsed_item, partial=True
                        )
                    else:
                        serialized_item = ReportingSerializer(data=parsed_item)

                    excel_reporting = {
                        "reporting_id": parsed_item.get("uuid", ""),
                        "excel_import_id": excel_import_id,
                        "row": row,
                        "operation": operation,
                    }
                    serialized_excel_reporting = ExcelReportingSerializer(
                        data=excel_reporting
                    )

                    if (
                        serialized_item.is_valid()
                        and serialized_excel_reporting.is_valid()
                    ):
                        # Update the existing instance
                        incoming_form_data = parsed_item.get("form_data")
                        if is_edit:
                            if incoming_form_data is not None:
                                if not isinstance(instance.form_data, dict):
                                    instance.form_data = incoming_form_data
                                else:
                                    update_form_data(
                                        is_array_edit,
                                        instance.form_data,
                                        incoming_form_data,
                                        instance.form_metadata,
                                    )
                            update_reporting_instance(instance, parsed_item)
                            reporting_instances.append(instance)
                        else:
                            reporting_instances.append(Reporting(**parsed_item))

                        # Create a new instance
                        excel_reportings.append(ExcelReporting(**excel_reporting))

                        processed_items += 1
                    else:
                        # NOTE: We break since there's no point processing the next items
                        # because we require every instance to be valid with len_objs == len(objs)
                        break

                # Create reporting files drafts
                # Only process reporting files if all previous items were valid
                processed_items = len(reporting_instances) + len(excel_reportings)

                reporting_files: List[ReportingFile] = []
                if len_objs == processed_items:
                    images = data.get("images", [])
                    len_objs += len(images)
                    for item in images:
                        item.pop("row", "")
                        item.pop("format", "")
                        item.pop("file_name", "")
                        item["upload"] = {"filename": item.get("upload", "")}
                        parsed_item = parse_dict_dates(item, date_fields)

                        serialized_item = ReportingFileSerializer(data=parsed_item)
                        if serialized_item.is_valid():
                            rf_uuid = parsed_item.get("uuid", "")
                            upload = parsed_item.pop("upload", {})
                            images_urls[rf_uuid] = upload.get("filename", "")
                            reporting_files.append(ReportingFile(**parsed_item))

                    processed_items += len(reporting_files)

                procedure_resources: List[ProcedureResource] = []
                # Create procedure resources drafts
                # Only process procedure resources if all previous items were valid
                if len_objs == processed_items:
                    resources = data.get("resources", [])
                    len_objs += len(resources)
                    for item in resources:
                        parsed_item = parse_dict_dates(item, date_fields)

                        serialized_item = ProcedureResourceSerializer(data=parsed_item)
                        if serialized_item.is_valid():
                            procedure_resources.append(ProcedureResource(**parsed_item))

                    processed_items += len(procedure_resources)

                # Create all objects
                # Only create if all previous items were valid
                if len_objs == processed_items:
                    create_reportings(reporting_instances, is_edit, imported_by)

                    create_excel_reportings(excel_reportings, operation)

                    create_reporting_files(reporting_files)

                    create_procedure_resources(procedure_resources)

                    s3 = boto3.client(
                        "s3",
                        aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
                        aws_session_token=credentials.AWS_SESSION_TOKEN,
                    )
                    data_file = upload_progress(s3, excel_import_id, None, images_urls)
                    save_reporting_files(excel_import_id, data_file)
                    return True
            except Exception as e:
                logging.error(str(e))
                sentry_sdk.capture_exception(e)

        # Set done to True
        excel_import.error = error
        excel_import.done = True
        excel_import.save()

        logging.info("Parse JSON Done")
    return True


def upload_image(s3, uuid, count_images, image_dict, image, part_index: int = 0):
    try:
        image_name = "{}-{}-{}.{}".format(uuid, part_index, count_images, image.format)

        bucket_name = settings.AWS_STORAGE_BUCKET_NAME
        expires = timezone.now() + timedelta(hours=6)
        object_name = "media/private/{}".format(image_name)

        image.ref.seek(0)

        s3.upload_fileobj(
            image.ref, bucket_name, object_name, ExtraArgs={"Expires": expires}
        )
        url_s3 = s3.generate_presigned_url(
            "get_object", Params={"Bucket": bucket_name, "Key": object_name}
        )
    except Exception as e:
        sentry_sdk.capture_exception(e)
        url_s3 = ""

    image_dict["upload"] = url_s3


class ImportExcel:
    """
    Rules:

    1) All datetime objects must use function to_utc_string
    """

    temp_path = "/tmp/excel_import/"

    def __init__(self, excel_import, user, inventory_code):
        global REQUIRED_FIELDS
        self.file_name = ""
        self.uuid = str(excel_import.pk)
        self.excel_import = excel_import
        self.company_id = str(excel_import.company_id)
        self.company = excel_import.company
        self.user_id = str(user.uuid)
        self.request_user = user
        self.created_at = to_utc_string(datetime.now())
        self.count_images = 0
        self.ARRAY_USED_POSITIONS = defaultdict(list)
        self.original_inventory_code = inventory_code
        self.inventory_code = inventory_code
        self.is_edit = False
        self.number_to_reporting = defaultdict(list)
        self.occ_types: Dict[str, OccurrenceType] = {}
        self.firms: Dict[str, Firm] = {}
        self.road_names: Dict[str, str] = {}
        self.status_uuids: Dict[str, str] = {}
        self.resources: Dict[Firm, dict] = {}
        self.jobs: Dict[str, str] = {}
        self.road_objects: Dict[str, Road] = {}
        self.part_index = 0

        if self.inventory_code and self.inventory_code not in ["uuid", "number"]:
            self.inventory_code = "form_data__{}".format(
                to_snake_case(self.inventory_code)
            )
        self.inventory_cache: Dict[str, dict] = {}
        self.required_fields_import = [
            field for field in REQUIRED_FIELDS if field != "company_id"
        ]
        # Set default approval_step
        approval_step = ApprovalStep.objects.filter(
            approval_flow__company=self.company_id,
            approval_flow__target_model="reportings.Reporting",
            previous_steps__isnull=True,
        ).first()
        if approval_step:
            self.approval_step_id = str(approval_step.uuid)
        else:
            self.approval_step_id = None

        self.hide_reporting_location = get_obj_from_path(
            self.company.metadata, "hide_reporting_location", default_return=False
        )
        self.show_coordinate_input = get_obj_from_path(
            self.company.metadata, "show_coordinate_input", default_return=False
        )
        if self.hide_reporting_location:
            remove_required_fields = ["km", "road_name", "direction", "lane"]
            self.required_fields_import.append("latitude")
            self.required_fields_import.append("longitude")
            REQUIRED_FIELDS = [
                field
                for field in REQUIRED_FIELDS
                if field not in remove_required_fields
            ]
        else:
            self.required_fields_import.append("km")
            self.required_fields_import.append("road_name")

        # Prepare all the data needed for menu parsing with one query
        suitable_menus = RecordMenuRelation.objects.filter(
            user=self.user_id,
            company=self.company_id,
            record_menu__system_default=False,
        ).values_list("record_menu__name", "record_menu", "hide_menu")
        self.visible_menus = {
            menu_name.lower(): str(menu_id)
            for menu_name, menu_id, hide_menu in suitable_menus
            if not hide_menu
        }
        self.hidden_menus = {
            menu_name.lower(): str(menu_id)
            for menu_name, menu_id, hide_menu in suitable_menus
            if hide_menu
        }
        self.fallback_menu = (
            next(iter(self.visible_menus.values()))
            if len(self.visible_menus) == 1
            else None
        )

        # Is the mass edit import feature be used?
        self.can_use_edit = self.company.metadata.get("allow_mass_edit_import") is True
        self.image_kinds = self.get_image_kinds()

    def get_image_kinds(self):
        possible_image_path = "reportingfile__fields__kind__selectoptions__options"
        kinds = get_obj_from_path(self.company.custom_options, possible_image_path)
        kind_translation = {
            clean_latin_string(item["name"]).lower(): item["value"]
            for item in kinds
            if "value" in item and "name" in item
        }
        return kind_translation

    def load_data(self):
        self.wb = shared_load_data(self.file_name)
        return

    def preload_inventory_cache(self, inventory_values: List) -> None:
        """
        Preloads all inventory_codes in a single query to optimize
        """
        if not self.inventory_code or not inventory_values:
            return

        valid_values = [v for v in inventory_values if v]
        if not valid_values:
            return

        inventory_filter = {f"{self.inventory_code}__in": valid_values}
        inventory_queryset = (
            Reporting.objects.filter(
                company=self.company_id,
                occurrence_type__occurrence_kind="2",
                **inventory_filter,
            )
            .prefetch_related("occurrence_type")
            .values(
                self.inventory_code,
                "occurrence_type__name",
                "number",
                "created_at",
                "road_name",
                "km",
                "uuid",
            )
        )

        for item in inventory_queryset:
            raw_inventory_value = item[self.inventory_code]

            if isinstance(raw_inventory_value, (int, float)):
                cache_key = raw_inventory_value
            else:
                cache_key = str(raw_inventory_value)

            if cache_key not in self.inventory_cache:
                self.inventory_cache[cache_key] = []
            self.inventory_cache[cache_key].append(item)

    def get_lane_and_direction(self):
        possible_lane_path = "reporting__fields__lane__selectoptions__options"
        lanes = get_obj_from_path(self.company.custom_options, possible_lane_path)
        self.lane_translation = {
            clean_latin_string(item["name"]).lower(): item["value"]
            for item in lanes
            if "value" in item and "name" in item
        }

        possible_direction_path = "reporting__fields__direction__selectoptions__options"
        directions = get_obj_from_path(
            self.company.custom_options, possible_direction_path
        )
        self.direction_translation = {
            clean_latin_string(item["name"]).lower(): item["value"]
            for item in directions
            if "value" in item and "name" in item
        }
        return

    def update_column_errors(self, reporting_dict, column_errors):
        return shared_update_column_errors(
            reporting_dict, column_errors, PROPERTY_TO_NAME
        )

    def format_row_location(self, row, num):
        return "{}:{}".format(row, num)

    def parse_km(self, value):
        if value and isinstance(value, str):
            try:
                value = float(value.replace("+", ".").replace(",", "."))
            except Exception:
                value = None
        elif value and not isinstance(value, (int, float)):
            value = None
        elif value == "":
            value = None
        return value

    def parse_road(self, value):
        if value:
            value = value.strip()
            road_name = self.road_names.get(value)
            if road_name is None:
                road = Road.objects.filter(company=self.company, name=value).first()

                # Se a road precisa de trecho padrão, cria clone
                if road and should_add_default_marks(road):
                    road = create_default_segment_road(road, self.company)

                road_name = road.name if road else None
                self.road_names[value] = road_name
                self.road_objects[value] = road
            self.road = self.road_objects.get(value)
            return road_name
        return None

    def parse_job(self, value):
        if value:
            value = value.strip()
            job_uuid = self.jobs.get(value)
            if job_uuid is None:
                job = Job.objects.filter(company=self.company, number=value).first()
                job_uuid = str(job.uuid) if job else None
                self.jobs[value] = job_uuid
            return job_uuid
        return None

    def parse_menu(self, value: str) -> Tuple[str, bool]:
        """
        Parse the menu entry according to the provided menu name.

        Args:
            value (str): Name of the menu

        Returns:
            Tuple[str, bool]: A tuple of the ID of the menu and whether that menu is visible or not
        """

        if value:
            value = value.strip().lower()

            if value in self.visible_menus:
                return self.visible_menus[value], True
            elif value in self.hidden_menus:
                return self.hidden_menus[value], False

        return (None, None)

    def parse_firm(self, row_dict):
        """
        Determine the Firm if provided, if not, attempt to use the reporting number if provided
        """

        firm_name = row_dict.get("firm_id", None)
        rep_number = row_dict.get("number", None)

        firm = None
        if isinstance(firm_name, str):
            firm_name = firm_name.strip()
            firm = self.firms.get(firm_name)

        if firm is None:
            query = {
                "company": self.company,
                "active": True,
            }

            if firm_name:
                query["name"] = firm_name
                firm = Firm.objects.filter(**query).first()
                if firm is not None:
                    self.firms[firm_name] = firm
            elif rep_number:
                query["reportings__number"] = rep_number.strip()
                firm = Firm.objects.filter(**query).first()

        return firm

    def parse_occurrence_type(self, row_dict):
        """
        Determine the OccurrenceType if provided, if not, attempt to use the reporting number if provided
        """

        occ_type_name = row_dict.get("occurrence_type_id", None)
        rep_number = row_dict.get("number", None)

        occ_type = None
        if isinstance(occ_type_name, str):
            occ_type_name = occ_type_name.strip()
            occ_type = self.occ_types.get(occ_type_name)

        if occ_type is None:
            query = {
                "company": self.company,
                "next_version__isnull": True,
                "active": True,
            }

            if occ_type_name:
                query["name"] = occ_type_name
                occ_type = OccurrenceType.objects.filter(**query).first()
                if occ_type is not None:
                    self.occ_types[occ_type_name] = occ_type
            elif rep_number:
                query["reporting_occurrence__number"] = rep_number.strip()
                occ_type = OccurrenceType.objects.filter(**query).first()

        return occ_type

    def parse_status(self, value):
        if value:
            value = value.strip()
            status_uuid = self.status_uuids.get(value)
            if status_uuid is None:
                status = ServiceOrderActionStatus.objects.filter(
                    companies=self.company, name=value
                ).first()
                status_uuid = str(status.uuid) if status else None
                self.status_uuids[value] = status_uuid
            return status_uuid
        return None

    def parse_track(self, value):
        if value:
            possible_track_path = "reporting__fields__track__selectoptions__options"
            tracks = get_obj_from_path(self.company.custom_options, possible_track_path)
            track_translation = {
                clean_latin_string(item["name"]).lower(): item["value"]
                for item in tracks
                if "value" in item and "name" in item
            }
            final_value = track_translation.get(value, "")
            return final_value
        return ""

    def parse_branch(self, value):
        if value:
            possible_branch_path = "reporting__fields__branch__selectoptions__options"
            branchs = get_obj_from_path(
                self.company.custom_options, possible_branch_path
            )
            branch_translation = {
                clean_latin_string(item["name"]).lower(): item["value"]
                for item in branchs
                if "value" in item and "name" in item
            }
            final_value = branch_translation.get(value, "")
            return final_value
        return ""

    def parse_coordinates(self, value):
        if value is None:
            return value
        elif isinstance(value, str):
            value = value.strip()
            try:
                value = float(value)
            except Exception:
                try:
                    deg, minutes, seconds, direction = re.split("[°'\"]", value)
                    value = (
                        float(deg)
                        + float(minutes) / 60
                        + float(seconds.replace(",", ".")) / (60 * 60)
                    ) * (-1 if direction in ["O", "W", "S"] else 1)
                except Exception:
                    return "erro"
                else:
                    return value
            else:
                return value
        elif isinstance(value, (float, int)):
            return value

    def parse_form_data(self, row_dict, occ_type, header):
        column_errors = []
        form_data = {}
        if occ_type and row_dict:
            array_pattern = re.compile(r"[\w ,-]+_\d+: [\w ,]+")
            array_fields = list(filter(array_pattern.match, header))
            fields = occ_type.form_fields.get("fields", [])
            for field in fields:
                field = keys_to_snake_case(field)
                api_name = to_snake_case(field.get("api_name", ""))
                data_type = to_snake_case(field.get("data_type", ""))
                display_name = field.get("import_name", "") or field.get(
                    "display_name", ""
                )
                field_value = row_dict.get(display_name, None)

                if data_type in ["select_multiple", "select"]:
                    select_options = keys_to_snake_case(field.get("select_options", {}))
                    if "reference" in select_options:
                        select_options = keys_to_snake_case(
                            select_options.get("reference", {})
                        )
                        if select_options:
                            resource = select_options.get("resource", "")
                            if resource == "OccurrenceType" and field_value:
                                try:
                                    if data_type == "select_multiple":
                                        field_value = field_value.split(",")
                                        value = [
                                            str(
                                                OccurrenceType.objects.get(
                                                    company=self.company,
                                                    name=a.strip(),
                                                    next_version__isnull=True,
                                                    active=True,
                                                ).uuid
                                            )
                                            for a in field_value
                                        ]
                                        form_data[api_name] = value
                                    else:
                                        value = OccurrenceType.objects.get(
                                            company=self.company,
                                            name=field_value.strip(),
                                            next_version__isnull=True,
                                            active=True,
                                        )
                                        form_data[api_name] = str(value.uuid)
                                except Exception:
                                    column_errors.append(display_name)
                    else:
                        select_options = select_options.get("options", [])
                        options = {
                            clean_latin_string(item["name"]).lower(): item["value"]
                            for item in select_options
                            if "value" in item and "name" in item
                        }
                        if field_value is not None:
                            value = clean_latin_string(str(field_value)).lower()
                            if data_type == "select_multiple":
                                option_value = [
                                    options.get(temp_value.strip(), "")
                                    for temp_value in value.split(",")
                                ]
                                is_valid = all(option_value)
                            else:
                                option_value = options.get(value.strip(), "")
                                is_valid = bool(option_value)

                            if is_valid:
                                form_data[api_name] = option_value
                            else:
                                column_errors.append(display_name)
                elif data_type == "float":
                    if isinstance(field_value, (int, float)):
                        form_data[api_name] = float(field_value)
                    elif field_value:
                        column_errors.append(display_name)
                elif data_type == "number":
                    if isinstance(field_value, (int, float)):
                        form_data[api_name] = int(field_value)
                    elif field_value:
                        column_errors.append(display_name)
                elif data_type in ["text_area", "license_plate", "string"]:
                    if field_value:
                        form_data[api_name] = str(field_value)
                elif data_type == "boolean":
                    if isinstance(field_value, bool):
                        form_data[api_name] = field_value
                    elif isinstance(field_value, str):
                        if field_value.lower() in ["sim", "true", "verdadeiro"]:
                            form_data[api_name] = True
                        else:
                            form_data[api_name] = False
                    elif field_value:
                        column_errors.append(display_name)
                elif data_type == "timestamp":
                    if isinstance(field_value, datetime):
                        aware_field_value = make_aware(
                            field_value, timezone=pytz.timezone(settings.TIME_ZONE)
                        )
                        form_data[api_name] = aware_field_value.isoformat()
                    elif isinstance(field_value, str):
                        try:
                            datetime_value = datetime.strptime(field_value, "%d/%m/%Y")
                            datetime_value = make_aware(
                                datetime_value,
                                timezone=pytz.timezone(settings.TIME_ZONE),
                            )
                        except ValueError:
                            column_errors.append(display_name)
                        else:
                            form_data[api_name] = datetime_value.isoformat()
                    elif field_value:
                        column_errors.append(display_name)
                elif data_type == "cpf":
                    if field_value is not None:
                        try:
                            field_value = validate_CPF(field_value)
                            field_value = format_cpf_brazilin(str(field_value))
                        except Exception:
                            column_errors.append(display_name)
                        finally:
                            form_data[api_name] = field_value
                elif data_type == "phone":
                    if field_value is not None:
                        try:
                            field_value = phone_validation(field_value)
                            if len(field_value) == 10:
                                field_value = format_phone_number_brazilin(
                                    str(field_value)
                                )
                            elif len(field_value) == 11:
                                field_value = format_mobile_number_brazilin(
                                    str(field_value)
                                )
                        except Exception:
                            column_errors.append(display_name)
                        finally:
                            form_data[api_name] = field_value

                elif data_type == "array_of_objects":
                    if array_fields:
                        inner_fields = field.get("inner_fields", [])
                        if inner_fields:
                            (parsed_array, column_errors) = self.parse_array_of_objects(
                                row_dict,
                                display_name,
                                api_name,
                                inner_fields,
                                array_fields,
                                column_errors,
                            )
                            if parsed_array:
                                form_data[api_name] = parsed_array

        row_dict["form_data"] = form_data
        # Update column_errors
        row_dict = self.update_column_errors(row_dict, column_errors)

        return row_dict

    def parse_array_of_objects(
        self,
        row_dict,
        display_name,
        api_name_array,
        inner_fields,
        array_fields,
        column_errors,
    ):
        ARRAY_COLUMNS[clean_latin_string(display_name).lower()] = api_name_array

        position_list = []
        for array_field in array_fields:
            position = int(array_field.split(":")[0].split("_")[-1].strip())
            position_list.append(position)

        parsed_array = [
            {"index_array": i + 1} for i in range(max(position_list, default=50))
        ]

        for array_field in array_fields:
            position = int(array_field.split(":")[0].split("_")[-1].strip())
            inner_name = array_field.split(":")[-1].strip().split(" - ")[0].strip()
            inner_value = row_dict.get(array_field, None)
            inner_dict = next(
                (
                    field
                    for field in inner_fields
                    if field.get("display_name") == inner_name
                    or field.get("displayName") == inner_name
                ),
                None,
            )
            if inner_dict:
                inner_dict = keys_to_snake_case(inner_dict)
                api_name = to_snake_case(inner_dict.get("api_name", ""))
                data_type = to_snake_case(inner_dict.get("data_type", ""))
                if inner_value is not None:
                    self.ARRAY_USED_POSITIONS[row_dict.get("row", "")].append(position)
                    if data_type in ["select_multiple", "select"]:
                        select_options = keys_to_snake_case(
                            inner_dict.get("select_options", {})
                        )
                        if "reference" in select_options:
                            select_options = keys_to_snake_case(
                                select_options.get("reference", {})
                            )
                            if select_options:
                                resource = select_options.get("resource", "")
                                if resource == "OccurrenceType" and inner_value:
                                    try:
                                        if data_type == "select_multiple":
                                            inner_value = inner_value.split(",")
                                            value = [
                                                str(
                                                    OccurrenceType.objects.get(
                                                        company=self.company,
                                                        name=a.strip(),
                                                        next_version__isnull=True,
                                                        active=True,
                                                    ).uuid
                                                )
                                                for a in inner_value
                                            ]
                                            parsed_array[position - 1][api_name] = value
                                        else:
                                            value = OccurrenceType.objects.get(
                                                company=self.company,
                                                name=inner_value.strip(),
                                                next_version__isnull=True,
                                                active=True,
                                            )
                                            parsed_array[position - 1][api_name] = str(
                                                value.uuid
                                            )
                                    except Exception:
                                        column_errors.append(array_field)

                        else:
                            select_options = select_options.get("options", [])
                            options = {
                                clean_latin_string(item["name"]).lower(): item["value"]
                                for item in select_options
                                if "value" in item and "name" in item
                            }
                            if inner_value is not None:
                                value = clean_latin_string(str(inner_value)).lower()
                                if data_type == "select_multiple":
                                    option_value = [
                                        options.get(temp_value.strip(), "")
                                        for temp_value in value.split(",")
                                    ]
                                    is_valid = all(option_value)
                                else:
                                    option_value = options.get(value.strip(), "")
                                    is_valid = bool(option_value)

                                if is_valid:
                                    parsed_array[position - 1][api_name] = option_value
                                else:
                                    column_errors.append(array_field)
                    elif data_type == "float":
                        if isinstance(inner_value, (int, float)):
                            parsed_array[position - 1][api_name] = float(inner_value)
                        elif inner_value:
                            column_errors.append(array_field)
                    elif data_type == "number":
                        if isinstance(inner_value, (int, float)):
                            parsed_array[position - 1][api_name] = int(inner_value)
                        elif inner_value:
                            column_errors.append(array_field)
                    elif data_type in ["text_area", "license_plate", "string"]:
                        parsed_array[position - 1][api_name] = str(inner_value)
                    elif data_type == "boolean":
                        if isinstance(inner_value, bool):
                            parsed_array[position - 1][api_name] = inner_value
                        elif isinstance(inner_value, str):
                            if inner_value.lower() in ["sim", "true", "verdadeiro"]:
                                parsed_array[position - 1][api_name] = True
                            else:
                                parsed_array[position - 1][api_name] = False
                        elif inner_value:
                            column_errors.append(array_field)
                    elif data_type == "timestamp":
                        if isinstance(inner_value, datetime):
                            aware_inner_value = make_aware(
                                inner_value, timezone=pytz.timezone(settings.TIME_ZONE)
                            )
                            parsed_array[position - 1][
                                api_name
                            ] = aware_inner_value.isoformat()
                        elif isinstance(inner_value, str):
                            try:
                                datetime_value = datetime.strptime(
                                    inner_value, "%d/%m/%Y"
                                )
                                datetime_value = make_aware(
                                    datetime_value,
                                    timezone=pytz.timezone(settings.TIME_ZONE),
                                )
                            except ValueError:
                                column_errors.append(array_field)
                            else:
                                parsed_array[position - 1][
                                    api_name
                                ] = datetime_value.isoformat()
                        elif inner_value:
                            column_errors.append(array_field)
                    elif data_type == "cpf":
                        try:
                            inner_value = validate_CPF(inner_value)
                            inner_value = format_cpf_brazilin(str(inner_value))
                        except Exception:
                            column_errors.append(array_field)
                        finally:
                            parsed_array[position - 1][api_name] = inner_value
                    elif data_type == "phone":
                        try:
                            inner_value = phone_validation(inner_value)
                            if len(inner_value) == 10:
                                inner_value = format_phone_number_brazilin(
                                    str(inner_value)
                                )
                            elif len(inner_value) == 11:
                                inner_value = format_mobile_number_brazilin(
                                    str(inner_value)
                                )
                        except Exception:
                            column_errors.append(array_field)
                        finally:
                            parsed_array[position - 1][api_name] = inner_value
                if data_type == "inner_images_array":
                    parsed_array[position - 1][api_name] = []
                    array_full_name = f"{clean_latin_string(display_name).lower()}***{clean_latin_string(inner_name).lower()}"
                    ARRAY_PHOTO_COLUMNS[array_full_name] = api_name
        return parsed_array, column_errors

    def translate_image_kind(self, value):
        if value:
            final_value = self.image_kinds.get(clean_latin_string(value).lower(), "")
            return final_value
        return ""

    def get_header(self, worksheet):
        try:
            header = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            header_map = [
                NAME_TO_PROPERTY[clean_latin_string(item).lower().strip()]
                if clean_latin_string(item).lower().strip() in NAME_TO_PROPERTY
                else item
                for item in header
            ]
        except Exception:
            header_map = []

        return header_map

    def is_image_column(self, header):
        return bool(re.search(IMG_PATTERN, header))

    def get_images(self, worksheet):
        images_list = {}
        if self.excel_import.zip_file:

            s3 = boto3.client(
                "s3",
                aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
                aws_session_token=credentials.AWS_SESSION_TOKEN,
            )
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            image_columns = {}
            for col_idx, header in enumerate(header_row, start=1):
                if isinstance(header, str) and self.is_image_column(header):
                    image_columns[col_idx] = header

            if image_columns:
                for row_idx, row in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True),
                    start=1,
                ):
                    for col_idx, header in image_columns.items():
                        try:
                            cell_value = (
                                row[col_idx - 1] if col_idx <= len(row) else None
                            )
                            if not cell_value:
                                continue
                            image_dict = {}
                            image_name = str(cell_value).strip()
                            if not image_name:
                                continue
                            img_num = "".join(header.split("_")[1:])
                            row_img_num = self.format_row_location(row_idx + 1, img_num)
                            file_name = f"{str(self.excel_import.uuid)}_{image_name}"
                            upload = f"media/private/{file_name}"
                            ext = image_name.split(".")[-1].lower()
                            exists = False
                            try:
                                exists = s3.head_object(
                                    Bucket=settings.AWS_STORAGE_BUCKET_NAME,
                                    Key=upload,
                                )
                            except Exception:
                                pass

                            ext = None
                            if exists:
                                ext = image_name.split(".")[-1].lower()

                            image_dict["md5"] = ""
                            image_dict["uuid"] = str(uuid.uuid4())
                            image_dict["format"] = ext
                            image_dict["created_by_id"] = self.user_id
                            image_dict[
                                "upload"
                            ] = f"{settings.AWS_STORAGE_BUCKET_NAME}/{upload}"
                            image_dict["uploaded_at"] = self.created_at
                            image_dict["file_name"] = image_name
                        except Exception:
                            pass
                        else:
                            images_list[row_img_num] = image_dict

        images = worksheet._images
        if images:
            thread_pool = ThreadPoolExecutor(UPLOAD_THREADING_LIMIT)
            for image in images:
                try:
                    image_dict = {}
                    header_value = worksheet.cell(
                        row=1, column=image.anchor._from.col + 1
                    ).value
                    img_num = "".join(str(header_value).split("_")[1:])

                    row_img_num = self.format_row_location(
                        image.anchor._from.row + 1, img_num
                    )

                    s3 = boto3.client(
                        "s3",
                        aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
                        aws_session_token=credentials.AWS_SESSION_TOKEN,
                    )
                    image_dict["md5"] = hashlib.md5(image._data()).hexdigest()
                    image_dict["uuid"] = str(uuid.uuid4())
                    image_dict["format"] = image.format
                    thread_pool.submit(
                        upload_image,
                        s3,
                        self.uuid,
                        self.count_images,
                        image_dict,
                        image,
                        self.part_index,
                    )
                    self.count_images += 1
                    image_dict["created_by_id"] = self.user_id
                    image_dict["uploaded_at"] = self.created_at
                except Exception:
                    pass
                else:
                    images_list[row_img_num] = image_dict

            thread_pool.shutdown()
        return images_list

    def parse_resources(self, row_dict, firm, reporting_id):
        column_errors = []
        resources_list = []
        if firm and row_dict and "resources" in row_dict:
            resources = self.resources.get(firm)
            if resources is None:
                resources = Resource.objects.filter(
                    company=self.company,
                    resource_service_orders__contract__unit_price_services__firms=firm,
                ).distinct()
                self.resources[firm] = resources
            for key, resource in row_dict["resources"].items():
                resource_num = key.split(":")[-1]
                resource_name = resource.get("name", None)
                section_name = resource.get("section", None)
                quantity = resource.get("quantity", None)

                if section_name and not resource_name:
                    column_errors.append("Recurso_{}".format(resource_num))
                    continue

                if not resource_name and not section_name:
                    continue

                if not quantity or not isinstance(quantity, (int, float)):
                    if resource_name:
                        column_errors.append("Recurso_{}".format(resource_num))
                    continue

                try:
                    if section_name:
                        sub = ContractService.objects.filter(
                            description__icontains=section_name.lower().strip(),
                            contract_item_unit_prices__resource__resource=OuterRef(
                                "pk"
                            ),
                        ).values("pk")[:1]

                        found_resource = resources.filter(
                            name__icontains=resource_name.lower().strip(),
                            resource_service_orders__contract__unit_price_services=Subquery(
                                sub
                            ),
                        )[0]
                    else:
                        found_resource = resources.filter(
                            name__icontains=resource_name.lower().strip()
                        )[0]
                    so_resource = found_resource.resource_service_orders.all()[0]
                except Exception:
                    column_errors.append("Recurso_{}".format(resource_num))
                    continue

                unit_price = so_resource.unit_price
                resources_list.append(
                    {
                        "uuid": str(uuid.uuid4()),
                        "total_price": quantity * unit_price,
                        "unit_price": unit_price,
                        "amount": quantity,
                        "reporting_id": reporting_id,
                        "resource_id": str(found_resource.uuid),
                        "service_order_resource_id": str(so_resource.uuid),
                        "created_by_id": self.user_id,
                        "creation_date": self.created_at,
                        "firm_id": str(firm.uuid),
                    }
                )

        row_dict["resources"] = resources_list

        # Update column_errors
        row_dict = self.update_column_errors(row_dict, column_errors)

        return row_dict

    def add_value(self, key, value, reporting_dict, column_errors, force_error=False):
        if key == "lot":
            if not value:
                column_errors.append("km")
            return reporting_dict, column_errors
        elif key == "inventory_value":
            column_errors.append("import_inventory_code")
            reporting_dict[key] = value
            return reporting_dict, column_errors
        else:
            reporting_dict[key] = value
            if (self.is_edit is False and not value) or force_error:
                column_errors.append(key)
            return reporting_dict, column_errors

    def parse_obj(self, row_dict, basic_columns, header):
        reporting_id = str(uuid.uuid4()) if not self.is_edit else None
        column_errors = []
        reporting_dict = {}
        reporting_dict["point"] = {}
        # Get ocurrence_type and firm first
        occ_type = self.parse_occurrence_type(row_dict)
        if occ_type:
            is_inventory = occ_type.occurrence_kind == "2"
        else:
            is_inventory = False
        firm = self.parse_firm(row_dict)

        # Parse form_data (mandatory)
        row_dict = self.parse_form_data(row_dict, occ_type, header)

        # Parse resources (mandatory)
        row_dict = self.parse_resources(row_dict, firm, reporting_id)
        for key, value in row_dict.items():
            # Validation
            if key in basic_columns:
                reporting_dict[key] = value
            elif key in REQUIRED_FIELDS and value is None:
                if self.is_edit is False:
                    column_errors.append(key)
            # Parse
            elif key in ["executed_at", "due_at", "found_at"]:
                if isinstance(value, datetime):
                    reporting_dict[key] = to_utc_string(value)
                elif isinstance(value, str):
                    try:
                        datetime_value = datetime.strptime(value, "%d/%m/%Y")
                    except ValueError:
                        column_errors.append(key)
                        reporting_dict[key] = None
                    else:
                        reporting_dict[key] = to_utc_string(datetime_value)
                elif not value and key == "found_at":
                    if self.is_edit is False:
                        reporting_dict[key] = self.created_at
                elif value:
                    column_errors.append(key)
                    reporting_dict[key] = None
            elif key == "km":
                reporting_dict[key] = self.parse_km(value)
                if reporting_dict[key] is None and not self.hide_reporting_location:
                    if self.is_edit is False:
                        column_errors.append(key)
                        reporting_dict["km_error"] = {
                            "error": "Preencha corretamente os campos obrigatórios: km e Rodovia."
                        }
            elif key in ["end_km", "project_km", "project_end_km", "km_reference"]:
                if key != "km_reference":
                    reporting_dict[key] = self.parse_km(value)
                elif key == "km_reference":
                    if get_obj_from_path(self.company.metadata, "show_track"):
                        reporting_dict[key] = self.parse_km(value)
            elif key == "road_name":
                road_value = self.parse_road(value)
                if not self.hide_reporting_location:
                    reporting_dict, column_errors = self.add_value(
                        key, road_value, reporting_dict, column_errors
                    )
                    if road_value is None and self.is_edit is False:
                        reporting_dict["road_error"] = {
                            "error": "Preencha corretamente os campos obrigatórios: km e Rodovia."
                        }
            elif key == "direction":
                value_clean = clean_latin_string(str(value)).lower()
                direction_value = self.direction_translation.get(value_clean, "")

                # Try using to get default value
                default_direction_value = get_obj_from_path(
                    self.company.custom_options,
                    "reporting__fields__direction__defaultvalue",
                )
                if default_direction_value and not direction_value:
                    direction_value = default_direction_value

                reporting_dict, column_errors = self.add_value(
                    key, direction_value, reporting_dict, column_errors
                )
            elif key == "lane":
                value_clean = clean_latin_string(str(value)).lower()
                lane_value = self.lane_translation.get(value_clean, "")
                reporting_dict, column_errors = self.add_value(
                    key, lane_value, reporting_dict, column_errors
                )
            elif key == "track":
                if get_obj_from_path(self.company.metadata, "show_track"):
                    if value:
                        value_clean = clean_latin_string(str(value)).lower()
                        track_value = self.parse_track(value_clean)
                        reporting_dict, column_errors = self.add_value(
                            key, track_value, reporting_dict, column_errors
                        )
            elif key == "branch":
                if get_obj_from_path(self.company.metadata, "show_track"):
                    if value:
                        value_clean = clean_latin_string(str(value)).lower()
                        branch_value = self.parse_branch(value_clean)
                        reporting_dict, column_errors = self.add_value(
                            key, branch_value, reporting_dict, column_errors
                        )
            elif key == "job_id":
                value = self.parse_job(value)
                reporting_dict, column_errors = self.add_value(
                    key, value, reporting_dict, column_errors
                )
            elif key == "menu_id" and value:  # Falsy values should go to fallback
                value, is_visible = self.parse_menu(value)

                # Non visible menus should trigger error even when found in DB
                force_error = False
                if value and not is_visible:
                    force_error = True
                    reporting_dict["menu_error"] = {
                        "error": "O sistema não permite importar apontamentos com menus que estão ocultos. Preencha essa coluna com um menu visível."
                    }
                elif value is None and self.is_edit is False and not is_inventory:
                    reporting_dict["menu_error"] = {
                        "error": "O menu informado não existe. Preencha essa coluna com um menu visível."
                    }

                reporting_dict, column_errors = self.add_value(
                    key, value, reporting_dict, column_errors, force_error=force_error
                )
            elif key == "occurrence_type_id":
                value = str(occ_type.uuid) if occ_type else ""
                reporting_dict, column_errors = self.add_value(
                    key, value, reporting_dict, column_errors
                )
            elif key == "firm_id":
                value = str(firm.uuid) if firm else ""
                reporting_dict, column_errors = self.add_value(
                    key, value, reporting_dict, column_errors
                )
            elif key == "status_id":
                value = self.parse_status(value)
                reporting_dict, column_errors = self.add_value(
                    key, value, reporting_dict, column_errors
                )
            elif key in ["form_data", "resources"]:
                reporting_dict[key] = value
            elif key == "inventory_value":
                if self.inventory_code is not None:
                    if value:
                        if isinstance(value, (int, float)):
                            cache_key = value
                        else:
                            cache_key = str(value)
                        cached_inventories = self.inventory_cache.get(cache_key, [])

                        if not cached_inventories:
                            reporting_dict["import_inventory_code"] = {
                                "error": "O código não consta em nenhum item de inventário. Adicione o código em um item de inventário."
                            }
                            reporting_dict, column_errors = self.add_value(
                                key, value, reporting_dict, column_errors
                            )
                        elif len(cached_inventories) > 1:
                            reporting_dict["import_inventory_code"] = {
                                "error": "O código foi encontrado em dois ou mais inventários. Verifique se existe dois ou mais itens de inventário com o mesmo código."
                            }
                            reporting_dict, column_errors = self.add_value(
                                key, value, reporting_dict, column_errors
                            )
                        else:
                            inventory_data = cached_inventories[0]
                            reporting_dict["import_inventory_code"] = {
                                "occurrence_type": {
                                    "name": "Classe",
                                    "value": inventory_data.get(
                                        "occurrence_type__name"
                                    ),
                                },
                                "number": {
                                    "name": "Serial",
                                    "value": inventory_data.get("number"),
                                },
                                "created_at": {
                                    "name": "Criado em",
                                    "value": to_utc_string(
                                        inventory_data.get("created_at")
                                    ),
                                },
                                "road_name": {
                                    "name": "Rodovia",
                                    "value": inventory_data.get("road_name"),
                                },
                                "km": {
                                    "name": "km",
                                    "value": "{:07.3f}".format(
                                        inventory_data.get("km")
                                    ),
                                },
                                "link": "{}/#/Inventory/{}/show".format(
                                    settings.FRONTEND_URL,
                                    str(inventory_data.get("uuid")),
                                ),
                                "uuid": str(inventory_data.get("uuid")),
                            }
                            reporting_dict[key] = value
                    else:
                        reporting_dict["import_inventory_code"] = {}
                        reporting_dict[key] = value
                else:
                    reporting_dict["import_inventory_code"] = {}
                    reporting_dict[key] = value
            elif key in ["latitude", "longitude"]:
                if self.show_coordinate_input or self.hide_reporting_location:
                    reporting_dict[key] = value
                    coord = self.parse_coordinates(value)
                    reporting_dict["point"].update(
                        {"x" if key == "longitude" else "y": coord}
                    )
            elif key == "number":
                rep_number = value.strip() if value and isinstance(value, str) else None
                reporting_dict[key] = rep_number

                matches = self.number_to_reporting[rep_number]
                match_id, is_editable = matches[0] if matches else (None, None)
                if not rep_number:
                    reporting_dict[
                        "number_error"
                    ] = "O código do serial é obrigatório na edição via importação."
                elif len(matches) > 1:
                    reporting_dict[
                        "number_error"
                    ] = "O código do serial foi encontrado em dois ou mais apontamentos, ou itens de inventários. Verifique se existe dois ou mais apontamentos, ou itens de inventários com o mesmo código."
                elif len(matches) == 0:
                    reporting_dict["number_error"] = "Serial não encontrado"
                elif is_editable is False:
                    reporting_dict[
                        "number_error"
                    ] = "O serial não está disponível para edição. O passo de aprovação no qual ele se encontra não permite mais alterações."
                else:
                    reporting_id = str(match_id)

                if "number_error" in reporting_dict:
                    column_errors.append(key)

        if "point" in reporting_dict:
            lat = reporting_dict["point"].get("y")
            long = reporting_dict["point"].get("x")
            if lat is None or long is None:
                if self.hide_reporting_location and not self.is_edit:
                    reporting_dict["point"] = {
                        "error": "Preencha os campos obrigatórios: latitude e longitude."
                    }
                    column_errors.append("point")
                else:
                    del reporting_dict["point"]
            elif lat == "erro" or long == "erro":
                reporting_dict["point"] = {
                    "error": "Preenchimento incorreto da latitude ou longitude."
                }

                column_errors.append("point")
            else:
                if (lat > 90 or lat < -90) or (long > 180 or long < -180):
                    reporting_dict["point"] = {
                        "error": "Preenchimento incorreto da latitude ou longitude."
                    }
                    column_errors.append("point")

        if (
            "menu_id" not in reporting_dict
            and self.is_edit is False
            and not is_inventory
        ):
            if self.fallback_menu is None:
                reporting_dict["menu_error"] = {
                    "error": "O menu informado não existe. Verifique os dados e importe o arquivo novamente."
                }

            # NOTE: This will add a column error if we couldn't find a suitable menu in the __init__()
            reporting_dict, column_errors = self.add_value(
                "menu_id", self.fallback_menu, reporting_dict, column_errors
            )

        # Autofill OccurrenceType ID if not provided but determined by parse_occurrence_type()
        if "occurrence_type_id" not in reporting_dict and occ_type:
            reporting_dict["occurrence_type_id"] = str(occ_type.uuid)

        # Ensure number is different from linked inventory
        inv_num = reporting_dict.get("inventory_value")
        if self.is_edit and inv_num and reporting_dict["number"] == inv_num:
            column_errors.append("number")
            column_errors.append("inventory_value")
            reporting_dict[
                "number_error"
            ] = "O mesmo código de serial foi inserido tanto na coluna 'Código do Inventário para vínculo com apontamento' quanto na coluna 'Serial do apontamento ou do item de inventário que será editado'. Para editar o serial correspondente, insira o código apenas na coluna 'Serial do apontamento ou do item de inventário que será editado'."

        # Inject default approval_step_id
        if self.approval_step_id:
            reporting_dict["approval_step_id"] = self.approval_step_id

        # Try to calculate lot
        calculate_lot_for_edit = (
            not self.is_edit or reporting_dict.get("km") is not None
        )
        if (
            hasattr(self, "road")
            and self.road is not None
            and self.road.lot_logic
            and self.road.lot_logic != {}
            and calculate_lot_for_edit
        ):
            try:
                lot = apply_json_logic(
                    self.road.lot_logic,
                    {"data": {"km": reporting_dict.get("km", None)}},
                )
            except Exception:
                lot = ""
            reporting_dict, column_errors = self.add_value(
                "lot", lot, reporting_dict, column_errors
            )
        if self.hide_reporting_location:
            reporting_dict.update(
                {"direction": "0", "km": 0, "lane": "X", "road_name": "X"}
            )
        reporting_keys = list(reporting_dict.keys())

        if self.is_edit is False:
            for item in self.required_fields_import:
                if item not in reporting_keys:
                    column_errors.append(item)
                    reporting_dict[item] = None

        # Update parsed resources with matched ID (if any) if is_edit
        # NOTE: We do this because when the resource is first parsed we don't have the ID match yet
        pars_resources = reporting_dict.get("resources")
        if self.is_edit and reporting_id and pars_resources:
            for pars_res in pars_resources:
                pars_res["reporting_id"] = reporting_id
            reporting_dict["resources"] = pars_resources

        # Update column_errors
        reporting_dict = self.update_column_errors(reporting_dict, column_errors)

        return {
            **reporting_dict,
            "uuid": reporting_id,
            "company_id": self.company_id,
            "created_by_id": self.user_id if self.is_edit is False else None,
            "created_at": self.created_at if self.is_edit is False else None,
        }

    def is_hidden_sheet(self, worksheet) -> bool:
        return shared_is_hidden_sheet(worksheet)

    def is_edit_import(self) -> bool:
        for sheet_name in self.wb.sheetnames:
            worksheet = self.wb[sheet_name]
            if self.is_hidden_sheet(worksheet):
                continue

            headers = self.get_header(worksheet)
            if headers and headers[0] == "number":
                return True

        return False

    def get_provided_numbers(self) -> Tuple[str]:
        """
        Returns all "number" column values in the workbook. Useful for edit imports.
        Returns:
            Tuple[str]: Tuple of all provided numbers
        """

        provided_numbers = []
        try:
            for sheet_name in self.wb.sheetnames:
                worksheet = self.wb[sheet_name]
                if self.is_hidden_sheet(worksheet):
                    continue

                # We take a slice without the header row
                sheet_numbers = []
                for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    if row and isinstance(row[0], str):
                        sheet_numbers.append(row[0])
                provided_numbers.extend(sheet_numbers)
        except Exception:
            return set()
        else:
            return set(provided_numbers)

    def get_data(self, starting_row=0, use_parts_filter=False, rows_per_part=None):
        is_array_edit = "edit_export" in self.wb.sheetnames

        if not self.hide_reporting_location:
            self.get_lane_and_direction()
            if not self.lane_translation or not self.direction_translation:
                return {}

        user_permissions = PermissionManager(
            self.company_id, self.request_user, "Reporting"
        )
        can_edit = user_permissions.has_permission("can_edit")

        # Edit import preparations and queryset retrieval
        self.is_edit = self.is_edit_import()
        if self.is_edit:
            self.approval_step_id = None

            if self.can_use_edit is False or not can_edit:
                self.excel_import.is_forbidden = True
                return {}

            provided_numbers = self.get_provided_numbers()
            if provided_numbers:
                # Limit the amount of reportings being processed before we do anything
                if len(provided_numbers) > REPORTING_LIMIT:
                    self.excel_import.is_over_limit = True
                    return {}

                qs = set(
                    get_reporting_queryset(
                        self.company_id, self.request_user, user_permissions
                    )
                    .filter(number__in=provided_numbers)
                    .values_list("uuid", "editable", "number")
                )
                # If not all instances were found, try Inventory instances too
                if len(qs) < len(provided_numbers):
                    qs = qs.union(
                        set(
                            get_inventory_queryset(
                                self.company_id, self.request_user, user_permissions
                            )
                            .filter(number__in=provided_numbers)
                            .values_list("uuid", "editable", "number")
                        )
                    )

                for item_id, editable, number in qs:
                    if (
                        number
                        and (item_id, editable) not in self.number_to_reporting[number]
                    ):
                        self.number_to_reporting[number].append((item_id, editable))

        # Preload inventory cache from all sheets at once to avoid duplicates
        # when the same inventory code appears in multiple sheets
        if self.inventory_code:
            all_inventory_values = []
            for sheet_name in self.wb.sheetnames:
                worksheet = self.wb[sheet_name]

                if self.is_hidden_sheet(worksheet):
                    continue

                header = self.get_header(worksheet)
                if "inventory_value" in header:
                    inventory_col_index = header.index("inventory_value")
                    for i, row in enumerate(worksheet.iter_rows()):
                        if i == 0 or not any(cell.value for cell in row):
                            continue
                        if (
                            len(row) > inventory_col_index
                            and row[inventory_col_index].value
                        ):
                            raw_value = row[inventory_col_index].value
                            if isinstance(raw_value, (int, float)):
                                all_inventory_values.append(raw_value)
                            else:
                                all_inventory_values.append(str(raw_value))

            if all_inventory_values:
                self.preload_inventory_cache(all_inventory_values)

        # Iterate on each sheet
        reportings_dict_list = []
        images_dict_list = []
        for sheet_name in self.wb.sheetnames:
            worksheet = self.wb[sheet_name]

            if self.is_hidden_sheet(worksheet):
                continue

            header = self.get_header(worksheet)
            len_header = len(header)
            images = self.get_images(worksheet)

            for i, row in enumerate(worksheet.iter_rows()):
                if i == 0:
                    continue
                if use_parts_filter:
                    if i <= starting_row:
                        continue
                    if i > starting_row + rows_per_part:
                        break
                row_is_not_empty = any(cell.value for cell in row)
                if not row_is_not_empty:
                    continue
                row_dict = {
                    "row": i + 1,
                    "column_errors": [],
                    "formula_errors": [],
                    "resources": {},
                }
                basic_columns = list(row_dict.keys())
                # Iterate on each cell
                for cell in row:
                    if cell.column <= len_header:
                        column_name = header[cell.column - 1]
                        initial_column_name = clean_latin_string(
                            column_name.split("_")[0]
                        ).lower()

                        if cell.data_type == "f":
                            row_dict["formula_errors"].append(column_name)

                        # Images cells
                        if initial_column_name in PHOTO_COLUMNS:
                            img_num = column_name.split("_")[-1]
                            row_img_num = self.format_row_location(cell.row, img_num)
                            img = images.get(row_img_num, {})
                            if img and img.get("upload", False):
                                if (
                                    initial_column_name == "foto"
                                    and img.get("format", "") not in COMMON_IMAGE_TYPE
                                ):
                                    row_dict["column_errors"].append("Foto")
                                elif (
                                    initial_column_name == "descricao foto"
                                    and cell.value
                                ):
                                    try:
                                        description = str(cell.value)
                                    except Exception:
                                        row_dict["column_errors"].append("Foto")
                                    else:
                                        img["description"] = description
                                elif initial_column_name == "tipo foto":
                                    img["kind"] = self.translate_image_kind(cell.value)
                                elif initial_column_name == "data foto":
                                    if cell.value and isinstance(cell.value, datetime):
                                        img["datetime"] = to_utc_string(cell.value)
                                    elif isinstance(cell.value, str):
                                        try:
                                            datetime_value = datetime.strptime(
                                                cell.value, "%d/%m/%Y"
                                            )
                                        except ValueError:
                                            row_dict["column_errors"].append("Foto")
                                        else:
                                            img["datetime"] = to_utc_string(
                                                datetime_value
                                            )
                                    elif cell.value:
                                        row_dict["column_errors"].append("Foto")
                                        img["datetime"] = None
                                if "description" not in img or not img["description"]:
                                    img["description"] = "Imagem enviada pela web"

                                img["row"] = i + 1
                            elif not img and cell.value:
                                row_dict["column_errors"].append("Foto")
                        elif (
                            "foto" in column_name.lower()
                            and " - " in column_name
                            and ":" in column_name
                        ):
                            img_name = (
                                "".join(column_name.split("_")[1:])
                                .replace("Descrição ", "")
                                .replace("Tipo ", "")
                                .replace("Data ", "")
                            )
                            row_img_num = self.format_row_location(cell.row, img_name)
                            img = images.get(row_img_num, {})
                            if img and img.get("upload", False):
                                position = str(
                                    column_name.split(":")[0].split("_")[-1].strip()
                                )
                                self.ARRAY_USED_POSITIONS[
                                    row_dict.get("row", "")
                                ].append(int(position))
                                photo_attribute = clean_latin_string(
                                    column_name.split(" - ")[-1].split("_")[0].strip()
                                ).lower()
                                if (
                                    photo_attribute == "foto"
                                    and img.get("format", "") not in COMMON_IMAGE_TYPE
                                ):
                                    row_dict["column_errors"].append(
                                        column_name.replace("Descrição ", "")
                                        .replace("Tipo ", "")
                                        .replace("Data ", "")
                                    )
                                elif photo_attribute == "descricao foto" and cell.value:
                                    try:
                                        description = str(cell.value)
                                    except Exception:
                                        row_dict["column_errors"].append(
                                            column_name.replace("Descrição ", "")
                                            .replace("Tipo ", "")
                                            .replace("Data ", "")
                                        )
                                    else:
                                        img["description"] = description
                                elif photo_attribute == "tipo foto":
                                    img["kind"] = self.translate_image_kind(cell.value)
                                elif photo_attribute == "data foto":
                                    if cell.value and isinstance(cell.value, datetime):
                                        img["datetime"] = to_utc_string(cell.value)
                                    elif isinstance(cell.value, str):
                                        try:
                                            datetime_value = datetime.strptime(
                                                cell.value, "%d/%m/%Y"
                                            )
                                        except ValueError:
                                            row_dict["column_errors"].append(
                                                column_name.replace("Descrição ", "")
                                                .replace("Tipo ", "")
                                                .replace("Data ", "")
                                            )
                                        else:
                                            img["datetime"] = to_utc_string(
                                                datetime_value
                                            )
                                    elif cell.value:
                                        row_dict["column_errors"].append(
                                            column_name.replace("Descrição ", "")
                                            .replace("Tipo ", "")
                                            .replace("Data ", "")
                                        )
                                        img["datetime"] = None
                                if "description" not in img or not img["description"]:
                                    img["description"] = "Imagem enviada pela web"
                                img["row"] = i + 1
                                img["array_photo"] = {initial_column_name: position}
                                img["inner_array_photo"] = (
                                    clean_latin_string(
                                        column_name.split(":")[1].split(" - ")[0]
                                    )
                                    .lower()
                                    .strip()
                                )
                            elif not img and cell.value:
                                row_dict["column_errors"].append(
                                    column_name.replace("Descrição ", "")
                                    .replace("Tipo ", "")
                                    .replace("Data ", "")
                                )
                        # Resource cells
                        elif initial_column_name in RESOURCE_COLUMNS:
                            resource_num = column_name.split("_")[-1]
                            row_key = self.format_row_location(cell.row, resource_num)
                            if initial_column_name == "secao do recurso" and cell.value:
                                if row_key not in row_dict["resources"]:
                                    row_dict["resources"][row_key] = {}
                                row_dict["resources"][row_key]["section"] = cell.value
                            elif initial_column_name == "recurso" and cell.value:
                                try:
                                    quantity = row[cell.column].value
                                except Exception:
                                    quantity = 0
                                if row_key not in row_dict["resources"]:
                                    row_dict["resources"][row_key] = {}
                                row_dict["resources"][row_key]["name"] = cell.value
                                row_dict["resources"][row_key]["quantity"] = quantity
                        # Other cells
                        else:
                            row_dict[column_name] = cell.value

                # Parse values
                if row_dict:
                    parsed_row = self.parse_obj(row_dict, basic_columns, header)
                    if parsed_row and "uuid" in parsed_row:
                        reportings_dict_list.append(parsed_row)
                        row_images = list(
                            filter(
                                lambda x: "row" in x and x["row"] == parsed_row["row"],
                                images.values(),
                            )
                        )

                        # Add reporting_id in row_images
                        row_images = [
                            {"reporting_id": parsed_row["uuid"], **item}
                            for item in row_images
                        ]
                        images_dict_list.extend(row_images)

        # Remove images with wrong format or not upload
        images_dict_list = list(
            filter(
                lambda x: x["upload"] and x["format"] in COMMON_IMAGE_TYPE,
                images_dict_list,
            )
        )
        images_to_remove = []
        for image in images_dict_list:
            array_photo = image.get("array_photo", None)
            if array_photo:
                reporting = next(
                    (
                        a
                        for a in reportings_dict_list
                        if image["reporting_id"] == a["uuid"]
                    ),
                    None,
                )
                if reporting and reporting.get("occurrence_type_id", "") == "":
                    images_to_remove.append(image["uuid"])
                if (
                    reporting
                    and reporting.get("occurrence_type_id", "") != ""
                    and reporting.get("form_data", {}) != {}
                ):
                    form_data_name, position = next(iter(array_photo.items()))
                    array_api_name = ARRAY_COLUMNS.get(form_data_name)
                    array_inner_name = image.get("inner_array_photo")
                    array_full_name = f"{form_data_name}***{array_inner_name}"
                    array_inner_api_name = ARRAY_PHOTO_COLUMNS.get(array_full_name)
                    reporting["form_data"][array_api_name][int(position) - 1][
                        array_inner_api_name
                    ].append(image["uuid"])
                    image.pop("array_photo", None)
                    image.pop("inner_array_photo", None)
        images_dict_list = [
            item for item in images_dict_list if item["uuid"] not in images_to_remove
        ]

        # Parse resources
        resources_dict_list = [
            a for item in reportings_dict_list for a in item.pop("resources", [])
        ]
        if not is_array_edit:
            for parsed_row in reportings_dict_list:
                if parsed_row.get("form_data", {}) != {}:
                    for k, v in parsed_row["form_data"].copy().items():
                        if isinstance(v, list) and v != [] and isinstance(v[0], dict):
                            array_position_filtered = list(
                                set(
                                    self.ARRAY_USED_POSITIONS.get(
                                        parsed_row.get("row"), ""
                                    )
                                )
                            )
                            v[:] = [
                                {key: value for key, value in item.items()}
                                for item in copy(v)
                                if item.get("index_array", "")
                                in array_position_filtered
                            ]

        if self.is_edit:
            number_to_indexes = defaultdict(list)
            for i, reporting_dict in enumerate(reportings_dict_list):
                # Report missing number fields (also blocks hybrid imports)
                if "number" not in reporting_dict:
                    reporting_dict["number"] = None
                    reporting_dict[
                        "number_error"
                    ] = "O código do serial é obrigatório na edição via importação."
                    reportings_dict_list[i] = self.update_column_errors(
                        reporting_dict, ["number"]
                    )
                # Count number duplicates
                else:
                    number_to_indexes[reporting_dict["number"]].append(i)

            # Raise error for duplicated numbers in different rows
            for indexes in number_to_indexes.values():
                if len(indexes) > 1:
                    for i in indexes:
                        reportings_dict_list[i][
                            "number_error"
                        ] = "O mesmo código de serial foi encontrado em duas ou mais linhas da planilha Excel"
                        reportings_dict_list[i] = self.update_column_errors(
                            reportings_dict_list[i], ["number"]
                        )

        if reportings_dict_list:
            return {
                "is_edit": self.is_edit,
                "reportings": reportings_dict_list,
                "images": images_dict_list,
                "resources": resources_dict_list,
                "is_array_edit": is_array_edit,
            }
        return {}

    def download_excel_file(self):
        if self.excel_import.excel_file:
            try:
                unquoted_file_path = parse.unquote(self.excel_import.excel_file.url)
                file_path = unquoted_file_path.split("?")[0].split(".com/")[1]
                bucket_name = unquoted_file_path.split(".s3")[0].split("/")[-1]
                file_format = file_path.split(".")[-1]
                file_name = file_path.split("/")[-1].split(".")[0]
            except Exception:
                return ""

            file_temp_path = "{}{}{}.{}".format(
                self.temp_path, file_name, self.uuid, file_format
            )

            try:
                self.s3.download_file(bucket_name, file_path, file_temp_path)
            except Exception:
                return ""
            else:
                return file_temp_path
        return ""

    def get_data_part(self, sheet_title: str, starting_row: int, rows_per_part: int):
        data = {}
        # Load data
        self.load_data()
        if not self.wb or not self.company_id or not self.user_id:
            data = {}
        else:
            delete_other_sheets_and_rows(
                self.wb, sheet_title, starting_row, rows_per_part
            )
            data = self.get_data(
                starting_row=starting_row,
                use_parts_filter=True,
                rows_per_part=rows_per_part,
            )

        return data

    def get_parts(self) -> Tuple[List[Tuple[str, int]], int]:
        """
        Get the parts of the excel import.
        Returns a list of tuples, each containing the sheet index and the starting row
        that should be processed by each excel import part invocation
        """

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )
        os.makedirs(self.temp_path, exist_ok=True)
        # Download excel
        self.file_name = shared_download_excel_file(
            self.excel_import, self.temp_path, self.s3
        )
        if self.file_name:
            # Load data
            self.load_data()
            if not self.wb or not self.company_id or not self.user_id:
                return None

            real_max_rows = get_real_max_rows_per_sheet(self.file_name)

            # Primeiro, calcula o total de linhas para determinar o tamanho de cada parte
            total_rows = 0
            sheet_rows = {}
            for ws in self.wb.worksheets:
                if ws.sheet_state == Worksheet.SHEETSTATE_VISIBLE:
                    max_row = real_max_rows.get(ws.title, ws.max_row)
                    rows_count = max(0, max_row - 1)  # Subtrai header
                    sheet_rows[ws.title] = rows_count
                    total_rows += rows_count

            # Calcula quantos batches seriam criados com MAX_ROWS_PER_PART
            estimated_batches = (
                total_rows + MAX_ROWS_PER_PART - 1
            ) // MAX_ROWS_PER_PART
            # Se exceder o limite, ajusta o tamanho de cada parte
            if estimated_batches > MAX_BATCHES_LIMIT:
                rows_per_part = (
                    total_rows + MAX_BATCHES_LIMIT - 1
                ) // MAX_BATCHES_LIMIT

                sentry_sdk.capture_message(
                    f"A importação do Excel {self.excel_import.uuid} excedeu o limite esperado de batches "
                    f"({estimated_batches}). O processamento foi ajustado automaticamente para "
                    f"{MAX_BATCHES_LIMIT} batches com {rows_per_part} linhas por parte para evitar "
                    f"sobrecarga no sistema."
                )

            else:
                rows_per_part = MAX_ROWS_PER_PART

            # Gera as partes com o tamanho calculado
            parts = []
            for ws in self.wb.worksheets:
                if ws.sheet_state == Worksheet.SHEETSTATE_VISIBLE:
                    max_row = real_max_rows.get(ws.title, ws.max_row)
                    for starting_row in range(0, max_row - 1, rows_per_part):
                        parts.append((ws.title, starting_row))
            return parts, rows_per_part
        return None, None

    def import_excel(self):
        parts, rows_per_part = self.get_parts()

        if parts:
            ExcelImport.objects.filter(uuid=self.excel_import.uuid).update(
                remaining_parts=len(parts)
            )

            for idx, (sheet_title, starting_row) in enumerate(parts):
                parse_excel_part_to_json(
                    str(self.excel_import.uuid),
                    self.user_id,
                    self.original_inventory_code,
                    sheet_title,
                    starting_row,
                    idx,
                    len(parts),
                    rows_per_part,
                )
        else:
            raise Exception(
                "helpers.import_excel.read_excel.get_parts: failed to get parts"
            )

    def save_part_data(self, data: dict, idx: int):

        file_name = f"{str(self.excel_import.uuid)}-{idx}.json"
        file_path = self.temp_path + file_name

        with open(file_path, "w") as outfile:
            json.dump(data, outfile)

        object_name = f"media/private/{file_name}"
        bucket_name = settings.AWS_STORAGE_BUCKET_NAME

        self.s3.upload_file(
            file_path,
            bucket_name,
            object_name,
        )

    def dec_and_fetch_remaining_parts(self):

        with transaction.atomic():
            # Lock the row for update to prevent concurrent modifications
            excel_import = ExcelImport.objects.select_for_update().get(
                uuid=self.excel_import.uuid
            )
            # Atomically decrement and update
            ExcelImport.objects.filter(uuid=self.excel_import.uuid).update(
                remaining_parts=F("remaining_parts") - 1
            )
            # Refresh to get the new value
            excel_import.refresh_from_db()
            remaining_parts = excel_import.remaining_parts
            # Update the instance attribute for consistency
            self.excel_import.remaining_parts = remaining_parts
            return remaining_parts

    def merge_data_parts(self, parts_count: int):
        data = {}
        for idx in range(parts_count):
            file_name = f"{str(self.excel_import.uuid)}-{idx}.json"
            object_name = f"media/private/{file_name}"
            bucket_name = settings.AWS_STORAGE_BUCKET_NAME
            file_path = self.temp_path + file_name
            self.s3.download_file(bucket_name, object_name, file_path)
            data_part = {}
            with open(file_path, "r") as infile:
                data_part = json.load(infile)
            data["is_edit"] = data.get("is_edit", False) or data_part.get(
                "is_edit", False
            )
            data["reportings"] = data.get("reportings", []) + data_part.get(
                "reportings", []
            )
            data["images"] = data.get("images", []) + data_part.get("images", [])
            data["resources"] = data.get("resources", []) + data_part.get(
                "resources", []
            )
            data["is_array_edit"] = data.get("is_array_edit", False) or data_part.get(
                "is_array_edit", False
            )
        return data

    def complete_excel_import(self, data: dict, parsing_error: bool):
        errors = False
        try:
            if data:
                json_name = "{}.json".format(self.uuid)
                json_file_path = self.temp_path + json_name
                # if cell contains formula change its value to "columns_errors"
                for item in data["reportings"]:
                    for formula_errors in item.get("formula_errors", []):
                        item["column_errors"].append(formula_errors)
                    item.pop("formula_errors")
                # Camelize data
                data = dict_to_casing(data)

                with open(json_file_path, "w") as outfile:
                    json.dump(data, outfile)

                json_file = open(json_file_path, "rb")
                self.excel_import.preview_file.save(
                    json_name, ContentFile(json_file.read())
                )

                errors = bool(
                    [
                        error_column
                        for item in data["reportings"]
                        for error_column in item.get("columnErrors", [])
                    ]
                )
            else:
                errors = True
        except Exception as e:
            sentry_sdk.capture_exception(e)
            errors = True

        self.excel_import.error = errors
        self.excel_import.generating_preview = False
        self.excel_import.save()
        logging.info("Parse Done")

    def import_excel_part(
        self,
        sheet_title: str,
        starting_row: int,
        idx: int,
        parts_count: int,
        rows_per_part: int,
    ):
        self.part_index = idx
        parsing_error = False
        data = {}
        try:

            self.s3 = boto3.client(
                "s3",
                aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
                aws_session_token=credentials.AWS_SESSION_TOKEN,
            )

            os.makedirs(self.temp_path, exist_ok=True)

            # Download excel
            self.file_name = shared_download_excel_file(
                self.excel_import, self.temp_path, self.s3
            )
            if self.file_name:
                data_part = self.get_data_part(sheet_title, starting_row, rows_per_part)
                self.save_part_data(data_part, idx)

                remaining_parts = self.dec_and_fetch_remaining_parts()
                if remaining_parts == 0:
                    data = self.merge_data_parts(parts_count)
                    self.complete_excel_import(data, parsing_error)

            else:
                parsing_error = True
        except SyntaxError as e:
            parsing_error = True
            sentry_sdk.capture_exception(e)

        shared_clean_up(self.file_name, self.temp_path)

    def get_excel_import(self):
        error = True

        self.s3 = boto3.client(
            "s3",
            aws_access_key_id=credentials.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=credentials.AWS_SECRET_ACCESS_KEY,
            aws_session_token=credentials.AWS_SESSION_TOKEN,
        )

        os.makedirs(self.temp_path, exist_ok=True)
        json_name = "{}.json".format(self.uuid)
        json_file_path = self.temp_path + json_name

        # Download excel
        self.file_name = shared_download_excel_file(
            self.excel_import, self.temp_path, self.s3
        )
        if self.file_name:
            # Load data
            self.load_data()
            data = {}
            if self.wb and self.company_id and self.user_id:
                data = self.get_data()

            if data:
                # if cell contains formula change its value to "columns_errors"
                for item in data["reportings"]:
                    for formula_errors in item.get("formula_errors", []):
                        item["column_errors"].append(formula_errors)
                    item.pop("formula_errors")
                # Camelize data
                data = dict_to_casing(data)

                with open(json_file_path, "w") as outfile:
                    json.dump(data, outfile)

                json_file = open(json_file_path, "rb")
                self.excel_import.preview_file.save(
                    json_name, ContentFile(json_file.read())
                )

                has_errors = [
                    error_column
                    for item in data["reportings"]
                    for error_column in item.get("columnErrors", [])
                ]
                if not has_errors:
                    error = False

        self.excel_import.error = error

        shared_clean_up(self.file_name, self.temp_path)

        return self.excel_import


MAX_ROWS_PER_PART = 200
MAX_BATCHES_LIMIT = 100  # Limite máximo de batches para evitar sobrecarga do sistema


def delete_other_sheets_and_rows(
    wb: Workbook, sheet_title: str, starting_row: int, rows_per_part: int
):
    ws = wb[sheet_title]
    for worksheet in wb.worksheets:
        if (
            worksheet.sheet_state == Worksheet.SHEETSTATE_VISIBLE
            and worksheet.title != sheet_title
        ):
            wb.remove(worksheet)

    # Remove images from rows before starting_row and after the kept interval
    images_to_remove = []
    if hasattr(ws, "_images") and ws._images:
        for image in ws._images:
            if hasattr(image, "anchor") and hasattr(image.anchor, "_from"):
                row_idx = image.anchor._from.row
                # Remove images before starting_row (but keep row 1 which is index 0)
                # Remove images after the kept interval
                if (
                    row_idx > 0 and row_idx < starting_row + 1
                ) or row_idx > starting_row + rows_per_part:
                    images_to_remove.append(image)
        for image in images_to_remove:
            ws._images.remove(image)
